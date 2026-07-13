"""
scripts/build_fmv_csv.py -- one-off converter: the two 31-01-2018 reference
workbooks under Data/GNUCashReports/ (public market data, NOT PII) into
compact CSVs committed under src/agents/skill_itr_workbook/data/ (plan
section 6.1/OQ-3R, decision D14a).

NSE bhavcopy (NSEBHAVCOPY31012018.xlsx): one row per symbol; FMV = HIGH.
AMFI all-schemes NAV (MFNAV31012018.xlsx): a scheme-name-SECTIONED sheet --
AMC name / scheme name / full option-name text rows, each immediately
followed by one numeric data row (NAV, Repurchase Price, Sale Price,
timestamp). The full option-name text row immediately preceding a numeric
row is the scheme key; FMV = Net Asset Value.

Run once from the repo root: python scripts/build_fmv_csv.py
Not part of the runtime pipeline -- rules.py/schedules.py read the
committed CSV output, never these source workbooks.
"""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC_NSE = ROOT / "Data" / "GNUCashReports" / "NSEBHAVCOPY31012018.xlsx"
SRC_MF = ROOT / "Data" / "GNUCashReports" / "MFNAV31012018.xlsx"
OUT_DIR = ROOT / "src" / "agents" / "skill_itr_workbook" / "data"
OUT_NSE = OUT_DIR / "nse_bhavcopy_31jan2018.csv"
OUT_MF = OUT_DIR / "mf_nav_31jan2018.csv"


def convert_nse() -> int:
    wb = openpyxl.load_workbook(SRC_NSE, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}
    out_rows = []
    for row in rows:
        if row[idx["SYMBOL"]] is None:
            continue
        out_rows.append({
            "symbol": row[idx["SYMBOL"]],
            "isin": row[idx["ISIN"]],
            "fmv_31jan2018": row[idx["HIGH"]],
        })
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUT_NSE, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["symbol", "isin", "fmv_31jan2018"])
        w.writeheader()
        w.writerows(out_rows)
    return len(out_rows)


def convert_mf() -> int:
    wb = openpyxl.load_workbook(SRC_MF, read_only=True, data_only=True)
    ws = wb.active
    out_rows = []
    last_text: str | None = None
    for row in ws.iter_rows(values_only=True):
        first = row[0] if row else None
        if first is None:
            continue
        if isinstance(first, (int, float)):
            if last_text is not None:
                out_rows.append({"scheme_name": last_text, "fmv_31jan2018": float(first)})
        else:
            last_text = str(first).strip()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUT_MF, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["scheme_name", "fmv_31jan2018"])
        w.writeheader()
        w.writerows(out_rows)
    return len(out_rows)


def main() -> None:
    if not SRC_NSE.is_file() or not SRC_MF.is_file():
        print(f"Source workbooks not found under {SRC_NSE.parent} -- nothing to convert.")
        return
    n_nse = convert_nse()
    n_mf = convert_mf()
    print(f"Wrote {OUT_NSE} ({n_nse} rows)")
    print(f"Wrote {OUT_MF} ({n_mf} rows)")


if __name__ == "__main__":
    main()
