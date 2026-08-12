"""
tests/test_gnucash_coverage.py -- agents.skill_gnucash_coverage.agent (the
Coverage Gap Detector). Deterministic, no LLM, no network, no gradio, no
native binaries -- pure XML-in, workbook-out.

Synthetic-only fixtures (a single made-up "Test Person", no real taxpayer
names, PANs, or data files). All books are written into tmp_path.

Covers:
  * an interior zero-transaction month inside an account's active window
    is reported;
  * the calendar month BEFORE an account's first transaction is not
    reported, even though it is inside the book's own FY;
  * a trailing gap (after the account's last transaction) is detected and
    labelled;
  * HIGH vs LOW confidence grading follows the account's own median,
    against the module's own named constant (not a hardcoded "4");
  * a credit-card (LIABILITY/CREDIT) account is in scope and labelled
    "Credit/Liability" in the report;
  * an FY-boundary gap is suppressed when the adjacent FY's registered book
    holds a transaction for the same account in that exact month;
  * opening-balance transactions are excluded from the median (they would
    otherwise skew a quiet account's genesis month upward).
"""
from __future__ import annotations

import gzip
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "src"
ITR_SCRIPTS = SRC / "agents" / "skill_itr_workbook" / "scripts"
COVERAGE_DIR = SRC / "agents" / "skill_gnucash_coverage"

for _p in (str(SRC), str(ITR_SCRIPTS), str(COVERAGE_DIR)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import configs  # noqa: E402
from agents.skill_gnucash_coverage import agent as cov  # noqa: E402

PAN = "ABCDE1234X"

# ---------------------------------------------------------------------------
# Synthetic .gnucash book builder (accounts + transactions).
# ---------------------------------------------------------------------------

_NS_DECL = (
    'xmlns:gnc="http://www.gnucash.org/XML/gnc" '
    'xmlns:act="http://www.gnucash.org/XML/act" '
    'xmlns:trn="http://www.gnucash.org/XML/trn" '
    'xmlns:split="http://www.gnucash.org/XML/split" '
    'xmlns:ts="http://www.gnucash.org/XML/ts" '
    'xmlns:slot="http://www.gnucash.org/XML/slot"'
)

# Fixed account IDs shared across both FY books so account.path matching in
# the FY-boundary check lines up.
ACC = {
    "root": ("Root Account", "ROOT", None, None),
    "assets": ("Assets", "ASSET", "root", [("placeholder", "true")]),
    "bank_parent": ("Bank", "ASSET", "assets", [("placeholder", "true")]),
    "bank": ("TestBank", "BANK", "bank_parent", None),
    "savings": ("SavingsQuarterly", "BANK", "bank_parent", None),
    "liab": ("Liabilities", "LIABILITY", "root", [("placeholder", "true")]),
    "card": ("TestCard", "CREDIT", "liab", None),
    "equity": ("Equity", "EQUITY", "root", [("placeholder", "true")]),
    "ob": ("Opening Balances", "EQUITY", "equity",
           [("equity-type", "opening-balance")]),
    "expense": ("Misc Expense", "EXPENSE", "root", None),
}


def _account_xml(aid: str) -> str:
    name, atype, parent, slots = ACC[aid]
    parts = [
        '  <gnc:account version="2.0.0">',
        f"    <act:name>{name}</act:name>",
        f'    <act:id type="guid">{aid}</act:id>',
        f"    <act:type>{atype}</act:type>",
    ]
    if parent is not None:
        parts.append(f'    <act:parent type="guid">{parent}</act:parent>')
    if slots:
        parts.append("    <act:slots>")
        for key, value in slots:
            parts.append(
                "      <slot>"
                f"<slot:key>{key}</slot:key>"
                f'<slot:value type="string">{value}</slot:value>'
                "</slot>"
            )
        parts.append("    </act:slots>")
    parts.append("  </gnc:account>")
    return "\n".join(parts)


def _txn_xml(date_str: str, legs: list[str], desc: str = "txn") -> str:
    """legs: account-ids for a simple N-way split (each gets an equal and
    opposite dummy value -- date-bucketing is all this skill reads)."""
    splits = []
    for i, aid in enumerate(legs):
        value = "10000/100" if i == 0 else "-10000/100"
        splits.append(
            "   <trn:split>"
            f"<split:account>{aid}</split:account>"
            f"<split:value>{value}</split:value>"
            "</trn:split>"
        )
    return (
        '  <gnc:transaction version="2.0.0">\n'
        f"   <trn:description>{desc}</trn:description>\n"
        f"   <trn:date-posted><ts:date>{date_str} 00:00:00 +0000</ts:date></trn:date-posted>\n"
        "   <trn:splits>\n" + "\n".join(splits) + "\n   </trn:splits>\n"
        "  </gnc:transaction>"
    )


def _book_xml(account_ids: list[str], txns: list[str]) -> str:
    body = [_account_xml(a) for a in account_ids] + txns
    return (
        '<?xml version="1.0" encoding="utf-8"?>\n'
        f"<gnc-v2 {_NS_DECL}>\n"
        '<gnc:book version="2.0.0">\n'
        + "\n".join(body)
        + "\n</gnc:book>\n</gnc-v2>\n"
    )


def _write_gz(path: Path, xml: str) -> None:
    path.write_bytes(gzip.compress(xml.encode("utf-8")))


def _dates_in_month(year: int, month: int, days: list[int]) -> list[str]:
    return [f"{year:04d}-{month:02d}-{d:02d}" for d in days]


# ---------------------------------------------------------------------------
# FY2024-25 book: TestBank (busy, one interior gap, one trailing gap),
# SavingsQuarterly (quiet, LOW confidence), TestCard (credit-card scope).
# ---------------------------------------------------------------------------

_ALL_ACCOUNTS = ["root", "assets", "bank_parent", "bank", "savings",
                 "liab", "card", "equity", "ob", "expense"]


def _busy_month_txns(aid: str, year: int, month: int, n: int = 5) -> list[str]:
    days = [1, 5, 10, 15, 20][:n]
    return [_txn_xml(d, [aid, "expense"]) for d in _dates_in_month(year, month, days)]


def _build_fy2425_book() -> str:
    txns = []
    # TestBank: opening-balance-only genesis month (2024-04), then 5
    # non-OB txns/month straight through except October (interior gap) and
    # March (trailing gap, also the FY's own last calendar month).
    txns.append(_txn_xml("2024-04-05", ["bank", "ob"], desc="opening balance"))
    for (y, m) in [(2024, 5), (2024, 6), (2024, 7), (2024, 8), (2024, 9),
                   (2024, 11), (2024, 12), (2025, 1), (2025, 2)]:
        txns += _busy_month_txns("bank", y, m)
    # (2024, 10) and (2025, 3) deliberately left with zero TestBank txns.

    # SavingsQuarterly: one interest credit every quarter -- quiet cadence,
    # median 0, so its zero months must grade LOW.
    for (y, m, d) in [(2024, 4, 10), (2024, 7, 10), (2024, 10, 10), (2025, 1, 10)]:
        txns.append(_txn_xml(f"{y:04d}-{m:02d}-{d:02d}", ["savings", "expense"],
                              desc="quarterly interest"))

    # TestCard: monthly except September (its own interior gap), to prove
    # a LIABILITY/CREDIT account is in scope and labelled correctly.
    for (y, m) in [(2024, 5), (2024, 6), (2024, 7), (2024, 8),
                   (2024, 10), (2024, 11), (2024, 12), (2025, 1), (2025, 2)]:
        txns += _busy_month_txns("card", y, m, n=4)

    return _book_xml(_ALL_ACCOUNTS, txns)


def _build_fy2526_adjacent_book() -> str:
    """Only what's needed for the FY-boundary check: a TestBank transaction
    dated 2025-03-28 -- i.e. filed into the FY2025-26 book even though the
    date falls in FY2024-25's own last month, the "wrong side of the
    rollover" scenario the boundary check exists to catch."""
    txns = [_txn_xml("2025-03-28", ["bank", "expense"], desc="misfiled into next book")]
    # Give the book at least one txn safely inside its own FY too.
    txns.append(_txn_xml("2025-04-15", ["bank", "expense"], desc="normal FY25-26 txn"))
    return _book_xml(_ALL_ACCOUNTS, txns)


@pytest.fixture()
def fy2425_book(tmp_path) -> Path:
    p = tmp_path / "TestPersonCoverage2425.gnucash"
    _write_gz(p, _build_fy2425_book())
    return p


@pytest.fixture()
def fy2526_book(tmp_path) -> Path:
    p = tmp_path / "TestPersonCoverage2526.gnucash"
    _write_gz(p, _build_fy2526_adjacent_book())
    return p


def _entities_path(tmp_path: Path, fy2425: Path, fy2526: Path | None = None) -> Path:
    books = {"2024-25": str(fy2425)}
    if fy2526 is not None:
        books["2025-26"] = str(fy2526)
    entities = {
        "TestPersonCoverage": configs.EntityProfile(
            key="TestPersonCoverage", name="Test Person Coverage", pan=PAN,
            status="Individual", books=books,
        ),
    }
    out = tmp_path / "entities.yaml"
    out.write_text(configs.dump_entities(entities), encoding="utf-8")
    return out


def _gap_map(gaps: list[cov.GapRow]) -> dict:
    """{(account_path, month): GapRow} for easy lookup by test assertions."""
    return {(g.account_path, g.month): g for g in gaps}


def _run_scan(books, tmp_path, entities_path=None):
    out_path = tmp_path / "out.xlsx"
    entities_path = entities_path if entities_path is not None else (tmp_path / "no-such-entities.yaml")
    summary = cov.run(
        books=books if isinstance(books, str) else "\n".join(str(b) for b in books),
        output_path=str(out_path),
        entities_path=str(entities_path),
    )
    return summary, out_path


# ---------------------------------------------------------------------------
# Core scenarios: interior gap, pre-genesis exclusion, trailing gap, median
# grading, credit-card scope, opening-balance exclusion -- all against the
# FY2024-25 book alone, WITHOUT a registered entity (so the FY-boundary
# check never fires and the March trailing gap is reported un-suppressed).
# ---------------------------------------------------------------------------

def test_interior_zero_month_is_reported(fy2425_book, tmp_path):
    _summary, out_path = _run_scan([fy2425_book], tmp_path)
    wb = load_workbook(out_path)
    ws = wb["Gaps"]
    months = {row[4].value for row in ws.iter_rows(min_row=2)
              if row[2].value == "Assets:Bank:TestBank"}
    assert "2024-10" in months


def test_month_before_first_transaction_not_reported(fy2425_book, tmp_path):
    """TestBank's first transaction (incl. its opening balance) is
    2024-04-05, so the active window starts at April 2024 -- no earlier
    month may appear, even though the book's own FY starts 2024-04-01."""
    _summary, out_path = _run_scan([fy2425_book], tmp_path)
    wb = load_workbook(out_path)
    ws = wb["Gaps"]
    months = {row[4].value for row in ws.iter_rows(min_row=2)
              if row[2].value == "Assets:Bank:TestBank"}
    assert "2024-03" not in months
    assert "2024-02" not in months
    # April itself is NOT a gap either: the opening-balance txn keeps its
    # transaction COUNT non-zero even though it is excluded from the median.
    assert "2024-04" not in months


def test_trailing_gap_detected_and_labelled(fy2425_book, tmp_path):
    _summary, out_path = _run_scan([fy2425_book], tmp_path)
    wb = load_workbook(out_path)
    ws = wb["Gaps"]
    rows = {(row[2].value, row[4].value): row for row in ws.iter_rows(min_row=2)}
    march = rows[("Assets:Bank:TestBank", "2025-03")]
    assert march[7].value == "TRAILING"
    # The interior October gap must NOT be mislabelled trailing.
    october = rows[("Assets:Bank:TestBank", "2024-10")]
    assert october[7].value != "TRAILING"


def test_high_vs_low_confidence_follows_median_threshold(fy2425_book, tmp_path):
    summary, out_path = _run_scan([fy2425_book], tmp_path)
    wb = load_workbook(out_path)
    ws = wb["Gaps"]
    rows = {(row[2].value, row[4].value): row for row in ws.iter_rows(min_row=2)}

    # TestBank runs 5 txns/month most months -> median well above the
    # module's own named threshold -> HIGH.
    bank_gap = rows[("Assets:Bank:TestBank", "2024-10")]
    assert bank_gap[6].value >= cov.HIGH_CONFIDENCE_MEDIAN_THRESHOLD
    assert bank_gap[5].value == "HIGH"

    # SavingsQuarterly posts once a quarter -> median 0 -> LOW. June 2024 is
    # an interior zero month for it (not FY-boundary, no OB txn nearby).
    savings_gap = rows[("Assets:Bank:SavingsQuarterly", "2024-06")]
    assert savings_gap[6].value < cov.HIGH_CONFIDENCE_MEDIAN_THRESHOLD
    assert savings_gap[5].value == "LOW"


def test_credit_card_account_in_scope_and_labelled(fy2425_book, tmp_path):
    _summary, out_path = _run_scan([fy2425_book], tmp_path)
    wb = load_workbook(out_path)
    ws = wb["Gaps"]
    rows = [row for row in ws.iter_rows(min_row=2)
            if row[2].value == "Liabilities:TestCard"]
    assert rows, "TestCard (CREDIT/LIABILITY) produced no gap rows at all"
    assert any(row[4].value == "2024-09" for row in rows)
    assert all(row[3].value == "Credit/Liability" for row in rows)

    # And the Summary sheet carries the same class label.
    sm = wb["Summary"]
    card_summary = [row for row in sm.iter_rows(min_row=2)
                     if row[2].value == "Liabilities:TestCard"]
    assert card_summary and card_summary[0][3].value == "Credit/Liability"


def test_opening_balance_excluded_from_median(fy2425_book, tmp_path):
    """If the opening-balance txn were NOT excluded, TestBank's April count
    would be treated as 1 real transaction rather than 0, nudging the
    median. Assert the reported median matches the OB-excluded value (5,
    the steady non-OB monthly cadence) rather than something skewed by
    counting April's OB-only entry as real activity."""
    _summary, out_path = _run_scan([fy2425_book], tmp_path)
    wb = load_workbook(out_path)
    sm = wb["Summary"]
    bank_row = next(row for row in sm.iter_rows(min_row=2)
                     if row[2].value == "Assets:Bank:TestBank")
    assert bank_row[7].value == 5.0


# ---------------------------------------------------------------------------
# FY-boundary suppression: needs the entity registered with BOTH the
# FY2024-25 book and the adjacent FY2025-26 book.
# ---------------------------------------------------------------------------

def test_fy_boundary_gap_suppressed_via_adjacent_book(fy2425_book, fy2526_book, tmp_path):
    entities_path = _entities_path(tmp_path, fy2425_book, fy2526_book)
    _summary, out_path = _run_scan([fy2425_book], tmp_path, entities_path=entities_path)
    wb = load_workbook(out_path)

    gaps_ws = wb["Gaps"]
    months = {row[4].value for row in gaps_ws.iter_rows(min_row=2)
              if row[2].value == "Assets:Bank:TestBank"}
    # March would ordinarily be a trailing+boundary gap, but the adjacent
    # FY2025-26 book has a TestBank transaction dated 2025-03-28 -> suppressed.
    assert "2025-03" not in months
    # The interior October gap is unaffected by the boundary check.
    assert "2024-10" in months

    sm = wb["Summary"]
    bank_row = next(row for row in sm.iter_rows(min_row=2)
                     if row[2].value == "Assets:Bank:TestBank")
    assert bank_row[11].value == 1  # FY-Boundary Suppressed count


def test_fy_boundary_gap_reported_without_adjacent_evidence(fy2425_book, tmp_path):
    """Same entity registered, but with NO FY2025-26 book at all -- nothing
    to consult, so March must be reported (not silently dropped)."""
    entities_path = _entities_path(tmp_path, fy2425_book, fy2526=None)
    _summary, out_path = _run_scan([fy2425_book], tmp_path, entities_path=entities_path)
    wb = load_workbook(out_path)
    gaps_ws = wb["Gaps"]
    months = {row[4].value for row in gaps_ws.iter_rows(min_row=2)
              if row[2].value == "Assets:Bank:TestBank"}
    assert "2025-03" in months
