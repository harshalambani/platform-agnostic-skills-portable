"""Tests for agents.skill_partner_comp_recon.

Fixture data (tests/fixtures/partner_comp_recon_fy2025_26.yaml) is entirely
invented and self-consistent -- no figure is copied or derived from any
real document. See that file's header comment and AGENT.md for how each
number was chosen and what it is meant to exercise.
"""
from __future__ import annotations

import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import pytest
import yaml

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SRC = PROJECT_ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from agents.skill_partner_comp_recon import engine, writer
from agents.skill_partner_comp_recon.agent import run
from agents.skill_partner_comp_recon.engine import (
    CANNOT_RECONCILE,
    build_report,
    classify_cohort_instalments,
    derive_misc,
    detect_mid_year_rate_change,
    driver,
    field_or_reason,
    fy_of_date,
    gross_up_one_off,
    required_cumulative_capital,
)
from agents.skill_partner_comp_recon.jv_emitter import (
    JOURNAL_HEADERS,
    build_journals,
    fy_prefix,
    write_journal_csv,
)
from agents.skill_partner_comp_recon.parsers.payout_advice import (
    NotAnL1DocumentError,
    parse_l1_text,
)
from agents.skill_partner_comp_recon.parsers.advisory import (
    NotAnL3DocumentError,
    parse_l3_text,
)
from agents.skill_partner_comp_recon.parsers.llp_statement import (
    NotAnL5DocumentError,
    parse_l5_words,
)

FIXTURE_PATH = Path(__file__).parent / "fixtures" / "partner_comp_recon_fy2025_26.yaml"


def _load_fixture() -> dict:
    import yaml

    return yaml.safe_load(FIXTURE_PATH.read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# Package plumbing
# ---------------------------------------------------------------------------

def test_skill_partner_comp_recon_imports():
    import agents.skill_partner_comp_recon as pkg
    assert pkg.__doc__
    assert hasattr(run, "__call__")


# ---------------------------------------------------------------------------
# N1 -- every rate/percentage/period is an input, never a constant/fallback.
# A missing driver must produce an explicit cannot-reconcile row, never a
# default or a computed guess.
# ---------------------------------------------------------------------------

def test_missing_driver_produces_cannot_reconcile_not_a_default():
    value, reason = driver({}, "firms_tax_rate", "2025-26", "firm's tax rate")
    assert value is None
    assert reason is not None
    assert CANNOT_RECONCILE in reason
    assert "firm's tax rate" in reason
    assert "2025-26" in reason


def test_missing_field_produces_cannot_reconcile():
    value, reason = field_or_reason({}, "bank_credits_total", "bank credits total")
    assert value is None
    assert CANNOT_RECONCILE in reason


def test_capital_rule_missing_any_input_is_cannot_reconcile_never_a_guess():
    result = required_cumulative_capital(
        target_compensation=10_000_000, months_achieved=12, months_total=None,
        rate=0.40, fy="2025-26",
    )
    assert result.status == CANNOT_RECONCILE
    assert result.required_cumulative_capital is None
    assert "capital months total" in result.reason


def test_build_report_with_no_drivers_at_all_never_computes_a_capital_figure():
    data = {"financial_year": "2025-26", "drivers": {}, "monthly": [], "cohorts": []}
    report = build_report(data)
    assert report.capital_rule.status == CANNOT_RECONCILE
    assert report.capital_rule.required_cumulative_capital is None


# ---------------------------------------------------------------------------
# s.3.1 -- the misc adjustment is always derived, never read off the advice.
# ---------------------------------------------------------------------------

def test_derive_misc():
    # total_paid - remuneration - share_of_profit_gross - additional_share_of_profit
    assert derive_misc(595280, 300000, 500000, 0) == pytest.approx(-204720)
    assert derive_misc(896400, 300000, 500000, 650560) == pytest.approx(-554160)


# ---------------------------------------------------------------------------
# s.3.2 -- grossing up a net one-off, with the roundness check.
# ---------------------------------------------------------------------------

def test_gross_up_one_off_confirmed_when_it_lands_on_a_round_lakh():
    result = gross_up_one_off(650560, 0.349440, "2025-26")
    assert result.gross == pytest.approx(1_000_000, abs=1)
    assert result.roundness == pytest.approx(0, abs=1)
    assert result.status == "CONFIRMED"


def test_gross_up_one_off_suspect_when_rate_is_wrong_for_the_year():
    result = gross_up_one_off(650560, 0.30, "2025-26")
    assert result.gross == pytest.approx(929_371.43, abs=0.01)
    # nearest lakh to 929,371.43 is 900,000 -- distance is ~29,371.43, well
    # past ROUNDNESS_TOLERANCE (1000), so this is correctly flagged SUSPECT.
    assert result.roundness == pytest.approx(29_371.43, abs=0.01)
    assert result.status == "SUSPECT"


def test_gross_up_one_off_missing_rate_is_cannot_reconcile():
    result = gross_up_one_off(650560, None, "2025-26")
    assert result.status == CANNOT_RECONCILE
    assert result.gross is None


# ---------------------------------------------------------------------------
# s.3.3 -- firm's tax must never show up as a TDS credit in Form 26AS.
# ---------------------------------------------------------------------------

def test_firms_tax_conflation_detected():
    note = engine.firms_tax_conflated_with_26as(
        firms_tax_total=-3_144_960, form_26as_total_credit=3_504_960,
        computed_creditable_tds=360_000,
    )
    assert note is not None
    assert "conflat" in note.lower() or "Form 26AS" in note


def test_firms_tax_no_conflation_when_26as_matches_tds_only():
    note = engine.firms_tax_conflated_with_26as(
        firms_tax_total=-3_144_960, form_26as_total_credit=360_000,
        computed_creditable_tds=360_000,
    )
    assert note is None


# ---------------------------------------------------------------------------
# s.4 -- cohort FY straddle: an instalment is assigned to its PAYMENT fy,
# never its award fy, and a following-year instalment is excluded.
# ---------------------------------------------------------------------------

def test_fy_of_date():
    assert fy_of_date("2025-07-31") == "2025-26"
    assert fy_of_date("2026-01-15") == "2025-26"
    assert fy_of_date("2026-06-30") == "2026-27"


def test_cohort_instalment_paid_next_fy_is_labelled_future_and_excluded():
    cohort = _load_fixture()["cohorts"][0]
    rows = classify_cohort_instalments(cohort, reporting_fy="2025-26")
    assert len(rows) == 3
    reporting = [r for r in rows if r.membership == "reporting"]
    future = [r for r in rows if r.membership == "future"]
    assert len(reporting) == 2
    assert len(future) == 1
    assert future[0].instalment_fy == "2026-27"
    assert future[0].label == "future (FY 2026-27)"
    # The award FY is preserved for traceability but never used for FY
    # assignment.
    assert future[0].award_fy == "2024-25"


# ---------------------------------------------------------------------------
# s.5.1 -- the capital rule.
# ---------------------------------------------------------------------------

def test_required_cumulative_capital():
    result = required_cumulative_capital(
        target_compensation=10_000_000, months_achieved=12, months_total=48,
        rate=0.40, fy="2025-26",
    )
    assert result.status == "OK"
    assert result.required_cumulative_capital == pytest.approx(1_000_000)


# ---------------------------------------------------------------------------
# s.5.2 -- the mid-year capital-rate-change detector.
# ---------------------------------------------------------------------------

def test_mid_year_rate_change_detected():
    suspect = detect_mid_year_rate_change(
        [416667, 291667, 291667], target_compensation=10_000_000,
        months_achieved=12, months_total=48,
    )
    assert suspect is not None
    assert suspect.implied_old_rate == pytest.approx(0.500, abs=0.001)
    assert suspect.implied_new_rate == pytest.approx(0.400, abs=0.001)
    assert "RATE CHANGE SUSPECTED" in suspect.note


def test_no_rate_change_when_all_instalment_capitals_equal():
    suspect = detect_mid_year_rate_change(
        [291667, 291667, 291667], target_compensation=10_000_000,
        months_achieved=12, months_total=48,
    )
    assert suspect is None


# ---------------------------------------------------------------------------
# End-to-end: build_report() over the full fixture.
# ---------------------------------------------------------------------------

def test_build_report_end_to_end_against_fixture():
    data = _load_fixture()
    report = build_report(data)

    assert report.financial_year == "2025-26"
    assert len(report.monthly) == 12
    # Only the two in-FY instalments count toward the reporting FY; the
    # third (2026-06-30) is next FY and must not appear in cohort totals.
    reporting = [i for i in report.cohort_instalments if i.membership == "reporting"]
    assert len(reporting) == 2

    by_category = {r.category: r for r in report.reconciliation}
    for cat, r in by_category.items():
        if cat.startswith("Incentive schedule"):
            assert r.agree is None, "Leg 2 has no independent PDF source in this build"
        else:
            assert r.agree is True, f"expected AGREE for {cat!r}, got {r.agree!r} ({r.note})"

    # The rate-change detector must fire on the fixture's one cohort even
    # though the aggregate capital rule (above) agreed exactly -- an
    # aggregate match must never mask an instalment-level anomaly.
    assert len(report.rate_change_suspects) == 1
    assert "RATE CHANGE SUSPECTED" in report.rate_change_suspects[0].note

    assert len(report.one_offs) == 1
    assert report.one_offs[0].status == "CONFIRMED"

    assert report.tds_month_exceptions == []


# ---------------------------------------------------------------------------
# Parser guard tests -- s.7. These are deliberately NOT implemented. Do not
# "finish" them opportunistically -- see AGENT.md's Stage 1/Stage 2 split.
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    "modname, needle",
    [
        ("payment_schedule", "specimen"),
    ],
)
def test_parser_is_a_guarded_placeholder(modname, needle):
    module = __import__(
        f"agents.skill_partner_comp_recon.parsers.{modname}", fromlist=["parse"]
    )
    with pytest.raises(NotImplementedError) as excinfo:
        module.parse("does-not-matter.pdf")
    assert needle in str(excinfo.value).lower()


# ---------------------------------------------------------------------------
# L1 (monthly partner payout certificate) parser -- payout_advice.py.
# parse_l1_text() is the PURE core; every fixture below is a synthetic,
# self-invented text block shaped like real specimens' extracted text
# (see payout_advice.py's module docstring), never a real document.
# ---------------------------------------------------------------------------

def _l1_text(
    *,
    month_phrase="Jan-26",
    remuneration="1,20,000",
    share_of_profit="80,000",
    additional_share_of_profit=None,
    tds=None,
    misc="5,000",
    total="2,05,000",
    words=None,
    issue_date="05-Feb-26",
    include_month_phrase=True,
    extra_lines=None,
):
    """Build a synthetic L1 body text block. Row order/spacing is not
    load-bearing -- parse_l1_text() maps by label, not position."""
    lines = ["To Whomsoever It may concern"]
    if include_month_phrase:
        lines.append(
            f"This is to certify that the amount payable for the month of "
            f"{month_phrase} is as under."
        )
    lines.append(f"Issued on {issue_date}")
    lines.append("Particulars                AMOUNTS")
    lines.append(f"Remuneration                {remuneration}")
    lines.append(f"Share of Profit             {share_of_profit}")
    if additional_share_of_profit is not None:
        lines.append(f"Add. Share of Profit         {additional_share_of_profit}")
    if tds is not None:
        lines.append(f"TDS on Remuneration          {tds}")
    lines.append(f"Misc Adjustments             {misc}")
    lines.append(f"Total                        {total}")
    if words is not None:
        lines.append(f"Amount in Words: {words}")
    if extra_lines:
        lines.extend(extra_lines)
    if not include_month_phrase:
        lines.append(month_phrase)
    return "\n".join(lines)


# 1 -- a TDS row absent vs present; absent must come through as None, not 0.
def test_l1_tds_row_absent_is_none_not_zero():
    without_tds = _l1_text(tds=None, misc="5,000", total="2,05,000")
    record = parse_l1_text(without_tds, source_name="jan26_without_tds.pdf")
    assert record["tds"] is None
    assert not any("ERROR" in d for d in record["diagnostics"])

    with_tds = _l1_text(tds="(5,000)", misc="5,000", total="2,00,000")
    record2 = parse_l1_text(with_tds, source_name="jan26_with_tds.pdf")
    assert record2["tds"] == -5000.0
    assert not any("ERROR" in d for d in record2["diagnostics"])


# 2 -- a parenthesised amount parses negative and is never abs()-ed.
def test_l1_parenthesised_amount_parses_negative():
    text = _l1_text(share_of_profit="(30,000)", misc="5,000", total="95,000")
    record = parse_l1_text(text, source_name="jan26.pdf")
    assert record["share_of_profit_gross"] == -30000.0


# 3 -- a "#N/A" token in the amount column never parses as a value.
def test_l1_na_token_in_amount_column_is_not_a_value():
    text = _l1_text(misc="#N/A", total="2,00,000")
    record = parse_l1_text(text, source_name="jan26.pdf")
    assert record["misc_printed"] is None
    assert not any("ERROR" in d for d in record["diagnostics"])


# 4 -- two L1 blocks with the same body month but different filenames
# resolve to one month.
def test_l1_same_body_month_different_filenames_resolve_to_one_month():
    text_a = _l1_text(month_phrase="Jan-26")
    text_b = _l1_text(month_phrase="Jan-26")
    record_a = parse_l1_text(text_a, source_name="advice_batch1_xyz123.pdf")
    record_b = parse_l1_text(text_b, source_name="completely_different_name.pdf")
    assert record_a["month"] == record_b["month"] == "2026-01"


# 5 -- rows that do not sum to Total produce the fail-loud diagnostic,
# never a silent pass and never an exception.
def test_l1_row_sum_mismatch_is_a_fail_loud_diagnostic_not_silent():
    text = _l1_text(total="9,99,999")  # deliberately wrong
    record = parse_l1_text(text, source_name="jan26.pdf")
    assert record["total_paid"] == 999999.0
    assert any("ERROR" in d and "sum" in d.lower() for d in record["diagnostics"])


# 6 -- a non-L1 document (a payroll salary statement) is skipped cleanly,
# never misparsed as a payout certificate.
def test_l1_non_l1_document_is_skipped_not_misparsed():
    salary_text = (
        "SALARY STATEMENT FOR THE MONTH OF JANUARY 2026\n"
        "Basic Pay              1,00,000\n"
        "Net Pay                1,00,000\n"
    )
    with pytest.raises(NotAnL1DocumentError):
        parse_l1_text(salary_text, source_name="jan26_salary.pdf")


# 7 -- month is taken from the body even when the filename says a
# different month.
def test_l1_month_from_body_ignores_filename():
    text = _l1_text(month_phrase="Jan-26")
    record = parse_l1_text(text, source_name="payout_advice_MARCH.pdf")
    assert record["month"] == "2026-01"


# Extra coverage: the body-month fallback (no "for the month of" phrase)
# must not be fooled by the issue date's "DD-Mon-YY" shape.
def test_l1_bare_month_line_not_confused_with_issue_date():
    text = _l1_text(include_month_phrase=False, month_phrase="Jan-26", issue_date="05-Feb-26")
    record = parse_l1_text(text, source_name="jan26.pdf")
    assert record["month"] == "2026-01"
    assert record["issue_date"] == "05-Feb-26"


# Extra coverage: an earlier-year template's literal "#N/A" standing in
# for the amount-in-words line is skipped, not parsed as a value.
def test_l1_na_placeholder_for_words_line_is_skipped():
    text = _l1_text(words="#N/A")
    record = parse_l1_text(text, source_name="jan26.pdf")
    assert record["amount_in_words_value"] is None
    assert not any("words" in d.lower() for d in record["diagnostics"])


# Extra coverage: the amount-in-words line carries paise, Total is
# rounded -- cross-checked to +/- 1.00, never exact equality.
def test_l1_amount_in_words_paise_cross_check_within_tolerance():
    text = _l1_text(
        total="2,05,000",
        words="Rupees Two Lakh Five Thousand and Forty Paise Only 2,05,000.40",
    )
    record = parse_l1_text(text, source_name="jan26.pdf")
    assert record["amount_in_words_value"] == 205000.40
    assert not any("words" in d.lower() for d in record["diagnostics"])


def test_l1_amount_in_words_mismatch_beyond_tolerance_is_fail_loud():
    text = _l1_text(
        total="2,05,000",
        words="Rupees Two Lakh Seven Thousand Only 2,07,000.00",
    )
    record = parse_l1_text(text, source_name="jan26.pdf")
    assert any("words" in d.lower() and "ERROR" in d for d in record["diagnostics"])


# Misc Adjustments is parsed and carried as a printed CHECK figure, but
# never fed into the model's input side under a different name.
def test_l1_misc_adjustments_is_printed_only_never_an_input_field():
    text = _l1_text(misc="5,000", total="2,05,000")
    record = parse_l1_text(text, source_name="jan26.pdf")
    assert record["misc_printed"] == 5000.0
    assert "misc" not in record or record.get("misc") is None
    assert "misc_adjustments" not in record


# A run with an unparseable/malformed required document (garbage-byte
# "PDF") still fails loud end-to-end, naming the document -- now for a
# pdfplumber-level "can't open this" reason rather than the old
# NotImplementedError-placeholder reason, since advisory.py is
# implemented. Optional legs still degrade to "not available" rather than
# fail, even though a required leg failed.
def test_agent_run_document_driven_still_fails_loud_on_unparseable_advisory(tmp_path):
    advices_dir = _advices_dir_with_one_pdf(tmp_path)
    advisory_path = tmp_path / "advisory.pdf"
    advisory_path.write_bytes(b"%PDF-1.4 not a real pdf")

    result = run(
        entity="Harshal",
        advices_dir=str(advices_dir),
        advisory_path=str(advisory_path),
        output_path=str(tmp_path / "out.xlsx"),
    )
    assert result.startswith("ERROR")
    assert "advisory" in result.lower()
    # Optional-leg status is still surfaced beneath the fail-loud header,
    # never silently dropped just because a required leg failed.
    assert "not available" in result.lower() or "optional" in result.lower()


# ---------------------------------------------------------------------------
# L3 (annual Compensation Advisory letter) parser -- advisory.py.
# parse_l3_text() is the PURE core; every fixture below is a synthetic,
# self-invented text block shaped like real specimens' extracted text (see
# advisory.py's module docstring), never a real document.
# ---------------------------------------------------------------------------

def _l3_text(
    *,
    report_year="2026",
    salary="12,00,000",
    remuneration="24,00,000",
    share_of_profit="18,00,000",
    arrears=None,
    incentive_gross="3,00,000",
    target_compensation="57,00,000",
    prior_year_target_compensation="52,00,000",
    interest_on_capital=None,
    drawings="40,00,000",
    interest_paid="1,00,000",
    balance="18,00,000",
    less_firms_tax="(6,00,000)",
    less_capital_contribution="(2,00,000)",
    net_payable="10,00,000",
    opening_balance="20,00,000",
    closing_balance="24,00,000",
    instalments=(
        {"no": 1, "gross": "2,50,000", "firms_tax": "(1,00,000)", "capital_contribution": "(50,000)", "net": "1,00,000"},
        {"no": 2, "gross": "2,50,000", "firms_tax": "(1,00,000)", "capital_contribution": "(50,000)", "net": "1,00,000"},
    ),
    extra_part1_lines=None,
    extra_part2_lines=None,
    extra_part3_lines=None,
    include_payments_header=True,
    include_schedule_header=True,
    fy_phrase=None,
):
    """Build a synthetic L3 Advisory body text block. Row order/spacing is
    not load-bearing -- parse_l3_text() maps by label, not position."""
    lines = ["Compensation Advisory"]
    lines.append(
        fy_phrase
        if fy_phrase is not None
        else f"For the year ended 31 March {report_year}"
    )
    lines.append("")
    if salary is not None:
        lines.append(f"Salary                              {salary}")
    if remuneration is not None:
        lines.append(f"Remuneration                        {remuneration}")
    if share_of_profit is not None:
        lines.append(f"Share of Profit                     {share_of_profit}")
    if arrears is not None:
        lines.append(f"Arrears                             {arrears}")
    if incentive_gross is not None:
        lines.append(f"Incentive                           {incentive_gross}")
    if target_compensation is not None:
        lines.append(f"Target Compensation                 {target_compensation}")
    if prior_year_target_compensation is not None:
        lines.append(f"Prior Year Target Compensation      {prior_year_target_compensation}")
    if interest_on_capital is not None:
        lines.append(f"Interest on Capital                 {interest_on_capital}")
    if extra_part1_lines:
        lines.extend(extra_part1_lines)

    if include_payments_header:
        lines.append("PAYMENTS")
    if drawings is not None:
        lines.append(f"Drawings                            {drawings}")
    if interest_paid is not None:
        lines.append(f"Interest Paid                       {interest_paid}")
    if balance is not None:
        lines.append(f"Balance                             {balance}")
    if less_firms_tax is not None:
        lines.append(f"Less: Firm's Tax / TDS              {less_firms_tax}")
    if less_capital_contribution is not None:
        lines.append(f"Less: Capital Contribution          {less_capital_contribution}")
    if net_payable is not None:
        lines.append(f"Net Payable                         {net_payable}")
    if extra_part2_lines:
        lines.extend(extra_part2_lines)

    if include_schedule_header:
        lines.append("SCHEDULE")
    if opening_balance is not None:
        lines.append(f"Opening Balance                     {opening_balance}")
    for inst in instalments:
        lines.append(
            f"Instalment No. {inst['no']}   {inst['gross']}   {inst['firms_tax']}   "
            f"{inst['capital_contribution']}   {inst['net']}"
        )
    if closing_balance is not None:
        lines.append(f"Projected Closing Balance            {closing_balance}")
    if extra_part3_lines:
        lines.extend(extra_part3_lines)

    return "\n".join(lines)


# 1 -- the reported FY comes from the document text, never the filename.
def test_l3_reported_fy_from_document_text_not_filename():
    text = _l3_text(report_year="2026")
    record = parse_l3_text(text, source_name="advisory_fy2099-00_wrong_name.pdf")
    assert record["financial_year"] == "2025-26"


# 2 -- "interest on capital" absent parses as None, not 0 (earlier years).
def test_l3_interest_on_capital_absent_is_none_not_zero():
    text = _l3_text(interest_on_capital=None)
    record = parse_l3_text(text, source_name="advisory.pdf")
    assert record["interest_on_capital"] is None
    assert not any("ERROR" in d for d in record["diagnostics"])

    text2 = _l3_text(interest_on_capital="45,000")
    record2 = parse_l3_text(text2, source_name="advisory.pdf")
    assert record2["interest_on_capital"] == 45000.0


# 3 -- a parenthesised amount parses negative, never abs()-ed.
def test_l3_parenthesised_amount_parses_negative():
    text = _l3_text(less_firms_tax="(6,00,000)")
    record = parse_l3_text(text, source_name="advisory.pdf")
    assert record["less_firms_tax"] == -600000.0


# 4 -- both the schedule's opening and projected closing balance are
# exposed, so a caller can later chain them across consecutive Advisories.
def test_l3_schedule_opening_and_closing_balance_both_exposed():
    text = _l3_text(opening_balance="20,00,000", closing_balance="24,00,000")
    record = parse_l3_text(text, source_name="advisory.pdf")
    assert record["schedule_opening_balance"] == 2000000.0
    assert record["schedule_projected_closing_balance"] == 2400000.0


# 5 -- a non-reconciling PAYMENTS block produces a fail-loud diagnostic,
# never a silent plug.
def test_l3_non_reconciling_payments_block_is_fail_loud():
    text = _l3_text(
        balance="18,00,000",
        less_firms_tax="(6,00,000)",
        less_capital_contribution="(2,00,000)",
        net_payable="99,99,999",  # deliberately wrong
    )
    record = parse_l3_text(text, source_name="advisory.pdf")
    assert any("ERROR" in d and "PAYMENTS" in d for d in record["diagnostics"])


# 6 -- a non-reconciling SCHEDULE instalment row produces a fail-loud
# diagnostic, never a silent plug.
def test_l3_non_reconciling_schedule_instalment_is_fail_loud():
    text = _l3_text(
        instalments=(
            {"no": 1, "gross": "2,50,000", "firms_tax": "(1,00,000)", "capital_contribution": "(50,000)", "net": "9,99,999"},
        ),
    )
    record = parse_l3_text(text, source_name="advisory.pdf")
    assert any("ERROR" in d and "instalment 1" in d for d in record["diagnostics"])


# 7 -- an unrecognised label line is reported on unknown_labels, never
# silently dropped.
def test_l3_unrecognised_label_line_is_reported_not_dropped():
    text = _l3_text(extra_part1_lines=["Signing Bonus                       1,50,000"])
    record = parse_l3_text(text, source_name="advisory.pdf")
    assert any("Signing Bonus" in u for u in record["unknown_labels"])


# 8 -- a non-L3 document (an L2 salary statement) is skipped cleanly, not
# misparsed as an L3 Advisory.
def test_l3_non_l3_salary_statement_is_skipped_not_misparsed():
    text = "SALARY STATEMENT FOR the month of March 2026\nNet Pay   2,00,000"
    with pytest.raises(NotAnL3DocumentError) as excinfo:
        parse_l3_text(text, source_name="salary_march26.pdf")
    assert "salary statement" in str(excinfo.value).lower()


# an L1 payout certificate is also skipped cleanly, not misparsed.
def test_l3_non_l3_l1_certificate_is_skipped_not_misparsed():
    text = "To Whomsoever It may concern\nRemuneration   1,20,000"
    with pytest.raises(NotAnL3DocumentError):
        parse_l3_text(text, source_name="jan26.pdf")


# 9 -- no rate is ever hardcoded: the parsed firm's-tax amount tracks the
# synthetic document's own printed figure, not a computed-from-a-rate
# constant.
def test_l3_no_hardcoded_rate_firms_tax_tracks_the_printed_figure():
    text_a = _l3_text(
        balance="18,00,000", less_firms_tax="(6,00,000)",
        less_capital_contribution="(2,00,000)", net_payable="10,00,000",
    )
    record_a = parse_l3_text(text_a, source_name="advisory.pdf")

    text_b = _l3_text(
        balance="18,00,000", less_firms_tax="(9,00,000)",
        less_capital_contribution="(2,00,000)", net_payable="7,00,000",
    )
    record_b = parse_l3_text(text_b, source_name="advisory.pdf")

    assert record_a["less_firms_tax"] == -600000.0
    assert record_b["less_firms_tax"] == -900000.0
    assert record_a["less_firms_tax"] != record_b["less_firms_tax"]
    assert not any("ERROR" in d for d in record_a["diagnostics"])
    assert not any("ERROR" in d for d in record_b["diagnostics"])


# A document missing the FY phrase and/or PAYMENTS/SCHEDULE headers, with
# no recognisable L1/L2 marker either, is still skipped cleanly via the
# generic reason -- never crashes, never guesses an L3 parse.
def test_l3_document_missing_markers_skipped_with_generic_reason():
    text = "Some unrelated memo with no recognisable structure at all."
    with pytest.raises(NotAnL3DocumentError) as excinfo:
        parse_l3_text(text, source_name="memo.pdf")
    assert "not an l3" in str(excinfo.value).lower()


# ---------------------------------------------------------------------------
# L5 (LLP Statement of Account) parser -- llp_statement.py.
# parse_l5_words() is PURE and coordinate-based (not text-based): every
# fixture below is a synthetic, self-invented list of word dicts shaped
# like pdfplumber's page.extract_words() output ({"text","x0","x1","top",
# "bottom"}), laid out to mimic the real document's two-column
# (CAPITAL ACCOUNT / CURRENT ACCOUNT) layout and its confirmed traps --
# see llp_statement.py's module docstring. Never a real document; every
# figure here is invented.
# ---------------------------------------------------------------------------

def _l5_word(text, x0, x1, top, bottom=None):
    return {
        "text": text,
        "x0": x0,
        "x1": x1,
        "top": top,
        "bottom": bottom if bottom is not None else top + 10.0,
    }


def _l5_word_line(text, x0, top):
    """Lay a text string out left-to-right as whitespace-separated word
    tokens, starting at (x0, top). Used for metadata/footer lines and for
    dispatch-rejection fixtures, where exact geometry is not load-bearing."""
    words = []
    x = x0
    for tok in text.split(" "):
        w = 6.5 * len(tok)
        words.append(_l5_word(tok, x, x + w, top))
        x += w + 4.0
    return words


def _words_from_lines(lines, start_top=20.0, x0=40.0, step=20.0):
    words = []
    top = start_top
    for line in lines:
        words.extend(_l5_word_line(line, x0, top))
        top += step
    return words


_L5_LABEL_X0 = 40.0
_L5_CAPITAL_X0 = 300.0
_L5_CURRENT_X0 = 420.0


def _l5_header_words(top=100.0):
    return [
        _l5_word("CAPITAL", 300.0, 345.0, top),
        _l5_word("ACCOUNT", 348.0, 395.0, top),
        _l5_word("CURRENT", 420.0, 463.0, top),
        _l5_word("ACCOUNT", 466.0, 513.0, top),
    ]
# The header above puts the derived boundary at (395.0 + 420.0) / 2 = 407.5
# and the label/value region start at 300.0 -- every row builder below
# stays clear of both, and tests that probe the boundary do so explicitly.


def _l5_row_words(top, label, capital_text=None, current_text=None):
    words = _l5_word_line(label, _L5_LABEL_X0, top)
    if capital_text is not None:
        w = 6.5 * len(capital_text)
        words.append(_l5_word(capital_text, _L5_CAPITAL_X0, _L5_CAPITAL_X0 + w, top))
    if current_text is not None:
        w = 6.5 * len(current_text)
        words.append(_l5_word(current_text, _L5_CURRENT_X0, _L5_CURRENT_X0 + w, top))
    return words


# A fully self-consistent, invented statement: every roll-forward and
# section sum below reconciles exactly, and current_transfer_to_capital
# (4,00,000) deliberately differs from capital_introduced_transferred
# (5,00,000) so the NOTE (never ERROR) about that comparison is exercised.
_DEFAULT_L5_ROWS = [
    ("Opening Balance as on 1 April, 2025", "10,00,000", "2,00,000"),
    ("ADDITIONS:-", None, None),
    ("Introduced/Transferred", "5,00,000", None),
    ("Interest on Capital", "1,00,000", None),
    ("Profit Share for the year", None, "18,00,000"),
    ("Remuneration", None, "24,00,000"),
    ("Total additions", "6,00,000", "42,00,000"),
    ("WITHDRAWALS:-", None, None),
    ("Drawings", "(2,00,000)", "(30,00,000)"),
    ("Transfer to Capital Account", None, "(4,00,000)"),
    ("Total withdrawals", "(2,00,000)", "(34,00,000)"),
    ("Closing Balance as on 31 March, 2026", "14,00,000", "10,00,000"),
]


def _l5_words(
    *,
    as_on="31 MARCH, 2026",
    assessment_year="2026-27",
    previous_year="31 MARCH, 2025",
    statement_date="05-Sep-2026",
    rows=None,
    extra_rows=None,
    include_metadata=True,
):
    """Build a synthetic L5 word list. `rows` (defaulting to
    _DEFAULT_L5_ROWS) is a list of (label, capital_text, current_text)
    tuples placed one per row below the header; `extra_rows` are appended
    after them (before the footer) without disturbing the defaults."""
    words: list[dict] = []
    top = 20.0
    if include_metadata:
        words += _l5_word_line("STATEMENT OF ACCOUNT OF A PARTNER", 40.0, top)
        top += 20.0
        words += _l5_word_line(f"AS ON {as_on}", 40.0, top)
        top += 20.0
        words += _l5_word_line(f"ASSESSMENT YEAR : {assessment_year}", 40.0, top)
        top += 20.0
        words += _l5_word_line(f"PREVIOUS YEAR : {previous_year}", 40.0, top)
        top += 20.0
        words += _l5_word_line(statement_date, 40.0, top)
        top += 30.0
    words += _l5_header_words(top=top)
    top += 20.0

    row_specs = _DEFAULT_L5_ROWS if rows is None else rows
    for label, capital_text, current_text in row_specs:
        words += _l5_row_words(top, label, capital_text, current_text)
        top += 20.0
    if extra_rows:
        for row in extra_rows:
            words += row(top) if callable(row) else _l5_row_words(top, *row)
            top += 20.0

    top += 20.0
    words += _l5_word_line("KPMG India Services LLP", 40.0, top)
    return words


# 1 -- the financial year is derived from "AS ON <d> MARCH, <yyyy>", and
# the printed ASSESSMENT YEAR is cross-checked against FY+1 -- agreement
# produces no ERROR.
def test_l5_fy_and_assessment_year_agree_no_error():
    words = _l5_words(as_on="31 MARCH, 2026", assessment_year="2026-27")
    record = parse_l5_words(words, source_name="stmt.pdf")
    assert record["financial_year"] == "2025-26"
    assert record["assessment_year"] == "2026-27"
    assert not any("ERROR" in d and "ASSESSMENT YEAR" in d for d in record["diagnostics"])


# 2 -- a disagreeing ASSESSMENT YEAR is a fail-loud ERROR diagnostic,
# never silently trusted or silently ignored.
def test_l5_fy_and_assessment_year_disagreement_is_fail_loud():
    words = _l5_words(as_on="31 MARCH, 2026", assessment_year="2099-00")
    record = parse_l5_words(words, source_name="stmt.pdf")
    assert record["financial_year"] == "2025-26"
    assert any(
        "ERROR" in d and "ASSESSMENT YEAR" in d for d in record["diagnostics"]
    )


# 3 -- capital and current closing balances are distinct fields, both
# exposed, never conflated into a single "closing balance".
def test_l5_distinct_capital_and_current_closing_balances_exposed():
    words = _l5_words()
    record = parse_l5_words(words, source_name="stmt.pdf")
    assert record["capital_closing_balance"] == 1400000.0
    assert record["current_closing_balance"] == 1000000.0
    assert record["capital_closing_balance"] != record["current_closing_balance"]


# 4 -- Interest on Capital is a separate labelled row here (unlike L1) and
# is pinned to the CAPITAL column only -- never the CURRENT column.
def test_l5_interest_on_capital_pinned_to_capital_column():
    words = _l5_words()
    record = parse_l5_words(words, source_name="stmt.pdf")
    assert record["capital_interest_on_capital"] == 100000.0
    assert "current_interest_on_capital" not in record


# 5 -- trap (a): a printed negative amount split across x-adjacent word
# tokens, including a split-off leading "(", still parses correctly --
# and this pins the SIGN of a withdrawal (never abs()-ed).
def test_l5_split_token_negative_amount_parses_and_keeps_sign():
    rows = list(_DEFAULT_L5_ROWS)
    # Replace the "Drawings" row with one whose capital figure arrives as
    # two adjacent word tokens: "(" then "2,00,000)" with a 0.5pt gap.
    idx = next(i for i, r in enumerate(rows) if r[0] == "Drawings")
    rows[idx] = ("Drawings", None, "(30,00,000)")
    words = _l5_words(rows=rows)
    # Manually splice in the split-token capital figure for Drawings at
    # the same top as the row emitted above (12 rows in, 20pt apart,
    # header + metadata offsets mirrored from _l5_words()).
    drawings_top = 20.0 * 5 + 30.0 + 20.0 + 20.0 * idx
    words = [w for w in words if not (w["top"] == drawings_top and w["x0"] >= _L5_CAPITAL_X0 and w["x0"] < _L5_CURRENT_X0)]
    words.append(_l5_word("(", 300.0, 303.0, drawings_top))
    words.append(_l5_word("2,00,000)", 303.5, 340.0, drawings_top))
    record = parse_l5_words(words, source_name="stmt.pdf")
    assert record["capital_drawings"] == -200000.0
    assert record["current_drawings"] == -3000000.0


# 6 -- trap (b): a row's label and amount tops can disagree by ~1pt; a
# +/-2pt tolerance still groups them into one row rather than losing the
# amount or splitting a spurious extra row.
def test_l5_top_tolerance_groups_label_and_amount_despite_1pt_offset():
    words = _l5_words()
    header_top = 20.0 * 5 + 30.0
    row_top = header_top + 20.0 + 20.0 * len(_DEFAULT_L5_ROWS)
    label_words = _l5_word_line("Signing Fee Reimbursement", _L5_LABEL_X0, row_top)
    amount_word = _l5_word("50,000", _L5_CURRENT_X0, _L5_CURRENT_X0 + 40.0, row_top + 1.3)
    words = words + label_words + [amount_word]
    record = parse_l5_words(words, source_name="stmt.pdf")
    assert any("Signing Fee Reimbursement" in u for u in record["unknown_labels"])


# 7 -- trap (c): a printed "-" means the figure is 0.0 (a nil that WAS
# printed) -- distinct from a label that never appears at all (None).
def test_l5_printed_dash_is_nil_absent_label_is_none():
    rows = list(_DEFAULT_L5_ROWS)
    idx = next(i for i, r in enumerate(rows) if r[0] == "Introduced/Transferred")
    rows[idx] = ("Introduced/Transferred", "-", None)
    words = _l5_words(rows=rows)
    record = parse_l5_words(words, source_name="stmt.pdf")
    assert record["capital_introduced_transferred"] == 0.0
    # "Interest on Capital" never carries a current-column figure anywhere
    # in this fixture -- that stays genuinely absent (None), never 0.0.
    assert record.get("current_interest_on_capital") is None


# 8 -- the CAPITAL/CURRENT column boundary is derived at runtime from the
# header row's own coordinates, never hardcoded: a value token placed
# just left of the derived midpoint (407.5) lands in CAPITAL, and one
# placed just right of it lands in CURRENT.
def test_l5_column_boundary_is_derived_from_header_not_hardcoded():
    words = _l5_header_words(top=100.0)
    words += _l5_word_line("Drawings", _L5_LABEL_X0, 120.0)
    words.append(_l5_word("(1,00,000)", 380.0, 406.0, 120.0))  # mid 393 < 407.5
    words.append(_l5_word("(2,00,000)", 409.0, 440.0, 120.0))  # mid 424.5 > 407.5
    words += _l5_word_line("STATEMENT OF ACCOUNT OF A PARTNER AS ON 31 MARCH, 2026", 40.0, 20.0)
    record = parse_l5_words(words, source_name="stmt.pdf")
    assert record["capital_drawings"] == -100000.0
    assert record["current_drawings"] == -200000.0


# 9 -- a non-reconciling ADDITIONS section (row sum vs printed
# "Total additions") is a fail-loud ERROR, never a silent plug.
def test_l5_non_reconciling_additions_section_is_fail_loud():
    rows = list(_DEFAULT_L5_ROWS)
    idx = next(i for i, r in enumerate(rows) if r[0] == "Total additions")
    rows[idx] = ("Total additions", "99,99,999", "42,00,000")  # deliberately wrong
    words = _l5_words(rows=rows)
    record = parse_l5_words(words, source_name="stmt.pdf")
    assert any(
        "ERROR" in d and "capital" in d and "additions" in d
        for d in record["diagnostics"]
    )


# 10 -- a non-reconciling WITHDRAWALS section is likewise fail-loud.
def test_l5_non_reconciling_withdrawals_section_is_fail_loud():
    rows = list(_DEFAULT_L5_ROWS)
    idx = next(i for i, r in enumerate(rows) if r[0] == "Total withdrawals")
    rows[idx] = ("Total withdrawals", "(2,00,000)", "(99,99,999)")  # wrong
    words = _l5_words(rows=rows)
    record = parse_l5_words(words, source_name="stmt.pdf")
    assert any(
        "ERROR" in d and "current" in d and "withdrawals" in d
        for d in record["diagnostics"]
    )


# 11 -- an unrecognised row inside a section still contributes to that
# section's sum AND is separately reported in unknown_labels -- never
# silently dropped, never silently excluded from reconciliation.
def test_l5_unknown_row_in_section_still_sums_and_is_reported():
    rows = list(_DEFAULT_L5_ROWS)
    add_idx = next(i for i, r in enumerate(rows) if r[0] == "Total additions")
    rows.insert(add_idx, ("Bonus Contribution", "50,000", None))
    total_idx = next(i for i, r in enumerate(rows) if r[0] == "Total additions")
    rows[total_idx] = ("Total additions", "6,50,000", "42,00,000")
    # The capital column's roll-forward must still reconcile: opening
    # 10,00,000 + additions 6,50,000 - withdrawals 2,00,000 = 14,50,000.
    close_idx = next(
        i for i, r in enumerate(rows) if r[0] == "Closing Balance as on 31 March, 2026"
    )
    rows[close_idx] = ("Closing Balance as on 31 March, 2026", "14,50,000", "10,00,000")
    words = _l5_words(rows=rows)
    record = parse_l5_words(words, source_name="stmt.pdf")
    assert any("Bonus Contribution" in u for u in record["unknown_labels"])
    assert not any(
        "ERROR" in d and "capital" in d and "additions" in d
        for d in record["diagnostics"]
    )


# 12 -- both roll-forwards (opening + additions + withdrawals == closing)
# reconcile for a fully self-consistent statement -- no ERROR at all.
def test_l5_fully_reconciling_statement_has_no_error_diagnostics():
    words = _l5_words()
    record = parse_l5_words(words, source_name="stmt.pdf")
    assert not any(d.startswith("ERROR") for d in record["diagnostics"])
    assert any("NOTE" in d for d in record["diagnostics"])  # the transfer-vs-introduced NOTE


# 13 -- an L1 monthly payout certificate is skipped cleanly, never
# misparsed as an L5 statement.
def test_l5_non_l5_l1_certificate_is_skipped_not_misparsed():
    words = _words_from_lines(
        ["To Whomsoever It may concern", "Remuneration 1,20,000"]
    )
    with pytest.raises(NotAnL5DocumentError):
        parse_l5_words(words, source_name="jan26.pdf")


# 14 -- an L2 salary statement is skipped cleanly.
def test_l5_non_l5_salary_statement_is_skipped_not_misparsed():
    words = _words_from_lines(
        ["SALARY STATEMENT FOR the month of March 2026", "Net Pay 2,00,000"]
    )
    with pytest.raises(NotAnL5DocumentError) as excinfo:
        parse_l5_words(words, source_name="salary_march26.pdf")
    assert "salary statement" in str(excinfo.value).lower()


# 15 -- an L3 Compensation Advisory letter is skipped cleanly.
def test_l5_non_l5_l3_advisory_is_skipped_not_misparsed():
    words = _words_from_lines(
        [
            "Compensation Advisory",
            "For the year ended 31 March 2026",
            "PAYMENTS",
            "Net Payable 10,00,000",
            "SCHEDULE",
            "Opening Balance 20,00,000",
        ]
    )
    with pytest.raises(NotAnL5DocumentError):
        parse_l5_words(words, source_name="advisory.pdf")


# 16 -- a document missing both the header row and the statement-phrase
# is skipped with a generic, but still specific, reason -- never crashes,
# never guesses an L5 parse.
def test_l5_document_missing_markers_skipped_with_generic_reason():
    words = _words_from_lines(["Some unrelated memo with no recognisable structure at all."])
    with pytest.raises(NotAnL5DocumentError) as excinfo:
        parse_l5_words(words, source_name="memo.pdf")
    assert "not an l5" in str(excinfo.value).lower()


# ---------------------------------------------------------------------------
# s.6.2 -- the "=" trap. A label that happens to start with "=" must never
# be silently stored as an Excel formula.
# ---------------------------------------------------------------------------

def test_text_helper_escapes_a_leading_equals_sign():
    assert writer._text("=SUM(A1:A2)") == " =SUM(A1:A2)"
    assert writer._text("Plain label") == "Plain label"
    assert writer._text(1234) == 1234


_FORMULA_START = re.compile(r"^[A-Za-z_$(+-]")


def test_saved_workbook_has_no_accidental_formula_cells(tmp_path):
    """Every <f> element in the saved workbook's sheet XML is a genuine
    formula this package wrote on purpose (all of which start with a
    function name, a cell reference, '$', '(' or a sign) -- never a plain
    label that merely began with '=' and got stored as a formula by
    accident (the "=" trap; see writer.py's module docstring)."""
    data = _load_fixture()
    report = build_report(data)
    out_path = tmp_path / "workbook.xlsx"
    writer.write_report_workbook(report, str(out_path))

    with zipfile.ZipFile(out_path) as zf:
        sheet_names = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")]
        assert sheet_names
        for name in sheet_names:
            xml_bytes = zf.read(name)
            root = ET.fromstring(xml_bytes)
            for f_el in root.iter():
                if f_el.tag.endswith("}f") or f_el.tag == "f":
                    formula_text = (f_el.text or "").strip()
                    assert formula_text, f"empty formula element in {name}"
                    assert _FORMULA_START.match(formula_text), (
                        f"suspicious formula in {name}: {formula_text!r} -- "
                        "does not look like a genuine formula; check for a "
                        "label that started with '=' and was not escaped"
                    )


# ---------------------------------------------------------------------------
# End-to-end via agent.run() -- the public entry point.
# ---------------------------------------------------------------------------

_ALL_SHEETS = [
    "Logic", "Drivers", "Monthly grid", "One-offs", "Cohorts", "Capital",
    "Reconciliation", "Exceptions", "Open items",
]


def test_agent_run_end_to_end_writes_all_sheets(tmp_path):
    out_path = tmp_path / "out.xlsx"
    result = run(input_path=str(FIXTURE_PATH), output_path=str(out_path))

    assert "ERROR" not in result
    assert out_path.is_file()

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    for sheet in _ALL_SHEETS:
        assert sheet in wb.sheetnames
    # The fixture supplies no payroll rows -- that sheet must be omitted
    # entirely, not written empty.
    assert "Payroll stream" not in wb.sheetnames


def test_agent_run_reports_rate_change_warning_in_summary(tmp_path):
    out_path = tmp_path / "out.xlsx"
    result = run(input_path=str(FIXTURE_PATH), output_path=str(out_path))
    assert "rate change" in result.lower()


def test_agent_run_missing_file_returns_error_string_not_an_exception():
    result = run(input_path="does-not-exist.yaml", output_path="whatever.xlsx")
    assert result.startswith("ERROR")


def test_agent_run_unsupported_extension_returns_error_string(tmp_path):
    bad = tmp_path / "input.txt"
    bad.write_text("financial_year: '2025-26'", encoding="utf-8")
    result = run(input_path=str(bad), output_path=str(tmp_path / "out.xlsx"))
    assert result.startswith("ERROR")


# ---------------------------------------------------------------------------
# H3.5 -- manifest reshape: skill.yaml now takes the raw documents
# directly. `input_path` is demoted to a test-only run() parameter (used
# throughout this file above) and must never render in the UI; every
# optional document-backed leg must degrade to an explicit "not
# available" note rather than a zero/default; a missing REQUIRED input
# must fail loud, naming it, before any parsing is attempted.
# ---------------------------------------------------------------------------

def _skill_yaml_inputs() -> list[dict]:
    skill_yaml_path = (
        Path(__file__).resolve().parent.parent
        / "src" / "agents" / "skill_partner_comp_recon" / "skill.yaml"
    )
    manifest = yaml.safe_load(skill_yaml_path.read_text(encoding="utf-8"))
    return manifest["inputs"]


def test_input_path_absent_from_manifest_but_run_still_accepts_it():
    names = {inp["name"] for inp in _skill_yaml_inputs()}
    assert "input_path" not in names, (
        "input_path must be demoted to a test-only run() parameter -- it "
        "must never appear in skill.yaml's inputs: (else it renders in the UI)"
    )
    # run() must still accept it -- every e2e test above this section calls
    # run(input_path=...) and depends on that continuing to work.
    import inspect

    assert "input_path" in inspect.signature(run).parameters


def _advices_dir_with_one_pdf(tmp_path: Path) -> Path:
    advices_dir = tmp_path / "advices"
    advices_dir.mkdir()
    (advices_dir / "jan.pdf").write_bytes(b"%PDF-1.4 not a real pdf")
    return advices_dir


def test_missing_required_input_fails_loud_and_names_it(tmp_path):
    advices_dir = _advices_dir_with_one_pdf(tmp_path)
    advisory_path = tmp_path / "advisory.pdf"
    advisory_path.write_bytes(b"%PDF-1.4 not a real pdf")

    # entity missing
    result = run(
        entity="",
        advices_dir=str(advices_dir),
        advisory_path=str(advisory_path),
        output_path=str(tmp_path / "out.xlsx"),
    )
    assert result.startswith("ERROR")
    assert "entity" in result

    # advices_dir missing
    result = run(
        entity="Harshal",
        advices_dir="",
        advisory_path=str(advisory_path),
        output_path=str(tmp_path / "out.xlsx"),
    )
    assert result.startswith("ERROR")
    assert "advices_dir" in result

    # advisory_path missing
    result = run(
        entity="Harshal",
        advices_dir=str(advices_dir),
        advisory_path="",
        output_path=str(tmp_path / "out.xlsx"),
    )
    assert result.startswith("ERROR")
    assert "advisory_path" in result


def test_optional_inputs_absent_degrade_to_not_available_never_zero_or_default(tmp_path):
    advices_dir = _advices_dir_with_one_pdf(tmp_path)
    advisory_path = tmp_path / "advisory.pdf"
    advisory_path.write_bytes(b"%PDF-1.4 not a real pdf")

    result = run(
        entity="Harshal",
        advices_dir=str(advices_dir),
        advisory_path=str(advisory_path),
        llp_statement="",
        gnucash_path="",
        xlsx_26as="",
        output_path=str(tmp_path / "out.xlsx"),
    )
    # The required legs still hard-fail here (the garbage-byte advisory.pdf
    # is unparseable, not a real Advisory -- advisory.py itself is no
    # longer a Stage 2 placeholder), but the optional-leg status notes must
    # appear regardless, and every one of them must say "not available" --
    # never a zero, a blank, or a fabricated figure.
    lowered = result.lower()
    assert "llp statement of account: not available" in lowered
    assert "gnucash books tie-out: not available" in lowered
    assert "26as tds-credit tie-out: not available" in lowered
    assert "not available (no document supplied)" in lowered  # llp_statement
    assert "not available (no book supplied)" in lowered  # gnucash_path
    assert "not available (no workbook supplied)" in lowered  # xlsx_26as


# ---------------------------------------------------------------------------
# Stage 1b -- jv_emitter.py: the GnuCash multi-split journal CSV.
# ---------------------------------------------------------------------------

def _read_journal_csv(path: Path) -> list[dict]:
    import csv

    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _build_fixture_journal_rows(tmp_path: Path) -> list[dict]:
    data = _load_fixture()
    report = build_report(data)
    journals = build_journals(report, data["accounts"])
    out_path = tmp_path / "journal.csv"
    write_journal_csv(journals, str(out_path))
    return _read_journal_csv(out_path)


# 3.1 -- header is exactly the 7 columns, in order.
def test_journal_csv_header_exact(tmp_path):
    rows = _build_fixture_journal_rows(tmp_path)
    assert rows  # sanity: fixture must produce at least one row
    with (tmp_path / "journal.csv").open(encoding="utf-8") as f:
        header_line = f.readline().strip()
    assert header_line.split(",") == JOURNAL_HEADERS
    assert JOURNAL_HEADERS == [
        "Date", "Transaction ID", "Number", "Description", "Account",
        "Amount", "Currency",
    ]


# 3.2 -- no cell in Date/Transaction ID/Number/Description is ever blank
# (regression guard for the blank-continuation-row parse-error trap).
def test_journal_csv_never_has_blank_transaction_fields(tmp_path):
    rows = _build_fixture_journal_rows(tmp_path)
    for row in rows:
        assert row["Date"].strip(), row
        assert row["Transaction ID"].strip(), row
        assert row["Number"].strip(), row
        assert row["Description"].strip(), row


# 3.3 -- rows sharing a Transaction ID share Date/Number/Description, and
# their Amounts sum to 0.00 within half a paisa.
def test_journal_csv_transactions_are_internally_consistent_and_balanced(tmp_path):
    rows = _build_fixture_journal_rows(tmp_path)
    by_txn: dict[str, list[dict]] = {}
    for row in rows:
        by_txn.setdefault(row["Transaction ID"], []).append(row)

    for txn_id, txn_rows in by_txn.items():
        dates = {r["Date"] for r in txn_rows}
        numbers = {r["Number"] for r in txn_rows}
        descriptions = {r["Description"] for r in txn_rows}
        assert len(dates) == 1, f"{txn_id}: mixed Date values {dates}"
        assert len(numbers) == 1, f"{txn_id}: mixed Number values {numbers}"
        assert len(descriptions) == 1, f"{txn_id}: mixed Description values {descriptions}"
        assert numbers == {txn_id}

        total = sum(float(r["Amount"]) for r in txn_rows)
        assert total == pytest.approx(0.0, abs=0.005), f"{txn_id}: splits sum to {total}"


# 3.4 -- Transaction IDs are unique per transaction, all carry the FY
# prefix, and no ID is a prefix of another (so GnuCash's grouping can never
# fuse two distinct transactions).
def test_journal_csv_transaction_ids_unique_fy_prefixed_and_not_colliding(tmp_path):
    data = _load_fixture()
    fy_pfx = fy_prefix(data["financial_year"])
    rows = _build_fixture_journal_rows(tmp_path)
    txn_ids = sorted({row["Transaction ID"] for row in rows})
    assert len(txn_ids) >= 2  # at least the opening reclass + one month

    for txn_id in txn_ids:
        assert txn_id.startswith(fy_pfx + "-"), txn_id

    for a in txn_ids:
        for b in txn_ids:
            if a != b:
                assert not b.startswith(a), f"{a!r} is a prefix of {b!r}"


# 3.5 -- Currency is INR on every row; every Date matches YYYY-MM-DD.
def test_journal_csv_currency_and_date_format(tmp_path):
    rows = _build_fixture_journal_rows(tmp_path)
    date_re = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    for row in rows:
        assert row["Currency"] == "INR"
        assert date_re.match(row["Date"]), row["Date"]


# 3.6 -- no Account starts with "Root Account:"; every Account appears in
# the fixture's own accounts: block (no invented/typo'd account leaks in).
def test_journal_csv_accounts_have_no_root_prefix_and_are_all_known(tmp_path):
    data = _load_fixture()
    known_accounts = set(data["accounts"].values())
    rows = _build_fixture_journal_rows(tmp_path)
    for row in rows:
        assert not row["Account"].startswith("Root Account:"), row["Account"]
        assert row["Account"] in known_accounts, row["Account"]


# 3.7 -- no Amount is ever "0.00" (zero-value splits must be omitted, not
# emitted as a no-op row).
def test_journal_csv_never_emits_a_zero_amount_row(tmp_path):
    rows = _build_fixture_journal_rows(tmp_path)
    for row in rows:
        assert row["Amount"] != "0.00", row


# 3.8 -- no Transfer Amount / Transfer Account column is ever emitted.
def test_journal_csv_has_no_transfer_columns(tmp_path):
    rows = _build_fixture_journal_rows(tmp_path)
    assert rows
    for row in rows:
        assert "Transfer Amount" not in row
        assert "Transfer Account" not in row


# 3.9 -- the firm's tax never appears as its own leg, and the share-of-
# profit credit for a month is the NET figure (gross + firms_tax +
# additional), never the gross figure.
def test_journal_csv_share_of_profit_is_net_of_firms_tax(tmp_path):
    data = _load_fixture()
    report = build_report(data)
    journals = build_journals(report, data["accounts"])

    firms_tax_account = data["accounts"]["tds_expense"]  # sanity: distinct key
    share_account = data["accounts"]["share_of_profit_income"]

    # Firm's tax must never be booked to any account under its own name --
    # there is no accounts.firms_tax key at all, so this is really just
    # confirming ACCOUNT_KEYS has no such entry and no journal invents one.
    for j in journals:
        for s in j.splits:
            assert s.account != "firms_tax", s

    by_month = {m.month: m for m in report.monthly}
    april = by_month["2025-04"]
    expected_net = (
        april.share_of_profit_gross + april.firms_tax + april.additional_share_of_profit
    )

    april_journal = next(j for j in journals if j.txn_id.endswith("-M01"))
    share_split = next(s for s in april_journal.splits if s.account == share_account)
    # Credit leg -> stored as .credit, positive.
    assert share_split.credit == pytest.approx(expected_net, abs=0.01)
    assert share_split.debit == 0.0


# 3.9b -- prior_cohort_drawdown is POSITIVE (a prior-year incentive
# instalment RECEIVED this year) and must produce a CREDIT (negative
# Amount) on the current_account split equal to -prior_cohort_drawdown,
# with no income account carrying any split attributable to it. A
# balance-only check is not sufficient to pin this direction -- the
# transaction still balances even with the wrong sign, which is exactly
# how the wrong sign shipped once already.
def test_journal_csv_prior_cohort_drawdown_credits_current_account(tmp_path):
    data = _load_fixture()
    report = build_report(data)
    journals = build_journals(report, data["accounts"])

    by_month = {m.month: m for m in report.monthly}
    april = by_month["2025-04"]
    assert april.prior_cohort_drawdown > 0, "fixture must exercise the positive case"

    current_account = data["accounts"]["current_account"]
    income_accounts = {
        data["accounts"]["remuneration_income"],
        data["accounts"]["share_of_profit_income"],
        data["accounts"]["interest_on_capital"],
    }

    april_journal = next(j for j in journals if j.txn_id.endswith("-M01"))
    drawdown_split = next(s for s in april_journal.splits if s.account == current_account)

    # Credit leg: stored as .credit (positive), .debit is 0 -- and the CSV
    # row's signed Amount must be negative (Cr).
    assert drawdown_split.debit == 0.0
    assert drawdown_split.credit == pytest.approx(april.prior_cohort_drawdown, abs=0.01)

    out_path = tmp_path / "journal_drawdown.csv"
    write_journal_csv(journals, str(out_path))
    csv_rows = _read_journal_csv(out_path)
    april_rows = [r for r in csv_rows if r["Transaction ID"].endswith("-M01")]
    current_account_rows = [r for r in april_rows if r["Account"] == current_account]
    assert len(current_account_rows) == 1
    amount = float(current_account_rows[0]["Amount"])
    assert amount == pytest.approx(-april.prior_cohort_drawdown, abs=0.01)
    assert amount < 0, "prior_cohort_drawdown must book as a CREDIT (negative Amount)"

    # No income account may carry a split whose value traces to the
    # drawdown -- i.e. no income-account split in April's transaction
    # equals the drawdown amount (which would indicate double-counting it
    # as income).
    for row in april_rows:
        if row["Account"] in income_accounts:
            assert abs(float(row["Amount"])) != pytest.approx(
                april.prior_cohort_drawdown, abs=0.01
            ), row


# 3.10 -- the opening reclass entry is present, dated as given, carries the
# "<FY>-RECT" id, and balances.
def test_journal_csv_opening_reclass_present_and_balanced(tmp_path):
    data = _load_fixture()
    fy_pfx = fy_prefix(data["financial_year"])
    report = build_report(data)
    journals = build_journals(report, data["accounts"])

    rect = next(j for j in journals if j.txn_id == f"{fy_pfx}-RECT")
    assert rect.date == data["opening_reclass"]["date"]
    assert rect.balanced
    assert rect.total_debit == pytest.approx(218650, abs=0.01)
    assert rect.total_credit == pytest.approx(218650, abs=0.01)


# 3.11 -- a missing required account key produces an "ERROR: ..." string
# from run(), never a traceback, and names the key.
def test_missing_account_key_returns_error_string_not_exception(tmp_path):
    import yaml

    data = _load_fixture()
    del data["accounts"]["interest_on_capital"]  # April's line uses this key
    bad_input = tmp_path / "bad_input.yaml"
    bad_input.write_text(yaml.safe_dump(data), encoding="utf-8")

    out_path = tmp_path / "out.xlsx"
    journal_path = tmp_path / "journal.csv"
    result = run(
        input_path=str(bad_input), output_path=str(out_path),
        journal_path=str(journal_path),
    )
    assert result.startswith("ERROR")
    assert "interest_on_capital" in result
    assert not journal_path.exists()


# 3.12 -- with journal_path="" the workbook output is byte-identical to a
# run that never mentions Stage 1b at all (no CSV written, no behaviour
# change).
def test_empty_journal_path_leaves_workbook_output_unchanged(tmp_path):
    out_default = tmp_path / "out_default.xlsx"
    out_explicit_empty = tmp_path / "out_explicit_empty.xlsx"

    result_default = run(input_path=str(FIXTURE_PATH), output_path=str(out_default))
    result_explicit = run(
        input_path=str(FIXTURE_PATH), output_path=str(out_explicit_empty),
        journal_path="",
    )

    assert result_default == result_explicit.replace(
        str(out_explicit_empty), str(out_default)
    )
    # Compare workbook content, not raw bytes: openpyxl stamps docProps/core.xml
    # with a wall-clock created/modified timestamp on every save, so two
    # separate runs are never byte-identical even with identical inputs --
    # what "unchanged" actually means here is every sheet's content, which
    # this compares directly, ignoring only that timestamp-bearing part.
    with zipfile.ZipFile(out_default) as zf_a, zipfile.ZipFile(out_explicit_empty) as zf_b:
        names_a = {n for n in zf_a.namelist() if n != "docProps/core.xml"}
        names_b = {n for n in zf_b.namelist() if n != "docProps/core.xml"}
        assert names_a == names_b
        for name in sorted(names_a):
            assert zf_a.read(name) == zf_b.read(name), f"content differs in {name}"
    assert list(tmp_path.glob("*.csv")) == []  # journal_path unset -> no CSV written at all
