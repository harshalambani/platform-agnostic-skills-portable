"""
tests/skill_itr_workbook/test_advance_tax.py -- required PR1 tests for
scripts/advance_tax.py (2026-08 advance-tax-estimator work):

  (i)   year-end-unchanged regression: schedules.py's own compute path is
        untouched by the tax_core extraction.
  (ii)  anti-drift: advance_tax.build_estimate(...) and
        schedules.build_all_schedules(...) produce the SAME tax liability
        on identical full-year data, for both regimes.
  (iii) an unprojected head stays blank and flagged, never guessed.
  (iv)  a capital-gains item with no sale_date is flagged and gets NO
        s.234C first-proviso relief in any instalment.
  (v)   manual mode works with no book/resolved/node_by_guid at all.

Fully offline; synthetic fixtures only (fixture_gen.py + the existing
syn_ind.* fixtures already used by test_schedules.py). No access to any
Data/ path, any *.gnucash file outside tests/fixtures, or the live
PortableApps store.
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPTS = ROOT / "src" / "agents" / "skill_itr_workbook" / "scripts"
FIXTURES = Path(__file__).resolve().parent / "fixtures"
RULES_DIR = ROOT / "bundling" / "canonical" / "itr" / "rules"

for p in (str(SCRIPTS), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

import parse_eguile as pe  # noqa: E402
import parse_gnucash as pg  # noqa: E402
import configs  # noqa: E402
import mapping as mapping_engine  # noqa: E402
import rules as rules_engine  # noqa: E402
import schedules as sch  # noqa: E402
import fixture_gen  # noqa: E402
import advance_tax as adv  # noqa: E402

YEAR_KEY = "2024-25"


@pytest.fixture(scope="module")
def syn_ind_resolved():
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    result = mapping_engine.resolve_tree(tree, loaded)
    assert not result.blocked
    return tree, result.resolved


@pytest.fixture(scope="module")
def syn_ind_book():
    return pg.parse_book(FIXTURES / "syn_ind.gnucash")


@pytest.fixture(scope="module")
def entity_and_scrips():
    entities = configs.load_entities(ROOT / "bundling" / "canonical" / "itr" / "entities.example.yaml")
    scrips = configs.load_scrips(ROOT / "bundling" / "canonical" / "itr" / "scrips.example.yaml")
    return entities["SYN-IND"], scrips


@pytest.fixture(scope="module")
def rules():
    return rules_engine.load_rules(RULES_DIR, YEAR_KEY)


@pytest.fixture(scope="module")
def fmv_tables():
    return sch.load_fmv_tables()


# ---------------------------------------------------------------------------
# (i) year-end-unchanged regression
# ---------------------------------------------------------------------------

def test_year_end_compute_tax_still_matches_tax_core(rules):
    # schedules.compute_tax is a thin re-export of tax_core.compute_tax_on --
    # the extraction must not have changed year-end behavior at all.
    import tax_core

    fy_end = date(2025, 3, 31)
    a = sch.compute_tax(
        normal_income=1_000_000.0, special_rate_tax=15_000.0,
        special_rate_income_amount=200_000.0, rules=rules, regime="new",
        status="Individual", dob="1990-01-01", fy_end=fy_end,
    )
    b = tax_core.compute_tax_on(
        normal_income=1_000_000.0, special_rate_tax=15_000.0,
        special_rate_income_amount=200_000.0, rules=rules, regime="new",
        status="Individual", dob="1990-01-01", fy_end=fy_end,
    )
    assert a == b
    assert a.tax_liability > 0


# ---------------------------------------------------------------------------
# (ii) anti-drift: estimator and year-end workbook must agree
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("regime", ["new", "old"])
def test_estimator_matches_year_end_tax_liability(
    syn_ind_resolved, syn_ind_book, entity_and_scrips, rules, fmv_tables, regime,
):
    tree, resolved = syn_ind_resolved
    entity, scrips = entity_and_scrips
    node_by_guid = sch._node_by_guid(tree)
    fy_end = date(2025, 3, 31)

    # Year-end path: form16=None so both engines read the identical
    # book-derived salary figure (form16 handling is out of scope for the
    # estimator's from_book() populator, which never reads a Form 16).
    model = sch.build_all_schedules(
        tree, resolved, syn_ind_book, None, YEAR_KEY, rules, regime,
        entity.status, entity.dob, scrips, fmv_tables, residency=entity.residency,
    )
    year_end_liability = model.computation.tax_block.tax_liability

    est_input = adv.from_book(
        resolved, node_by_guid, syn_ind_book, YEAR_KEY, rules,
        entity_name=entity.name, regime=regime, status=entity.status, dob=entity.dob,
        scrips=scrips, fmv_tables=fmv_tables, residency=entity.residency, as_of=fy_end,
    )
    # No VIA deductions in this fixture's book, so the estimator's
    # documented old-regime-only VIA simplification cannot introduce drift
    # here either way.
    est_input.deductions_via = 0.0
    estimate = adv.build_estimate(est_input, rules)
    regime_estimate = estimate.new_regime if regime == "new" else estimate.old_regime

    assert regime_estimate.tax_liability == pytest.approx(year_end_liability, abs=0.01)


# ---------------------------------------------------------------------------
# (iii) unprojected head stays blank and flagged
# ---------------------------------------------------------------------------

def test_unprojected_head_stays_blank_and_flagged():
    inp = adv.EstimateInput(
        entity_name="SYN-TEST", fy=YEAR_KEY, regime="new",
        salary=adv.ProjectedHead(actual_to_date=400_000.0),  # no projected_remainder
        remuneration=adv.ProjectedHead(actual_to_date=0.0, projected_remainder=0.0),
        rent=adv.ProjectedHead(actual_to_date=0.0, projected_remainder=0.0),
        interest=adv.ProjectedHead(actual_to_date=0.0, projected_remainder=0.0),
    )
    assert inp.salary.unprojected
    assert inp.salary.full_year == 400_000.0  # never guesses the remainder
    assert inp.unprojected_heads() == ["salary"]

    rules_cfg = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    estimate = adv.build_estimate(inp, rules_cfg)
    assert estimate.unprojected_heads == ["salary"]


def test_annualise_only_fills_genuinely_blank_heads():
    inp = adv.EstimateInput(
        entity_name="SYN-TEST", fy=YEAR_KEY, regime="new", as_of=date(2024, 10, 1),
        salary=adv.ProjectedHead(actual_to_date=300_000.0),
        remuneration=adv.ProjectedHead(actual_to_date=50_000.0, projected_remainder=999.0),
    )
    out = adv.annualise(inp)
    assert not out.salary.unprojected
    assert out.salary.projected_remainder > 0
    # already-filled head must not be touched by annualise()
    assert out.remuneration.projected_remainder == 999.0


# ---------------------------------------------------------------------------
# (iv) undated CG item is flagged and gets no s.234C relief
# ---------------------------------------------------------------------------

def test_undated_capital_gain_flagged_and_no_234c_relief():
    dated = adv.CapitalGainItem(amount=100_000.0, gain_type="LTCG_112A", sale_date=date(2024, 8, 1))
    undated = adv.CapitalGainItem(amount=200_000.0, gain_type="STCG_111A", sale_date=None)
    inp = adv.EstimateInput(
        entity_name="SYN-TEST", fy=YEAR_KEY, regime="new",
        salary=adv.ProjectedHead(actual_to_date=0.0, projected_remainder=0.0),
        remuneration=adv.ProjectedHead(actual_to_date=0.0, projected_remainder=0.0),
        rent=adv.ProjectedHead(actual_to_date=0.0, projected_remainder=0.0),
        interest=adv.ProjectedHead(actual_to_date=0.0, projected_remainder=0.0),
        capital_gains=[dated, undated],
    )
    assert undated.undated
    assert inp.undated_capital_gains() == [undated]

    rules_cfg = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    estimate = adv.build_estimate(inp, rules_cfg)
    assert estimate.undated_capital_gains == [undated]
    assert estimate.undated_cg_amount == 200_000.0

    excluded, any_undated = adv._cg_proviso_exclusions_by_item(inp.capital_gains, rules_cfg, YEAR_KEY)
    assert any_undated is True
    # The undated STCG item's tax must never be excluded from ANY
    # instalment's base -- i.e. it gets zero relief throughout, the
    # conservative (never a silent "assumed Q1") direction.
    stcg_rate = adv._cg_rates(rules_cfg)[1]
    undated_item_tax = undated.amount * stcg_rate
    for excl in excluded:
        assert excl <= dated.amount * adv._cg_rates(rules_cfg)[0] + 0.01
        assert excl < undated_item_tax + dated.amount * adv._cg_rates(rules_cfg)[0]


# ---------------------------------------------------------------------------
# (v) manual mode works with no book present
# ---------------------------------------------------------------------------

def test_manual_mode_works_with_no_book_at_all():
    inp = adv.EstimateInput(
        entity_name="Manual Entry, No Book", fy=YEAR_KEY, regime="new", status="Individual",
        dob="1985-05-05", as_of=date(2024, 11, 1),
        salary=adv.ProjectedHead(actual_to_date=600_000.0, projected_remainder=200_000.0),
        remuneration=adv.ProjectedHead(actual_to_date=0.0, projected_remainder=0.0),
        rent=adv.ProjectedHead(actual_to_date=0.0, projected_remainder=0.0),
        interest=adv.ProjectedHead(actual_to_date=5_000.0, projected_remainder=2_000.0),
        capital_gains=[adv.CapitalGainItem(amount=50_000.0, gain_type="LTCG_112A", sale_date=date(2024, 9, 1))],
        tds_to_date=30_000.0,
        advance_tax_paid=[(date(2024, 6, 15), 10_000.0)],
    )
    rules_cfg = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    estimate = adv.build_estimate(inp, rules_cfg)

    assert estimate.new_regime.tax_liability >= 0
    assert estimate.old_regime.tax_liability >= 0
    assert estimate.unprojected_heads == []
    assert estimate.undated_capital_gains == []


def test_save_load_roundtrip_manual_mode(tmp_path):
    inp = adv.EstimateInput(
        entity_name="Roundtrip Test", fy=YEAR_KEY, regime="old",
        salary=adv.ProjectedHead(actual_to_date=100_000.0),
        capital_gains=[adv.CapitalGainItem(amount=1_000.0, gain_type="OTHER", sale_date=None, note="test note")],
        advance_tax_paid=[(date(2024, 6, 15), 5_000.0)],
    )
    adv.save_estimate("rt", inp, data_root=tmp_path / "Data")
    loaded = adv.load_estimate("rt", data_root=tmp_path / "Data")

    assert loaded.entity_name == "Roundtrip Test"
    assert loaded.salary.actual_to_date == 100_000.0
    assert loaded.capital_gains[0].note == "test note"
    assert loaded.capital_gains[0].sale_date is None
    assert loaded.advance_tax_paid == [(date(2024, 6, 15), 5_000.0)]
    assert adv.list_estimates(data_root=tmp_path / "Data") == ["rt"]


# ---------------------------------------------------------------------------
# xlsx writer smoke test (live-formula instalment schedule)
# ---------------------------------------------------------------------------

def test_write_workbook_produces_formulas_not_baked_values(tmp_path):
    import write_advance_tax_workbook as wr

    rules_cfg = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    inp = adv.EstimateInput(
        entity_name="SYN-TEST", fy=YEAR_KEY, regime="new", as_of=date(2024, 9, 1),
        salary=adv.ProjectedHead(actual_to_date=500_000.0),
        remuneration=adv.ProjectedHead(actual_to_date=0.0, projected_remainder=0.0),
        rent=adv.ProjectedHead(actual_to_date=0.0, projected_remainder=0.0),
        interest=adv.ProjectedHead(actual_to_date=10_000.0, projected_remainder=10_000.0),
        capital_gains=[adv.CapitalGainItem(amount=200_000.0, gain_type="LTCG_112A", sale_date=date(2024, 8, 1))],
        tds_to_date=40_000.0,
        advance_tax_paid=[(date(2024, 6, 10), 20_000.0)],
    )
    estimate = adv.build_estimate(inp, rules_cfg)
    out_path = tmp_path / "advance_tax_estimate.xlsx"
    wr.write_workbook(estimate, out_path)
    assert out_path.is_file()

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    assert {"Computation (New)", "Computation (Old)", "Instalments (New)", "Instalments (Old)", "Flags"} <= set(wb.sheetnames)

    ws = wb["Instalments (New)"]
    formula_cells = [
        c.value for row in ws.iter_rows() for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]
    assert any("MAX(0," in f for f in formula_cells)
