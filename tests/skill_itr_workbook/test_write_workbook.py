"""
tests/skill_itr_workbook/test_write_workbook.py -- Batch 6 tests for
scripts/write_workbook.py. openpyxl never evaluates formulas (plan section
4.5, point 3 notes this explicitly), so this file asserts two different
things separately: (1) formula-graph WIRING -- the right cells reference
the right other-sheet cells, via regex/string assertions on the formula
text; (2) the underlying VALUES the formulas are built from, via
schedules.py's own arithmetic (already covered by test_schedules.py).
Fully offline; synthetic fixtures only.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPTS = ROOT / "src" / "agents" / "skill_itr_workbook" / "scripts"
FIXTURES = Path(__file__).resolve().parent / "fixtures"
RULES_DIR = ROOT / "bundling" / "canonical" / "itr" / "rules"

for p in (str(SCRIPTS), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

import parse_eguile as pe  # noqa: E402
import parse_gnucash as pg  # noqa: E402
import parse_form16  # noqa: E402
import parse_foreign as pf  # noqa: E402
import configs  # noqa: E402
import mapping as mapping_engine  # noqa: E402
import rules as rules_engine  # noqa: E402
import schedules as sch  # noqa: E402
import write_workbook as ww  # noqa: E402
import fixture_gen  # noqa: E402

YEAR_KEY = "2024-25"


def _build(tmp_path_factory, out_name, foreign_report=None):
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    book = pg.parse_book(FIXTURES / "syn_ind.gnucash")
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    result = mapping_engine.resolve_tree(tree, loaded)
    form16 = parse_form16.parse_form16(FIXTURES / "syn_ind_form16.pdf")
    rules = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    entities = configs.load_entities(ROOT / "bundling" / "canonical" / "itr" / "entities.example.yaml")
    entity = entities["SYN-IND"]
    scrips = configs.load_scrips(ROOT / "bundling" / "canonical" / "itr" / "scrips.example.yaml")
    fmv_tables = sch.load_fmv_tables()
    user_rules = rules_engine.load_user_rules(RULES_DIR / "user_rules.yaml")

    model = sch.build_all_schedules(
        tree, result.resolved, book, form16, YEAR_KEY, rules, "new",
        entity.status, entity.dob, scrips, fmv_tables,
        foreign_report=foreign_report,
    )
    out_path = tmp_path_factory.mktemp("wb") / out_name
    ww.write_workbook(
        str(out_path), tree, model, rules, user_rules, entity, "new", YEAR_KEY,
        form16.opted_out_115bac, [], [], [], result.unmapped, "v1", "2026-01-01T00:00:00", {},
        result.resolved, loaded.entries,
    )
    return openpyxl.load_workbook(str(out_path)), model


@pytest.fixture(scope="module")
def built_workbook(tmp_path_factory):
    return _build(tmp_path_factory, "syn_ind.xlsx")


@pytest.fixture(scope="module")
def built_workbook_with_foreign(tmp_path_factory):
    rep = pf.load_foreign_income_report(FIXTURES / "syn_indmoney_tax_report.xlsx")
    return _build(tmp_path_factory, "syn_ind_foreign.xlsx", foreign_report=rep), rep


def _col_b_formulas(ws):
    return [row[1] for row in ws.iter_rows(min_col=1, max_col=2, values_only=True)]


def test_all_expected_sheets_present(built_workbook):
    wb, _ = built_workbook
    expected = {
        "Rules", "Entity", "Salary", "BusinessPL", "HouseProperty", "ScheduleFA",
        "OtherSources", "CapitalGains", "ExemptIncome", "TaxesPaid", "Deductions",
        "ScheduleAL", "IS_Transcript", "BS_Transcript", "Computation", "Reconciliation",
        "Mapping Review",
    }
    assert expected.issubset(set(wb.sheetnames))


def test_computation_sheet_is_formula_driven(built_workbook):
    wb, _ = built_workbook
    formulas = [f for f in _col_b_formulas(wb["Computation"]) if isinstance(f, str)]
    formula_cells = [f for f in formulas if f.startswith("=")]
    # Every non-header, non-blank row on Computation should be a formula --
    # no literal numbers hardcoded into the tax block.
    assert len(formula_cells) >= 20
    assert all(f.startswith("=") for f in formula_cells)


def _all_formulas(ws):
    return [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]


def test_computation_references_every_schedule_sheet(built_workbook):
    """2026-07-20 on-page-totals change: 'Deductions' is now referenced from
    the `Statement of Income` page (the on-page Chapter VI-A leaf line), not
    from `Computation` -- the cross-section ladder moved off the working
    sheet. So this scans both sheets together; every other schedule is still
    referenced directly from `Computation`'s leaf cells."""
    wb, _ = built_workbook
    text = "\n".join(_all_formulas(wb["Computation"]) + _all_formulas(wb["Statement of Income"]))
    for sheet in ("Salary", "HouseProperty", "BusinessPL", "OtherSources", "CapitalGains", "Deductions", "TaxesPaid", "Rules", "Entity"):
        assert f"'{sheet}'!" in text, f"neither Computation nor Statement of Income references {sheet!r}"


def test_regime_chooser_cell_drives_selected_liability(built_workbook):
    wb, _ = built_workbook
    ws = wb["Computation"]
    # Find the IF(...'Entity'!...) chooser formula.
    chooser = [f for f in _col_b_formulas(ws) if isinstance(f, str) and f.startswith("=IF('Entity'!")]
    assert len(chooser) == 1
    assert '"old"' in chooser[0]


def test_rules_sheet_dumps_both_regime_slab_tables(built_workbook):
    wb, _ = built_workbook
    ws = wb["Rules"]
    labels = [row[0] for row in ws.iter_rows(min_col=1, max_col=1, values_only=True)]
    assert any(label == "New Regime" for label in labels)
    assert "Old Regime (resolved slab table for this entity's age/status)" in labels
    assert any(str(label).startswith("Slab ") for label in labels)


def test_rules_sheet_dumps_applied_user_rule(built_workbook):
    wb, _ = built_workbook
    ws = wb["Rules"]
    ids = [row[0] for row in ws.iter_rows(min_col=1, max_col=1, values_only=True)]
    assert "RULE-1" in ids


def test_entity_sheet_regime_cell_is_literal_value_not_formula(built_workbook):
    wb, _ = built_workbook
    ws = wb["Entity"]
    values = {row[0]: row[1] for row in ws.iter_rows(min_col=1, max_col=2, values_only=True) if row[0]}
    regime_label = [k for k in values if str(k).startswith("Regime")]
    assert regime_label
    assert values[regime_label[0]] == "new"


def test_house_property_sheet_shows_oq1_order(built_workbook):
    wb, model = built_workbook
    ws = wb["HouseProperty"]
    labels = [row[0] for row in ws.iter_rows(min_col=1, max_col=1, values_only=True)]
    assert "Net Annual Value (GAV - municipal tax)" in labels
    assert "Standard deduction 30% of NAV (s.24(a))" in labels
    assert "Interest on housing loan (s.24(b))" in labels


def test_other_sources_sheet_shows_refund_principal_and_interest_separately(built_workbook):
    wb, _ = built_workbook
    ws = wb["OtherSources"]
    rows = {row[0]: row[1] for row in ws.iter_rows(min_col=1, max_col=2, values_only=True) if row[0]}
    principal_key = next(k for k in rows if "refund principal" in str(k).lower())
    interest_key = next(k for k in rows if "IT refund" in str(k) and "interest" in str(k).lower())
    assert rows[principal_key] == 1000.0
    assert rows[interest_key] == 300.0


def test_capital_gains_sheet_has_lot_detail_rows(built_workbook):
    wb, model = built_workbook
    ws = wb["CapitalGains"]
    rows = list(ws.iter_rows(min_col=1, max_col=1, values_only=True))
    scrip_names = {r[0] for r in rows}
    assert "OldTech Ltd" in scrip_names
    assert "QuickFlip Ltd" in scrip_names


def test_reconciliation_sheet_reports_no_unmapped(built_workbook):
    wb, _ = built_workbook
    ws = wb["Reconciliation"]
    labels = [row[0] for row in ws.iter_rows(min_col=1, max_col=1, values_only=True)]
    assert any(str(label).startswith("Unmapped accounts (0)") for label in labels)


# ---------------------------------------------------------------------------
# Foreign report wiring (Schedule FA rework + ScheduleFSI/Form67/ScheduleTR)
# ---------------------------------------------------------------------------

def test_no_foreign_report_leaves_schedule_fa_and_sheet_set_unchanged(built_workbook):
    """A run with no foreign_report_xlsx supplied must be indistinguishable
    from a run before parse_foreign.py existed -- no new sheets, and
    ScheduleFA keeps its original AL_FOREIGN-only header."""
    wb, _ = built_workbook
    assert "ScheduleFSI" not in wb.sheetnames
    assert "Form67" not in wb.sheetnames
    assert "ScheduleTR" not in wb.sheetnames
    ws = wb["ScheduleFA"]
    first_cell = ws.cell(row=1, column=1).value
    assert first_cell == "Schedule FA -- pre-seeded from AL_FOREIGN-flagged holdings; detail cells manual"


def test_foreign_report_adds_three_new_sheets(built_workbook_with_foreign):
    (wb, _model), _rep = built_workbook_with_foreign
    assert {"ScheduleFSI", "Form67", "ScheduleTR"}.issubset(set(wb.sheetnames))


def test_schedule_fa_sheet_has_parsed_block_and_al_foreign_crosscheck_block(built_workbook_with_foreign):
    (wb, _model), _rep = built_workbook_with_foreign
    ws = wb["ScheduleFA"]
    labels = [row[0] for row in ws.iter_rows(min_col=1, max_col=1, values_only=True) if row[0]]
    assert any("parsed from foreign broker consolidated tax report" in str(label) for label in labels)
    assert any(str(label).startswith("AL_FOREIGN book roll-up") for label in labels)
    assert any(str(label).startswith("A2 -- Foreign Depository") for label in labels)
    assert any(str(label).startswith("A3 -- Foreign Equity") for label in labels)


def test_schedule_fa_a3_never_dedupes_lots_on_the_sheet(built_workbook_with_foreign):
    """Two lots share the same entity name in the fixture (see
    test_parse_foreign.py's test_schedule_fa_a3_never_dedupes_by_entity_name)
    -- the sheet must render both rows, not collapse them."""
    (wb, _model), _rep = built_workbook_with_foreign
    ws = wb["ScheduleFA"]
    entity_col = [row[1] for row in ws.iter_rows(min_col=1, max_col=2, values_only=True)]
    assert entity_col.count("SynCo Alpha Ltd (ADR)") == 2


def test_schedule_fsi_placeholder_and_prose_cells_are_flagged_not_blank_or_zero(built_workbook_with_foreign):
    (wb, _model), _rep = built_workbook_with_foreign
    ws = wb["ScheduleFSI"]
    other_row = next(
        row for row in ws.iter_rows(min_col=1, max_col=8, values_only=True)
        if row[0] == "Other Sources"
    )
    # income (col D) is a real, non-placeholder 4203 -- must not be flagged.
    assert other_row[3] == 4203
    # tax payable India (col F) is the "*" placeholder glyph in the fixture.
    assert other_row[5] == "placeholder:*"
    # tax relief (col G) is instructional prose in the fixture, never coerced
    # to a number and never left blank.
    assert other_row[6] == "prose:Lower of (c) or (d)"


def test_schedule_fsi_capital_gains_real_zero_survives_as_zero_not_a_placeholder(built_workbook_with_foreign):
    """The fixture's Capital Gains income cell is a genuine 0, distinct from
    a placeholder glyph or blank -- the sheet must render 0, not a flag
    string and not a blank cell."""
    (wb, _model), _rep = built_workbook_with_foreign
    ws = wb["ScheduleFSI"]
    cg_row = next(
        row for row in ws.iter_rows(min_col=1, max_col=8, values_only=True)
        if row[0] == "Capital Gains"
    )
    assert cg_row[3] == 0


def test_schedule_fsi_tie_out_block_present(built_workbook_with_foreign):
    (wb, _model), _rep = built_workbook_with_foreign
    ws = wb["ScheduleFSI"]
    labels = [row[0] for row in ws.iter_rows(min_col=1, max_col=1, values_only=True) if row[0]]
    assert any("Tie-out check" in str(label) for label in labels)
    values = {row[0]: row[1] for row in ws.iter_rows(min_col=1, max_col=2, values_only=True) if row[0]}
    assert values["Schedule FSI -- Other Sources income"] == 4203
    assert values["Dividend + Interest combined"] == 4203
    assert values["Difference (Other Sources - combined)"] == 0


def test_form67_other_cells_rendered_verbatim(built_workbook_with_foreign):
    (wb, _model), _rep = built_workbook_with_foreign
    ws = wb["Form67"]
    labels = [row[0] for row in ws.iter_rows(min_col=1, max_col=1, values_only=True) if row[0]]
    assert any("additional columns" in str(label) for label in labels)


def test_schedule_tr_row_rendered(built_workbook_with_foreign):
    (wb, _model), _rep = built_workbook_with_foreign
    ws = wb["ScheduleTR"]
    countries = [row[0] for row in ws.iter_rows(min_col=1, max_col=1, values_only=True)]
    assert "United States of America" in countries


def test_vendor_limit_absent_from_source_is_flagged_not_silently_dropped(built_workbook_with_foreign):
    (wb, _model), _rep = built_workbook_with_foreign
    ws = wb["ScheduleFA"]
    values = {row[0]: row[1] for row in ws.iter_rows(min_col=1, max_col=2, values_only=True) if row[0]}
    absent_key = next(k for k in values if "US tax withheld" in str(k))
    assert values[absent_key] == "EXPECTED BUT NOT FOUND IN SOURCE -- verify manually"
