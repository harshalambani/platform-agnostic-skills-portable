"""
tests/skill_itr_workbook/test_mapping_review_sheet.py -- golden test for
write_workbook.write_mapping_review_sheet against the SYN-IND fixture, with
one entry dropped (unmapped row) and one entry re-tagged PERSONAL in
memory only (never touching the committed fixture file) so a single test
run exercises: an unmapped row, a PERSONAL row, and the RULE-1 refund pair
(NONTAX_REFUND_PRINCIPAL / OS_REFUND_INTEREST) rendering with the correct
plain-English treatments from tags.py.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPTS = ROOT / "src" / "agents" / "skill_itr_workbook" / "scripts"
FIXTURES = Path(__file__).resolve().parent / "fixtures"

for p in (str(SCRIPTS), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

import parse_eguile as pe  # noqa: E402
import configs  # noqa: E402
import mapping as mapping_engine  # noqa: E402
import tags as tag_vocab  # noqa: E402
import write_workbook as ww  # noqa: E402
import fixture_gen  # noqa: E402

UNMAPPED_GUID = "ca85f6240bb66d5d38a0e4f40d908525"        # Assets/Misc Holding (unmapped)
PERSONAL_GUID = "2239368f6b27f15392c15b66544ca0f2"          # Expense/Business Expenses
REFUND_PRINCIPAL_GUID = "0403526e9bfd848676046e7c26556bb9"  # Income/IT Refund Principal
REFUND_INTEREST_GUID = "f1cba329edfd1b9bb836eb67534b93f4"   # Income/IT Refund Interest


def _sheet_rows(ws):
    return list(ws.iter_rows(values_only=True))


def _build_review_sheet():
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")

    entries = dict(loaded.entries)
    del entries[UNMAPPED_GUID]  # deliberately drop -> unmapped row
    entries[PERSONAL_GUID] = configs.MappingEntry(
        guid=PERSONAL_GUID, path="Expense/Business Expenses", tag="PERSONAL",
    )
    modified = configs.MappingLoadResult(entries=entries, warnings=[])
    result = mapping_engine.resolve_tree(tree, modified)

    wb = Workbook()
    ww.write_mapping_review_sheet(wb, tree, result.resolved, result.unmapped, entries)
    return wb["Mapping Review"], result


def test_unmapped_row_is_highlighted_at_top():
    ws, result = _build_review_sheet()
    assert len(result.unmapped) == 1
    labels = [row[0] for row in _sheet_rows(ws)]
    assert any(str(label).startswith("UNMAPPED (1)") for label in labels)
    rows_by_path = {row[0]: row for row in _sheet_rows(ws)}
    unmapped_row = rows_by_path["Assets/Misc Holding (unmapped)"]
    assert unmapped_row[2] == "REPLACE_ME"


def test_personal_row_shows_correct_treatment():
    ws, _ = _build_review_sheet()
    rows_by_path = {row[0]: row for row in _sheet_rows(ws)}
    row = rows_by_path["Expense/Business Expenses"]
    assert row[2] == "PERSONAL"
    assert row[4] == tag_vocab.TAGS["PERSONAL"].treatment


def test_rule1_refund_pair_renders_with_correct_treatments_and_amounts():
    ws, _ = _build_review_sheet()
    rows_by_path = {row[0]: row for row in _sheet_rows(ws)}

    principal = rows_by_path["Income/IT Refund Principal"]
    assert principal[2] == "NONTAX_REFUND_PRINCIPAL"
    assert principal[1] == 1000.0
    assert principal[4] == tag_vocab.TAGS["NONTAX_REFUND_PRINCIPAL"].treatment

    interest = rows_by_path["Income/IT Refund Interest"]
    assert interest[2] == "OS_REFUND_INTEREST"
    assert interest[1] == 300.0
    assert interest[4] == tag_vocab.TAGS["OS_REFUND_INTEREST"].treatment


def test_groups_have_subtotal_rows():
    ws, _ = _build_review_sheet()
    labels = [row[0] for row in _sheet_rows(ws)]
    subtotal_labels = [label for label in labels if isinstance(label, str) and label.startswith("Subtotal -- ")]
    assert subtotal_labels
    assert "Subtotal -- OtherSources" in subtotal_labels


# ---------------------------------------------------------------------------
# Fail-loud provenance (2026-07-19 mapping-precedence prompt, item 1b)
# ---------------------------------------------------------------------------

BANK_INTEREST_GUID = "11fc5a724efb9eb161ac039825b3e6dd"       # Income/Bank Interest (RE-Income)
BUS_REMUNERATION_GUID = "e605d8704bc37ebb37e724769e83bfb6"    # Income/Business Remuneration (RE-Income)


def _build_resolution_with_one_heuristic_income_leaf():
    """Marks Income/Bank Interest as an unverified heuristic guess (the same
    note text bootstrap tooling has historically used) while leaving every
    other entry, including another RE-Income leaf, as a plain approved
    entry -- so the provenance helpers can be checked for both true- and
    false-positive discrimination."""
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    entries = dict(loaded.entries)
    entries[BANK_INTEREST_GUID] = configs.MappingEntry(
        guid=BANK_INTEREST_GUID, path="Income/Bank Interest", tag="OS_INTEREST_BANK",
        note="HEURISTIC AUTO-SUGGESTION -- unverified, keyword-based, no LLM/network used.",
    )
    modified = configs.MappingLoadResult(entries=entries, warnings=[])
    result = mapping_engine.resolve_tree(tree, modified)
    return tree, result, entries


def test_provenance_counts_tallies_heuristic_vs_approved():
    tree, result, entries = _build_resolution_with_one_heuristic_income_leaf()
    counts = ww.provenance_counts(result.resolved, entries)
    assert counts["heuristic"] == 1
    assert counts["approved"] == len(result.resolved) - 1


def test_heuristic_income_leaves_flags_only_the_heuristic_income_row():
    tree, result, entries = _build_resolution_with_one_heuristic_income_leaf()
    leaves = ww.heuristic_income_leaves(tree, result.resolved, entries)
    # Only the one entry deliberately marked heuristic is flagged -- the
    # other RE-Income leaf (Business Remuneration), plainly approved, is not.
    assert [path for path, _tag, _amt in leaves] == ["Income/Bank Interest"]
    assert result.resolved[BUS_REMUNERATION_GUID].path not in [p for p, _t, _a in leaves]


def test_reconciliation_sheet_surfaces_heuristic_income_warning():
    tree, result, entries = _build_resolution_with_one_heuristic_income_leaf()
    wb = Workbook()
    ww.write_reconciliation_sheet(
        wb, tree, [], [], [], result.unmapped, "v1", "v1", "2026-07-19T00:00:00", {}, True,
        resolved=result.resolved, mapping_entries=entries,
    )
    ws = wb["Reconciliation"]
    cells = [row[0] for row in ws.iter_rows(values_only=True) if row and row[0] is not None]
    assert any("Tag provenance" in str(c) for c in cells)
    assert any("unverified INCOME tag" in str(c) for c in cells)
    assert any("Income/Bank Interest" in str(c) for c in cells)
