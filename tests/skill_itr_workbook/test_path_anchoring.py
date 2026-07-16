"""
tests/skill_itr_workbook/test_path_anchoring.py -- Batch 8 tests.

Defect A (Data/itr paths resolve to Data/Data/itr in the frozen build):
  - gate 1: the {data_root} run_args token (skill.yaml + ui/tabs/_generic.py)
    anchors entities_path/rules_dir/scrips_path to the same root
    ui/_config.data_root_dir() returns, in both a simulated frozen layout
    (CWD = <tmp>/Data) and a simulated source layout -- never doubling up to
    <root>/Data/itr/...
  - gate 2: an explicitly selected entity_key that can't be resolved
    (entities.yaml missing, or the key absent from it) fails loud -- no
    silent UNKNOWN/Individual/new-regime fallback, no green stub.

Defect B (mapping-less run silently emits an empty stub):
  - gate 3: a selected entity with an existing
    <data_root>/itr/mappings/<entity>.mapping.yaml and a blank mapping_file
    input auto-derives that mapping and proceeds as if it had been supplied.
  - gate 4: a selected entity with NO mapping file anywhere is a true cold
    start -- BLOCKED-FOR-REVIEW, a proposed-mappings snippet, never a green
    OK stub.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest
import yaml

ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPTS = ROOT / "src" / "agents" / "skill_itr_workbook" / "scripts"
AGENT_DIR = ROOT / "src" / "agents" / "skill_itr_workbook"
FIXTURES = Path(__file__).resolve().parent / "fixtures"

for p in (str(SCRIPTS), str(AGENT_DIR), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import agent  # noqa: E402
import configs  # noqa: E402
import fixture_gen  # noqa: E402
import ui._config as ui_config  # noqa: E402


SYN_TEST_ENTITY = {
    "SYN-TEST": {"name": "Synthetic Test", "pan": "ZZZZZ0000Z", "status": "Individual"},
}


def _write_entities_yaml(path: Path, entities: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(yaml.safe_dump(entities), encoding="utf-8")


# ---------------------------------------------------------------------------
# Gate 1 -- path anchoring
# ---------------------------------------------------------------------------

def test_data_root_token_resolves_entities_path_in_simulated_frozen_layout(tmp_path):
    """CWD = <tmp>/Data, data_root_dir() -> <tmp>/Data (Launcher convention):
    the {data_root}/itr/entities.yaml template -- resolved the same way
    ui/tabs/_generic.py substitutes it into a skill's run_args -- must read
    <tmp>/Data/itr/entities.yaml, never <tmp>/Data/Data/itr/entities.yaml."""
    data_dir = tmp_path / "Data"
    _write_entities_yaml(data_dir / "itr" / "entities.yaml", SYN_TEST_ENTITY)
    doubled = data_dir / "Data" / "itr" / "entities.yaml"

    with patch("ui._config.data_root_dir", return_value=data_dir):
        entities_path = "{data_root}/itr/entities.yaml".replace(
            "{data_root}", str(ui_config.data_root_dir())
        )

    assert not doubled.exists()
    entities = configs.load_entities(entities_path)
    assert "SYN-TEST" in entities


def test_data_root_token_resolves_entities_path_in_simulated_source_layout(tmp_path):
    """CWD = repo root (dev convenience) -- data_root_dir() resolves to
    <project>/Data there; simulated here as an arbitrary root. Same
    template, same anchor, same result: no special-casing between the two
    layouts is needed once entities_path is data_root_dir()-anchored."""
    data_dir = tmp_path / "src-mode-data"
    _write_entities_yaml(data_dir / "itr" / "entities.yaml", SYN_TEST_ENTITY)

    with patch("ui._config.data_root_dir", return_value=data_dir):
        entities_path = "{data_root}/itr/entities.yaml".replace(
            "{data_root}", str(ui_config.data_root_dir())
        )

    entities = configs.load_entities(entities_path)
    assert "SYN-TEST" in entities


def test_skill_yaml_uses_data_root_token_not_bare_data_prefix():
    """Regression guard: skill.yaml's entities_path/rules_dir/scrips_path
    run_args must route through the {data_root} anchor, not a bare
    'Data/...' string -- a bare prefix is exactly what doubled up to
    Data/Data/itr/... under the frozen Launcher's Data\\ CWD."""
    raw = yaml.safe_load((AGENT_DIR / "skill.yaml").read_text(encoding="utf-8"))
    run_args = raw["run_args"]
    assert run_args["entities_path"] == "{data_root}/itr/entities.yaml"
    assert run_args["rules_dir"] == "{data_root}/itr/rules"
    assert run_args["scrips_path"] == "{data_root}/itr/scrips.yaml"


# ---------------------------------------------------------------------------
# Gate 2 -- fail-loud entity resolution
# ---------------------------------------------------------------------------

def test_fail_loud_when_entity_key_set_but_entities_yaml_missing(tmp_path):
    html_path = tmp_path / "bs.html"
    html_path.write_text(fixture_gen.build_syn_ind_html(), encoding="utf-8")
    out_path = tmp_path / "out.xlsx"
    missing_entities = tmp_path / "no_such_dir" / "entities.yaml"

    summary = agent.run(
        str(html_path), str(out_path),
        entity_key="SYN-IND", entities_path=str(missing_entities),
    )

    assert "ERROR" in summary
    assert "SYN-IND" in summary
    assert str(missing_entities.resolve()) in summary
    assert "STATUS: OK" not in summary
    assert "Workbook: full schedule model built" not in summary

    wb = openpyxl.load_workbook(str(out_path))
    assert wb.sheetnames == ["Reconciliation"]  # stub only, error text


def test_fail_loud_when_entity_key_absent_from_entities_yaml(tmp_path):
    html_path = tmp_path / "bs.html"
    html_path.write_text(fixture_gen.build_syn_ind_html(), encoding="utf-8")
    out_path = tmp_path / "out.xlsx"
    entities_path = tmp_path / "entities.yaml"
    _write_entities_yaml(entities_path, {"SOME-OTHER-ENTITY": SYN_TEST_ENTITY["SYN-TEST"]})

    summary = agent.run(
        str(html_path), str(out_path),
        entity_key="SYN-IND", entities_path=str(entities_path),
    )

    assert "ERROR" in summary
    assert "SYN-IND" in summary
    assert str(entities_path.resolve()) in summary
    assert "STATUS: OK" not in summary


def test_entity_key_unset_still_falls_back_gracefully(tmp_path):
    """Only an EXPLICIT entity_key fails loud. An ad hoc run with no entity
    selected at all (entity_key omitted, no mapping_file to derive a stem
    from either) must keep degrading gracefully -- there's nothing the user
    actually chose that could be silently wrong."""
    html_path = tmp_path / "bs.html"
    html_path.write_text(fixture_gen.build_syn_ind_html(), encoding="utf-8")
    out_path = tmp_path / "out.xlsx"
    missing_entities = tmp_path / "no_such_dir" / "entities.yaml"

    summary = agent.run(str(html_path), str(out_path), entities_path=str(missing_entities))

    assert "ERROR: entity" not in summary


# ---------------------------------------------------------------------------
# Gate 3 -- B(i) auto-derive the entity's existing mapping
# ---------------------------------------------------------------------------

def _entities_yaml_with_syn_ind(path: Path) -> None:
    _write_entities_yaml(path, {
        "SYN-IND": {
            "name": "Synthetic Individual", "pan": "AAAAA0000A", "status": "Individual",
            "default_regime": "new",
        },
    })


def test_auto_derives_existing_entity_mapping_when_blank(tmp_path):
    html_path = tmp_path / "bs.html"
    html_path.write_text(fixture_gen.build_syn_ind_html(), encoding="utf-8")
    out_path = tmp_path / "out.xlsx"

    entities_path = tmp_path / "itr" / "entities.yaml"
    _entities_yaml_with_syn_ind(entities_path)
    mappings_dir = entities_path.parent / "mappings"
    mappings_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy(FIXTURES / "syn_ind.mapping.yaml", mappings_dir / "SYN-IND.mapping.yaml")

    summary = agent.run(
        str(html_path), str(out_path),
        entity_key="SYN-IND", entities_path=str(entities_path),
        # mapping_file intentionally omitted -- the Entity mapping box was empty.
    )

    assert "Mapping: auto-derived" in summary
    assert "SYN-IND.mapping.yaml" in summary
    assert "STATUS: OK" in summary
    assert "Workbook: full schedule model built" in summary
    wb = openpyxl.load_workbook(str(out_path))
    assert "Computation" in wb.sheetnames


def test_auto_derive_does_not_override_explicit_mapping_file(tmp_path):
    """An explicitly supplied mapping_file always wins -- auto-derive only
    kicks in when the box was actually left blank."""
    html_path = tmp_path / "bs.html"
    html_path.write_text(fixture_gen.build_syn_ind_html(), encoding="utf-8")
    out_path = tmp_path / "out.xlsx"

    entities_path = tmp_path / "itr" / "entities.yaml"
    _entities_yaml_with_syn_ind(entities_path)
    mappings_dir = entities_path.parent / "mappings"
    mappings_dir.mkdir(parents=True, exist_ok=True)
    # A decoy auto-derive target that would resolve every leaf, so if
    # auto-derive wrongly overrode the explicit mapping_file this test would
    # see STATUS: OK instead of BLOCKED-FOR-REVIEW.
    shutil.copy(FIXTURES / "syn_ind.mapping.yaml", mappings_dir / "SYN-IND.mapping.yaml")

    summary = agent.run(
        str(html_path), str(out_path),
        entity_key="SYN-IND", entities_path=str(entities_path),
        mapping_file=str(FIXTURES / "syn_ind_unmapped.mapping.yaml"),
    )

    assert "Mapping: auto-derived" not in summary
    assert "STATUS: BLOCKED-FOR-REVIEW" in summary


# ---------------------------------------------------------------------------
# Gate 4 -- B(ii) true cold start (no mapping anywhere for the entity)
# ---------------------------------------------------------------------------

def test_cold_start_no_mapping_anywhere_blocks_for_review(tmp_path):
    html_path = tmp_path / "bs.html"
    html_path.write_text(fixture_gen.build_syn_ind_html(), encoding="utf-8")
    out_path = tmp_path / "out.xlsx"

    entities_path = tmp_path / "itr" / "entities.yaml"
    _entities_yaml_with_syn_ind(entities_path)
    # No mappings/ directory at all -- nothing to auto-derive.

    summary = agent.run(
        str(html_path), str(out_path),
        entity_key="SYN-IND", entities_path=str(entities_path),
    )

    assert "STATUS: BLOCKED-FOR-REVIEW" in summary
    assert "STATUS: OK" not in summary
    assert "cold start" in summary
    snippet_path = Path(str(out_path) + "-proposed-mappings.yaml")
    assert snippet_path.exists()
    assert "REPLACE_ME" in snippet_path.read_text(encoding="utf-8")

    wb = openpyxl.load_workbook(str(out_path))
    assert wb.sheetnames == ["Reconciliation"]  # never a green full workbook
