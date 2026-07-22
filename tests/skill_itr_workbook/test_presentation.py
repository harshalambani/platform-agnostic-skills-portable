"""
tests/skill_itr_workbook/test_presentation.py -- the four deliverable sheets
(`Statement of Income`, `IS`, `BS`, `CG`) added by scripts/presentation.py.

openpyxl never evaluates formulas, so -- as in test_write_workbook.py -- these
tests assert the formula GRAPH (which cell points at which other sheet's cell)
rather than evaluated numbers. That is the right target here anyway: the whole
point of the presentation layer is that it recomputes nothing, so "is this cell
a formula pointing back at the engine" IS the property under test.

ALL fixture identity data is synthetic: the SYN-IND profile from
Data/itr/entities.example.yaml (fake name/PAN/DOB/address) plus invented
account paths and scrip names. No real PAN, Aadhaar, DOB, address, account
number or holding appears anywhere in this file.
"""
from __future__ import annotations

import dataclasses
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPTS = ROOT / "src" / "agents" / "skill_itr_workbook" / "scripts"
FIXTURES = Path(__file__).resolve().parent / "fixtures"
RULES_DIR = ROOT / "Data" / "itr" / "rules"

for p in (str(SCRIPTS), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

import parse_eguile as pe  # noqa: E402
import parse_gnucash as pg  # noqa: E402
import parse_form16  # noqa: E402
import configs  # noqa: E402
import mapping as mapping_engine  # noqa: E402
import presentation  # noqa: E402
import rules as rules_engine  # noqa: E402
import schedules as sch  # noqa: E402
import write_workbook as ww  # noqa: E402
import fixture_gen  # noqa: E402

YEAR_KEY = "2024-25"

# Sheets that existed before the presentation layer -- none of these may lose
# values or disappear; four of them become hidden (never deleted).
ORIGINAL_SHEETS = {
    "Rules", "Entity", "Salary", "BusinessPL", "HouseProperty", "ScheduleFA",
    "OtherSources", "CapitalGains", "ExemptIncome", "TaxesPaid", "Deductions",
    "ScheduleAL", "IS_Transcript", "BS_Transcript", "Computation",
    "Reconciliation", "Mapping Review",
}


def _build(tmp_path, *, dob: str | None = None, no_capital_gains: bool = False,
           entity_overrides: dict | None = None):
    """Build the full synthetic workbook, optionally overriding the synthetic
    entity's (already fake) DOB to exercise a different age class, other
    entity fields (father_name/aadhaar/residency, all still synthetic) via
    `entity_overrides`, or blanking this FY's capital-gains activity to
    exercise the no-CG path."""
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    book = pg.parse_book(FIXTURES / "syn_ind.gnucash")
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    result = mapping_engine.resolve_tree(tree, loaded)
    form16 = parse_form16.parse_form16(FIXTURES / "syn_ind_form16.pdf")
    rules = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    entity = configs.load_entities(ROOT / "Data" / "itr" / "entities.example.yaml")["SYN-IND"]
    if dob is not None:
        entity = dataclasses.replace(entity, dob=dob)
    if entity_overrides:
        entity = dataclasses.replace(entity, **entity_overrides)
    scrips = configs.load_scrips(ROOT / "Data" / "itr" / "scrips.example.yaml")
    fmv_tables = sch.load_fmv_tables()
    user_rules = rules_engine.load_user_rules(RULES_DIR / "user_rules.yaml")

    model = sch.build_all_schedules(
        tree, result.resolved, book, form16, YEAR_KEY, rules, "new",
        entity.status, entity.dob, scrips, fmv_tables, residency=entity.residency,
    )
    if no_capital_gains:
        # A financial year with no disposals at all -- synthetic, and applied
        # per run, exactly as a real quiet year would arrive from the books.
        model.capital_gains = sch.CapitalGainsSchedule()
    out_path = tmp_path / "syn_ind.xlsx"
    ww.write_workbook(
        str(out_path), tree, model, rules, user_rules, entity, "new", YEAR_KEY,
        form16.opted_out_115bac, [], [], [], result.unmapped, "v1",
        "2026-01-01T00:00:00", {}, result.resolved, loaded.entries,
    )
    return openpyxl.load_workbook(str(out_path))


@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    return _build(tmp_path_factory.mktemp("pres"))


# ---------------------------------------------------------------------------
# Gate 1 -- sheet order, hiding, additive-only
# ---------------------------------------------------------------------------

def test_four_presentation_sheets_are_first_in_order(wb):
    assert wb.sheetnames[:4] == ["Statement of Income", "BS", "IS", "CG"]


def test_working_sheets_are_hidden_not_deleted(wb):
    for name in ("Rules", "Mapping Review", "IS_Transcript", "BS_Transcript"):
        assert name in wb.sheetnames, f"{name} was deleted -- must be hidden, not gone"
        assert wb[name].sheet_state == "hidden"


def test_presentation_layer_is_purely_additive(wb):
    """Every pre-existing sheet survives, and the ONLY sheets that are hidden
    are the four named working sheets. (That the surviving sheets' *values*
    are unchanged is held by the pre-existing suites -- test_write_workbook.py,
    test_agent_full_pipeline.py, test_best_effort_workbook.py and
    test_mapping_review_sheet.py all still assert their contents.)"""
    assert ORIGINAL_SHEETS.issubset(set(wb.sheetnames))
    hidden = {n for n in wb.sheetnames if wb[n].sheet_state == "hidden"}
    assert hidden == set(presentation.HIDDEN_SHEETS)


# ---------------------------------------------------------------------------
# Gate 2 -- every money cell is a formula, never a literal
# ---------------------------------------------------------------------------

def _is_input_cell(c) -> bool:
    """True for the one deliberate exception to "every money cell is a
    formula": the b/f-loss set-off cell (2026-07-20 on-page-totals change) is
    a REAL, editable, literal-valued input cell by design -- distinguished by
    its `_INPUT_FILL` styling (see presentation._input_cell)."""
    fill = c.fill
    rgb = getattr(fill.start_color, "rgb", None) or getattr(fill.fgColor, "rgb", None)
    return rgb == "00DDEBF7"


@pytest.mark.parametrize("sheet", ["Statement of Income", "IS", "BS", "CG"])
def test_no_numeric_literal_anywhere_on_presentation_sheets(wb, sheet):
    """The strongest form of "every money cell is a formula": no cell on these
    sheets holds a numeric type at all. Amounts, quantities, dates and even the
    CG Sr. No. column are formulas, so any int/float here is by definition a
    hardcoded value that escaped the formula link. The single deliberate
    exception is the b/f-loss set-off INPUT cell (2026-07-20 change) --
    identified by its distinct styling, not by position, so this stays a
    strong test."""
    offenders = [
        (c.coordinate, c.value)
        for row in wb[sheet].iter_rows() for c in row
        if isinstance(c.value, (int, float)) and not isinstance(c.value, bool)
        and not _is_input_cell(c)
    ]
    assert offenders == [], f"{sheet} has hardcoded numeric literal(s) at {offenders}"


def _caption_row(ws, col: int = 3) -> int:
    for row in ws.iter_rows(min_col=col, max_col=col):
        for c in row:
            if c.value == "Rs.":
                return c.row
    raise AssertionError("no 'Rs.' caption row found")


def test_statement_money_columns_below_caption_are_all_formulas(wb):
    """Same deliberate exception as test_no_numeric_literal_anywhere_on_
    presentation_sheets: the b/f-loss INPUT cell, identified by styling."""
    ws = wb["Statement of Income"]
    start = _caption_row(ws) + 1
    seen = 0
    for row in ws.iter_rows(min_row=start, min_col=3, max_col=5):
        for c in row:
            if c.value is None:
                continue
            if _is_input_cell(c):
                continue
            assert isinstance(c.value, str) and c.value.startswith("="), \
                f"{c.coordinate} is not a formula: {c.value!r}"
            seen += 1
    assert seen >= 10


@pytest.mark.parametrize("sheet", ["IS", "BS"])
def test_hierarchy_money_columns_are_all_formulas(wb, sheet):
    ws = wb[sheet]
    seen = 0
    for row in ws.iter_rows(min_row=_caption_row(ws, 2) + 1, min_col=2):
        for c in row:
            if c.value is None:
                continue
            assert isinstance(c.value, str) and c.value.startswith("=")
            seen += 1
    assert seen >= 5


# ---------------------------------------------------------------------------
# Gate 3 / 4 -- parked items, residency assumption, age class
# ---------------------------------------------------------------------------

def _labels(ws, col: int = 1):
    return [c.value for row in ws.iter_rows(min_col=col, max_col=col) for c in row if c.value]


def _find_label(ws, needle: str):
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and needle.lower() in c.value.lower():
                return c
    raise AssertionError(f"label {needle!r} not found")


def test_brought_forward_loss_buckets_are_real_editable_input_cells_not_parked(wb):
    """2026-07-21 on-page-totals change (design doc section 11.2): the old
    single lump b/f-loss cell is replaced by FOUR statutory per-bucket
    input cells (HP s.71B, Business s.72, STCL/LTCL s.74). Each is a real,
    editable cell defaulting to 0, styled distinctly from a parked cell
    (solid thin border + _INPUT_FILL, not dotted + _PARKED_FILL). Routing
    is proven via the Business bucket, which SYN-IND actually renders on
    the page (non-zero business income) -- it must be wired into the
    "Net business income" head item, NOT into Total Income directly (the
    pre-2026-07-21 lump-subtraction wiring)."""
    ws = wb["Statement of Income"]
    section_header = _find_label(ws, "Brought forward losses set off")
    assert "(to be filled)" not in section_header.value, "b/f loss section must no longer read as PARKED"

    bucket_needles = (
        "b/f House Property loss (s.71B)",
        "b/f Business loss (s.72)",
        "b/f Short-term capital loss (s.74)",
        "b/f Long-term capital loss (s.74)",
    )
    biz_cell = None
    for needle in bucket_needles:
        label = _find_label(ws, needle)
        cell = ws.cell(row=label.row, column=4)
        assert cell.value == 0, f"{needle!r} input cell must default to literal 0, got {cell.value!r}"
        assert cell.fill.start_color.rgb == "00DDEBF7" or cell.fill.fgColor.rgb == "00DDEBF7"
        assert cell.border.bottom.style == "thin"
        if needle == "b/f Business loss (s.72)":
            biz_cell = cell

    biz_item_label = _find_label(ws, "Net business income")
    biz_formula = ws.cell(row=biz_item_label.row, column=3).value
    assert isinstance(biz_formula, str) and biz_formula.startswith("=")
    assert biz_cell.coordinate in biz_formula, \
        "Business bucket must be wired into the Business head item, not a lump Total Income term"

    total_income_label = _find_label(ws, "Total Income")
    total_income_formula = ws.cell(row=total_income_label.row, column=5).value
    assert isinstance(total_income_formula, str) and total_income_formula.startswith("=")
    assert biz_cell.coordinate not in total_income_formula, \
        "b/f buckets must not be a lump Total-Income deduction any more -- routing happens at the head level"


# ---------------------------------------------------------------------------
# Gate 1/1a -- Father's Name and Aadhaar are unparked, present vs. absent
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("needle,value_col", [
    ("Father's Name", 2),
    ("Aadhaar No.", 5),
])
def test_father_name_and_aadhaar_render_as_parked_when_absent(tmp_path, needle, value_col):
    """When the entity has no father_name/aadhaar on file, the field falls
    back to the same styled-empty PARKED presentation as before -- per field,
    proven independently of the other unparked field."""
    built = _build(tmp_path, entity_overrides={"father_name": None, "aadhaar": None})
    ws = built["Statement of Income"]
    label = _find_label(ws, needle)
    assert "(to be filled)" in label.value
    cell = ws.cell(row=label.row, column=value_col)
    assert cell.value is None, f"{needle} parked cell must stay empty, got {cell.value!r}"
    assert cell.fill.start_color.rgb == "00FFF2CC" or cell.fill.fgColor.rgb == "00FFF2CC"
    assert cell.border.bottom.style == "dotted"


def test_father_name_renders_a_real_value_and_drops_the_parked_note(wb):
    """SYN-IND now carries a synthetic father_name -- the label loses
    "(to be filled)" and the value cell is a formula back to the Entity
    sheet, not a hardcoded literal (audit-trail rule)."""
    ws = wb["Statement of Income"]
    label = _find_label(ws, "Father's Name")
    assert "(to be filled)" not in label.value
    value = ws.cell(row=label.row, column=2).value
    assert isinstance(value, str) and value.startswith("='Entity'!")


def test_aadhaar_renders_space_grouped_formula_and_drops_the_parked_note(wb):
    """SYN-IND now carries a synthetic aadhaar -- the label loses
    "(to be filled)" and the value cell is a LEFT/MID formula over the raw
    digits on the Entity sheet, space-grouped NNNN NNNN NNNN CA-file style."""
    ws = wb["Statement of Income"]
    label = _find_label(ws, "Aadhaar No.")
    assert "(to be filled)" not in label.value
    value = ws.cell(row=label.row, column=5).value
    assert isinstance(value, str) and value.startswith("=LEFT('Entity'!")
    assert 'MID(' in value and '" "&' in value


def test_aadhaar_value_never_appears_literally_outside_its_own_entity_cell(wb):
    """The raw Aadhaar digit string is stored once, on the Entity sheet, the
    same way PAN/DOB are (plaintext, no at-rest masking -- see PR report).
    It must never appear a second time anywhere else in the workbook -- not
    as a duplicated literal, and not inside any warning/error/log-style cell
    -- only inside formulas that reference the Entity sheet's own cell by
    coordinate."""
    import configs as configs_mod
    entity = configs_mod.load_entities(
        ROOT / "Data" / "itr" / "entities.example.yaml"
    )["SYN-IND"]
    raw = entity.aadhaar
    assert raw is not None and raw.isdigit()
    literal_hits = []
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and raw in v and not v.startswith("="):
                    literal_hits.append((sheet, c.coordinate, v))
    assert len(literal_hits) == 1, f"Aadhaar digits found outside its single Entity source cell: {literal_hits}"
    assert literal_hits[0][0] == "Entity"


def test_status_line_is_assumed_ror_with_footnote_marker(wb):
    ws = wb["Statement of Income"]
    label = _find_label(ws, "Residential Status")
    value = ws.cell(row=label.row, column=5).value
    assert value.startswith("R/OR")
    assert value.rstrip().endswith("*"), "R/OR must carry a footnote marker -- it is an assumption"


def test_assumptions_note_renders_and_names_the_assumption(wb):
    ws = wb["Statement of Income"]
    note = _find_label(ws, "Residential status is ASSUMED")
    assert "R/OR" in note.value and "not determined by this tool" in note.value
    assert _find_label(ws, "Assumptions") is not None


# ---------------------------------------------------------------------------
# Gate 2/3 -- real residency: declared vs. defaulted, RNOR/NR rendering
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("residency,expected_prefix", [
    ("R/OR", "R/OR"),
    ("RNOR", "RNOR"),
    ("NR", "NR"),
])
def test_status_line_renders_all_three_declared_residency_categories(residency, expected_prefix):
    assert presentation.status_line("general", residency, declared=True) == expected_prefix


def test_status_line_carries_no_footnote_when_declared():
    value = presentation.status_line("general", "NR", declared=True)
    assert not value.rstrip().endswith("*")


def test_status_line_carries_footnote_when_not_declared():
    value = presentation.status_line("general", "R/OR", declared=False)
    assert value.rstrip().endswith("*")


def test_resolve_residency_only_accepts_the_three_statutory_tokens():
    assert rules_engine.resolve_residency("R/OR") == ("R/OR", True)
    assert rules_engine.resolve_residency("RNOR") == ("RNOR", True)
    assert rules_engine.resolve_residency("NR") == ("NR", True)
    # Legacy free text (every real and synthetic entities.yaml today) and an
    # unset value are both undeclared -- default to R/OR, exactly as before.
    assert rules_engine.resolve_residency("Resident") == ("R/OR", False)
    assert rules_engine.resolve_residency(None) == ("R/OR", False)


def test_declared_nr_entity_drops_the_assumptions_footnote_end_to_end(tmp_path):
    """SYN-IND-NR (Data/itr/entities.example.yaml) declares residency: NR.
    End to end, the header line shows 'NR' with no footnote marker and the
    whole Assumptions block is gone -- someone asserted this."""
    entity = configs.load_entities(ROOT / "Data" / "itr" / "entities.example.yaml")["SYN-IND-NR"]
    assert entity.residency == "NR"
    built = _build(tmp_path, entity_overrides={
        "residency": entity.residency, "dob": entity.dob,
        "father_name": entity.father_name, "aadhaar": entity.aadhaar,
    })
    ws = built["Statement of Income"]
    label = _find_label(ws, "Residential Status")
    value = ws.cell(row=label.row, column=5).value
    assert value.startswith("NR")
    assert not value.rstrip().endswith("*")
    with pytest.raises(AssertionError):
        _find_label(ws, "Assumptions")


def test_status_line_age_half_comes_from_resolve_age_class():
    """The age half of the status line is whatever the EXISTING resolver
    returns -- no new age logic. Call site: write_workbook.write_workbook()
    passes rules.resolve_age_class(entity.status, entity.dob, fy_end) into
    presentation.build_presentation_layer()."""
    from datetime import date
    fy_end = date(2025, 3, 31)
    # Synthetic DOBs only.
    assert rules_engine.resolve_age_class("Individual", "1990-01-01", fy_end) == "general"
    assert rules_engine.resolve_age_class("Individual", "1960-01-01", fy_end) == "senior"
    assert rules_engine.resolve_age_class("Individual", "1940-01-01", fy_end) == "super_senior"

    # 'general' deliberately renders NO age suffix (prompt 2b).
    assert presentation.status_line("general").rstrip(" *") == "R/OR"
    assert presentation.status_line("senior").startswith("R/OR - Senior Citizen")
    assert presentation.status_line("super_senior").startswith("R/OR - Super Senior Citizen")


@pytest.mark.parametrize("dob,expected", [
    ("1940-01-01", "R/OR - Super Senior Citizen"),   # synthetic
    ("1990-01-01", "R/OR"),                           # synthetic
])
def test_age_class_renders_end_to_end_with_synthetic_dobs(tmp_path, dob, expected):
    built = _build(tmp_path, dob=dob)
    ws = built["Statement of Income"]
    label = _find_label(ws, "Residential Status")
    assert ws.cell(row=label.row, column=5).value.rstrip(" *") == expected


def test_only_the_selected_regime_is_shown(wb):
    """The both-regimes comparison stays on the Computation working sheet."""
    ws = wb["Statement of Income"]
    labels = " ".join(str(v) for v in _labels(ws, 2))
    assert "Old Regime" not in labels
    tax_row = _find_label(ws, "Tax on total income")
    formula = ws.cell(row=tax_row.row, column=4).value
    assert formula.startswith("=IF('Entity'!"), "regime must be picked from the Entity chooser cell"


# ---------------------------------------------------------------------------
# Gate 5 -- hierarchy, tiered subtotals, sibling groups, depth from path
# ---------------------------------------------------------------------------

def _render(entries, source="BS_Transcript"):
    wb_ = openpyxl.Workbook()
    ws = wb_.active
    root = presentation.build_hierarchy(entries)
    presentation.render_hierarchy(ws, 1, root, source)
    return ws


def _row_of(ws, label_fragment: str):
    for row in ws.iter_rows(min_col=1, max_col=1):
        for c in row:
            if isinstance(c.value, str) and c.value.strip() == label_fragment:
                return c.row
    raise AssertionError(f"row {label_fragment!r} not rendered")


def _money_col(ws, row: int) -> int:
    for c in ws[row]:
        if c.column > 1 and c.value is not None:
            return c.column
    raise AssertionError(f"row {row} has no amount cell")


# Synthetic BS shaped like a real book: Current Assets with four sibling
# groups. Account numbers and bank names are invented.
_SIBLING_ENTRIES = [
    ("Assets/Current Assets/Brokers/SynBroker Ltd", "B3"),
    ("Assets/Current Assets/Cash and Bank/SynBank - 900001", "B4"),
    ("Assets/Current Assets/Cash and Bank/SynBank - 900002", "B5"),
    ("Assets/Current Assets/Fixed Deposits/SynBank FD - 900003", "B6"),
    ("Assets/Current Assets/Loans and Advances/SynLoan Advance", "B7"),
]


def test_current_assets_keeps_its_four_sibling_groups_separate():
    """`BS` is a GnuCash view. Each sibling group keeps its own row and its own
    subtotal -- sibling groups are never merged."""
    ws = _render(_SIBLING_ENTRIES)
    for group in ("Brokers", "Cash and Bank", "Fixed Deposits", "Loans and Advances"):
        assert _row_of(ws, group), f"{group} lost its own group row"
        assert _row_of(ws, f"Total {group}"), f"{group} lost its own subtotal row"


def test_fixed_deposits_is_not_folded_into_cash_and_bank():
    """Schedule AL's statutory buckets DO combine bank and FDs ("Bank
    including all deposits"). That is a different sheet with a different
    purpose and does not license combining them here."""
    ws = _render(_SIBLING_ENTRIES)
    cash_total_row = _row_of(ws, "Total Cash and Bank")
    fd_total_row = _row_of(ws, "Total Fixed Deposits")
    assert cash_total_row != fd_total_row

    cash_formula = ws.cell(row=cash_total_row, column=_money_col(ws, cash_total_row)).value
    fd_formula = ws.cell(row=fd_total_row, column=_money_col(ws, fd_total_row)).value
    # The FD leaf must contribute to the FD subtotal and to nothing else.
    fd_leaf_row = _row_of(ws, "SynBank FD - 900003")
    assert f"B{fd_leaf_row}" in fd_formula.replace("SUM(", "").replace(")", "") or \
        f"B{fd_leaf_row}" in fd_formula
    assert str(fd_leaf_row) not in cash_formula.split("SUM(")[-1].split(")")[0]


def test_tiered_subtotal_columns_occupy_three_different_columns():
    """A leaf, its group total and a higher-level total must land in three
    different columns -- not flattened into one."""
    ws = _render(_SIBLING_ENTRIES)
    leaf_col = _money_col(ws, _row_of(ws, "SynBank - 900001"))
    group_col = _money_col(ws, _row_of(ws, "Total Cash and Bank"))
    mid_col = _money_col(ws, _row_of(ws, "Total Current Assets"))
    top_col = _money_col(ws, _row_of(ws, "Total Assets"))
    assert leaf_col < group_col < mid_col < top_col
    assert len({leaf_col, group_col, mid_col, top_col}) == 4


def test_depth_is_derived_from_the_path_not_a_hardcoded_level_count():
    """A synthetic 5-level book must render 5 levels -- depth is driven by the
    `Path` column, never capped at the levels present in one entity's data."""
    entries = [("L1/L2/L3/L4/SynLeaf", "B3"), ("L1/L2/L3/L4/SynLeaf2", "B4")]
    ws = _render(entries)
    for level_name in ("L1", "L2", "L3", "L4"):
        assert _row_of(ws, level_name)
        assert _row_of(ws, f"Total {level_name}")
    # Four nested groups => four distinct subtotal columns above the leaf.
    cols = [_money_col(ws, _row_of(ws, f"Total {n}")) for n in ("L4", "L3", "L2", "L1")]
    assert cols == sorted(cols) and len(set(cols)) == 4
    assert presentation.max_group_level(presentation.build_hierarchy(entries)) == 3


def test_leaf_amounts_are_formulas_into_the_transcript():
    ws = _render([("Assets/Current Assets/Brokers/SynBroker Ltd", "B3")])
    leaf_row = _row_of(ws, "SynBroker Ltd")
    assert ws.cell(row=leaf_row, column=2).value == "='BS_Transcript'!B3"


def test_is_and_bs_titles_are_built_from_entity_by_formula(wb):
    assert wb["IS"]["A1"].value.startswith("='Entity'!")
    assert "Income Statement For Period Covering" in wb["IS"]["A1"].value
    assert wb["BS"]["A1"].value.startswith("='Entity'!")
    assert "Balance Sheet as at" in wb["BS"]["A1"].value


def test_indentation_deepens_per_level():
    ws = _render(_SIBLING_ENTRIES)
    def indent(row):
        v = ws.cell(row=row, column=1).value
        return len(v) - len(v.lstrip(" "))
    assert indent(_row_of(ws, "Assets")) == 0
    assert indent(_row_of(ws, "Current Assets")) == 6
    assert indent(_row_of(ws, "Brokers")) == 12
    assert indent(_row_of(ws, "SynBroker Ltd")) == 18


# ---------------------------------------------------------------------------
# Gate 6 -- CG traces to CapitalGains, no CA-file arithmetic
# ---------------------------------------------------------------------------

def _cg_formulas(wb):
    return [c.value for row in wb["CG"].iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith("=")]


def test_cg_figures_all_trace_to_the_capitalgains_sheet(wb):
    # The title cell legitimately references Entity for the entity name; the
    # Sr. No. column and the block totals are self-referential.
    formulas = [f for f in _cg_formulas(wb)
                if not f.startswith(("=ROW(", "=SUM(", "='Entity'!"))]
    assert formulas
    for f in formulas:
        assert "'CapitalGains'!" in f, f"CG figure does not trace to CapitalGains: {f!r}"


def test_cg_has_no_inline_fmv_literal(wb):
    """The CA reference hardcodes 31-Jan-2018 FMV prices as literals inside
    each row's formula. This sheet reads FMV from the CapitalGains column that
    the real fmv_tables lookup populated."""
    import re
    for f in _cg_formulas(wb):
        stripped = re.sub(r"'CapitalGains'!\w+\d+", "", f)
        stripped = re.sub(r"\bROW\(\)", "", stripped)
        numbers = re.findall(r"\d+\.\d+", stripped)
        assert not numbers, f"CG formula carries an inline price literal: {f!r}"


def test_cg_never_reimplements_the_ca_grandfathering_arithmetic(wb):
    """The CA file's K10 = J10 - G10 subtracts the 31-Jan-2018 FMV *in
    addition to* cost, which is not a valid basis, and is inconsistent with
    its own K9. Neither that formula nor any other cross-column subtraction is
    recreated here: every CG figure is a direct reference."""
    for f in _cg_formulas(wb):
        if f.startswith(("=ROW(", "=SUM(", "=IF(")):
            continue
        assert "-" not in f, f"CG must not do its own arithmetic: {f!r}"
    # And specifically: no <col><row> - <col><row> style formula anywhere.
    import re
    for f in _cg_formulas(wb):
        assert not re.search(r"\b[A-J]\d+\s*-\s*[A-J]\d+", f), \
            f"CG recreates a CA-style difference formula: {f!r}"


def test_cg_sheet_is_present_when_the_year_has_gains(wb):
    assert "CG" in wb.sheetnames


def test_cg_sheet_is_omitted_when_the_year_has_no_gains(tmp_path):
    """A CA reference workbook for a year with no gains has no capital-gains
    sheet at all. The decision is per-run and per-FY -- never an entity-level
    flag, and never inferred from a prior year or a reference workbook."""
    built = _build(tmp_path, no_capital_gains=True)
    assert "CG" not in built.sheetnames


def test_remaining_sheets_still_render_and_position_correctly_without_cg(tmp_path):
    """Sheet ordering must not assume CG is present."""
    built = _build(tmp_path, no_capital_gains=True)
    assert built.sheetnames[:3] == ["Statement of Income", "BS", "IS"]
    assert ORIGINAL_SHEETS.issubset(set(built.sheetnames))
    hidden = {n for n in built.sheetnames if built[n].sheet_state == "hidden"}
    assert hidden == set(presentation.HIDDEN_SHEETS)
    for sheet in ("Statement of Income", "IS", "BS"):
        assert built[sheet].print_area
        assert built[sheet].column_dimensions["A"].width


def test_capital_gains_activity_is_decided_from_this_years_data_only():
    empty = sch.CapitalGainsSchedule()
    assert presentation.has_capital_gains_activity(empty) is False
    # Any one of: a disposal lot, a book control total, or a taxable figure.
    assert presentation.has_capital_gains_activity(
        sch.CapitalGainsSchedule(st_taxable_gross=1234.0)) is True
    assert presentation.has_capital_gains_activity(
        sch.CapitalGainsSchedule(lt_control=-500.0)) is True


def test_cg_splits_long_term_and_short_term_blocks(wb):
    ws = wb["CG"]
    banners = [c.value for row in ws.iter_rows(max_col=1) for c in row
               if isinstance(c.value, str) and c.value.startswith("Details of")]
    assert len(banners) == 2
    assert "Long Term" in banners[0]
    assert "Short Term" in banners[1]


def test_cg_date_columns_are_date_formatted(wb):
    ws = wb["CG"]
    date_cells = [c for row in ws.iter_rows() for c in row
                  if isinstance(c.value, str) and "DATEVALUE(" in c.value]
    assert date_cells
    for c in date_cells:
        assert c.number_format == presentation.DATE_FORMAT


# ---------------------------------------------------------------------------
# PL for Business (2026-07-19 "PL for Business" prompt) -- subtree walk,
# isolation, omit-vs-raise, and sheet ordering. All lightweight: these hit
# presentation.py's functions directly with hand-built entries, exactly like
# the Gate-5 _render()/_SIBLING_ENTRIES pattern above, so none of them need
# the full synthetic-book pipeline. Account names below are structural/generic
# (a partnership remuneration line + a handful of common business expense
# heads) -- no real amounts, names or PII of any kind.
# ---------------------------------------------------------------------------

_BIZ_SUBTREE = "Income/xBusiness Income"

_BIZ_ENTRIES = [
    ("Income/xBusiness Income/Remuneration from Partnership", "B10"),
    ("Income/xBusiness Income/Business Expenses/Bank Service Charge", "B11"),
    ("Income/xBusiness Income/Business Expenses/Staff Salary", "B12"),
    # Decoy: sounds business-related, but lives OUTSIDE the configured
    # subtree -- mirrors Expense/Professional Tax. Must never appear in a
    # filtered/rendered PL for Business result.
    ("Expense/Professional Tax", "B13"),
]


def test_business_subtree_walk_excludes_an_out_of_subtree_business_sounding_account():
    """The highest-value test here: subtree isolation is a plain path-prefix
    walk, not a keyword/name match. `Expense/Professional Tax` sounds
    business-related but is outside the configured subtree and must not leak
    into the filtered entries."""
    matches = presentation.resolve_business_entries(_BIZ_ENTRIES, _BIZ_SUBTREE)
    paths = [p for p, _ in matches]
    assert "Expense/Professional Tax" not in paths
    assert len(matches) == 3
    assert all(p.startswith(_BIZ_SUBTREE + "/") for p in paths)


def test_pl_for_business_renders_income_leaf_and_nested_expense_group_with_net_row():
    wb_ = openpyxl.Workbook()
    matches = presentation.resolve_business_entries(_BIZ_ENTRIES, _BIZ_SUBTREE)
    presentation.write_pl_for_business_sheet(wb_, matches, {"name": _Coord("A1")}, "period", "title")
    ws = wb_["PL for Business"]

    # Income leaf, the nested expense group, its subtotal, and the top-level
    # net row all render.
    assert _row_of(ws, "Remuneration from Partnership")
    assert _row_of(ws, "Business Expenses")
    assert _row_of(ws, "Total Business Expenses")
    net_row = _row_of(ws, "Net Business Income / (Loss)")

    # The decoy never appears anywhere on the rendered sheet.
    all_labels = [c.value for row in ws.iter_rows(min_col=1, max_col=1) for c in row]
    assert not any(isinstance(v, str) and "Professional Tax" in v for v in all_labels)

    # Net row is a plain SUM/addition of the top-level cells -- expenses are
    # already negative (HTML sign convention), so this nets correctly as-is.
    # It must not be rewritten as a subtraction.
    net_cell = ws.cell(row=net_row, column=_money_col(ws, net_row))
    assert net_cell.value.startswith("=")
    assert "-" not in net_cell.value
    assert net_cell.number_format == presentation.INR_FORMAT


# ---------------------------------------------------------------------------
# BS tally -- the year's net income carried over from IS
#
# A GnuCash balance sheet exported mid-year does NOT balance on its own: income
# and expenses are still sitting in the IS and have not been closed to capital,
# so Assets exceed Liabilities + Equity by exactly the net income. These tests
# prove the BS closes that gap and says so on the page.
#
# openpyxl never evaluates formulas, so the tally is proven by evaluating the
# rendered formulas here (they are simple enough: SUM ranges, + / - chains and
# cross-sheet refs) against known transcript numbers.
# ---------------------------------------------------------------------------

def _mini_eval(wb, sheet: str, coord: str, _seen=None):
    """Evaluate one cell of the rendered workbook, following formulas across
    sheets. Supports exactly the shapes the hierarchy writer emits."""
    import re as _re
    _seen = _seen or set()
    key = (sheet, coord)
    assert key not in _seen, f"formula cycle at {sheet}!{coord}"
    _seen = _seen | {key}

    v = wb[sheet][coord].value
    if v is None:
        return 0.0
    if not (isinstance(v, str) and v.startswith("=")):
        return float(v)

    expr = v[1:]

    def sum_range(m):
        sh, col, r1, r2 = m.group(1), m.group(2), int(m.group(3)), int(m.group(4))
        sh = (sh or sheet).strip("'!") or sheet
        total = sum(_mini_eval(wb, sh, f"{col}{r}", _seen) for r in range(r1, r2 + 1))
        return f"({total!r})"

    expr = _re.sub(r"SUM\((?:('[^']+'!))?([A-Z]+)(\d+):[A-Z]+(\d+)\)", sum_range, expr)

    def one_ref(m):
        sh = (m.group(1) or "").strip("'!") or sheet
        return f"({_mini_eval(wb, sh, m.group(2), _seen)!r})"

    expr = _re.sub(r"(?:('[^']+'!))?\$?([A-Z]{1,3}\$?\d+)\b", one_ref, expr)
    return float(eval(expr))  # noqa: S307 - test-local, operands are all numeric


#: Synthetic book. Assets 1000; Liabilities 200; Equity 500. Income 400,
#: expenses -100 (leaves carry the book's negative sign), so net income is 300
#: and 200 + 500 + 300 == 1000 exactly.
_IS_TALLY_ENTRIES = [("Income/Interest", "B4"), ("Expense/Bank Charges", "B5")]
_BS_TALLY_ENTRIES = [("Assets/Cash and Bank/BOB", "B4"),
                     ("Liabilities/Loans/Rent Payable", "B5"),
                     ("Equity/Capital Account", "B6")]


def _tally_workbook():
    wb_ = openpyxl.Workbook()
    wb_.remove(wb_.active)
    ist = wb_.create_sheet("IS_Transcript")
    ist["B4"], ist["B5"] = 400.0, -100.0
    bst = wb_.create_sheet("BS_Transcript")
    bst["B4"], bst["B5"], bst["B6"] = 1000.0, 200.0, 500.0

    _ws, net_ref = presentation.write_is_sheet(
        wb_, _IS_TALLY_ENTRIES, {"name": _Coord("A1")}, "period", "title")
    presentation.write_bs_sheet(
        wb_, _BS_TALLY_ENTRIES, {"name": _Coord("A1")}, "31-03-2026", "title",
        net_income_ref=net_ref)
    return wb_, net_ref


def test_is_sheet_closes_with_a_net_income_bottom_line():
    wb_, net_ref = _tally_workbook()
    ws = wb_["IS"]
    row = _row_of(ws, presentation.NET_INCOME_LABEL)
    assert net_ref, "write_is_sheet must report the bottom-line cell ref"
    assert ws[net_ref].row == row
    # Expenses already carry the book's negative sign, so the bottom line is a
    # plain addition -- never restated as a subtraction.
    assert "-" not in ws[net_ref].value
    assert _mini_eval(wb_, "IS", net_ref) == pytest.approx(300.0)


def test_balance_sheet_tallies_to_nil_once_net_income_is_carried_over():
    """The whole point: Assets 1000 vs Liabilities 200 + Equity 500 is short by
    300 until the year's income is brought in under capital."""
    wb_, _ = _tally_workbook()
    ws = wb_["BS"]

    # Without the carry-over the sheet is out by exactly the net income.
    assets = _mini_eval(wb_, "BS", _money_ref(ws, "Total Assets"))
    equity_before = _mini_eval(wb_, "BS", _money_ref(ws, "Total Equity"))
    liabilities = _mini_eval(wb_, "BS", _money_ref(ws, "Total Liabilities"))
    assert assets - (liabilities + equity_before) == pytest.approx(300.0)

    # With it, the sheet tallies and the difference row proves it.
    equity_after = _mini_eval(
        wb_, "BS", _money_ref(ws, "Total Equity including Current Year Income"))
    le_total = _mini_eval(wb_, "BS", _money_ref(ws, "Total Liabilities and Equity"))
    assert equity_after == pytest.approx(800.0)
    assert le_total == pytest.approx(1000.0)

    diff_label = "Difference (Assets less Liabilities and Equity) -- must be nil"
    assert _mini_eval(wb_, "BS", _money_ref(ws, diff_label)) == pytest.approx(0.0)


def test_bs_net_income_row_points_at_the_is_bottom_line_not_a_baked_number():
    """The carry-over must be a live cross-sheet reference: edit a leaf on the
    transcript and both sheets must move together."""
    wb_, net_ref = _tally_workbook()
    ws = wb_["BS"]
    ref = _money_ref(ws, f"Add: {presentation.NET_INCOME_LABEL} (per IS)")
    assert ws[ref].value == f"='IS'!{net_ref}"

    wb_["IS_Transcript"]["B4"] = 900.0          # income 400 -> 900
    assert _mini_eval(wb_, "BS", ref) == pytest.approx(800.0)
    diff_label = "Difference (Assets less Liabilities and Equity) -- must be nil"
    # Assets did not move, so the sheet is now legitimately out by 500 and the
    # difference row SAYS so rather than hiding it.
    assert _mini_eval(wb_, "BS", _money_ref(ws, diff_label)) == pytest.approx(-500.0)


def test_bs_omits_the_tally_when_there_is_no_is_sheet_to_source_it_from():
    """No IS bottom line must mean no tally rows -- never a tally that quietly
    treats the year's income as nil."""
    wb_ = openpyxl.Workbook()
    wb_.remove(wb_.active)
    bst = wb_.create_sheet("BS_Transcript")
    bst["B4"], bst["B5"], bst["B6"] = 1000.0, 200.0, 500.0
    presentation.write_bs_sheet(wb_, _BS_TALLY_ENTRIES, {"name": _Coord("A1")},
                                "31-03-2026", "title", net_income_ref="")
    labels = [c.value for row in wb_["BS"].iter_rows(min_col=1, max_col=1) for c in row]
    assert not any(isinstance(v, str) and "Total Liabilities and Equity" in v
                   for v in labels)


def test_bs_sections_are_found_by_name_not_by_position():
    """A book that orders Equity before Assets must still tally -- a positional
    lookup would silently compare the wrong sections."""
    wb_ = openpyxl.Workbook()
    wb_.remove(wb_.active)
    ist = wb_.create_sheet("IS_Transcript")
    ist["B4"], ist["B5"] = 400.0, -100.0
    bst = wb_.create_sheet("BS_Transcript")
    bst["B4"], bst["B5"], bst["B6"] = 500.0, 200.0, 1000.0

    _ws, net_ref = presentation.write_is_sheet(
        wb_, _IS_TALLY_ENTRIES, {"name": _Coord("A1")}, "period", "title")
    reordered = [("Equity/Capital Account", "B4"),
                 ("Liabilities/Loans/Rent Payable", "B5"),
                 ("Assets/Cash and Bank/BOB", "B6")]
    presentation.write_bs_sheet(wb_, reordered, {"name": _Coord("A1")},
                                "31-03-2026", "title", net_income_ref=net_ref)
    ws = wb_["BS"]
    diff_label = "Difference (Assets less Liabilities and Equity) -- must be nil"
    assert _mini_eval(wb_, "BS", _money_ref(ws, diff_label)) == pytest.approx(0.0)


def _money_ref(ws, label: str) -> str:
    """A1 ref of the amount cell on the row whose column-A label is `label`."""
    row = _row_of(ws, label)
    for cell in ws[row]:
        if cell.column > 1 and cell.value is not None:
            return cell.coordinate
    raise AssertionError(f"no amount cell on row {label!r}")


class _Coord:
    def __init__(self, coordinate):
        self.coordinate = coordinate


def _stub_model():
    import types
    return types.SimpleNamespace(
        salary=sch.SalarySchedule(), house_property=sch.HousePropertySchedule(),
        business=sch.BusinessSchedule(), capital_gains=sch.CapitalGainsSchedule(),
        other_sources=sch.OtherSourcesSchedule(), taxes_paid=sch.TaxesPaidSchedule(),
        # Present but not computed: write_statement_of_income renders the
        # interest-u/s-234 block unconditionally, because the tax ladder
        # references fixed coordinates inside it, so the stub must carry the
        # field even though this path exercises the not-computed branch.
        interest_234=sch.Interest234Schedule(),
    )


_STUB_ENTITY_LAYOUT = {k: _Coord(f"B{i}") for i, k in
                       enumerate(("name", "pan", "address", "dob", "status", "regime"), start=1)}
# 2026-07-20 on-page-totals change: comp_layout passed into
# build_presentation_layer/write_statement_of_income is now LEAF-only (the
# cross-section ladder -- GTI, Total Income, normal/special-CG split -- is
# computed on the page itself). The tax-slab "tail" (selected_liability,
# refund, new/old cess, new/old tax_before_cess) is supplied by a
# computation_tail_fn callback instead -- see _stub_computation_tail_fn.
_STUB_COMP_LAYOUT = {
    "salary": "'Computation'!B1", "hp": "'Computation'!B2", "business": "'Computation'!B3",
    "os": "'Computation'!B4", "cg_lt": "'Computation'!B5", "cg_st": "'Computation'!B6",
}
_STUB_TP_LAYOUT = {"total": "B1"}
_STUB_DED_LAYOUT = {"total": "B1"}
# 2026-07-21 on-page-totals change: write_statement_of_income now builds the
# full standard tax computation (slab/rebate/marginal-relief/surcharge/cess)
# directly on the page, so it needs the full Rules-sheet layout, not just
# round_ti_nearest -- a minimal one-slab-band table is enough to exercise the
# formula builders without asserting on their output (these tests don't
# assert on tax-block wiring).
_STUB_SLABS = {0: ("B100", "B101")}
_STUB_SURCHARGE = {0: ("B110", "B111")}
_STUB_RULES_LAYOUT = {
    "round_ti_nearest": _Coord("B99"),
    "round_tax_nearest": _Coord("B98"),
    "cess_rate": _Coord("B97"),
    "new_slabs": _STUB_SLABS, "old_slabs": _STUB_SLABS,
    "new_rebate_max_ti": _Coord("B102"), "new_rebate_max_amt": _Coord("B103"),
    "new_rebate_marginal": _Coord("B104"), "new_surcharge": _STUB_SURCHARGE,
    "old_rebate_max_ti": _Coord("B105"), "old_rebate_max_amt": _Coord("B106"),
    "old_rebate_marginal": _Coord("B107"), "old_surcharge": _STUB_SURCHARGE,
}
_STUB_CG_LAYOUT = {"special_tax_cell": "B120"}


def _stub_computation_tail_fn(page_layout):
    """Stand-in for write_computation_tail: ignores page_layout (these tests
    don't assert on Computation-sheet wiring) and returns a fixed
    comp_layout-shaped tail dict so write_statement_of_income's tax/cess/
    refund lines resolve."""
    return {
        "selected_liability": "'Computation'!B10",
        "refund": "'Computation'!B11",
        "new_cess": "'Computation'!B12",
        "old_cess": "'Computation'!B13",
        "new_tax_before_cess": "'Computation'!B14",
        "old_tax_before_cess": "'Computation'!B15",
    }


def _minimal_presentation_layer(is_entries, bs_entries, business_subtree):
    """Calls build_presentation_layer directly with a zeroed stub model and
    dummy layouts -- no salary/HP/CG activity, so only the always-emitted
    rows of Statement of Income execute. This proves sheet
    presence/order/omission without needing the full synthetic-book pipeline
    (parse_eguile/parse_gnucash/mapping/Form16/schedules)."""
    wb_ = openpyxl.Workbook()
    wb_.remove(wb_.active)
    presentation.build_presentation_layer(
        wb_, _stub_model(), _STUB_ENTITY_LAYOUT, _STUB_COMP_LAYOUT, {}, _STUB_TP_LAYOUT,
        _STUB_DED_LAYOUT, _STUB_RULES_LAYOUT, _STUB_CG_LAYOUT, is_entries, bs_entries, 10, "general",
        "2024-25", "AY 2025-26", _stub_computation_tail_fn, business_subtree=business_subtree,
    )
    return wb_


def test_pl_for_business_is_omitted_when_entity_has_no_business_subtree_configured():
    """No entity-level flag anywhere -- the decision is made fresh, per call,
    from the `business_subtree` argument alone."""
    wb_ = _minimal_presentation_layer(_BIZ_ENTRIES, [], business_subtree=None)
    assert "PL for Business" not in wb_.sheetnames
    assert wb_.sheetnames[:3] == ["Statement of Income", "BS", "IS"]


def test_pl_for_business_is_present_and_correctly_positioned_when_configured():
    wb_ = _minimal_presentation_layer(_BIZ_ENTRIES, [], business_subtree=_BIZ_SUBTREE)
    assert wb_.sheetnames[:4] == ["Statement of Income", "BS", "IS", "PL for Business"]
    assert "CG" not in wb_.sheetnames   # no CG activity in the stub model -- still omittable


def test_business_subtree_decision_holds_no_cross_year_or_cross_call_state():
    """Calling with business_subtree=None then immediately with a configured,
    matching subtree must not be influenced by the prior call -- proves the
    per-FY decision is not cached or persisted anywhere."""
    first = _minimal_presentation_layer(_BIZ_ENTRIES, [], business_subtree=None)
    second = _minimal_presentation_layer(_BIZ_ENTRIES, [], business_subtree=_BIZ_SUBTREE)
    assert "PL for Business" not in first.sheetnames
    assert "PL for Business" in second.sheetnames


def test_configured_but_missing_business_subtree_raises_not_silently_omits():
    """Gate 4: if business_subtree IS configured but this FY's data matches
    nothing under it, the code must raise -- never silently render a
    zero/omitted sheet as if it were simply 'no activity'."""
    no_match_entries = [("Expense/Professional Tax", "B13")]   # decoy only
    with pytest.raises(presentation.BusinessSubtreeError):
        presentation.resolve_business_entries(no_match_entries, _BIZ_SUBTREE)
    with pytest.raises(presentation.BusinessSubtreeError):
        _minimal_presentation_layer(no_match_entries, [], business_subtree=_BIZ_SUBTREE)


def test_business_subtree_config_field_round_trips_through_load_entities():
    entities = configs.load_entities(ROOT / "Data" / "itr" / "entities.example.yaml")
    assert entities["SYN-IND-BIZ"].business_subtree == "Income/xBusiness Income"
    # Unconfigured entities keep the field absent, not a hardcoded literal.
    assert entities["SYN-IND"].business_subtree is None


# ---------------------------------------------------------------------------
# Gate 7 -- widths and print setup on every sheet
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("sheet", ["Statement of Income", "IS", "BS", "CG"])
def test_every_presentation_sheet_has_explicit_widths_and_print_setup(wb, sheet):
    ws = wb[sheet]
    widths = {k: d.width for k, d in ws.column_dimensions.items() if d.width}
    assert "A" in widths and widths["A"] > 0, f"{sheet} ships with default column widths"
    assert len(widths) >= 3
    assert ws.print_area, f"{sheet} has no print area"
    # openpyxl round-trips paperSize as an int but exposes the constant as str.
    assert int(ws.page_setup.paperSize) == int(ws.PAPERSIZE_A4)
    assert ws.page_setup.fitToWidth == 1
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True
    assert ws.sheet_view.showGridLines is False


def test_cg_is_landscape_and_others_portrait(wb):
    assert wb["CG"].page_setup.orientation == "landscape"
    for sheet in ("Statement of Income", "IS", "BS"):
        assert wb[sheet].page_setup.orientation == "portrait"


def test_indian_number_format_is_preserved(wb):
    ws = wb["BS"]
    money = [c for row in ws.iter_rows(min_row=4, min_col=2) for c in row if c.value]
    assert money
    assert all(c.number_format == presentation.INR_FORMAT for c in money)
