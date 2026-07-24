"""
tests/skill_itr_workbook/test_apply_mapping_corrections.py -- round-trip
test for scripts/apply_mapping_corrections.py: write a Mapping Review sheet,
simulate a reviewer filling in Correction cells (one previously-unmapped row
gets a valid tag, one already-mapped row gets re-tagged, one row gets an
invalid tag), run the correction script, and check the row shows up
approved in the updated YAML -- and that the invalid one is reported, not
applied.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPTS = ROOT / "src" / "agents" / "skill_itr_workbook" / "scripts"
FIXTURES = Path(__file__).resolve().parent / "fixtures"

for p in (str(SCRIPTS), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

import parse_eguile as pe  # noqa: E402
import configs  # noqa: E402
import mapping as mapping_engine  # noqa: E402
import write_workbook as ww  # noqa: E402
import apply_mapping_corrections as amc  # noqa: E402
import fixture_gen  # noqa: E402

UNMAPPED_GUID = "ca85f6240bb66d5d38a0e4f40d908525"   # Assets/Misc Holding (unmapped)
RETAG_GUID = "11fc5a724efb9eb161ac039825b3e6dd"        # Income/Bank Interest, was OS_INTEREST_BANK
BAD_GUID = "84486cc24c1d61165711e1b881ef1bab"          # Equity/Capital Account


def _find_row(ws, guid):
    for row in ws.iter_rows(values_only=True):
        if row and row[-1] == guid:
            return row
    return None


def test_correction_round_trip(tmp_path):
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    entries = dict(loaded.entries)
    del entries[UNMAPPED_GUID]
    modified = configs.MappingLoadResult(entries=entries, warnings=[])
    result = mapping_engine.resolve_tree(tree, modified)

    wb = Workbook()
    ww.write_mapping_review_sheet(wb, tree, result.resolved, result.unmapped, entries)
    ws = wb["Mapping Review"]

    # Reviewer fills in three Correction cells.
    for row in ws.iter_rows():
        guid_cell = row[7]
        if guid_cell.value == UNMAPPED_GUID:
            row[6].value = "AL_CASH_BANK"          # previously unmapped -> valid tag
        elif guid_cell.value == RETAG_GUID:
            row[6].value = "OS_INTEREST_SB"         # already-mapped -> re-tagged
        elif guid_cell.value == BAD_GUID:
            row[6].value = "NOT_A_REAL_TAG"         # invalid -> must be rejected

    reviewed_xlsx = tmp_path / "reviewed.xlsx"
    wb.save(reviewed_xlsx)

    # apply_mapping_corrections never touches the committed fixture file --
    # write the "current mapping" copy to tmp_path first.
    mapping_copy = tmp_path / "syn_ind.mapping.yaml"
    mapping_copy.write_text((FIXTURES / "syn_ind.mapping.yaml").read_text(encoding="utf-8"), encoding="utf-8")

    output_yaml = tmp_path / "syn_ind.mapping.updated.yaml"
    applied, invalid, backup_path = amc.apply_corrections(str(mapping_copy), str(reviewed_xlsx), str(output_yaml))

    assert applied == 2
    assert len(invalid) == 1
    assert invalid[0][1] == BAD_GUID
    # An explicit output_yaml is a deliberate dry run: no backup is taken and
    # mapping_copy (the "live" file) must be left untouched.
    assert backup_path is None
    assert mapping_copy.read_text(encoding="utf-8") == (FIXTURES / "syn_ind.mapping.yaml").read_text(encoding="utf-8")

    updated = configs.load_mapping(output_yaml)
    assert updated.entries[UNMAPPED_GUID].tag == "AL_CASH_BANK"
    assert updated.entries[UNMAPPED_GUID].note == amc._APPROVED_NOTE
    assert updated.entries[UNMAPPED_GUID].suggested_by_llm is None
    assert updated.entries[RETAG_GUID].tag == "OS_INTEREST_SB"
    # The invalid correction must not have touched the Equity/Capital Account entry.
    assert updated.entries[BAD_GUID].tag == "EQUITY_CAPITAL"


def test_correction_default_writes_in_place_with_backup(tmp_path):
    """This is the regression test for the root-cause bug: real corrections
    were never reaching the run because the old CLI always required a
    manual review-and-rename step that, in practice, never happened. The
    default (no output_yaml) must now persist straight into mapping_file,
    with a timestamped backup of the pre-correction file, so an approved
    correction is live on the very next run without any manual step."""
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    entries = dict(loaded.entries)
    del entries[UNMAPPED_GUID]
    modified = configs.MappingLoadResult(entries=entries, warnings=[])
    result = mapping_engine.resolve_tree(tree, modified)

    wb = Workbook()
    ww.write_mapping_review_sheet(wb, tree, result.resolved, result.unmapped, entries)
    ws = wb["Mapping Review"]
    for row in ws.iter_rows():
        if row[7].value == UNMAPPED_GUID:
            row[6].value = "AL_CASH_BANK"

    reviewed_xlsx = tmp_path / "reviewed.xlsx"
    wb.save(reviewed_xlsx)

    mapping_copy = tmp_path / "syn_ind.mapping.yaml"
    original_text = (FIXTURES / "syn_ind.mapping.yaml").read_text(encoding="utf-8")
    mapping_copy.write_text(original_text, encoding="utf-8")

    applied, invalid, backup_path = amc.apply_corrections(str(mapping_copy), str(reviewed_xlsx))

    assert applied == 1
    assert invalid == []
    assert backup_path is not None
    assert Path(backup_path).read_text(encoding="utf-8") == original_text

    # The live mapping file itself now carries the correction -- no manual
    # rename step required.
    reloaded = configs.load_mapping(mapping_copy)
    assert reloaded.entries[UNMAPPED_GUID].tag == "AL_CASH_BANK"
    assert reloaded.entries[UNMAPPED_GUID].note == amc._APPROVED_NOTE

    # And it survives a fresh resolve on the next run, taking precedence
    # over whatever heuristic/inherited tag would otherwise have applied --
    # this is the "approved always outranks heuristic" guarantee, proven
    # end-to-end through the real persistence path rather than asserted in
    # the abstract.
    known_paths = {n.guid: n.path for n in tree.all_nodes() if n.guid}
    reloaded_for_run = configs.load_mapping(mapping_copy, known_paths=known_paths)
    next_run_result = mapping_engine.resolve_tree(tree, reloaded_for_run)
    assert next_run_result.resolved[UNMAPPED_GUID].tag == "AL_CASH_BANK"


# ---------------------------------------------------------------------------
# Path-drift self-healing (2026-07-23): refresh-on-write, not
# auto-write-on-run. GUID is identity; `path:` is descriptive metadata that
# self-heals only when a correction/refresh run actually writes the mapping
# file -- never silently, and never for a GUID that's genuinely gone from
# the tree.
# ---------------------------------------------------------------------------

_DRIFT_MAPPING_YAML = (
    "- guid: g1\n  path: Old/Path/One\n  tag: AL_CASH_BANK\n  note: some note\n"
    "- guid: g2\n  path: Stable/Path/Two\n  tag: AL_SECURITIES\n"
    "- guid: g3\n  path: Deleted/Path/Three\n  tag: PERSONAL\n"
)


def _write_drift_mapping(tmp_path):
    mapping_file = tmp_path / "drift.mapping.yaml"
    mapping_file.write_text(_DRIFT_MAPPING_YAML, encoding="utf-8")
    return mapping_file


def test_refresh_drifted_paths_refreshes_only_guids_present_in_known_paths():
    from configs import MappingEntry
    entries = {
        "g1": MappingEntry(guid="g1", path="Old/Path/One", tag="AL_CASH_BANK"),
        "g2": MappingEntry(guid="g2", path="Stable/Path/Two", tag="AL_SECURITIES"),
        "g3": MappingEntry(guid="g3", path="Deleted/Path/Three", tag="PERSONAL"),
    }
    known_paths = {"g1": "New/Path/One", "g2": "Stable/Path/Two"}  # g3 absent -- deleted
    refreshed, count = amc.refresh_drifted_paths(entries, known_paths)
    assert count == 1
    assert refreshed["g1"].path == "New/Path/One"
    assert refreshed["g2"].path == "Stable/Path/Two"       # unchanged (no drift)
    assert refreshed["g3"].path == "Deleted/Path/Three"    # untouched (guid missing from tree)
    # original dict must not be mutated.
    assert entries["g1"].path == "Old/Path/One"


def test_refresh_paths_writes_backup_and_refreshes_drifted_path(tmp_path):
    mapping_file = _write_drift_mapping(tmp_path)
    original_text = mapping_file.read_text(encoding="utf-8")
    known_paths = {"g1": "New/Path/One", "g2": "Stable/Path/Two"}  # g3 absent -- deleted

    count, backup_path = amc.refresh_paths(str(mapping_file), known_paths)

    assert count == 1
    assert backup_path is not None
    assert Path(backup_path).read_text(encoding="utf-8") == original_text

    updated = configs.load_mapping(mapping_file)
    assert updated.entries["g1"].path == "New/Path/One"
    assert updated.entries["g1"].tag == "AL_CASH_BANK"      # rest of the entry untouched
    assert updated.entries["g2"].path == "Stable/Path/Two"
    assert updated.entries["g3"].path == "Deleted/Path/Three"


def test_refresh_paths_is_true_noop_when_nothing_drifted(tmp_path):
    mapping_file = _write_drift_mapping(tmp_path)
    original_text = mapping_file.read_text(encoding="utf-8")
    # known_paths matches every stored path exactly -- nothing to refresh.
    known_paths = {"g1": "Old/Path/One", "g2": "Stable/Path/Two"}

    count, backup_path = amc.refresh_paths(str(mapping_file), known_paths)

    assert count == 0
    assert backup_path is None
    # A true no-op: no backup file written, mapping file byte-for-byte unchanged.
    assert mapping_file.read_text(encoding="utf-8") == original_text
    assert list(tmp_path.glob("*.bak-*")) == []


def test_refresh_paths_never_rewrites_missing_guid_entry_which_still_warns_loudly(tmp_path):
    mapping_file = _write_drift_mapping(tmp_path)
    known_paths = {"g1": "New/Path/One", "g2": "Stable/Path/Two"}  # g3 (deleted) absent

    amc.refresh_paths(str(mapping_file), known_paths)

    reloaded = configs.load_mapping(mapping_file, known_paths=known_paths)
    assert reloaded.entries["g3"].path == "Deleted/Path/Three"    # never rewritten
    assert any("g3" in w and "NOT FOUND" in w for w in reloaded.warnings)
    # And the (now-healed) rename no longer produces a warning at all.
    assert not any("g1" in w for w in reloaded.warnings)


def test_refresh_paths_leaves_unrelated_entries_byte_stable(tmp_path):
    mapping_file = _write_drift_mapping(tmp_path)
    before = configs.load_mapping(mapping_file)
    known_paths = {"g1": "New/Path/One", "g2": "Stable/Path/Two"}

    amc.refresh_paths(str(mapping_file), known_paths)

    after = configs.load_mapping(mapping_file)
    # g2 and g3 are entirely unaffected by g1's refresh -- same value, field for field.
    assert after.entries["g2"] == before.entries["g2"]
    assert after.entries["g3"] == before.entries["g3"]
    # note on g1 (unrelated to path) survives the refresh too.
    assert after.entries["g1"].note == "some note"


def test_apply_corrections_heals_drifted_paths_for_free_on_a_correction_run(tmp_path):
    """apply_corrections() must refresh every drifted path when it writes
    the mapping file for a correction -- not only the entries being
    corrected -- since read_current_paths() sees the whole reviewed
    workbook, not just the corrected rows."""
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    entries = dict(loaded.entries)
    del entries[UNMAPPED_GUID]
    modified = configs.MappingLoadResult(entries=entries, warnings=[])
    result = mapping_engine.resolve_tree(tree, modified)

    wb = Workbook()
    ww.write_mapping_review_sheet(wb, tree, result.resolved, result.unmapped, entries)
    ws = wb["Mapping Review"]
    for row in ws.iter_rows():
        if row[7].value == UNMAPPED_GUID:
            row[6].value = "AL_CASH_BANK"

    reviewed_xlsx = tmp_path / "reviewed.xlsx"
    wb.save(reviewed_xlsx)

    mapping_copy = tmp_path / "syn_ind.mapping.yaml"
    # Simulate a stale path on the RETAG_GUID entry -- as if that account
    # had been renamed in GnuCash since this mapping file was last written.
    entries_before = dict(configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml").entries)
    stale_path = "Income/Old Bank Interest Name (renamed)"
    real_path = entries_before[RETAG_GUID].path
    entries_before[RETAG_GUID] = configs.MappingEntry(
        guid=RETAG_GUID, path=stale_path, tag=entries_before[RETAG_GUID].tag,
        flags=entries_before[RETAG_GUID].flags, note=entries_before[RETAG_GUID].note,
        suggested_by_llm=entries_before[RETAG_GUID].suggested_by_llm,
    )
    from configs import dump_mapping_entries
    mapping_copy.write_text(dump_mapping_entries(list(entries_before.values())), encoding="utf-8")

    applied, invalid, backup_path = amc.apply_corrections(str(mapping_copy), str(reviewed_xlsx))

    assert applied == 1
    assert backup_path is not None
    updated = configs.load_mapping(mapping_copy)
    # The corrected entry applies as usual...
    assert updated.entries[UNMAPPED_GUID].tag == "AL_CASH_BANK"
    # ...and the untouched RETAG_GUID entry's stale path healed for free,
    # back to the tree's current path, even though it wasn't corrected.
    assert updated.entries[RETAG_GUID].path == real_path
    assert updated.entries[RETAG_GUID].path != stale_path


def test_main_refresh_paths_cli_flag_reports_no_drift_as_noop(tmp_path, monkeypatch, capsys):
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    result = mapping_engine.resolve_tree(tree, loaded)

    wb = Workbook()
    ww.write_mapping_review_sheet(wb, tree, result.resolved, result.unmapped, dict(loaded.entries))
    reviewed_xlsx = tmp_path / "reviewed.xlsx"
    wb.save(reviewed_xlsx)

    mapping_copy = tmp_path / "syn_ind.mapping.yaml"
    mapping_copy.write_text((FIXTURES / "syn_ind.mapping.yaml").read_text(encoding="utf-8"), encoding="utf-8")

    monkeypatch.setattr(sys, "argv", ["amc", str(mapping_copy), str(reviewed_xlsx), "--refresh-paths"])
    rc = amc.main()
    assert rc == 0
    captured = capsys.readouterr()
    assert "nothing to refresh" in captured.out
    assert list(tmp_path.glob("*.bak-*")) == []
