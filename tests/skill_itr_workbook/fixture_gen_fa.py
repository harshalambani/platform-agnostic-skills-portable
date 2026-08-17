"""
fixture_gen_fa.py -- synthetic INDmoney "consolidated tax report" fixture
generator for the ITR Workbook skill's foreign-asset tests.

NO real family data appears here: the client name, PAN, broker, account
number, entity names, addresses and every amount are made up for testing.

What IS faithful is the *layout*, transcribed cell-by-cell from a real
report so the parser is exercised against the real shape rather than a
description of it:

  - three blank rows, then ``Client Name:`` / ``Client PAN:`` /
    (``Financial Year:`` or ``Report as on Date:``) label-value rows, a
    title row, and only then the real header -- so headers must be found
    by label scan, never a fixed row index;
  - two-row headers with horizontally merged parents (Schedule FA's
    "Gross amount paid/credited" over "Nature of Amount"/"Amount";
    Form 67's "Tax paid outside India" over "Amount"/"Rate");
  - Schedule FSI's vertically merged country block over four fixed
    head-of-income rows, plus its (a)..(f) column-letter row;
  - country cells carrying an embedded newline ("002\\nUnited States of
    America");
  - placeholder glyphs where a reader expects a number -- "-", "#", "*" --
    and cells holding *instructional prose* rather than a value
    ("To be filled by user basis Individual slab rates"). A parser that
    calls float() on these blows up; one that coerces them silently is
    worse.

Two deliberate non-ties are encoded, because they are correct and a
parser that "fixes" them would be wrong:

  - Schedule FA reports the **calendar** year and the Dividend/Interest
    sheets report the **financial** year, so FA's gross-amount figures do
    not match the Dividend/Interest totals;
  - Schedule FSI's "Other Sources" income does tie to the FY dividend
    plus FY interest totals, and that tie is worth asserting.

Schedule FA's A3 block also carries the same entity on more than one row
(one row per acquisition lot). Deduplicating by entity name would destroy
holdings, so the fixture keeps a duplicated entity to catch it.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

CLIENT_NAME = "SYNTHETIC TEST USER"
CLIENT_PAN = "AAAPA0000A"
FINANCIAL_YEAR = "2025-26"
REPORT_AS_ON = "31st December, 2025"

BROKER = "SynBroker LLC"
COUNTRY = "United States of America"
COUNTRY_CODE = "002"

# Calendar-year (Schedule FA) gross amounts credited, per lot.
FA_GROSS_LOT_1 = 1500
FA_GROSS_LOT_2 = 75
FA_GROSS_LOT_3 = 2500

# Financial-year totals -- deliberately NOT equal to the calendar-year
# figures above. See the module docstring.
FY_DIVIDEND_TOTAL = 4200
FY_INTEREST_TOTAL = 3
FSI_OTHER_SOURCES_INCOME = FY_DIVIDEND_TOTAL + FY_INTEREST_TOTAL  # 4203
FSI_TAX_PAID_OUTSIDE = 735

# Prose that occupies value cells in the real report.
PROSE_SLAB = "(To be filled by user basis Individual slab rates)"
PROSE_EFFECTIVE_RATE = "This is calculated based on your effective tax rate."
PROSE_NOT_APPLICABLE = (
    "Not applicable for salaried individuals: Enter 0 (zero). "
    "For business owners: Please contact your tax advisor."
)
PROSE_LOWER_OF = (
    "Lower of \n 1. Tax payable under normal provisions in India or \n "
    "2. Tax paid outside India"
)
PROSE_FROM_FSI = "From 'Schedule FSI' sheet: Total of column (e)"

DISCLAIMERS = [
    "1. This statement has been prepared on a best-effort basis using trade and "
    "account information available at the time of report generation.",
    "4. The tax document is prepared as per the US timezone. So, the numbers "
    "presented in this statement may not be equal to other statements which are "
    "prepared as per Indian timezone.",
    "6. Corporate Actions covered in data/report are limited to only Stock Splits "
    "and Reverse Stock Splits",
    "7. SBI TT Buy Rate have been used for conversion from USD to INR.",
    "9. The Peak Value and Closing Value of delisted stock will not be included in "
    "the reporting.",
    "16. This tax report excludes INDglobal dividends as they are credited directly "
    "to your bank. Please add these manually when filing your return.",
]


def _write_row(ws: Worksheet, row: int, values: list) -> None:
    for col, value in enumerate(values, start=1):
        if value is not None:
            ws.cell(row=row, column=col, value=value)


def _preamble(ws: Worksheet, third_label: str, third_value: str, title: str) -> int:
    """Rows 1-3 blank, then the label-value block, blank, title, blank.

    Returns the next free row.
    """
    _write_row(ws, 4, ["Client Name:", CLIENT_NAME])
    _write_row(ws, 5, ["Client PAN:", CLIENT_PAN])
    _write_row(ws, 6, [third_label, third_value])
    _write_row(ws, 8, [title])
    return 10


def _disclaimers(ws: Worksheet, row: int, width: int) -> None:
    ws.cell(row=row, column=1, value="Disclaimer:-")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    for offset, text in enumerate(DISCLAIMERS, start=1):
        target = row + offset
        ws.cell(row=target, column=1, value=text)
        ws.merge_cells(
            start_row=target, start_column=1, end_row=target, end_column=width
        )


def _build_schedule_fa(ws: Worksheet) -> None:
    row = _preamble(ws, "Report as on Date:", REPORT_AS_ON, "Schedule FA (Foreign Assets)")

    _write_row(
        ws,
        row,
        [
            "A2. Details of Foreign Custodial Accounts held (including any beneficial "
            "interest) at any time during the calendar year ending as on "
            f"{REPORT_AS_ON}."
        ],
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)

    header = row + 2
    _write_row(
        ws,
        header,
        [
            "Country Name",
            "Country Code",
            "Name of financial institution",
            "Address of financial institution",
            "ZIP Code",
            "Account Number",
            "Status",
            "Account Opening date",
            "Peak Balance During the Period",
            "Closing Balance",
            "Gross amount paid/credited to the account during the period",
        ],
    )
    # The gross-amount parent spans two sub-columns on the row beneath.
    ws.merge_cells(start_row=header, start_column=11, end_row=header, end_column=12)
    _write_row(ws, header + 1, [None] * 10 + ["Nature of Amount", "Amount"])
    _write_row(
        ws,
        header + 2,
        [
            COUNTRY,
            COUNTRY_CODE,
            BROKER,
            "1 Test Street, Test City, NJ 07000, USA",
            "07000",
            "SYN000001",
            "Beneficial owner",
            "2024-04-17",
            5000,
            800,
            "Interest",
            9,
        ],
    )

    row = header + 4
    _write_row(
        ws,
        row,
        [
            "A3. Details of Foreign Equity and Debt Interest held (including any "
            "beneficial interest) in any entity at any time during the calendar year "
            f"ending as on {REPORT_AS_ON}."
        ],
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)

    header = row + 2
    _write_row(
        ws,
        header,
        [
            "Country Name",
            "Country Code",
            "Name of the entity",
            "Address of the entity",
            "ZIP Code",
            "Nature of entity",
            "Date of acquiring interest",
            "Initial value of investment",
            "Peak value of investment",
            "Closing Balance",
            "Total gross amount paid/credited to the holding during the period",
            "Total gross proceeds from sale or redemption of investment during the "
            "period",
            "Broker Name",
        ],
    )

    # Two lots of the SAME entity, then a third holding. Do not dedupe.
    lots = [
        ("SynCo Alpha Ltd (ADR)", "9 Alpha Road, Alpha City, Testland", "311121",
         "2021-11-22", 100000, 160000, 120000, FA_GROSS_LOT_1, 0),
        ("SynCo Alpha Ltd (ADR)", "9 Alpha Road, Alpha City, Testland", "311121",
         "2021-11-22", 5000, 8000, 6000, FA_GROSS_LOT_2, 0),
        ("SynCo Beta Inc.", "42 Beta Drive, Beta City, CA, United States", "92121",
         "2024-12-02", 110000, 140000, 130000, FA_GROSS_LOT_3, 0),
    ]
    for offset, lot in enumerate(lots, start=1):
        name, address, zip_code, acquired, initial, peak, closing, gross, proceeds = lot
        _write_row(
            ws,
            header + offset,
            [
                COUNTRY,
                COUNTRY_CODE,
                name,
                address,
                zip_code,
                "Company",
                acquired,
                initial,
                peak,
                closing,
                gross,
                proceeds,
                BROKER,
            ],
        )

    _disclaimers(ws, header + len(lots) + 2, 13)


def _build_schedule_fsi(ws: Worksheet) -> None:
    row = _preamble(
        ws, "Financial Year:", FINANCIAL_YEAR, "Schedule FSI (Foreign Sourced Income)"
    )
    _write_row(
        ws,
        row,
        [
            "Details of income from outside India and tax relief (applicable for "
            "resident Indians only)"
        ],
    )

    header = row + 2
    _write_row(
        ws,
        header,
        [
            "Sl. No.",
            "Country Code",
            "Tax Identification Number",
            "Sl. No.",
            "Head of Income",
            "Income (INR)",
            "Tax paid outside India (INR)",
            "Tax payable in India (INR)",
            "Tax relief [Lower of (c) or (d)]",
            "Relevant article of DTAA where relief is claimed",
        ],
    )
    _write_row(ws, header + 1, [None, None, None, None, "(a)", "(b)", "(c)", "(d)",
                                "(e)", "(f)"])

    # Four fixed heads of income. Placeholders "-" and "#" and "*" sit where a
    # naive reader expects numbers.
    heads = [
        ("i", "Salary", "-", "-", "-", "-", None),
        ("ii", "House Property", "-", "-", "-", "-", None),
        ("iii", "Capital Gains", 0, "-", "#", "-", None),
        ("iv", "Other Sources", FSI_OTHER_SOURCES_INCOME, FSI_TAX_PAID_OUTSIDE, "*",
         "Lower of (c) or (d)", '"10, 25"'),
    ]
    first = header + 2
    for offset, head in enumerate(heads):
        sl, name, income, tax_paid, payable, relief, article = head
        _write_row(
            ws,
            first + offset,
            [None, None, None, sl, name, income, tax_paid, payable, relief, article],
        )

    # Country identity is merged vertically across all four head rows, and the
    # cell itself carries an embedded newline.
    last = first + len(heads) - 1
    ws.cell(row=first, column=1, value=1)
    ws.cell(row=first, column=2, value=f"{COUNTRY_CODE}\n{COUNTRY}")
    ws.cell(
        row=first,
        column=3,
        value="Your TIN in USA (if available) or your Passport Number",
    )
    for col in (1, 2, 3):
        ws.merge_cells(start_row=first, start_column=col, end_row=last, end_column=col)

    _write_row(
        ws,
        last + 1,
        ["Total", None, None, None, None, FSI_OTHER_SOURCES_INCOME,
         FSI_TAX_PAID_OUTSIDE],
    )
    _disclaimers(ws, last + 3, 10)


def _build_form_67(ws: Worksheet) -> None:
    row = _preamble(ws, "Financial Year:", FINANCIAL_YEAR, "Form 67")

    header = row + 1
    _write_row(
        ws,
        header,
        [
            "Sr No",
            "Name of the country",
            "Source of Income",
            "Income from outside India",
            "Tax paid outside India",
            None,
            "Tax payable on such income under normal provisions in India",
            "Tax payable on such income under section 115JB/JC",
            "Credit claimed under section 90/90A",
            None,
            None,
            "Credit claimed under section 91 amount",
            "Total foreign tax credit claimed",
        ],
    )
    # "Tax paid outside India" parents Amount/Rate on the row beneath.
    ws.merge_cells(start_row=header, start_column=5, end_row=header, end_column=6)
    _write_row(
        ws,
        header + 1,
        [
            None,
            None,
            None,
            None,
            "Amount",
            "Rate",
            PROSE_SLAB,
            PROSE_SLAB,
            "Article No. of Double Taxation Avoidance Agreements",
            "Rate of tax as per Double Taxation Avoidance Agreements",
            "Amount",
        ],
    )

    rows = [
        (1, "Dividend", FY_DIVIDEND_TOTAL, 732, "17.43%"),
        (2, "Interest", FY_INTEREST_TOTAL, "-", "-"),
    ]
    for offset, entry in enumerate(rows, start=2):
        sr, source, income, tax_amount, rate = entry
        _write_row(
            ws,
            header + offset,
            [
                sr,
                "USA",
                source,
                income,
                tax_amount,
                rate,
                PROSE_EFFECTIVE_RATE,
                PROSE_NOT_APPLICABLE,
                "*10, 25*",
                "25%",
                PROSE_LOWER_OF,
                "-",
                PROSE_LOWER_OF,
            ],
        )

    _disclaimers(ws, header + len(rows) + 3, 13)


def _build_schedule_tr(ws: Worksheet) -> None:
    row = _preamble(ws, "Financial Year:", FINANCIAL_YEAR, "Schedule TR")
    _write_row(ws, row, ["Tax Relief Claimed"])

    header = row + 2
    _write_row(
        ws,
        header,
        [
            "Country Code",
            "Tax Identification Number",
            "Total taxes paid outside India",
            "Total tax relief available",
            "Tax relief claimed under section",
            "Total Tax relief available in respect of country where DTAA is "
            "applicable (section 90/90A) (Part of total of 1(d))",
            "Total Tax relief available in respect of country where DTAA is not "
            "applicable (section 91) (Part of total of 1(d))",
            "Whether any tax paid outside India has been refunded or credited",
        ],
    )
    _write_row(
        ws,
        header + 1,
        ["1(a)", "1(b)", "1(c)", "1(d)", "1(e)", "2", "3", "4"],
    )
    _write_row(
        ws,
        header + 2,
        [
            f"{COUNTRY_CODE}\n {COUNTRY}",
            "Your TIN in USA (if available) or your Passport Number",
            FSI_TAX_PAID_OUTSIDE,
            PROSE_FROM_FSI,
            90,
            PROSE_FROM_FSI,
            0,
            "No",
        ],
    )
    _disclaimers(ws, header + 4, 8)


def _build_dividend(ws: Worksheet) -> None:
    row = _preamble(ws, "Financial Year:", FINANCIAL_YEAR, "Dividend Summary")

    _write_row(ws, row, ["Particulars", "Amount (₹)", "ITR Section"])
    _write_row(ws, row + 1, ["Indian Stocks", 0, "Schedule OS U/s Dividend"])
    _write_row(ws, row + 2, ["US Stocks", FY_DIVIDEND_TOTAL,
                             "Schedule OS U/s Dividend"])
    _write_row(ws, row + 3, ["Mutual Funds", 0, "Schedule OS U/s Dividend"])
    _write_row(ws, row + 4, ["Buybacks", 0, "Schedule OS u/s Dividend 2(22)(f)"])
    _write_row(ws, row + 5, ["Total", FY_DIVIDEND_TOTAL])

    row += 7
    _write_row(ws, row, ["Quaterly Break - Up"])
    _write_row(
        ws,
        row + 1,
        [
            "Particulars",
            "1-Apr to 15-Jun",
            "16-Jun to 15-Sep",
            "16-Sep to 15-Dec",
            "16-Dec to 15-Mar",
            "16-Mar to 31-Mar",
        ],
    )
    # Quarterly cells use "-" for nil, and the five buckets must sum to the total.
    _write_row(ws, row + 2, ["Dividend Income", "-", 2200, 700, 650, 650])
    _write_row(ws, row + 3, ["Dividend u/s 2(22)(f)", "-", "-", "-", "-", "-"])

    row += 5
    _write_row(ws, row, ["Indian Stocks Dividend"])
    _write_row(ws, row + 1, ["Name of Stock", "Ex-Date", "Amount (₹)"])
    _write_row(ws, row + 2, ["-", "-", "-"])
    _write_row(ws, row + 3, ["Total", None, 0])

    row += 5
    _write_row(ws, row, ["Dividends From US Stocks"])
    _write_row(
        ws,
        row + 1,
        [
            "Name of the company",
            "Date Received",
            "Exchange Rates (SBI TT Buying Rates)",
            "Amount",
            None,
            "Tax deducted",
            None,
            "Broker Name",
        ],
    )
    ws.merge_cells(start_row=row + 1, start_column=4, end_row=row + 1, end_column=5)
    ws.merge_cells(start_row=row + 1, start_column=6, end_row=row + 1, end_column=7)
    _write_row(ws, row + 2, [None, None, None, "US$", "INR", "US$", "INR"])

    payments = [
        ("SynCo Beta Inc. Common Stock", "2025-06-26", 85.1, 7.73, 2200, 1.93, 380),
        ("SynCo Alpha Ltd American Depositary Shares", "2025-07-10", 85.7, 8.96, 700,
         0, 0),
        ("SynCo Alpha Ltd American Depositary Shares", "2025-10-10", 88.0, 7.39, 650,
         0, 0),
        ("SynCo Beta Inc. Common Stock", "2025-12-24", 89.0, 7.30, 650, 0, 0),
    ]
    for offset, payment in enumerate(payments, start=3):
        name, date, rate, usd, inr, tax_usd, tax_inr = payment
        _write_row(
            ws,
            row + offset,
            [name, date, rate, usd, inr, tax_usd, tax_inr, BROKER],
        )
    _write_row(ws, row + 3 + len(payments), ["Total", None, None, None,
                                             FY_DIVIDEND_TOTAL])

    _disclaimers(ws, row + 6 + len(payments), 14)


def _build_interest(ws: Worksheet) -> None:
    row = _preamble(ws, "Financial Year:", FINANCIAL_YEAR, "Interest")
    _write_row(ws, 8, ["Total Interest received from US Stocks", FY_INTEREST_TOTAL])

    _write_row(ws, row, ["Transactions"])
    _write_row(
        ws,
        row + 1,
        [
            "Date Received",
            "Exchange Rates (SBI TT Buying Rates)",
            "Amount",
            None,
            "Broker Name",
        ],
    )
    ws.merge_cells(start_row=row + 1, start_column=3, end_row=row + 1, end_column=4)
    _write_row(ws, row + 2, [None, None, "US$", "INR"])

    entries = [
        ("2025-08-04", 87.15, 0.01, 1),
        ("2025-09-03", 87.70, 0.01, 1),
        ("2025-10-02", 88.35, 0.01, 1),
    ]
    for offset, entry in enumerate(entries, start=3):
        _write_row(ws, row + offset, list(entry))
    _write_row(ws, row + 3 + len(entries),
               ["Total", None, 0.03, FY_INTEREST_TOTAL])


def build_workbook() -> Workbook:
    """Return a synthetic INDmoney consolidated tax report workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet order matches the real report, including the sheets this skill does
    # not read -- lookup must be by name, not position.
    builders = {
        "LTCG": None,
        "STCG": None,
        "F&O": None,
        "Intraday": None,
        "Dividend": _build_dividend,
        "Interest": _build_interest,
        "Bonds & SGB": None,
        "Schedule FA": _build_schedule_fa,
        "Schedule FSI": _build_schedule_fsi,
        "Form 67": _build_form_67,
        "Schedule TR": _build_schedule_tr,
    }
    for name, builder in builders.items():
        ws = wb.create_sheet(name)
        if builder is None:
            _preamble(ws, "Financial Year:", FINANCIAL_YEAR, name)
        else:
            builder(ws)
    return wb


def write_fixture(path) -> None:
    build_workbook().save(str(path))


if __name__ == "__main__":  # pragma: no cover
    import pathlib

    target = (
        pathlib.Path(__file__).parent / "fixtures" / "syn_indmoney_tax_report.xlsx"
    )
    write_fixture(target)
    print(f"wrote {target}")
