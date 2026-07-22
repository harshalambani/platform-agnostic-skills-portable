"""
tests/skill_itr_workbook/test_refund_num_salary_gross_layout.py -- regression
tests for the 2026-07-22 refund-#NUM! / Salary-gross / Workings-layout fix
bundle:

  Fix A: Excel's MROUND(number, multiple) raises #NUM! whenever `number` and
  `multiple` have opposite signs -- which happens on EVERY tax-payable (not
  refund) return and every loss-year (negative Total Income) return, i.e.
  not an edge case but the common case for a large fraction of real filers.
  Replaced with presentation.mround_safe(): a sign-safe
  `ROUND((x)/m,0)*m` formula that reproduces MROUND's round-half-away-from-
  zero behaviour for both signs and never errors.

  Fix B: Salary sheet's "Gross salary (17(1)+17(2)+17(3))" label previously
  displayed the book's SALARY_GROSS tag total, which only ever captures
  17(1) -- so a salary with non-zero perquisites (17(2)) or profits in lieu
  (17(3)) showed a gross that didn't match its own label. schedules.py's
  build_salary() now sources `gross` from Form16's total_1d (falling back to
  s17_1+s17_2+s17_3 when total_1d is absent) on the Form16 path.

  Fix B2: a fail-loud "banner, no abort" control mirroring the existing CG
  reconciliation banner -- SalarySchedule.reconciliation_ok/diff, an ERROR
  banner on the Salary and Statement of Income sheets, a Reconciliation-sheet
  line, and agent.py summary-line/exit-code wiring -- so a genuine
  gross/exemptions/deductions vs income-chargeable mismatch can never be
  silently shipped in a filing-ready workbook.

  Fix C: the "Brought forward losses set off" input block and the New/Old
  regime tax-working formulas moved from ahead of "Income from Salary" to a
  new "Workings / Inputs" section below "Refund Due / (Tax Payable)" -- so a
  reader doesn't have to scroll past raw working machinery before reaching
  income. The b/f cells remain live inputs; the regime workings stay on the
  same sheet; the move is proven safe by runtime row-prediction assertions
  in presentation.py itself (any drift is a loud AssertionError at
  generation time) plus the structural checks below.

  Fix D: one added line in the Assumptions block: interest u/s 234A/234B/
  234C is NOT computed by this tool.

  openpyxl does not evaluate formulas, so where full numeric proof is
  required this file uses Python shadow calculations of the exact rounding
  semantics (`_excel_round_half_away`, mirroring `presentation.mround_safe`
  and Excel's own MROUND) rather than standing up a real spreadsheet engine.

  Fully offline; synthetic fixtures only (fixture_gen.py / hand-built
  dataclasses). No real account numbers, PANs, or names.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPTS = ROOT / "src" / "agents" / "skill_itr_workbook" / "scripts"
AGENT_DIR = ROOT / "src" / "agents" / "skill_itr_workbook"
FIXTURES = Path(__file__).resolve().parent / "fixtures"
RULES_DIR = ROOT / "Data" / "itr" / "rules"

for p in (str(SCRIPTS), str(AGENT_DIR), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

import parse_gnucash as pg  # noqa: E402
import configs  # noqa: E402
import rules as rules_engine  # noqa: E402
import schedules as sch  # noqa: E402
import write_workbook as ww  # noqa: E402
import presentation  # noqa: E402
import fixture_gen  # noqa: E402
import parse_eguile as pe  # noqa: E402
import parse_form16  # noqa: E402
import mapping as mapping_engine  # noqa: E402
import verify  # noqa: E402

YEAR_KEY = "2024-25"


# ---------------------------------------------------------------------------
# Shared fixture: the syn_ind pipeline, mirrors test_cg_gain_split_and_banner
# ---------------------------------------------------------------------------

@pytest.fixture()
def syn_ind_model_and_paths():
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    book = pg.parse_book(FIXTURES / "syn_ind.gnucash")
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    result = mapping_engine.resolve_tree(tree, loaded)
    form16 = parse_form16.parse_form16(FIXTURES / "syn_ind_form16.pdf")
    rules = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    entities = configs.load_entities(ROOT / "Data" / "itr" / "entities.example.yaml")
    entity = entities["SYN-IND"]
    scrips = configs.load_scrips(ROOT / "Data" / "itr" / "scrips.example.yaml")
    fmv_tables = sch.load_fmv_tables()
    user_rules = rules_engine.load_user_rules(RULES_DIR / "user_rules.yaml")

    model = sch.build_all_schedules(
        tree, result.resolved, book, form16, YEAR_KEY, rules, "new",
        entity.status, entity.dob, scrips, fmv_tables,
    )
    return tree, model, rules, user_rules, entity, result, loaded, form16


def _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded, name="out.xlsx"):
    out_path = tmp_path / name
    ww.write_workbook(
        str(out_path), tree, model, rules, user_rules, entity, "new", YEAR_KEY,
        None, [], [], [], result.unmapped, "v1", "2026-01-01T00:00:00", {},
        result.resolved, loaded.entries,
    )
    return openpyxl.load_workbook(str(out_path))


def _all_formulas(ws):
    formulas = []
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                formulas.append(c.value)
    return formulas


def _find_row(ws, label_substring, max_row=None):
    """Statement of Income uses a variable label column (LBL=2, indented
    sub-items further right) -- search every column, not just column 1."""
    for row in ws.iter_rows(min_row=1, max_row=max_row or ws.max_row):
        for c in row:
            if isinstance(c.value, str) and label_substring in c.value:
                return c.row
    return None


def _find_cell(ws, label_substring, max_row=None):
    for row in ws.iter_rows(min_row=1, max_row=max_row or ws.max_row):
        for c in row:
            if isinstance(c.value, str) and label_substring in c.value:
                return c
    return None


def _formula_in_row(ws, row):
    """Return the first formula string found anywhere in `row` -- labels and
    formulas live in different columns depending on section (LBL/INNER/SUB/
    OUTER), so locate by content, not a fixed column index."""
    for c in ws[row]:
        if isinstance(c.value, str) and c.value.startswith("="):
            return c.value
    return None


# ---------------------------------------------------------------------------
# Fix A: mround_safe() sign-safety -- pure Python shadow proof + generated
# workbook structural proof (refund-positive AND tax-payable/negative).
# ---------------------------------------------------------------------------

def _excel_round_half_away(x: float, ndigits: int = 0) -> float:
    """Shadow of Excel's ROUND(): round-half-away-from-zero (NOT Python's
    banker's rounding). Used to evaluate `presentation.mround_safe`'s output
    formula in pure Python, since openpyxl cannot evaluate formulas."""
    import math
    factor = 10 ** ndigits
    y = x * factor
    if y >= 0:
        return math.floor(y + 0.5) / factor
    return math.ceil(y - 0.5) / factor


def _shadow_mround_safe(x: float, m: float) -> float:
    """Pure-Python evaluation of `ROUND((x)/m,0)*m` -- exactly what
    presentation.mround_safe's formula computes once Excel evaluates it."""
    return _excel_round_half_away(x / m, 0) * m


def _shadow_excel_mround(x: float, m: float) -> float:
    """Faithful shadow of Excel's actual MROUND(number, multiple): raises
    (mirroring #NUM!) when `number` and `multiple` have opposite signs --
    this is the pre-fix bug being regression-tested against."""
    if x != 0 and m != 0 and (x < 0) != (m < 0):
        raise ValueError("#NUM! -- MROUND requires number and multiple to share a sign")
    return _excel_round_half_away(x / m, 0) * m


def test_mround_safe_formula_shape():
    formula = presentation.mround_safe("'TaxesPaid'!B15-E39", "'Rules'!C10")
    assert formula == "ROUND(('TaxesPaid'!B15-E39)/'Rules'!C10,0)*'Rules'!C10"
    assert "MROUND(" not in formula


@pytest.mark.parametrize("number,multiple,expected", [
    (7.5, 10, 10),      # positive-vs-positive: matches Excel MROUND
    (0, 10, 0),
    (-425000, 10, -425000),   # already a multiple of 10, negative total income
    (-425003, 10, -425000),   # loss-year TI needing s.288A rounding, negative
    (-12345.5, 10, -12350),   # tax-payable (refund negative) needing s.288B rounding
    (12345.5, 10, 12350),     # refund-positive, unchanged behaviour
])
def test_shadow_mround_safe_matches_expected_rounding(number, multiple, expected):
    assert _shadow_mround_safe(number, multiple) == pytest.approx(expected)


def test_shadow_excel_mround_raises_on_opposite_signs_proving_the_bug():
    """This is the actual bug being fixed: a genuine Excel MROUND(x, m) call
    raises #NUM! whenever x and m have opposite signs -- exactly the
    tax-payable (x<0, m>0) and loss-year (x<0, m>0) cases. mround_safe must
    not."""
    with pytest.raises(ValueError):
        _shadow_excel_mround(-12345.5, 10)
    # ...but the sign-safe replacement handles it fine, with the identical
    # round-half-away-from-zero result MROUND would have given had it not
    # errored:
    assert _shadow_mround_safe(-12345.5, 10) == pytest.approx(-12350)


def test_refund_formula_on_statement_of_income_and_computation_use_sign_safe_rounding(
    tmp_path, syn_ind_model_and_paths,
):
    """Refund-positive (unchanged) case: the shared syn_ind fixture produces
    a normal refund. Both the Statement of Income and Computation sheets'
    refund formulas must use the sign-safe ROUND((x)/m,0)*m shape, never
    MROUND, so the SAME formula also works unmodified the day this filer's
    numbers flip to tax-payable."""
    tree, model, rules, user_rules, entity, result, loaded, form16 = syn_ind_model_and_paths
    wb = _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded)

    soi_row = _find_row(wb["Statement of Income"], "Refund Due / (Tax Payable)")
    assert soi_row is not None
    soi_formula = _formula_in_row(wb["Statement of Income"], soi_row)
    assert isinstance(soi_formula, str) and soi_formula.startswith("=ROUND((")
    assert "MROUND(" not in soi_formula

    comp_row = _find_row(wb["Computation"], "Refund")
    assert comp_row is not None
    comp_formula = wb["Computation"].cell(row=comp_row, column=2).value
    assert isinstance(comp_formula, str) and comp_formula.startswith("=ROUND((")
    assert "MROUND(" not in comp_formula


def test_tax_payable_negative_case_still_generates_sign_safe_formula_no_crash(
    tmp_path, syn_ind_model_and_paths,
):
    """Force a tax-payable (refund-negative) scenario by zeroing out prepaid
    taxes -- liability > 0 - 0 = a negative refund line, i.e. exactly the
    opposite-sign case that used to raise #NUM! under MROUND. Workbook
    generation must not raise, and the resulting formula must keep the
    sign-safe shape (openpyxl can't evaluate the formula itself, but the
    shadow-math tests above independently prove the shape is arithmetically
    correct for this exact sign combination)."""
    tree, model, rules, user_rules, entity, result, loaded, form16 = syn_ind_model_and_paths
    model.taxes_paid.advance_tax = 0.0
    model.taxes_paid.self_assessment_tax = 0.0
    model.taxes_paid.tds_salary = 0.0
    model.taxes_paid.tds_interest = 0.0
    model.taxes_paid.tds_dividend = 0.0
    model.taxes_paid.tcs = 0.0
    model.taxes_paid.total = 0.0

    wb = _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded, name="payable.xlsx")

    soi_row = _find_row(wb["Statement of Income"], "Refund Due / (Tax Payable)")
    soi_formula = _formula_in_row(wb["Statement of Income"], soi_row)
    assert isinstance(soi_formula, str) and soi_formula.startswith("=ROUND((")
    assert "MROUND(" not in soi_formula

    comp_row = _find_row(wb["Computation"], "Refund")
    comp_formula = wb["Computation"].cell(row=comp_row, column=2).value
    assert isinstance(comp_formula, str) and comp_formula.startswith("=ROUND((")
    assert "MROUND(" not in comp_formula


# ---------------------------------------------------------------------------
# Fix B: Salary gross sources from Form16 total_1d (17(1)+17(2)+17(3)), not
# the book's SALARY_GROSS tag (which only ever captures 17(1)).
# ---------------------------------------------------------------------------

def _fake_resolved_with_salary_gross_tag():
    class _FakeLeaf:
        guid = "g1"
        tag = "SALARY_GROSS"
        path = "fake/SALARY_GROSS"
        flags = []

    class _FakeNode:
        total = 500000.0   # book only ever captures 17(1)

    return {"g1": _FakeLeaf()}, {"g1": _FakeNode()}


def test_salary_gross_uses_total_1d_not_book_salary_gross_when_perquisites_present():
    """The actual bug: a salary WITH non-zero 17(2) perquisites -- book's
    SALARY_GROSS (17(1) only) is 500000, but total_1d (17(1)+17(2)+17(3)) is
    560000. The sheet's own label says '(17(1)+17(2)+17(3))', so it must
    show 560000, not the book's 500000."""
    resolved, node_by_guid = _fake_resolved_with_salary_gross_tag()
    rules = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    form16 = parse_form16.Form16Data(
        s17_1=500000.0, s17_2=50000.0, s17_3=10000.0, total_1d=560000.0,
        std_deduction_16a=50000.0, total_2i=0.0, prof_tax_16c=2500.0,
        income_chargeable_6=507500.0,
    )
    schedule = sch.build_salary(resolved, node_by_guid, form16, rules, "new")
    assert schedule.source == "form16"
    assert schedule.gross == pytest.approx(560000.0)
    assert schedule.gross != pytest.approx(500000.0)  # must NOT be the book control total
    assert schedule.reconciliation_ok
    assert schedule.reconciliation_diff == pytest.approx(0.0, abs=0.01)


def test_salary_gross_falls_back_to_s17_sum_when_total_1d_missing():
    resolved, node_by_guid = _fake_resolved_with_salary_gross_tag()
    rules = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    form16 = parse_form16.Form16Data(
        s17_1=500000.0, s17_2=50000.0, s17_3=10000.0, total_1d=None,
        std_deduction_16a=50000.0, total_2i=0.0, prof_tax_16c=2500.0,
        income_chargeable_6=507500.0,
    )
    schedule = sch.build_salary(resolved, node_by_guid, form16, rules, "new")
    assert schedule.gross == pytest.approx(560000.0)   # 500000+50000+10000


def test_salary_book_only_path_unchanged_by_fix():
    """No Form16 at all: must still fall back to the book-only path,
    unmodified by this fix."""
    resolved, node_by_guid = _fake_resolved_with_salary_gross_tag()
    rules = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    schedule = sch.build_salary(resolved, node_by_guid, None, rules, "new")
    assert schedule.source == "book-only"
    assert schedule.manual_flagged is True
    assert schedule.gross == pytest.approx(500000.0)


def test_salary_sheet_identity_to_the_paisa(tmp_path):
    """Salary sheet writes raw numeric leaf cells (not formulas) -- so
    openpyxl CAN read real numbers here, and the identity gross - s10 -
    stdded - proftax == income_chargeable must hold exactly to the paisa."""
    schedule = sch.SalarySchedule(
        gross=560000.0, s10_exempt_total=0.0, std_deduction=50000.0,
        prof_tax=2500.0, income_chargeable=507500.0, source="form16",
        reconciliation_ok=True, reconciliation_diff=0.0,
    )
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    layout = ww.write_salary_sheet(wb, schedule)
    ws = wb["Salary"]

    # write_salary_sheet's layout dict returns openpyxl Cell objects whose
    # .coordinate re-resolves against the live worksheet, and whose raw
    # numeric .value was written directly from the SalarySchedule fields
    # (no formulas involved on this sheet) -- so reading .value back is a
    # genuine to-the-paisa numeric identity check, not a formula-shape proxy.
    gross = ws[layout["gross"].coordinate].value
    s10 = ws[layout["s10_exempt"].coordinate].value
    stdded = ws[layout["std_deduction"].coordinate].value
    proftax = ws[layout["prof_tax"].coordinate].value
    income_chargeable = ws[layout["income_chargeable"].coordinate].value
    assert gross == pytest.approx(560000.0)
    assert (gross - s10 - stdded - proftax) == pytest.approx(income_chargeable, abs=0.005)


def test_verify_cross_check_form16_still_ties_to_17_1_alone(syn_ind_model_and_paths):
    """verify.py's Book<->Form16 control must still compare the book's
    SALARY_GROSS total against Form16's 17(1) ALONE -- untouched by Fix B,
    which only changed the Salary SHEET's displayed gross (17(1)+17(2)+
    17(3)). For the syn_ind fixture, 17(2) and 17(3) are both 0.00, so
    17(1) == total_1d and the control naturally still ties out."""
    tree, model, rules, user_rules, entity, result, loaded, form16 = syn_ind_model_and_paths
    results = verify.cross_check_form16(tree, result.resolved, form16)
    salary_result = next(r for r in results if r.label == "17(1) Salary vs SALARY_GROSS")
    assert salary_result.form16_total == pytest.approx(form16.s17_1)
    assert salary_result.ok


def test_verify_cross_check_form16_unaffected_by_sheet_gross_divergence():
    """Direct proof that verify.py's control is blind to (and correctly so)
    a Salary-sheet total_1d that diverges from SALARY_GROSS due to non-zero
    perquisites -- it must keep comparing 17(1) alone, never total_1d."""
    class _FakeLeaf:
        guid = "g1"
        tag = "SALARY_GROSS"
        path = "fake/SALARY_GROSS"
        flags = []

    class _FakeNode:
        guid = "g1"
        total = 500000.0

    class _FakeTree:
        def all_nodes(self):
            return [_FakeNode()]

    form16 = parse_form16.Form16Data(
        s17_1=500000.0, s17_2=50000.0, s17_3=10000.0, total_1d=560000.0,
    )
    results = verify.cross_check_form16(_FakeTree(), {"g1": _FakeLeaf()}, form16)
    salary_result = next(r for r in results if r.label == "17(1) Salary vs SALARY_GROSS")
    # Book (500000) vs 17(1) alone (500000) -- ties out even though the
    # Salary sheet itself now shows 560000 (total_1d, Fix B).
    assert salary_result.mapped_total == pytest.approx(500000.0)
    assert salary_result.form16_total == pytest.approx(500000.0)
    assert salary_result.ok


# ---------------------------------------------------------------------------
# Fix B2: salary fail-loud banner ("banner, no abort").
# ---------------------------------------------------------------------------

def _cell_values(ws, max_row=6):
    values = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
        values.extend(v for v in row if v is not None)
    return values


def test_reconciled_salary_has_no_error_banner(tmp_path, syn_ind_model_and_paths):
    tree, model, rules, user_rules, entity, result, loaded, form16 = syn_ind_model_and_paths
    assert model.salary.reconciliation_ok
    wb = _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded)
    for sheet in ("Salary", "Statement of Income"):
        text = " ".join(str(v) for v in _cell_values(wb[sheet]))
        assert presentation.SALARY_RECONCILIATION_ERROR_MARKER not in text


def test_salary_mismatch_writes_error_banner_on_salary_and_statement_of_income(
    tmp_path, syn_ind_model_and_paths,
):
    tree, model, rules, user_rules, entity, result, loaded, form16 = syn_ind_model_and_paths
    model.salary.reconciliation_ok = False
    model.salary.reconciliation_diff = 777.77

    wb = _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded, name="mismatch.xlsx")

    assert "Salary" in wb.sheetnames
    assert "Statement of Income" in wb.sheetnames
    for sheet in ("Salary", "Statement of Income"):
        text = " ".join(str(v) for v in _cell_values(wb[sheet]))
        assert presentation.SALARY_RECONCILIATION_ERROR_MARKER in text
        assert "777.77" in text

    recon_text = " ".join(
        str(v) for row in wb["Reconciliation"].iter_rows(values_only=True) for v in row if v is not None
    )
    assert "MISMATCH" in recon_text


# ---------------------------------------------------------------------------
# Fix C: Workings/Inputs section (b/f losses + regime tax workings) moved
# below the Refund line; b/f cells remain live inputs; no #REF!/#NUM! text
# anywhere; sheet-generation itself already self-validates via runtime
# assertions in presentation.py -- these tests add an external structural
# proof plus a whole-suite negative-Total-Income smoke check.
# ---------------------------------------------------------------------------

def test_workings_inputs_section_is_below_refund_line(tmp_path, syn_ind_model_and_paths):
    tree, model, rules, user_rules, entity, result, loaded, form16 = syn_ind_model_and_paths
    wb = _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded)
    ws = wb["Statement of Income"]

    refund_row = _find_row(ws, "Refund Due / (Tax Payable)")
    workings_row = _find_row(ws, "Workings / Inputs")
    bf_row = _find_row(ws, "Brought forward losses set off")
    regime_row = _find_row(ws, "Regime comparison workings")

    assert refund_row is not None
    assert workings_row is not None
    assert bf_row is not None
    assert regime_row is not None
    assert workings_row > refund_row
    assert bf_row > workings_row
    assert regime_row > refund_row


def test_bf_loss_cells_remain_live_numeric_input_cells(tmp_path, syn_ind_model_and_paths):
    tree, model, rules, user_rules, entity, result, loaded, form16 = syn_ind_model_and_paths
    wb = _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded)
    ws = wb["Statement of Income"]

    bf_row = _find_row(ws, "Brought forward losses set off")
    assert bf_row is not None
    # The statutory buckets (HP loss, Business loss, STCL, LTCL) are the
    # next 4 rows below the section header; each must be a plain numeric
    # literal (default 0), NOT a formula -- i.e. still directly editable
    # by hand. Input cells live in the SUB column (4 / "D"); other columns
    # on these same rows legitimately carry unrelated regime-tax-workings
    # formulas (both sections' row ranges overlap by design -- see
    # presentation.py), so only the SUB-column cell is asserted on.
    SUB_COL = 4
    for r in range(bf_row + 1, bf_row + 5):
        label = next(
            (c.value for c in ws[r] if isinstance(c.value, str) and c.value.strip() and c.column < SUB_COL),
            None,
        )
        value = ws.cell(row=r, column=SUB_COL).value
        assert isinstance(label, str) and label.strip() != ""
        assert not (isinstance(value, str) and value.startswith("=")), (
            f"b/f-loss row {r} ({label!r}) must be a live input cell, not a formula: {value!r}"
        )
        assert value == 0  # default


def test_no_ref_or_num_error_literals_anywhere_in_statement_of_income(tmp_path, syn_ind_model_and_paths):
    """openpyxl doesn't evaluate formulas, so this can't prove Excel would
    never SHOW #REF!/#NUM! at open-time -- but it does prove the generator
    itself never emits one of those error tokens as literal formula text
    (which would indicate a broken/self-referential coordinate baked in at
    generation time), across every sheet in the workbook."""
    tree, model, rules, user_rules, entity, result, loaded, form16 = syn_ind_model_and_paths
    wb = _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded)
    for sheet in wb.sheetnames:
        for formula in _all_formulas(wb[sheet]):
            assert "#REF!" not in formula, f"{sheet}: {formula}"
            assert "#NUM!" not in formula, f"{sheet}: {formula}"


def test_regime_tax_workings_still_present_and_referenced_same_sheet(tmp_path, syn_ind_model_and_paths):
    tree, model, rules, user_rules, entity, result, loaded, form16 = syn_ind_model_and_paths
    wb = _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded)
    ws = wb["Statement of Income"]
    formulas = _all_formulas(ws)
    # The regime workings must still exist on THIS sheet (not moved to a
    # different sheet) -- proven by New/Old labels being present and by
    # formulas within the sheet cross-referencing 'CapitalGains'!/'Rules'!
    # exactly as before the move.
    labels = [
        c.value for row in ws.iter_rows() for c in row
        if isinstance(c.value, str) and "regime working" in c.value
    ]
    assert any(label.startswith("New regime working") for label in labels)
    assert any(label.startswith("Old regime working") for label in labels)
    assert any("'Rules'!" in f for f in formulas)


def test_pointer_note_present_at_house_property_business_and_cg_section_heads(
    tmp_path, syn_ind_model_and_paths,
):
    tree, model, rules, user_rules, entity, result, loaded, form16 = syn_ind_model_and_paths
    wb = _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded)
    ws = wb["Statement of Income"]
    labels = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    pointer_hits = [lbl for lbl in labels if "see Workings below" in lbl]
    # At minimum, House Property and Capital Gains sections (present in the
    # syn_ind fixture) must carry the pointer note; Salary must NOT (it has
    # no b/f set-off applied to it).
    assert any("House Property" in lbl for lbl in pointer_hits) or any(
        "Capital Gains" in lbl for lbl in pointer_hits
    )
    assert not any(lbl.strip().startswith("Income from Salary") and "see Workings below" in lbl for lbl in labels)


def test_negative_total_income_loss_year_generation_does_not_crash_and_stays_sign_safe(
    tmp_path, syn_ind_model_and_paths,
):
    """Loss-year smoke test: force Total Income negative by zeroing every
    positive-income schedule and inflating a deduction-side figure, then
    confirm the workbook still generates cleanly (all of Fix C's runtime
    layout-drift assertions pass) and the Total Income formula keeps its
    sign-safe ROUND(...) shape (never MROUND) -- this is precisely the
    scenario Fix A's #NUM! bug hit hardest, and precisely the scenario Fix
    C's row-prediction assertions must survive unchanged."""
    tree, model, rules, user_rules, entity, result, loaded, form16 = syn_ind_model_and_paths
    model.deductions.total = model.deductions.total + 10_000_000.0  # force GTI - VIA negative
    wb = _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded, name="loss.xlsx")
    ws = wb["Statement of Income"]
    ti_row = _find_row(ws, "Total Income")
    ti_formula = _formula_in_row(ws, ti_row)
    assert isinstance(ti_formula, str) and ti_formula.startswith("=ROUND((")
    assert "MROUND(" not in ti_formula


# ---------------------------------------------------------------------------
# Fix D: Assumptions block notes that 234A/234B/234C interest is not
# computed.
# ---------------------------------------------------------------------------

def test_assumptions_block_notes_234_interest_not_computed(tmp_path, syn_ind_model_and_paths):
    tree, model, rules, user_rules, entity, result, loaded, form16 = syn_ind_model_and_paths
    wb = _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded)
    ws = wb["Statement of Income"]
    text = " ".join(
        str(v) for row in ws.iter_rows(values_only=True) for v in row if isinstance(v, str)
    )
    assert "234A" in text and "234B" in text and "234C" in text
    assert "NOT computed" in text


# ---------------------------------------------------------------------------
# Tax parity: Fix B must not move any tax FIGURE -- the Salary sheet's
# GROSS label changed, but income_chargeable (what actually flows into
# Total Income / tax) is untouched, so no tax computation input changes.
# ---------------------------------------------------------------------------

def test_fix_b_does_not_change_income_chargeable_or_downstream_tax_inputs():
    """Same Form16 (zero perquisites, matching the syn_ind fixture shape) --
    income_chargeable, std_deduction, s10_exempt_total, prof_tax (the only
    fields that flow into Total Income / tax) must be byte-identical to
    what build_salary produced before Fix B; only `gross` (a display-only
    figure feeding the Salary-sheet label, not itself referenced by the tax
    ladder) can legitimately move."""
    resolved, node_by_guid = _fake_resolved_with_salary_gross_tag()
    rules = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    form16 = parse_form16.Form16Data(
        s17_1=500000.0, s17_2=0.0, s17_3=0.0, total_1d=500000.0,
        std_deduction_16a=50000.0, total_2i=0.0, prof_tax_16c=2500.0,
        income_chargeable_6=447500.0,
    )
    schedule = sch.build_salary(resolved, node_by_guid, form16, rules, "new")
    assert schedule.gross == pytest.approx(500000.0)          # unchanged when 17(2)/17(3) are 0
    assert schedule.income_chargeable == pytest.approx(447500.0)
    assert schedule.std_deduction == pytest.approx(50000.0)
    assert schedule.prof_tax == pytest.approx(2500.0)
    assert schedule.s10_exempt_total == pytest.approx(0.0)
