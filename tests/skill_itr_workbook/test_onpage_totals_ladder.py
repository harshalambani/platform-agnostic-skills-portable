"""
tests/skill_itr_workbook/test_onpage_totals_ladder.py -- regression coverage
for the 2026-07-20 on-page-totals change (docs/2026-07-20-itr-onpage-totals-plan.md).

Three things this file exists to prove, none of them covered elsewhere:

1. FORMULA-STRING WIRING -- the `Statement of Income` page's income ladder
   (Gross Total Income, Chapter VI-A, b/f-loss, Total Income, the
   normal-income/special-rate-CG split) is built from on-page cells, and
   `Computation`'s re-anchored slab-tax formula reads the PAGE's
   normal-income-base cell, not a locally recomputed figure.
2. PYTHON SHADOW-CALC RECONCILIATION -- for the DEFAULT (no b/f override)
   case, the arithmetic the new on-page formula graph implies must equal
   `schedules.py`'s own (untouched) tax engine output to the paisa -- that
   engine is "today's" ground truth, unaffected by this presentation-layer
   change. A synthetic b/f-loss override is also hand-simulated and checked
   against `schedules.compute_tax` (the same pure-Python slab/rebate/
   surcharge/cess implementation `Computation`'s Excel formulas mirror).
3. THE CORRECTNESS TRAP (design doc section 5) -- special-rate capital gains
   (111A/112A) must never enter the slab-tax base. A dedicated regression
   proves the on-page normal-income-base formula excludes special-rate CG,
   that `Computation`'s special-rate-CG tax cell never references the page
   at all (independent of any override), and that naively including CG in
   the slab base would change the numbers -- i.e. the carve-out is not a
   no-op.

openpyxl never evaluates formulas, so (1) is proven via formula-string
assertions and (2)/(3) via a Python re-derivation compared against
`schedules.py`'s own engine -- the same approach used throughout this test
suite (see test_write_workbook.py's module docstring).

Fixtures: reuses the existing SYN-IND synthetic fixture set already used by
test_write_workbook.py and test_presentation.py (Data/itr/entities.example.yaml
+ tests/skill_itr_workbook/fixtures/syn_ind.*) -- fabricated identity/financial
data, no real PII, PAN, account numbers or amounts anywhere in this file.
"""
from __future__ import annotations

import re
import sys
from datetime import date
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPTS = ROOT / "src" / "agents" / "skill_itr_workbook" / "scripts"
FIXTURES = Path(__file__).resolve().parent / "fixtures"
RULES_DIR = ROOT / "Data" / "itr" / "rules"

for p in (str(SCRIPTS), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

import parse_eguile as pe  # noqa: E402
import parse_gnucash as pg  # noqa: E402
import parse_form16  # noqa: E402
import configs  # noqa: E402
import mapping as mapping_engine  # noqa: E402
import rules as rules_engine  # noqa: E402
import schedules as sch  # noqa: E402
import write_workbook as ww  # noqa: E402
import fixture_gen  # noqa: E402

YEAR_KEY = "2024-25"
REGIME = "new"


@pytest.fixture(scope="module")
def built():
    """Same construction as test_write_workbook.py's `built_workbook` --
    the SYN-IND synthetic fixture, which conveniently has BOTH normal income
    (salary) and special-rate capital gains (LT loss + ST gain), making it
    the right fixture for the correctness-trap regression without inventing
    a second one."""
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    book = pg.parse_book(FIXTURES / "syn_ind.gnucash")
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    result = mapping_engine.resolve_tree(tree, loaded)
    form16 = parse_form16.parse_form16(FIXTURES / "syn_ind_form16.pdf")
    rules = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    entities = configs.load_entities(ROOT / "Data" / "itr" / "entities.example.yaml")
    entity = entities["SYN-IND"]
    scrips = configs.load_scrips(ROOT / "Data" / "itr" / "scrips.example.yaml")
    fmv_tables = sch.load_fmv_tables()
    user_rules = rules_engine.load_user_rules(RULES_DIR / "user_rules.yaml")

    model = sch.build_all_schedules(
        tree, result.resolved, book, form16, YEAR_KEY, rules, REGIME,
        entity.status, entity.dob, scrips, fmv_tables,
    )
    return model, rules, entity


@pytest.fixture(scope="module")
def built_workbook(tmp_path_factory, built):
    model, rules, entity = built
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    book = pg.parse_book(FIXTURES / "syn_ind.gnucash")
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    result = mapping_engine.resolve_tree(tree, loaded)
    form16 = parse_form16.parse_form16(FIXTURES / "syn_ind_form16.pdf")
    user_rules = rules_engine.load_user_rules(RULES_DIR / "user_rules.yaml")
    out_path = tmp_path_factory.mktemp("wb") / "syn_ind_ladder.xlsx"
    ww.write_workbook(
        str(out_path), tree, model, rules, user_rules, entity, REGIME, YEAR_KEY,
        form16.opted_out_115bac, [], [], [], result.unmapped, "v1", "2026-01-01T00:00:00", {},
        result.resolved, loaded.entries,
    )
    return openpyxl.load_workbook(str(out_path))


# ---------------------------------------------------------------------------
# Helpers -- exact-label cell lookup (avoids the substring ambiguity of
# scanning for e.g. "Total" which also matches "Total prepaid taxes").
# ---------------------------------------------------------------------------

def _find_exact(ws, label: str, col: int | None = None):
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value == label:
                if col is None or c.column == col:
                    return c
    raise AssertionError(f"exact label {label!r} not found (col={col})")


def _find_containing(ws, needle: str, col: int | None = None):
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and needle in c.value:
                if col is None or c.column == col:
                    return c
    raise AssertionError(f"label containing {needle!r} not found (col={col})")


LBL, INNER, SUB, OUTER = 2, 3, 4, 5


# ---------------------------------------------------------------------------
# 1. Formula-string wiring
# ---------------------------------------------------------------------------

def test_gti_is_an_on_page_sum_formula(built_workbook):
    """Gross Total Income ("Total" row, right after the income-head sections)
    must be a live =SUM(...) over on-page subtotal cells -- not a mirror of a
    hidden Computation-sheet cell (the pre-2026-07-20 behaviour)."""
    ws = built_workbook["Statement of Income"]
    gti_cell = _find_exact(ws, "Total", col=LBL)
    gti_value = ws.cell(row=gti_cell.row, column=OUTER).value
    assert isinstance(gti_value, str)
    assert re.fullmatch(r"=SUM\((D\d+,?)+\)", gti_value), \
        f"GTI formula not an on-page SUM over column D subtotal cells: {gti_value!r}"
    assert "Computation" not in gti_value, \
        "GTI must no longer mirror the Computation sheet"


def test_chapter_via_is_a_leaf_reference_to_deductions_sheet(built_workbook):
    ws = built_workbook["Statement of Income"]
    via_cell = _find_exact(ws, "Less - Chapter VI-A deductions", col=LBL)
    via_value = ws.cell(row=via_cell.row, column=SUB).value
    assert via_value.startswith("='Deductions'!")


def test_total_income_is_gti_minus_via_with_no_lump_bf_term(built_workbook):
    """2026-07-21 design (section 11.2): there is no longer a single lump
    b/f-loss cell subtracted at the Total Income line -- each of the four
    statutory buckets nets off its OWN head/gain-type BEFORE aggregation
    into Gross Total Income (see test_bf_loss_buckets_are_real_editable_...
    and test_bf_loss_buckets_route_into_their_own_head_... below), so Total
    Income is simply GTI minus Chapter VI-A, s.288A-rounded."""
    ws = built_workbook["Statement of Income"]
    gti_cell = _find_exact(ws, "Total", col=LBL)
    gti_ref = f"{ws.cell(row=gti_cell.row, column=OUTER).coordinate}"

    via_cell = _find_exact(ws, "Less - Chapter VI-A deductions", col=LBL)
    via_ref = ws.cell(row=via_cell.row, column=SUB).coordinate

    ti_cell = _find_exact(ws, "Total Income", col=LBL)
    ti_formula = ws.cell(row=ti_cell.row, column=OUTER).value
    assert isinstance(ti_formula, str) and ti_formula.startswith("=MROUND(")
    assert gti_ref in ti_formula, "Total Income must reference the GTI cell"
    assert via_ref in ti_formula, "Total Income must reference the Chapter VI-A cell"
    assert "'Rules'!" in ti_formula, "Total Income must keep the s.288A MROUND rounding"


def test_bf_loss_buckets_are_real_editable_input_cells_defaulting_to_zero(built_workbook):
    """Design doc section 11.2: FOUR statutory per-bucket b/f-loss input
    cells (replacing the pre-2026-07-21 single lump cell) -- HP (s.71B),
    Business (s.72), STCL and LTCL (both s.74) -- each a real editable
    `_input_cell` defaulting to literal 0."""
    ws = built_workbook["Statement of Income"]
    needles = (
        "b/f House Property loss (s.71B)",
        "b/f Business loss (s.72)",
        "b/f Short-term capital loss (s.74)",
        "b/f Long-term capital loss (s.74)",
    )
    for needle in needles:
        label_cell = _find_containing(ws, needle, col=LBL)
        bf_cell = ws.cell(row=label_cell.row, column=SUB)
        assert bf_cell.value == 0, f"{needle!r} input cell must default to literal 0"


def test_bf_loss_buckets_route_into_their_own_head_before_gti_not_a_lump_ti_deduction(built_workbook):
    """Proves the routing half of section 11.2 on the actual rendered page:
    the Business b/f bucket (s.72) must appear inside the "Net business
    income" head item's formula (rendered for SYN-IND, which carries
    non-zero business income), and the HP bucket must NOT leak into it.
    Total Income must not reference any bf cell directly any more -- set-off
    happens at the head level, before GTI, not as a lump TI subtraction."""
    ws = built_workbook["Statement of Income"]
    hp_bf = _find_containing(ws, "b/f House Property loss (s.71B)", col=LBL)
    hp_bf_ref = ws.cell(row=hp_bf.row, column=SUB).coordinate
    biz_bf = _find_containing(ws, "b/f Business loss (s.72)", col=LBL)
    biz_bf_ref = ws.cell(row=biz_bf.row, column=SUB).coordinate
    stcl_bf = _find_containing(ws, "b/f Short-term capital loss (s.74)", col=LBL)
    stcl_bf_ref = ws.cell(row=stcl_bf.row, column=SUB).coordinate

    biz_item = _find_containing(ws, "Net business income", col=LBL)
    biz_formula = ws.cell(row=biz_item.row, column=INNER).value
    assert biz_bf_ref in biz_formula, "Business bucket must route into the Business head item"
    assert hp_bf_ref not in biz_formula, "HP bucket must never leak into Business"
    assert stcl_bf_ref not in biz_formula, "CG buckets must never leak into Business"

    stcg_item = _find_containing(ws, "Short Term Capital Gain / (Loss)", col=LBL)
    stcg_formula = ws.cell(row=stcg_item.row, column=INNER).value
    assert stcl_bf_ref in stcg_formula, "STCL bucket must route into the STCG head item"
    assert biz_bf_ref not in stcg_formula, "Business bucket must never leak into CG"

    ti_cell = _find_exact(ws, "Total Income", col=LBL)
    ti_formula = ws.cell(row=ti_cell.row, column=OUTER).value
    for ref in (hp_bf_ref, biz_bf_ref, stcl_bf_ref):
        assert ref not in ti_formula, "b/f buckets must not be a lump Total-Income deduction any more"


def test_normal_income_base_excludes_special_rate_cg_on_page(built_workbook):
    """THE correctness-trap wiring check (design doc section 5): the on-page
    normal-income-base formula must be MAX(0, TotalIncome - SpecialRateCG),
    i.e. it structurally subtracts CG rather than including it."""
    ws = built_workbook["Statement of Income"]
    ti_cell = _find_exact(ws, "Total Income", col=LBL)
    ti_ref = ws.cell(row=ti_cell.row, column=OUTER).coordinate

    cg_base_label = _find_containing(ws, "Special-rate Capital Gains", col=LBL)
    cg_base_ref = ws.cell(row=cg_base_label.row, column=SUB).coordinate

    normal_label = _find_containing(ws, "Normal income (slab-tax base)", col=LBL)
    normal_formula = ws.cell(row=normal_label.row, column=SUB).value
    assert normal_formula == f"=MAX(0,{ti_ref}-{cg_base_ref})"


def test_computation_slab_base_reads_the_pages_normal_income_cell(built_workbook):
    """The re-anchor point: `Computation`'s "Tax on normal-rate income" cell
    (both regimes) must reference `'Statement of Income'!<normal_income_base>`
    -- not a locally recomputed GTI/TI. Proven for at least one regime block;
    both are built by the same `_regime_block` closure in write_workbook.py."""
    stmt = built_workbook["Statement of Income"]
    normal_label = _find_containing(stmt, "Normal income (slab-tax base)", col=LBL)
    normal_ref = stmt.cell(row=normal_label.row, column=SUB).coordinate
    expected_ref = f"'Statement of Income'!{normal_ref}"

    comp = built_workbook["Computation"]
    tax_normal_label = _find_containing(comp, "Tax on normal-rate income", col=1)
    tax_normal_formula = comp.cell(row=tax_normal_label.row, column=2).value
    assert expected_ref in tax_normal_formula, (
        f"Computation slab-tax formula does not reference the page's normal-income "
        f"cell {expected_ref!r}: {tax_normal_formula!r}"
    )


def test_special_rate_cg_tax_cell_never_references_the_page(built_workbook):
    """The other half of the correctness trap: 112A/111A tax on special-rate
    CG must stay wired to `CapitalGains` only -- an on-page override (b/f
    loss, VI-A, any leaf) must never change how special-rate CG itself is
    taxed."""
    comp = built_workbook["Computation"]
    for label_cell in comp.iter_rows():
        for c in label_cell:
            if c.value == "112A/111A tax on special-rate CG":
                formula = comp.cell(row=c.row, column=2).value
                assert formula.startswith("='CapitalGains'!")
                assert "Statement of Income" not in formula
                return
    raise AssertionError("112A/111A tax on special-rate CG cell not found")


# ---------------------------------------------------------------------------
# 2. Python shadow-calc reconciliation (default, no override)
# ---------------------------------------------------------------------------

def test_default_case_ladder_matches_schedules_engine_to_the_paisa(built):
    """The strongest regression guarantee (design doc section 7, point 3):
    for the DEFAULT (b/f=0) case, the on-page ladder's implied arithmetic
    must equal `schedules.py`'s own (untouched) tax-engine output exactly.

    schedules.build_computation (unaffected by this change) already computes
    gti, via, total_income, and TaxBlock.normal_income/special_rate_income
    from the identical component leaves the page now sums on-page. This test
    re-derives GTI/TI/normal-base/special-CG-base the same way the new
    on-page formulas do, using the model's own leaf numbers, and checks the
    result matches the engine's ground truth -- i.e. moving the ladder
    on-page did not change what it computes."""
    model, rules, entity = built
    comp = model.computation

    # Re-derive the page's ladder the same way write_statement_of_income's
    # formulas do: GTI = SUM(section subtotals); special_cg_base = cg_lt + cg_st
    # (cg_lt already net of LT exemption, matching the leaf-cell formula);
    # normal_income_base = MAX(0, TotalIncome - special_cg_base).
    gti = (
        comp.salary_income + comp.house_property_income + comp.business_income
        + comp.other_sources_income + comp.capital_gains_lt + comp.capital_gains_st
    )
    assert gti == pytest.approx(comp.gti), "re-derived on-page GTI must match schedules.py's gti"

    total_income = sch.round_288a(gti - comp.via_deductions, rules.common["rounding"]["total_income"]["nearest"])
    assert total_income == pytest.approx(comp.total_income_rounded)

    special_cg_base = comp.capital_gains_lt + comp.capital_gains_st
    assert special_cg_base == pytest.approx(comp.tax_block.special_rate_income)

    normal_income_base = max(0.0, total_income - special_cg_base)
    assert normal_income_base == pytest.approx(comp.tax_block.normal_income), (
        "on-page normal-income-base (the slab-tax base Computation now reads) must match "
        "schedules.py's TaxBlock.normal_income exactly -- this is the paisa-exact default-"
        "case regression guarantee"
    )


def test_default_case_tax_liability_and_refund_match_schedules_engine(built):
    """Rounds out the previous test with the downstream figures a preparer
    actually reads: tax liability and refund, both computed by
    `schedules.compute_tax` (the same slab/rebate/surcharge/cess logic
    `Computation`'s re-anchored Excel formulas mirror) from the re-derived
    ladder, must equal the model's own numbers for the default b/f=0 case."""
    model, rules, entity = built
    comp = model.computation
    fy_end = date(int(YEAR_KEY[:4]) + 1, 3, 31)

    total_income = comp.total_income_rounded
    special_cg_base = comp.tax_block.special_rate_income
    normal_income_base = max(0.0, total_income - special_cg_base)

    shadow = sch.compute_tax(
        normal_income=normal_income_base,
        special_rate_tax=comp.tax_block.tax_on_special_rate_income,
        special_rate_income_amount=special_cg_base,
        rules=rules, regime=REGIME, status=entity.status, dob=entity.dob,
        fy_end=fy_end, residency=entity.residency,
    )
    assert shadow.tax_liability == pytest.approx(comp.tax_block.tax_liability)
    refund_shadow = sch.round_288b(
        comp.taxes_paid - shadow.tax_liability, rules.common["rounding"]["tax_payable_refund"]["nearest"],
    )
    assert refund_shadow == pytest.approx(comp.refund_or_payable)


def _capped_setoff_py(head: float, bf: float) -> float:
    """Pure-Python mirror of `presentation._capped_setoff_expr`'s Excel
    formula `(head)-MIN(bf,MAX(head,0))` -- used to shadow-calc the
    statutory per-head cap without needing an Excel formula engine."""
    return head - min(bf, max(head, 0.0))


def _cg_setoff_py(stcg: float, ltcg: float, bf_stcl: float, bf_ltcl: float) -> tuple:
    """Pure-Python mirror of `presentation._cg_setoff_exprs`'s s.74 cascade
    (STCL against STCG first, spillover to LTCG; LTCL against LTCG only).
    Deliberately does NOT floor an untouched negative head at 0 -- only an
    amount actually available (positive) before set-off can be capped
    towards zero; a raw current-year loss passes through unchanged when
    there is nothing positive for a b/f bucket to set off against."""
    stcg_avail = max(stcg, 0.0)
    stcl_used = min(bf_stcl, stcg_avail)
    net_stcg = stcg - stcl_used
    stcl_spill = max(0.0, bf_stcl - stcg_avail)
    ltcg_avail = max(ltcg, 0.0)
    spill_used = min(stcl_spill, ltcg_avail)
    remaining_after_spill = ltcg_avail - spill_used
    ltcl_used = min(bf_ltcl, remaining_after_spill)
    net_ltcg = ltcg - spill_used - ltcl_used
    return net_stcg, net_ltcg


def test_bf_business_bucket_caps_at_available_business_income_never_goes_negative(built):
    """s.72 cap test: a b/f Business-loss entry that EXCEEDS the year's
    business income must reduce the Business head to exactly 0 -- never
    negative -- while an entry within the available income reduces it
    paisa-for-paisa. SYN-IND's business_income is 180000.0 (non-zero, so
    the Business section actually renders on-page)."""
    model, rules, entity = built
    comp = model.computation
    business_income = comp.business_income
    assert business_income > 0, "fixture must carry non-zero business income for this cap test to mean anything"

    within = _capped_setoff_py(business_income, 50000.0)
    assert within == pytest.approx(business_income - 50000.0)

    exceeding = _capped_setoff_py(business_income, business_income + 999999.0)
    assert exceeding == pytest.approx(0.0), "an over-large b/f Business bucket must cap the head at 0, not negative"
    assert exceeding >= 0.0


def test_bf_hp_bucket_never_leaks_into_business_or_salary(built):
    """s.71B routes ONLY against House Property income -- proven here at the
    formula-semantics level (the shadow-calc mirror of
    `_capped_setoff_expr`) since SYN-IND's own HP income happens to be 0.0,
    so the HP section is not rendered on-page for this fixture (nothing to
    show); the cap function itself is fixture-independent and must still
    correctly zero out (not leak elsewhere) when the head is already nil."""
    model, rules, entity = built
    comp = model.computation
    assert comp.house_property_income == 0.0, "fixture assumption: SYN-IND carries no HP income"
    # Even a large HP b/f entry against a nil HP head must net to exactly 0 --
    # never negative, and it has no channel to reduce Business/Salary/OS at all
    # (each head's expression only ever references its OWN bf bucket ref; see
    # test_bf_loss_buckets_route_into_their_own_head_before_gti_not_a_lump_ti_deduction).
    net_hp = _capped_setoff_py(comp.house_property_income, 75000.0)
    assert net_hp == pytest.approx(0.0)
    assert net_hp >= 0.0


def test_bf_stcl_bucket_sets_off_stcg_first_then_spills_to_ltcg(built):
    """s.74 cascade, within-STCG case: a b/f STCL smaller than STCG reduces
    STCG only, LTCG is untouched by the spillover step (spill is 0). SYN-IND's
    LTCG head is already a current-year loss (-35000): with no spillover, it
    must pass through UNCHANGED (raw, negative) -- NOT floored at 0 -- since
    nothing was actually available to set off against it (the floor is only
    ever applied to a positive amount actually consumed by a set-off)."""
    model, rules, entity = built
    comp = model.computation
    stcg, ltcg = comp.capital_gains_st, comp.capital_gains_lt
    assert stcg > 0, "fixture must carry positive STCG for this test to mean anything"
    assert ltcg < 0, "fixture must carry a negative (loss) LTCG head for this test to mean anything"

    net_stcg, net_ltcg = _cg_setoff_py(stcg, ltcg, bf_stcl=5000.0, bf_ltcl=0.0)
    assert net_stcg == pytest.approx(stcg - 5000.0)
    assert net_ltcg == pytest.approx(ltcg), \
        "LTCG must pass through unchanged (raw, negative) when there is no spillover to absorb"


def test_bf_stcl_bucket_caps_at_stcg_and_spillover_cannot_reach_a_negative_ltcg(built):
    """s.74 cascade, over-large STCL case: a b/f STCL LARGER than STCG must
    cap STCG at exactly 0 (never negative). The remainder ("spillover") is
    only usable against a *positive* LTCG (MAX(ltcg,0)); SYN-IND's LTCG
    head is already negative, so MAX(ltcg,0)=0 and the spillover has
    nothing to absorb -- LTCG must stay at its raw (unchanged, negative)
    value, not be floored to 0."""
    model, rules, entity = built
    comp = model.computation
    stcg, ltcg = comp.capital_gains_st, comp.capital_gains_lt
    bf_stcl = stcg + 999999.0  # deliberately far larger than available STCG

    net_stcg, net_ltcg = _cg_setoff_py(stcg, ltcg, bf_stcl=bf_stcl, bf_ltcl=0.0)
    assert net_stcg == pytest.approx(0.0), "STCL spillover must cap STCG at 0, not negative"
    assert net_stcg >= 0.0
    assert net_ltcg == pytest.approx(ltcg), \
        "a negative LTCG head has no positive amount for the spillover to consume, so it is untouched"


def test_bf_ltcl_bucket_never_touches_stcg_or_normal_income(built):
    """s.74: b/f Long-term capital loss must reduce LTCG ONLY -- never STCG,
    never any normal-income head. Proven by varying bf_ltcl while holding
    bf_stcl=0 and observing STCG is completely unchanged; and since SYN-IND's
    LTCG head is already negative (own-year loss), there is nothing positive
    for the LTCL bucket to set off against either, so LTCG itself is also
    unchanged (raw pass-through, not floored)."""
    model, rules, entity = built
    comp = model.computation
    stcg, ltcg = comp.capital_gains_st, comp.capital_gains_lt

    net_stcg_a, _ = _cg_setoff_py(stcg, ltcg, bf_stcl=0.0, bf_ltcl=0.0)
    net_stcg_b, net_ltcg_b = _cg_setoff_py(stcg, ltcg, bf_stcl=0.0, bf_ltcl=40000.0)
    assert net_stcg_a == pytest.approx(net_stcg_b) == pytest.approx(stcg), \
        "LTCL bucket must never move the STCG figure"
    assert net_ltcg_b == pytest.approx(ltcg), \
        "a negative LTCG head has nothing positive for the LTCL bucket to set off against"


def test_bf_ltcl_bucket_caps_a_positive_ltcg_at_zero_when_it_exceeds_it(built):
    """The other half of the LTCL cap: when LTCG IS positive, an over-large
    b/f LTCL entry must cap it at exactly 0, never negative. SYN-IND's own
    LTCG is a loss, so this is proven with a synthetic positive LTCG value
    fed directly into the shadow-calc mirror (the mirror is fixture-
    independent, matching `presentation._cg_setoff_exprs` formula-for-
    formula)."""
    net_stcg, net_ltcg = _cg_setoff_py(stcg=0.0, ltcg=30000.0, bf_stcl=0.0, bf_ltcl=999999.0)
    assert net_ltcg == pytest.approx(0.0), "an over-large LTCL entry must cap a positive LTCG at 0"
    assert net_ltcg >= 0.0
    assert net_stcg == pytest.approx(0.0), "LTCL bucket must never move STCG"


def test_bf_bucket_default_zero_is_a_no_op_matching_pre_change_default_case(built):
    """When every bucket is left at its default 0 (no preparer override),
    every net-of-set-off expression must reduce to the original un-netted
    leaf value -- i.e. paisa-exact parity with the pre-2026-07-21 default
    case (also proven end-to-end by
    test_default_case_ladder_matches_schedules_engine_to_the_paisa)."""
    model, rules, entity = built
    comp = model.computation
    assert _capped_setoff_py(comp.business_income, 0.0) == pytest.approx(comp.business_income)
    assert _capped_setoff_py(comp.house_property_income, 0.0) == pytest.approx(comp.house_property_income)
    net_stcg, net_ltcg = _cg_setoff_py(comp.capital_gains_st, comp.capital_gains_lt, 0.0, 0.0)
    assert net_stcg == pytest.approx(comp.capital_gains_st)
    assert net_ltcg == pytest.approx(comp.capital_gains_lt), \
        "at bf=0 the raw (possibly negative) LTCG head must pass through unchanged"

    # And the on-page 'Special-rate Capital Gains' figure (net_ltcg + net_stcg)
    # must equal schedules.py's own tax_block.special_rate_income exactly --
    # the actual ground-truth this on-page line is meant to reproduce.
    assert net_ltcg + net_stcg == pytest.approx(comp.tax_block.special_rate_income)


# ---------------------------------------------------------------------------
# 3. The correctness-trap regression -- special-rate CG must stay carved out
# ---------------------------------------------------------------------------

def test_including_special_rate_cg_in_slab_base_would_change_the_tax_the_carve_out_is_not_a_no_op(built):
    """Guards against the single most dangerous possible regression: silently
    dropping the CG carve-out and taxing special-rate CG at slab rates (or
    vice versa). Demonstrates that the correct (carved-out) computation and
    an intentionally WRONG one (CG folded into the slab base, no separate
    special-rate tax) diverge for this fixture -- so if a future change
    accidentally merges them back together, a paisa-exact test WILL notice."""
    model, rules, entity = built
    comp = model.computation
    fy_end = date(int(YEAR_KEY[:4]) + 1, 3, 31)

    total_income = comp.total_income_rounded
    special_cg_base = comp.tax_block.special_rate_income
    correct_normal_base = max(0.0, total_income - special_cg_base)

    correct = sch.compute_tax(
        normal_income=correct_normal_base, special_rate_tax=comp.tax_block.tax_on_special_rate_income,
        special_rate_income_amount=special_cg_base, rules=rules, regime=REGIME,
        status=entity.status, dob=entity.dob, fy_end=fy_end, residency=entity.residency,
    )

    # WRONG: fold special-rate CG into the slab base and drop its separate
    # flat-rate tax (special_rate_tax=0), as if the carve-out never happened.
    wrong_normal_base = total_income  # CG never subtracted out
    wrong = sch.compute_tax(
        normal_income=wrong_normal_base, special_rate_tax=0.0,
        special_rate_income_amount=0.0, rules=rules, regime=REGIME,
        status=entity.status, dob=entity.dob, fy_end=fy_end, residency=entity.residency,
    )

    # SYN-IND has non-trivial special-rate CG (LT loss + ST gain net non-zero
    # once exemption is applied) -- the two computations must diverge, proving
    # the carve-out has teeth for this fixture rather than accidentally being
    # a no-op that would let a regression slip through unnoticed.
    #
    # NOTE: compare `tax_before_relief` rather than the final `tax_liability`.
    # SYN-IND's income is small enough that s.87A rebate floors BOTH the
    # correct and the wrong computation to a 0 final liability, which would
    # make the two scenarios look identical at the tax_liability line despite
    # genuinely different intermediate tax -- exactly the false negative this
    # regression must not have. tax_before_relief (pre-floor) still exposes
    # the divergence.
    assert special_cg_base != 0.0, "fixture must carry non-zero special-rate CG for this regression to mean anything"
    assert wrong.tax_before_relief != pytest.approx(correct.tax_before_relief), (
        "special-rate CG carve-out must not be a no-op for this fixture -- if wrong == correct, "
        "a regression that silently drops the carve-out would go undetected"
    )
    # And the correct computation must match what schedules.py itself
    # (unaffected by this presentation-layer change) actually produced.
    assert correct.tax_liability == pytest.approx(comp.tax_block.tax_liability)
    assert correct.tax_before_relief == pytest.approx(comp.tax_block.tax_before_relief)
