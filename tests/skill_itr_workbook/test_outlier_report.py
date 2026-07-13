"""
tests/skill_itr_workbook/test_outlier_report.py -- Part 3 tests for
scripts/outlier_report.py. Two layers: (1) unit tests on compare_lines/
classify/drill_down with small hand-built inputs, so the MATCH/DIFF/
NO-REFERENCE and classification logic is verified independent of any real
fixture's actual numbers; (2) one integration test building a real ITRModel
from the SYN-IND synthetic fixtures and diffing it against the synthetic
SYN-IND filed-return JSON fixture, checking the workbook comes out with the
expected sheets/shape. Fully offline; synthetic fixtures only, no real data.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPTS = ROOT / "src" / "agents" / "skill_itr_workbook" / "scripts"
FIXTURES = Path(__file__).resolve().parent / "fixtures"
RULES_DIR = ROOT / "Data" / "itr" / "rules"

for p in (str(SCRIPTS), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

import parse_eguile as pe  # noqa: E402
import parse_gnucash as pg  # noqa: E402
import parse_form16  # noqa: E402
import configs  # noqa: E402
import mapping as mapping_engine  # noqa: E402
import rules as rules_engine  # noqa: E402
import schedules as sch  # noqa: E402
import filed_return as fr  # noqa: E402
import outlier_report as orpt  # noqa: E402
import fixture_gen  # noqa: E402

YEAR_KEY = "2024-25"


# ---------------------------------------------------------------------------
# Unit tests -- compare_lines / classify / drill_down against hand-built
# inputs (numbers chosen for the test, not tied to any real/fixture figures).
# ---------------------------------------------------------------------------

def _ref(lines):
    return orpt.ReferenceSet(entity_key="X", source_kind="filed-json", source_path="x.json", lines=lines)


def test_compare_lines_match_within_tolerance():
    our = {"A": 1000.4}
    reference = _ref({"A": orpt.ReferenceLine(1000.0)})
    diffs = orpt.compare_lines(our, reference)
    assert diffs[0].status == "MATCH"


def test_compare_lines_diff_beyond_tolerance():
    our = {"A": 1050.0}
    reference = _ref({"A": orpt.ReferenceLine(1000.0)})
    diffs = orpt.compare_lines(our, reference)
    assert diffs[0].status == "DIFF"
    assert diffs[0].diff == 50.0


def test_compare_lines_no_reference_when_not_extracted():
    our = {"A": 1000.0}
    reference = _ref({"A": orpt.ReferenceLine(orpt.NOT_EXTRACTED)})
    diffs = orpt.compare_lines(our, reference)
    assert diffs[0].status == "NO-REFERENCE"


def test_compare_lines_no_reference_when_line_missing():
    our = {"A": 1000.0}
    reference = _ref({})
    diffs = orpt.compare_lines(our, reference)
    assert diffs[0].status == "NO-REFERENCE"


def test_classify_ca_working_always_filed_side_note():
    diff = orpt.LineDiff("Salary income", 100.0, 90.0, 10.0, "DIFF", orpt.CA_WORKING_CONFIDENCE)
    assert orpt.classify(diff, []) == orpt.FILED_SIDE_NOTE


def test_classify_drillable_line_with_no_mapped_accounts_is_data_gap():
    diff = orpt.LineDiff("Salary income", 100.0, 90.0, 10.0, "DIFF", orpt.FILED_RETURN_CONFIDENCE)
    assert orpt.classify(diff, []) == orpt.DATA_GAP


def test_classify_drillable_line_with_mapped_accounts_is_unclassified_not_asserted():
    diff = orpt.LineDiff("Salary income", 100.0, 90.0, 10.0, "DIFF", orpt.FILED_RETURN_CONFIDENCE)
    drill_rows = [orpt.DrillRow("Assets:Salary", "SALARY_GROSS", 100.0)]
    assert orpt.classify(diff, drill_rows) == orpt.UNCLASSIFIED


def test_classify_composite_line_never_data_gap_even_with_no_drill_rows():
    # Gross Total Income etc. have no single tag set -- absence of drill rows
    # is expected, not a data gap.
    diff = orpt.LineDiff("Gross Total Income", 100.0, 90.0, 10.0, "DIFF", orpt.FILED_RETURN_CONFIDENCE)
    assert orpt.classify(diff, []) == orpt.UNCLASSIFIED


def test_classify_match_has_no_classification():
    diff = orpt.LineDiff("Salary income", 100.0, 100.0, 0.0, "MATCH", orpt.FILED_RETURN_CONFIDENCE)
    assert orpt.classify(diff, []) == ""


# ---------------------------------------------------------------------------
# Integration test -- real ITRModel (SYN-IND fixtures) vs synthetic filed
# return JSON fixture.
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def syn_ind_model_and_mapping():
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    book = pg.parse_book(FIXTURES / "syn_ind.gnucash")
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    result = mapping_engine.resolve_tree(tree, loaded)
    form16 = parse_form16.parse_form16(FIXTURES / "syn_ind_form16.pdf")
    rules = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    entities = configs.load_entities(ROOT / "Data" / "itr" / "entities.example.yaml")
    entity = entities["SYN-IND"]
    scrips = configs.load_scrips(ROOT / "Data" / "itr" / "scrips.example.yaml")
    fmv_tables = sch.load_fmv_tables()

    model = sch.build_all_schedules(
        tree, result.resolved, book, form16, YEAR_KEY, rules, "new",
        entity.status, entity.dob, scrips, fmv_tables,
    )
    node_by_guid = {n.guid: n for n in tree.all_nodes() if n.guid}
    return model, result.resolved, node_by_guid


def test_outlier_report_end_to_end(tmp_path, syn_ind_model_and_mapping):
    model, resolved, node_by_guid = syn_ind_model_and_mapping
    filed = fr.parse_filed_return(FIXTURES / "syn_ind_filed_return.json", "SYN-IND")
    reference = orpt.reference_from_filed_return(filed)

    out_path = tmp_path / "SYN-IND-outliers.xlsx"
    diffs = orpt.run_outlier_report(model, reference, resolved, node_by_guid, str(out_path))

    assert out_path.exists()
    line_names = {d.line_name for d in diffs}
    assert line_names == set(orpt.build_our_lines(model).keys())
    assert all(d.status in ("MATCH", "DIFF", "NO-REFERENCE") for d in diffs)

    wb = openpyxl.load_workbook(str(out_path))
    assert set(wb.sheetnames) == {
        "Section A - Comparison", "Section B - Drilldown",
        "Section C - Classification", "Summary",
    }
    summary_rows = {row[0]: row[1] for row in wb["Summary"].iter_rows(values_only=True) if row and row[0]}
    assert summary_rows["Entity"] == "SYN-IND"
    assert summary_rows["Reference source"] == "filed-json"
    assert summary_rows["Lines compared"] == len(diffs)


def test_reference_from_filed_return_flags_not_extracted_lines():
    itr_only_salary = {"entity_key": "SYN-IND", "source_format": "json", "source_path": "x"}
    filed = fr.FiledReturn(**itr_only_salary, heads={"Salaries": 500000}, tax_liability={})
    reference = orpt.reference_from_filed_return(filed)
    assert reference.lines["Salary income"].value == 500000
    assert reference.lines["Gross Total Income"].value == orpt.NOT_EXTRACTED
    assert reference.lines["Net Tax Liability"].value == orpt.NOT_EXTRACTED
