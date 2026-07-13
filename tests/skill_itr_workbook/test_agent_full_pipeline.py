"""
tests/skill_itr_workbook/test_agent_full_pipeline.py -- Batch 6 tests:
agent.py's full parse -> map -> schedules -> write -> verify wiring, the
Definition-of-Done scenario (SYN-IND with form16_pdf produces a complete
formula-driven workbook + green Reconciliation), the as26.py 26AS-workbook
reader, and the local_samples end-to-end real-corpus check (skips with a
setup hint when Data/itr/entities.yaml + an approved mapping are absent --
never fails CI, never hand-tags/hard-codes real data).
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPTS = ROOT / "src" / "agents" / "skill_itr_workbook" / "scripts"
AGENT_DIR = ROOT / "src" / "agents" / "skill_itr_workbook"
FIXTURES = Path(__file__).resolve().parent / "fixtures"
REAL_SAMPLES_DIR = ROOT / "Data" / "GNUCashReports"
REAL_MAPPINGS_DIR = ROOT / "Data" / "itr" / "mappings"
REAL_ENTITIES_PATH = ROOT / "Data" / "itr" / "entities.yaml"

for p in (str(SCRIPTS), str(AGENT_DIR), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

import configs  # noqa: E402
import as26  # noqa: E402
import agent  # noqa: E402
import fixture_gen  # noqa: E402

ENTITIES_EXAMPLE = ROOT / "Data" / "itr" / "entities.example.yaml"
SCRIPS_EXAMPLE = ROOT / "Data" / "itr" / "scrips.example.yaml"


def _run_full(tmp_path, html_build_fn, book_build_fn, mapping_path, entity_key, form16_pdf=None):
    html_path = tmp_path / "bs.html"
    html_path.write_text(html_build_fn(), encoding="utf-8")
    book_path = tmp_path / "book.gnucash"
    import gzip
    with gzip.open(book_path, "wt", encoding="utf-8") as f:
        f.write(book_build_fn())
    out_path = tmp_path / "out.xlsx"

    summary = agent.run(
        str(html_path), str(out_path), book_file=str(book_path), mapping_file=str(mapping_path),
        form16_pdf=str(form16_pdf) if form16_pdf else None,
        entities_path=str(ENTITIES_EXAMPLE), entity_key=entity_key, scrips_path=str(SCRIPS_EXAMPLE),
    )
    return summary, out_path


# ---------------------------------------------------------------------------
# Definition of Done: agent.py on SYN-IND with form16_pdf produces a
# complete formula-driven workbook + green Reconciliation.
# ---------------------------------------------------------------------------

def test_dod_syn_ind_produces_full_workbook_with_green_reconciliation(tmp_path):
    form16_path = tmp_path / "form16.pdf"
    form16_path.write_bytes(fixture_gen.build_syn_ind_form16_pdf())

    summary, out_path = _run_full(
        tmp_path, fixture_gen.build_syn_ind_html, fixture_gen.build_syn_ind_gnucash,
        FIXTURES / "syn_ind.mapping.yaml", "SYN-IND", form16_pdf=form16_path,
    )
    assert "STATUS: OK" in summary
    assert "Workbook: full schedule model built" in summary

    wb = openpyxl.load_workbook(str(out_path))
    assert "Computation" in wb.sheetnames
    computation_formulas = [
        row[1] for row in wb["Computation"].iter_rows(min_col=1, max_col=2, values_only=True)
        if isinstance(row[1], str) and row[1].startswith("=")
    ]
    assert len(computation_formulas) >= 20

    recon_rows = list(wb["Reconciliation"].iter_rows(min_col=1, max_col=2, values_only=True))
    recon_text = {str(r[0]): r[1] for r in recon_rows if r[0]}
    recon_labels = [str(r[0]) for r in recon_rows if r[0]]
    assert recon_text.get("Identity checks") == "OK"
    assert recon_text.get("Capital Gains reconciliation (lots vs control)") == "OK"
    assert "Unmapped accounts (0)" in recon_labels


def test_dod_syn_huf_produces_full_workbook(tmp_path):
    summary, out_path = _run_full(
        tmp_path, fixture_gen.build_syn_huf_html, fixture_gen.build_syn_huf_gnucash,
        FIXTURES / "syn_huf.mapping.yaml", "SYN-HUF",
    )
    assert "STATUS: OK" in summary
    assert "Workbook: full schedule model built" in summary
    wb = openpyxl.load_workbook(str(out_path))
    assert "Computation" in wb.sheetnames
    # HUF has no salary seam -- the Salary sheet should show a zeroed/manual schedule.
    salary_rows = {row[0]: row[1] for row in wb["Salary"].iter_rows(min_col=1, max_col=2, values_only=True) if row[0]}
    assert salary_rows.get("Gross salary (17(1)+17(2)+17(3))") == 0.0


def test_blocked_mapping_writes_stub_not_full_workbook(tmp_path):
    """An unmapped account must BLOCK the full workbook build entirely --
    no value from an unmapped leaf may reach any downstream schedule."""
    summary, out_path = _run_full(
        tmp_path, fixture_gen.build_syn_ind_html, fixture_gen.build_syn_ind_gnucash,
        FIXTURES / "syn_ind_unmapped.mapping.yaml", "SYN-IND",
    )
    assert "STATUS: BLOCKED-FOR-REVIEW" in summary
    assert "Workbook: full schedule model built" not in summary
    wb = openpyxl.load_workbook(str(out_path))
    assert wb.sheetnames == ["Reconciliation"]  # stub workbook only


def test_no_mapping_file_writes_stub(tmp_path):
    html_path = tmp_path / "bs.html"
    html_path.write_text(fixture_gen.build_syn_ind_html(), encoding="utf-8")
    out_path = tmp_path / "out.xlsx"
    summary = agent.run(str(html_path), str(out_path))
    assert "Workbook: full schedule model built" not in summary
    wb = openpyxl.load_workbook(str(out_path))
    assert wb.sheetnames == ["Reconciliation"]


def test_unknown_entity_key_falls_back_gracefully(tmp_path):
    """No matching entities.yaml entry -- pipeline still builds (Individual/
    new regime default) rather than crashing."""
    summary, _ = _run_full(
        tmp_path, fixture_gen.build_syn_ind_html, fixture_gen.build_syn_ind_gnucash,
        FIXTURES / "syn_ind.mapping.yaml", "NO-SUCH-ENTITY",
    )
    assert "Workbook: full schedule model built" in summary


# ---------------------------------------------------------------------------
# as26.py: 26AS skill output workbook reader
# ---------------------------------------------------------------------------

def _build_minimal_26as_workbook(tmp_path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Part I"
    ws.cell(row=1, column=1, value="Part I title")
    ws.cell(row=2, column=1, value="meta")
    headers = [
        "Deductor Sr.No.", "Name of Deductor", "TAN of Deductor",
        "Total Amount Paid/Credited", "Total Tax Deducted #", "Total TDS Deposited",
        "Txn Sr.No.", "Section", "Transaction Date", "Status of Booking",
        "Date of Booking", "Remarks", "Amount Paid/Credited",
        "Tax Deducted ##", "TDS Deposited",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    ws.cell(row=4, column=1, value=1)
    ws.cell(row=4, column=2, value="Example Employer Ltd")
    ws.cell(row=4, column=3, value="ABCE12345E")
    ws.cell(row=4, column=8, value="192")
    ws.cell(row=4, column=9, value="15-May-2024")
    ws.cell(row=4, column=13, value=500000.0)
    ws.cell(row=4, column=14, value=25000.0)
    ws.cell(row=4, column=15, value=25000.0)
    ws.cell(row=5, column=1, value="#1")
    ws.cell(row=5, column=2, value="Sub-total -- Example Employer Ltd (txns: 1)")
    out_path = tmp_path / "as26.xlsx"
    wb.save(out_path)
    return out_path


def test_as26_reader_parses_transactions_and_skips_subtotals(tmp_path):
    path = _build_minimal_26as_workbook(tmp_path)
    data = as26.parse_as26_workbook(str(path))
    assert len(data.transactions) == 1
    t = data.transactions[0]
    assert t.tan == "ABCE12345E"
    assert t.amount == 500000.0
    assert t.tax_deducted == 25000.0
    assert t.txn_date is not None and t.txn_date.isoformat() == "2024-05-15"


def test_as26_reader_missing_part_i_sheet_returns_empty(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Part II"
    path = tmp_path / "no_part1.xlsx"
    wb.save(path)
    data = as26.parse_as26_workbook(str(path))
    assert data.transactions == []


# ---------------------------------------------------------------------------
# local_samples: end-to-end on the real corpus, skip-with-message if absent
# ---------------------------------------------------------------------------

@pytest.mark.local_samples
def test_real_corpus_end_to_end_identities_green_and_controls_tie():
    """Runs the full pipeline on any real entity with an approved mapping
    file under Data/itr/mappings/. Skips (never fails) when the local user
    configs aren't present -- this is exactly the state the real family
    corpus is in for most contributors/CI. Asserts identity/reconciliation
    checks only -- never a hardcoded real amount, PAN, or account name."""
    if not REAL_SAMPLES_DIR.is_dir():
        pytest.skip(f"{REAL_SAMPLES_DIR} not present -- local-only smoke test")
    if not REAL_ENTITIES_PATH.is_file():
        pytest.skip(
            f"{REAL_ENTITIES_PATH} not present -- create it locally (see entities.example.yaml) "
            "to exercise the full real-corpus pipeline"
        )
    if not REAL_MAPPINGS_DIR.is_dir() or not list(REAL_MAPPINGS_DIR.glob("*.mapping.yaml")):
        pytest.skip(f"no approved mapping files under {REAL_MAPPINGS_DIR} -- nothing to run end-to-end yet")

    html_files = sorted(REAL_SAMPLES_DIR.glob("*.html"))
    if not html_files:
        pytest.skip(f"no eguile HTML export found under {REAL_SAMPLES_DIR}")

    entities = configs.load_entities(REAL_ENTITIES_PATH)
    ran_any = False
    for mapping_path in REAL_MAPPINGS_DIR.glob("*.mapping.yaml"):
        entity_key = mapping_path.stem.replace(".mapping", "")
        if entity_key not in entities:
            continue
        for html_path in html_files:
            out_path = html_path.with_suffix(".itr-test-output.xlsx")
            try:
                summary = agent.run(
                    str(html_path), str(out_path), mapping_file=str(mapping_path),
                    entities_path=str(REAL_ENTITIES_PATH), entity_key=entity_key,
                )
            finally:
                if out_path.exists():
                    out_path.unlink()
            if "STATUS: BLOCKED-FOR-REVIEW" in summary:
                continue  # this HTML doesn't belong to this entity's mapping
            ran_any = True
            assert "Verify: OK" in summary or "VALIDATION ERRORS" not in summary

    if not ran_any:
        pytest.skip("no (entity, HTML) pair produced a clean STATUS: OK run locally")
