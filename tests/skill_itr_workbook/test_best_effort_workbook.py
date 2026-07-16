"""
tests/skill_itr_workbook/test_best_effort_workbook.py -- 2026-07-16 Part 1
acceptance gates ("Best-effort workbook instead of block-to-nothing", see
2026-07-16-itr-besteffort-workbook-and-mapping-ui-prompt.md):

  1. Incomplete mapping (3 dropped tags: one BS-side, two RE-Income-side)
     still builds a full workbook; BS Assets == Equity+Liabilities (a
     tree-level invariant, independent of mapping -- parse_eguile's own
     identity check) AND the 3 unclassified amounts land in a visible
     UNCLASSIFIED bucket that is included in that bucket's own total.
  2. The IT working shows DRAFT tax (resolved items only) + a worst-case
     upper bound (unclassified income-type items taxed at the top slab),
     both labelled, with the unclassified count/amount called out, and no
     cell claims to be a final/filing-ready total.
  3. A fully-mapped run (0 unmapped) is behaviourally identical to a
     pre-Part-1 run: no DRAFT stamp, no Unclassified sheet, tax shown as
     final.
  4. Hard-error paths (bad HTML / unresolved entity / AY mismatch) are
     unchanged -- covered in test_agent_full_pipeline.py /
     test_parse_eguile.py; re-asserted lightly here for completeness.
  5. <output>-proposed-mappings.yaml is still written whenever N > 0.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPTS = ROOT / "src" / "agents" / "skill_itr_workbook" / "scripts"
AGENT_DIR = ROOT / "src" / "agents" / "skill_itr_workbook"
FIXTURES = Path(__file__).resolve().parent / "fixtures"

for p in (str(SCRIPTS), str(AGENT_DIR), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

import agent  # noqa: E402
import configs  # noqa: E402
import fixture_gen  # noqa: E402
import mapping as mapping_engine  # noqa: E402
import parse_eguile as pe  # noqa: E402
import rules as rules_engine  # noqa: E402
import schedules as sch  # noqa: E402

ENTITIES_EXAMPLE = ROOT / "Data" / "itr" / "entities.example.yaml"
SCRIPS_EXAMPLE = ROOT / "Data" / "itr" / "scrips.example.yaml"
RULES_DIR = ROOT / "Data" / "itr" / "rules"
YEAR_KEY = "2024-25"

# From fixture_gen.build_syn_ind_html: the 3 leaves dropped by
# syn_ind_partial.mapping.yaml.
MISC_HOLDING_AMOUNT = 5000.00          # Assets/Misc Holding (unmapped) -- BS-side
BUSINESS_REMUNERATION_AMOUNT = 300000.00  # Income/Business Remuneration -- RE-Income
BANK_INTEREST_AMOUNT = 20000.00        # Income/Bank Interest -- RE-Income
UNCLASSIFIED_TOTAL = MISC_HOLDING_AMOUNT + BUSINESS_REMUNERATION_AMOUNT + BANK_INTEREST_AMOUNT
UNCLASSIFIED_INCOME_TYPE_TOTAL = BUSINESS_REMUNERATION_AMOUNT + BANK_INTEREST_AMOUNT


def _run_partial(tmp_path):
    html_path = tmp_path / "bs.html"
    html_path.write_text(fixture_gen.build_syn_ind_html(), encoding="utf-8")
    out_path = tmp_path / "out.xlsx"
    summary = agent.run(
        str(html_path), str(out_path), mapping_file=str(FIXTURES / "syn_ind_partial.mapping.yaml"),
        entities_path=str(ENTITIES_EXAMPLE), entity_key="SYN-IND", scrips_path=str(SCRIPS_EXAMPLE),
    )
    return summary, out_path


# ---------------------------------------------------------------------------
# Gate 1 -- incomplete mapping still builds; BS tallies; UNCLASSIFIED bucket
# includes the 3 dropped amounts.
# ---------------------------------------------------------------------------

def test_partial_mapping_builds_workbook_and_bs_identity_holds(tmp_path):
    summary, out_path = _run_partial(tmp_path)
    assert "STATUS: BUILT -- 3 REVIEW ITEM(S)" in summary
    assert "Workbook: full schedule model built" in summary

    tree = pe.parse_file(str(Path(tmp_path / "bs.html")))
    assert abs(tree.imbalance) < 0.01
    total_assets = tree.section_totals["Assets Accounts"]
    etl = tree.section_totals["Equity, Trading, and Liabilities"]
    total_re = tree.section_totals["Retained Earnings"]
    assert total_assets == pytest.approx(etl + total_re, abs=0.01)

    wb = openpyxl.load_workbook(str(out_path))
    assert "Unclassified" in wb.sheetnames
    ws = wb["Unclassified"]
    rows = {row[0]: row[1] for row in ws.iter_rows(min_col=1, max_col=2, values_only=True) if row[0]}
    assert rows["TOTAL unclassified (all sections)"] == pytest.approx(UNCLASSIFIED_TOTAL)
    assert rows["  of which income-type (worst-case tax base -- see Computation)"] == pytest.approx(
        UNCLASSIFIED_INCOME_TYPE_TOTAL
    )
    paths = [row[0] for row in ws.iter_rows(min_col=1, max_col=1, values_only=True)]
    assert "Assets/Misc Holding (unmapped)" in paths
    assert "Income/Business Remuneration" in paths
    assert "Income/Bank Interest" in paths


def test_build_unclassified_schedule_sums_dropped_leaves():
    """Unit-level: schedules.build_unclassified routes every unmapped leaf
    (all sections) into the bucket and separates out the RE-Income subset
    used for the worst-case tax base."""
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    loaded = configs.load_mapping(FIXTURES / "syn_ind_partial.mapping.yaml")
    result = mapping_engine.resolve_tree(tree, loaded)
    assert len(result.unmapped) == 3

    uncl = sch.build_unclassified(tree, result.unmapped)
    assert uncl.count == 3
    assert uncl.total_amount == pytest.approx(UNCLASSIFIED_TOTAL)
    assert uncl.income_type_total == pytest.approx(UNCLASSIFIED_INCOME_TYPE_TOTAL)
    income_type_paths = {i.path for i in uncl.items if i.is_income_type}
    assert income_type_paths == {"Income/Business Remuneration", "Income/Bank Interest"}


# ---------------------------------------------------------------------------
# Gate 2 -- IT working: DRAFT + worst-case upper bound, both labelled, no
# final/filing-ready total.
# ---------------------------------------------------------------------------

def test_computation_sheet_shows_draft_and_worst_case_labelled(tmp_path):
    summary, out_path = _run_partial(tmp_path)
    assert "3 account(s) unclassified" in summary

    wb = openpyxl.load_workbook(str(out_path))
    ws = wb["Computation"]
    labels = [row[0] for row in ws.iter_rows(min_col=1, max_col=1, values_only=True) if row[0]]
    joined = "\n".join(str(l) for l in labels)

    draft_labels = [l for l in labels if "DRAFT" in str(l)]
    worst_case_labels = [l for l in labels if "Worst-case upper bound" in str(l)]
    assert draft_labels, "no DRAFT-labelled row found on Computation"
    assert worst_case_labels, "no worst-case-upper-bound row found on Computation"
    assert "not filing-ready" in joined.lower() or "NOT filing-ready" in joined

    # The plain "Tax liability (selected regime)" row (no DRAFT/worst-case
    # qualifier) must not appear -- every liability-shaped row is qualified.
    assert not any(
        str(l) == "Tax liability (selected regime)" for l in labels
    ), "an unqualified 'Tax liability' row would look filing-ready while unclassified items exist"


def test_worst_case_tax_liability_is_at_least_draft(tmp_path):
    """The worst-case upper bound must never be LESS than the DRAFT figure
    -- unclassified expense/deduction-side items are never assumed to
    reduce tax (conservative, plan decision locked 2026-07-16)."""
    entities = configs.load_entities(ENTITIES_EXAMPLE)
    entity = entities["SYN-IND"]
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    loaded = configs.load_mapping(FIXTURES / "syn_ind_partial.mapping.yaml")
    result = mapping_engine.resolve_tree(tree, loaded)
    rules = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    scrips = configs.load_scrips(SCRIPS_EXAMPLE)
    fmv_tables = sch.load_fmv_tables()

    model = sch.build_all_schedules(
        tree, result.resolved, None, None, YEAR_KEY, rules, "new",
        entity.status, entity.dob, scrips, fmv_tables, None, result.unmapped,
    )
    assert model.unclassified.count == 3
    assert model.computation.worst_case_tax_liability >= model.computation.tax_block.tax_liability
    # The two RE-Income drops are worth 320,000.00 -- the worst-case extra
    # tax must be strictly positive (some of that lands above a 0% slab).
    assert model.computation.worst_case_extra_tax > 0.0
    assert model.computation.worst_case_tax_liability == pytest.approx(
        model.computation.tax_block.tax_liability + model.computation.worst_case_extra_tax
    )


# ---------------------------------------------------------------------------
# Gate 3 -- fully-mapped run stays behaviourally identical: no DRAFT stamp,
# no Unclassified sheet, tax shown as final.
# ---------------------------------------------------------------------------

def test_fully_mapped_run_has_no_draft_stamp_or_unclassified_sheet(tmp_path):
    html_path = tmp_path / "bs.html"
    html_path.write_text(fixture_gen.build_syn_ind_html(), encoding="utf-8")
    out_path = tmp_path / "out.xlsx"
    summary = agent.run(
        str(html_path), str(out_path), mapping_file=str(FIXTURES / "syn_ind.mapping.yaml"),
        entities_path=str(ENTITIES_EXAMPLE), entity_key="SYN-IND", scrips_path=str(SCRIPS_EXAMPLE),
    )
    assert "STATUS: OK" in summary
    assert "DRAFT" not in summary
    assert "unclassified" not in summary.lower()

    wb = openpyxl.load_workbook(str(out_path))
    assert "Unclassified" not in wb.sheetnames
    ws = wb["Computation"]
    labels = [row[0] for row in ws.iter_rows(min_col=1, max_col=1, values_only=True) if row[0]]
    assert "Tax liability (selected regime)" in labels
    assert "Refund (+) / Payable (-), s.288B rounded" in labels
    assert not any("DRAFT" in str(l) for l in labels)
    assert not any("Worst-case" in str(l) for l in labels)


def test_fully_mapped_unclassified_schedule_is_empty():
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    result = mapping_engine.resolve_tree(tree, loaded)
    assert not result.blocked
    uncl = sch.build_unclassified(tree, result.unmapped)
    assert uncl.count == 0
    assert uncl.total_amount == 0.0
    assert uncl.income_type_total == 0.0


# ---------------------------------------------------------------------------
# Gate 4 -- hard-error paths unchanged (lightly re-asserted here; primary
# coverage lives in test_agent_full_pipeline.py / test_path_anchoring.py).
# ---------------------------------------------------------------------------

def test_hard_error_unresolved_entity_still_fails_loud_with_stub(tmp_path):
    html_path = tmp_path / "bs.html"
    html_path.write_text(fixture_gen.build_syn_ind_html(), encoding="utf-8")
    out_path = tmp_path / "out.xlsx"
    summary = agent.run(
        str(html_path), str(out_path), mapping_file=str(FIXTURES / "syn_ind_partial.mapping.yaml"),
        entities_path=str(ENTITIES_EXAMPLE), entity_key="NO-SUCH-ENTITY", scrips_path=str(SCRIPS_EXAMPLE),
    )
    assert "ERROR:" in summary
    assert "Workbook: full schedule model built" not in summary
    wb = openpyxl.load_workbook(str(out_path))
    assert wb.sheetnames == ["Reconciliation"]


# ---------------------------------------------------------------------------
# Gate 5 -- proposed-mappings.yaml still written whenever N > 0.
# ---------------------------------------------------------------------------

def test_proposed_mappings_snippet_still_written_when_built_with_review_items(tmp_path):
    summary, out_path = _run_partial(tmp_path)
    snippet_path = Path(str(out_path) + "-proposed-mappings.yaml")
    assert snippet_path.exists()
    text = snippet_path.read_text(encoding="utf-8")
    assert "Assets/Misc Holding (unmapped)" in text
    assert "Income/Business Remuneration" in text
    assert "Income/Bank Interest" in text
