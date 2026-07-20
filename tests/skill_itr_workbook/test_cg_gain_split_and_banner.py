"""
tests/skill_itr_workbook/test_cg_gain_split_and_banner.py -- regression
tests for the 2026-07-19 CG gain-split-vs-action fix.

Part 1 (scripts/lots.py): real GnuCash books never stamp an LTCG/STCG
`action` on the booked-gain INCOME split -- only the stock assistant's
Buy/Sell actions get auto-stamped. The pre-fix classifier in
`_sale_transactions()` required `sp.action in {"LTCG", "STCG"}` to treat a
split as the gain leg, so a real book's `action=None` gain split fell into
`other_splits` (proceeds) instead, cancelling real proceeds and collapsing
computed gain to ~0. `test_income_split_with_action_none_yields_correct_gain`
below is the regression test: it constructs a synthetic disposal
transaction with an `action=None` INCOME split directly (parse_gnucash's
Book/Account/Split/Transaction dataclasses -- no XML/fixture-file needed)
and asserts the correct non-zero gain is recovered.

Part 2 (scripts/presentation.py + agent.py): a genuine CG reconciliation
mismatch must "fail loud, banner, no abort" -- the workbook is still
written in full, but with a prominent ERROR banner at the top of both the
CG and Statement of Income sheets, and the run's summary carries
presentation.CG_RECONCILIATION_ERROR_MARKER so a process-level caller
(agent.py's new `main()` CLI wrapper) can exit non-zero without the
workbook itself being blocked or deleted.

Fully offline; synthetic fixtures only (fixture_gen.py / hand-built
Book objects). No real account numbers, PANs, or names.
"""
from __future__ import annotations

import sys
from datetime import date
from fractions import Fraction
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPTS = ROOT / "src" / "agents" / "skill_itr_workbook" / "scripts"
AGENT_DIR = ROOT / "src" / "agents" / "skill_itr_workbook"
FIXTURES = Path(__file__).resolve().parent / "fixtures"
RULES_DIR = ROOT / "Data" / "itr" / "rules"

for p in (str(SCRIPTS), str(AGENT_DIR), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

import parse_gnucash as pg  # noqa: E402
import configs  # noqa: E402
import rules as rules_engine  # noqa: E402
import schedules as sch  # noqa: E402
import lots as lots_engine  # noqa: E402
import write_workbook as ww  # noqa: E402
import presentation  # noqa: E402
import fixture_gen  # noqa: E402
import parse_eguile as pe  # noqa: E402
import parse_form16  # noqa: E402
import mapping as mapping_engine  # noqa: E402
import agent  # noqa: E402

YEAR_KEY = "2024-25"


# ---------------------------------------------------------------------------
# Part 1: lots.py -- gain split detected by account TYPE, not `action`
# ---------------------------------------------------------------------------

def _build_synthetic_disposal_book() -> pg.Book:
    """A minimal, hand-built book: buy 100 units of TestScrip at cost 10000
    on 2020-01-01, then sell all 100 on 2024-06-01 (> 1yr later, so LT) for
    proceeds 15000, booking a gain of 5000 through a plain INCOME split that
    carries NO `action` (action=None) -- exactly the real-book shape the
    bug report describes; GnuCash's stock assistant only ever auto-stamps
    Buy/Sell, never LTCG/STCG."""
    accounts = {
        "root": pg.Account(guid="root", name="Root", type="ROOT", parent_guid=None),
        "stock": pg.Account(guid="stock", name="TestScrip", type="STOCK", parent_guid="root"),
        "bank": pg.Account(guid="bank", name="TestBank", type="ASSET", parent_guid="root"),
        "gain": pg.Account(guid="gain", name="TestCapGain", type="INCOME", parent_guid="root"),
    }
    pg._build_paths(accounts)

    buy_txn = pg.Transaction(
        guid="buy-1", date_posted=date(2020, 1, 1), description="Buy TestScrip",
        splits=[
            pg.Split(guid="s-buy-stock", account_guid="stock", value=Fraction(10000), quantity=Fraction(100)),
            pg.Split(guid="s-buy-bank", account_guid="bank", value=Fraction(-10000), quantity=Fraction(-10000)),
        ],
    )
    sell_txn = pg.Transaction(
        guid="sell-1", date_posted=date(2024, 6, 1), description="Sell TestScrip",
        splits=[
            pg.Split(guid="s-sell-stock", account_guid="stock", value=Fraction(-10000), quantity=Fraction(-100)),
            pg.Split(guid="s-sell-bank", account_guid="bank", value=Fraction(15000), quantity=Fraction(15000)),
            # The booked-gain leg -- action=None, exactly as real books enter it.
            pg.Split(guid="s-sell-gain", account_guid="gain", value=Fraction(-5000), quantity=Fraction(-5000), action=None),
        ],
    )
    return pg.Book(accounts=accounts, transactions=[buy_txn, sell_txn])


def test_income_split_with_action_none_yields_correct_gain():
    book = _build_synthetic_disposal_book()
    reconciliations = lots_engine.reconstruct_lots(book, YEAR_KEY)
    assert len(reconciliations) == 1
    recon = reconciliations[0]

    # This is the actual regression check: pre-fix, the action=None gain
    # split fell into `other_splits` (proceeds), so total_proceeds became
    # 15000 + (-5000) = 10000 == cost, and gain_splits was empty -- both
    # booked_gain and every lot's gain collapsed to 0.0 instead of 5000.0.
    assert recon.booked_gain == pytest.approx(5000.0)
    assert len(recon.lots) == 1
    lot = recon.lots[0]
    assert lot.attribution == "matched"
    assert lot.gain == pytest.approx(5000.0)
    assert lot.proceeds == pytest.approx(15000.0)
    assert lot.cost == pytest.approx(10000.0)
    assert recon.ok  # booked_gain (5000) matches lot_gain_sum (5000)


def test_income_split_with_explicit_ltcg_action_still_classified_as_gain():
    """Books that DO set action="LTCG"/"STCG" on the gain split must keep
    working exactly as before -- the fix must not regress that case."""
    book = _build_synthetic_disposal_book()
    for sp in book.transactions[1].splits:
        if sp.account_guid == "gain":
            sp.action = "LTCG"
    reconciliations = lots_engine.reconstruct_lots(book, YEAR_KEY)
    recon = reconciliations[0]
    assert recon.booked_gain == pytest.approx(5000.0)
    assert recon.lots[0].gain == pytest.approx(5000.0)


# ---------------------------------------------------------------------------
# Part 2a: schedules.py -- a genuine control mismatch sets
# reconciliation_ok False (this already worked pre-fix; asserted here so
# the banner/exit-code tests below build on a demonstrably real signal).
# ---------------------------------------------------------------------------

class _FakeLeaf:
    def __init__(self, guid, tag):
        self.guid = guid
        self.tag = tag
        self.path = f"fake/{tag}"
        self.flags = []


class _FakeNode:
    def __init__(self, total):
        self.total = total


def test_genuine_control_mismatch_sets_reconciliation_not_ok():
    rules = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    scrips = configs.load_scrips(ROOT / "Data" / "itr" / "scrips.example.yaml")
    fmv_tables = sch.load_fmv_tables()
    book = pg.parse_book(FIXTURES / "syn_ind.gnucash")

    # A deliberately wrong CG_LT_CONTROL total (real books' control comes
    # from a mapped BS/computation leaf tagged CG_LT_CONTROL -- here it's
    # forced to disagree with the book's own booked-gain splits).
    resolved = {"g1": _FakeLeaf("g1", "CG_LT_CONTROL")}
    node_by_guid = {"g1": _FakeNode(999999.0)}

    cg = sch.build_capital_gains(resolved, node_by_guid, book, YEAR_KEY, rules, scrips, fmv_tables)
    assert not cg.reconciliation_ok
    assert abs(cg.reconciliation_diff) > 0.01


# ---------------------------------------------------------------------------
# Part 2b: presentation.py -- top-of-sheet ERROR banner on CG and Statement
# of Income when reconciliation_ok is False; workbook is still fully written.
# ---------------------------------------------------------------------------

@pytest.fixture()
def syn_ind_model_and_paths():
    tree = pe.parse_html(fixture_gen.build_syn_ind_html())
    book = pg.parse_book(FIXTURES / "syn_ind.gnucash")
    loaded = configs.load_mapping(FIXTURES / "syn_ind.mapping.yaml")
    result = mapping_engine.resolve_tree(tree, loaded)
    form16 = parse_form16.parse_form16(FIXTURES / "syn_ind_form16.pdf")
    rules = rules_engine.load_rules(RULES_DIR, YEAR_KEY)
    entities = configs.load_entities(ROOT / "Data" / "itr" / "entities.example.yaml")
    entity = entities["SYN-IND"]
    scrips = configs.load_scrips(ROOT / "Data" / "itr" / "scrips.example.yaml")
    fmv_tables = sch.load_fmv_tables()
    user_rules = rules_engine.load_user_rules(RULES_DIR / "user_rules.yaml")

    model = sch.build_all_schedules(
        tree, result.resolved, book, form16, YEAR_KEY, rules, "new",
        entity.status, entity.dob, scrips, fmv_tables,
    )
    return tree, model, rules, user_rules, entity, result, loaded


def _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded):
    out_path = tmp_path / "out.xlsx"
    ww.write_workbook(
        str(out_path), tree, model, rules, user_rules, entity, "new", YEAR_KEY,
        None, [], [], [], result.unmapped, "v1", "2026-01-01T00:00:00", {},
        result.resolved, loaded.entries,
    )
    return openpyxl.load_workbook(str(out_path))


def _cell_values(ws, max_row=5):
    values = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
        values.extend(v for v in row if v is not None)
    return values


def test_reconciled_book_has_no_error_banner(tmp_path, syn_ind_model_and_paths):
    """Sanity/control case: the shared syn_ind fixture reconciles cleanly
    (test_schedules.py already asserts this), so no banner should appear."""
    tree, model, rules, user_rules, entity, result, loaded = syn_ind_model_and_paths
    assert model.capital_gains.reconciliation_ok
    wb = _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded)
    for sheet in ("CG", "Statement of Income"):
        text = " ".join(str(v) for v in _cell_values(wb[sheet]))
        assert presentation.CG_RECONCILIATION_ERROR_MARKER not in text


def test_cg_mismatch_writes_error_banner_on_cg_and_statement_of_income(tmp_path, syn_ind_model_and_paths):
    tree, model, rules, user_rules, entity, result, loaded = syn_ind_model_and_paths

    # Simulate a genuine, already-detected control mismatch (the mechanics
    # of HOW reconciliation_ok/reconciliation_diff get computed are
    # build_capital_gains' job, exercised above and unchanged by this fix --
    # this test is about what the write/presentation layer DOES with that
    # signal: banner, no abort).
    model.capital_gains.reconciliation_ok = False
    model.capital_gains.reconciliation_diff = 1234.56

    wb = _write_and_load(tmp_path, tree, model, rules, user_rules, entity, result, loaded)

    # The workbook must still be fully produced -- never refused/stubbed.
    assert "Computation" in wb.sheetnames
    assert "CG" in wb.sheetnames
    assert "Statement of Income" in wb.sheetnames

    for sheet in ("CG", "Statement of Income"):
        text = " ".join(str(v) for v in _cell_values(wb[sheet]))
        assert presentation.CG_RECONCILIATION_ERROR_MARKER in text
        assert "1,234.56" in text or "1234.56" in text


# ---------------------------------------------------------------------------
# Part 2c: agent.py -- run() summary carries the marker, and the new main()
# CLI wrapper turns that into a non-zero process exit code + stderr line,
# without ever refusing to write the workbook.
# ---------------------------------------------------------------------------

def test_build_and_write_workbook_appends_marker_line_on_mismatch(tmp_path, monkeypatch, syn_ind_model_and_paths):
    """agent._build_and_write_workbook's returned summary lines must carry
    presentation.CG_RECONCILIATION_ERROR_MARKER when the schedule model it
    just built+wrote has reconciliation_ok False -- this is the line run()
    folds into its returned summary, which main() later greps."""
    tree, model, rules, user_rules, entity, result, loaded = syn_ind_model_and_paths

    def fake_build_all_schedules(*args, **kwargs):
        model.capital_gains.reconciliation_ok = False
        model.capital_gains.reconciliation_diff = 42.0
        return model

    monkeypatch.setattr(sch, "build_all_schedules", fake_build_all_schedules)

    out_path = tmp_path / "out.xlsx"
    lines = agent._build_and_write_workbook(
        tree, pg.parse_book(FIXTURES / "syn_ind.gnucash"), result, None, YEAR_KEY,
        [], [], str(out_path), FIXTURES / "syn_ind.mapping.yaml", entity,
        str(RULES_DIR), ROOT / "Data" / "itr" / "scrips.example.yaml",
    )
    joined = "\n".join(lines)
    assert presentation.CG_RECONCILIATION_ERROR_MARKER in joined
    assert out_path.exists()  # workbook still written -- banner, no abort


def test_main_exits_nonzero_and_prints_stderr_when_marker_present(monkeypatch, capsys):
    monkeypatch.setattr(
        agent, "run",
        lambda *a, **kw: f"STATUS: OK\n\n{presentation.CG_RECONCILIATION_ERROR_MARKER} (diff 42.00) -- DO NOT FILE without review.",
    )
    rc = agent.main(["bs.html", "out.xlsx"])
    assert rc == 1
    captured = capsys.readouterr()
    assert "ERROR" in captured.err
    assert presentation.CG_RECONCILIATION_ERROR_MARKER in captured.err


def test_main_exits_zero_when_marker_absent(monkeypatch, capsys):
    monkeypatch.setattr(agent, "run", lambda *a, **kw: "STATUS: OK\n\nWorkbook: full schedule model built.")
    rc = agent.main(["bs.html", "out.xlsx"])
    assert rc == 0
    captured = capsys.readouterr()
    assert captured.err == ""
