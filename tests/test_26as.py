"""
tests/test_26as.py — Tests for the 26AS extraction script.

Focus: the sub-total / reconciliation behaviour.

The core tests use synthetic deductor data (no PDF, no LLM) so they always run
in CI. They guard the specific regression that was fixed: deductor sub-total
cells must hold COMPUTED NUMERIC VALUES, not bare =SUM formulas — a bare formula
reads back as blank in any data_only consumer (the app preview, pandas, etc.).

An optional end-to-end test runs only when a fixture PDF is dropped at
tests/fixtures/sample_26as.pdf (none is committed, to avoid storing a real PAN).

Run with:
    cd src && python -m pytest ../tests/test_26as.py -v
"""
from __future__ import annotations

import importlib.util
import sys
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "src"
SCRIPT = SRC / "agents" / "skill_26as" / "scripts" / "extract_26as_to_xlsx.py"
FIXTURE = Path(__file__).resolve().parent / "fixtures" / "sample_26as.pdf"


def _load_module():
    """Load the extraction script as a module straight from its file path."""
    if str(SRC) not in sys.path:
        sys.path.insert(0, str(SRC))
    spec = importlib.util.spec_from_file_location("extract_26as_to_xlsx", SCRIPT)
    mod = importlib.util.module_from_spec(spec)
    # Register before exec: dataclass annotation resolution looks the module up
    # in sys.modules by name while the class body executes.
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


m = _load_module()


def _deductor(sr, name, tan, tot_amt, tot_tax, tot_tds, txns):
    """txns: list of (amount, tax, tds) tuples."""
    d = m.P1Deductor(sr=sr, name=name, tan=tan,
                     tot_amt=tot_amt, tot_tax=tot_tax, tot_tds=tot_tds)
    for i, (a, t, td) in enumerate(txns, 1):
        d.txns.append(m.P1Txn(sr=i, section="194A", txn_date="01-Apr-2025",
                              status="F", date_booking="01-May-2025",
                              remarks="-", amount=a, tax=t, tds=td))
    return d


# ---------------------------------------------------------------------------
# Script presence
# ---------------------------------------------------------------------------

def test_script_exists():
    assert SCRIPT.exists(), f"Script not found: {SCRIPT}"


# ---------------------------------------------------------------------------
# Sub-totals must be real numbers (the regression that was fixed)
# ---------------------------------------------------------------------------

def test_subtotals_are_numeric_in_data_only():
    """Sub-total and grand-total cells must read back as numbers when the file
    is opened with data_only=True (i.e. without recalculating formulas)."""
    deductors = [
        _deductor(1, "ALPHA LTD", "AAAA11111A", 1000, 100, 100,
                  [(600, 60, 60), (400, 40, 40)]),
        _deductor(2, "BETA LTD", "BBBB22222B", 500, 50, 50, [(500, 50, 50)]),
    ]
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Part I")
    m.build_part_i(ws, m.Assessee(name="X", pan="ABCDE1234F"), deductors)

    out = Path(tempfile.gettempdir()) / "test_26as_subtotals.xlsx"
    wb.save(out)

    ws2 = load_workbook(out, data_only=True)["Part I"]
    sub_vals, grand_vals = [], None
    for r in range(4, ws2.max_row + 1):
        label = ws2.cell(r, 2).value
        if isinstance(label, str) and label.startswith("Sub-total"):
            sub_vals.append((ws2.cell(r, 13).value,
                             ws2.cell(r, 14).value,
                             ws2.cell(r, 15).value))
        elif isinstance(label, str) and label.startswith("GRAND TOTAL"):
            grand_vals = (ws2.cell(r, 13).value,
                          ws2.cell(r, 14).value,
                          ws2.cell(r, 15).value)

    assert sub_vals == [(1000, 100, 100), (500, 50, 50)], sub_vals
    assert grand_vals == (1500, 150, 150), grand_vals
    # None of them may be blank or a leftover formula string.
    for trip in sub_vals + [grand_vals]:
        for v in trip:
            assert isinstance(v, (int, float)), f"sub-total cell not numeric: {v!r}"


# ---------------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------------

def test_reconcile_clean():
    """When every header total equals its transaction sum, no mismatches."""
    deductors = [
        _deductor(1, "ALPHA LTD", "AAAA11111A", 1000, 100, 100,
                  [(600, 60, 60), (400, 40, 40)]),
        _deductor(2, "BETA LTD", "BBBB22222B", 500, 50, 50, [(500, 50, 50)]),
    ]
    assert m.reconcile_part_i(deductors) == []


def test_reconcile_flags_mismatch():
    """A header total that disagrees with the transaction sum must be reported,
    field by field, with the signed difference."""
    bad = _deductor(1, "ACME LTD", "ABCD12345E", 100000, 10000, 10000,
                    [(45750, 4575, 4575)])
    good = _deductor(2, "GOOD CO", "WXYZ54321A", 45750, 4575, 4575,
                     [(45750, 4575, 4575)])
    mm = m.reconcile_part_i([bad, good])

    assert len(mm) == 1
    entry = mm[0]
    assert entry["sr"] == 1 and entry["name"] == "ACME LTD"
    fields = {f["field"]: f for f in entry["fields"]}
    assert set(fields) == {"Amount Paid/Credited", "Tax Deducted", "TDS Deposited"}
    assert fields["Amount Paid/Credited"]["header_total"] == 100000
    assert fields["Amount Paid/Credited"]["computed_subtotal"] == 45750
    assert fields["Amount Paid/Credited"]["difference"] == -54250


def test_reconcile_zero_txn_deductor_is_flagged():
    """A deductor whose transactions failed to parse (0 rows) must surface as a
    mismatch rather than silently producing a blank/broken sub-total."""
    empty = _deductor(1, "DROPPED ROWS LTD", "AAAA11111A", 9999, 999, 999, [])
    mm = m.reconcile_part_i([empty])
    assert len(mm) == 1
    assert all(f["computed_subtotal"] == 0 for f in mm[0]["fields"])


# ---------------------------------------------------------------------------
# Part VI — TCS
#
# Part VI used to be written as an always-empty sheet, so TCS never reached the
# workbook and the tax credit was silently lost downstream. These tests pin that
# it is now parsed and rendered with the same geometry as Part I (which is what
# lets the journal builder read either sheet by column index).
# ---------------------------------------------------------------------------

PART_VI_TEXT = """
                                                                Sr. No.   Name of Collector   TAN of Collector
     1   THOMAS COOK INDIA                     ABCD12345E          500,000.00      25,000.00      25,000.00
         LIMITED
         1   206CQ     15-Jun-2025   F   30-Jun-2025   -    300,000.00   15,000.00   15,000.00
         2   206CQ     20-Nov-2025   F   30-Nov-2025   -    200,000.00   10,000.00   10,000.00
"""


def test_parse_part_vi_reads_collectors_and_transactions():
    collectors = m.parse_part_vi(PART_VI_TEXT)
    assert len(collectors) == 1
    c = collectors[0]
    # The name wrapped onto a second line and must be re-joined.
    assert c.name == "THOMAS COOK INDIA LIMITED"
    assert c.tan == "ABCD12345E"
    assert (c.tot_amt, c.tot_tax, c.tot_tds) == (500000.0, 25000.0, 25000.0)
    assert [t.section for t in c.txns] == ["206CQ", "206CQ"]
    assert sum(t.tax for t in c.txns) == 25000.0


def test_part_vi_column_geometry_matches_part_i():
    """The journal builder reads Part I and Part VI by the SAME column indices;
    if the two ever diverge it would read the wrong figures silently."""
    assert len(m.P6_HEADERS) == len(m.P1_HEADERS) == 15


def test_build_part_vi_renders_numeric_subtotals():
    collectors = [_deductor(1, "THOMAS COOK INDIA LIMITED", "ABCD12345E",
                            500000, 25000, 25000,
                            [(300000, 15000, 15000), (200000, 10000, 10000)])]
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Part VI")
    m.build_part_vi(ws, m.Assessee(name="X", pan="ABCDE1234F"), collectors)
    out = Path(tempfile.gettempdir()) / "test_26as_part_vi.xlsx"
    wb.save(out)

    ws2 = load_workbook(out, data_only=True)["Part VI"]
    assert ws2.cell(3, 2).value == "Name of Collector"
    assert ws2.cell(3, 14).value == "Tax Collected ++"
    subs = [(ws2.cell(r, 13).value, ws2.cell(r, 14).value, ws2.cell(r, 15).value)
            for r in range(4, ws2.max_row + 1)
            if isinstance(ws2.cell(r, 2).value, str)
            and ws2.cell(r, 2).value.startswith("Sub-total")]
    assert subs == [(500000, 25000, 25000)], subs


def test_build_part_vi_empty_still_renders_headers():
    """No TCS in the year must produce the banner, not a crash or a blank sheet."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Part VI")
    m.build_part_vi(ws, m.Assessee(name="X"), [])
    assert ws.cell(3, 1).value == "Collector Sr.No."
    assert ws.cell(4, 1).value == "No Transactions Present"


# ---------------------------------------------------------------------------
# Part II — Details of Tax Deducted at Source for 15G / 15H
#
# Part II was never parsed at all, so 15G/15H interest never reached the
# workbook (and downstream, the journals). GEOMETRY WARNING pinned here: Part
# II's transaction rows have NO "Status of Booking" column, unlike Part I/VI
# — these tests assert the missing column doesn't shift any later field.
# ---------------------------------------------------------------------------

PART_II_TEXT = """
                                                                Sr. No.   Name of Deductor   TAN of Deductor
     1   BAJAJ FINANCE                        ABCD12345E           50,000.00           0.00           0.00
         LIMITED
         1   194A     10-Apr-2025   05-May-2025   -     30,000.00        0.00        0.00
         2   194A     10-Oct-2025   05-Nov-2025   -     20,000.00        0.00        0.00
     2   SHRIRAM FINANCE LIMITED               WXYZ54321A           75,000.00       7,500.00       7,500.00
         1   194A     15-Jun-2025   20-Jun-2025   -     75,000.00     7,500.00     7,500.00
"""


def test_parse_part_ii_reads_deductors_and_transactions():
    deductors = m.parse_part_ii(PART_II_TEXT)
    assert len(deductors) == 2

    d1 = deductors[0]
    # The name wrapped onto a second line and must be re-joined.
    assert d1.name == "BAJAJ FINANCE LIMITED"
    assert d1.tan == "ABCD12345E"
    assert (d1.tot_amt, d1.tot_tax, d1.tot_tds) == (50000.0, 0.0, 0.0)
    assert len(d1.txns) == 2
    # Missing "Status of Booking" column must not shift Date of Booking,
    # Remarks, Amount, Tax or TDS out of place.
    t1 = d1.txns[0]
    assert t1.section == "194A"
    assert t1.txn_date == "10-Apr-2025"
    assert t1.date_booking == "05-May-2025"
    assert t1.remarks == "-"
    assert (t1.amount, t1.tax, t1.tds) == (30000.0, 0.0, 0.0)

    d2 = deductors[1]
    assert d2.name == "SHRIRAM FINANCE LIMITED"
    assert (d2.tot_amt, d2.tot_tax, d2.tot_tds) == (75000.0, 7500.0, 7500.0)
    assert len(d2.txns) == 1
    assert (d2.txns[0].amount, d2.txns[0].tax, d2.txns[0].tds) == (75000.0, 7500.0, 7500.0)


def test_part_ii_column_geometry_shares_the_shared_prefix_with_part_i():
    """Part II has no 'Status of Booking' column, so it is 14 columns, not
    Part I/VI's 15 — but columns (1-indexed) 1/2/4/5/6/8 (Sr, Name, the three
    header totals, Section) hold the SAME field at the SAME position across
    all three, which is what lets the journal builder's _parse_party_sheet
    read any of the three sheets unchanged."""
    p2 = m.EMPTY_HEADERS["Part II"]
    assert len(p2) == 14
    assert len(m.P1_HEADERS) == len(m.P6_HEADERS) == 15

    def field_at(headers, one_indexed_col):
        return headers[one_indexed_col - 1]

    # Wording legitimately differs across parts (Paid/Credited vs
    # Paid/Debited, Deducted vs Collected, Deductor vs Collector...) — the
    # invariant that matters is column INDEX, since that's all
    # _parse_party_sheet ever reads by.
    for headers in (p2, m.P1_HEADERS, m.P6_HEADERS):
        assert "Name of" in field_at(headers, 2)
        assert field_at(headers, 4).startswith("Total Amount")
        assert field_at(headers, 5).startswith("Total Tax")
        assert "Deposited" in field_at(headers, 6)
        assert field_at(headers, 8) == "Section"


def test_build_part_ii_renders_numeric_subtotals():
    d = m.P2Deductor(sr=1, name="BAJAJ FINANCE LIMITED", tan="ABCD12345E",
                     tot_amt=50000, tot_tax=0, tot_tds=0)
    d.txns = [
        m.P2Txn(sr=1, section="194A", txn_date="10-Apr-2025",
               date_booking="05-May-2025", remarks="-",
               amount=30000, tax=0, tds=0),
        m.P2Txn(sr=2, section="194A", txn_date="10-Oct-2025",
               date_booking="05-Nov-2025", remarks="-",
               amount=20000, tax=0, tds=0),
    ]
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Part II")
    m.build_part_ii(ws, m.Assessee(name="X", pan="ABCDE1234F"), [d])
    out = Path(tempfile.gettempdir()) / "test_26as_part_ii.xlsx"
    wb.save(out)

    ws2 = load_workbook(out, data_only=True)["Part II"]
    assert ws2.cell(3, 2).value == "Name of Deductor"
    subs = [(ws2.cell(r, 12).value, ws2.cell(r, 13).value, ws2.cell(r, 14).value)
            for r in range(4, ws2.max_row + 1)
            if isinstance(ws2.cell(r, 2).value, str)
            and ws2.cell(r, 2).value.startswith("Sub-total")]
    assert subs == [(50000, 0, 0)], subs
    for trip in subs:
        for v in trip:
            assert isinstance(v, (int, float)), f"sub-total cell not numeric: {v!r}"


def test_build_part_ii_empty_still_renders_headers():
    """No 15G/15H deductors in the year must produce the banner, not a crash
    or a spurious blank sheet."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Part II")
    m.build_part_ii(ws, m.Assessee(name="X"), [])
    assert ws.cell(3, 1).value == "Sr.No."
    assert ws.cell(4, 1).value == "No Transactions Present"


def test_parse_part_ii_absent_from_pdf_returns_empty_list():
    """A 26AS with no Part II text at all must not crash — just no deductors."""
    assert m.parse_part_ii("") == []


# ---------------------------------------------------------------------------
# Tier 4 — hard fail on impossibly-empty extraction
#
# check_extraction_not_vacuous() is the backstop for every other way
# extraction can silently yield nothing (a future binary swap, a changed PDF
# layout, a regex that stops matching). It must raise when ALL FOUR parts
# this script actually parses (I, II, VI, VIII) are simultaneously zero, but
# must NOT raise when any one of them alone is zero -- Part VI (TCS) being
# zero is the normal case for most taxpayers.
# ---------------------------------------------------------------------------

def test_check_extraction_not_vacuous_all_empty_raises():
    """All four parsed-part counts zero -- this is the impossible case (a
    real 26AS always has at least one populated part) -- must raise with
    diagnostics naming the resolved binary and the character count."""
    stats = {
        "part_i_deductors": 0,
        "part_ii_deductors": 0,
        "part_vi_collectors": 0,
        "part_viii_rows": 0,
    }
    with pytest.raises(m.Empty26ASExtractionError) as exc_info:
        m.check_extraction_not_vacuous(stats, "some unrelated text", "/fake/path/pdftotext")
    msg = str(exc_info.value)
    assert "/fake/path/pdftotext" in msg
    assert "Characters of text extracted" in msg
    assert "Part I (deductors) = 0" in msg


def test_check_extraction_not_vacuous_all_empty_no_text_raises():
    """Same all-zero case but with genuinely empty extracted text -- the
    diagnostic should say the binary produced no text at all, not conflate
    it with 'text but no regex match'."""
    stats = {
        "part_i_deductors": 0,
        "part_ii_deductors": 0,
        "part_vi_collectors": 0,
        "part_viii_rows": 0,
    }
    with pytest.raises(m.Empty26ASExtractionError) as exc_info:
        m.check_extraction_not_vacuous(stats, "", "/fake/path/pdftotext")
    assert "NO text at all" in str(exc_info.value)


def test_check_extraction_not_vacuous_one_part_populated_does_not_raise():
    """Only Part I populated, the rest zero -- a perfectly normal document
    (most taxpayers have no 15G/15H, TCS, or Part VIII rows) -- must NOT
    raise."""
    stats = {
        "part_i_deductors": 5,
        "part_ii_deductors": 0,
        "part_vi_collectors": 0,
        "part_viii_rows": 0,
    }
    m.check_extraction_not_vacuous(stats, "some text", "/fake/path/pdftotext")  # no raise


def test_check_extraction_not_vacuous_part_vi_legitimately_zero_does_not_raise():
    """Part VI (TCS) = 0 is the common, correct case for most taxpayers --
    pinned explicitly per the boundary this check must get right."""
    stats = {
        "part_i_deductors": 3,
        "part_ii_deductors": 1,
        "part_vi_collectors": 0,
        "part_viii_rows": 0,
    }
    m.check_extraction_not_vacuous(stats, "some text", "/fake/path/pdftotext")  # no raise


def test_run_raises_before_writing_output_when_all_parts_empty(monkeypatch):
    """run() must call the vacuous-extraction gate BEFORE building/writing the
    workbook, so a broken extraction never produces a misleading .xlsx file,
    and the caller never gets a chance to print a vacuous '0/0 OK' summary."""
    monkeypatch.setattr(m, "pdf_to_text", lambda pdf_path: ("no matching rows here", "/fake/pdftotext"))
    out = Path(tempfile.gettempdir()) / "test_26as_vacuous_run.xlsx"
    if out.exists():
        out.unlink()
    with pytest.raises(m.Empty26ASExtractionError):
        m.run(Path("unused.pdf"), out)
    assert not out.exists(), "workbook must not be written when extraction is vacuous"


def test_vacuous_0_0_ok_message_can_no_longer_be_emitted(monkeypatch, capsys):
    """End-to-end through main(): an all-empty extraction must print a FATAL
    diagnostic to stderr and return a non-zero exit code -- never the old
    reassuring '0/0 deductors OK' reconciliation success message."""
    monkeypatch.setattr(m, "pdf_to_text", lambda pdf_path: ("", "/fake/pdftotext"))
    out = Path(tempfile.gettempdir()) / "test_26as_vacuous_main.xlsx"
    if out.exists():
        out.unlink()
    rc = m.main(["extract_26as_to_xlsx.py", "unused.pdf", str(out)])
    captured = capsys.readouterr()
    assert rc != 0
    assert "0/0" not in captured.out
    assert "OK" not in captured.out
    assert "FATAL" in captured.err
    assert not out.exists()


def test_wrong_pdftotext_flavour_prints_clean_fatal_and_exit_4(monkeypatch, capsys):
    """End-to-end through main(): a WrongPdftextFlavourError raised from
    pdf_to_text()'s binary resolution (e.g. Xpdf found instead of Poppler)
    must print a clean 'FATAL: <message>' line to stderr -- NOT a raw Python
    traceback -- and return the distinct exit code 4 (3 is reserved for
    Empty26ASExtractionError). pdf_to_text() wraps the underlying
    WrongPdftextFlavourError in a plain RuntimeError (see its docstring), so
    that is the shape simulated here."""
    message = "Wrong pdftotext found at '/fake/vendor/poppler/bin/pdftotext': this is Xpdf (Glyph & Cog), not Poppler."

    def fake_pdf_to_text(pdf_path):
        raise RuntimeError(f"pdftotext resolution failed: {message}")

    monkeypatch.setattr(m, "pdf_to_text", fake_pdf_to_text)
    out = Path(tempfile.gettempdir()) / "test_26as_wrong_flavour_main.xlsx"
    if out.exists():
        out.unlink()
    rc = m.main(["extract_26as_to_xlsx.py", "unused.pdf", str(out)])
    captured = capsys.readouterr()
    assert rc == 4
    assert captured.err.startswith("FATAL:")
    assert message in captured.err
    # Must not be a raw traceback.
    assert "Traceback" not in captured.err
    assert not out.exists()


def test_empty_extraction_error_still_wins_exit_code_3_not_4(monkeypatch, capsys):
    """Empty26ASExtractionError is itself a RuntimeError subclass -- the
    generic RuntimeError except-clause added for WrongPdftextFlavourError
    must not shadow it. This pins the ordering: exit code 3 (empty
    extraction) must still be returned, never 4."""
    monkeypatch.setattr(m, "pdf_to_text", lambda pdf_path: ("", "/fake/pdftotext"))
    out = Path(tempfile.gettempdir()) / "test_26as_empty_still_3.xlsx"
    if out.exists():
        out.unlink()
    rc = m.main(["extract_26as_to_xlsx.py", "unused.pdf", str(out)])
    assert rc == 3


# ---------------------------------------------------------------------------
# Optional end-to-end (only if a fixture PDF is provided)
# ---------------------------------------------------------------------------

@pytest.mark.skipif(not FIXTURE.exists(),
                    reason="No fixture PDF at tests/fixtures/sample_26as.pdf")
def test_extract_end_to_end():
    out = Path(tempfile.gettempdir()) / "test_26as_e2e.xlsx"
    stats = m.run(FIXTURE, out)
    assert out.exists()
    assert stats["part_i_deductors"] >= 1
    assert "reconciliation" in stats
    # Sub-totals must be numeric in a data_only read.
    ws = load_workbook(out, data_only=True)["Part I"]
    for r in range(4, ws.max_row + 1):
        label = ws.cell(r, 2).value
        if isinstance(label, str) and label.startswith(("Sub-total", "GRAND TOTAL")):
            for c in (13, 14, 15):
                v = ws.cell(r, c).value
                assert v is None or isinstance(v, (int, float))
