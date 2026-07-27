"""
tests/test_ais_reconcile.py — Tests for the AIS decrypt + normalize
foundation layer (src/agents/skill_ais_reconcile/).

Uses a SYNTHETIC AIS JSON dict (fake PAN, fake dates, fake amounts) that we
encrypt ourselves with a throwaway password using the same scheme
decrypt.decrypt_ais expects, then round-trip it through decrypt -> normalize.
No real taxpayer data — nothing under Data/AIS/ is read or referenced.

Two synthetic fixture families are covered:
  - synthetic_ais(): hand-built DICT-keyed columnData rows (legacy/simple
    shape some fixtures use).
  - synthetic_ais_positional(): mirrors the REAL AIS export shape, where
    columnData rows are POSITIONAL LISTS of scalars, keyed by `seq` in a
    dict-descriptor columnLabel (l1-style) or by plain-string columnLabel
    headers aligned 1:1 by index (l2-style), with columnDataType as the
    authoritative parallel type list.

Run with:
    cd src && python -m pytest ../tests/test_ais_reconcile.py -v
"""
from __future__ import annotations

import copy
import json
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))
# configs.py (skill_itr_workbook) does a bare `import tags` assuming its own
# scripts/ dir is on sys.path directly (the convention used throughout
# tests/skill_itr_workbook/) -- add it so the package-path import below of
# configs.py (for Phase C's MappingEntry/MappingLoadResult) succeeds too.
_ITR_SCRIPTS = SRC / "agents" / "skill_itr_workbook" / "scripts"
if str(_ITR_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_ITR_SCRIPTS))

from agents.skill_ais_reconcile import decrypt as D  # noqa: E402
from agents.skill_ais_reconcile import normalize as N  # noqa: E402
from agents.skill_ais_reconcile import reconcile as R  # noqa: E402
from agents.skill_ais_reconcile import excel_writer as X  # noqa: E402
from agents.skill_ais_reconcile import reconcile_26as as R26  # noqa: E402
from agents.skill_itr_workbook.scripts.as26 import As26Data, As26Transaction  # noqa: E402

PAN = "ABCDE1234F"
DOB_ISO = "1985-06-15"
DOB_DDMMYYYY = "15061985"


# ---------------------------------------------------------------------------
# Synthetic fixture builder
# ---------------------------------------------------------------------------

def _column_label(field_, name, type_, sum_required="Y", display_feedback=False):
    return {
        "field": field_, "name": name, "type": type_, "seq": 1,
        "sumRequired": sum_required, "isEditable": True, "isMandatory": False,
        "displayFeedback": display_feedback, "infoCode": None,
        "validationRegex": None,
    }


def _tds_element():
    column_label = [
        _column_label("deductorName", "Name of Deductor", "string"),
        _column_label("section", "Section", "string"),
        _column_label("amtPaid", "Amount Paid", "decimal", display_feedback=True),
        _column_label("taxDeducted", "Tax Deducted", "decimal"),
        _column_label("dateOfPayment", "Date of Payment", "date"),
    ]
    l1 = {
        "columnLabel": column_label,
        "columnData": [
            {
                "deductorName": "SYNTHETIC BANK LIMITED",
                "section": "194A",
                "amtPaid": "250237.00",
                "taxDeducted": "25024.00",
                "dateOfPayment": "2026-01-15",
                "status": "Active",
                "transFeedback": None,
            },
            {
                # blank/None decimal must coerce to None, not crash
                "deductorName": "SYNTHETIC NBFC LTD",
                "section": "193",
                "amtPaid": "",
                "taxDeducted": None,
                "dateOfPayment": "2026-02-01",
                "status": "Active",
                "transFeedback": "Feedback submitted",
            },
        ],
        "columnDataType": {},
    }
    return {
        "title": "Interest from deposit",
        "l1Src": "TDS",
        "infoSrcId": "SFT-001",
        "l1": l1,
        "l2": None,
    }


def _sft_element():
    column_label = [
        _column_label("sellerName", "Name of Seller", "string"),
        _column_label("salesConsideration", "Sales Consideration", "decimal",
                       display_feedback=True),
    ]
    l1 = {
        "columnLabel": column_label,
        "columnData": [
            {
                "sellerName": "SYNTHETIC BROKING PVT LTD",
                "salesConsideration": "1000000.00",
                "status": "Active",
                "transFeedback": None,
            },
        ],
        "columnDataType": {},
    }
    return {
        "title": "Sale of securities",
        "l1Src": "SFT",
        "infoSrcId": "SFT-002",
        "l1": l1,
        "l2": l1,   # exercise the l2 branch too (same shape as l1)
    }


def _section_with_string_column_labels():
    """A header-like section where columnLabel entries are plain strings, not
    descriptor dicts — the normalizer must guard against this, not crash."""
    return {
        "title": "Other info",
        "l1Src": "OTHER",
        "infoSrcId": None,
        "l1": {
            "columnLabel": ["Some Header", "Another Header"],
            "columnData": [{"Some Header": "x", "Another Header": "y"}],
            "columnDataType": {},
        },
        "l2": None,
    }


def _column_label_seq(field_, name, type_, seq, display_feedback=False):
    """Dict-descriptor columnLabel entry for the POSITIONAL row shape: `seq`
    is the column index into the row, and (per real AIS files) only a subset
    of a row's columns are described this way."""
    return {
        "field": field_, "name": name, "type": type_, "seq": seq,
        "sumRequired": "Y", "isEditable": True, "isMandatory": False,
        "displayFeedback": display_feedback, "infoCode": None,
        "validationRegex": None,
    }


def _positional_tds_element():
    """Mirrors the REAL l1 shape: columnData rows are 15-element positional
    lists of scalars; columnLabel is a list of dict descriptors (only ~8 of
    the 15 columns described, keyed by `seq`); columnDataType is a parallel
    15-entry list of type strings that is authoritative for coercion. Also
    exercises status/transFeedback arriving as descriptors (popped into
    AisRow.status/trans_feedback, not left in `fields`)."""
    column_label = [
        _column_label_seq("deductorName", "Name of Deductor", "string", 0),
        _column_label_seq("section", "Section", "string", 1),
        _column_label_seq("amtPaid", "Amount Paid", "decimal", 3, display_feedback=True),
        _column_label_seq("taxDeducted", "Tax Deducted", "decimal", 5),
        _column_label_seq("dateOfPayment", "Date of Payment", "date", 7),
        _column_label_seq("status", "Status", "string", 9),
        _column_label_seq("transFeedback", "Transaction Feedback", "string", 10),
        _column_label_seq("infoCode", "Information Code", "string", 12),
    ]
    column_data_type = [
        "string", "string", "string", "decimal", "string", "decimal", "string",
        "date", "string", "string", "string", "string", "string", "string", "string",
    ]
    row_full = [
        "SYNTHETIC BANK LIMITED", "194A", None, "250237.00", None, "25024.00",
        None, "2026-01-15", None, "Active", "Feedback submitted", None,
        "INFO123", None, None,
    ]
    # Blank/None amtPaid + taxDeducted must coerce to None, not crash.
    row_blank_amounts = [
        "SYNTHETIC NBFC LTD", "193", None, "", None, None,
        None, "2026-02-01", None, "Active", None, None,
        "INFO456", None, None,
    ]
    l1 = {
        "columnLabel": column_label,
        "columnData": [row_full, row_blank_amounts],
        "columnDataType": column_data_type,
    }
    return {
        "title": "Interest from deposit (positional)",
        "l1Src": "TDS",
        "infoSrcId": "SFT-003",
        "l1": l1,
        "l2": None,
    }


def _positional_l2_element():
    """Mirrors the REAL l2 shape: columnData rows are 9-element positional
    lists; columnLabel is a list of PLAIN STRING headers aligned 1:1 by
    position (no dict descriptors at all). The reconciliation money fields
    (Amount / Derived Amount) live here."""
    column_label = [
        "Information Category", "Information Code", "Information Description",
        "Information Source", "Count", "Amount", "Information Category Code",
        "Derived Amount", "Qualifies For",
    ]
    column_data_type = [
        "string", "string", "string", "string", "numeric", "decimal",
        "string", "decimal", "string",
    ]
    row = [
        "Interest", "SFT-017", "Interest income reported by deductor", "AIS",
        "1", "50000.00", "SFT017", "48000.00", "Yes",
    ]
    l2 = {
        "columnLabel": column_label,
        "columnData": [row],
        "columnDataType": column_data_type,
    }
    return {
        "title": "Interest from deposit (positional l2)",
        "l1Src": "TDS",
        "infoSrcId": "SFT-004",
        "l1": None,
        "l2": l2,
    }


def synthetic_ais_positional() -> dict:
    """A minimal AIS dict exercising ONLY the real positional columnData
    shapes (see _positional_tds_element / _positional_l2_element) — kept
    separate from synthetic_ais() so the dict-row fixture's row-count
    assertions don't have to change."""
    return {
        "metadata": {
            "loggedInPan": PAN,
            "jsonVersion": "1.0",
            "downloadDate": "27-Jul-2026",
            "utilityVersion": "1.0.0",
            "sourceSharedFeedbackFeatureEnabled": True,
        },
        "header": {},
        "partA": {"columnLabel": ["PAN"], "columnData": [PAN]},
        "partB": {
            "sections": [
                {
                    "sectionKey": "tdsTcs",
                    "title": "TDS/TCS Information",
                    "heading": "B1",
                    "elements": [_positional_tds_element(), _positional_l2_element()],
                },
            ],
        },
        "rejectedFeedbacks": [],
        "footer": {},
        "fileSize": 999,
    }


def synthetic_ais() -> dict:
    return {
        "metadata": {
            "loggedInPan": PAN,
            "jsonVersion": "1.0",
            "downloadDate": "27-Jul-2026",
            "utilityVersion": "1.0.0",
            "sourceSharedFeedbackFeatureEnabled": True,
        },
        "header": {},
        "partA": {
            "columnLabel": ["PAN", "Aadhaar", "Name", "DOB", "Mobile", "Email", "Address"],
            "columnData": [PAN, "XXXX-XXXX-1234", "SYNTHETIC TAXPAYER", "15-Jun-1985",
                           "+91-9999999999", "synthetic@example.test", "123 Synthetic Street"],
        },
        "partB": {
            "sections": [
                {
                    "sectionKey": "tdsTcs",
                    "title": "TDS/TCS Information",
                    "heading": "B1",
                    "elements": [_tds_element()],
                },
                {
                    "sectionKey": "sft",
                    "title": "SFT Information",
                    "heading": "B2",
                    "elements": [_sft_element()],
                },
                {
                    # sectionKey with no `elements` key at all -- must be skipped, not crash
                    "sectionKey": "demandAndRefund",
                    "title": "Demand and Refund",
                    "heading": "B4",
                },
                {
                    # `elements` explicitly None -- same guard
                    "sectionKey": "paymentOfTaxes",
                    "title": "Payment of Taxes",
                    "heading": "B3",
                    "elements": None,
                },
                {
                    "sectionKey": "other-info",
                    "title": "Other Information",
                    "heading": "B7",
                    "elements": [_section_with_string_column_labels()],
                },
            ],
        },
        "rejectedFeedbacks": [],
        "footer": {},
        "fileSize": 12345,
    }


# ---------------------------------------------------------------------------
# decrypt.py
# ---------------------------------------------------------------------------

def test_derive_password_default_pepper():
    pw = D.derive_password(PAN, DOB_DDMMYYYY)
    assert pw == f"{PAN.lower()}GQ39%*g{DOB_DDMMYYYY}"


def test_derive_password_empty_pepper_variant():
    pw = D.derive_password(PAN, DOB_DDMMYYYY, pepper="")
    assert pw == f"{PAN.lower()}{DOB_DDMMYYYY}"


def test_derive_password_from_iso_date_matches_ddmmyyyy():
    assert D.derive_password_from_iso_date(PAN, DOB_ISO) == D.derive_password(PAN, DOB_DDMMYYYY)


def test_encrypt_decrypt_round_trip():
    ais = synthetic_ais()
    password = D.derive_password(PAN, DOB_DDMMYYYY)
    blob = D._encrypt_for_test(ais, password)
    out = D.decrypt_ais(blob, password)
    assert out == ais


def test_decrypt_wrong_password_raises_ais_decrypt_error():
    ais = synthetic_ais()
    password = D.derive_password(PAN, DOB_DDMMYYYY)
    blob = D._encrypt_for_test(ais, password)
    with pytest.raises(D.AisDecryptError):
        D.decrypt_ais(blob, "totally-wrong-password")


def test_decrypt_garbage_header_raises_ais_decrypt_error():
    with pytest.raises(D.AisDecryptError):
        D.decrypt_ais("not a valid ais blob at all", "whatever")


def test_decrypt_too_short_raises_ais_decrypt_error():
    with pytest.raises(D.AisDecryptError):
        D.decrypt_ais("abcd1234", "whatever")


def test_encrypt_decrypt_deterministic_with_fixed_iv_salt():
    ais = {"a": 1}
    password = "pw"
    iv = b"\x00" * 16
    salt = b"\x01" * 16
    blob1 = D._encrypt_for_test(ais, password, iv=iv, salt=salt)
    blob2 = D._encrypt_for_test(ais, password, iv=iv, salt=salt)
    assert blob1 == blob2
    assert D.decrypt_ais(blob1, password) == ais


# ---------------------------------------------------------------------------
# normalize.py
# ---------------------------------------------------------------------------

def test_normalize_metadata():
    n = N.normalize(synthetic_ais())
    assert n.json_version == "1.0"
    assert n.download_date == "27-Jul-2026"
    assert n.logged_in_pan == PAN


def test_normalize_part_a_zips_labels_and_data():
    n = N.normalize(synthetic_ais())
    assert n.part_a["PAN"] == PAN
    assert n.part_a["Name"] == "SYNTHETIC TAXPAYER"
    assert n.part_a["DOB"] == "15-Jun-1985"


def test_normalize_flattens_rows_across_sections():
    n = N.normalize(synthetic_ais())
    # 2 rows in tdsTcs l1, 1 row in sft l1 + 1 row in sft l2 (same block used
    # twice), 1 row in other-info l1 -> 5 total.
    assert len(n.rows) == 5


def test_normalize_decimal_coercion():
    n = N.normalize(synthetic_ais())
    tds_rows = [r for r in n.rows if r.section_key == "tdsTcs"]
    row0 = next(r for r in tds_rows if r.fields["deductorName"] == "SYNTHETIC BANK LIMITED")
    from decimal import Decimal
    assert row0.fields["amtPaid"] == Decimal("250237.00")
    assert row0.fields["taxDeducted"] == Decimal("25024.00")
    assert row0.category == "Interest from deposit"
    assert row0.section_title == "TDS/TCS Information"
    assert row0.info_src_id == "SFT-001"
    assert row0.level == "l1"


def test_normalize_blank_and_none_decimal_become_none_not_crash():
    n = N.normalize(synthetic_ais())
    row1 = next(r for r in n.rows if r.fields.get("deductorName") == "SYNTHETIC NBFC LTD")
    assert row1.fields["amtPaid"] is None
    assert row1.fields["taxDeducted"] is None


def test_normalize_feedback_fields_and_editable_flag():
    n = N.normalize(synthetic_ais())
    row0 = next(r for r in n.rows if r.fields.get("deductorName") == "SYNTHETIC BANK LIMITED")
    assert row0.status == "Active"
    assert row0.feedback_editable_field == "amtPaid"

    row1 = next(r for r in n.rows if r.fields.get("deductorName") == "SYNTHETIC NBFC LTD")
    assert row1.trans_feedback == "Feedback submitted"


def test_normalize_sft_feedback_editable_field_is_sales_consideration():
    n = N.normalize(synthetic_ais())
    sft_rows = [r for r in n.rows if r.section_key == "sft"]
    assert sft_rows
    for r in sft_rows:
        assert r.feedback_editable_field == "salesConsideration"


def test_normalize_l2_processed_same_shape_as_l1():
    n = N.normalize(synthetic_ais())
    levels = {r.level for r in n.rows if r.section_key == "sft"}
    assert levels == {"l1", "l2"}


def test_normalize_section_missing_elements_key_is_skipped_not_crashed():
    n = N.normalize(synthetic_ais())
    assert not any(r.section_key == "demandAndRefund" for r in n.rows)


def test_normalize_section_elements_none_is_skipped_not_crashed():
    n = N.normalize(synthetic_ais())
    assert not any(r.section_key == "paymentOfTaxes" for r in n.rows)


def test_normalize_string_column_labels_guarded_no_crash():
    """Some columnLabel entries are plain strings (header-like sections);
    those rows still show up but with no type-driven coercion applied."""
    n = N.normalize(synthetic_ais())
    other_rows = [r for r in n.rows if r.section_key == "other-info"]
    assert len(other_rows) == 1
    assert other_rows[0].fields == {"Some Header": "x", "Another Header": "y"}
    assert other_rows[0].feedback_editable_field is None


def test_normalize_summary_counts_per_section_category():
    n = N.normalize(synthetic_ais())
    summary = {(s.section_key, s.category): s.row_count for s in n.summary}
    assert summary[("tdsTcs", "Interest from deposit")] == 2
    assert summary[("sft", "Sale of securities")] == 2  # l1 + l2
    assert summary[("other-info", "Other info")] == 1  # element title, not section title


# ---------------------------------------------------------------------------
# REAL positional columnData shape (rows are lists, not dicts) — regression
# coverage for the bug where _normalize_level dropped every row because it
# only accepted dict-shaped rows. See _positional_tds_element /
# _positional_l2_element for the fixture shapes.
# ---------------------------------------------------------------------------

def test_normalize_positional_l1_rows_present_not_dropped():
    n = N.normalize(synthetic_ais_positional())
    assert len(n.rows) > 0
    l1_rows = [r for r in n.rows if r.level == "l1"]
    assert len(l1_rows) == 2


def test_normalize_positional_l1_seq_mapped_fields_and_decimal_coercion():
    n = N.normalize(synthetic_ais_positional())
    row0 = next(r for r in n.rows
                if r.level == "l1" and r.fields.get("deductorName") == "SYNTHETIC BANK LIMITED")
    from decimal import Decimal
    assert row0.fields["section"] == "194A"
    assert row0.fields["amtPaid"] == Decimal("250237.00")
    assert row0.fields["taxDeducted"] == Decimal("25024.00")
    assert row0.fields["dateOfPayment"] == "2026-01-15"
    assert row0.fields["infoCode"] == "INFO123"
    # Unmapped positional indices (no seq descriptor) must not leak into fields.
    assert len(row0.fields) == 6  # deductorName, section, amtPaid, taxDeducted, dateOfPayment, infoCode


def test_normalize_positional_l1_status_and_trans_feedback_land_on_row_not_fields():
    n = N.normalize(synthetic_ais_positional())
    row0 = next(r for r in n.rows
                if r.level == "l1" and r.fields.get("deductorName") == "SYNTHETIC BANK LIMITED")
    assert row0.status == "Active"
    assert row0.trans_feedback == "Feedback submitted"
    assert "status" not in row0.fields
    assert "transFeedback" not in row0.fields


def test_normalize_positional_l1_editable_field_from_seq_descriptor():
    n = N.normalize(synthetic_ais_positional())
    l1_rows = [r for r in n.rows if r.level == "l1"]
    for r in l1_rows:
        assert r.feedback_editable_field == "amtPaid"


def test_normalize_positional_l1_blank_decimal_becomes_none_not_crash():
    n = N.normalize(synthetic_ais_positional())
    row1 = next(r for r in n.rows
                if r.level == "l1" and r.fields.get("deductorName") == "SYNTHETIC NBFC LTD")
    assert row1.fields["amtPaid"] is None
    assert row1.fields["taxDeducted"] is None
    assert row1.status == "Active"
    assert row1.trans_feedback is None


def test_normalize_positional_l2_plain_string_headers_present_not_dropped():
    n = N.normalize(synthetic_ais_positional())
    l2_rows = [r for r in n.rows if r.level == "l2"]
    assert len(l2_rows) == 1


def test_normalize_positional_l2_amount_and_derived_amount_decimal_coercion():
    n = N.normalize(synthetic_ais_positional())
    row = next(r for r in n.rows if r.level == "l2")
    from decimal import Decimal
    assert row.fields["Amount"] == Decimal("50000.00")
    assert row.fields["Derived Amount"] == Decimal("48000.00")
    assert row.fields["Information Category"] == "Interest"
    assert row.fields["Information Code"] == "SFT-017"
    assert row.fields["Qualifies For"] == "Yes"


def test_normalize_positional_summary_counts_nonzero():
    n = N.normalize(synthetic_ais_positional())
    summary = {(s.section_key, s.category): s.row_count for s in n.summary}
    assert summary[("tdsTcs", "Interest from deposit (positional)")] == 2
    assert summary[("tdsTcs", "Interest from deposit (positional l2)")] == 1


def test_normalize_legacy_dict_row_path_still_covered():
    """Existing hand-built dict-row fixtures (synthetic_ais()) must keep
    working unchanged now that positional rows are also supported."""
    n = N.normalize(synthetic_ais())
    assert len(n.rows) == 5
    assert all(isinstance(r.fields, dict) for r in n.rows)


def test_normalize_open_ended_unknown_section_key_still_flattens():
    """A future/unknown sectionKey (e.g. GST/business receipts) must still
    normalize via the schema-driven path, not require a code change."""
    ais = synthetic_ais()
    ais["partB"]["sections"].append({
        "sectionKey": "futureBusinessReceipts",
        "title": "Future Business Receipts",
        "heading": "B99",
        "elements": [{
            "title": "GST turnover",
            "l1Src": "GST",
            "infoSrcId": "GST-001",
            "l1": {
                "columnLabel": [_column_label("turnover", "Turnover", "decimal")],
                "columnData": [{"turnover": "500000.00", "status": "Active",
                                "transFeedback": None}],
                "columnDataType": {},
            },
            "l2": None,
        }],
    })
    n = N.normalize(ais)
    from decimal import Decimal
    row = next(r for r in n.rows if r.section_key == "futureBusinessReceipts")
    assert row.fields["turnover"] == Decimal("500000.00")


def test_normalize_missing_sections_list_returns_empty_rows():
    ais = synthetic_ais()
    ais["partB"] = {}
    n = N.normalize(ais)
    assert n.rows == []
    assert n.summary == []


def test_normalize_does_not_mutate_input():
    ais = synthetic_ais()
    snapshot = copy.deepcopy(ais)
    N.normalize(ais)
    assert ais == snapshot


# ---------------------------------------------------------------------------
# Full round trip: encrypt synthetic -> decrypt -> normalize
# ---------------------------------------------------------------------------

def test_full_round_trip_decrypt_then_normalize():
    ais = synthetic_ais()
    password = D.derive_password(PAN, DOB_DDMMYYYY)
    blob = D._encrypt_for_test(ais, password)

    decrypted = D.decrypt_ais(blob, password)
    n = N.normalize(decrypted)

    assert n.logged_in_pan == PAN
    assert len(n.rows) == 5
    assert json.loads(json.dumps(decrypted)) == ais  # sanity: still plain JSON-able


# ---------------------------------------------------------------------------
# reconcile.py — AIS-internal reconciliation (Phase A)
#
# Unit tests build a NormalizedAis directly (bypassing decrypt/normalize) so
# the reconciliation math is exercised precisely and independent of the
# normalizer's own row-shape handling (already covered above).
# ---------------------------------------------------------------------------

from decimal import Decimal  # noqa: E402


def _row(section_key, category, level, fields, *, info_src_id="SRC-1",
         status=None, trans_feedback=None, feedback_editable_field=None,
         section_title=""):
    return N.AisRow(
        section_key=section_key, section_title=section_title, category=category,
        info_src_id=info_src_id, level=level, fields=fields, status=status,
        trans_feedback=trans_feedback, feedback_editable_field=feedback_editable_field,
    )


def _normalized(rows, *, pan=PAN, part_a=None):
    return N.NormalizedAis(
        json_version="1.0", download_date="27-Jul-2026", logged_in_pan=pan,
        part_a=part_a or {"Name": "SYNTHETIC TAXPAYER"}, rows=rows, summary=[],
    )


def test_reconcile_matching_element_no_flag():
    rows = [
        _row("tdsTcs", "Interest from deposit", "l1",
             {"deductorName": "BANK A", "amtPaid": Decimal("600")},
             feedback_editable_field="amtPaid", status="Active"),
        _row("tdsTcs", "Interest from deposit", "l1",
             {"deductorName": "BANK B", "amtPaid": Decimal("400")},
             feedback_editable_field="amtPaid", status="Active"),
        _row("tdsTcs", "Interest from deposit", "l2",
             {"Amount": Decimal("1000")}),
    ]
    report = R.reconcile_internal(_normalized(rows))
    assert len(report.elements) == 1
    el = report.elements[0]
    assert el.detail_sum == Decimal("1000")
    assert el.reported == Decimal("1000")
    assert el.delta_detail_vs_reported == Decimal("0")
    assert not el.flagged
    assert not el.flag_detail_mismatch
    assert report.flagged_element_count == 0


def test_reconcile_deliberate_mismatch_flags():
    rows = [
        _row("tdsTcs", "Interest from deposit", "l1",
             {"deductorName": "BANK A", "amtPaid": Decimal("600")},
             feedback_editable_field="amtPaid"),
        _row("tdsTcs", "Interest from deposit", "l1",
             {"deductorName": "BANK B", "amtPaid": Decimal("400")},
             feedback_editable_field="amtPaid"),
        _row("tdsTcs", "Interest from deposit", "l2",
             {"Amount": Decimal("900")}),   # 100 short of the detail sum
    ]
    report = R.reconcile_internal(_normalized(rows))
    el = report.elements[0]
    assert el.delta_detail_vs_reported == Decimal("100")
    assert el.flag_detail_mismatch is True
    assert el.flagged is True
    assert report.flagged_element_count == 1


def test_reconcile_delta_within_tolerance_not_flagged():
    rows = [
        _row("tdsTcs", "Interest from deposit", "l1",
             {"amtPaid": Decimal("1000.40")}, feedback_editable_field="amtPaid"),
        _row("tdsTcs", "Interest from deposit", "l2", {"Amount": Decimal("1000")}),
    ]
    el = R.reconcile_internal(_normalized(rows)).elements[0]
    assert abs(el.delta_detail_vs_reported) <= R.TOLERANCE
    assert el.flag_detail_mismatch is False


def test_reconcile_derived_present_flags_derivation():
    rows = [
        _row("sft", "Sale of securities", "l1",
             {"sellerName": "BROKER X", "salesConsideration": Decimal("1000")},
             feedback_editable_field="salesConsideration"),
        _row("sft", "Sale of securities", "l2",
             {"Amount": Decimal("1000"), "Derived Amount": Decimal("950")}),
    ]
    report = R.reconcile_internal(_normalized(rows))
    el = report.elements[0]
    assert el.reported == Decimal("1000")
    assert el.derived == Decimal("950")
    assert el.effective == Decimal("950")          # derived wins when present
    assert el.flag_derivation is True
    assert el.flagged is True
    # detail matches reported exactly, so ONLY the derivation flag fired
    assert el.flag_detail_mismatch is False


def test_reconcile_derived_equal_to_reported_no_derivation_flag():
    rows = [
        _row("sft", "Sale of securities", "l2",
             {"Amount": Decimal("1000"), "Derived Amount": Decimal("1000")}),
    ]
    el = R.reconcile_internal(_normalized(rows)).elements[0]
    assert el.flag_derivation is False
    assert el.effective == Decimal("1000")


def test_reconcile_derived_none_effective_falls_back_to_reported():
    rows = [_row("sft", "Sale of securities", "l2", {"Amount": Decimal("1000")})]
    el = R.reconcile_internal(_normalized(rows)).elements[0]
    assert el.derived is None
    assert el.effective == Decimal("1000")
    assert el.flag_derivation is False


def test_reconcile_count_mismatch_is_informational_not_a_loud_flag():
    """l2 "Count" is a coarser grain than l1 detail on real AIS (Count counts
    deductors/entries, l1 is per-transaction), so l1 rows legitimately
    outnumber Count. flag_count_mismatch is still COMPUTED (informational
    column) but must NOT contribute to `flagged` / the DIFFERENCE banner --
    otherwise nearly every taxpayer would fire a false positive. Here the money
    (detail vs reported) ties out, so the element is NOT flagged despite the
    count difference."""
    rows = [
        _row("tdsTcs", "Interest from deposit", "l1", {"amtPaid": Decimal("100")},
             feedback_editable_field="amtPaid", info_src_id="A"),
        _row("tdsTcs", "Interest from deposit", "l1", {"amtPaid": Decimal("200")},
             feedback_editable_field="amtPaid", info_src_id="A"),
        _row("tdsTcs", "Interest from deposit", "l2",
             {"Amount": Decimal("300"), "Count": "3"}, info_src_id="A"),
    ]
    report = R.reconcile_internal(_normalized(rows))
    el = report.elements[0]
    assert el.l1_row_count == 2
    assert el.l2_count == "3"
    assert el.flag_count_mismatch is True
    assert el.flagged is False
    assert report.flagged_element_count == 0


def test_reconcile_count_match_no_flag():
    rows = [
        _row("tdsTcs", "X", "l1", {"amtPaid": Decimal("100")},
             feedback_editable_field="amtPaid"),
        _row("tdsTcs", "X", "l2", {"Amount": Decimal("100"), "Count": "1"}),
    ]
    el = R.reconcile_internal(_normalized(rows)).elements[0]
    assert el.flag_count_mismatch is False


def test_reconcile_l1_only_no_l2_no_crash_no_flag():
    """An element with only l1 rows and no l2 aggregate at all (nothing to
    compare against) must not crash and must not spuriously flag."""
    rows = [
        _row("other-info", "Some category with no rollup", "l1",
             {"grossSalary": Decimal("50000")}, feedback_editable_field="grossSalary"),
    ]
    report = R.reconcile_internal(_normalized(rows))
    el = report.elements[0]
    assert el.reported is None
    assert el.detail_sum == Decimal("50000")
    assert el.delta_detail_vs_reported is None
    assert el.flagged is False


def test_reconcile_empty_paymentoftaxes_section_handled_gracefully():
    """A section with NO rows at all (modern AIS: no advance/self-assessment
    tax means paymentOfTaxes emits nothing) must not produce a group, crash,
    or appear in the rollups -- there's simply nothing to reconcile."""
    rows = [
        _row("tdsTcs", "Interest from deposit", "l1", {"amtPaid": Decimal("100")},
             feedback_editable_field="amtPaid"),
        _row("tdsTcs", "Interest from deposit", "l2", {"Amount": Decimal("100")}),
    ]
    report = R.reconcile_internal(_normalized(rows))
    assert not any(el.section_key == "paymentOfTaxes" for el in report.elements)
    assert "paymentOfTaxes" not in report.total_reported_by_section


def test_reconcile_empty_normalized_no_rows_at_all():
    report = R.reconcile_internal(_normalized([]))
    assert report.elements == []
    assert report.total_reported_by_section == {}
    assert report.total_tds_credit == Decimal("0")
    assert report.flagged_element_count == 0


def test_reconcile_total_tds_credit_sums_amount_deposited_and_deducted():
    rows = [
        _row("tdsTcs", "Interest from deposit", "l1",
             {"amtPaid": Decimal("1000"), "amountDeposited": Decimal("100")},
             feedback_editable_field="amtPaid", info_src_id="A"),
        _row("tdsTcs", "Interest from deposit", "l1",
             {"amtPaid": Decimal("2000"), "amountDeducted": Decimal("200")},
             feedback_editable_field="amtPaid", info_src_id="B"),
        # non-tdsTcs l1 rows must NOT contribute to the TDS credit total
        _row("sft", "Sale of securities", "l1", {"salesConsideration": Decimal("999")},
             feedback_editable_field="salesConsideration"),
    ]
    report = R.reconcile_internal(_normalized(rows))
    assert report.total_tds_credit == Decimal("300")


def test_reconcile_total_reported_by_section_rollup():
    rows = [
        _row("tdsTcs", "A", "l2", {"Amount": Decimal("100")}, info_src_id="A"),
        _row("tdsTcs", "B", "l2", {"Amount": Decimal("50")}, info_src_id="B"),
        _row("sft", "C", "l2", {"Amount": Decimal("10")}, info_src_id="C"),
    ]
    report = R.reconcile_internal(_normalized(rows))
    assert report.total_reported_by_section == {
        "tdsTcs": Decimal("150"), "sft": Decimal("10"),
    }


def test_reconcile_status_breakdown_counts_l1_statuses():
    rows = [
        _row("tdsTcs", "A", "l1", {"amtPaid": Decimal("1")}, status="Active",
             feedback_editable_field="amtPaid"),
        _row("tdsTcs", "A", "l1", {"amtPaid": Decimal("2")}, status="Active",
             feedback_editable_field="amtPaid"),
        _row("tdsTcs", "A", "l1", {"amtPaid": Decimal("3")}, status="Denied",
             feedback_editable_field="amtPaid"),
    ]
    el = R.reconcile_internal(_normalized(rows)).elements[0]
    assert el.status_breakdown == {"Active": 2, "Denied": 1}


def test_reconcile_external_matches_extension_point_defaults_empty():
    """Phase B (AIS <-> 26AS) plugs in by populating this field later; for
    Phase A it must always default to an empty list, never None/missing."""
    rows = [_row("tdsTcs", "A", "l1", {"amtPaid": Decimal("1")},
                 feedback_editable_field="amtPaid")]
    el = R.reconcile_internal(_normalized(rows)).elements[0]
    assert el.external_matches == []


def test_reconcile_report_metadata_and_taxpayer_name():
    rows = [_row("tdsTcs", "A", "l1", {"amtPaid": Decimal("1")},
                 feedback_editable_field="amtPaid")]
    report = R.reconcile_internal(_normalized(rows))
    assert report.json_version == "1.0"
    assert report.download_date == "27-Jul-2026"
    assert report.logged_in_pan == PAN
    assert report.taxpayer_name == "SYNTHETIC TAXPAYER"


def test_reconcile_numeric_string_editable_field_summed_into_detail():
    """Regression (found on real files): the feedback-editable field is chosen
    by displayFeedback, not by type, so a genuine money amount can arrive as a
    NUMERIC STRING (real case: other-info/Salary `grossSalary`, columnDataType
    "string"). It must still be coerced and summed into detail_sum, not
    dropped and not crash on Decimal + str."""
    rows = [
        _row("other-info", "Salary", "l1", {"grossSalary": "1200000"},
             feedback_editable_field="grossSalary", info_src_id="EMP1"),
        _row("other-info", "Salary", "l1", {"grossSalary": "300000.50"},
             feedback_editable_field="grossSalary", info_src_id="EMP1"),
        _row("other-info", "Salary", "l2", {"Amount": Decimal("1500000.50")},
             info_src_id="EMP1"),
    ]
    report = R.reconcile_internal(_normalized(rows))
    el = report.elements[0]
    assert el.detail_sum == Decimal("1500000.50")
    assert el.delta_detail_vs_reported == Decimal("0")
    assert not el.flagged


def test_reconcile_non_numeric_editable_field_skipped_not_crash():
    """A feedback-editable field holding genuine text (not a number) drops out
    of detail_sum rather than crashing; with no numeric detail at all and an
    l2 present, detail_sum is None so no detail-vs-reported flag can fire."""
    rows = [
        _row("other-info", "Remarks", "l1", {"remark": "SEE NOTE"},
             feedback_editable_field="remark", info_src_id="R1"),
        _row("other-info", "Remarks", "l2", {"Amount": Decimal("0")},
             info_src_id="R1"),
    ]
    report = R.reconcile_internal(_normalized(rows))
    el = report.elements[0]
    assert el.detail_sum is None
    assert el.flag_detail_mismatch is False


def test_reconcile_taxpayer_name_from_name_of_assessee_label():
    """Real AIS partA labels the name "Name of Assessee" / "Name of Entity",
    not "Name"; the name lookup must match on the "name" substring and skip
    the PAN/Aadhaar identity lines."""
    rows = [_row("tdsTcs", "A", "l1", {"amtPaid": Decimal("1")},
                 feedback_editable_field="amtPaid")]
    part_a = {
        "Permanent Account Number (PAN)": PAN,
        "Aadhaar Number": "XXXX",
        "Name of Assessee": "SYNTHETIC ASSESSEE",
    }
    report = R.reconcile_internal(_normalized(rows, part_a=part_a))
    assert report.taxpayer_name == "SYNTHETIC ASSESSEE"


def test_reconcile_multiple_elements_grouped_independently():
    rows = [
        _row("tdsTcs", "Interest", "l1", {"amtPaid": Decimal("100")},
             feedback_editable_field="amtPaid", info_src_id="A"),
        _row("tdsTcs", "Interest", "l2", {"Amount": Decimal("100")}, info_src_id="A"),
        _row("tdsTcs", "Dividend", "l1", {"amtPaid": Decimal("50")},
             feedback_editable_field="amtPaid", info_src_id="B"),
        _row("tdsTcs", "Dividend", "l2", {"Amount": Decimal("999")}, info_src_id="B"),
    ]
    report = R.reconcile_internal(_normalized(rows))
    assert len(report.elements) == 2
    by_cat = {el.category: el for el in report.elements}
    assert not by_cat["Interest"].flagged
    assert by_cat["Dividend"].flagged


def test_reconcile_full_pipeline_from_positional_synthetic_ais():
    """Integration: decrypt -> normalize -> reconcile over the REAL-shaped
    positional fixture (synthetic_ais_positional), to make sure reconcile
    works on normalize.py's actual output, not just hand-built AisRow lists."""
    n = N.normalize(synthetic_ais_positional())
    report = R.reconcile_internal(n)
    assert len(report.elements) == 2   # the l1-only element + the l2-only element
    assert report.logged_in_pan == PAN


# ---------------------------------------------------------------------------
# excel_writer.py — standalone .xlsx AIS-internal reconciliation report
# ---------------------------------------------------------------------------

from openpyxl import load_workbook  # noqa: E402


def _reco_report(elements, **overrides):
    defaults = dict(
        json_version="1.0", download_date="27-Jul-2026", logged_in_pan=PAN,
        taxpayer_name="SYNTHETIC TAXPAYER",
        total_reported_by_section={}, total_tds_credit=Decimal("0"),
        flagged_element_count=sum(1 for e in elements if e.flagged),
    )
    defaults.update(overrides)
    return R.AisRecoReport(elements=elements, **defaults)


def _element(section_key="tdsTcs", category="Interest", info_src_id="A", *,
             detail_sum=Decimal("100"), reported=Decimal("100"), derived=None,
             l1_row_count=1, l2_count=None, flag_detail_mismatch=False,
             flag_derivation=False, flag_count_mismatch=False, status_breakdown=None):
    from collections import Counter
    delta = (detail_sum - reported) if (detail_sum is not None and reported is not None) else None
    return R.ElementReco(
        section_key=section_key, category=category, info_src_id=info_src_id,
        l1_row_count=l1_row_count, detail_sum=detail_sum, reported=reported,
        derived=derived, effective=derived if derived is not None else reported,
        l2_count=l2_count, delta_detail_vs_reported=delta,
        flag_detail_mismatch=flag_detail_mismatch, flag_derivation=flag_derivation,
        flag_count_mismatch=flag_count_mismatch,
        status_breakdown=status_breakdown or Counter({"Active": l1_row_count}),
    )


def test_excel_writer_smoke_banner_flags_and_sheet_names(tmp_path):
    clean = _element("tdsTcs", "Interest", detail_sum=Decimal("100"), reported=Decimal("100"))
    mismatched = _element("tdsTcs", "Dividend", detail_sum=Decimal("500"),
                          reported=Decimal("300"), flag_detail_mismatch=True)
    report = _reco_report(
        [clean, mismatched],
        total_reported_by_section={"tdsTcs": Decimal("400")},
        total_tds_credit=Decimal("40"),
    )

    out = tmp_path / "ais-reco.xlsx"
    X.write_reco_workbook(report, str(out))
    assert out.exists()

    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Summary", "tdsTcs", "Flags"}

    summary = wb["Summary"]
    assert summary["A1"].value == "1 DIFFERENCE FOUND"

    flags = wb["Flags"]
    assert flags["B2"].value == "Dividend"   # only the flagged element made it in

    # A red-flagged cell exists on the mismatched row in the detail sheet.
    detail = wb["tdsTcs"]
    red_cells = [
        c for row in detail.iter_rows(min_row=2, max_row=detail.max_row)
        for c in row
        if c.fill is not None and c.fill.fgColor and c.fill.fgColor.rgb == "00FFC7CE"
    ]
    assert red_cells, "expected at least one red-filled cell for the flagged row"


def test_excel_writer_no_differences_green_banner(tmp_path):
    clean = _element("tdsTcs", "Interest", detail_sum=Decimal("100"), reported=Decimal("100"))
    report = _reco_report([clean])

    out = tmp_path / "ais-reco-clean.xlsx"
    X.write_reco_workbook(report, str(out))

    wb = load_workbook(out)
    summary = wb["Summary"]
    assert summary["A1"].value == "NO DIFFERENCES"

    flags = wb["Flags"]
    assert flags["A2"].value == "No flagged elements."


def test_excel_writer_creates_one_sheet_per_section_present():
    tds = _element("tdsTcs", "Interest")
    sft = _element("sft", "Sale of securities")
    other = _element("other-info", "Salary")
    report = _reco_report([tds, sft, other])

    import tempfile
    out = Path(tempfile.gettempdir()) / "test_ais_reco_sections.xlsx"
    X.write_reco_workbook(report, str(out))
    wb = load_workbook(out)
    assert {"Summary", "tdsTcs", "sft", "other-info", "Flags"} <= set(wb.sheetnames)


def test_excel_writer_unknown_future_section_still_gets_a_sheet():
    """A sectionKey this writer has never seen (e.g. a future GST section)
    must still get its own detail sheet, not be dropped."""
    el = _element("futureBusinessReceipts", "GST turnover")
    report = _reco_report([el])

    import tempfile
    out = Path(tempfile.gettempdir()) / "test_ais_reco_future_section.xlsx"
    X.write_reco_workbook(report, str(out))
    wb = load_workbook(out)
    assert "futureBusinessReceipts" in wb.sheetnames


def test_excel_writer_full_pipeline_from_positional_synthetic_ais(tmp_path):
    """Integration: decrypt -> normalize -> reconcile -> write, over the
    REAL-shaped positional fixture, end to end."""
    n = N.normalize(synthetic_ais_positional())
    report = R.reconcile_internal(n)

    out = tmp_path / "ais-reco-e2e.xlsx"
    X.write_reco_workbook(report, str(out))

    wb = load_workbook(out)
    assert "Summary" in wb.sheetnames
    assert "Flags" in wb.sheetnames
    assert "tdsTcs" in wb.sheetnames
    # 2 elements in that fixture (one l1-only, one l2-only) -> 2 data rows on the tdsTcs sheet
    detail = wb["tdsTcs"]
    data_rows = [r for r in detail.iter_rows(min_row=2) if r[0].value is not None]
    assert len(data_rows) == 2


# ---------------------------------------------------------------------------
# reconcile_26as.py — Phase B: AIS <-> 26AS TDS-credit tie-out
# ---------------------------------------------------------------------------

from datetime import date  # noqa: E402


def _tds_row(title, *, quarter, amount_deposited, info_src_id="A", amt_paid=None):
    return _row(
        "tdsTcs", title, "l1",
        {
            "amtPaid": amt_paid if amt_paid is not None else amount_deposited * 10,
            "amountDeposited": amount_deposited,
            "quarter": quarter,
            "transactionDate": "15-Jun-2025",
        },
        feedback_editable_field="amtPaid", info_src_id=info_src_id, status="Active",
    )


def _txn(deductor_name, *, txn_date, tds_deposited, tan="TAN00001A", section="194A"):
    return As26Transaction(
        tan=tan, deductor_name=deductor_name, section=section, txn_date=txn_date,
        amount=tds_deposited * 10.0, tax_deducted=tds_deposited, tds_deposited=tds_deposited,
    )


def test_reconcile_26as_aggregate_matches_no_flag():
    rows = [_tds_row("BANK A", quarter="Q1", amount_deposited=Decimal("100"))]
    n = _normalized(rows)
    as26 = As26Data(transactions=[_txn("BANK A", txn_date=date(2025, 5, 10), tds_deposited=100.0)])

    report = R26.reconcile_ais_vs_26as(n, as26)
    assert report.total_ais_tds_credit == Decimal("100")
    assert report.total_26as_tds_deposited == Decimal("100")
    assert report.delta_aggregate == Decimal("0")
    assert report.flag_aggregate_mismatch is False


def test_reconcile_26as_aggregate_mismatch_flags():
    rows = [_tds_row("BANK A", quarter="Q1", amount_deposited=Decimal("100"))]
    n = _normalized(rows)
    as26 = As26Data(transactions=[_txn("BANK A", txn_date=date(2025, 5, 10), tds_deposited=60.0)])

    report = R26.reconcile_ais_vs_26as(n, as26)
    assert report.delta_aggregate == Decimal("40")
    assert report.flag_aggregate_mismatch is True


def test_reconcile_26as_aggregate_within_tolerance_not_flagged():
    rows = [_tds_row("BANK A", quarter="Q1", amount_deposited=Decimal("100.40"))]
    n = _normalized(rows)
    as26 = As26Data(transactions=[_txn("BANK A", txn_date=date(2025, 5, 10), tds_deposited=100.0)])
    report = R26.reconcile_ais_vs_26as(n, as26)
    assert report.flag_aggregate_mismatch is False


def test_reconcile_26as_per_quarter_split_correctness():
    rows = [
        _tds_row("BANK A", quarter="Q1", amount_deposited=Decimal("100"), info_src_id="A"),
        _tds_row("BANK B", quarter="Quarter 3", amount_deposited=Decimal("50"), info_src_id="B"),
    ]
    n = _normalized(rows)
    as26 = As26Data(transactions=[
        _txn("BANK A", txn_date=date(2025, 5, 1), tds_deposited=100.0),   # May -> FY Q1
        _txn("BANK B", txn_date=date(2025, 11, 1), tds_deposited=50.0),  # Nov -> FY Q3
    ])
    report = R26.reconcile_ais_vs_26as(n, as26)
    by_q = {q.quarter: q for q in report.quarters}
    assert len(report.quarters) == 4
    assert by_q[1].ais_tds_credit == Decimal("100")
    assert by_q[1].as26_tds_deposited == Decimal("100")
    assert by_q[1].flagged is False
    assert by_q[3].ais_tds_credit == Decimal("50")
    assert by_q[3].as26_tds_deposited == Decimal("50")
    # untouched quarters are present with zeros, not omitted
    assert by_q[2].ais_tds_credit == Decimal("0")
    assert by_q[2].as26_tds_deposited == Decimal("0")
    assert by_q[2].flagged is False


def test_reconcile_26as_per_quarter_mismatch_flags_only_that_quarter():
    rows = [_tds_row("BANK A", quarter="Q1", amount_deposited=Decimal("100"))]
    n = _normalized(rows)
    as26 = As26Data(transactions=[
        _txn("BANK A", txn_date=date(2025, 5, 1), tds_deposited=100.0),   # Q1: matches
        _txn("BANK X", txn_date=date(2025, 11, 1), tds_deposited=25.0),  # Q3: AIS side has nothing
    ])
    report = R26.reconcile_ais_vs_26as(n, as26)
    by_q = {q.quarter: q for q in report.quarters}
    assert by_q[1].flagged is False
    assert by_q[3].flagged is True
    assert report.flagged_quarters == [3]
    # aggregate also breaks since the totals differ overall
    assert report.flag_aggregate_mismatch is True


def test_reconcile_26as_quarter_extraction_lenient_formats():
    assert R26._extract_fy_quarter("Q1") == 1
    assert R26._extract_fy_quarter("Quarter 3") == 3
    assert R26._extract_fy_quarter("4") == 4
    assert R26._extract_fy_quarter("QTR-2") == 2
    assert R26._extract_fy_quarter(None) is None
    assert R26._extract_fy_quarter("unparseable") is None


def test_reconcile_26as_decimal_float_boundary():
    """as26 amounts are native floats (per As26Transaction); the tie-out must
    convert them to Decimal via str() at the boundary, never mixing float
    and Decimal arithmetic (which would raise TypeError)."""
    rows = [_tds_row("BANK A", quarter="Q1", amount_deposited=Decimal("33.33"))]
    n = _normalized(rows)
    as26 = As26Data(transactions=[
        _txn("BANK A", txn_date=date(2025, 5, 1), tds_deposited=33.33),
    ])
    report = R26.reconcile_ais_vs_26as(n, as26)
    assert isinstance(report.total_26as_tds_deposited, Decimal)
    assert isinstance(report.delta_aggregate, Decimal)
    assert report.total_26as_tds_deposited == Decimal("33.33")
    assert report.flag_aggregate_mismatch is False


def test_reconcile_26as_empty_as26_no_crash():
    rows = [_tds_row("BANK A", quarter="Q1", amount_deposited=Decimal("100"))]
    n = _normalized(rows)
    report = R26.reconcile_ais_vs_26as(n, As26Data())
    assert report.total_26as_tds_deposited == Decimal("0")
    assert report.flag_aggregate_mismatch is True   # 100 vs 0 is a real difference


def test_reconcile_26as_empty_ais_no_crash():
    n = _normalized([])
    as26 = As26Data(transactions=[_txn("BANK A", txn_date=date(2025, 5, 1), tds_deposited=100.0)])
    report = R26.reconcile_ais_vs_26as(n, as26)
    assert report.total_ais_tds_credit == Decimal("0")


# ---------------------------------------------------------------------------
# reconcile_26as.py — per-income-category tie-out (interest/dividend/other)
# ---------------------------------------------------------------------------

# A minimal tds_sections map, same shape as20 classify_section expects:
# category -> list of section codes.
TDS_SECTIONS = {
    "interest": ["194A"],
    "dividend": ["194"],
}


def test_reconcile_26as_category_classification_keyword_based():
    assert R26._classify_ais_category("Interest from deposit") == "interest"
    assert R26._classify_ais_category("Interest on Income Tax Refund") == "interest"
    assert R26._classify_ais_category("Dividend") == "dividend"
    assert R26._classify_ais_category("Dividend (Domestic Company)") == "dividend"
    assert R26._classify_ais_category("Salary") == "salary"
    assert R26._classify_ais_category("Salary received from employer") == "salary"
    assert R26._classify_ais_category("Rent received") == "other"
    assert R26._classify_ais_category(None) == "other"


def test_reconcile_26as_category_tieout_interest_and_dividend_match():
    rows = [
        _tds_row("Interest from deposit", quarter="Q1", amount_deposited=Decimal("100"),
                 info_src_id="A"),
        _tds_row("Dividend", quarter="Q1", amount_deposited=Decimal("50"), info_src_id="B"),
    ]
    n = _normalized(rows)
    as26 = As26Data(transactions=[
        _txn("BANK A", txn_date=date(2025, 5, 1), tds_deposited=100.0, section="194A"),
        _txn("COMPANY X", txn_date=date(2025, 5, 1), tds_deposited=50.0, section="194"),
    ])
    report = R26.reconcile_ais_vs_26as(n, as26, tds_sections=TDS_SECTIONS)
    by_cat = {c.category: c for c in report.categories}
    assert by_cat["interest"].ais_credit == Decimal("100")
    assert by_cat["interest"].as26_deposited == Decimal("100")
    assert by_cat["interest"].flagged is False
    assert by_cat["dividend"].ais_credit == Decimal("50")
    assert by_cat["dividend"].as26_deposited == Decimal("50")
    assert by_cat["dividend"].flagged is False


def test_reconcile_26as_category_tieout_deliberate_mismatch_flags():
    rows = [_tds_row("Interest from deposit", quarter="Q1", amount_deposited=Decimal("100"))]
    n = _normalized(rows)
    as26 = As26Data(transactions=[
        _txn("BANK A", txn_date=date(2025, 5, 1), tds_deposited=40.0, section="194A"),
    ])
    report = R26.reconcile_ais_vs_26as(n, as26, tds_sections=TDS_SECTIONS)
    by_cat = {c.category: c for c in report.categories}
    assert by_cat["interest"].delta == Decimal("60")
    assert by_cat["interest"].flagged is True
    assert report.flagged_categories == ["interest"]


def test_reconcile_26as_category_tieout_unknown_category_falls_to_other():
    rows = [_tds_row("Winnings from lottery", quarter="Q1", amount_deposited=Decimal("20"))]
    n = _normalized(rows)
    as26 = As26Data(transactions=[
        _txn("PAYER Z", txn_date=date(2025, 5, 1), tds_deposited=20.0, section="194B"),
    ])
    report = R26.reconcile_ais_vs_26as(n, as26, tds_sections=TDS_SECTIONS)
    by_cat = {c.category: c for c in report.categories}
    assert "other" in by_cat
    assert by_cat["other"].ais_credit == Decimal("20")
    assert by_cat["other"].as26_deposited == Decimal("20")   # 194B classifies to None -> "other"
    assert by_cat["other"].flagged is False


def test_reconcile_26as_category_tieout_present_on_only_one_side():
    """A category present on only one side still shows up (union), with a
    zero on the missing side, rather than being silently dropped."""
    rows = [_tds_row("Dividend", quarter="Q1", amount_deposited=Decimal("75"))]
    n = _normalized(rows)
    as26 = As26Data(transactions=[])
    report = R26.reconcile_ais_vs_26as(n, as26, tds_sections=TDS_SECTIONS)
    by_cat = {c.category: c for c in report.categories}
    assert by_cat["dividend"].ais_credit == Decimal("75")
    assert by_cat["dividend"].as26_deposited == Decimal("0")
    assert by_cat["dividend"].flagged is True


def test_reconcile_26as_category_tieout_skipped_when_no_tds_sections():
    """tds_sections=None (the default) -> no section->category mapping to
    classify the 26AS side by, so the category tie-out is skipped entirely
    -- aggregate and per-quarter still run normally."""
    rows = [_tds_row("Interest from deposit", quarter="Q1", amount_deposited=Decimal("100"))]
    n = _normalized(rows)
    as26 = As26Data(transactions=[
        _txn("BANK A", txn_date=date(2025, 5, 1), tds_deposited=100.0, section="194A"),
    ])
    report = R26.reconcile_ais_vs_26as(n, as26)   # no tds_sections
    assert report.categories == []
    assert report.flagged_categories == []
    # aggregate + quarter unaffected
    assert report.flag_aggregate_mismatch is False
    assert len(report.quarters) == 4


# ---------------------------------------------------------------------------
# excel_writer.py — "26AS Tie-out" sheet (Phase B)
# ---------------------------------------------------------------------------

def test_excel_writer_26as_tieout_sheet_smoke(tmp_path):
    rows = [
        _tds_row("Interest from deposit", quarter="Q1", amount_deposited=Decimal("100"),
                 info_src_id="A"),
    ]
    n = _normalized(rows)
    as26 = As26Data(transactions=[
        _txn("BANK A", txn_date=date(2025, 5, 1), tds_deposited=60.0, section="194A"),   # mismatch
    ])
    ais_report = R.reconcile_internal(n)
    tieout_report = R26.reconcile_ais_vs_26as(n, as26, tds_sections=TDS_SECTIONS)

    out = tmp_path / "ais-reco-with-26as.xlsx"
    X.write_reco_workbook(ais_report, str(out), report_26as=tieout_report)

    wb = load_workbook(out)
    assert "26AS Tie-out" in wb.sheetnames
    ws = wb["26AS Tie-out"]
    assert ws["A1"].value == "1 DIFFERENCE FOUND"

    # a red-filled cell exists somewhere on the sheet (the flagged aggregate
    # and/or the flagged Q1 row and/or the flagged interest category row)
    red_cells = [
        c for row in ws.iter_rows()
        for c in row
        if c.fill is not None and c.fill.fgColor and c.fill.fgColor.rgb == "00FFC7CE"
    ]
    assert red_cells, "expected at least one red-filled cell on the tie-out sheet"


def test_excel_writer_26as_tieout_no_differences_green_banner(tmp_path):
    rows = [_tds_row("Interest from deposit", quarter="Q1", amount_deposited=Decimal("100"),
                      info_src_id="A")]
    n = _normalized(rows)
    as26 = As26Data(transactions=[
        _txn("BANK A", txn_date=date(2025, 5, 1), tds_deposited=100.0, section="194A"),
    ])
    ais_report = R.reconcile_internal(n)
    tieout_report = R26.reconcile_ais_vs_26as(n, as26, tds_sections=TDS_SECTIONS)

    out = tmp_path / "ais-reco-with-26as-clean.xlsx"
    X.write_reco_workbook(ais_report, str(out), report_26as=tieout_report)

    wb = load_workbook(out)
    ws = wb["26AS Tie-out"]
    assert ws["A1"].value == "NO DIFFERENCES"


def test_excel_writer_26as_tieout_no_tds_sections_shows_placeholder(tmp_path):
    rows = [_tds_row("Interest from deposit", quarter="Q1", amount_deposited=Decimal("100"),
                      info_src_id="A")]
    n = _normalized(rows)
    as26 = As26Data(transactions=[
        _txn("BANK A", txn_date=date(2025, 5, 1), tds_deposited=100.0, section="194A"),
    ])
    ais_report = R.reconcile_internal(n)
    tieout_report = R26.reconcile_ais_vs_26as(n, as26)   # no tds_sections

    out = tmp_path / "ais-reco-with-26as-no-categories.xlsx"
    X.write_reco_workbook(ais_report, str(out), report_26as=tieout_report)

    wb = load_workbook(out)
    ws = wb["26AS Tie-out"]
    cell_values = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert any("tds_sections not supplied" in str(v) for v in cell_values)


def test_excel_writer_omits_26as_sheet_when_not_provided(tmp_path):
    """Existing single-argument call keeps working unchanged: no 26AS
    report -> no 26AS Tie-out sheet, everything else as before."""
    clean = _element("tdsTcs", "Interest", detail_sum=Decimal("100"), reported=Decimal("100"))
    report = _reco_report([clean])

    out = tmp_path / "ais-reco-no-26as.xlsx"
    X.write_reco_workbook(report, str(out))

    wb = load_workbook(out)
    assert "26AS Tie-out" not in wb.sheetnames
    assert set(wb.sheetnames) == {"Summary", "tdsTcs", "Flags"}


# ---------------------------------------------------------------------------
# reconcile_books.py -- Phase C: AIS <-> GnuCash books reconciliation
# ---------------------------------------------------------------------------

from agents.skill_ais_reconcile import reconcile_books as RB  # noqa: E402
from agents.skill_itr_workbook.scripts.parse_gnucash import (  # noqa: E402
    Account, Book, Split, Transaction,
)
from agents.skill_itr_workbook.scripts.configs import MappingEntry, MappingLoadResult  # noqa: E402

YEAR_KEY = "2025-26"


def _acct(guid, name, type_, parent_guid=None):
    return Account(guid=guid, name=name, type=type_, parent_guid=parent_guid)


def _split(account_guid, value):
    from fractions import Fraction
    return Split(guid=f"sp-{account_guid}-{value}", account_guid=account_guid,
                 value=Fraction(value), quantity=Fraction(value))


def _txn_book(guid, *, posted, splits):
    return Transaction(guid=guid, date_posted=posted, description="synthetic", splits=splits)


def _mapping(entries: dict[str, str]) -> MappingLoadResult:
    """entries: guid -> tag"""
    return MappingLoadResult(
        entries={guid: MappingEntry(guid=guid, path=guid, tag=tag) for guid, tag in entries.items()},
        warnings=[],
    )


def _books_element(category, detail_sum, *, info_src_id="A"):
    return _element("tdsTcs", category, info_src_id, detail_sum=detail_sum, reported=detail_sum)


def test_reconcile_books_nearest_ancestor_tag_resolution():
    """A tag on a parent account covers an untagged child: the child's FY
    sum rolls up into the parent's tag bucket."""
    root = _acct("root", "Root", "ROOT")
    parent = _acct("p1", "Interest Income", "INCOME", parent_guid="root")
    child = _acct("c1", "Bank FD Interest", "INCOME", parent_guid="p1")
    book = Book(
        accounts={"root": root, "p1": parent, "c1": child},
        transactions=[
            _txn_book("t1", posted=date(2025, 5, 1),
                      splits=[_split("c1", -100)]),  # INCOME flips sign -> +100
        ],
    )
    mapping = _mapping({"p1": "OS_INTEREST_BANK"})

    report = RB.reconcile_ais_vs_books(
        _reco_report([_books_element("Interest from deposit", Decimal("100"))],
                     total_tds_credit=Decimal("0")),
        book, mapping, YEAR_KEY,
    )
    interest = next(c for c in report.categories if c.category == "interest")
    assert interest.books_income == Decimal("100")
    assert interest.ais_income == Decimal("100")
    assert not interest.flagged


def test_reconcile_books_category_mismatch_flags():
    acct = _acct("a1", "Dividend Income", "INCOME")
    book = Book(
        accounts={"a1": acct},
        transactions=[_txn_book("t1", posted=date(2025, 6, 1), splits=[_split("a1", -50)])],
    )
    mapping = _mapping({"a1": "OS_DIVIDEND"})

    report = RB.reconcile_ais_vs_books(
        _reco_report([_books_element("Dividend", Decimal("500"))], total_tds_credit=Decimal("0")),
        book, mapping, YEAR_KEY,
    )
    dividend = next(c for c in report.categories if c.category == "dividend")
    assert dividend.ais_income == Decimal("500")
    assert dividend.books_income == Decimal("50")
    assert dividend.delta == Decimal("450")
    assert dividend.flagged


def test_reconcile_books_tds_credit_tie_out_match():
    acct = _acct("t1", "TDS on Interest", "ASSET")
    book = Book(
        accounts={"t1": acct},
        transactions=[_txn_book("tx1", posted=date(2025, 7, 1), splits=[_split("t1", 74)])],
    )
    mapping = _mapping({"t1": "TAXPAID_TDS_INTEREST"})

    report = RB.reconcile_ais_vs_books(
        _reco_report([], total_tds_credit=Decimal("74")),
        book, mapping, YEAR_KEY,
    )
    assert report.ais_tds_credit == Decimal("74")
    assert report.books_tds_credit == Decimal("74")
    assert not report.flag_tds_mismatch


def test_reconcile_books_tds_credit_tie_out_mismatch():
    acct = _acct("t1", "TDS on Interest", "ASSET")
    book = Book(
        accounts={"t1": acct},
        transactions=[_txn_book("tx1", posted=date(2025, 7, 1), splits=[_split("t1", 50)])],
    )
    mapping = _mapping({"t1": "TAXPAID_TDS_INTEREST"})

    report = RB.reconcile_ais_vs_books(
        _reco_report([], total_tds_credit=Decimal("74")),
        book, mapping, YEAR_KEY,
    )
    assert report.delta_tds == Decimal("24")
    assert report.flag_tds_mismatch


def test_reconcile_books_income_not_in_ais_completeness_list():
    """Books show salary income; AIS reports none -- informational list."""
    acct = _acct("s1", "Salary Income", "INCOME")
    book = Book(
        accounts={"s1": acct},
        transactions=[_txn_book("tx1", posted=date(2025, 5, 1), splits=[_split("s1", -1000)])],
    )
    mapping = _mapping({"s1": "SALARY_GROSS"})

    report = RB.reconcile_ais_vs_books(
        _reco_report([], total_tds_credit=Decimal("0")),
        book, mapping, YEAR_KEY,
    )
    assert "salary" in report.books_income_not_in_ais
    assert "salary" not in report.ais_income_not_in_books


def test_reconcile_books_ais_income_not_in_books_completeness_list():
    """AIS reports interest income; books have no corresponding posting --
    the loud call-out since books is primary."""
    report = RB.reconcile_ais_vs_books(
        _reco_report([_books_element("Interest from deposit", Decimal("300"))],
                     total_tds_credit=Decimal("0")),
        Book(accounts={}, transactions=[]),
        _mapping({}),
        YEAR_KEY,
    )
    assert "interest" in report.ais_income_not_in_books
    assert "interest" not in report.books_income_not_in_ais


def test_reconcile_books_untagged_income_account_count():
    tagged = _acct("a1", "Salary Income", "INCOME")
    untagged = _acct("a2", "Misc Income", "INCOME")
    book = Book(
        accounts={"a1": tagged, "a2": untagged},
        transactions=[
            _txn_book("t1", posted=date(2025, 5, 1), splits=[_split("a1", -100)]),
            _txn_book("t2", posted=date(2025, 5, 2), splits=[_split("a2", -200)]),
        ],
    )
    mapping = _mapping({"a1": "SALARY_GROSS"})

    report = RB.reconcile_ais_vs_books(
        _reco_report([], total_tds_credit=Decimal("0")), book, mapping, YEAR_KEY,
    )
    assert report.untagged_income_account_count == 1


def test_reconcile_books_decimal_float_boundary():
    """account_fy_sum returns a native float (via Fraction); the report's
    books_income must come back as an exact Decimal, not a float artifact."""
    acct = _acct("a1", "Bank FD Interest", "INCOME")
    from fractions import Fraction
    book = Book(
        accounts={"a1": acct},
        transactions=[_txn_book("t1", posted=date(2025, 5, 1),
                                 splits=[Split(guid="sp1", account_guid="a1",
                                               value=Fraction(-1, 10), quantity=Fraction(-1, 10))])],
    )
    mapping = _mapping({"a1": "OS_INTEREST_BANK"})

    report = RB.reconcile_ais_vs_books(
        _reco_report([], total_tds_credit=Decimal("0")), book, mapping, YEAR_KEY,
    )
    interest = next(c for c in report.categories if c.category == "interest")
    assert isinstance(interest.books_income, Decimal)
    assert interest.books_income == Decimal("0.1")


def test_reconcile_books_untagged_child_excludes_tagged_parent_from_untagged_count():
    """An account whose ancestor IS mapped should not count as untagged, even
    though the account itself carries no direct mapping entry."""
    parent = _acct("p1", "Interest Income", "INCOME")
    child = _acct("c1", "Bank FD Interest", "INCOME", parent_guid="p1")
    book = Book(
        accounts={"p1": parent, "c1": child},
        transactions=[_txn_book("t1", posted=date(2025, 5, 1), splits=[_split("c1", -10)])],
    )
    mapping = _mapping({"p1": "OS_INTEREST_BANK"})

    report = RB.reconcile_ais_vs_books(
        _reco_report([], total_tds_credit=Decimal("0")), book, mapping, YEAR_KEY,
    )
    assert report.untagged_income_account_count == 0


def test_excel_writer_books_sheet_smoke(tmp_path):
    acct = _acct("a1", "Bank FD Interest", "INCOME")
    book = Book(
        accounts={"a1": acct},
        transactions=[_txn_book("t1", posted=date(2025, 5, 1), splits=[_split("a1", -500)])],
    )
    mapping = _mapping({"a1": "OS_INTEREST_BANK"})
    ais_report = _reco_report([_books_element("Interest from deposit", Decimal("100"))],
                               total_tds_credit=Decimal("0"))
    books_report = RB.reconcile_ais_vs_books(ais_report, book, mapping, YEAR_KEY)

    out = tmp_path / "ais-reco-with-books.xlsx"
    X.write_reco_workbook(ais_report, str(out), report_books=books_report)

    wb = load_workbook(out)
    assert "Books Reconciliation" in wb.sheetnames
    # placed right after Summary
    assert wb.sheetnames[0] == "Summary"
    assert wb.sheetnames[1] == "Books Reconciliation"

    ws = wb["Books Reconciliation"]
    red_cells = [
        c for row in ws.iter_rows()
        for c in row
        if c.fill is not None and c.fill.fgColor and c.fill.fgColor.rgb == "00FFC7CE"
    ]
    assert red_cells, "expected at least one red-filled cell (interest category flagged)"


def test_excel_writer_omits_books_sheet_when_not_provided(tmp_path):
    """Existing calls without report_books keep working unchanged."""
    clean = _element("tdsTcs", "Interest", detail_sum=Decimal("100"), reported=Decimal("100"))
    report = _reco_report([clean])

    out = tmp_path / "ais-reco-no-books.xlsx"
    X.write_reco_workbook(report, str(out))

    wb = load_workbook(out)
    assert "Books Reconciliation" not in wb.sheetnames
    assert set(wb.sheetnames) == {"Summary", "tdsTcs", "Flags"}


# ---------------------------------------------------------------------------
# feedback.py -- Phase D: advisory AIS portal-feedback suggestions
# ---------------------------------------------------------------------------

from agents.skill_ais_reconcile import feedback as F  # noqa: E402


def test_feedback_derivation_flag_suggestion():
    el = _element("tdsTcs", "Interest", detail_sum=Decimal("100"), reported=Decimal("100"),
                  derived=Decimal("80"), flag_derivation=True)
    report = _reco_report([el])

    suggestions = F.suggest_feedback(report)
    assert len(suggestions) == 1
    s = suggestions[0]
    assert s.suggested_action == F.FEEDBACK_NOT_FULLY_CORRECT
    assert s.confidence == "low"
    assert s.section_key == "tdsTcs"
    assert s.category == "Interest"


def test_feedback_detail_mismatch_falls_back_to_review():
    el = _element("tdsTcs", "Dividend", detail_sum=Decimal("500"), reported=Decimal("300"),
                  flag_detail_mismatch=True)
    report = _reco_report([el])

    suggestions = F.suggest_feedback(report)
    assert len(suggestions) == 1
    s = suggestions[0]
    assert s.suggested_action == F.FEEDBACK_REVIEW
    assert s.confidence == "low"


def test_feedback_ais_income_not_in_books_review_medium():
    report = _reco_report([], total_tds_credit=Decimal("0"))
    books = RB.AisBooksReport(
        categories=[], ais_tds_credit=Decimal("0"), books_tds_credit=Decimal("0"),
        delta_tds=Decimal("0"), flag_tds_mismatch=False,
        books_income_not_in_ais=[], ais_income_not_in_books=["interest"],
        untagged_income_account_count=0,
    )

    suggestions = F.suggest_feedback(report, books=books)
    assert len(suggestions) == 1
    s = suggestions[0]
    assert s.category == "interest"
    assert s.suggested_action == F.FEEDBACK_REVIEW
    assert s.confidence == "medium"


def test_feedback_books_income_not_in_ais_review_medium():
    report = _reco_report([], total_tds_credit=Decimal("0"))
    books = RB.AisBooksReport(
        categories=[], ais_tds_credit=Decimal("0"), books_tds_credit=Decimal("0"),
        delta_tds=Decimal("0"), flag_tds_mismatch=False,
        books_income_not_in_ais=["salary"], ais_income_not_in_books=[],
        untagged_income_account_count=0,
    )

    suggestions = F.suggest_feedback(report, books=books)
    assert len(suggestions) == 1
    s = suggestions[0]
    assert s.category == "salary"
    assert s.suggested_action == F.FEEDBACK_REVIEW
    assert s.confidence == "medium"


def test_feedback_ais_income_not_in_books_duplicate_variant_when_also_derivation_flagged():
    el = _element("tdsTcs", "interest", detail_sum=Decimal("100"), reported=Decimal("100"),
                  derived=Decimal("80"), flag_derivation=True)
    report = _reco_report([el])
    books = RB.AisBooksReport(
        categories=[], ais_tds_credit=Decimal("0"), books_tds_credit=Decimal("0"),
        delta_tds=Decimal("0"), flag_tds_mismatch=False,
        books_income_not_in_ais=[], ais_income_not_in_books=["interest"],
        untagged_income_account_count=0,
    )

    suggestions = F.suggest_feedback(report, books=books)
    actions = {s.suggested_action for s in suggestions}
    assert F.FEEDBACK_NOT_FULLY_CORRECT in actions   # from the derivation flag itself
    assert F.FEEDBACK_REVIEW in actions               # ais_income_not_in_books
    assert F.FEEDBACK_DUPLICATE in actions             # the duplicate-hint variant
    dup = next(s for s in suggestions if s.suggested_action == F.FEEDBACK_DUPLICATE)
    assert dup.confidence == "low"


def test_feedback_26as_flagged_category_review_medium():
    report = _reco_report([], total_tds_credit=Decimal("0"))
    as26_report = R26.Ais26asReport(
        total_ais_tds_credit=Decimal("100"), total_26as_tds_deposited=Decimal("50"),
        delta_aggregate=Decimal("50"), flag_aggregate_mismatch=True,
        quarters=[], categories=[
            R26.CategoryTieOut(category="interest", ais_credit=Decimal("100"),
                                as26_deposited=Decimal("50"), delta=Decimal("50"), flagged=True),
        ],
    )

    suggestions = F.suggest_feedback(report, as26=as26_report)
    assert len(suggestions) == 1
    s = suggestions[0]
    assert s.category == "interest"
    assert s.suggested_action == F.FEEDBACK_REVIEW
    assert s.confidence == "medium"


def test_feedback_26as_aggregate_mismatch_no_category_granularity():
    """flag_aggregate_mismatch true but no flagged categories (tds_sections
    wasn't supplied) -- still surfaces one suggestion, not silently dropped."""
    report = _reco_report([], total_tds_credit=Decimal("0"))
    as26_report = R26.Ais26asReport(
        total_ais_tds_credit=Decimal("100"), total_26as_tds_deposited=Decimal("50"),
        delta_aggregate=Decimal("50"), flag_aggregate_mismatch=True,
        quarters=[], categories=[],
    )

    suggestions = F.suggest_feedback(report, as26=as26_report)
    assert len(suggestions) == 1
    assert suggestions[0].suggested_action == F.FEEDBACK_REVIEW


def test_feedback_no_flags_produces_no_suggestions():
    el = _element("tdsTcs", "Interest", detail_sum=Decimal("100"), reported=Decimal("100"))
    report = _reco_report([el])
    assert F.suggest_feedback(report) == []


def test_feedback_dedupe_same_section_category_info_src_action():
    el = _element("tdsTcs", "Interest", info_src_id="A", detail_sum=Decimal("100"),
                  reported=Decimal("100"), derived=Decimal("80"), flag_derivation=True)
    report = _reco_report([el, el])   # duplicate element -> would double-fire without dedupe
    suggestions = F.suggest_feedback(report)
    assert len(suggestions) == 1


def test_feedback_sort_order_medium_before_low():
    el = _element("tdsTcs", "Interest", detail_sum=Decimal("100"), reported=Decimal("100"),
                  derived=Decimal("80"), flag_derivation=True)   # low confidence
    report = _reco_report([el])
    books = RB.AisBooksReport(
        categories=[], ais_tds_credit=Decimal("0"), books_tds_credit=Decimal("0"),
        delta_tds=Decimal("0"), flag_tds_mismatch=False,
        books_income_not_in_ais=["salary"], ais_income_not_in_books=[],  # medium confidence
        untagged_income_account_count=0,
    )

    suggestions = F.suggest_feedback(report, books=books)
    confidences = [s.confidence for s in suggestions]
    assert confidences.index("medium") < confidences.index("low")


def test_feedback_confidence_never_high():
    el_deriv = _element("tdsTcs", "Interest", detail_sum=Decimal("100"), reported=Decimal("100"),
                         derived=Decimal("80"), flag_derivation=True)
    el_detail = _element("tdsTcs", "Dividend", detail_sum=Decimal("500"), reported=Decimal("300"),
                          flag_detail_mismatch=True)
    report = _reco_report([el_deriv, el_detail])
    books = RB.AisBooksReport(
        categories=[], ais_tds_credit=Decimal("0"), books_tds_credit=Decimal("0"),
        delta_tds=Decimal("0"), flag_tds_mismatch=False,
        books_income_not_in_ais=["salary"], ais_income_not_in_books=["other"],
        untagged_income_account_count=0,
    )
    as26_report = R26.Ais26asReport(
        total_ais_tds_credit=Decimal("100"), total_26as_tds_deposited=Decimal("50"),
        delta_aggregate=Decimal("50"), flag_aggregate_mismatch=True,
        quarters=[], categories=[
            R26.CategoryTieOut(category="interest", ais_credit=Decimal("100"),
                                as26_deposited=Decimal("50"), delta=Decimal("50"), flagged=True),
        ],
    )

    suggestions = F.suggest_feedback(report, books=books, as26=as26_report)
    assert suggestions
    assert all(s.confidence in ("low", "medium") for s in suggestions)


def test_excel_writer_feedback_sheet_smoke_and_advisory_header(tmp_path):
    el = _element("tdsTcs", "Interest", detail_sum=Decimal("100"), reported=Decimal("100"),
                  derived=Decimal("80"), flag_derivation=True)
    report = _reco_report([el])
    suggestions = F.suggest_feedback(report)

    out = tmp_path / "ais-reco-with-feedback.xlsx"
    X.write_reco_workbook(report, str(out), suggestions=suggestions)

    wb = load_workbook(out)
    assert "Feedback Suggestions" in wb.sheetnames
    ws = wb["Feedback Suggestions"]

    cell_values = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert any("ADVISORY ONLY" in str(v) for v in cell_values)
    assert any(F.FEEDBACK_NOT_FULLY_CORRECT == v for v in cell_values)
    # never a scare-red fill on this sheet -- these are suggestions, not errors
    red_cells = [
        c for row in ws.iter_rows()
        for c in row
        if c.fill is not None and c.fill.fgColor and c.fill.fgColor.rgb == "00FFC7CE"
    ]
    assert not red_cells


def test_excel_writer_omits_feedback_sheet_when_not_provided(tmp_path):
    clean = _element("tdsTcs", "Interest", detail_sum=Decimal("100"), reported=Decimal("100"))
    report = _reco_report([clean])

    out = tmp_path / "ais-reco-no-feedback.xlsx"
    X.write_reco_workbook(report, str(out))

    wb = load_workbook(out)
    assert "Feedback Suggestions" not in wb.sheetnames


# ---------------------------------------------------------------------------
# agent.py -- Phase E: run() orchestration (direct-mode skill entry point)
# ---------------------------------------------------------------------------

import yaml  # noqa: E402
from agents.skill_ais_reconcile import agent as AG  # noqa: E402

ENTITY_KEY = "harshal"
# PAN[3:9] = "DE1234" -> masked "XXX" + "DE1234" + "X"
AIS_FILENAME = "XXXDE1234X_2025-26_AIS.json"


def _write_entities_yaml(tmp_path, *, key=ENTITY_KEY, pan=PAN, status="Individual",
                          dob=DOB_ISO, doi=None) -> Path:
    fields = {"name": "Synthetic Taxpayer", "pan": pan, "status": status}
    if dob:
        fields["dob"] = dob
    if doi:
        fields["doi"] = doi
    path = tmp_path / "entities.yaml"
    path.write_text(yaml.safe_dump({key: fields}), encoding="utf-8")
    return path


def _write_encrypted_ais(tmp_path, *, filename=AIS_FILENAME) -> Path:
    ais = synthetic_ais()
    password = D.derive_password_from_iso_date(PAN, DOB_ISO)
    blob = D._encrypt_for_test(ais, password)
    path = tmp_path / filename
    path.write_text(blob, encoding="utf-8")
    return path


def _write_mapping_stub(tmp_path, *, key=ENTITY_KEY) -> Path:
    """agent.py only checks mapping_path.is_file() before calling the
    (monkeypatched) configs.load_mapping -- content is irrelevant here."""
    mdir = tmp_path / "mappings"
    mdir.mkdir(exist_ok=True)
    path = mdir / f"{key}.mapping.yaml"
    path.write_text("[]", encoding="utf-8")
    return path


def _synthetic_book(*, in_fy: bool):
    """A tiny one-account book with a single INCOME transaction either
    inside or outside YEAR_KEY's FY window, for the FY-match check."""
    root = _acct("root", "Root", "ROOT")
    interest = _acct("i1", "Interest Income", "INCOME", parent_guid="root")
    posted = date(2025, 6, 1) if in_fy else date(2020, 6, 1)
    book = Book(
        accounts={"root": root, "i1": interest},
        transactions=[_txn_book("t1", posted=posted, splits=[_split("i1", -1000)])],
    )
    return book


def _empty_26as_workbook(tmp_path) -> Path:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Part I"
    ws.append(["Sr.No", "Name of Deductor", "TAN", "Section"])  # header-only, no data rows
    path = tmp_path / "synthetic-26AS.xlsx"
    wb.save(path)
    return path


def test_run_no_entity_match_returns_error(tmp_path):
    entities_path = _write_entities_yaml(tmp_path, pan="ZZZZZ0000Z")  # won't mask to AIS_FILENAME's prefix
    ais_path = _write_encrypted_ais(tmp_path)
    out = tmp_path / "out.xlsx"

    summary = AG.run(str(ais_path), str(out), entities_path=str(entities_path))

    assert summary.startswith("ERROR:")
    assert "XXXDE1234X" in summary
    assert not out.exists()


def test_run_ais_internal_only_resolves_entity_and_writes_workbook(tmp_path, monkeypatch):
    entities_path = _write_entities_yaml(tmp_path)
    ais_path = _write_encrypted_ais(tmp_path)
    out = tmp_path / "out.xlsx"

    summary = AG.run(str(ais_path), str(out), entities_path=str(entities_path))

    assert "Synthetic Taxpayer" in summary
    assert "2025-26" in summary
    assert "AIS-internal:" in summary
    assert "Feedback:" in summary
    assert out.is_file()
    wb = load_workbook(out)
    assert "Summary" in wb.sheetnames
    assert "Books Reconciliation" not in wb.sheetnames


def test_run_with_matching_fy_book_runs_books_phase_no_warning(tmp_path, monkeypatch):
    entities_path = _write_entities_yaml(tmp_path)
    ais_path = _write_encrypted_ais(tmp_path)
    _write_mapping_stub(tmp_path)
    gnucash_path = tmp_path / "book.gnucash"
    gnucash_path.write_text("not real xml -- parse_book is monkeypatched", encoding="utf-8")
    out = tmp_path / "out.xlsx"

    book = _synthetic_book(in_fy=True)
    mapping = _mapping({"i1": "OS_INTEREST_BANK"})
    monkeypatch.setattr(AG.pg, "parse_book", lambda path: book)
    monkeypatch.setattr(AG.configs, "load_mapping", lambda path: mapping)

    summary = AG.run(str(ais_path), str(out), gnucash_path=str(gnucash_path),
                      entities_path=str(entities_path))

    assert "WARNING" not in summary
    assert "Books:" in summary
    wb = load_workbook(out)
    assert wb.sheetnames[1] == "Books Reconciliation"


def test_run_fy_mismatch_warning_fires(tmp_path, monkeypatch):
    entities_path = _write_entities_yaml(tmp_path)
    ais_path = _write_encrypted_ais(tmp_path)
    _write_mapping_stub(tmp_path)
    gnucash_path = tmp_path / "wrong-year-book.gnucash"
    gnucash_path.write_text("not real xml -- parse_book is monkeypatched", encoding="utf-8")
    out = tmp_path / "out.xlsx"

    book = _synthetic_book(in_fy=False)
    mapping = _mapping({"i1": "OS_INTEREST_BANK"})
    monkeypatch.setattr(AG.pg, "parse_book", lambda path: book)
    monkeypatch.setattr(AG.configs, "load_mapping", lambda path: mapping)

    summary = AG.run(str(ais_path), str(out), gnucash_path=str(gnucash_path),
                      entities_path=str(entities_path))

    assert "WARNING" in summary
    assert "no transactions in FY 2025-26" in summary
    assert "Books:" in summary  # Phase C still ran (sheet still exists)
    wb = load_workbook(out)
    assert "Books Reconciliation" in wb.sheetnames


def test_run_with_26as_workbook_runs_26as_phase(tmp_path):
    entities_path = _write_entities_yaml(tmp_path)
    ais_path = _write_encrypted_ais(tmp_path)
    xlsx_path = _empty_26as_workbook(tmp_path)
    out = tmp_path / "out.xlsx"

    summary = AG.run(str(ais_path), str(out), xlsx_path=str(xlsx_path), entities_path=str(entities_path))

    assert "26AS:" in summary
    wb = load_workbook(out)
    assert "26AS Tie-out" in wb.sheetnames


def test_run_huf_entity_uses_doi_for_password(tmp_path):
    """HUF entities derive the decrypt password from doi, not dob -- confirm
    the wrong-field case fails loud rather than silently trying dob."""
    entities_path = _write_entities_yaml(
        tmp_path, key="hufentity", pan=PAN, status="HUF", dob=None, doi=None,
    )
    ais_path = _write_encrypted_ais(tmp_path)
    out = tmp_path / "out.xlsx"

    summary = AG.run(str(ais_path), str(out), entities_path=str(entities_path))

    assert summary.startswith("ERROR:")
    assert "doi" in summary
