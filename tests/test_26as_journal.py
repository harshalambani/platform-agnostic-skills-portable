"""
tests/test_26as_journal.py — Tests for the 26AS -> GnuCash TDS journal builder.

Uses synthetic deductors + a synthetic account tree (no PII, no external
files) so the matcher, split math, balancing, overrides and CSV output are
all exercised in CI. An optional end-to-end test runs only if a fixture
workbook + .gnucash are dropped under tests/fixtures/.

Run with:
    cd src && python -m pytest ../tests/test_26as_journal.py -v
"""
from __future__ import annotations

import csv
import datetime as dt
import importlib.util
import sys
import tempfile
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "src"
SCRIPT = SRC / "agents" / "skill_26as_journal" / "scripts" / "build_tds_journals.py"


def _load():
    if str(SRC) not in sys.path:
        sys.path.insert(0, str(SRC))
    spec = importlib.util.spec_from_file_location("build_tds_journals", SCRIPT)
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod  # needed for dataclass annotation resolution
    spec.loader.exec_module(mod)
    return mod


m = _load()


def _accounts():
    """Synthetic account tree mirroring the real MyFinances2425 structure."""
    inc = [
        "Income:Interest Income:Interest on FD",
        "Income:Interest Income:Interest on BOB - FD",
        "Income:Interest Income:Interest on ICICI Bank - FD",
        "Income:Interest Income:Interest on Chola Bond",
        "Income:Interest Income:Interest on EPF Taxable",
        "Income:Interest Income:Interest from HSBC Bank",
        "Income:Dividend - MF",
        "Income:Dividend - Shares",
        "Income:Dividend - Shares:Dividend - DRL",
        "Income:Dividend - Shares:Dividend - Ramco Cements",
        "Income:Dividend - Shares:Dividend - JB Chemicals",
        "Income:xBusiness Income:Remuneration from Partnership",
    ]
    accts = [m.Account(path=p, leaf=p.split(":")[-1], type="INCOME") for p in inc]
    accts += [
        m.Account("Expense:TDS on Interest", "TDS on Interest", "EXPENSE"),
        m.Account("Expense:TDS on Dividend", "TDS on Dividend", "EXPENSE"),
        m.Account("Liabilities:Suspense", "Suspense", "LIABILITY"),
    ]
    return accts


def _deductor(sr, name, section, amt, tax):
    return m.Deductor(sr=sr, name=name, sections=(section,),
                      amount_paid=amt, tax_deducted=tax, tds_deposited=tax)


# ---------------------------------------------------------------------------
# Categorisation
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("section,cat", [
    ("194A", "A"), ("193", "A"), ("194", "B"), ("194T", "C"),
])
def test_categorize(section, cat):
    assert m.categorize((section,))[0] == cat


def test_categorize_unknown():
    assert m.categorize(("194I",))[0] is None  # unhandled section


# ---------------------------------------------------------------------------
# Matching — the 8 sample deductors land on the right account
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("name,cat,expected", [
    ("BANK OF BARODA", "A", "Income:Interest Income:Interest on BOB - FD"),
    ("ICICI BANK LIMITED", "A", "Income:Interest Income:Interest on ICICI Bank - FD"),
    ("CHOLAMANDALAM INVESTMENT AND FINANCE COMPANY LIMITED", "A",
     "Income:Interest Income:Interest on Chola Bond"),
    ("OFFICE OF REGIONAL PROVIDENT FUND COMMISSIONER BANDRA EAST", "A",
     "Income:Interest Income:Interest on EPF Taxable"),
    ("DR REDDY'S LABORATORIES LTD.", "B", "Income:Dividend - Shares:Dividend - DRL"),
    ("THE RAMCO CEMENTS LIMITED", "B", "Income:Dividend - Shares:Dividend - Ramco Cements"),
    ("J.B.CHEMICALS & PHARMACEUTICALS LTD.", "B", "Income:Dividend - Shares:Dividend - JB Chemicals"),
    ("ACME CONSULTING LLP", "C", "Income:xBusiness Income:Remuneration from Partnership"),
])
def test_match_sample_deductors(name, cat, expected):
    acct, conf, basis, cands = m.match_credit_account(name, cat, _accounts())
    assert acct == expected, f"{name}: got {acct} ({basis})"


def test_match_no_candidate_goes_suspense():
    """An interest deductor with no resembling account stays unmatched (Suspense)."""
    acct, conf, basis, cands = m.match_credit_account("ZZ UNKNOWN ENTITY XQ", "A", _accounts())
    assert acct is None and conf == "Suspense"


# ---------------------------------------------------------------------------
# Placeholder / hidden accounts are never valid credit candidates (the bug:
# 'Income:Interest Income' is a GnuCash placeholder and cannot be posted to).
# ---------------------------------------------------------------------------

def test_placeholder_parent_excluded_from_candidates():
    accts = [
        m.Account("Income:Interest Income", "Interest Income", "INCOME", special=True),
        m.Account("Income:Interest Income:Interest on HDFC - FD",
                  "Interest on HDFC - FD", "INCOME"),
    ]
    paths = [c.path for c in m._candidates_for("A", accts)]
    assert "Income:Interest Income" not in paths          # placeholder header
    assert "Income:Interest Income:Interest on HDFC - FD" in paths


def test_placeholder_only_interest_account_routes_to_suspense():
    """If the ONLY category-A account is the placeholder parent, the deductor
    must go to Suspense — never post directly to the placeholder."""
    accts = [m.Account("Income:Interest Income", "Interest Income", "INCOME",
                       special=True)]
    acct, conf, basis, cands = m.match_credit_account("HDFC BANK LIMITED", "A", accts)
    assert acct is None and conf == "Suspense"
    assert cands == []


def test_hidden_account_never_matched():
    """A hidden (retired) account must not be offered even if the name matches."""
    accts = [
        m.Account("Income:Interest Income:Interest on Old Bank",
                  "Interest on Old Bank", "INCOME", special=True),
        m.Account("Income:Interest Income:Interest on HDFC - FD",
                  "Interest on HDFC - FD", "INCOME"),
    ]
    acct, conf, basis, cands = m.match_credit_account("OLD BANK", "A", accts)
    assert "Income:Interest Income:Interest on Old Bank" not in cands
    assert acct != "Income:Interest Income:Interest on Old Bank"


def test_special_account_not_taken_as_generic_fd():
    """A placeholder/hidden FD account must not be picked as the generic FD
    debit account either."""
    def acc(p, special=False):
        return m.Account(p, p.split(":")[-1], "INCOME", special=special)
    accts = [
        acc("Income:Interest Income:Interest on FD", special=True),   # placeholder
        acc("Income:Interest Income:Interest on Fixed Deposit"),      # real generic
    ]
    assert m.find_generic_fd_account(accts) == \
        "Income:Interest Income:Interest on Fixed Deposit"


def test_build_journals_places_placeholder_only_deductor_on_suspense():
    """End-to-end at the journal level: a deductor whose category has only a
    placeholder account lands its credit split on Suspense and is flagged."""
    accts = [
        m.Account("Income:Interest Income", "Interest Income", "INCOME", special=True),
        m.Account("Expense:TDS on Interest", "TDS on Interest", "EXPENSE"),
        m.Account("Liabilities:Suspense", "Suspense", "LIABILITY"),
    ]
    d = _deductor(1, "SOME OBSCURE PAYER", "194A", 10000.0, 1000.0)
    journals = m.build_journals([d], accts)
    j = journals[0]
    assert j.credit_account == "Liabilities:Suspense"
    assert j.needs_review is True
    assert j.balanced


def test_generic_interest_on_fd_never_a_credit_match():
    """The generic FD-interest account must never be returned as a credit match."""
    fd = m.find_generic_fd_account(_accounts())
    for name in ("BANK OF BARODA", "SOME RANDOM FD HOLDER"):
        acct, _c, _b, cands = m.match_credit_account(name, "A", _accounts(), fd)
        assert acct != fd
        assert fd not in cands


def test_find_generic_fd_account_variants():
    """The generic FD account is fuzzy-found regardless of its exact name, and a
    deductor-specific '... - FD' account is NOT mistaken for the generic one."""
    def acc(p):
        return m.Account(p, p.split(":")[-1], "INCOME")
    # Harshal-style
    assert m.find_generic_fd_account([acc("Income:Interest Income:Interest on FD"),
                                      acc("Income:Interest Income:Interest on BOB - FD")]) \
        == "Income:Interest Income:Interest on FD"
    # Vaikunth-style ('Interest on Fixed Deposit') + specific accounts present
    fd = m.find_generic_fd_account([
        acc("Income:Interest Income:Interest on Fixed Deposit"),
        acc("Income:Interest Income:Interest on HDFC - FD"),
        acc("Income:Interest Income:Interest on ICICI Bank - FD"),
    ])
    assert fd == "Income:Interest Income:Interest on Fixed Deposit"
    # No FD account at all -> None (caller falls back to the canonical name)
    assert m.find_generic_fd_account([acc("Income:Interest Income:Interest on Bonds")]) is None


def test_category_a_debit_uses_fuzzy_found_fd_account():
    """Category A's second debit posts to the chart's actual FD account, even
    when it isn't literally named 'Interest on FD'."""
    def acc(p, t="INCOME"):
        return m.Account(p, p.split(":")[-1], t)
    accts = [
        acc("Income:Interest Income:Interest on Fixed Deposit"),
        acc("Income:Interest Income:Interest on HDFC - FD"),
        acc("Expense:TDS on Interest", "EXPENSE"),
        acc("Liabilities:Suspense", "LIABILITY"),
    ]
    d = _deductor(2, "HDFC BANK LIMITED", "193", 18023.06, 0.0)
    j = m.build_journals([d], accts)[0]
    debit_accts = {s.account for s in j.splits if s.debit}
    assert "Income:Interest Income:Interest on Fixed Deposit" in debit_accts
    assert j.credit_account == "Income:Interest Income:Interest on HDFC - FD"
    assert j.balanced


# ---------------------------------------------------------------------------
# Split construction + balancing
# ---------------------------------------------------------------------------

def test_category_a_three_splits():
    d = _deductor(4, "BANK OF BARODA", "194A", 250237, 25024)
    j = m.build_journals([d], _accounts())[0]
    assert j.category == "A" and len(j.splits) == 3
    accts = {s.account: (s.debit, s.credit) for s in j.splits}
    assert accts[m.ACC_TDS_INTEREST] == (25024, 0)
    assert accts[m.ACC_INTEREST_ON_FD] == (round(250237 - 25024, 2), 0)
    assert accts["Income:Interest Income:Interest on BOB - FD"] == (0, 250237)
    assert j.balanced


def test_category_b_two_splits():
    d = _deductor(2, "DR REDDY'S LABORATORIES LTD.", "194", 208000, 20800)
    j = m.build_journals([d], _accounts())[0]
    assert j.category == "B" and len(j.splits) == 2
    accts = {s.account: (s.debit, s.credit) for s in j.splits}
    assert accts[m.ACC_TDS_DIVIDEND] == (20800, 0)
    assert accts["Income:Dividend - Shares:Dividend - DRL"] == (0, 20800)
    assert j.balanced


def test_category_c_emits_partnership_tds_as_is():
    d = _deductor(8, "ACME CONSULTING LLP", "194T", 3656276, 365628)
    j = m.build_journals([d], _accounts())[0]
    assert j.category == "C" and len(j.splits) == 2
    accts = {s.account: (s.debit, s.credit) for s in j.splits}
    assert accts[m.ACC_TDS_PARTNERSHIP] == (365628, 0)
    assert j.balanced


def test_all_sample_journals_balanced():
    deds = [
        _deductor(1, "CHOLAMANDALAM INVESTMENT AND FINANCE COMPANY LIMITED", "193", 45750, 4575),
        _deductor(2, "DR REDDY'S LABORATORIES LTD.", "194", 208000, 20800),
        _deductor(4, "BANK OF BARODA", "194A", 250237, 25024),
        _deductor(8, "ACME CONSULTING LLP", "194T", 3656276, 365628),
    ]
    for j in m.build_journals(deds, _accounts()):
        assert j.balanced, j.deductor


# ---------------------------------------------------------------------------
# Category G -- 15G/15H (Part II)
#
# The 15G/15H interest is already booked in a generic FD-interest bucket --
# this journal is a RECLASSIFICATION (Dr generic FD interest, Cr the specific
# NBFC account), not new income. With tax_deducted == 0 (the statutory case)
# that's a 2-split; if tax was ever withheld anyway it degenerates to the
# identical 3-split Category A uses.
# ---------------------------------------------------------------------------

def test_category_g_zero_tax_two_split_reclass():
    d = _deductor(1, "BANK OF BARODA", "194A", 50000, 0)
    j = m.build_15g_journals([d], _accounts())[0]
    assert j.category == "G" and len(j.splits) == 2
    accts = {s.account: (s.debit, s.credit) for s in j.splits}
    assert accts[m.ACC_INTEREST_ON_FD] == (50000, 0)
    assert accts["Income:Interest Income:Interest on BOB - FD"] == (0, 50000)
    assert j.balanced


def test_category_g_nonzero_tax_three_split_like_category_a():
    """Statutory tax on a 15G/15H deductor 'can't be' non-zero, but if the
    26AS ever shows one, it must post exactly like Category A's 3-split."""
    d = _deductor(1, "BANK OF BARODA", "194A", 250237, 25024)
    j = m.build_15g_journals([d], _accounts())[0]
    assert j.category == "G" and len(j.splits) == 3
    accts = {s.account: (s.debit, s.credit) for s in j.splits}
    assert accts[m.ACC_TDS_INTEREST] == (25024, 0)
    assert accts[m.ACC_INTEREST_ON_FD] == (round(250237 - 25024, 2), 0)
    assert accts["Income:Interest Income:Interest on BOB - FD"] == (0, 250237)
    assert j.balanced


def test_category_g_never_collides_with_category_a_tdsj_series():
    """A Part II 194A deductor must NOT become Category A / TDSJ -- category
    is decided by which Part the row came from, never by a section lookup."""
    part_i = _deductor(1, "BANK OF BARODA", "194A", 1000, 100)
    part_ii = _deductor(1, "BANK OF BARODA", "194A", 50000, 0)  # same Sr, same section
    journals = m.build_journals([part_i], _accounts())
    journals += m.build_15g_journals([part_ii], _accounts())
    assert [j.category for j in journals] == ["A", "G"]

    out = Path(tempfile.gettempdir()) / "test_category_g_ids.csv"
    m.write_csv(journals, out, "2025-26")
    txns = {}
    with out.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            txns.setdefault(row["Transaction ID"], 0.0)
            txns[row["Transaction ID"]] += float(row["Amount"])
    assert set(txns) == {"2526-TDSJ01", "2526-15GJ01"}, \
        "Category G must get its own 15GJ series, never fold into TDSJ"
    for tid, total in txns.items():
        assert abs(total) < 0.01, f"{tid} does not sum to zero: {total}"


def test_category_g_empty_part_ii_is_a_noop():
    """No 15G/15H deductors (Part II absent or empty) must not crash and must
    not emit any Category G journals."""
    assert m.build_15g_journals([], _accounts()) == []


def test_series_for_category_maps_g_to_15gj_and_rejects_unknown():
    assert m.series_for_category("G") == "15GJ"
    assert m.series_for_category("A") == "TDSJ"
    assert m.series_for_category("T") == "TCSJ"
    with pytest.raises(ValueError):
        m.series_for_category("Z")


# ---------------------------------------------------------------------------
# Overrides + unknown section
# ---------------------------------------------------------------------------

def test_override_wins():
    d = _deductor(7, "OFFICE OF REGIONAL PROVIDENT FUND COMMISSIONER BANDRA EAST", "194A", 19380, 1938)
    j = m.build_journals([d], _accounts(), overrides={7: "Liabilities:Suspense"})[0]
    assert j.credit_account == "Liabilities:Suspense"
    assert j.credit_confidence == "Override" and not j.needs_review
    assert j.balanced


def test_unknown_section_flags_and_uses_suspense():
    d = _deductor(9, "MYSTERY CO", "194I", 1000, 100)
    j = m.build_journals([d], _accounts())[0]
    assert j.needs_review and j.credit_account == "Liabilities:Suspense"
    assert j.balanced


# ---------------------------------------------------------------------------
# Category T — TCS (Part VI)
#
# TCS is a tax credit like TDS: if it never reaches the books the return
# understates taxes paid and overstates the balance due. The journal moves the
# TAX only, out of the personal spending it was collected on.
# ---------------------------------------------------------------------------

def _tcs_accounts():
    return _accounts() + [
        m.Account("Expense:TCS on Foreign Trip", "TCS on Foreign Trip", "EXPENSE"),
        m.Account("Expense:Drawings", "Drawings", "EXPENSE"),
        m.Account("Assets:Current Assets:HDFC Bank", "HDFC Bank", "BANK"),
    ]


def _collector(sr, name, section, amt, tax):
    return m.Deductor(sr=sr, name=name, sections=(section,),
                      amount_paid=amt, tax_deducted=tax, tds_deposited=tax)


@pytest.mark.parametrize("section", ["206CQ", "206CR", "206CL", "206C"])
def test_tcs_sections_recognised(section):
    assert m.is_tcs_section((section,))


def test_non_tcs_section_not_recognised():
    assert not m.is_tcs_section(("194A",))
    assert not m.is_tcs_section(())


def test_tcs_journal_posts_tax_only_dr_tcs_cr_drawings():
    c = _collector(1, "THOMAS COOK INDIA LIMITED", "206CQ", 500000, 25000)
    j = m.build_tcs_journals([c], _tcs_accounts())[0]
    assert j.category == "T" and len(j.splits) == 2
    accts = {s.account: (s.debit, s.credit) for s in j.splits}
    # The 500000 spend is already in the books — only the 25000 tax moves.
    assert accts["Expense:TCS on Foreign Trip"] == (25000, 0)
    assert accts["Expense:Drawings"] == (0, 25000)
    assert j.balanced


def test_tcs_accounts_discovered_from_chart_not_hardcoded():
    """A book that names the account differently must still be found."""
    accts = _accounts() + [
        m.Account("Assets:TCS Receivable", "TCS Receivable", "ASSET"),
        m.Account("Equity:Drawings", "Drawings", "EQUITY"),
    ]
    j = m.build_tcs_journals([_collector(1, "X TOURS", "206CQ", 100, 10)], accts)[0]
    paths = {s.account for s in j.splits}
    assert paths == {"Assets:TCS Receivable", "Equity:Drawings"}


def test_tcs_falls_back_to_canonical_names_when_chart_has_neither():
    """Missing accounts must surface as names to CREATE, not vanish."""
    j = m.build_tcs_journals([_collector(1, "X TOURS", "206CQ", 100, 10)],
                             _accounts())[0]
    paths = {s.account for s in j.splits}
    assert paths == {m.ACC_TCS_DEFAULT, m.ACC_DRAWINGS}


def test_tcs_credit_account_is_configurable():
    """TCS paid across separately credits Bank, not Drawings."""
    bank = "Assets:Current Assets:HDFC Bank"
    j = m.build_tcs_journals([_collector(1, "X TOURS", "206CQ", 100, 10)],
                             _tcs_accounts(), credit_account=bank)[0]
    assert j.credit_account == bank
    assert {s.account for s in j.splits} == {"Expense:TCS on Foreign Trip", bank}
    assert j.balanced


def test_tcs_override_wins():
    j = m.build_tcs_journals([_collector(3, "X TOURS", "206CQ", 100, 10)],
                             _tcs_accounts(),
                             overrides={3: "Liabilities:Suspense"})[0]
    assert j.credit_account == "Liabilities:Suspense"
    assert j.credit_confidence == "Override" and not j.needs_review


def test_non_206c_section_in_part_vi_goes_suspense():
    """A section we can't justify as TCS must not silently claim a tax credit."""
    j = m.build_tcs_journals([_collector(1, "ODD CO", "194A", 100, 10)],
                             _tcs_accounts())[0]
    assert j.needs_review and j.credit_account == "Liabilities:Suspense"
    assert j.balanced


def test_tcs_never_uses_the_income_matcher():
    """match_credit_account searches INCOME subtrees; a collection at source
    has no income leg, so no TCS split may land on one."""
    j = m.build_tcs_journals([_collector(1, "BANK OF BARODA", "206CQ", 100, 10)],
                             _tcs_accounts())[0]
    income = {a.path for a in _tcs_accounts() if a.type == "INCOME"}
    assert not ({s.account for s in j.splits} & income)


def _party_sheet(ws, title, rows):
    """Write a Convert-shaped Part I / Part VI sheet: title band, meta strip,
    header row, then one row per transaction with the party's header totals
    repeated in cols 2/4/5/6 and the section in col 8."""
    ws.cell(1, 1, f"{title} - Details")
    ws.cell(2, 1, "Assessee Name: X  |  PAN: AAAAA1111A  |  Financial Year: 2025-26")
    ws.cell(3, 1, "Sr.No.")
    r = 4
    for sr, name, section, amt, tax in rows:
        ws.cell(r, 1, sr)
        ws.cell(r, 2, name)
        ws.cell(r, 4, amt)
        ws.cell(r, 5, tax)
        ws.cell(r, 6, tax)
        ws.cell(r, 8, section)
        r += 1


def _make_workbook(path, parts):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for title, rows in parts.items():
        _party_sheet(wb.create_sheet(title=title), title, rows)
    wb.save(path)
    return path


def test_parse_parts_reads_both_tds_and_tcs():
    p = Path(tempfile.gettempdir()) / "test_26as_both.xlsx"
    _make_workbook(p, {
        "Part I": [(1, "BANK OF BARODA", "194A", 1000.0, 100.0)],
        "Part VI": [(1, "THOMAS COOK INDIA LIMITED", "206CQ", 500000.0, 25000.0)],
    })
    deductors, g_deductors, collectors, fy = m.parse_parts(p)
    assert fy == "2025-26"
    assert [d.name for d in deductors] == ["BANK OF BARODA"]
    assert g_deductors == []
    assert [c.name for c in collectors] == ["THOMAS COOK INDIA LIMITED"]
    assert collectors[0].tax_deducted == 25000.0
    assert collectors[0].sections == ("206CQ",)


def test_parse_parts_accepts_tcs_only_workbook():
    """A 26AS with TCS but no TDS is valid and must not be rejected."""
    p = Path(tempfile.gettempdir()) / "test_26as_tcs_only.xlsx"
    _make_workbook(p, {"Part VI": [(1, "X TOURS", "206CQ", 100.0, 10.0)]})
    deductors, g_deductors, collectors, fy = m.parse_parts(p)
    assert deductors == [] and g_deductors == [] and len(collectors) == 1
    assert fy == "2025-26"


def test_parse_parts_reads_part_ii_15g_15h():
    """A 26AS with only 15G/15H (Part II) deductors is valid — no TDS/TCS
    required — and Part II shares columns 1/2/4/5/6/8 with Part I/VI."""
    p = Path(tempfile.gettempdir()) / "test_26as_part_ii_only.xlsx"
    _make_workbook(p, {
        "Part II": [(1, "BAJAJ FINANCE LIMITED", "194A", 50000.0, 0.0)],
    })
    deductors, g_deductors, collectors, fy = m.parse_parts(p)
    assert deductors == [] and collectors == []
    assert [d.name for d in g_deductors] == ["BAJAJ FINANCE LIMITED"]
    assert g_deductors[0].amount_paid == 50000.0
    assert g_deductors[0].tax_deducted == 0.0
    assert fy == "2025-26"


def test_parse_parts_rejects_workbook_with_neither_part():
    p = Path(tempfile.gettempdir()) / "test_26as_neither.xlsx"
    _make_workbook(p, {"Part VII": []})
    with pytest.raises(ValueError, match="Part I"):
        m.parse_parts(p)


def test_tcs_and_tds_transaction_ids_never_collide():
    """Part I Sr.1 and Part VI Sr.1 are different parties — a shared ID would
    make GnuCash fuse their splits into one unbalanced transaction."""
    journals = m.build_journals([_deductor(1, "BANK OF BARODA", "194A", 1000, 100)],
                                _tcs_accounts())
    journals += m.build_tcs_journals([_collector(1, "X TOURS", "206CQ", 500, 50)],
                                     _tcs_accounts())
    out = Path(tempfile.gettempdir()) / "test_tcs_ids.csv"
    m.write_csv(journals, out, "2025-26")

    txns = {}
    with out.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            txns.setdefault(row["Transaction ID"], 0.0)
            txns[row["Transaction ID"]] += float(row["Amount"])
    assert set(txns) == {"2526-TDSJ01", "2526-TCSJ01"}
    for tid, total in txns.items():
        assert abs(total) < 0.01, f"{tid} does not sum to zero: {total}"


# ---------------------------------------------------------------------------
# Date + CSV output
# ---------------------------------------------------------------------------

def test_journal_date_is_march_31_current_year():
    assert m.journal_date() == f"{dt.date.today().year}-03-31"


def test_csv_roundtrips_and_balances():
    deds = [
        _deductor(4, "BANK OF BARODA", "194A", 250237, 25024),
        _deductor(2, "DR REDDY'S LABORATORIES LTD.", "194", 208000, 20800),
    ]
    journals = m.build_journals(deds, _accounts())
    out = Path(tempfile.gettempdir()) / "test_tds_journals.csv"
    m.write_csv(journals, out, "2025-26")

    # Re-parse: splits group by Transaction ID (repeated on every row); each
    # transaction balances when its signed Amount splits sum to zero. Debits
    # are positive, credits negative (GnuCash "Amount" convention).
    txns = {}
    with out.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            tid = row["Transaction ID"]
            assert tid.strip(), "every split row must carry the Transaction ID"
            assert tid.startswith("2526-"), f"Transaction ID needs FY prefix: {tid}"
            assert row["Date"].strip(), "every split row must carry the Date"
            assert row["Amount"].strip(), "every split row must carry an Amount"
            txns.setdefault(tid, 0.0)
            txns[tid] += float(row["Amount"])
    assert len(txns) == 2
    for tid, total in txns.items():
        assert abs(total) < 0.01, f"{tid} does not sum to zero: {total}"


# ---------------------------------------------------------------------------
# Optional end-to-end (only if fixtures provided — no PII committed)
# ---------------------------------------------------------------------------

_FX_XLSX = Path(__file__).resolve().parent / "fixtures" / "sample_26as.xlsx"
_FX_GNC = Path(__file__).resolve().parent / "fixtures" / "sample.gnucash"


@pytest.mark.skipif(not (_FX_XLSX.exists() and _FX_GNC.exists()),
                    reason="No fixtures at tests/fixtures/sample_26as.xlsx + sample.gnucash")
def test_end_to_end():
    out = Path(tempfile.gettempdir()) / "test_tds_e2e.csv"
    stats = m.run(_FX_XLSX, _FX_GNC, out)
    assert out.exists()
    assert stats["balanced_all"]
    assert stats["deductors"] >= 1


# ---------------------------------------------------------------------------
# apply_overrides graceful degradation
#
# A weak tool-calling model sometimes invokes apply_overrides with no arguments.
# When `overrides` is a REQUIRED tool parameter, strict endpoints (e.g. Groq)
# reject the whole request with HTTP 400 `tool_use_failed` before the tool body
# ever runs. These tests pin the fix: `overrides` is optional in the tool schema,
# and a None / empty payload normalizes to a harmless no-op.
# ---------------------------------------------------------------------------

TOOLS = SRC / "agents" / "skill_26as_journal" / "tools.py"


def _load_tools():
    if str(SRC) not in sys.path:
        sys.path.insert(0, str(SRC))
    spec = importlib.util.spec_from_file_location("t26as_tools", TOOLS)
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


tl = _load_tools()


@pytest.mark.parametrize("value", [None, "", "   ", {}])
def test_normalize_overrides_empty_is_noop(value):
    assert tl._normalize_overrides(value) == {}


def test_normalize_overrides_dict_passthrough_stringifies_keys():
    assert tl._normalize_overrides({2: "Income:Interest Income:Interest on FD"}) == {
        "2": "Income:Interest Income:Interest on FD"}


def test_normalize_overrides_drops_falsy_values():
    assert tl._normalize_overrides({"2": "Income:X", "3": "", "4": None}) == {"2": "Income:X"}


def test_normalize_overrides_parses_json_string():
    assert tl._normalize_overrides('{"2": "Income:X"}') == {"2": "Income:X"}


def test_normalize_overrides_extracts_number_from_label_keys():
    # A tool-calling model often echoes the display label ("Sr 7") instead of
    # the bare number. Normalize to the digits so the subprocess int() succeeds.
    assert tl._normalize_overrides({"Sr 7": "Income:X"}) == {"7": "Income:X"}
    assert tl._normalize_overrides({"sr7": "Income:X", " 2 ": "Income:Y"}) == {
        "7": "Income:X", "2": "Income:Y"}


def test_normalize_overrides_drops_keys_without_a_number():
    # No number in the key -> can't resolve to a deductor Sr, so drop it rather
    # than pass a key that would crash int() in the builder subprocess.
    assert tl._normalize_overrides({"total": "Income:X"}) == {}


def test_normalize_overrides_bad_string_returns_error():
    out = tl._normalize_overrides("not an object")
    assert isinstance(out, str) and out.startswith("ERROR")


def test_run_apply_none_is_noop_and_touches_nothing():
    # Dummy, nonexistent paths: run_apply must short-circuit on the empty
    # override BEFORE it ever reads the workbook or rebuilds the CSV.
    missing = str(Path(tempfile.gettempdir()) / "does-not-exist-26as.csv")
    out = tl.run_apply(missing, missing, missing, None)
    assert out == "No overrides supplied; the existing CSV is unchanged and valid."
    assert not Path(missing).exists()


def test_apply_overrides_tool_param_is_optional():
    """The real tool schema must NOT list `overrides` as required — that is what
    prevents strict endpoints from 400-ing an argument-less call."""
    from agents.skill_26as_journal.agent import _make_tools

    tools = {t.name: t for t in _make_tools("x.xlsx", "y.gnucash", "z.csv")}
    schema = tools["apply_overrides"].args_schema.model_json_schema()
    assert "overrides" in schema.get("properties", {}), "param must still exist"
    assert "overrides" not in schema.get("required", []), \
        "overrides must be optional so a no-arg tool call is accepted, not 400'd"
