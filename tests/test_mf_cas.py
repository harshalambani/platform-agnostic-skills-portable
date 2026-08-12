"""
tests/test_mf_cas.py -- agents.skill_mf_cas.

Two independent input shapes are exercised in this one file:

1. The OLD CAS-text shape (parser.py + lots.py, FIFO lot derivation) --
   kept intact and still fully tested below, even though agent.run() no
   longer uses it (see AGENT.md's "Kept-but-unused modules"). Synthetic
   text lines only.
2. The CURRENT KFintech Capital Gain / Loss Statement (.xlsx) shape
   (cg_parser.py + cg_writer.py, driven by agent.run()) -- synthetic
   openpyxl-built workbooks only, shaped like the real statement's four
   sheets but with entirely invented scheme names/folios/ISINs/amounts.

No real, encrypted, or password-protected xlsx fixture is ever generated or
committed -- xlsx-boundary tests monkeypatch msoffcrypto.OfficeFile with a
fake that treats an ordinary (unencrypted) openpyxl-built test workbook as
"the decrypted result", the same monkeypatch-the-library-entry-point
pattern the old PDF-boundary tests used for pdfplumber.

Covers (CAS-text/FIFO shape, unchanged):
  * multi-lot FIFO disposal splitting (one row per consumed lot)
  * partial-lot disposal (a lot only partially consumed)
  * a pre-2018-02-01 buy raises the grandfathering flag
  * a units-reconciliation breach is reported, not silently absorbed
  * a disposal reaching into a nonzero Opening Unit Balance is flagged
    "unattributed -- review", never guessed
  * switch-in / switch-out classification

Covers (KFintech Capital Gain / Loss Statement xlsx shape, new):
  * matched-pair row mapping (Section A acquisition + Section B outflow +
    Section C gain, joined by row) from Trasaction_Details
  * ISIN extraction from the scheme-name string
  * Summary sheet parsing by row-label text, not fixed row number
  * the five s.234C period buckets landing in the correct columns
  * Equity vs NonEquity separation feeding the three-way reconciliation
  * three-way reconciliation both agreeing and disagreeing, and the
    explicit CANNOT_RECONCILE case when a leg is absent
  * grandfathered cost values carried straight through, without ever
    raising lots.GRANDFATHER_FLAG (this path does no FIFO re-derivation)
  * a wrong statement password -> clean message, password never echoed
  * a workbook missing the expected sheets fails loud (never an empty
    0-rows report) -- the fix for the skill's original silent-empty defect
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from agents.skill_mf_cas import agent as mf_agent  # noqa: E402
from agents.skill_mf_cas import cg_parser  # noqa: E402
from agents.skill_mf_cas import lots  # noqa: E402
from agents.skill_mf_cas.parser import classify_row, parse_cas_text  # noqa: E402


def _folio_block(
    folio="TEST/001", amc="Test AMC Mutual Fund", isin="INF000A00123",
    scheme="Test Equity Growth Fund", rta="Test RTA", scheme_type="EQUITY",
    opening="0.0000", closing="0.0000", txn_lines=(), rta_gain=None,
):
    lines = [
        f"Folio No: {folio}",
        f"AMC: {amc}",
        f"Scheme: {scheme} (ISIN: {isin})",
        f"Registrar: {rta}",
        f"Scheme Type: {scheme_type}",
        f"Opening Unit Balance: {opening}",
        *txn_lines,
        f"Closing Unit Balance: {closing}",
    ]
    if rta_gain is not None:
        lines.append(f"RTA Realised Gain: {rta_gain}")
    return lines


def _txn(date, desc, amount, units, nav, balance):
    return f"{date} {desc} {amount} {units} {nav} {balance}"


# ---------------------------------------------------------------------------
# classify_row
# ---------------------------------------------------------------------------

def test_classify_row_switch_in_is_acquisition():
    assert classify_row("Switch in from Test Debt Fund", 50.0) == "ACQUISITION"


def test_classify_row_switch_out_is_disposal():
    assert classify_row("Switch out to Test Debt Fund", -50.0) == "DISPOSAL"


def test_classify_row_zero_units_is_neither():
    assert classify_row("IDCW Payout", 0.0) == "NEITHER"


def test_classify_row_dividend_reinvestment_is_acquisition():
    assert classify_row("Dividend Reinvestment", 5.25) == "ACQUISITION"


# ---------------------------------------------------------------------------
# parse_cas_text
# ---------------------------------------------------------------------------

def test_parse_basic_block_fields():
    lines = _folio_block(
        txn_lines=[_txn("01-Apr-2020", "Purchase", "10000.00", "100.0000", "100.00", "100.0000")],
        closing="100.0000",
    )
    blocks = parse_cas_text(lines)
    assert len(blocks) == 1
    b = blocks[0]
    assert b.folio == "TEST/001"
    assert b.isin == "INF000A00123"
    assert b.scheme_type == "EQUITY"
    assert len(b.transactions) == 1
    assert b.transactions[0].kind == "ACQUISITION"


def test_parse_never_captures_pan_or_name_lines():
    lines = [
        "Investor Name: Some Person",
        "PAN: ABCDE1234X",
        "Email: someone@example.com",
        "Mobile: 9999999999",
        *_folio_block(txn_lines=[
            _txn("01-Apr-2020", "Purchase", "10000.00", "100.0000", "100.00", "100.0000")
        ], closing="100.0000"),
    ]
    blocks = parse_cas_text(lines)
    assert len(blocks) == 1
    # None of the PII lines produced any field anywhere in the parsed block.
    dumped = repr(blocks[0])
    assert "Some Person" not in dumped
    assert "ABCDE1234X" not in dumped
    assert "someone@example.com" not in dumped
    assert "9999999999" not in dumped


# ---------------------------------------------------------------------------
# FIFO derivation
# ---------------------------------------------------------------------------

def test_multi_lot_disposal_splits_into_one_row_per_lot():
    lines = _folio_block(
        txn_lines=[
            _txn("01-Apr-2020", "Purchase", "10000.00", "100.0000", "100.00", "100.0000"),
            _txn("01-May-2020", "Purchase", "11000.00", "100.0000", "110.00", "200.0000"),
            # Redeem 150 units -- must draw 100 from lot 1, 50 from lot 2.
            _txn("01-Jun-2020", "Redemption", "-18000.00", "-150.0000", "120.00", "50.0000"),
        ],
        closing="50.0000",
    )
    blocks = parse_cas_text(lines)
    recon = lots.derive_scheme(blocks[0])

    assert len(recon.disposal_lots) == 2
    units_taken = sorted(dl.units for dl in recon.disposal_lots)
    assert units_taken == [50.0, 100.0]
    # FIFO order: first lot (100 units @ NAV 100) consumed first and fully.
    first = next(dl for dl in recon.disposal_lots if dl.units == 100.0)
    assert first.buy_nav == 100.0
    second = next(dl for dl in recon.disposal_lots if dl.units == 50.0)
    assert second.buy_nav == 110.0
    assert recon.units_ok
    assert recon.matched_vs_disposed_ok


def test_partial_lot_disposal_leaves_remainder_in_lot():
    lines = _folio_block(
        txn_lines=[
            _txn("01-Apr-2020", "Purchase", "10000.00", "100.0000", "100.00", "100.0000"),
            # Redeem only 40 of the 100 units in the single lot.
            _txn("01-Jun-2020", "Redemption", "-4800.00", "-40.0000", "120.00", "60.0000"),
        ],
        closing="60.0000",
    )
    blocks = parse_cas_text(lines)
    recon = lots.derive_scheme(blocks[0])

    assert len(recon.disposal_lots) == 1
    dl = recon.disposal_lots[0]
    assert dl.units == 40.0
    assert dl.buy_cost == pytest.approx(4000.00)  # 40/100 * 10000
    assert dl.sale_proceeds == pytest.approx(4800.00)
    assert dl.gain == pytest.approx(800.00)
    assert recon.units_ok


def test_grandfathering_flag_on_pre_2018_buy():
    lines = _folio_block(
        txn_lines=[
            _txn("15-Jan-2017", "Purchase", "5000.00", "50.0000", "100.00", "50.0000"),
            _txn("01-Jun-2020", "Redemption", "-6000.00", "-50.0000", "120.00", "0.0000"),
        ],
        closing="0.0000",
    )
    blocks = parse_cas_text(lines)
    recon = lots.derive_scheme(blocks[0])

    assert len(recon.disposal_lots) == 1
    assert lots.GRANDFATHER_FLAG in recon.disposal_lots[0].flags


def test_post_2018_buy_has_no_grandfathering_flag():
    lines = _folio_block(
        txn_lines=[
            _txn("15-Mar-2018", "Purchase", "5000.00", "50.0000", "100.00", "50.0000"),
            _txn("01-Jun-2020", "Redemption", "-6000.00", "-50.0000", "120.00", "0.0000"),
        ],
        closing="0.0000",
    )
    blocks = parse_cas_text(lines)
    recon = lots.derive_scheme(blocks[0])

    assert lots.GRANDFATHER_FLAG not in recon.disposal_lots[0].flags


def test_units_reconciliation_breach_is_reported():
    # Opening 0 + acquired 100 - disposed 40 = 60, but statement claims 55.
    lines = _folio_block(
        txn_lines=[
            _txn("01-Apr-2020", "Purchase", "10000.00", "100.0000", "100.00", "100.0000"),
            _txn("01-Jun-2020", "Redemption", "-4800.00", "-40.0000", "120.00", "60.0000"),
        ],
        closing="55.0000",  # deliberately wrong
    )
    blocks = parse_cas_text(lines)
    recon = lots.derive_scheme(blocks[0])

    assert not recon.units_ok
    assert recon.units_diff == pytest.approx(5.0)


def test_unattributed_flag_when_disposal_draws_on_opening_balance():
    # Opening balance of 30 units with no visible acquisition in this
    # statement's window; a 30-unit redemption must draw entirely on it.
    lines = _folio_block(
        opening="30.0000",
        txn_lines=[
            _txn("01-Jun-2020", "Redemption", "-3600.00", "-30.0000", "120.00", "0.0000"),
        ],
        closing="0.0000",
    )
    blocks = parse_cas_text(lines)
    recon = lots.derive_scheme(blocks[0])

    assert len(recon.disposal_lots) == 1
    dl = recon.disposal_lots[0]
    assert lots.UNATTRIBUTED in dl.flags
    assert dl.buy_date is None
    assert dl.buy_cost is None
    assert dl.gain is None  # never guessed
    assert recon.units_ok  # 30 opening - 30 disposed == 0 closing, still balances


def test_unattributed_flag_never_fabricates_cost_or_date():
    lines = _folio_block(
        opening="30.0000",
        txn_lines=[_txn("01-Jun-2020", "Redemption", "-3600.00", "-30.0000", "120.00", "0.0000")],
        closing="0.0000",
    )
    blocks = parse_cas_text(lines)
    recon = lots.derive_scheme(blocks[0])
    dl = recon.disposal_lots[0]
    # Sale proceeds/units are known from the statement row itself, but
    # anything requiring the (unknown) buy side must stay None, not guessed.
    assert dl.sale_proceeds == pytest.approx(3600.00)
    assert dl.units == 30.0
    assert dl.buy_nav is None


def test_rta_gain_variance_reported_when_present():
    lines = _folio_block(
        txn_lines=[
            _txn("01-Apr-2020", "Purchase", "10000.00", "100.0000", "100.00", "100.0000"),
            _txn("01-Jun-2020", "Redemption", "-12000.00", "-100.0000", "120.00", "0.0000"),
        ],
        closing="0.0000",
        rta_gain="1500.00",  # FIFO derives 2000.00 -- deliberate variance
    )
    blocks = parse_cas_text(lines)
    recon = lots.derive_scheme(blocks[0])

    assert recon.rta_gain_reported == pytest.approx(1500.00)
    assert recon.rta_gain_derived == pytest.approx(2000.00)
    assert recon.rta_gain_variance == pytest.approx(500.00)


def test_no_rta_gain_line_leaves_reported_as_none():
    lines = _folio_block(
        txn_lines=[_txn("01-Apr-2020", "Purchase", "10000.00", "100.0000", "100.00", "100.0000")],
        closing="100.0000",
    )
    blocks = parse_cas_text(lines)
    recon = lots.derive_scheme(blocks[0])
    assert recon.rta_gain_reported is None
    assert recon.rta_gain_variance is None


# ---------------------------------------------------------------------------
# Review-finding regression tests (PR #176 follow-up: matched_vs_disposed_ok
# was decorative, rta_gain_derived treated unknown gain as zero, and a
# DISPOSAL row could silently vanish if its sign guarantee ever broke).
# ---------------------------------------------------------------------------

def test_shortfall_disposal_makes_matched_vs_disposed_ok_false():
    # Only 50 units ever exist (one 50-unit lot), but the statement's own
    # redemption row claims 80 units sold -- a 30-unit phantom shortfall
    # that exceeds every available lot, including the (absent) opening
    # balance. Under the OLD code, the phantom top-up used to cover this
    # shortfall was folded straight into total_matched_units, which forced
    # matched_vs_disposed_ok to True for every possible input -- the
    # invariant could never fail. It must now be able to report False.
    lines = _folio_block(
        txn_lines=[
            _txn("01-Apr-2020", "Purchase", "5000.00", "50.0000", "100.00", "50.0000"),
            _txn("01-Jun-2020", "Redemption", "-9600.00", "-80.0000", "120.00", "-30.0000"),
        ],
        closing="-30.0000",
    )
    blocks = parse_cas_text(lines)
    recon = lots.derive_scheme(blocks[0])

    assert recon.unattributed_shortfall_units == pytest.approx(30.0)
    assert recon.matched_vs_disposed_ok is False


def test_all_unattributed_disposals_cannot_cross_check_rta_gain():
    # Every disposed unit here draws on the unknown-cost OPENING lot, so
    # nothing about this scheme's realised gain is actually derivable.
    # Under the OLD code, rta_gain_derived silently became 0.0 (treating
    # "unknown" as "zero gain"), and rta_gain_variance then read as a real
    # discrepancy against the RTA's stated figure -- a false signal about
    # genuinely undecidable input.
    lines = _folio_block(
        opening="40.0000",
        txn_lines=[
            _txn("01-Jun-2020", "Redemption", "-4800.00", "-40.0000", "120.00", "0.0000"),
        ],
        closing="0.0000",
        rta_gain="900.00",
    )
    blocks = parse_cas_text(lines)
    recon = lots.derive_scheme(blocks[0])

    assert lots.UNATTRIBUTED in recon.disposal_lots[0].flags
    assert recon.rta_gain_derived is None
    assert recon.rta_gain_variance is None
    assert recon.rta_gain_note == lots.CANNOT_CROSS_CHECK

    # And this must reach the rendered workbook -- a blank/absent Exceptions
    # row on undecidable input would misread as "no variance found".
    from openpyxl import Workbook

    from agents.skill_mf_cas import excel_writer

    wb = Workbook()
    wb.remove(wb.active)
    excel_writer._write_exceptions_sheet(wb, [recon])
    ws = wb["Exceptions"]
    details = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
    assert any(lots.CANNOT_CROSS_CHECK in (d or "") for d in details)


def test_disposal_with_nonnegative_units_is_flagged_not_dropped():
    # qty_needed = -txn.units (and the whole matching loop) relies on
    # classify_row only ever returning DISPOSAL for negative units. This
    # test builds a TxnRow directly (bypassing the parser, which today
    # cannot produce this shape) to prove that IF that guarantee ever broke
    # upstream, derive_scheme would not let the row vanish with no lot and
    # no flag -- under the OLD code, a non-negative qty_needed would break
    # the matching loop immediately (remaining_to_consume <= 0) AND fail
    # the shortfall check (remaining_to_consume > 0 also false), so the
    # transaction would disappear silently.
    from datetime import date as _date

    from agents.skill_mf_cas.parser import DISPOSAL, SchemeBlock, TxnRow

    txn = TxnRow(
        date=_date(2020, 6, 1), description="Malformed disposal row",
        amount=-1000.0, units=25.0,  # non-negative units on a DISPOSAL row
        nav=100.0, balance=0.0, kind=DISPOSAL,
    )
    block = SchemeBlock(
        folio="TEST/999", amc="Test AMC", rta="Test RTA",
        scheme_name="Test Fund", isin="INF000A00999", scheme_type="EQUITY",
        opening_units=25.0, closing_units=0.0, transactions=[txn],
    )
    recon = lots.derive_scheme(block)

    flagged = [dl for dl in recon.disposal_lots if lots.INVALID_DISPOSAL_SIGN in dl.flags]
    assert len(flagged) == 1
    assert flagged[0].units == 25.0


def test_switch_out_then_switch_in_different_scheme_each_folio():
    lines = [
        *_folio_block(
            folio="TEST/001", scheme="Source Fund", isin="INF000A00001",
            txn_lines=[
                _txn("01-Apr-2020", "Purchase", "10000.00", "100.0000", "100.00", "100.0000"),
                _txn("01-Jul-2020", "Switch out to Target Fund", "-11000.00", "-100.0000", "110.00", "0.0000"),
            ],
            closing="0.0000",
        ),
        *_folio_block(
            folio="TEST/002", scheme="Target Fund", isin="INF000A00002",
            txn_lines=[
                _txn("01-Jul-2020", "Switch in from Source Fund", "11000.00", "55.0000", "200.00", "55.0000"),
            ],
            closing="55.0000",
        ),
    ]
    blocks = parse_cas_text(lines)
    assert len(blocks) == 2
    source_recon = lots.derive_scheme(blocks[0])
    target_recon = lots.derive_scheme(blocks[1])

    assert len(source_recon.disposal_lots) == 1
    assert source_recon.disposal_lots[0].gain == pytest.approx(1000.00)
    assert target_recon.units_ok
    assert blocks[1].transactions[0].kind == "ACQUISITION"


# ---------------------------------------------------------------------------
# cg_parser -- synthetic KFintech Capital Gain / Loss Statement workbooks.
# Every scheme name / folio / ISIN / amount below is invented for this test
# file; none of it comes from any real statement.
# ---------------------------------------------------------------------------

_TXN_HEADERS = [
    "Fund", "Fund Name", "Folio Number", "Scheme Name",
    "Trxn.Type", "Date", "Current Units", "Source Scheme Units",
    "Original Purchase Cost", "Original Cost Amount",
    "Grandfathered NAV as on 31/01/2018", "Grandfathered Cost Value",
    "IT Applicable NAV", "IT Applicable Cost Value",
    "Trxn.Type", "Date", "Units", "Amount", "Price", "Tax Perc", "Tax",
    "Short Term", "Indexed Cost", "Long Term With Index", "Long Term Without Index",
]


def _cg_txn_details_sheet(wb, rows, sheet_name=cg_parser.SHEET_TXN_DETAILS):
    """rows: list of 25-tuples, in _TXN_HEADERS column order."""
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value="Holder Name Placeholder")
    ws.cell(row=2, column=1, value="PAN Placeholder")
    ws.cell(row=3, column=5, value="Section A :Subscriptions")
    ws.cell(row=3, column=15, value="Section B : Outflows")
    ws.cell(row=3, column=22, value="Section C : Gains/Losses")
    for i, h in enumerate(_TXN_HEADERS):
        ws.cell(row=4, column=i + 1, value=h)
    r = 5
    for row in rows:
        for i, v in enumerate(row):
            ws.cell(row=r, column=i + 1, value=v)
        r += 1
    return ws


def _cg_summary_sheet(wb, sheet_name, section_rows, header_row=3):
    """section_rows: {section -> [(label, [5 buckets], total), ...]}. Every
    section from cg_parser._SUMMARY_SECTIONS is always written as a marker
    row (even with zero field rows) so callers only need to supply the
    sections they care about; parse_summary_sheet still requires at least
    one field row per section, so tests that need a fully valid sheet must
    supply all three."""
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value="KFintech Capital Gain / Loss Statement")
    ws.cell(row=header_row, column=1, value="Summary Of Capital Gains")
    for i, b in enumerate(cg_parser.BUCKET_LABELS):
        ws.cell(row=header_row, column=2 + i, value=b)
    ws.cell(row=header_row, column=7, value="Total")
    r = header_row + 1
    for section in cg_parser._SUMMARY_SECTIONS:
        ws.cell(row=r, column=1, value=section)
        r += 1
        for label, buckets, total in section_rows.get(section, []):
            ws.cell(row=r, column=1, value=label)
            for i, v in enumerate(buckets):
                ws.cell(row=r, column=2 + i, value=v)
            ws.cell(row=r, column=7, value=total)
            r += 1
    return ws


def _cg_scheme_summary_sheet(wb, rows, total=None):
    """rows: list of (scheme_name_with_isin, count, outflow, net_value,
    fmv_nav, fmv, short_gain, lt_with, lt_without). total, if given, is the
    same 8-value tuple without the scheme-name column."""
    ws = wb.create_sheet(cg_parser.SHEET_SCHEME_SUMMARY)
    header_row = 3
    for i, h in enumerate(cg_parser._SCHEME_SUMMARY_HEADERS):
        # write the real KFintech label text, not the normalised match text
        ws.cell(row=header_row, column=i + 1, value=h)
    ws.cell(row=header_row, column=3, value="Outflow\nAmount")  # embedded-newline quirk
    r = header_row + 1
    for row in rows:
        for i, v in enumerate(row):
            ws.cell(row=r, column=i + 1, value=v)
        r += 1
    if total is not None:
        ws.cell(row=r, column=1, value="Total")
        for i, v in enumerate(total, start=2):
            ws.cell(row=r, column=i, value=v)
    return ws


def _build_cg_workbook(*, summary_rows=None, scheme_rows=None, scheme_total=None,
                        txn_rows=None, txn_sheet_name=cg_parser.SHEET_TXN_DETAILS):
    wb = Workbook()
    wb.remove(wb.active)
    summary_rows = summary_rows or {}
    _cg_summary_sheet(wb, cg_parser.SHEET_SUMMARY_EQUITY, summary_rows.get("equity", {}))
    _cg_summary_sheet(wb, cg_parser.SHEET_SUMMARY_NONEQUITY, summary_rows.get("nonequity", {}))
    _cg_scheme_summary_sheet(wb, scheme_rows or [], scheme_total)
    _cg_txn_details_sheet(wb, txn_rows or [], sheet_name=txn_sheet_name)
    return wb


def _all_three_sections_minimal():
    """A section_rows dict with one minimal field row per section, all
    buckets/total zero -- enough to satisfy parse_summary_sheet's
    "all three sections present" requirement without contributing to any
    reconciliation total."""
    return {
        s: [(cg_parser._SECTION_GAIN_ROW_LABEL[s], [0.0] * 5, 0.0)]
        for s in cg_parser._SUMMARY_SECTIONS
    }


def test_extract_isin_from_scheme_name():
    name, isin = cg_parser.extract_isin(
        "Test Debt Short Term Fund - Growth Plan (INF999Z00111)"
    )
    assert name == "Test Debt Short Term Fund - Growth Plan"
    assert isin == "INF999Z00111"


def test_extract_isin_missing_returns_none():
    name, isin = cg_parser.extract_isin("Test Fund With No ISIN Suffix")
    assert name == "Test Fund With No ISIN Suffix"
    assert isin is None


def test_parse_transaction_details_matched_pair_and_isin():
    wb = Workbook()
    wb.remove(wb.active)
    ws = _cg_txn_details_sheet(wb, [
        (
            "F1", "Test Fund House", "FOLIO/CG/001",
            "Test Equity Growth Fund - Regular Plan (INF000A00123)",
            "Purchase", date(2019, 5, 10), 100.0, 100.0, 100.0, 10000.0,
            105.0, 10500.0, 108.0, 10800.0,
            "Redemption", date(2021, 6, 15), 100.0, 13000.0, 130.0, 0.0, 0.0,
            500.0, None, 0.0, 0.0,
        ),
    ])
    rows = cg_parser.parse_transaction_details(ws)
    assert len(rows) == 1
    r = rows[0]
    assert r.folio == "FOLIO/CG/001"
    assert r.scheme_name == "Test Equity Growth Fund - Regular Plan"
    assert r.isin == "INF000A00123"
    assert r.acq_trxn_type == "Purchase"
    assert r.acq_date == date(2019, 5, 10)
    assert r.outflow_trxn_type == "Redemption"
    assert r.outflow_date == date(2021, 6, 15)
    assert r.units == 100.0
    assert r.amount == 13000.0
    assert r.short_term == 500.0
    assert r.long_term_with_index == 0.0
    assert r.grandfathered_nav == 105.0
    assert r.grandfathered_cost_value == 10500.0


def test_parse_transaction_details_skips_pii_and_blank_rows():
    wb = Workbook()
    wb.remove(wb.active)
    ws = _cg_txn_details_sheet(wb, [
        (
            "F1", "Test Fund House", "FOLIO/CG/002",
            "Test Debt Fund - Direct Plan (INF000A00456)",
            "Purchase", date(2020, 1, 1), 50.0, 50.0, 20.0, 1000.0,
            None, None, None, None,
            "Redemption", date(2022, 1, 1), 50.0, 1200.0, 24.0, 0.0, 0.0,
            0.0, None, 200.0, 0.0,
        ),
        (None,) * 25,  # a trailing blank row must be skipped, not crash
    ])
    rows = cg_parser.parse_transaction_details(ws)
    assert len(rows) == 1
    # rows 1-2 (holder name / PAN placeholders) must never surface anywhere
    holder_row_value = ws.cell(row=1, column=1).value
    assert holder_row_value not in (rows[0].fund_code, rows[0].fund_name, rows[0].folio)


def test_summary_sheet_header_located_by_label_not_fixed_row_number():
    """Same content, header row shifted from 3 to 6 -- proves the header is
    found by matching label text, not a hardcoded row offset."""
    wb = Workbook()
    wb.remove(wb.active)
    section_rows = _all_three_sections_minimal()
    section_rows[cg_parser._SECTION_ST] = [
        (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_ST], [100.0, 0, 0, 0, 0], 100.0),
    ]
    ws = _cg_summary_sheet(wb, cg_parser.SHEET_SUMMARY_EQUITY, section_rows, header_row=6)
    data = cg_parser.parse_summary_sheet(ws, "Equity")
    row = data.gain_loss_row(cg_parser._SECTION_ST)
    assert row is not None
    assert row.total == pytest.approx(100.0)


def test_summary_sheet_five_234c_buckets_map_correctly():
    wb = Workbook()
    wb.remove(wb.active)
    buckets = [10.0, 20.0, 30.0, 40.0, 50.0]
    section_rows = _all_three_sections_minimal()
    section_rows[cg_parser._SECTION_ST] = [
        (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_ST], buckets, 150.0),
    ]
    ws = _cg_summary_sheet(wb, cg_parser.SHEET_SUMMARY_EQUITY, section_rows)
    data = cg_parser.parse_summary_sheet(ws, "Equity")
    row = data.gain_loss_row(cg_parser._SECTION_ST)
    for label, expected in zip(cg_parser.BUCKET_LABELS, buckets):
        assert row.buckets[label] == pytest.approx(expected)


def test_summary_sheet_missing_a_section_fails_loud():
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Bad")
    ws.cell(row=3, column=1, value="Summary Of Capital Gains")
    for i, b in enumerate(cg_parser.BUCKET_LABELS):
        ws.cell(row=3, column=2 + i, value=b)
    ws.cell(row=3, column=7, value="Total")
    ws.cell(row=4, column=1, value=cg_parser._SECTION_ST)
    ws.cell(row=5, column=1, value="Short Term Capital Gain/Loss")
    for i in range(5):
        ws.cell(row=5, column=2 + i, value=0.0)
    ws.cell(row=5, column=7, value=0.0)
    # Long Term sections deliberately omitted.
    with pytest.raises(cg_parser.CGParseError):
        cg_parser.parse_summary_sheet(ws, "Equity")


def _matched_row(short_term=0.0, lt_with=0.0, lt_without=0.0):
    return (
        "F1", "Test Fund House", "FOLIO/CG/900",
        "Test Hybrid Fund - Growth (INF000A00999)",
        "Purchase", date(2017, 3, 1), 10.0, 10.0, 50.0, 500.0,
        60.0, 600.0, 62.0, 620.0,
        "Redemption", date(2022, 3, 1), 10.0, 900.0, 90.0, 0.0, 0.0,
        short_term, None, lt_with, lt_without,
    )


def test_three_way_reconciliation_agrees_when_all_legs_match():
    wb = _build_cg_workbook(
        summary_rows={
            "equity": {
                cg_parser._SECTION_ST: [
                    (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_ST], [500.0, 0, 0, 0, 0], 500.0),
                ],
                cg_parser._SECTION_LT_WITH_INDEX: [
                    (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_LT_WITH_INDEX], [200.0, 0, 0, 0, 0], 200.0),
                ],
                cg_parser._SECTION_LT_WITHOUT_INDEX: [
                    (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_LT_WITHOUT_INDEX], [0, 0, 0, 0, 0], 0.0),
                ],
            },
            "nonequity": {
                cg_parser._SECTION_ST: [
                    (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_ST], [0, 0, 0, 0, 0], 0.0),
                ],
                cg_parser._SECTION_LT_WITH_INDEX: [
                    (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_LT_WITH_INDEX], [0, 0, 0, 0, 0], 0.0),
                ],
                cg_parser._SECTION_LT_WITHOUT_INDEX: [
                    (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_LT_WITHOUT_INDEX], [100.0, 0, 0, 0, 0], 100.0),
                ],
            },
        },
        scheme_total=(9.0, 0.0, 0.0, None, None, 500.0, 200.0, 100.0),
        txn_rows=[_matched_row(short_term=500.0, lt_with=200.0, lt_without=100.0)],
    )
    summary_equity = cg_parser.parse_summary_sheet(cg_parser.find_sheet(wb, cg_parser.SHEET_SUMMARY_EQUITY), "Equity")
    summary_nonequity = cg_parser.parse_summary_sheet(cg_parser.find_sheet(wb, cg_parser.SHEET_SUMMARY_NONEQUITY), "NonEquity")
    scheme_rows, scheme_total = cg_parser.parse_scheme_level_summary(cg_parser.find_sheet(wb, cg_parser.SHEET_SCHEME_SUMMARY))
    txn_rows = cg_parser.parse_transaction_details(cg_parser.find_sheet(wb, cg_parser.SHEET_TXN_DETAILS))

    results = cg_parser.reconcile(summary_equity, summary_nonequity, scheme_rows, scheme_total, txn_rows)
    by_cat = {r.category: r for r in results}
    assert by_cat[cg_parser.CATEGORY_ST].agree is True
    assert by_cat[cg_parser.CATEGORY_LT_WITH_INDEX].agree is True
    assert by_cat[cg_parser.CATEGORY_LT_WITHOUT_INDEX].agree is True


def test_three_way_reconciliation_reports_variance():
    wb = _build_cg_workbook(
        summary_rows={
            "equity": {
                cg_parser._SECTION_ST: [
                    (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_ST], [500.0, 0, 0, 0, 0], 500.0),
                ],
                cg_parser._SECTION_LT_WITH_INDEX: [
                    (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_LT_WITH_INDEX], [0, 0, 0, 0, 0], 0.0),
                ],
                cg_parser._SECTION_LT_WITHOUT_INDEX: [
                    (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_LT_WITHOUT_INDEX], [0, 0, 0, 0, 0], 0.0),
                ],
            },
            "nonequity": {
                cg_parser._SECTION_ST: [
                    (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_ST], [0, 0, 0, 0, 0], 0.0),
                ],
                cg_parser._SECTION_LT_WITH_INDEX: [
                    (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_LT_WITH_INDEX], [0, 0, 0, 0, 0], 0.0),
                ],
                cg_parser._SECTION_LT_WITHOUT_INDEX: [
                    (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_LT_WITHOUT_INDEX], [0, 0, 0, 0, 0], 0.0),
                ],
            },
        },
        # Scheme_Level_Summary Total deliberately wrong vs. Trasaction_Details/Summary (500.0).
        scheme_total=(9.0, 0.0, 0.0, None, None, 400.0, 0.0, 0.0),
        txn_rows=[_matched_row(short_term=500.0)],
    )
    summary_equity = cg_parser.parse_summary_sheet(cg_parser.find_sheet(wb, cg_parser.SHEET_SUMMARY_EQUITY), "Equity")
    summary_nonequity = cg_parser.parse_summary_sheet(cg_parser.find_sheet(wb, cg_parser.SHEET_SUMMARY_NONEQUITY), "NonEquity")
    scheme_rows, scheme_total = cg_parser.parse_scheme_level_summary(cg_parser.find_sheet(wb, cg_parser.SHEET_SCHEME_SUMMARY))
    txn_rows = cg_parser.parse_transaction_details(cg_parser.find_sheet(wb, cg_parser.SHEET_TXN_DETAILS))

    results = cg_parser.reconcile(summary_equity, summary_nonequity, scheme_rows, scheme_total, txn_rows)
    by_cat = {r.category: r for r in results}
    st = by_cat[cg_parser.CATEGORY_ST]
    assert st.agree is False
    assert st.note is not None and "variance" in st.note.lower()


def test_three_way_reconciliation_cannot_reconcile_when_scheme_summary_empty():
    summary_equity = cg_parser.SummarySheetData(sheet_label="Equity", rows=[
        cg_parser.SummaryFieldRow(
            section=cg_parser._SECTION_ST,
            label=cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_ST],
            buckets={b: 0.0 for b in cg_parser.BUCKET_LABELS}, total=500.0,
        ),
        cg_parser.SummaryFieldRow(
            section=cg_parser._SECTION_LT_WITH_INDEX,
            label=cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_LT_WITH_INDEX],
            buckets={b: 0.0 for b in cg_parser.BUCKET_LABELS}, total=0.0,
        ),
        cg_parser.SummaryFieldRow(
            section=cg_parser._SECTION_LT_WITHOUT_INDEX,
            label=cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_LT_WITHOUT_INDEX],
            buckets={b: 0.0 for b in cg_parser.BUCKET_LABELS}, total=0.0,
        ),
    ])
    summary_nonequity = cg_parser.SummarySheetData(sheet_label="NonEquity", rows=[
        cg_parser.SummaryFieldRow(
            section=s, label=cg_parser._SECTION_GAIN_ROW_LABEL[s],
            buckets={b: 0.0 for b in cg_parser.BUCKET_LABELS}, total=0.0,
        )
        for s in cg_parser._SUMMARY_SECTIONS
    ])
    results = cg_parser.reconcile(summary_equity, summary_nonequity, [], None, [])
    by_cat = {r.category: r for r in results}
    st = by_cat[cg_parser.CATEGORY_ST]
    assert st.agree is None
    assert st.note is not None and cg_parser.CANNOT_RECONCILE in st.note


def test_workbook_missing_required_sheets_fails_loud():
    """The fix for the skill's original silent-empty defect: an
    unrecognised workbook must never produce a 0-schemes/0-transactions
    report -- it must raise, naming what was expected vs. found."""
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Not The Right Sheet At All")
    with pytest.raises(cg_parser.CGParseError) as excinfo:
        cg_parser.validate_workbook_shape(wb)
    message = str(excinfo.value)
    assert cg_parser.SHEET_SUMMARY_EQUITY in message
    assert "Not The Right Sheet At All" in message


# ---------------------------------------------------------------------------
# agent.py -- xlsx boundary (password / file-type rejection / loud failure /
# end-to-end). msoffcrypto.OfficeFile is monkeypatched with a fake that
# treats an ordinary openpyxl-built workbook as "the decrypted result" --
# no real encrypted fixture is ever generated or committed.
# ---------------------------------------------------------------------------

class _FakeOfficeFile:
    """Stand-in for msoffcrypto.OfficeFile: accepts exactly one password and
    "decrypts" by handing back the original (already-plaintext) bytes."""

    def __init__(self, f, expected_password="right-password"):
        self._f = f
        self._expected_password = expected_password

    def load_key(self, password="", verify_password=True):
        if password != self._expected_password:
            import msoffcrypto.exceptions as exc
            raise exc.InvalidKeyError("Failed to verify password")

    def decrypt(self, buf):
        self._f.seek(0)
        buf.write(self._f.read())


def test_wrong_password_error_never_echoes_password(monkeypatch, tmp_path):
    import msoffcrypto

    secret = "hunter2-secret-statement-password"
    monkeypatch.setattr(
        msoffcrypto, "OfficeFile",
        lambda f: _FakeOfficeFile(f, expected_password=secret),
    )

    xlsx_path = tmp_path / "fake_statement.xlsx"
    Workbook().save(xlsx_path)  # contents never matter -- load_key raises first

    with pytest.raises(ValueError) as excinfo:
        mf_agent._load_workbook(str(xlsx_path), "wrong-password")

    message = str(excinfo.value)
    assert secret not in message
    assert "password" in message.lower()


def test_run_rejects_pdf_input(tmp_path):
    pdf_path = tmp_path / "statement.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 not a real pdf")
    out_path = tmp_path / "out.xlsx"

    summary = mf_agent.run(str(pdf_path), str(out_path), cas_password="whatever")

    assert "ERROR" in summary
    assert "xlsx" in summary.lower()
    assert not out_path.exists()


def test_run_fails_loud_on_unrecognised_workbook(monkeypatch, tmp_path):
    import msoffcrypto

    monkeypatch.setattr(
        msoffcrypto, "OfficeFile",
        lambda f: _FakeOfficeFile(f, expected_password="pw"),
    )

    bad_wb = Workbook()
    bad_wb.active.title = "Nothing Recognisable"
    xlsx_path = tmp_path / "bad.xlsx"
    bad_wb.save(xlsx_path)
    out_path = tmp_path / "out.xlsx"

    summary = mf_agent.run(str(xlsx_path), str(out_path), cas_password="pw")

    assert "ERROR" in summary
    assert cg_parser.SHEET_SUMMARY_EQUITY in summary
    assert not out_path.exists()


def test_run_end_to_end_writes_workbook_reconciliation_and_csv_sidecar(monkeypatch, tmp_path):
    import msoffcrypto

    monkeypatch.setattr(
        msoffcrypto, "OfficeFile",
        lambda f: _FakeOfficeFile(f, expected_password="pw"),
    )

    # A pre-2018-02-01 acquisition with RTA-supplied grandfathering values --
    # this path must carry them through as-is and never touch lots.py, so
    # lots.GRANDFATHER_FLAG must never appear anywhere in the output.
    minimal_sections = _all_three_sections_minimal()
    minimal_sections[cg_parser._SECTION_ST] = [
        (cg_parser._SECTION_GAIN_ROW_LABEL[cg_parser._SECTION_ST], [500.0, 0, 0, 0, 0], 500.0),
    ]
    src_wb = _build_cg_workbook(
        summary_rows={"equity": minimal_sections, "nonequity": _all_three_sections_minimal()},
        scheme_total=(9.0, 0.0, 0.0, None, None, 500.0, 0.0, 0.0),
        txn_rows=[
            (
                "F1", "Test Fund House", "FOLIO/CG/777",
                "Test Grandfathered Fund - Growth (INF000A00777)",
                "Purchase", date(2016, 6, 1), 10.0, 10.0, 50.0, 500.0,
                62.5, 625.0, 65.0, 650.0,
                "Redemption", date(2022, 3, 1), 10.0, 1000.0, 100.0, 0.0, 0.0,
                500.0, None, 0.0, 0.0,
            ),
        ],
    )
    src_path = tmp_path / "source.xlsx"
    src_wb.save(src_path)

    out_path = tmp_path / "report.xlsx"
    summary = mf_agent.run(str(src_path), str(out_path), cas_password="pw")

    assert "ERROR" not in summary
    assert out_path.is_file()
    csv_path = tmp_path / "report_matched_lots.csv"
    assert csv_path.is_file()

    import openpyxl
    report_wb = openpyxl.load_workbook(out_path)
    assert set(report_wb.sheetnames) == {
        "MatchedLots", "SchemeSummary", "PeriodSummary", "Reconciliation", "Exceptions",
    }

    matched_ws = report_wb["MatchedLots"]
    # Grandfathered NAV/Cost Value columns (12, 13 -- see cg_writer._MATCHED_LOT_HEADERS)
    # carry the RTA-supplied values through unchanged.
    header_row = [c.value for c in matched_ws[1]]
    gf_nav_col = header_row.index("Grandfathered NAV") + 1
    gf_cost_col = header_row.index("Grandfathered Cost Value") + 1
    assert matched_ws.cell(row=2, column=gf_nav_col).value == pytest.approx(62.5)
    assert matched_ws.cell(row=2, column=gf_cost_col).value == pytest.approx(625.0)

    recon_ws = report_wb["Reconciliation"]
    statuses = [recon_ws.cell(row=r, column=5).value for r in range(2, recon_ws.max_row + 1)]
    assert "AGREE" in statuses

    # This path must never invoke lots.py's FIFO/grandfathering-flag logic.
    for ws in report_wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                assert cell.value != lots.GRANDFATHER_FLAG
