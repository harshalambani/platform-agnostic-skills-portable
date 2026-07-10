#!/usr/bin/env python3
"""
excel_report.py -- render an intercompany reconciliation result (from
reconcile_intercompany.reconcile) to a formatted .xlsx workbook.

Sheets:
    Summary        -- owners, period, opening/movement/closing per side, tie diff
    Matched        -- paired FY movements, side by side
    Exceptions A   -- movements in A's contra a/c with no match in B (+ hunt)
    Exceptions B   -- movements in B's contra a/c with no match in A (+ hunt)
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Styling helpers.
# --------------------------------------------------------------------------- #
NUM = "#,##0.00;(#,##0.00)"      # Indian-friendly; negatives in parentheses
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUB_FONT = Font(italic=True, color="595959")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_row(ws, row, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _money(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = NUM
    cell.border = BORDER
    return cell


def _txt(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    return cell


def _fmt_suggestions(suggs) -> str:
    if not suggs:
        return "No candidate found in the other book (likely genuinely unrecorded)."
    lines = []
    for s in suggs:
        tag = "bank/cash" if s.liquid else "ledger"
        gap = "same day" if s.day_gap == 0 else f"{s.day_gap}d away"
        lines.append(f"{s.d} {s.amount:,.2f} -> {s.account_path} "
                     f"[{tag}, {gap}] {s.desc[:40]}")
    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# Sheets.
# --------------------------------------------------------------------------- #
def _summary_sheet(ws, r):
    a, b = r["book_a"], r["book_b"]
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Intercompany Reconciliation"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"{a.owner}  <->  {b.owner}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A3"] = (f"Period: {r['fy_label']}  ({r['fy_start']} to {r['fy_end']})   "
                f"|   date tolerance: +/-{r['tol_days']} days")
    ws["A3"].font = SUB_FONT
    ws["A4"] = (f"Book A: {a.path.name}   |   Book B: {b.path.name}")
    ws["A4"].font = SUB_FONT

    row = 6
    ws.cell(row=row, column=1, value="Contra accounts used").font = Font(bold=True)
    row += 1
    _txt(ws, row, 1, f"{a.owner}'s book")
    _txt(ws, row, 2, ", ".join(x.name for x in r["a_contra"]))
    row += 1
    _txt(ws, row, 1, f"{b.owner}'s book")
    _txt(ws, row, 2, ", ".join(x.name for x in r["b_contra"]))

    row += 2
    _header_row(ws, row, ["Balance movement", f"{a.owner} (Book A)",
                          f"{b.owner} (Book B)"],
                widths=[34, 22, 22])
    lines = [
        ("Opening balance b/f (before period)", r["a_open"], r["b_open"]),
        ("Movements during period", r["a_move_sum"], r["b_move_sum"]),
        ("Closing balance c/f", r["a_close"], r["b_close"]),
    ]
    for label, av, bv in lines:
        row += 1
        _txt(ws, row, 1, label)
        _money(ws, row, 2, round(av, 2))
        _money(ws, row, 3, round(bv, 2))
        if label.startswith("Closing"):
            for cc in (1, 2, 3):
                ws.cell(row=row, column=cc).font = Font(bold=True)

    row += 2
    ws.cell(row=row, column=1, value="Difference (Closing A + Closing B; a matched "
            "intercompany balance nets to 0)").font = Font(bold=True)
    diff_cell = _money(ws, row, 2, r["difference"])
    diff_cell.font = Font(bold=True)
    diff_cell.fill = OK_FILL if r["difference"] == 0 else BAD_FILL
    row += 1
    verdict = ("TIES OUT - the two books agree." if r["difference"] == 0
               else "OUT OF BALANCE - see exceptions below.")
    ws.cell(row=row, column=1, value=verdict).font = SUB_FONT

    row += 2
    _header_row(ws, row, ["Counts", "Value"], widths=[34, 22])
    for label, val in [
        ("Matched pairs", len(r["pairs"])),
        (f"Exceptions in {a.owner}'s book (A)", len(r["a_exc"])),
        (f"Exceptions in {b.owner}'s book (B)", len(r["b_exc"])),
    ]:
        row += 1
        _txt(ws, row, 1, label)
        c = ws.cell(row=row, column=2, value=val)
        c.border = BORDER


def _matched_sheet(ws, r):
    a, b = r["book_a"], r["book_b"]
    headers = ["Date (A)", f"Amount ({a.owner})", "Description (A)", "Account (A)",
               "Date (B)", f"Amount ({b.owner})", "Description (B)", "Account (B)",
               "Day gap", "Match basis"]
    _header_row(ws, 1, headers,
                widths=[12, 16, 34, 20, 12, 16, 34, 20, 9, 20])
    row = 2
    for p in r["pairs"]:
        _txt(ws, row, 1, str(p.a.d))
        _money(ws, row, 2, round(p.a.amount, 2))
        _txt(ws, row, 3, p.a.desc)
        _txt(ws, row, 4, p.a.account_name)
        _txt(ws, row, 5, str(p.b.d))
        _money(ws, row, 6, round(p.b.amount, 2))
        _txt(ws, row, 7, p.b.desc)
        _txt(ws, row, 8, p.b.account_name)
        c = ws.cell(row=row, column=9, value=p.day_gap)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        _txt(ws, row, 10, p.basis)
        row += 1
    if not r["pairs"]:
        ws.cell(row=2, column=1, value="No matched pairs in this period.")


def _exceptions_sheet(ws, exc, suggestions, side_owner, other_owner):
    headers = [f"Date ({side_owner})", "Amount", "Description", "Contra account",
               f"Probable posting in {other_owner}'s book (mis-posting hunt)"]
    _header_row(ws, 1, headers, widths=[13, 16, 36, 22, 60])
    row = 2
    for m in exc:
        _txt(ws, row, 1, str(m.d))
        _money(ws, row, 2, round(m.amount, 2))
        _txt(ws, row, 3, m.desc)
        _txt(ws, row, 4, m.account_name)
        _txt(ws, row, 5, _fmt_suggestions(suggestions.get(id(m), [])))
        row += 1
    if not exc:
        ws.cell(row=2, column=1, value="No exceptions - every movement matched.")


# --------------------------------------------------------------------------- #
# Entry point.
# --------------------------------------------------------------------------- #
def write_workbook(result: dict, out_path: str) -> str:
    wb = Workbook()
    _summary_sheet(wb.active, result)
    wb.active.title = "Summary"
    _matched_sheet(wb.create_sheet("Matched"), result)
    a_owner = result["book_a"].owner
    b_owner = result["book_b"].owner
    _exceptions_sheet(wb.create_sheet(f"Exceptions {a_owner.split()[0]}"),
                      result["a_exc"], result["a_suggestions"], a_owner, b_owner)
    _exceptions_sheet(wb.create_sheet(f"Exceptions {b_owner.split()[0]}"),
                      result["b_exc"], result["b_suggestions"], b_owner, a_owner)
    wb.save(out_path)
    return out_path
