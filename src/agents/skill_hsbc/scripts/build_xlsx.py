"""Stage 4: Render the enriched JSON into the final Excel workbook.

The workbook has two sheets:

  1. "HSBC Savings <period>" — eight columns, one row per transaction,
     chronologically sorted. The column order was chosen after iteration
     with the user; keep it stable so downstream consumers don't break:

        Date | Transaction Details | Transaction Date | Transaction Number
        | Extra Information | Deposit | Withdrawals | Balance

     Formatting choices (each one has a reason):
       - Arial 10 for body, Arial 11 bold white-on-navy for header.
       - Consolas 9 for Transaction Number — fixed-width helps scanning IDs.
       - Italic grey (Arial 9 #595959) for Extra Information — visually
         subordinate to the primary description.
       - Date columns use dd-mmm-yyyy (unambiguous, locale-free).
       - Amount columns use `#,##0.00;(#,##0.00);"-"` — negatives in
         parentheses, zero shown as a dash.
       - Freeze pane at A2 so the header stays visible while scrolling.
       - Rows where the balance was auto-corrected during reconciliation
         get a soft yellow fill (`FFF2CC`) as a flag for manual review.

  2. "Summary" — opening/closing balance, total deposit/withdrawal counts
     and sums, and how many rows had OCR corrections or enrichment hits.

Usage:
    python build_xlsx.py --in enriched.json --out out.xlsx [--title "HSBC Savings Apr2025-Mar2026"]
"""
import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


HEADERS = [
    "Date", "Transaction Details", "Transaction Date", "Transaction Number",
    "Extra Information", "Deposit", "Withdrawals", "Balance",
]

HEADER_FILL = PatternFill('solid', start_color='1F4E78')
THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CORRECTED_FILL = PatternFill('solid', start_color='FFF2CC')
NUM_FMT = '#,##0.00;(#,##0.00);"-"'
DATE_FMT = 'dd-mmm-yyyy'


def sort_key(t):
    d = t.get('date') or '0000-00-00'
    try:
        return datetime.strptime(d, '%Y-%m-%d')
    except ValueError:
        return datetime(1970, 1, 1)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--in", dest="in_path", required=True, type=Path,
                    help="enriched.json from enrich.py")
    ap.add_argument("--out", dest="out_path", required=True, type=Path,
                    help="Output .xlsx path")
    ap.add_argument("--title", default="HSBC Savings",
                    help="Sheet title (default: 'HSBC Savings')")
    args = ap.parse_args()

    with open(args.in_path) as f:
        data = json.load(f)

    # Drop trailing CLOSING BALANCE rows and any non-transactional stragglers.
    rows = []
    for t in data:
        desc = (t.get('desc') or '').strip()
        if desc.upper() == 'CLOSING BALANCE':
            continue
        if t.get('type') != 'brought_forward' and t.get('deposit') is None and t.get('withdrawal') is None:
            continue
        rows.append(t)

    rows.sort(key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = args.title[:31]  # Excel sheet-name limit

    ws.append(HEADERS)

    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col_idx in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = HEADER_FILL
        c.alignment = header_align
        c.border = BORDER

    flagged_rows = []
    for i, t in enumerate(rows, start=2):
        date_str = t.get('date')
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else None
        except ValueError:
            dt = None

        txn_date_str = t.get('txn_date') or date_str
        try:
            txn_dt = datetime.strptime(txn_date_str, '%Y-%m-%d').date() if txn_date_str else None
        except ValueError:
            txn_dt = None

        if t.get('type') == 'brought_forward':
            details = 'BALANCE BROUGHT FORWARD'
        else:
            cd = t.get('cleaned_desc')
            details = cd if cd is not None else (t.get('desc') or '').strip()
        details = re.sub(r'\s+', ' ', details).strip()

        ws.cell(row=i, column=1, value=dt if dt else date_str)
        ws.cell(row=i, column=2, value=details)
        ws.cell(row=i, column=3, value=txn_dt if txn_dt else txn_date_str)
        ws.cell(row=i, column=4, value=t.get('txn_no') or '')
        ws.cell(row=i, column=5, value=t.get('extra_info') or '')
        ws.cell(row=i, column=6, value=t.get('deposit'))
        ws.cell(row=i, column=7, value=t.get('withdrawal'))
        ws.cell(row=i, column=8, value=t.get('balance'))

        if t.get('balance_corrected'):
            flagged_rows.append(i)
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=i, column=col).fill = CORRECTED_FILL

    # Body formatting
    arial = Font(name='Arial', size=10)
    mono = Font(name='Consolas', size=9)
    extra_font = Font(name='Arial', size=9, italic=True, color='595959')

    for r in range(2, ws.max_row + 1):
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = arial
            cell.border = BORDER

        ws.cell(row=r, column=1).number_format = DATE_FMT
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='top')

        ws.cell(row=r, column=2).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        ws.cell(row=r, column=3).number_format = DATE_FMT
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='left', vertical='top')

        ws.cell(row=r, column=4).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.cell(row=r, column=4).font = mono

        ws.cell(row=r, column=5).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.cell(row=r, column=5).font = extra_font

        for col in (6, 7, 8):
            ws.cell(row=r, column=col).number_format = NUM_FMT
            ws.cell(row=r, column=col).alignment = Alignment(horizontal='right', vertical='top')

    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 26
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 17

    ws.freeze_panes = 'A2'

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    for c in ws2[1]:
        c.font = header_font; c.fill = HEADER_FILL; c.alignment = header_align; c.border = BORDER

    data_only = [r for r in rows if r.get('type') != 'brought_forward']
    with_txn_id = sum(1 for r in data_only if r.get('txn_ids'))
    with_embed_date = sum(1 for r in data_only if r.get('txn_date') and r['txn_date'] != r['date'])
    with_extra = sum(1 for r in data_only if r.get('extra_info'))
    period_start = data_only[0]['date'] if data_only else None
    period_end = data_only[-1]['date'] if data_only else None

    summary_data = [
        ("Period", f"{period_start} to {period_end}" if period_start else ""),
        ("Opening balance", rows[0]['balance'] if rows else None),
        ("Closing balance", rows[-1]['balance'] if rows else None),
        ("Total transactions", len(data_only)),
        ("Transactions with extracted Txn Number", with_txn_id),
        ("Transactions with separate Txn Date", with_embed_date),
        ("Transactions with Extra Information", with_extra),
        ("Total deposits count", sum(1 for r in data_only if r.get('deposit'))),
        ("Total withdrawals count", sum(1 for r in data_only if r.get('withdrawal'))),
        ("Sum of deposits", sum((r.get('deposit') or 0) for r in data_only)),
        ("Sum of withdrawals", sum((r.get('withdrawal') or 0) for r in data_only)),
        ("Net change", sum((r.get('deposit') or 0) for r in data_only) -
                       sum((r.get('withdrawal') or 0) for r in data_only)),
        ("Rows auto-corrected for OCR errors", len(flagged_rows)),
    ]
    for k, v in summary_data:
        ws2.append([k, v])

    for r in range(2, ws2.max_row + 1):
        for col in range(1, 3):
            ws2.cell(row=r, column=col).font = arial
            ws2.cell(row=r, column=col).border = BORDER
        if isinstance(ws2.cell(row=r, column=2).value, (int, float)):
            ws2.cell(row=r, column=2).number_format = NUM_FMT

    ws2.column_dimensions['A'].width = 42
    ws2.column_dimensions['B'].width = 32

    args.out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out_path)
    print(f"Saved {args.out_path}")
    print(f"Data rows: {len(rows)}  (1 brought-forward + {len(data_only)} transactions)")
    print(f"Rows with Txn Number: {with_txn_id}/{len(data_only)}")
    print(f"Rows with Extra Information: {with_extra}/{len(data_only)}")
    print(f"Rows auto-corrected for OCR errors: {len(flagged_rows)}")


if __name__ == "__main__":
    main()
