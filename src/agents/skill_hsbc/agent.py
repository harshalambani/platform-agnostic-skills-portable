"""
agent.py — HSBC bank statement entry point.

``run()`` drives the OCR -> parse -> enrich -> Excel pipeline for the
standalone UI tab, deterministically (no LLM). :class:`HSBCSkill` is the
``BankSkill`` implementation: ``parse()`` accepts a single PDF, a folder of
PDFs (HSBC's multi-statement consolidation -- date-ordered, continuity
checked, see ``scripts/parse_tsv.py``), or -- as a fast-path/test seam that
also keeps the current GnuCash-pipeline call site working unmodified -- an
already-enriched ``.xlsx``/``.xlsm`` workbook. HSBC is the OCR bank: every
row's fidelity is ``"ocr-approx"``, never ``"exact"``.
"""
from __future__ import annotations

import logging
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from agents.bank_common import normalize as _normalize
from agents.bank_common import password as _password
from agents.bank_contract import BankResult, BankStatementMeta
from agents.canonical_io import run_balance_check, write_canonical_csv, write_sidecar

log = logging.getLogger(__name__)

BANK_KEY = "hsbc"

_SCRIPTS = Path(__file__).parent / "scripts"
_PIPELINE = _SCRIPTS / "run_pipeline.py"


def run(
    pdf_dir: str,
    work_dir: str,
    output_path: str,
    title: str = "HSBC Statement",
    config_path: str = None,
    model_override: str = None,
) -> str:
    """
    Run the HSBC OCR -> parse -> enrich -> Excel pipeline directly (no LLM)
    and return a summary string, or raise with the real stderr on failure.

    Args:
        pdf_dir:        Directory containing HSBC PDF statements.
        work_dir:       Scratch directory for intermediate files.
        output_path:    Path where the output .xlsx should be saved.
        title:          Workbook title shown in the Summary sheet.
        config_path:    Unused (kept for run_args compatibility with the
                         other bank skills' entry-point signature).
        model_override: Unused (ditto).
    """
    result = subprocess.run(
        [
            sys.executable, str(_PIPELINE),
            "--pdf-dir", pdf_dir,
            "--work-dir", work_dir,
            "--out", output_path,
            "--title", title,
        ],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(
            "HSBC pipeline failed:\n"
            + (result.stderr.strip() or result.stdout.strip() or "(no output captured)")
        )
    return result.stdout.strip() or "Pipeline complete."


# ---------------------------------------------------------------------------
# BankSkill implementation
# ---------------------------------------------------------------------------
#
# Maps the enriched HSBC workbook to the canonical schema. This replaces the
# retired ``adapter_hsbc`` AND fixes its column-mapping bug: the adapter looked
# for ``Particulars``/``Debit``/``Cheque`` columns, but the enriched workbook
# emits ``Transaction Details``/``Withdrawals``/``Transaction Number``. The old
# code therefore produced blank descriptions, dropped every withdrawal, and
# turned the brought-forward row's NaN amounts into the string "nan" (hence
# ``opening=nan``). The mapping below reads the real column names and coerces
# NaN/blank amounts to "0", so the running balance reconciles.

# Canonical-to-workbook column resolution, in preference order.
_DATE_COLS = ("Date", "Transaction Date")
_DESC_COLS = ("Transaction Details", "Particulars", "Description")
_EXTRA_INFO_COLS = ("Extra Information",)
_TXN_ID_COLS = ("Transaction Number", "Cheque", "Reference", "Ref")
_DEPOSIT_COLS = ("Deposit", "Credit")
_WITHDRAWAL_COLS = ("Withdrawals", "Withdrawal", "Debit")
_BALANCE_COLS = ("Balance",)


def _parse_date_hsbc(date_val: Any) -> Optional[str]:
    """Parse an HSBC date (datetime/Timestamp or string) to ISO YYYY-MM-DD."""
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val.strftime("%Y-%m-%d")
    date_str = str(date_val).strip()
    if not date_str or date_str.lower() == "nat":
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y",
                "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _parse_number_hsbc(num_val: Any) -> str:
    """Parse an amount to a decimal string. NaN/None/blank all become "0".

    Comma/Cr-Dr-suffix cleanup is delegated to ``bank_common.clean_amount``
    (the same primitive HDFC/BoB/ICICI use) — HSBC just guards NaN/None first
    and re-normalizes through ``float()`` to keep its existing decimal-string
    convention (e.g. "50000.0", not "50000").
    """
    if num_val is None:
        return "0"
    # Catch float('nan') (and pandas NaN, which is a float) — nan != nan.
    if isinstance(num_val, float) and num_val != num_val:
        return "0"
    s = str(num_val).strip()
    if not s or s.lower() == "nan":
        return "0"
    cleaned = _normalize.clean_amount(s, blank_zero=False)
    try:
        return str(float(cleaned))
    except ValueError:
        return "0"


def _clean_str(val: Any) -> str:
    """Coerce a cell to a trimmed string; NaN/None/"nan" become ""."""
    if val is None:
        return ""
    if isinstance(val, float) and val != val:  # float('nan')
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


def _pick(columns, candidates) -> Optional[str]:
    """Return the first candidate present in ``columns`` (else None)."""
    return next((c for c in candidates if c in columns), None)


def _read_enriched_rows(input_xlsx: str) -> list[dict]:
    """Read the enriched HSBC workbook's transaction sheet as a list of dicts."""
    import pandas as pd  # noqa: PLC0415

    xls = pd.ExcelFile(input_xlsx)
    # The transaction sheet is the first one ("HSBC Savings <period>"); the
    # trailing "Summary" sheet is metadata. Prefer an explicit "Transactions".
    sheet = "Transactions" if "Transactions" in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(input_xlsx, sheet_name=sheet)
    if df.empty:
        return []

    cols = set(df.columns)
    date_col = _pick(cols, _DATE_COLS)
    desc_col = _pick(cols, _DESC_COLS)
    extra_col = _pick(cols, _EXTRA_INFO_COLS)
    txn_col = _pick(cols, _TXN_ID_COLS)
    dep_col = _pick(cols, _DEPOSIT_COLS)
    wdl_col = _pick(cols, _WITHDRAWAL_COLS)
    bal_col = _pick(cols, _BALANCE_COLS)

    rows: list[dict] = []
    for _, row in df.iterrows():
        date_str = _parse_date_hsbc(row.get(date_col)) if date_col else None
        if not date_str:
            continue
        desc = _clean_str(row.get(desc_col)) if desc_col else ""
        extra = _clean_str(row.get(extra_col)) if extra_col else ""
        # Extra Information (channel/time stamps, stray IMPS markers, etc. --
        # see enrich.py's clean_desc) is real content the enrichment stage
        # extracted; the canonical schema has no dedicated column for it, so
        # fold it into Description rather than silently dropping it.
        description = f"{desc} | {extra}" if desc and extra else (desc or extra)
        rows.append({
            "Date": date_str,
            "Transaction ID": _clean_str(row.get(txn_col)) if txn_col else "",
            "Description": description,
            "Account": "",
            "Deposit": _parse_number_hsbc(row.get(dep_col)) if dep_col else "0",
            "Withdrawal": _parse_number_hsbc(row.get(wdl_col)) if wdl_col else "0",
            "Balance": _parse_number_hsbc(row.get(bal_col)) if bal_col else "0",
            "Currency": "INR",
        })
    return rows


def _run_ocr_pipeline(pdf_paths: list[Path], password: str | None = None) -> Path:
    """Run OCR -> parse -> enrich -> xlsx over ``pdf_paths`` in a fresh scratch
    dir and return the built enriched workbook's path.

    This is the "uniform PDF-in boundary": it's the same four-stage pipeline
    ``run()``/``run_pipeline.py`` drives, just invoked with an in-memory list
    of PDFs (one or many -- HSBC's multi-statement date-ordering and
    continuity detection in ``scripts/parse_tsv.py`` kick in automatically
    for folders with more than one PDF) rather than a pre-existing directory.
    """
    work_dir = Path(tempfile.mkdtemp(prefix="hsbc_bankskill_"))
    pdf_dir = work_dir / "pdfs"
    pdf_dir.mkdir()
    for src in pdf_paths:
        (pdf_dir / src.name).write_bytes(Path(src).read_bytes())
    out_xlsx = work_dir / "enriched.xlsx"

    cmd = [
        sys.executable, str(_PIPELINE),
        "--pdf-dir", str(pdf_dir),
        "--work-dir", str(work_dir / "work"),
        "--out", str(out_xlsx),
    ]
    if password:
        cmd += ["--password", password]

    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        stderr = result.stderr.strip() or result.stdout.strip() or "(no output captured)"
        if _password.is_password_error(RuntimeError(stderr)):
            raise ValueError(
                _password.password_error_message("the statement-open password")
            )
        raise RuntimeError("HSBC pipeline failed:\n" + stderr)
    return out_xlsx


class HSBCSkill:
    """HSBC parser implementing the ``BankSkill`` protocol.

    ``parse`` accepts:
      - a single PDF statement,
      - a folder of PDF statements (HSBC's multi-statement consolidation),
      - or, as a fast-path/test seam -- and to keep the current
        ``skill_gnucash_pipeline`` call site (which builds the enriched
        workbook itself, then calls ``HSBCSkill().parse(xlsx,
        output_path=...)``) working unmodified -- an already-enriched
        ``.xlsx``/``.xlsm`` workbook, which skips OCR entirely.

    Every path OCRs at some stage, so ``fidelity`` is always ``"ocr-approx"``.
    """

    bank_key = BANK_KEY

    def formats(self) -> tuple[str, ...]:
        return (".pdf",)

    def detect(self, path: str | Path) -> float:
        """Confidence that ``path`` is an HSBC statement (PDF) or an already-
        enriched HSBC workbook (the fast-path/test seam)."""
        p = Path(path)
        suffix = p.suffix.lower()
        if suffix in (".xlsx", ".xlsm"):
            try:
                import openpyxl  # noqa: PLC0415

                wb = openpyxl.load_workbook(str(p), read_only=True)
                ws = wb[wb.sheetnames[0]]
                header = {str(c.value).strip() for c in next(ws.iter_rows(max_row=1))}
                wb.close()
            except Exception:  # noqa: BLE001 — detection must never raise
                return 0.0
            return 0.9 if ("Transaction Details" in header and "Withdrawals" in header) else 0.0
        if suffix != ".pdf":
            return 0.0
        if not p.is_file():
            return 0.0
        try:
            import pdfplumber  # noqa: PLC0415

            with pdfplumber.open(str(p)) as pdf:
                text = (pdf.pages[0].extract_text() or "") if pdf.pages else ""
            return 0.8 if "hsbc" in text.lower() else 0.5
        except Exception:  # noqa: BLE001 — scanned/encrypted PDFs have no text layer; that's the norm for HSBC
            return 0.5

    def parse(
        self,
        path: str | Path,
        password: str | None = None,
        output_path: str | Path | None = None,
    ) -> BankResult:
        """Parse HSBC statement(s) into a canonical BankResult.

        ``path`` may be a single PDF, a folder of PDFs, or an already-enriched
        ``.xlsx``/``.xlsm`` workbook (see class docstring). ``output_path``,
        when given, writes the canonical CSV + sidecar as a side effect (the
        existing ``skill_gnucash_pipeline`` call convention).
        """
        p = Path(path)
        suffix = p.suffix.lower()

        if suffix in (".xlsx", ".xlsm"):
            enriched_xlsx = p
        elif p.is_dir():
            pdfs = sorted(p.glob("*.pdf"))
            if not pdfs:
                raise ValueError(f"No PDFs found in {path}")
            enriched_xlsx = _run_ocr_pipeline(pdfs, password)
        elif suffix == ".pdf":
            enriched_xlsx = _run_ocr_pipeline([p], password)
        else:
            raise ValueError(f"Unsupported file type: {suffix or path}")

        rows = _read_enriched_rows(str(enriched_xlsx))
        if not rows:
            raise ValueError(f"No transaction rows found in HSBC statement: {path}")

        balance_check = run_balance_check(rows)
        warnings: list[str] = []
        if not balance_check.ok:
            warnings.extend(f"Balance: {d}" for d in balance_check.details)

        sidecar_path = None
        if output_path is not None:
            write_canonical_csv(rows, output_path)
            sidecar_path = write_sidecar(
                output_path, "HSBC", "derived",
                balance_check.opening_balance, balance_check.closing_balance,
                len(rows),
            )

        dates = [r["Date"] for r in rows if r.get("Date")]
        meta = BankStatementMeta(
            bank_key=BANK_KEY,
            account_number=None,
            period_from=min(dates) if dates else None,
            period_to=max(dates) if dates else None,
            source_format="pw-pdf" if password else "pdf",
            fidelity="ocr-approx",
            password_used=bool(password),
        )

        return BankResult(
            rows=rows,
            bank_key=BANK_KEY,
            account_label="HSBC",
            currency="INR",
            opening_balance=balance_check.opening_balance,
            closing_balance=balance_check.closing_balance,
            balance_check=balance_check,
            sidecar_path=sidecar_path,
            warnings=warnings,
            meta=meta,
        )


bank_skill = HSBCSkill()
