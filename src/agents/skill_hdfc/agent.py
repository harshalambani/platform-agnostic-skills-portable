#!/usr/bin/env python3
"""
HDFC Bank statement -> canonical 8-column CSV.

Supports four input shapes, all deterministic (no LLM):
  PDF (digital, password-protected, or scanned/garbled):
    1. (Primary) pdfplumber text extraction (optionally password-protected)
    2. Garbled-text-layer detection -- if the extracted text is unusable
       (custom font encoding producing "(cid:NN)" junk, low printable-ASCII
       ratio, or missing structural anchors), fall back to OCR
    3. (Fallback) OCR with pytesseract
    4. Regex-parse transaction lines from either source
    5. Statement summary validation (DrCount, CrCount, Opening, Closing)

  XLS/XLSX (net-banking download):
    1. Read with xlrd (.xls) or openpyxl (.xlsx)
    2. Auto-detect header row (tolerant of preamble rows and '****' separators)
    3. Deterministic column mapping (alias table)

  CSV (net-banking download, possibly with renamed headers):
    1. Read with csv.reader
    2. Same header-row detection + alias-table column mapping as XLS/XLSX
    3. Accepts both ISO (YYYY-MM-DD) and DD/MM/YY(YY) dates

Canonical output columns:
  Date, Transaction ID, Description, Account, Deposit, Withdrawal, Balance, Currency
"""
import csv
import json
import logging
import re
from pathlib import Path

from agents.balance_utils import (
    verify_running_balance,
    verify_closing_balance,
    format_balance_summary,
)
from agents.bank_common import normalize as _normalize
from agents.bank_common import password as _password
from agents.bank_common import tabular as _tabular
from agents.bank_common import text_quality as _text_quality
from agents.bank_contract import BankResult, BankStatementMeta
from agents.canonical_io import CANONICAL_FIELDS, run_balance_check

log = logging.getLogger(__name__)

BANK_KEY = "hdfc"

# Single source of truth for the canonical schema lives in canonical_io; keep
# the local name as an alias so existing references don't churn.
CANONICAL_COLS = list(CANONICAL_FIELDS)

# Local aliases onto agents.bank_common — HDFC is the reference implementation
# these were promoted from verbatim; kept as module-level names so the many
# call-sites below don't churn.
_clean_amount = _normalize.clean_amount
_normalise_date = _normalize.normalise_date


def _extract_statement_summary(text):
    m = re.search(
        r'STATEMENT\s*SUMMARY\s*[:\-]*\s*'
        r'(?:Opening\s*Balance\s*Dr\s*Count\s*Cr\s*Count\s*Debits\s*Credits\s*Closing\s*Bal\w*\s*)?'
        r'([\d,]+\.?\d*)\s+(\d+)\s+(\d+)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)',
        text, re.IGNORECASE,
    )
    if not m:
        return {}
    return {
        "opening": float(m.group(1).replace(",", "")),
        "dr_count": int(m.group(2)),
        "cr_count": int(m.group(3)),
        "debits": float(m.group(4).replace(",", "")),
        "credits": float(m.group(5).replace(",", "")),
        "closing": float(m.group(6).replace(",", "")),
    }


# ---------------------------------------------------------------------------
# Garbled-text-layer detection (item 2): some HDFC PDFs use a custom font
# encoding that pdfplumber/pdfminer can't map to real characters, leaving
# "(cid:NN)" placeholders instead of text. Detect this BEFORE attempting the
# regex parse and route to OCR instead.
# ---------------------------------------------------------------------------

# HDFC's required structural anchors -- absence of either indicates a
# font-encoding problem that will make the transaction-line regexes fail
# silently. See agents.bank_common.text_quality for the shared heuristic.
_HDFC_TEXT_ANCHORS = (r'\bdate\b', r'narration')


def _text_layer_usable(full_text):
    return _text_quality.text_layer_usable(full_text, _HDFC_TEXT_ANCHORS)


_PB_DATE_RE = re.compile(r'^(\d{2}/\d{2}/\d{2})\s+(.+)')
# HDFC PDF exports vary: some print a literal "0.00" placeholder for the
# blank Withdrawal/Deposit column (3 trailing numbers: withdrawal, deposit,
# balance); others omit the blank column entirely (2 trailing numbers:
# amount, balance) -- capture 2-or-3 and let _build_pb_txn disambiguate.
_PB_TAIL_RE = re.compile(
    r'(\S+)\s+'
    r'(\d{2}/\d{2}/\d{2})\s+'
    r'((?:[\d,]+\.\d{2}\s*){2,3})$'
)

_PB_SKIP_RE = re.compile(r'|'.join([
    r'closing\s*balance', r'opening\s*balance', r'page\s*no',
    r'statement\s*of\s*account', r'generated\s*on', r'generated\s*by',
    r'hdfc\s*bank\s*limited', r'registered\s*office',
    r'contents\s*of\s*this', r'state\s*account\s*branch',
    r'gstin\s*number', r'narration', r'chq\.\s*/?\s*ref',
    r'withdrawal\s*amt', r'deposit\s*amt', r'statement\s*summary',
    r'earmarked\s*for\s*hold', r'considered\s*correct',
    r'branch\s*gstn', r'senapati\s*bapat', r'account\s*branch',
    r'preferred\s*customer', r'account\s*type', r'branch\s*code',
    r'nomination', r'rtgs.*ifsc', r'joint\s*holders',
    r'a/c\s*open\s*date', r'account\s*status', r'account\s*no',
    r'cust\s*id', r'od\s*limit', r'currency\s*:', r'email\s*:',
    r'phone\s*no', r'city\s*:', r'address\s*:', r'landmark',
    r's\.v\.\s*road', r'requesting\s*branch', r'from\s*:\s*\d{2}/\d{2}',
    r'^\d{2}/\d{2}/\d{4}$',
    r'^\d[\d,]+\.\d{2}\s+\d+\s+\d+\s+\d[\d,]+\.\d{2}',
    r'^MAHARASHTRA$', r'^MUMBAI$', r'^INDIA$',
]), re.IGNORECASE)

_PB_FOOTER_RE = re.compile(
    r'hdfcbank|closingbal|senapati|registeredo|stateaccount'
    r'|gstinnumber|contentsof|thisstatement|taxpayment'
    r'|understand\s*your|\.com/personal'
    r'|computergenerated|doesnotrequire|signature'
    r'|^MR\.|^MRS\.|^MS\.'
    r'|^State\s*:',
    re.IGNORECASE,
)


_AMBIGUOUS_AMOUNT_KEY = "_ambiguous_amount"


def _build_pb_txn(date_str, rest, m_tail, cont):
    """Build a transaction dict from one pdfplumber-parsed line.

    "Date" is emitted as the Value Dt (m_tail.group(2)), falling back to the
    leading posting date only when the value date is blank/unparseable —
    HDFC statements have distinct posting and value dates (e.g. cheque
    clearing) and downstream balance/dedup logic keys on this field.

    The trailing numeric tail (m_tail.group(3)) holds either 3 numbers
    (withdrawal, deposit, balance -- some exports print a literal "0.00" for
    the blank column) or 2 (amount, balance -- other exports omit the blank
    column entirely). The 2-number case is genuinely ambiguous from the line
    alone; it's flagged with _AMBIGUOUS_AMOUNT_KEY and resolved afterwards by
    _resolve_ambiguous_amounts() using the running-balance direction.
    """
    ref = m_tail.group(1)
    ref_pos = rest.rfind(ref)
    desc = rest[:ref_pos].strip() if ref_pos > 0 else ""
    full_desc = (desc + " " + " ".join(cont)).strip() if cont else desc
    value_date = _normalise_date(m_tail.group(2)) if m_tail.group(2) else ""
    posting_date = _normalise_date(date_str)
    amounts = m_tail.group(3).split()
    ref_clean = ref.lstrip("0") or ref
    txn = {
        "Date": value_date or posting_date,
        "Transaction ID": ref_clean,
        "Description": full_desc,
        "Account": "",
        "Currency": "INR",
    }
    if len(amounts) >= 3:
        withdrawal, deposit, balance = amounts[-3], amounts[-2], amounts[-1]
        txn["Withdrawal"] = _clean_amount(withdrawal)
        txn["Deposit"] = _clean_amount(deposit)
        txn["Balance"] = balance.replace(",", "")
    else:
        amount, balance = amounts[-2], amounts[-1]
        txn["Withdrawal"] = ""
        txn["Deposit"] = ""
        txn["Balance"] = balance.replace(",", "")
        txn[_AMBIGUOUS_AMOUNT_KEY] = amount
    return txn


def _resolve_ambiguous_amounts(transactions, opening_balance=None):
    """Resolve single-amount transaction lines (see _build_pb_txn) into a
    Deposit or Withdrawal using the running-balance direction: Balance is
    always present and exact, so comparing consecutive rows' Balance tells us
    whether the amount increased (deposit) or decreased (withdrawal) the
    account -- regardless of which printed column it came from.

    Mutates and returns `transactions`. `opening_balance` seeds the first
    row's comparison when available (from the statement summary); if absent,
    the first ambiguous row falls back to treating the amount as a
    withdrawal only if that reconciles with its own balance change against
    itself being impossible to know, so it's left as a withdrawal by
    convention (rare: only the very first row of a statement can hit this).
    """
    prev_balance = opening_balance
    for txn in transactions:
        cur_balance = float(txn["Balance"].replace(",", "") or "0")
        if _AMBIGUOUS_AMOUNT_KEY in txn:
            amount = txn.pop(_AMBIGUOUS_AMOUNT_KEY)
            if prev_balance is not None and cur_balance < prev_balance:
                txn["Withdrawal"] = _clean_amount(amount)
                txn["Deposit"] = ""
            else:
                txn["Deposit"] = _clean_amount(amount)
                txn["Withdrawal"] = ""
        prev_balance = cur_balance
    return transactions


def _parse_pdf_pdfplumber(pdf_path, password=None):
    """Extract transactions via pdfplumber's text layer.

    Returns (transactions, summary, usable) where `usable` reports whether
    the text layer itself looked sane (see _text_layer_usable) -- callers
    use this to decide whether to fall back to OCR even when a handful of
    (garbage) transactions happened to match the regexes.
    """
    import pdfplumber

    try:
        pdf = pdfplumber.open(pdf_path, password=password or "")
    except Exception as e:
        if _password.is_password_error(e):
            raise ValueError(
                _password.password_error_message("for HDFC often the Cust ID")
            ) from e
        raise

    all_lines = []
    for page in pdf.pages:
        text = page.extract_text(x_tolerance=1) or ""
        for line in text.split("\n"):
            stripped = line.strip()
            if stripped:
                all_lines.append(stripped)
    pdf.close()

    if not all_lines:
        return [], {}, False

    full_text = "\n".join(all_lines)
    usable = _text_layer_usable(full_text)
    if not usable:
        return [], {}, False

    summary = _extract_statement_summary(full_text)

    transactions = []
    i = 0
    while i < len(all_lines):
        line = all_lines[i]
        m_date = _PB_DATE_RE.match(line)
        m_tail = _PB_TAIL_RE.search(m_date.group(2)) if m_date else None
        # Skip/footer text (e.g. "HDFC BANK LIMITED", "...contents of this
        # statement...") is filtered out UNLESS the line already looks like a
        # complete transaction row (date + ref + value-dt + amounts) --
        # UPI narrations routinely embed "HDFCBANK" as a VPA domain suffix
        # (e.g. "BLINKIT.PAYU@HDFCBANK"), which would otherwise false-match
        # _PB_FOOTER_RE and silently drop real transactions.
        if not m_tail and (_PB_SKIP_RE.search(line) or _PB_FOOTER_RE.search(line)):
            i += 1
            continue
        if m_date:
            date_str = m_date.group(1)
            rest = m_date.group(2)
            if m_tail:
                cont = []
                j = i + 1
                while j < len(all_lines):
                    nl = all_lines[j]
                    if _PB_DATE_RE.match(nl):
                        break
                    if _PB_SKIP_RE.search(nl) or _PB_FOOTER_RE.search(nl):
                        j += 1
                        continue
                    if len(nl) < 5 and nl.isdigit():
                        j += 1
                        continue
                    cont.append(nl)
                    j += 1
                transactions.append(_build_pb_txn(date_str, rest, m_tail, cont))
                i = j
                continue
        i += 1
    return transactions, summary, usable


_DATE_RE = re.compile(r'^(\d{2}/\d{2}/\d{2})\s*[.\s]*\|')
# Tesseract inconsistently recognises the "|" column separator between the
# ref number and the value date (sometimes a real "|", sometimes just
# whitespace, depending on scan/glyph quality) -- accept either. The
# trailing amounts are likewise 2-or-3 numbers (see _PB_TAIL_RE for why) and
# resolved the same way via _resolve_ambiguous_amounts.
_TAIL_RE = re.compile(
    r'([^\s|]{10,25})[|\s]+'
    r'(\d{2}/\d{2}/\d{2})\s+'
    r'((?:[\d,]+\.\d{2}\s*){2,3})$'
)


def _ocr_pdf(pdf_path, password=None):
    import pypdfium2 as pdfium
    import pytesseract
    pdf = pdfium.PdfDocument(pdf_path, password=password or None)
    pages_text = []
    for i in range(len(pdf)):
        page = pdf[i]
        bitmap = page.render(scale=2.5)
        pil_img = bitmap.to_pil()
        text = pytesseract.image_to_string(pil_img, config="--psm 6")
        pages_text.append(text)
    return "\n".join(pages_text), len(pdf)


def _parse_pdf_transactions(ocr_text):
    lines = ocr_text.splitlines()
    start_idx = 0
    for i, line in enumerate(lines):
        if "statement of account" in line.lower():
            start_idx = i + 1
            break
    transactions = []
    current = None
    current_text = ""
    for line in lines[start_idx:]:
        line = line.strip()
        if not line:
            continue
        low = line.lower()
        if any(k in low for k in [
            "closing balance", "opening balance", "total deposits",
            "total withdrawals", "page no", "statement of account",
            "generated on", "generated by", "statement summary",
        ]):
            continue
        m_date = _DATE_RE.match(line)
        if m_date:
            if current and current_text:
                m_tail = _TAIL_RE.search(current_text)
                if m_tail:
                    transactions.append(_build_pdf_txn(current, current_text, m_tail))
            current = {"date": m_date.group(1)}
            current_text = line
        elif current is not None:
            current_text += " " + line
        if current and current_text:
            m_tail = _TAIL_RE.search(current_text)
            if m_tail:
                transactions.append(_build_pdf_txn(current, current_text, m_tail))
                current = None
                current_text = ""
    if current and current_text:
        m_tail = _TAIL_RE.search(current_text)
        if m_tail:
            transactions.append(_build_pdf_txn(current, current_text, m_tail))
    return transactions


def _clean_ocr_desc(desc):
    desc = re.sub(r'[\|]+\s*[\-\s]*$', "", desc)
    desc = re.sub(r'\s+0$', "", desc)
    desc = re.sub(r'\s+\d[\d,]*\.\d{2}\s+\d[\d,]*\.\d{2}\s+\d[\d,]*\.\d{2}', "", desc)
    desc = re.sub(r'\s*\|\s*', " ", desc)
    return re.sub(r'\s+', " ", desc).strip(" -|")


def _build_pdf_txn(current, text, m_tail):
    """Build a transaction dict from one OCR-parsed line.

    "Date" is emitted as the Value Dt (m_tail.group(2)), falling back to the
    leading posting date (current["date"]) only when the value date is
    blank/unparseable — same rationale as _build_pb_txn for the pdfplumber path.

    The trailing amounts blob (m_tail.group(3)) holds 2 or 3 numbers for the
    same reason as _build_pb_txn: some scans/exports carry a withdrawal +
    deposit + balance triple, others only amount + balance. The 2-number case
    is flagged with _AMBIGUOUS_AMOUNT_KEY and resolved by
    _resolve_ambiguous_amounts() afterwards.
    """
    ref_no = m_tail.group(1)
    ref_clean = ref_no.lstrip("0") or ref_no
    amounts = m_tail.group(3).split()
    pipe_idx = text.find("|")
    if pipe_idx >= 0:
        pipe_idx += 1
        ref_idx = text.rfind(ref_no)
        raw_desc = text[pipe_idx:ref_idx].strip() if ref_idx > pipe_idx else ""
    else:
        raw_desc = ""
    desc = re.sub(r'\s+', " ", raw_desc).strip(" -")
    desc = _clean_ocr_desc(desc)
    value_date = _normalise_date(m_tail.group(2)) if m_tail.group(2) else ""
    posting_date = _normalise_date(current["date"])
    txn = {
        "Date": value_date or posting_date,
        "Transaction ID": ref_clean,
        "Description": desc,
        "Account": "",
        "Currency": "INR",
    }
    if len(amounts) >= 3:
        withdrawal, deposit, balance = amounts[-3], amounts[-2], amounts[-1]
        txn["Withdrawal"] = _clean_amount(withdrawal)
        txn["Deposit"] = _clean_amount(deposit)
        txn["Balance"] = balance.replace(",", "")
    else:
        amount, balance = amounts[-2], amounts[-1]
        txn["Withdrawal"] = ""
        txn["Deposit"] = ""
        txn["Balance"] = balance.replace(",", "")
        txn[_AMBIGUOUS_AMOUNT_KEY] = amount
    return txn


# ---------------------------------------------------------------------------
# Tabular parsing (XLS / XLSX / CSV) — shared header-detection + alias-table
# column mapping. HDFC's native XLS export ("Date", "Narration", "Chq./Ref.No.",
# "Value Dt", "Withdrawal Amt.", "Deposit Amt.", "Closing Balance") and a
# renamed CSV export ("Value Date", "Description", "Number", "Withdrawal",
# "Deposit", "Balance") both resolve through the same alias table.
# ---------------------------------------------------------------------------

_HDFC_TABULAR_COLS = {
    "date":       ["date"],
    "value_date": ["value dt", "value date", "val date", "val dt"],
    "narration":  ["narration", "description", "particulars"],
    "nature":     ["nature of exp", "nature", "remarks"],
    "ref":        ["chq./ref.no.", "chq/ref no", "ref no", "cheque no",
                    "reference no", "number", "transaction id"],
    "withdrawal": ["withdrawal amt.", "withdrawal", "debit", "debit amt."],
    "deposit":    ["deposit amt.", "deposit", "credit", "credit amt."],
    "balance":    ["closing balance", "balance"],
}

_HEADER_DATE_ALIASES = ("date", "value date", "value dt", "val date", "val dt", "txn date")
_HEADER_DESC_ALIASES = ("narration", "description", "particulars")

_TABULAR_DATE_ROW_RE = re.compile(r'^\d{2}/\d{2}/\d{2,4}$|^\d{4}-\d{2}-\d{2}$')


def _read_xls_rows(file_path):
    p = Path(file_path)
    suffix = p.suffix.lower()
    if suffix == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c).strip() if c is not None else "" for c in row])
        return rows
    else:
        try:
            import xlrd
        except ImportError:
            raise ImportError("xlrd is required to read .xls files.")
        wb = xlrd.open_workbook(str(p))
        ws = wb.sheet_by_index(0)
        rows = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    row.append(dt.strftime("%d/%m/%Y"))
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    v = cell.value
                    row.append(str(int(v)) if v == int(v) else str(v))
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    row.append("")
                else:
                    row.append(str(cell.value).strip())
            rows.append(row)
        return rows


def _read_csv_rows(file_path):
    with open(file_path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        return [[str(c).strip() for c in row] for row in reader]


# Header-row detection is tolerant of preamble rows and a '****' separator
# (HDFC's native XLS export) by looking for the FIRST row containing both a
# date-like column name and a description-like column name, regardless of
# their position (CSV exports may rename/reorder the "Date" column to "Value
# Date", which sits in column 0 with no separate "Date" column). See
# agents.bank_common.tabular for the shared mechanics.

def _find_header_row(rows):
    return _tabular.find_header_row(rows, _HEADER_DATE_ALIASES, _HEADER_DESC_ALIASES)


def _map_columns(header_row):
    return _tabular.map_columns(header_row, _HDFC_TABULAR_COLS)


def _parse_tabular_transactions(rows, source_label):
    header_idx = _find_header_row(rows)
    if header_idx == -1:
        raise ValueError(
            "Could not find HDFC header row in " + source_label +
            " (expected a Date-like column and a Narration/Description column)"
        )
    col = _map_columns(rows[header_idx])
    for req in ("narration", "ref", "withdrawal", "deposit", "balance"):
        if col.get(req) is None:
            raise ValueError("Required column not found in " + source_label + ": " + req)
    if col.get("date") is None and col.get("value_date") is None:
        raise ValueError("Required column not found in " + source_label + ": date")
    date_col = col["value_date"] if col.get("value_date") is not None else col["date"]
    transactions = []
    for row in rows[header_idx + 1:]:
        if not row or not str(row[0]).strip():
            continue
        if date_col >= len(row):
            continue
        date_raw = str(row[date_col]).strip()
        if not _TABULAR_DATE_ROW_RE.match(date_raw):
            continue
        narration = str(row[col["narration"]]).strip() if col["narration"] is not None else ""
        nature = str(row[col["nature"]]).strip() if col.get("nature") is not None else ""
        ref = str(row[col["ref"]]).strip() if col["ref"] is not None else ""
        wdl = _clean_amount(row[col["withdrawal"]]) if col["withdrawal"] is not None else ""
        dep = _clean_amount(row[col["deposit"]]) if col["deposit"] is not None else ""
        bal = _clean_amount(row[col["balance"]]) if col["balance"] is not None else ""
        if nature and nature.lower() not in ("", "none", "nan"):
            desc = narration + " -- " + nature
        else:
            desc = narration
        ref_clean = ref.lstrip("0") or ref
        transactions.append({
            "Date": _normalise_date(date_raw),
            "Transaction ID": ref_clean,
            "Description": desc,
            "Account": "",
            "Deposit": dep,
            "Withdrawal": wdl,
            "Balance": bal,
            "Currency": "INR",
        })
    return transactions


# Backwards-compatible alias: existing tests/callers import _parse_xls_transactions.
def _parse_xls_transactions(rows):
    return _parse_tabular_transactions(rows, "XLS/XLSX")


class _UnsupportedFormat(Exception):
    """Raised by _extract_transactions for a suffix HDFC doesn't handle."""


def _extract_transactions(input_path, pdf_password=None):
    """Parse ``input_path`` (PDF/XLS/XLSX/CSV) into (transactions, summary, fmt).

    The shared core of both run() and parse(): sniffs the suffix, dispatches
    to the right reader, and falls back to OCR for PDFs with an unusable text
    layer. Raises ValueError (a parser-detected problem, e.g. bad password or
    missing header) or _UnsupportedFormat (unknown suffix) instead of
    returning an error string, so both callers can format/handle it their
    own way.
    """
    suffix = Path(input_path).suffix.lower()
    summary = {}

    if suffix == ".pdf":
        transactions, summary, usable = _parse_pdf_pdfplumber(input_path, password=pdf_password)
        if transactions:
            _resolve_ambiguous_amounts(transactions, summary.get("opening"))
        fmt = "PDF (pdfplumber)"
        if not usable or not transactions:
            page_count_note = ""
            try:
                import pdfplumber
                with pdfplumber.open(input_path, password=pdf_password or "") as _p:
                    page_count_note = " (~%dpp)" % len(_p.pages)
            except Exception:
                pass
            log.info(
                "pdfplumber text layer unusable or empty — falling back to OCR%s",
                page_count_note,
            )
            ocr_text, _page_count = _ocr_pdf(input_path, password=pdf_password)
            transactions = _parse_pdf_transactions(ocr_text)
            if not summary:
                summary = _extract_statement_summary(ocr_text)
            if transactions:
                _resolve_ambiguous_amounts(transactions, summary.get("opening"))
            fmt = "PDF (OCR" + page_count_note + ")"
    elif suffix in (".xls", ".xlsx"):
        rows = _read_xls_rows(input_path)
        transactions = _parse_tabular_transactions(rows, suffix.upper())
        fmt = suffix.upper()
    elif suffix == ".csv":
        rows = _read_csv_rows(input_path)
        transactions = _parse_tabular_transactions(rows, "CSV")
        fmt = "CSV"
    else:
        raise _UnsupportedFormat(suffix)
    return transactions, summary, fmt


def run(pdf_path, output_path, config_path=None, model_override=None, pdf_password=None):
    """HDFC statement (PDF, password-protected PDF, XLS/XLSX, or CSV) -> canonical CSV."""
    input_path = str(pdf_path)
    if not Path(input_path).is_file():
        return "File not found: " + input_path

    try:
        transactions, summary, fmt = _extract_transactions(input_path, pdf_password)
    except _UnsupportedFormat as e:
        return "Unsupported file type: " + str(e)
    except ValueError as e:
        return "Error processing " + Path(input_path).name + ": " + str(e)
    except Exception as e:
        return "Error processing " + Path(input_path).name + ": " + str(e)

    if not transactions:
        return "No transactions found in " + fmt + " file."

    running = verify_running_balance(transactions)

    if summary and "closing" in summary:
        expected_closing = summary["closing"]
    else:
        expected_closing = running["closing_balance"]

    closing = verify_closing_balance(transactions, expected_closing=expected_closing)

    count_warnings = []
    if summary:
        expected_total = summary.get("dr_count", 0) + summary.get("cr_count", 0)
        actual_total = len(transactions)
        if expected_total > 0 and actual_total != expected_total:
            msg = "TRANSACTION COUNT MISMATCH: extracted %d, statement says %d (Dr=%d + Cr=%d)" % (
                actual_total, expected_total, summary["dr_count"], summary["cr_count"])
            count_warnings.append(msg)
            log.warning(msg)

        actual_dr = sum(1 for t in transactions
                        if float(t.get("Withdrawal", "0").replace(",", "") or "0") > 0)
        actual_cr = sum(1 for t in transactions
                        if float(t.get("Deposit", "0").replace(",", "") or "0") > 0)
        if actual_dr != summary.get("dr_count", actual_dr):
            count_warnings.append("  Debit count: extracted %d, expected %d" % (
                actual_dr, summary["dr_count"]))
        if actual_cr != summary.get("cr_count", actual_cr):
            count_warnings.append("  Credit count: extracted %d, expected %d" % (
                actual_cr, summary["cr_count"]))

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CANONICAL_COLS)
        writer.writeheader()
        writer.writerows(transactions)

    # Write sidecar summary JSON for pipeline's independent balance verification
    sidecar_path = Path(output_path).with_suffix(".csv_summary.json")
    sidecar_data = {
        "bank": "HDFC",
        "source": "statement_summary" if summary else "derived",
        "opening_balance": summary.get("opening", running["opening_balance"]),
        "closing_balance": summary.get("closing", running["closing_balance"]),
        "dr_count": summary.get("dr_count", 0),
        "cr_count": summary.get("cr_count", 0),
        "row_count": len(transactions),
    }
    try:
        with open(sidecar_path, "w", encoding="utf-8") as sf:
            json.dump(sidecar_data, sf, indent=2)
        log.info("Wrote sidecar summary: %s", sidecar_path)
    except Exception as e:
        log.warning("Could not write sidecar summary: %s", e)

    balance_info = format_balance_summary(running, closing)
    result = "HDFC (%s): extracted %d transactions from %s -> canonical CSV\n%s" % (
        fmt, len(transactions), Path(input_path).name, balance_info)
    if count_warnings:
        result += "\n" + "\n".join(count_warnings)
    return result


# ---------------------------------------------------------------------------
# BankSkill protocol (agents.bank_contract) — re-expresses the parsing above
# as detect()/parse()/formats(), with zero change to run()'s own behavior
# (run() is still the skill's UI entry_point and is untouched above).
# ---------------------------------------------------------------------------

def formats() -> tuple[str, ...]:
    return (".pdf", ".xls", ".xlsx", ".csv")


def detect(path) -> float:
    """Cheap, conservative sniff: does this file look like an HDFC statement?

    Format mismatch (wrong suffix) -> 0.0. Otherwise looks for "hdfc" in the
    first slice of readable content; a positive hit is a strong signal, a
    miss (or an unreadable/encrypted file, e.g. a password-protected PDF we
    can't peek into without a password) still returns a low-but-nonzero
    confidence since the format itself matches.
    """
    p = Path(str(path))
    suffix = p.suffix.lower()
    if suffix not in formats():
        return 0.0
    try:
        if suffix == ".pdf":
            import pdfplumber
            with pdfplumber.open(str(p)) as pdf:
                text = (pdf.pages[0].extract_text() or "") if pdf.pages else ""
        elif suffix == ".csv":
            text = p.read_text(encoding="utf-8-sig", errors="ignore")[:4000]
        else:
            rows = _read_xls_rows(str(p))
            text = " ".join(str(c) for row in rows[:5] for c in row)
        return 0.8 if "hdfc" in text.lower() else 0.3
    except Exception:
        return 0.3


def parse(path, password=None) -> BankResult:
    """Parse ``path`` into a :class:`BankResult` of canonical rows.

    Delegates to the same ``_extract_transactions`` core as ``run()`` — no
    parsing logic is duplicated — but returns the rows in memory rather than
    writing a CSV/sidecar (that IO belongs to the caller, per the BankSkill
    protocol's contract).
    """
    input_path = str(path)
    if not Path(input_path).is_file():
        raise FileNotFoundError(input_path)
    suffix = Path(input_path).suffix.lower()
    if suffix not in formats():
        raise ValueError("Unsupported file type: " + suffix)

    transactions, summary, fmt = _extract_transactions(input_path, password)
    if not transactions:
        raise ValueError("No transactions found in " + fmt + " file.")

    balance_check = run_balance_check(transactions)
    opening = summary.get("opening", balance_check.opening_balance)
    closing = summary.get("closing", balance_check.closing_balance)

    warnings = []
    if summary:
        expected_total = summary.get("dr_count", 0) + summary.get("cr_count", 0)
        if expected_total > 0 and len(transactions) != expected_total:
            warnings.append(
                "TRANSACTION COUNT MISMATCH: extracted %d, statement says %d (Dr=%d + Cr=%d)" % (
                    len(transactions), expected_total, summary["dr_count"], summary["cr_count"])
            )

    meta = BankStatementMeta(
        bank_key=BANK_KEY,
        account_number=None,
        period_from=None,
        period_to=None,
        source_format=fmt,
        fidelity="ocr-approx" if "OCR" in fmt else "exact",
        password_used=bool(password),
    )

    return BankResult(
        rows=transactions,
        bank_key=BANK_KEY,
        currency="INR",
        opening_balance=opening,
        closing_balance=closing,
        balance_check=balance_check,
        sidecar_path=None,
        warnings=warnings,
        meta=meta,
    )


class HDFCBankSkill:
    """BankSkill implementation for HDFC — delegates to this module's
    detect()/parse()/formats(), which are the reference for other banks'
    future migration onto the same protocol."""

    bank_key = BANK_KEY

    def formats(self) -> tuple[str, ...]:
        return formats()

    def detect(self, path) -> float:
        return detect(path)

    def parse(self, path, password=None) -> BankResult:
        return parse(path, password=password)


bank_skill = HDFCBankSkill()
