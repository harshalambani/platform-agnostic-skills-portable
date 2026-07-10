#!/usr/bin/env python3
"""
matrix_report.py -- render an all-family matrix result (from
matrix_recon.run_matrix) to a formatted .xlsx workbook.

Reuses the styling helpers from the pairwise skill's excel_report.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Reuse styling + helpers from the pairwise skill.
_PAIR_SCRIPTS = (Path(__file__).resolve().parents[2]
                 / "skill_gnucash_intercompany" / "scripts")
if str(_PAIR_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_PAIR_SCRIPTS))

from excel_report import (  # noqa: E402
    NUM, HDR_FILL, HDR_FONT, TITLE_FONT, SUB_FONT, OK_FILL, BAD_FILL, BORDER,
    _header_row, _money, _txt,
)

DIAG_FILL = None  # set lazily to avoid import-time openpyxl style churn


def _matrix_sheet(ws, m):
    owners = m["owners"]
    n = len(owners)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Intercompany Matrix -- balance difference per pair"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"Period: {m['fy_label']}   |   date tolerance: +/-{m['tol_days']} days"
                f"   |   0 = ties, blank = same book, n/a = no mutual contra")
    ws["A2"].font = SUB_FONT

    top = 4
    # corner + column headers
    corner = ws.cell(row=top, column=1, value="A \\ B")
    corner.fill = HDR_FILL
    corner.font = HDR_FONT
    corner.border = BORDER
    corner.alignment = Alignment(horizontal="center")
    for j, o in enumerate(owners):
        c = ws.cell(row=top, column=2 + j, value=o)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(2 + j)].width = 16
    ws.column_dimensions["A"].width = 20

    # index pair results by (ia, ib)
    grid = {}
    for p in m["pairs"]:
        grid[(p["ia"], p["ib"])] = p
        grid[(p["ib"], p["ia"])] = p

    for i, o in enumerate(owners):
        r = top + 1 + i
        rc = ws.cell(row=r, column=1, value=o)
        rc.fill = HDR_FILL
        rc.font = HDR_FONT
        rc.border = BORDER
        for j in range(n):
            cell = ws.cell(row=r, column=2 + j)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            if i == j:
                cell.value = "-"
                continue
            p = grid.get((i, j))
            if not p or not p.get("ok"):
                cell.value = "n/a"
                cell.font = Font(color="808080")
                continue
            diff = p["res"]["difference"]
            cell.value = diff
            cell.number_format = NUM
            cell.fill = OK_FILL if diff == 0 else BAD_FILL
    ws.freeze_panes = ws.cell(row=top + 1, column=2)


def _pairs_sheet(ws, m):
    headers = ["Person A", "Person B", "Contra a/c (A)", "Contra a/c (B)",
               "Opening A", "Opening B", "Closing A", "Closing B",
               "Difference", "Status", "Matched", "Exc A", "Exc B"]
    _header_row(ws, 1, headers,
                widths=[20, 20, 24, 24, 14, 14, 14, 14, 14, 16, 9, 7, 7])
    row = 2
    for p in m["pairs"]:
        _txt(ws, row, 1, p["a_owner"])
        _txt(ws, row, 2, p["b_owner"])
        if p.get("ok"):
            r = p["res"]
            _txt(ws, row, 3, ", ".join(x.name for x in r["a_contra"]))
            _txt(ws, row, 4, ", ".join(x.name for x in r["b_contra"]))
            _money(ws, row, 5, round(r["a_open"], 2))
            _money(ws, row, 6, round(r["b_open"], 2))
            _money(ws, row, 7, round(r["a_close"], 2))
            _money(ws, row, 8, round(r["b_close"], 2))
            dc = _money(ws, row, 9, r["difference"])
            dc.fill = OK_FILL if r["difference"] == 0 else BAD_FILL
            status = "Ties" if r["difference"] == 0 else "Out of balance"
            if r["difference"] == 0 and (r["a_exc"] or r["b_exc"]):
                status = "Ties (exceptions net off)"
            _txt(ws, row, 10, status)
            for col, val in ((11, len(r["pairs"])), (12, len(r["a_exc"])),
                             (13, len(r["b_exc"]))):
                c = ws.cell(row=row, column=col, value=val)
                c.border = BORDER
                c.alignment = Alignment(horizontal="center")
        else:
            _txt(ws, row, 3, "n/a")
            _txt(ws, row, 10, f"n/a: {p['err']}")
        row += 1


def _all_exceptions_sheet(ws, m):
    headers = ["Pair", "Recorded in", "Date", "Amount", "Description",
               "Contra account", "Best probable posting in the other book"]
    _header_row(ws, 1, headers, widths=[30, 18, 12, 14, 34, 22, 56])
    row = 2
    for p in m["pairs"]:
        if not p.get("ok"):
            continue
        r = p["res"]
        pair_label = f"{p['a_owner']} <-> {p['b_owner']}"
        for who, excs, suggs in (
            (r["book_a"].owner, r["a_exc"], r["a_suggestions"]),
            (r["book_b"].owner, r["b_exc"], r["b_suggestions"]),
        ):
            for mv in excs:
                _txt(ws, row, 1, pair_label)
                _txt(ws, row, 2, who)
                _txt(ws, row, 3, str(mv.d))
                _money(ws, row, 4, round(mv.amount, 2))
                _txt(ws, row, 5, mv.desc)
                _txt(ws, row, 6, mv.account_name)
                s = suggs.get(id(mv), [])
                if s:
                    b = s[0]
                    tag = "bank/cash" if b.liquid else "ledger"
                    gap = "same day" if b.day_gap == 0 else f"{b.day_gap}d away"
                    hint = f"{b.d} {b.amount:,.2f} -> {b.account_path} [{tag}, {gap}]"
                else:
                    hint = "No candidate (unrecorded, or belongs to a related entity's book)."
                _txt(ws, row, 7, hint)
                row += 1
    if row == 2:
        ws.cell(row=2, column=1, value="No exceptions across any reconciled pair.")


def write_matrix_workbook(m: dict, out_path: str) -> str:
    wb = Workbook()
    _matrix_sheet(wb.active, m)
    wb.active.title = "Matrix"
    _pairs_sheet(wb.create_sheet("Pairs"), m)
    _all_exceptions_sheet(wb.create_sheet("All Exceptions"), m)
    wb.save(out_path)
    return out_path
