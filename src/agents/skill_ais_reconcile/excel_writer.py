"""
excel_writer.py — writes a standalone .xlsx AIS-internal reconciliation
report from an AisRecoReport (see reconcile.py).

Pure output layer: takes an already-built report and a destination path, and
writes exactly that one file. No knowledge of how the report was produced,
no decision-making about where it should live -- the tools/agent layer picks
`out_path`.

Styling follows the same openpyxl conventions used elsewhere in this repo
(see src/agents/skill_itr_workbook/scripts/presentation.py: Arial 10pt, INR
"#,##,##0.00" number format, red-fill/dark-red-font "error" styling for
flagged cells using the same FFC7CE/9C0006 colors). Reimplemented locally
here rather than importing presentation.py directly -- that module is an
ITR-workbook-internal script (no other skill imports it), not a shared
library, and pulling it in would couple this skill to ITR-workbook's
Workbook-model-specific helpers for no benefit.
"""
from __future__ import annotations

from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from agents.skill_ais_reconcile.reconcile import AisRecoReport, ElementReco
from agents.skill_ais_reconcile.reconcile_26as import Ais26asReport, CategoryTieOut, QuarterTieOut
from agents.skill_ais_reconcile.reconcile_books import AisBooksReport, CategoryBooksTieOut
from agents.skill_ais_reconcile.feedback import FeedbackSuggestion

FONT_NAME = "Arial"
INR_FORMAT = "#,##,##0.00"

_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_RED_FONT_COLOR = "9C0006"
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_GREEN_FONT_COLOR = "006100"
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# Sections written first, in this order, when present; any other/future
# sectionKey still gets a sheet -- just appended afterwards, alphabetically.
_DETAIL_SECTIONS_ORDER = ("tdsTcs", "sft", "other-info")

_DETAIL_COLUMNS = [
    "Category", "Info source", "Detail sum (l1)", "Reported (l2 Amount)",
    "Delta (detail - reported)", "Derived (l2 Derived Amount)",
    "Count (l1)", "Count (l2)", "Status breakdown",
]
_DETAIL_DELTA_COL = 5
_DETAIL_MONEY_COLS = (3, 4, 5, 6)

_FLAGS_COLUMNS = [
    "Section", "Category", "Info source", "Detail sum", "Reported", "Delta",
    "Derived", "Derivation flag", "Count mismatch", "Status breakdown",
]
_FLAGS_MONEY_COLS = (4, 5, 6, 7)


def _font(*, bold: bool = False, color: str | None = None, size: int = 10) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)


def _set(ws: Worksheet, row: int, col: int, value: Any, *, bold: bool = False,
         number_format: str | None = None, fill: PatternFill | None = None,
         color: str | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, color=color)
    if number_format:
        cell.number_format = number_format
    if fill is not None:
        cell.fill = fill
    return cell


def _num(value: Any):
    """Cell-safe numeric value: Decimal -> float (openpyxl handles Decimal
    fine, but float keeps this predictable across openpyxl versions),
    anything else passed through as-is (e.g. a raw "Count" string)."""
    if isinstance(value, Decimal):
        return float(value)
    return value


def _autosize(ws: Worksheet, ncols: int, nrows: int, min_width: float = 10.0,
              cap: float = 60.0) -> None:
    for col in range(1, ncols + 1):
        longest = min_width
        for row in range(1, nrows + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            longest = max(longest, len(str(v)) + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(longest, cap)


def _status_breakdown_text(el: ElementReco) -> str:
    if not el.status_breakdown:
        return ""
    return ", ".join(f"{k}: {v}" for k, v in sorted(el.status_breakdown.items()))


def _flag_magnitude(el: ElementReco) -> Decimal:
    """"Worst delta" ranking key for the Flags sheet: the largest magnitude
    among whichever deltas this element actually has (detail-vs-reported,
    and/or derived-vs-reported). Elements flagged only for a count mismatch
    (no numeric delta at all) sort last, at 0."""
    mags = []
    if el.delta_detail_vs_reported is not None:
        mags.append(abs(el.delta_detail_vs_reported))
    if el.flag_derivation and el.derived is not None and el.reported is not None:
        mags.append(abs(el.derived - el.reported))
    return max(mags) if mags else Decimal("0")


def _write_banner(ws: Worksheet, row: int, ncols: int, flagged_count: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols, 4))
    cell = ws.cell(row=row, column=1)
    if flagged_count > 0:
        plural = "S" if flagged_count != 1 else ""
        cell.value = f"{flagged_count} DIFFERENCE{plural} FOUND"
        cell.fill = _RED_FILL
        cell.font = _font(bold=True, color=_RED_FONT_COLOR, size=12)
    else:
        cell.value = "NO DIFFERENCES"
        cell.fill = _GREEN_FILL
        cell.font = _font(bold=True, color=_GREEN_FONT_COLOR, size=12)
    cell.alignment = Alignment(horizontal="center")
    return row + 2


def _write_summary_sheet(wb: Workbook, report: AisRecoReport) -> None:
    ws = wb.create_sheet("Summary")
    row = 1
    row = _write_banner(ws, row, 6, report.flagged_element_count)

    for label, value in (
        ("PAN", report.logged_in_pan),
        ("Name", report.taxpayer_name),
        ("AIS JSON version", report.json_version),
        ("Download date", report.download_date),
    ):
        _set(ws, row, 1, label, bold=True)
        _set(ws, row, 2, value)
        row += 1
    row += 1

    _set(ws, row, 1, "Section", bold=True, fill=_HEADER_FILL)
    _set(ws, row, 2, "Total reported (l2 Amount)", bold=True, fill=_HEADER_FILL)
    row += 1
    for section_key in sorted(report.total_reported_by_section):
        _set(ws, row, 1, section_key)
        _set(ws, row, 2, _num(report.total_reported_by_section[section_key]),
             number_format=INR_FORMAT)
        row += 1
    _set(ws, row, 1, "Total TDS credit (tdsTcs l1)", bold=True)
    _set(ws, row, 2, _num(report.total_tds_credit), number_format=INR_FORMAT)
    row += 2

    _set(ws, row, 1, "Elements checked", bold=True)
    _set(ws, row, 2, len(report.elements))
    row += 1
    _set(ws, row, 1, "Elements flagged", bold=True)
    _set(ws, row, 2, report.flagged_element_count)
    row += 1

    ws.freeze_panes = "A2"
    _autosize(ws, 6, row)


def _write_detail_sheet(wb: Workbook, section_key: str, elements: list[ElementReco]) -> None:
    title = (section_key or "unknown")[:31]
    ws = wb.create_sheet(title)
    for col, label in enumerate(_DETAIL_COLUMNS, start=1):
        _set(ws, 1, col, label, bold=True, fill=_HEADER_FILL)

    row = 2
    for el in elements:
        values = [
            el.category, el.info_src_id,
            _num(el.detail_sum), _num(el.reported),
            _num(el.delta_detail_vs_reported), _num(el.derived),
            el.l1_row_count, _num(el.l2_count),
            _status_breakdown_text(el),
        ]
        fill = _RED_FILL if el.flagged else None
        color = _RED_FONT_COLOR if el.flagged else None
        for col, value in enumerate(values, start=1):
            nf = INR_FORMAT if col in _DETAIL_MONEY_COLS else None
            _set(ws, row, col, value, number_format=nf, fill=fill, color=color,
                 bold=el.flagged)
        row += 1

    last_row = row - 1
    if last_row >= 2:
        col_letter = get_column_letter(_DETAIL_DELTA_COL)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{last_row}",
            CellIsRule(operator="notEqual", formula=["0"], fill=_RED_FILL),
        )

    ws.freeze_panes = "A2"
    _autosize(ws, len(_DETAIL_COLUMNS), last_row)


def _write_flags_sheet(wb: Workbook, report: AisRecoReport) -> None:
    ws = wb.create_sheet("Flags")
    flagged = sorted((el for el in report.elements if el.flagged),
                      key=_flag_magnitude, reverse=True)

    for col, label in enumerate(_FLAGS_COLUMNS, start=1):
        _set(ws, 1, col, label, bold=True, fill=_HEADER_FILL)

    row = 2
    for el in flagged:
        values = [
            el.section_key, el.category, el.info_src_id,
            _num(el.detail_sum), _num(el.reported),
            _num(el.delta_detail_vs_reported), _num(el.derived),
            "YES" if el.flag_derivation else "",
            "YES" if el.flag_count_mismatch else "",
            _status_breakdown_text(el),
        ]
        for col, value in enumerate(values, start=1):
            nf = INR_FORMAT if col in _FLAGS_MONEY_COLS else None
            _set(ws, row, col, value, number_format=nf, fill=_RED_FILL,
                 color=_RED_FONT_COLOR, bold=True)
        row += 1

    if not flagged:
        _set(ws, row, 1, "No flagged elements.", bold=True)
        row += 1

    ws.freeze_panes = "A2"
    _autosize(ws, len(_FLAGS_COLUMNS), row - 1)


_TIEOUT_QUARTER_COLUMNS = ["Quarter", "AIS TDS credit", "26AS TDS deposited", "Delta"]
_TIEOUT_QUARTER_MONEY_COLS = (2, 3, 4)

_TIEOUT_CATEGORY_COLUMNS = ["Category", "AIS TDS credit", "26AS TDS deposited", "Delta"]
_TIEOUT_CATEGORY_MONEY_COLS = (2, 3, 4)


def _write_26as_tieout_sheet(wb: Workbook, report: Ais26asReport) -> None:
    ws = wb.create_sheet("26AS Tie-out")
    row = 1
    row = _write_banner(ws, row, 4, 1 if report.flag_aggregate_mismatch else 0)

    # -- Aggregate tie-out -------------------------------------------------
    _set(ws, row, 1, "Aggregate TDS-credit tie-out", bold=True, fill=_HEADER_FILL)
    row += 1
    agg_flag = report.flag_aggregate_mismatch
    agg_fill = _RED_FILL if agg_flag else None
    agg_color = _RED_FONT_COLOR if agg_flag else None
    for label, value in (
        ("Total AIS TDS credit", report.total_ais_tds_credit),
        ("Total 26AS TDS deposited", report.total_26as_tds_deposited),
        ("Delta (AIS - 26AS)", report.delta_aggregate),
    ):
        _set(ws, row, 1, label, bold=agg_flag)
        _set(ws, row, 2, _num(value), number_format=INR_FORMAT,
             fill=agg_fill, color=agg_color, bold=agg_flag)
        row += 1
    row += 1

    # -- Per-quarter tie-out -------------------------------------------------
    _set(ws, row, 1, "Per-quarter tie-out", bold=True, fill=_HEADER_FILL)
    row += 1
    for col, label in enumerate(_TIEOUT_QUARTER_COLUMNS, start=1):
        _set(ws, row, col, label, bold=True, fill=_HEADER_FILL)
    row += 1
    quarter_header_row = row - 1
    quarter_first_row = row
    for q in report.quarters:
        fill = _RED_FILL if q.flagged else None
        color = _RED_FONT_COLOR if q.flagged else None
        values = [f"Q{q.quarter}", _num(q.ais_tds_credit), _num(q.as26_tds_deposited),
                  _num(q.delta)]
        for col, value in enumerate(values, start=1):
            nf = INR_FORMAT if col in _TIEOUT_QUARTER_MONEY_COLS else None
            _set(ws, row, col, value, number_format=nf, fill=fill, color=color,
                 bold=q.flagged)
        row += 1
    quarter_last_row = row - 1
    if quarter_last_row >= quarter_first_row:
        delta_col_letter = get_column_letter(_TIEOUT_QUARTER_COLUMNS.index("Delta") + 1)
        ws.conditional_formatting.add(
            f"{delta_col_letter}{quarter_first_row}:{delta_col_letter}{quarter_last_row}",
            CellIsRule(operator="notEqual", formula=["0"], fill=_RED_FILL),
        )
    row += 1

    # -- Per-income-category tie-out -----------------------------------------
    _set(ws, row, 1,
         "Per-income-category tie-out (interest/dividend/other; requires "
         "tds_sections -- omitted here means it wasn't supplied)", bold=True)
    row += 1
    for col, label in enumerate(_TIEOUT_CATEGORY_COLUMNS, start=1):
        _set(ws, row, col, label, bold=True, fill=_HEADER_FILL)
    row += 1
    category_first_row = row
    for c in report.categories:
        fill = _RED_FILL if c.flagged else None
        color = _RED_FONT_COLOR if c.flagged else None
        values = [c.category, _num(c.ais_credit), _num(c.as26_deposited), _num(c.delta)]
        for col, value in enumerate(values, start=1):
            nf = INR_FORMAT if col in _TIEOUT_CATEGORY_MONEY_COLS else None
            _set(ws, row, col, value, number_format=nf, fill=fill, color=color,
                 bold=c.flagged)
        row += 1
    category_last_row = row - 1
    if category_last_row >= category_first_row:
        delta_col_letter = get_column_letter(_TIEOUT_CATEGORY_COLUMNS.index("Delta") + 1)
        ws.conditional_formatting.add(
            f"{delta_col_letter}{category_first_row}:{delta_col_letter}{category_last_row}",
            CellIsRule(operator="notEqual", formula=["0"], fill=_RED_FILL),
        )
    elif not report.categories:
        _set(ws, row, 1, "(no category tie-out -- tds_sections not supplied)")
        row += 1

    ws.freeze_panes = f"A{quarter_header_row}"
    _autosize(ws, max(len(_TIEOUT_QUARTER_COLUMNS), len(_TIEOUT_CATEGORY_COLUMNS)), row - 1)


_BOOKS_CATEGORY_COLUMNS = ["Category", "AIS income", "Books income", "Delta"]
_BOOKS_CATEGORY_MONEY_COLS = (2, 3, 4)


def _write_books_sheet(wb: Workbook, report: AisBooksReport) -> None:
    ws = wb.create_sheet("Books Reconciliation")
    row = 1
    flagged_count = len(report.flagged_categories) + (1 if report.flag_tds_mismatch else 0)
    row = _write_banner(ws, row, 4, flagged_count)

    # -- TDS credit tie-out -------------------------------------------------
    _set(ws, row, 1, "TDS credit tie-out (AIS vs books)", bold=True, fill=_HEADER_FILL)
    row += 1
    tds_flag = report.flag_tds_mismatch
    tds_fill = _RED_FILL if tds_flag else None
    tds_color = _RED_FONT_COLOR if tds_flag else None
    for label, value in (
        ("AIS TDS credit", report.ais_tds_credit),
        ("Books TDS credit", report.books_tds_credit),
        ("Delta (AIS - books)", report.delta_tds),
    ):
        _set(ws, row, 1, label, bold=tds_flag)
        _set(ws, row, 2, _num(value), number_format=INR_FORMAT,
             fill=tds_fill, color=tds_color, bold=tds_flag)
        row += 1
    row += 1

    # -- Per-category income tie-out ----------------------------------------
    _set(ws, row, 1, "Per-category income tie-out (interest/dividend/salary)",
         bold=True, fill=_HEADER_FILL)
    row += 1
    for col, label in enumerate(_BOOKS_CATEGORY_COLUMNS, start=1):
        _set(ws, row, col, label, bold=True, fill=_HEADER_FILL)
    row += 1
    category_header_row = row - 1
    category_first_row = row
    for c in report.categories:
        fill = _RED_FILL if c.flagged else None
        color = _RED_FONT_COLOR if c.flagged else None
        values = [c.category, _num(c.ais_income), _num(c.books_income), _num(c.delta)]
        for col, value in enumerate(values, start=1):
            nf = INR_FORMAT if col in _BOOKS_CATEGORY_MONEY_COLS else None
            _set(ws, row, col, value, number_format=nf, fill=fill, color=color,
                 bold=c.flagged)
        row += 1
    category_last_row = row - 1
    if category_last_row >= category_first_row:
        delta_col_letter = get_column_letter(_BOOKS_CATEGORY_COLUMNS.index("Delta") + 1)
        ws.conditional_formatting.add(
            f"{delta_col_letter}{category_first_row}:{delta_col_letter}{category_last_row}",
            CellIsRule(operator="notEqual", formula=["0"], fill=_RED_FILL),
        )
    row += 1

    # -- Completeness call-outs (books is primary) ---------------------------
    _set(ws, row, 1, "AIS income not found in books (books is primary -- "
                      "check for missing postings)", bold=True, fill=_HEADER_FILL)
    row += 1
    if report.ais_income_not_in_books:
        for cat in report.ais_income_not_in_books:
            _set(ws, row, 1, cat, fill=_RED_FILL, color=_RED_FONT_COLOR, bold=True)
            row += 1
    else:
        _set(ws, row, 1, "(none)")
        row += 1
    row += 1

    _set(ws, row, 1, "Books income not found in AIS (informational)",
         bold=True, fill=_HEADER_FILL)
    row += 1
    if report.books_income_not_in_ais:
        for cat in report.books_income_not_in_ais:
            _set(ws, row, 1, cat)
            row += 1
    else:
        _set(ws, row, 1, "(none)")
        row += 1
    row += 1

    _set(ws, row, 1, "Untagged income accounts with nonzero FY activity",
         bold=True)
    _set(ws, row, 2, report.untagged_income_account_count)
    row += 1

    ws.freeze_panes = f"A{category_header_row}"
    _autosize(ws, len(_BOOKS_CATEGORY_COLUMNS), row - 1)


_FEEDBACK_COLUMNS = ["Section", "Category", "Info source", "Discrepancy",
                     "Suggested action", "Confidence", "Rationale"]
_FEEDBACK_RATIONALE_COL = 7
_ADVISORY_NOTE = ("ADVISORY ONLY -- review each item before acting on the portal; "
                   "nothing here is auto-submitted.")


def _write_feedback_sheet(wb: Workbook, suggestions: list[FeedbackSuggestion]) -> None:
    ws = wb.create_sheet("Feedback Suggestions")
    row = 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                    end_column=max(len(_FEEDBACK_COLUMNS), 4))
    note_cell = ws.cell(row=row, column=1, value=_ADVISORY_NOTE)
    note_cell.fill = _HEADER_FILL
    note_cell.font = _font(bold=True, size=11)
    note_cell.alignment = Alignment(horizontal="center", wrap_text=True)
    row += 2

    for col, label in enumerate(_FEEDBACK_COLUMNS, start=1):
        _set(ws, row, col, label, bold=True, fill=_HEADER_FILL)
    row += 1
    header_row = row - 1

    if not suggestions:
        _set(ws, row, 1, "No feedback suggestions -- no flagged discrepancies to review.")
        row += 1
    for s in suggestions:
        values = [s.section_key, s.category, s.info_src_id, s.delta_context,
                  s.suggested_action, s.confidence, s.rationale]
        for col, value in enumerate(values, start=1):
            cell = _set(ws, row, col, value)
            if col == _FEEDBACK_RATIONALE_COL:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize(ws, len(_FEEDBACK_COLUMNS), row - 1, cap=50.0)
    ws.column_dimensions[get_column_letter(_FEEDBACK_RATIONALE_COL)].width = 60.0


def write_reco_workbook(report: AisRecoReport, out_path: str,
                         report_26as: Ais26asReport | None = None,
                         report_books: AisBooksReport | None = None,
                         suggestions: list[FeedbackSuggestion] | None = None) -> None:
    """Write ONE standalone .xlsx from an AisRecoReport:
      - "Summary": per-entity header block, section rollups, TDS credit
        total, and a prominent flagged-count banner (red "N DIFFERENCES
        FOUND" / green "NO DIFFERENCES").
      - one detail sheet per section present in the report's elements
        (tdsTcs/sft/other-info first if present, any other sectionKey
        appended alphabetically), with a red fill + bold on flagged rows and
        conditional formatting on the delta column.
      - "Flags": only the flagged elements, worst-delta-first.
      - "Books Reconciliation" (Phase C, only when `report_books` is given,
        placed immediately after "Summary" since this is the PRIMARY
        reconciliation): AIS-vs-books TDS-credit tie-out, a per-category
        income tie-out (interest/dividend/salary), and two completeness
        call-outs -- "AIS income not found in books" (the loud one: books is
        primary, so this means a posting may be missing) and "Books income
        not found in AIS" (informational).
      - "26AS Tie-out" (Phase B, only when `report_26as` is given): aggregate
        AIS-vs-26AS TDS-credit banner, a per-quarter table (red fill on any
        quarter whose delta exceeds tolerance), and a per-income-category
        table (interest/dividend/other; red fill on any category whose delta
        exceeds tolerance; empty when `tds_sections` wasn't supplied to
        reconcile_ais_vs_26as, since there's nothing to classify the 26AS
        side by).
      - "Feedback Suggestions" (Phase D, only when `suggestions` is given):
        ADVISORY-ONLY portal-feedback suggestions for a human to review --
        never auto-submitted, never claimed as definitive. Neutral (non-red)
        styling, since these are suggestions, not errors.
    The tools/agent layer decides `out_path`; this function only writes it.
    The existing single-argument call (`write_reco_workbook(report, path)`)
    keeps working unchanged -- `report_26as`, `report_books`, and
    `suggestions` all default to None and their sheets are simply omitted.
    """
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet -- we create our own

    _write_summary_sheet(wb, report)

    if report_books is not None:
        _write_books_sheet(wb, report_books)

    by_section: dict[str, list[ElementReco]] = {}
    for el in report.elements:
        by_section.setdefault(el.section_key, []).append(el)

    ordered_sections = [s for s in _DETAIL_SECTIONS_ORDER if s in by_section]
    ordered_sections += sorted(s for s in by_section if s not in _DETAIL_SECTIONS_ORDER)
    for section_key in ordered_sections:
        _write_detail_sheet(wb, section_key, by_section[section_key])

    _write_flags_sheet(wb, report)

    if report_26as is not None:
        _write_26as_tieout_sheet(wb, report_26as)

    if suggestions is not None:
        _write_feedback_sheet(wb, suggestions)

    wb.save(out_path)
