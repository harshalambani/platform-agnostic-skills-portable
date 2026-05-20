"""
tools.py — LangChain tools for the 26AS skill.
The LLM calls these; they delegate to the deterministic extraction script.
"""
import subprocess
import sys
from pathlib import Path
from langchain_core.tools import tool

SCRIPT = Path(__file__).parent / "scripts" / "extract_26as_to_xlsx.py"


@tool
def extract_26as(pdf_path: str, output_path: str) -> str:
    """
    Extract a Form 26AS PDF into an Excel workbook with one sheet per Part.
    Returns a summary (assessee name, deductor count, transaction count) or an error message.
    Call this first before verify_26as_output.
    """
    result = subprocess.run(
        [sys.executable, str(SCRIPT), pdf_path, output_path],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        return f"ERROR: {result.stderr.strip()}"
    return result.stdout.strip() or "Extraction complete."


@tool
def verify_26as_output(xlsx_path: str) -> str:
    """
    Run a per-deductor reconciliation check on the extracted 26AS Excel file.
    Checks that transaction-level totals match the deductor header totals for
    Amount Paid/Credited, Tax Deducted, and TDS Deposited.
    Returns 'OK' if all deductors reconcile, or a description of mismatches.
    Call this after extract_26as.
    """
    try:
        import openpyxl
        from collections import defaultdict

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        if "Part I" not in wb.sheetnames:
            return "ERROR: Part I sheet not found in workbook."

        ws = wb["Part I"]
        groups = defaultdict(lambda: [0.0, 0.0, 0.0, 0.0, 0.0, 0.0])

        for r in range(4, ws.max_row + 1):
            sr = ws.cell(r, 1).value
            if not isinstance(sr, int):
                continue
            groups[sr][0] = float(ws.cell(r, 4).value or 0)   # header: total amt
            groups[sr][1] = float(ws.cell(r, 5).value or 0)   # header: total tax
            groups[sr][2] = float(ws.cell(r, 6).value or 0)   # header: total tds
            groups[sr][3] += float(ws.cell(r, 13).value or 0) # txn: amt
            groups[sr][4] += float(ws.cell(r, 14).value or 0) # txn: tax
            groups[sr][5] += float(ws.cell(r, 15).value or 0) # txn: tds

        mismatches = []
        for sr, v in groups.items():
            if abs(v[0] - v[3]) >= 0.01 or abs(v[1] - v[4]) >= 0.01 or abs(v[2] - v[5]) >= 0.01:
                mismatches.append(
                    f"Deductor Sr.No {sr}: "
                    f"Amt header={v[0]} txn_sum={v[3]:.2f}, "
                    f"Tax header={v[1]} txn_sum={v[4]:.2f}, "
                    f"TDS header={v[2]} txn_sum={v[5]:.2f}"
                )

        if mismatches:
            return "RECONCILIATION ERRORS:\n" + "\n".join(mismatches)
        return f"OK — all {len(groups)} deductors reconcile."

    except Exception as e:
        return f"ERROR during verification: {e}"
