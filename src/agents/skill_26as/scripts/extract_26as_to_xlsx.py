"""
Form 26AS (Indian Income Tax Annual Tax Statement, TRACES) -> xlsx.

Given a 26AS PDF, parse every Part (I..X) and emit one sheet per Part in a
single .xlsx workbook.

Design choices (baked in — override by asking explicitly):
  * Flat layout for Part I: one row per transaction, deductor info
    (name, TAN, header totals) repeated on every row of that deductor's block.
    This is the layout that filters/pivots cleanly.
  * Inline sub-total row (bold, shaded) after each deductor's transactions in
    Part I, with SUM formulas over that deductor's rows only.
  * Grand Total row at bottom summing the sub-totals (never the raw rows —
    avoids double-counting).
  * A sheet per Part, even empty ones, so the output is predictable (10 sheets).
    Empty Parts get their column headers + a single 'No Transactions Present'
    banner row, matching the PDF.
  * Title band + assessee metadata strip (Name, PAN, FY, AY, Data updated till)
    at the top of each sheet.
  * Arial 10, bordered cells, thousands-formatted numbers, frozen header row.

Usage:
  python extract_26as_to_xlsx.py <input.pdf> <output.xlsx>

Requires pdftotext (poppler-utils) and openpyxl.
"""

from __future__ import annotations

import re
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# This script always runs as a standalone subprocess (spawned via
# ``sys.executable``, both in dev and in the PyInstaller-frozen build) -- it
# never inherits a caller's sys.path/PYTHONPATH. Bootstrap our own path to
# ``src`` (or _MEIPASS, in frozen mode) so ``agents._native_resolve`` is
# importable regardless of how this script was invoked. Same convention as
# skill_hsbc/scripts/parse_tsv.py.
sys.path.insert(0, str(Path(__file__).resolve().parents[3]))
from agents._native_resolve import resolve_pdftotext, WrongPdftextFlavourError  # noqa: E402


# -------------------- regexes --------------------

# A deductor "data" row: Sr.No.  <Name (spaces ok)>  TAN  Total-Amt  Total-Tax  Total-TDS
# TAN is exactly 10 chars: 4 letters + 5 digits + 1 letter.
TAN_RE = r"[A-Z]{4}\d{5}[A-Z]"
NUM_RE = r"-?\d[\d,]*\.\d{2}"
DATE_RE = r"\d{2}-[A-Za-z]{3}-\d{4}"
STATUS_RE = r"[UMPFOZ]"
# PAN is 5 letters + 4 digits + 1 letter.
PAN_RE = r"[A-Z]{5}\d{4}[A-Z]"

DEDUCTOR_RX = re.compile(
    rf"^\s*(\d+)\s+(.+?)\s+({TAN_RE})\s+({NUM_RE})\s+({NUM_RE})\s+({NUM_RE})\s*$"
)

# A transaction row for Part I / VI:
#  Sr.No. Section TxnDate Status DateOfBooking Remarks Amount Tax TDS
TXN_RX = re.compile(
    rf"^\s*(\d+)\s+(\S+)\s+({DATE_RE})\s+({STATUS_RE})\s+({DATE_RE})\s+(\S+)\s+"
    rf"({NUM_RE})\s+({NUM_RE})\s+({NUM_RE})\s*$"
)

# A transaction row for Part II (15G/15H) — NO "Status of Booking" column,
# unlike Part I / Part VI:
#  Txn Sr.No. Section TxnDate DateOfBooking Remarks Amount Tax TDS
TXN2_RX = re.compile(
    rf"^\s*(\d+)\s+(\S+)\s+({DATE_RE})\s+({DATE_RE})\s+(\S+)\s+"
    rf"({NUM_RE})\s+({NUM_RE})\s+({NUM_RE})\s*$"
)

# Part VIII deductee data row:
#  Sr.No. AckNum Name PAN TxnDate TotalTxnAmt TotalTDS TotalOther
PART8_DEDUCTEE_RX = re.compile(
    rf"^\s*(\d+)\s+(\S+)\s+(.+?)\s+({PAN_RE})\s+({DATE_RE})\s+"
    rf"({NUM_RE})\s+({NUM_RE})\s+({NUM_RE})\s*$"
)

# Part VIII TDS certificate / txn row:
#  Sr.No. CertNum Section DateOfDeposit Status DateOfBooking DemandPayment TDSDeposited Other
PART8_TDS_RX = re.compile(
    rf"^\s*(\d+)\s+(\S+)\s+(\S+)\s+({DATE_RE})\s+({STATUS_RE})\s+({DATE_RE})\s+"
    rf"(\S+)\s+({NUM_RE})\s+({NUM_RE})\s*$"
)

PART_HEADER_RX = re.compile(r"^\s*PART[- ]?([IVX]+)\b(.*)$")
COL_HEADER_RX = re.compile(r"^\s*Sr\.?\s*No\.?\b")


def clean_num(s: str) -> float:
    return float(s.replace(",", ""))


# -------------------- data classes --------------------


@dataclass
class P1Txn:
    sr: int
    section: str
    txn_date: str
    status: str
    date_booking: str
    remarks: str
    amount: float
    tax: float
    tds: float


@dataclass
class P1Deductor:
    sr: int
    name: str
    tan: str
    tot_amt: float
    tot_tax: float
    tot_tds: float
    txns: list[P1Txn] = field(default_factory=list)


@dataclass
class P2Txn:
    sr: int
    section: str
    txn_date: str
    date_booking: str
    remarks: str
    amount: float
    tax: float
    tds: float


@dataclass
class P2Deductor:
    sr: int
    name: str
    tan: str
    tot_amt: float
    tot_tax: float
    tot_tds: float
    txns: list[P2Txn] = field(default_factory=list)


@dataclass
class P8Row:
    sr: int
    ack_num: str
    name: str
    pan: str
    txn_date: str
    tot_txn_amt: float
    tot_tds: float
    tot_other: float
    cert_num: str = ""
    section: str = ""
    date_deposit: str = ""
    status: str = ""
    date_booking: str = ""
    demand: str = ""
    tds_deposited: float = 0.0
    other_deposited: float = 0.0


@dataclass
class Assessee:
    name: str = ""
    pan: str = ""
    fy: str = ""
    ay: str = ""
    data_updated: str = ""


# -------------------- parsing --------------------


def pdf_to_text(pdf_path: Path) -> tuple[str, str]:
    """Run pdftotext -layout and return (text, resolved_pdftotext_path).

    Resolves the vendored Poppler pdftotext by absolute path (not a bare name
    trusted to PATH) and verifies it is actually Poppler before use — see
    agents/_native_resolve.py. Raises WrongPdftextFlavourError loudly (naming
    what was found and where) rather than silently parsing with the wrong
    tool, which is how the earlier "0 deductors" bug happened.

    The resolved path is returned (not just the text) so that a later
    zero-extraction failure (see check_extraction_not_vacuous) can name
    exactly which binary produced the (missing) text, without re-resolving.
    """
    try:
        pdftotext_path = resolve_pdftotext()
    except (FileNotFoundError, WrongPdftextFlavourError) as e:
        raise RuntimeError(f"pdftotext resolution failed: {e}") from e
    out = subprocess.check_output([pdftotext_path, "-layout", str(pdf_path), "-"])
    return out.decode("utf-8", errors="replace"), pdftotext_path


def parse_assessee(text: str) -> Assessee:
    a = Assessee()
    # PAN + FY + AY + name + data-updated line come from the front matter.
    m = re.search(r"Data updated till\s+(\d{2}-[A-Za-z]{3}-\d{4})", text)
    if m:
        a.data_updated = m.group(1)
    m = re.search(r"Permanent Account Number \(PAN\)\s+(" + PAN_RE + r")", text)
    if m:
        a.pan = m.group(1)
    m = re.search(r"Financial Year\s+(\d{4}-\d{2})", text)
    if m:
        a.fy = m.group(1)
    m = re.search(r"Assessment Year\s+(\d{4}-\d{2})", text)
    if m:
        a.ay = m.group(1)
    m = re.search(r"Name of Assessee\s+([^\n]+)", text)
    if m:
        a.name = m.group(1).strip()
    return a


def split_by_parts(text: str) -> dict[str, str]:
    """Return a dict mapping 'I' / 'II' / ... -> the text body of that Part.

    We split on lines matching 'PART-<roman>' or 'PART <roman>'.
    """
    parts: dict[str, str] = {}
    current_key: Optional[str] = None
    buf: list[str] = []
    for line in text.splitlines():
        m = PART_HEADER_RX.match(line)
        if m:
            # flush previous
            if current_key is not None:
                parts[current_key] = "\n".join(buf)
            current_key = m.group(1).upper()
            buf = []
            continue
        if current_key is not None:
            buf.append(line)
    if current_key is not None:
        parts[current_key] = "\n".join(buf)
    return parts


P1_HEADER_TOKENS = ("Name of Deductor", "TAN of Deductor", "Section",
                    "Transaction Date", "Remarks", "Credited", "Deposited",
                    "Tax Deducted", "TDS Deposited", "Status of Booking",
                    "Date of Booking", "Amount Paid")

# Part VI (TCS) has the same row geometry as Part I — a collector header row
# (Sr, Name, TAN, three totals) followed by transaction rows in the same nine
# columns — so the same two regexes parse it. Only the column-header wording
# differs, and those tokens must be recognised so a wrapped header line is
# never mistaken for a collector name-wrap tail.
P6_HEADER_TOKENS = ("Name of Collector", "TAN of Collector", "Section",
                    "Transaction Date", "Remarks", "Debited", "Deposited",
                    "Tax Collected", "TCS Deposited", "Status of Booking",
                    "Date of Booking", "Amount Paid")


def parse_part_i(body: str, header_tokens: tuple = P1_HEADER_TOKENS) -> list[P1Deductor]:
    """Parse Part I into a list of deductors each with their transactions.

    Strategy: walk lines. When we see a DEDUCTOR_RX match, open a new deductor.
    When we see a TXN_RX match, append to the current deductor.  Column-header
    rows and page banners get implicitly skipped because they don't match.
    Handles multi-line deductor names by buffering any line that doesn't match
    either regex and retrying it joined with the next non-matching line.

    Also parses Part VI (see parse_part_vi) — the layout is identical.
    """
    deductors: list[P1Deductor] = []
    current: Optional[P1Deductor] = None
    # Whether the *most recent* row we emitted was a deductor header (so the
    # next plain-text-only line is a name-wrap tail like "LIMITED" or "EAST").
    last_was_deductor = False

    HEADER_TOKENS = header_tokens

    for raw in body.splitlines():
        line = raw.rstrip()
        if not line.strip():
            continue
        # Skip page banners
        if "Assessee PAN:" in line and "Assessee Name:" in line:
            continue

        m = DEDUCTOR_RX.match(line)
        if m:
            current = P1Deductor(
                sr=int(m.group(1)),
                name=re.sub(r"\s+", " ", m.group(2).strip()),
                tan=m.group(3),
                tot_amt=clean_num(m.group(4)),
                tot_tax=clean_num(m.group(5)),
                tot_tds=clean_num(m.group(6)),
            )
            deductors.append(current)
            last_was_deductor = True
            continue

        m = TXN_RX.match(line)
        if m and current is not None:
            current.txns.append(
                P1Txn(
                    sr=int(m.group(1)),
                    section=m.group(2),
                    txn_date=m.group(3),
                    status=m.group(4),
                    date_booking=m.group(5),
                    remarks=m.group(6),
                    amount=clean_num(m.group(7)),
                    tax=clean_num(m.group(8)),
                    tds=clean_num(m.group(9)),
                )
            )
            last_was_deductor = False
            continue

        # Unmatched line — might be a name-wrap tail immediately after a
        # deductor row (e.g. the "LIMITED" suffix that spilled to line 2).
        # Only treat as a tail if it's pure text (no digits, not a header
        # label row) and we just emitted a deductor.
        stripped = line.strip()
        if (
            last_was_deductor
            and current is not None
            and stripped
            and not re.search(r"\d", stripped)
            and not any(tok in stripped for tok in HEADER_TOKENS)
        ):
            current.name = re.sub(r"\s+", " ", current.name + " " + stripped)
            # Name can wrap across multiple short lines — keep the flag on.
            continue

        # Anything else (column headers, etc.) clears the wrap state.
        last_was_deductor = False
    return deductors


def parse_part_vi(body: str) -> list[P1Deductor]:
    """Parse Part VI (Details of Tax Collected at Source) into collectors.

    Reuses the Part I walker: a TCS block is shaped exactly like a TDS block,
    with 'collector' in place of 'deductor' and Amount Paid/DEBITED in place of
    Amount Paid/Credited. The returned P1Deductor.tot_tax / P1Txn.tax carry the
    tax COLLECTED, and .tot_tds / .tds the TCS deposited.
    """
    return parse_part_i(body, P6_HEADER_TOKENS)


def parse_part_ii(body: str) -> list[P2Deductor]:
    """Parse Part II (Details of TDS for 15G / 15H) into deductors + their
    transactions.

    GEOMETRY WARNING: Part II is NOT a drop-in reuse of parse_part_i / TXN_RX.
    It matches Part I column-for-column on the deductor header row (Sr, Name,
    TAN, three totals — DEDUCTOR_RX applies unchanged), but its transaction
    rows carry NO "Status of Booking" column: Sr.No., Section, Transaction
    Date, Date of Booking, Remarks, Amount, Tax, TDS — eight fields, not
    Part I/VI's nine. Reusing TXN_RX here would silently misalign every field
    from Date of Booking onward, so a dedicated TXN2_RX is used instead. The
    deductor-name-wrap handling is otherwise identical to parse_part_i.
    """
    deductors: list[P2Deductor] = []
    current: Optional[P2Deductor] = None
    last_was_deductor = False

    for raw in body.splitlines():
        line = raw.rstrip()
        if not line.strip():
            continue
        if "Assessee PAN:" in line and "Assessee Name:" in line:
            continue

        m = DEDUCTOR_RX.match(line)
        if m:
            current = P2Deductor(
                sr=int(m.group(1)),
                name=re.sub(r"\s+", " ", m.group(2).strip()),
                tan=m.group(3),
                tot_amt=clean_num(m.group(4)),
                tot_tax=clean_num(m.group(5)),
                tot_tds=clean_num(m.group(6)),
            )
            deductors.append(current)
            last_was_deductor = True
            continue

        m = TXN2_RX.match(line)
        if m and current is not None:
            current.txns.append(
                P2Txn(
                    sr=int(m.group(1)),
                    section=m.group(2),
                    txn_date=m.group(3),
                    date_booking=m.group(4),
                    remarks=m.group(5),
                    amount=clean_num(m.group(6)),
                    tax=clean_num(m.group(7)),
                    tds=clean_num(m.group(8)),
                )
            )
            last_was_deductor = False
            continue

        stripped = line.strip()
        if (
            last_was_deductor
            and current is not None
            and stripped
            and not re.search(r"\d", stripped)
            and not any(tok in stripped for tok in P1_HEADER_TOKENS)
        ):
            current.name = re.sub(r"\s+", " ", current.name + " " + stripped)
            continue

        last_was_deductor = False
    return deductors


def parse_part_viii(body: str) -> list[P8Row]:
    """Parse Part VIII. Each deductee has one data row and one TDS row.

    We match them in order and zip.
    """
    deductees: list[tuple] = []
    tds_rows: list[tuple] = []
    for raw in body.splitlines():
        line = raw.rstrip()
        if not line.strip():
            continue
        if "Assessee PAN:" in line and "Assessee Name:" in line:
            continue
        if "Gross Total Across" in line:
            continue
        m = PART8_DEDUCTEE_RX.match(line)
        if m:
            deductees.append(m.groups())
            continue
        m = PART8_TDS_RX.match(line)
        if m:
            tds_rows.append(m.groups())

    rows: list[P8Row] = []
    for d, t in zip(deductees, tds_rows):
        rows.append(
            P8Row(
                sr=int(d[0]),
                ack_num=d[1],
                name=re.sub(r"\s+", " ", d[2].strip()),
                pan=d[3],
                txn_date=d[4],
                tot_txn_amt=clean_num(d[5]),
                tot_tds=clean_num(d[6]),
                tot_other=clean_num(d[7]),
                cert_num=t[1],
                section=t[2],
                date_deposit=t[3],
                status=t[4],
                date_booking=t[5],
                demand=t[6],
                tds_deposited=clean_num(t[7]),
                other_deposited=clean_num(t[8]),
            )
        )
    return rows


def is_empty_part(body: str) -> bool:
    return "No Transactions Present" in body


def reconcile_part_i(deductors: list[P1Deductor]) -> list[dict]:
    """Compare each deductor's transaction sub-total against the total the 26AS
    prints in its header row.

    The 26AS header line for every deductor carries a Total Amount Paid/Credited,
    Total Tax Deducted and Total TDS Deposited. Those are the document's own
    authoritative figures. The Excel sub-total is the sum of the transaction
    rows we parsed. If the two disagree, either a transaction row was dropped or
    duplicated during parsing, or the 26AS itself is internally inconsistent —
    either way the user needs to know.

    Returns one dict per deductor whose computed sub-total differs from its
    header total in any of the three fields (empty list == everything ties out).
    Comparison is to the paisa (rounded to 2 dp) to avoid float noise.
    """
    mismatches: list[dict] = []
    for d in deductors:
        comp_amt = round(sum(t.amount for t in d.txns), 2)
        comp_tax = round(sum(t.tax for t in d.txns), 2)
        comp_tds = round(sum(t.tds for t in d.txns), 2)
        diffs = []
        for field_name, header_val, comp_val in (
            ("Amount Paid/Credited", round(d.tot_amt, 2), comp_amt),
            ("Tax Deducted", round(d.tot_tax, 2), comp_tax),
            ("TDS Deposited", round(d.tot_tds, 2), comp_tds),
        ):
            if header_val != comp_val:
                diffs.append({
                    "field": field_name,
                    "header_total": header_val,
                    "computed_subtotal": comp_val,
                    "difference": round(comp_val - header_val, 2),
                })
        if diffs:
            mismatches.append({
                "sr": d.sr,
                "name": d.name,
                "tan": d.tan,
                "txns": len(d.txns),
                "fields": diffs,
            })
    return mismatches


class Empty26ASExtractionError(RuntimeError):
    """
    Raised when a 26AS PDF parses to ZERO rows across every part this script
    can populate (Part I, Part II, Part VI, Part VIII).

    Why this exists (tier 4 of the 26AS/Xpdf incident remediation — see
    agents/_native_resolve.py): tiers 1-3 stop the WRONG pdftotext being used.
    This tier is the backstop for every other way extraction can silently
    yield nothing (a future binary swap, a changed PDF layout, a regex that
    stops matching) — it must fail loudly instead of printing a reassuring
    "0/0 deductors OK — all sub-totals match" summary, which is vacuously
    true over an empty set and was exactly how the original incident hid
    itself from the user.

    Boundary this deliberately gets right: Part VI (TCS) alone being zero is
    completely normal — most taxpayers never have tax collected at source
    (see the real, correct output "Part VI: 0 collectors"). Only ALL FOUR
    populated-part counts being zero simultaneously is treated as impossible
    — a real Form 26AS always has at least a Part I entry (even a nil-TDS
    salary deductor prints a header row) or a Part II/VI/VIII row.
    """


def check_extraction_not_vacuous(stats: dict, text: str, pdftotext_path: str) -> None:
    """
    Raise Empty26ASExtractionError unless at least one of the four parts this
    script actually parses (I, II, VI, VIII) has a non-zero row count.

    Parts III/IV/V/VII/IX/X are never parsed for content by this script (they
    are always rendered via build_empty_part regardless of PDF content), so
    they carry no signal here and are intentionally excluded from the check.
    """
    counts = {
        "Part I (deductors)": stats["part_i_deductors"],
        "Part II (15G/15H deductors)": stats["part_ii_deductors"],
        "Part VI (TCS collectors)": stats["part_vi_collectors"],
        "Part VIII (buyer/tenant rows)": stats["part_viii_rows"],
    }
    if any(v > 0 for v in counts.values()):
        return

    chars_extracted = len(text)
    if len(text.strip()) == 0:
        text_diagnosis = "the binary produced NO text at all from this PDF"
    else:
        text_diagnosis = (
            "the binary DID produce text, but none of the row regexes matched — "
            "the PDF layout may have changed, or this is not a 26AS statement"
        )

    parts_summary = ", ".join(f"{k} = 0" for k in counts)
    raise Empty26ASExtractionError(
        "26AS extraction produced ZERO rows across every part that can "
        "legitimately be populated. A real Form 26AS always has at least one "
        "non-empty part; this is almost certainly a broken extraction, not a "
        "genuinely empty tax record. Diagnostics:\n"
        f"  Parts checked (all zero): {parts_summary}\n"
        f"  Resolved pdftotext binary: {pdftotext_path}\n"
        f"  Characters of text extracted from the PDF: {chars_extracted} "
        f"({text_diagnosis})"
    )


# -------------------- xlsx building --------------------

PART_DEFS = {
    "Part I":    "Details of Tax Deducted at Source",
    "Part II":   "Details of Tax Deducted at Source for 15G / 15H",
    "Part III":  "Details of Transactions under Proviso to section 194B / First Proviso to sub-section (1) of section 194R / Proviso to sub-section(1) of section 194S / Sub-section (2) of section 194BA",
    "Part IV":   "Details of Tax Deducted at Source u/s 194IA / 194IB / 194M / 194S (For Seller/Landlord of Property/Contractors or Professionals / Seller of Virtual Digital Asset)",
    "Part V":    "Details of Transactions under Proviso to sub-section (1) of section 194S as per Form-26QE (For Seller of Virtual Digital Asset)",
    "Part VI":   "Details of Tax Collected at Source",
    "Part VII":  "Details of Paid Refund (For which source is CPC TDS. For other details refer AIS at E-filing portal)",
    "Part VIII": "Details of Tax Deducted at Source u/s 194IA / 194IB / 194M / 194S (For Buyer/Tenant of Property / Person making payment to contractors or Professionals / Buyer of Virtual Digital Asset)",
    "Part IX":   "Details of Transactions / Demand Payments under Proviso to sub-section (1) of section 194S as per Form 26QE (For Buyer of Virtual Digital Asset)",
    "Part X":    "TDS/TCS Defaults (Processing of Statements)",
}

SHEET_ORDER = ["Part I", "Part II", "Part III", "Part IV", "Part V",
               "Part VI", "Part VII", "Part VIII", "Part IX", "Part X"]

# Map roman -> sheet name
ROMAN_TO_SHEET = {
    "I": "Part I", "II": "Part II", "III": "Part III", "IV": "Part IV",
    "V": "Part V", "VI": "Part VI", "VII": "Part VII", "VIII": "Part VIII",
    "IX": "Part IX", "X": "Part X",
}

EMPTY_HEADERS: dict[str, list[str]] = {
    "Part II": [
        "Sr.No.", "Name of Deductor", "TAN of Deductor",
        "Total Amount Paid/Credited", "Total Tax Deducted #", "Total TDS Deposited",
        "Txn Sr.No.", "Section", "Transaction Date", "Date of Booking",
        "Remarks", "Amount Paid/Credited", "Tax Deducted ##", "TDS Deposited",
    ],
    "Part III": [
        "Sr.No.", "Name of Deductor", "TAN of Deductor", "Total Amount Paid/Credited",
        "Txn Sr.No.", "Section", "Transaction Date", "Status of Booking",
        "Remarks", "Amount Paid/Credited",
    ],
    "Part IV": [
        "Sr.No.", "Acknowledgement Number", "Name of Deductor", "PAN of Deductor",
        "Transaction Date", "Total Transaction Amount", "Total TDS Deposited***",
        "TDS Certificate Number", "Section", "Date of Deposit",
        "Status of Booking", "Date of Booking", "Demand Payment", "TDS Deposited***",
    ],
    "Part V": [
        "Sr.No.", "Acknowledgement Number", "Name of Buyer", "PAN of Buyer",
        "Transaction Date", "Total Transaction Amount",
        "BSR Code", "Date of Deposit", "Challan Serial Number", "Total Tax Amount",
        "Status of Booking",
    ],
    "Part VI": [
        "Sr.No.", "Name of Collector", "TAN of Collector",
        "Total Amount Paid/Debited", "Total Tax Collected +", "Total TCS Deposited",
        "Txn Sr.No.", "Section", "Transaction Date", "Status of Booking",
        "Date of Booking", "Remarks", "Amount Paid/Debited",
        "Tax Collected ++", "TCS Deposited",
    ],
    "Part VII": [
        "Sr.No.", "Assessment Year", "Mode", "Refund Issued",
        "Nature of Refund", "Amount of Refund", "Interest",
        "Date of Payment", "Remarks",
    ],
    "Part IX": [
        "Sr.No.", "Acknowledgement Number", "Name of Seller", "PAN of Seller",
        "Transaction Date", "Total Transaction Amount",
        "Total Amount Deposited other than TDS",
        "BSR Code", "Date of Deposit", "Challan Serial Number", "Total Tax Amount",
        "Status of Booking", "Demand Payment",
    ],
    "Part X": [
        "Sr.No.", "Financial Year", "Short Payment", "Short Deduction/Collection",
        "Interest on TDS/TCS Payments Default",
        "Interest on TDS/TCS Deduction/Collection Default",
        "Late Filing Fee u/s 234E", "Interest u/s 220(2)", "Total Default",
        "TANs",
    ],
}

# Styling
THIN = Side(border_style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", start_color="1F4E78")
SUBHDR_FILL = PatternFill("solid", start_color="D9E2F3")
META_FILL = PatternFill("solid", start_color="F2F2F2")
SUBTOTAL_FILL = PatternFill("solid", start_color="FFF2CC")
GRAND_FILL = PatternFill("solid", start_color="1F4E78")

TITLE_FONT = Font(name="Arial", size=12, bold=True, color="FFFFFF")
META_FONT = Font(name="Arial", size=10, bold=True)
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
SUBTOTAL_FONT = Font(name="Arial", size=10, bold=True, color="1F4E78")
GRAND_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
RIGHT_WRAP = Alignment(horizontal="right", vertical="center", wrap_text=True)

NUM_FMT = '#,##0.00;(#,##0.00);"-"'


def _write_meta(ws, title: str, ncols: int, a: Assessee) -> None:
    c = ws.cell(row=1, column=1, value=f"{title} — {PART_DEFS[title]}")
    c.font = TITLE_FONT
    c.fill = HDR_FILL
    c.alignment = LEFT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.row_dimensions[1].height = 24

    meta = []
    if a.name: meta.append(("Assessee Name", a.name))
    if a.pan: meta.append(("PAN", a.pan))
    if a.fy: meta.append(("Financial Year", a.fy))
    if a.ay: meta.append(("Assessment Year", a.ay))
    if a.data_updated: meta.append(("Data updated till", a.data_updated))
    label = "  |  ".join(f"{k}: {v}" for k, v in meta) if meta else ""
    c2 = ws.cell(row=2, column=1, value=label)
    c2.font = META_FONT
    c2.fill = META_FILL
    c2.alignment = LEFT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.row_dimensions[2].height = 20


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 34


def _style_body_cell(cell, col: int, num_cols: set[int]) -> None:
    cell.font = BODY_FONT
    cell.border = BORDER
    if col in num_cols:
        cell.alignment = RIGHT
        cell.number_format = NUM_FMT
    else:
        cell.alignment = LEFT if col in (2, 3) else CENTER


P1_HEADERS = [
    "Deductor Sr.No.", "Name of Deductor", "TAN of Deductor",
    "Total Amount Paid/Credited", "Total Tax Deducted #", "Total TDS Deposited",
    "Txn Sr.No.", "Section", "Transaction Date", "Status of Booking",
    "Date of Booking", "Remarks", "Amount Paid/Credited",
    "Tax Deducted ##", "TDS Deposited",
]

# Part VI mirrors Part I column-for-column; only the wording changes. Keeping
# the geometry identical is what lets one parser and one builder serve both,
# and lets the journal builder read either sheet by column index.
P6_HEADERS = [
    "Collector Sr.No.", "Name of Collector", "TAN of Collector",
    "Total Amount Paid/Debited", "Total Tax Collected +", "Total TCS Deposited",
    "Txn Sr.No.", "Section", "Transaction Date", "Status of Booking",
    "Date of Booking", "Remarks", "Amount Paid/Debited",
    "Tax Collected ++", "TCS Deposited",
]


def build_part_i(ws, a: Assessee, deductors: list[P1Deductor],
                 title: str = "Part I", headers: Optional[list] = None,
                 entity: str = "deductors") -> None:
    """Render a deductor/collector part (Part I, or Part VI via build_part_vi)."""
    headers = headers or P1_HEADERS
    ncols = len(headers)
    _write_meta(ws, title, ncols, a)
    _write_header_row(ws, 3, headers)

    r = 4
    num_cols = {4, 5, 6, 13, 14, 15}
    subtotal_rows: list[int] = []

    if not deductors:
        cell = ws.cell(row=r, column=1, value="No Transactions Present")
        cell.font = Font(name="Arial", size=10, italic=True, color="7F7F7F")
        cell.alignment = CENTER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = BORDER
    else:
        for d in deductors:
            txn_start = r
            for t in d.txns:
                vals = [
                    d.sr, d.name, d.tan, d.tot_amt, d.tot_tax, d.tot_tds,
                    t.sr, t.section, t.txn_date, t.status, t.date_booking,
                    t.remarks, t.amount, t.tax, t.tds,
                ]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    _style_body_cell(cell, c, num_cols)
                r += 1
            txn_end = r - 1

            # Sub-total row for this deductor.
            # Write COMPUTED NUMERIC VALUES (not =SUM formulas): openpyxl never
            # caches a formula's result, so a bare =SUM reads back as blank in
            # any data_only consumer (the app preview, pandas, etc.). Summing in
            # Python guarantees a real number everywhere. It also makes a
            # zero-transaction deductor show 0 cleanly instead of a broken
            # backwards SUM range — and the reconciliation step then flags it.
            sub_amt = sum(t.amount for t in d.txns)
            sub_tax = sum(t.tax for t in d.txns)
            sub_tds = sum(t.tds for t in d.txns)
            sub_row = r
            subtotal_rows.append(sub_row)
            label = f"Sub-total — {d.name}  (txns: {len(d.txns)})"
            ws.cell(row=sub_row, column=1, value=f"#{d.sr}")
            ws.cell(row=sub_row, column=2, value=label)
            ws.merge_cells(start_row=sub_row, start_column=2,
                           end_row=sub_row, end_column=12)
            for col, val in ((13, sub_amt), (14, sub_tax), (15, sub_tds)):
                ws.cell(row=sub_row, column=col, value=val)
            for c in range(1, ncols + 1):
                cell = ws.cell(row=sub_row, column=c)
                cell.fill = SUBTOTAL_FILL
                cell.font = SUBTOTAL_FONT
                cell.border = BORDER
                if c in num_cols:
                    cell.alignment = RIGHT
                    cell.number_format = NUM_FMT
                elif c == 2:
                    cell.alignment = RIGHT_WRAP
                else:
                    cell.alignment = CENTER
            ws.row_dimensions[sub_row].height = 20
            r += 1

        # Grand Total — computed numeric values (sum of every transaction;
        # equals the sum of the deductor sub-totals, without double-counting).
        if subtotal_rows:
            grand_amt = sum(t.amount for d in deductors for t in d.txns)
            grand_tax = sum(t.tax for d in deductors for t in d.txns)
            grand_tds = sum(t.tds for d in deductors for t in d.txns)
            tot_row = r
            ws.cell(row=tot_row, column=2, value=f"GRAND TOTAL (all {entity})")
            for col, val in ((13, grand_amt), (14, grand_tax), (15, grand_tds)):
                ws.cell(row=tot_row, column=col, value=val)
            for c in range(1, ncols + 1):
                cell = ws.cell(row=tot_row, column=c)
                cell.fill = GRAND_FILL
                cell.font = GRAND_FONT
                cell.border = BORDER
                if c in num_cols:
                    cell.alignment = RIGHT
                    cell.number_format = NUM_FMT
                elif c == 2:
                    cell.alignment = RIGHT_WRAP
                else:
                    cell.alignment = CENTER
            ws.merge_cells(start_row=tot_row, start_column=2,
                           end_row=tot_row, end_column=12)
            ws.row_dimensions[tot_row].height = 24

    widths = [12, 46, 15, 18, 18, 18, 10, 10, 14, 10, 14, 10, 18, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def build_part_vi(ws, a: Assessee, collectors: list[P1Deductor]) -> None:
    """Render Part VI (Tax Collected at Source) with the Part I geometry."""
    build_part_i(ws, a, collectors, title="Part VI", headers=P6_HEADERS,
                 entity="collectors")


def build_part_ii(ws, a: Assessee, deductors: list[P2Deductor]) -> None:
    """Render Part II (TDS for 15G / 15H). 14 columns — NOT Part I's 15;
    there is no 'Status of Booking' column, so this does NOT reuse
    build_part_i's column geometry (num_cols, merge ranges, widths all
    shift by one column from Part I/VI)."""
    headers = EMPTY_HEADERS["Part II"]
    ncols = len(headers)
    _write_meta(ws, "Part II", ncols, a)
    _write_header_row(ws, 3, headers)

    r = 4
    num_cols = {4, 5, 6, 12, 13, 14}
    subtotal_rows: list[int] = []

    if not deductors:
        cell = ws.cell(row=r, column=1, value="No Transactions Present")
        cell.font = Font(name="Arial", size=10, italic=True, color="7F7F7F")
        cell.alignment = CENTER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = BORDER
    else:
        for d in deductors:
            for t in d.txns:
                vals = [
                    d.sr, d.name, d.tan, d.tot_amt, d.tot_tax, d.tot_tds,
                    t.sr, t.section, t.txn_date, t.date_booking,
                    t.remarks, t.amount, t.tax, t.tds,
                ]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    _style_body_cell(cell, c, num_cols)
                r += 1

            # Sub-total row — computed numeric values (see build_part_i for
            # why: a bare =SUM formula reads back blank under data_only).
            sub_amt = sum(t.amount for t in d.txns)
            sub_tax = sum(t.tax for t in d.txns)
            sub_tds = sum(t.tds for t in d.txns)
            sub_row = r
            subtotal_rows.append(sub_row)
            label = f"Sub-total — {d.name}  (txns: {len(d.txns)})"
            ws.cell(row=sub_row, column=1, value=f"#{d.sr}")
            ws.cell(row=sub_row, column=2, value=label)
            ws.merge_cells(start_row=sub_row, start_column=2,
                           end_row=sub_row, end_column=11)
            for col, val in ((12, sub_amt), (13, sub_tax), (14, sub_tds)):
                ws.cell(row=sub_row, column=col, value=val)
            for c in range(1, ncols + 1):
                cell = ws.cell(row=sub_row, column=c)
                cell.fill = SUBTOTAL_FILL
                cell.font = SUBTOTAL_FONT
                cell.border = BORDER
                if c in num_cols:
                    cell.alignment = RIGHT
                    cell.number_format = NUM_FMT
                elif c == 2:
                    cell.alignment = RIGHT_WRAP
                else:
                    cell.alignment = CENTER
            ws.row_dimensions[sub_row].height = 20
            r += 1

        if subtotal_rows:
            grand_amt = sum(t.amount for d in deductors for t in d.txns)
            grand_tax = sum(t.tax for d in deductors for t in d.txns)
            grand_tds = sum(t.tds for d in deductors for t in d.txns)
            tot_row = r
            ws.cell(row=tot_row, column=2, value="GRAND TOTAL (all deductors)")
            for col, val in ((12, grand_amt), (13, grand_tax), (14, grand_tds)):
                ws.cell(row=tot_row, column=col, value=val)
            for c in range(1, ncols + 1):
                cell = ws.cell(row=tot_row, column=c)
                cell.fill = GRAND_FILL
                cell.font = GRAND_FONT
                cell.border = BORDER
                if c in num_cols:
                    cell.alignment = RIGHT
                    cell.number_format = NUM_FMT
                elif c == 2:
                    cell.alignment = RIGHT_WRAP
                else:
                    cell.alignment = CENTER
            ws.merge_cells(start_row=tot_row, start_column=2,
                           end_row=tot_row, end_column=11)
            ws.row_dimensions[tot_row].height = 24

    widths = [12, 46, 15, 18, 18, 18, 10, 10, 14, 14, 10, 18, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def build_part_viii(ws, a: Assessee, rows: list[P8Row]) -> None:
    headers = [
        "Sr.No.", "Acknowledgement Number", "Name of Deductee", "PAN of Deductee",
        "Transaction Date", "Total Transaction Amount", "Total TDS Deposited***",
        "Total Amount Deposited other than TDS",
        "TDS Certificate Number", "Section", "Date of Deposit", "Status of Booking",
        "Date of Booking", "Demand Payment",
        "TDS Deposited*** (Txn)", "Amount Deposited other than TDS (Txn)",
    ]
    ncols = len(headers)
    _write_meta(ws, "Part VIII", ncols, a)
    _write_header_row(ws, 3, headers)

    r = 4
    num_cols = {6, 7, 8, 15, 16}
    if not rows:
        cell = ws.cell(row=r, column=1, value="No Transactions Present")
        cell.font = Font(name="Arial", size=10, italic=True, color="7F7F7F")
        cell.alignment = CENTER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = BORDER
    else:
        first_txn_row = r
        for row in rows:
            vals = [
                row.sr, row.ack_num, row.name, row.pan, row.txn_date,
                row.tot_txn_amt, row.tot_tds, row.tot_other,
                row.cert_num, row.section, row.date_deposit, row.status,
                row.date_booking, row.demand,
                row.tds_deposited, row.other_deposited,
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                _style_body_cell(cell, c, num_cols)
            r += 1
        last_txn_row = r - 1
        # If more than one deductee, add a grand total
        if len(rows) > 1:
            tot_row = r
            ws.cell(row=tot_row, column=2, value="GRAND TOTAL (all deductees)")
            # Computed numeric values, not =SUM (which reads back blank in any
            # data_only consumer because openpyxl doesn't cache formula results).
            col_totals = {
                6: sum(x.tot_txn_amt for x in rows),
                7: sum(x.tot_tds for x in rows),
                8: sum(x.tot_other for x in rows),
                15: sum(x.tds_deposited for x in rows),
                16: sum(x.other_deposited for x in rows),
            }
            for col, val in col_totals.items():
                ws.cell(row=tot_row, column=col, value=val)
            for c in range(1, ncols + 1):
                cell = ws.cell(row=tot_row, column=c)
                cell.fill = GRAND_FILL
                cell.font = GRAND_FONT
                cell.border = BORDER
                if c in num_cols:
                    cell.alignment = RIGHT
                    cell.number_format = NUM_FMT
                else:
                    cell.alignment = RIGHT_WRAP if c == 2 else CENTER
            ws.merge_cells(start_row=tot_row, start_column=2,
                           end_row=tot_row, end_column=5)
            ws.row_dimensions[tot_row].height = 24

    widths = [8, 22, 28, 16, 15, 20, 20, 24, 22, 10, 15, 14, 15, 14, 20, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def build_empty_part(ws, a: Assessee, title: str) -> None:
    headers = EMPTY_HEADERS[title]
    ncols = len(headers)
    _write_meta(ws, title, ncols, a)
    _write_header_row(ws, 3, headers)

    cell = ws.cell(row=4, column=1, value="No Transactions Present")
    cell.font = Font(name="Arial", size=10, italic=True, color="7F7F7F")
    cell.alignment = CENTER
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)
    ws.row_dimensions[4].height = 22
    for c in range(1, ncols + 1):
        ws.cell(row=4, column=c).border = BORDER

    widths = [max(12, min(22, len(h) + 2)) for h in headers]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


# -------------------- orchestration --------------------


def run(pdf_path: Path, out_path: Path) -> dict:
    text, pdftotext_path = pdf_to_text(pdf_path)
    assessee = parse_assessee(text)
    parts = split_by_parts(text)

    p1 = parse_part_i(parts.get("I", ""))
    p2 = parse_part_ii(parts.get("II", ""))
    p6 = parse_part_vi(parts.get("VI", ""))
    p8 = parse_part_viii(parts.get("VIII", ""))

    # Tier 4 hard-fail gate — BEFORE the workbook is built/written, so a
    # broken extraction never produces a plausible-looking .xlsx file. See
    # Empty26ASExtractionError's docstring for why this specific boundary
    # (all four parts zero, not any one part alone) is the right one.
    precheck_stats = {
        "part_i_deductors": len(p1),
        "part_ii_deductors": len(p2),
        "part_vi_collectors": len(p6),
        "part_viii_rows": len(p8),
    }
    check_extraction_not_vacuous(precheck_stats, text, pdftotext_path)

    wb = Workbook()
    wb.remove(wb.active)
    for title in SHEET_ORDER:
        ws = wb.create_sheet(title=title)
        if title == "Part I":
            build_part_i(ws, assessee, p1)
        elif title == "Part II":
            build_part_ii(ws, assessee, p2)
        elif title == "Part VI":
            build_part_vi(ws, assessee, p6)
        elif title == "Part VIII":
            build_part_viii(ws, assessee, p8)
        else:
            build_empty_part(ws, assessee, title)
    wb.save(out_path)

    recon_mismatches = reconcile_part_i(p1)
    recon_mismatches_ii = reconcile_part_i(p2)
    recon_mismatches_vi = reconcile_part_i(p6)

    stats = {
        "assessee": vars(assessee),
        "part_i_deductors": len(p1),
        "part_i_transactions": sum(len(d.txns) for d in p1),
        "part_ii_deductors": len(p2),
        "part_ii_transactions": sum(len(d.txns) for d in p2),
        "part_vi_collectors": len(p6),
        "part_vi_transactions": sum(len(c.txns) for c in p6),
        "part_viii_rows": len(p8),
        "output": str(out_path),
        "reconciliation": {
            "deductors_checked": len(p1),
            "deductors_ok": len(p1) - len(recon_mismatches),
            "mismatches": recon_mismatches,
        },
        "reconciliation_ii": {
            "deductors_checked": len(p2),
            "deductors_ok": len(p2) - len(recon_mismatches_ii),
            "mismatches": recon_mismatches_ii,
        },
        "reconciliation_vi": {
            "collectors_checked": len(p6),
            "collectors_ok": len(p6) - len(recon_mismatches_vi),
            "mismatches": recon_mismatches_vi,
        },
    }
    return stats


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print("Usage: python extract_26as_to_xlsx.py <input.pdf> <output.xlsx>",
              file=sys.stderr)
        return 2
    pdf, out = Path(argv[1]), Path(argv[2])
    try:
        stats = run(pdf, out)
    except Empty26ASExtractionError as e:
        print(f"FATAL: {e}", file=sys.stderr)
        return 3
    except (WrongPdftextFlavourError, RuntimeError) as e:
        # pdf_to_text() wraps a WrongPdftextFlavourError/FileNotFoundError
        # from resolve_pdftotext() in a plain RuntimeError (see pdf_to_text's
        # docstring) — catch both so either shape prints a clean FATAL line
        # instead of a raw traceback. Empty26ASExtractionError is also a
        # RuntimeError subclass, so it must be (and is) caught above this.
        print(f"FATAL: {e}", file=sys.stderr)
        return 4
    # Print a short summary (useful for verification)
    print(f"Assessee: {stats['assessee'].get('name','?')}  "
          f"PAN: {stats['assessee'].get('pan','?')}  "
          f"AY: {stats['assessee'].get('ay','?')}")
    print(f"Part I: {stats['part_i_deductors']} deductors, "
          f"{stats['part_i_transactions']} transactions")
    print(f"Part II: {stats['part_ii_deductors']} deductors, "
          f"{stats['part_ii_transactions']} transactions (15G/15H)")
    print(f"Part VI: {stats['part_vi_collectors']} collectors, "
          f"{stats['part_vi_transactions']} TCS transactions")
    print(f"Part VIII: {stats['part_viii_rows']} rows")

    # Reconciliation: computed sub-totals vs the totals the 26AS prints.
    recon = stats["reconciliation"]
    mismatches = recon["mismatches"]
    if not mismatches:
        print(f"Reconciliation: {recon['deductors_ok']}/"
              f"{recon['deductors_checked']} deductors OK — all sub-totals "
              f"match the 26AS header totals.")
    else:
        print(f"Reconciliation: {recon['deductors_ok']}/"
              f"{recon['deductors_checked']} deductors OK — "
              f"{len(mismatches)} MISMATCH(es) below:")
        for m in mismatches:
            print(f"  ! Sr.{m['sr']} {m['name']} (TAN {m['tan']}, "
                  f"{m['txns']} txns):")
            for f in m["fields"]:
                print(f"      {f['field']}: 26AS={f['header_total']:,.2f}  "
                      f"computed={f['computed_subtotal']:,.2f}  "
                      f"diff={f['difference']:+,.2f}")

    recon2 = stats["reconciliation_ii"]
    if recon2["deductors_checked"]:
        mm2 = recon2["mismatches"]
        if not mm2:
            print(f"Reconciliation (Part II): {recon2['deductors_ok']}/"
                  f"{recon2['deductors_checked']} deductors OK.")
        else:
            print(f"Reconciliation (Part II): {recon2['deductors_ok']}/"
                  f"{recon2['deductors_checked']} deductors OK — "
                  f"{len(mm2)} MISMATCH(es) below:")
            for m in mm2:
                print(f"  ! Sr.{m['sr']} {m['name']} (TAN {m['tan']}, "
                      f"{m['txns']} txns):")
                for f in m["fields"]:
                    print(f"      {f['field']}: 26AS={f['header_total']:,.2f}  "
                          f"computed={f['computed_subtotal']:,.2f}  "
                          f"diff={f['difference']:+,.2f}")

    recon6 = stats["reconciliation_vi"]
    if recon6["collectors_checked"]:
        mm6 = recon6["mismatches"]
        if not mm6:
            print(f"Reconciliation (Part VI): {recon6['collectors_ok']}/"
                  f"{recon6['collectors_checked']} collectors OK.")
        else:
            print(f"Reconciliation (Part VI): {recon6['collectors_ok']}/"
                  f"{recon6['collectors_checked']} collectors OK — "
                  f"{len(mm6)} MISMATCH(es) below:")
            for m in mm6:
                print(f"  ! Sr.{m['sr']} {m['name']} (TAN {m['tan']}, "
                      f"{m['txns']} txns):")
                for f in m["fields"]:
                    print(f"      {f['field']}: 26AS={f['header_total']:,.2f}  "
                          f"computed={f['computed_subtotal']:,.2f}  "
                          f"diff={f['difference']:+,.2f}")
    print(f"Saved: {stats['output']}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
