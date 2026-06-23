"""
agent.py — Form 26AS → Excel — DIRECT mode, no LLM.

Runs the deterministic extractor extract_26as_to_xlsx.py, then a per-deductor
reconciliation check on the produced workbook. No language model involved.
"""
from __future__ import annotations

import subprocess
import sys
from collections import defaultdict
from pathlib import Path

SCRIPT = Path(__file__).parent / "scripts" / "extract_26as_to_xlsx.py"


def _verify(xlsx_path: str) -> str:
    """Per-deductor reconciliation: txn-level totals vs deductor header totals."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        if "Part I" not in wb.sheetnames:
            return "Verify: ERROR — Part I sheet not found in workbook."
        ws = wb["Part I"]
        groups = defaultdict(lambda: [0.0, 0.0, 0.0, 0.0, 0.0, 0.0])
        for r in range(4, ws.max_row + 1):
            sr = ws.cell(r, 1).value
            if not isinstance(sr, int):
                continue
            groups[sr][0] = float(ws.cell(r, 4).value or 0)
            groups[sr][1] = float(ws.cell(r, 5).value or 0)
            groups[sr][2] = float(ws.cell(r, 6).value or 0)
            groups[sr][3] += float(ws.cell(r, 13).value or 0)
            groups[sr][4] += float(ws.cell(r, 14).value or 0)
            groups[sr][5] += float(ws.cell(r, 15).value or 0)
        mism = []
        for sr, v in groups.items():
            if abs(v[0]-v[3]) >= 0.01 or abs(v[1]-v[4]) >= 0.01 or abs(v[2]-v[5]) >= 0.01:
                mism.append(f"Deductor Sr.No {sr}: Amt {v[0]} vs {v[3]:.2f}, "
                            f"Tax {v[1]} vs {v[4]:.2f}, TDS {v[2]} vs {v[5]:.2f}")
        if mism:
            return "Verify: RECONCILIATION ERRORS:\n" + "\n".join(mism)
        return f"Verify: OK — all {len(groups)} deductors reconcile."
    except Exception as e:
        return f"Verify: ERROR during verification: {e}"


def run(
    pdf_path: str,
    output_path: str,
    config_path: str = "config.yaml",
    model_override: str = None,
) -> str:
    """
    Convert a Form 26AS PDF to an Excel workbook at output_path, then verify
    per-deductor totals. Direct mode — no LLM. config_path / model_override ignored.
    """
    result = subprocess.run(
        [sys.executable, str(SCRIPT), pdf_path, output_path],
        capture_output=True, text=True,
    )
    out = (result.stdout or "").strip()
    err = (result.stderr or "").strip()
    if result.returncode != 0:
        return f"ERROR: {err or out}"
    summary = out or "Extraction complete."
    if Path(output_path).is_file():
        summary += "\n\n" + _verify(output_path)
    return summary
