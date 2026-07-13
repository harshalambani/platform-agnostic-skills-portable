"""
write_workbook.py -- openpyxl writer for the standardized ITR output workbook
(plan section 2.2). Maximally formula-driven: every Computation figure is
`='<sheet>'!<cell>`; the tax block itself is built as two parallel formula
columns (New regime / Old regime), each referencing ONLY Rules-sheet cells
for its rates/caps/slabs (never a literal rate in a formula) -- flipping the
Entity sheet's regime cell re-drives which column the final Tax Liability /
Refund-or-Payable rows select (plan's "visible both-regimes comparison block
with a chooser cell", an explicitly acceptable implementation).

Hard constants (no formula) are confined to: the Entity sheet (profile
input), the Rules sheet (the versioned config dump), and the schedule
sheets' own detail rows (parsed input values -- these ARE the traceable
source data, plan section 3). Schedule-sheet SUBTOTAL cells are SUM
formulas over that sheet's own detail rows wherever there is more than one
contributing row.

Indian digit grouping is applied via a custom number_format; paise are kept
on every schedule sheet; the ONLY rounding is at the two statutory
checkpoints (s.288A on Total Income, s.288B on Tax Payable/Refund), both via
Excel MROUND() referencing the Rules sheet's rounding-nearest cells (never a
hardcoded ROUND(...,-1) in the workbook).
"""
from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import rules as rules_engine
import schedules as sch
import tags as tag_vocab

INR_FORMAT = "#,##,##0.00"

_MAX_SLAB_ROWS = 8
_MAX_SURCHARGE_ROWS = 4

_UNMAPPED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Line label shown on the Mapping Review sheet's Destination column, one per
# tag -- kept in sync BY HAND with the row labels write_*_sheet() above
# actually writes for that tag's contribution (this sheet is a read-only
# review surface; it doesn't reuse the writer functions' own row-label
# strings since several tags feed into one summed row there).
_TAG_LINE_LABEL: dict[str, str] = {
    "SALARY_GROSS": "Gross salary (17(1)+17(2)+17(3))",
    "BUS_REMUNERATION": "Partner/proprietor remuneration",
    "BUS_EXPENSE": "Business expenses (net)",
    "OS_INTEREST_SB": "Savings bank interest",
    "OS_INTEREST_BANK": "Bank FD interest",
    "OS_INTEREST_NBFC": "NBFC/HFC interest",
    "OS_INTEREST_EPF_TAXABLE": "EPF taxable interest",
    "OS_REFUND_INTEREST": "Interest on IT refund (taxable -- RULE-1)",
    "OS_DIVIDEND": "Dividend income (gross)",
    "OS_SLBS": "SLBS income",
    "NONTAX_REFUND_PRINCIPAL": "IT refund principal (excluded, not income -- RULE-1)",
    "CG_LT_CONTROL": "Books LTCG control total",
    "CG_ST_CONTROL": "Books STCG control total",
    "EXEMPT_PPF_INTEREST": "PPF interest",
    "EXEMPT_10_2A": "Share of firm profit (s.10(2A))",
    "HP_RENT": "Gross Annual Value (rent)",
    "HP_MUNICIPAL_TAX": "Municipal taxes paid by owner",
    "HP_INTEREST": "Interest on housing loan (s.24(b))",
    "TAXPAID_ADV": "Advance tax",
    "TAXPAID_SAT": "Self-assessment tax",
    "TAXPAID_TDS_SALARY": "TDS on salary",
    "TAXPAID_TDS_INTEREST": "TDS on interest",
    "TAXPAID_TDS_DIVIDEND": "TDS on dividend",
    "TAXPAID_TCS": "TCS",
    "PERSONAL": "(non-deductible / non-taxable, shown for completeness only)",
    "DED_80C_CANDIDATE": "80C candidate",
    "DED_80D_CANDIDATE": "80D candidate",
    "DED_80G_CANDIDATE": "80G candidate",
    "EQUITY_CAPITAL": "(balancing figure)",
    "TRADING": "(balancing figure)",
}
for _tag in tag_vocab.AL_TAGS:
    _TAG_LINE_LABEL.setdefault(_tag, _tag)


def _bold(cell) -> None:
    cell.font = Font(bold=True)


class _SheetWriter:
    """Thin row-cursor helper around an openpyxl worksheet."""

    def __init__(self, ws):
        self.ws = ws
        self.row = 1

    def cell(self, col: int, value=None, *, bold: bool = False, number_format: str | None = None):
        c = self.ws.cell(row=self.row, column=col, value=value)
        if bold:
            _bold(c)
        if number_format:
            c.number_format = number_format
        return c

    def label_value(self, label: str, value=None, *, number_format: str | None = INR_FORMAT):
        self.cell(1, label)
        c = self.cell(2, value, number_format=number_format)
        self.row += 1
        return c

    def header(self, text: str):
        self.cell(1, text, bold=True)
        self.row += 1

    def blank(self):
        self.row += 1


# ---------------------------------------------------------------------------
# Rules sheet
# ---------------------------------------------------------------------------

def _write_slab_block(sw: _SheetWriter, slabs: list) -> dict:
    """Write a slab table padded to _MAX_SLAB_ROWS rows (extra rows repeat
    the last upto/0-rate so they contribute nothing); returns {row_index:
    (upto_cell, rate_cell)}."""
    refs = {}
    padded = list(slabs) + [{"upto": slabs[-1]["upto"] if slabs else 0, "rate": 0.0}] * (_MAX_SLAB_ROWS - len(slabs))
    for i, band in enumerate(padded[:_MAX_SLAB_ROWS]):
        sw.cell(1, f"Slab {i + 1} upto")
        sw.cell(2, band["upto"] if band["upto"] is not None else 999999999)
        sw.cell(3, "rate")
        sw.cell(4, band["rate"])
        refs[i] = (f"B{sw.row}", f"D{sw.row}")
        sw.row += 1
    return refs


def _write_surcharge_block(sw: _SheetWriter, bands: list) -> dict:
    refs = {}
    padded = list(bands) + [{"above": 10**12, "rate": 0.0}] * (_MAX_SURCHARGE_ROWS - len(bands))
    for i, band in enumerate(padded[:_MAX_SURCHARGE_ROWS]):
        sw.cell(1, f"Surcharge band {i + 1} above")
        sw.cell(2, band["above"])
        sw.cell(3, "rate")
        sw.cell(4, band["rate"])
        refs[i] = (f"B{sw.row}", f"D{sw.row}")
        sw.row += 1
    return refs


def write_rules_sheet(wb: Workbook, rules: rules_engine.RulesConfig, user_rules: list, status: str, dob: str | None, fy_end: date) -> dict:
    """Dumps the versioned rules config (plan row 2). Returns a `layout` dict
    of A1 cell references the Computation sheet formulas key off of."""
    ws = wb.create_sheet("Rules")
    sw = _SheetWriter(ws)
    layout: dict = {}

    sw.header(f"Rules -- {rules.year_label} (act {rules.act}, config version {rules.version})")
    sw.blank()

    sw.header("New Regime")
    new_block = rules.regime("new")
    layout["new_slabs"] = _write_slab_block(sw, new_block["slabs"])
    layout["new_std_deduction"] = sw.label_value("Std deduction (salary)", new_block["std_deduction_salary"])
    layout["new_rebate_max_ti"] = sw.label_value("87A max total income", new_block["rebate_87a"]["max_total_income"])
    layout["new_rebate_max_amt"] = sw.label_value("87A max rebate", new_block["rebate_87a"]["max_rebate"])
    layout["new_rebate_marginal"] = sw.label_value("87A marginal relief", new_block["rebate_87a"].get("marginal_relief", False), number_format=None)
    layout["new_rebate_excl_special"] = sw.label_value(
        "87A excludes special-rate income", new_block["rebate_87a"].get("excludes_special_rate_income", False), number_format=None,
    )
    layout["new_surcharge"] = _write_surcharge_block(sw, new_block["surcharge"]["bands"])
    layout["new_surcharge_cg_cap"] = sw.label_value("Surcharge cap on CG/dividend", new_block["surcharge"]["cap_on_cg_dividend"])
    sw.blank()

    sw.header("Old Regime (resolved slab table for this entity's age/status)")
    old_block = rules.regime("old")
    old_slabs = rules_engine.resolve_slabs(rules, "old", status, dob, fy_end)
    layout["old_slabs"] = _write_slab_block(sw, old_slabs)
    layout["old_std_deduction"] = sw.label_value("Std deduction (salary)", old_block["std_deduction_salary"])
    layout["old_rebate_max_ti"] = sw.label_value("87A max total income", old_block["rebate_87a"]["max_total_income"])
    layout["old_rebate_max_amt"] = sw.label_value("87A max rebate", old_block["rebate_87a"]["max_rebate"])
    layout["old_rebate_marginal"] = sw.label_value("87A marginal relief", old_block["rebate_87a"].get("marginal_relief", False), number_format=None)
    layout["old_surcharge"] = _write_surcharge_block(sw, old_block["surcharge"]["bands"])
    layout["old_surcharge_cg_cap"] = sw.label_value("Surcharge cap on CG/dividend", old_block["surcharge"]["cap_on_cg_dividend"])
    sw.blank()

    sw.header("Chapter VI-A caps (old regime)")
    for key, val in old_block.get("vi_a_caps", {}).items():
        layout[f"cap_{key}"] = sw.label_value(key, val)
    sw.blank()

    sw.header("Common")
    common = rules.common
    layout["cess_rate"] = sw.label_value("Cess rate", common["cess_rate"], number_format="0.00%")
    layout["round_ti_nearest"] = sw.label_value("Total Income rounding (s.288A), nearest", common["rounding"]["total_income"]["nearest"])
    layout["round_tax_nearest"] = sw.label_value("Tax payable/refund rounding (s.288B), nearest", common["rounding"]["tax_payable_refund"]["nearest"])
    layout["hp_std_ded_pct"] = sw.label_value("House property std deduction % of NAV", common["house_property"]["std_deduction_pct_of_nav"], number_format="0.00%")
    layout["al_threshold"] = sw.label_value("Schedule AL total-income threshold", common["schedule_al"]["total_income_threshold"])
    sw.blank()

    sw.header("Capital gains rates (regime-independent -- 111A/112A)")
    cg = common["capital_gains"]

    def _rate_pair(block: dict) -> tuple:
        if "rate" in block:
            return block["rate"], block["rate"]
        return block["before_split"], block["on_after_split"]

    ltcg_before, ltcg_after = _rate_pair(cg["s112a_ltcg_equity_stt"])
    stcg_before, stcg_after = _rate_pair(cg["s111a_stcg_equity_stt"])
    layout["ltcg_rate_before"] = sw.label_value("112A LTCG rate -- before split date", ltcg_before, number_format="0.00%")
    layout["ltcg_rate_after"] = sw.label_value("112A LTCG rate -- on/after split date", ltcg_after, number_format="0.00%")
    layout["stcg_rate_before"] = sw.label_value("111A STCG rate -- before split date", stcg_before, number_format="0.00%")
    layout["stcg_rate_after"] = sw.label_value("111A STCG rate -- on/after split date", stcg_after, number_format="0.00%")
    layout["ltcg_exemption"] = sw.label_value("112A exemption limit", cg["s112a_ltcg_equity_stt"].get("exemption_limit", 0))
    if cg.get("split_date"):
        sw.label_value(
            "ASSUMPTION -- split-year 112A exemption allocation", (
                "When the 125,000 exemption only partially offsets LTCG (112A) gains "
                "spanning the split date, the unabsorbed exemption is allocated "
                "PRO-RATA across the before/after split-date buckets (by taxable-gain "
                "share), not to one bucket first -- the law does not mandate an "
                "ordering between the two in-year rate buckets. See Reconciliation "
                "for whether this allocation engaged on this run."
            ), number_format=None,
        )
    sw.blank()

    sw.header("Applied user rules")
    for r in user_rules:
        sw.label_value(r.id, r.statement, number_format=None)

    return layout


# ---------------------------------------------------------------------------
# Entity sheet
# ---------------------------------------------------------------------------

def write_entity_sheet(wb: Workbook, entity, year_key: str, rules: rules_engine.RulesConfig, regime: str, form16_election: str | None) -> dict:
    ws = wb.create_sheet("Entity")
    sw = _SheetWriter(ws)
    layout = {}

    sw.header("Entity")
    sw.label_value("Name", entity.name, number_format=None)
    sw.label_value("PAN", entity.pan, number_format=None)
    sw.label_value("Status", entity.status, number_format=None)
    sw.label_value("Residency", entity.residency, number_format=None)
    sw.label_value("DOB", entity.dob, number_format=None)
    sw.label_value("Address", entity.address, number_format=None)
    layout["year_label"] = sw.label_value("Year", rules.year_label, number_format=None)
    layout["regime"] = sw.label_value("Regime (flip this cell: new / old)", regime, number_format=None)
    if form16_election is not None:
        sw.label_value("Form16 115BAC(1A) opt-out election", form16_election, number_format=None)
    sw.label_value(
        "Suggested ITR form",
        "=IF('BusinessPL'!B2<>0,\"ITR-3\",\"ITR-2\")", number_format=None,
    )
    return layout


# ---------------------------------------------------------------------------
# Schedule sheets (data rows + SUM subtotals)
# ---------------------------------------------------------------------------

def write_salary_sheet(wb: Workbook, salary: sch.SalarySchedule) -> dict:
    ws = wb.create_sheet("Salary")
    sw = _SheetWriter(ws)
    layout = {}
    sw.header("Salary")
    layout["gross"] = sw.label_value("Gross salary (17(1)+17(2)+17(3))", salary.gross)
    sw.label_value("Source", salary.source, number_format=None)
    layout["s10_exempt"] = sw.label_value("s.10 exemptions total", salary.s10_exempt_total)
    layout["std_deduction"] = sw.label_value("Standard deduction (s.16(ia))", salary.std_deduction)
    layout["prof_tax"] = sw.label_value("Professional tax (s.16(iii))", salary.prof_tax)
    layout["income_chargeable"] = sw.label_value("Income chargeable under Salaries", salary.income_chargeable)
    if salary.manual_flagged:
        sw.label_value("Flag", "manual entry -- no Form16 supplied, CA to confirm", number_format=None)
    return layout


def write_business_sheet(wb: Workbook, business: sch.BusinessSchedule) -> dict:
    ws = wb.create_sheet("BusinessPL")
    sw = _SheetWriter(ws)
    layout = {}
    sw.header("Business P&L (ITR-3 only)")
    layout["remuneration"] = sw.label_value("Partner/proprietor remuneration", business.remuneration)
    layout["expenses"] = sw.label_value("Business expenses (net)", business.expenses_total)
    sw.cell(1, "Net business income")
    sw.cell(2, f"=B{sw.row - 2}+B{sw.row - 1}", number_format=INR_FORMAT)
    layout["net"] = f"B{sw.row}"
    sw.row += 1
    return layout


def write_house_property_sheet(wb: Workbook, hp: sch.HousePropertySchedule) -> dict:
    ws = wb.create_sheet("HouseProperty")
    sw = _SheetWriter(ws)
    layout = {}
    sw.header("House Property")
    layout["gav"] = sw.label_value("Gross Annual Value (rent)", hp.gav)
    layout["municipal_tax"] = sw.label_value("Municipal taxes paid by owner", hp.municipal_tax)
    layout["nav"] = sw.label_value("Net Annual Value (GAV - municipal tax)", hp.nav)
    layout["std_deduction"] = sw.label_value("Standard deduction 30% of NAV (s.24(a))", hp.std_deduction_24a)
    layout["interest"] = sw.label_value("Interest on housing loan (s.24(b))", hp.interest_24b)
    layout["income"] = sw.label_value("Income from house property", hp.income)
    return layout


def write_other_sources_sheet(wb: Workbook, os_: sch.OtherSourcesSchedule) -> dict:
    ws = wb.create_sheet("OtherSources")
    sw = _SheetWriter(ws)
    layout = {}
    sw.header("Other Sources")
    layout["sb"] = sw.label_value("Savings bank interest", os_.interest_sb)
    layout["bank"] = sw.label_value("Bank FD interest", os_.interest_bank)
    layout["nbfc"] = sw.label_value("NBFC/HFC interest", os_.interest_nbfc)
    layout["epf"] = sw.label_value("EPF taxable interest", os_.interest_epf_taxable)
    layout["refund_interest"] = sw.label_value("Interest on IT refund (taxable -- RULE-1)", os_.refund_interest)
    layout["refund_principal"] = sw.label_value("IT refund principal (excluded, not income -- RULE-1)", os_.refund_principal_excluded)
    layout["dividend"] = sw.label_value("Dividend income (gross)", os_.dividend_gross)
    sw.label_value("  Dividend quarter-bucket source", os_.dividend_quarters_source, number_format=None)
    for i, q in enumerate(os_.dividend_quarters, start=1):
        sw.label_value(f"  Dividend Q{i} (234C bucket)", q)
    sw.label_value("  Interest quarter-bucket source", os_.interest_quarters_source, number_format=None)
    for i, q in enumerate(os_.interest_quarters, start=1):
        sw.label_value(f"  Interest Q{i} (234C bucket)", q)
    layout["slbs"] = sw.label_value("SLBS income", os_.slbs)
    sw.cell(1, "Taxable Other Sources total")
    sw.cell(2, os_.taxable_total, number_format=INR_FORMAT)
    layout["taxable_total"] = f"B{sw.row}"
    sw.row += 1
    return layout


def write_capital_gains_sheet(wb: Workbook, cg: sch.CapitalGainsSchedule, rules_layout: dict) -> dict:
    ws = wb.create_sheet("CapitalGains")
    sw = _SheetWriter(ws)
    layout = {}
    sw.header("Capital Gains")
    layout["lt_control"] = sw.label_value("Books LTCG control total", cg.lt_control)
    layout["st_control"] = sw.label_value("Books STCG control total", cg.st_control)
    layout["lt_taxable_gross"] = sw.label_value("LTCG taxable (before 112A exemption)", cg.lt_taxable_gross)
    layout["lt_exemption"] = sw.label_value("112A exemption used", cg.lt_exemption_used)
    layout["lt_before_split"] = sw.label_value("LTCG taxable -- before split date", cg.lt_taxable_before_split)
    layout["lt_after_split"] = sw.label_value("LTCG taxable -- on/after split date", cg.lt_taxable_on_after_split)
    layout["st_before_split"] = sw.label_value("STCG taxable -- before split date", cg.st_taxable_before_split)
    layout["st_after_split"] = sw.label_value("STCG taxable -- on/after split date", cg.st_taxable_on_after_split)
    layout["st_taxable_gross"] = sw.label_value("STCG taxable (total)", cg.st_taxable_gross)

    sw.cell(1, "112A/111A tax on special-rate CG (regime-independent)")
    special_tax_cell = sw.cell(
        2,
        f"={layout['lt_before_split'].coordinate}*'Rules'!{rules_layout['ltcg_rate_before'].coordinate}"
        f"+{layout['lt_after_split'].coordinate}*'Rules'!{rules_layout['ltcg_rate_after'].coordinate}"
        f"+{layout['st_before_split'].coordinate}*'Rules'!{rules_layout['stcg_rate_before'].coordinate}"
        f"+{layout['st_after_split'].coordinate}*'Rules'!{rules_layout['stcg_rate_after'].coordinate}",
        number_format=INR_FORMAT,
    )
    layout["special_tax_cell"] = special_tax_cell.coordinate
    sw.row += 1

    sw.label_value("Reconciliation OK", cg.reconciliation_ok, number_format=None)
    sw.label_value("Reconciliation diff (lot sum - control)", cg.reconciliation_diff)
    if cg.unresolved_scrips:
        sw.label_value("Unresolved FMV scrips (fail-loud, review)", ", ".join(cg.unresolved_scrips), number_format=None)
    sw.blank()

    sw.header("Lot detail")
    lot_headers = ["Scrip", "Sale date", "Buy date", "Term", "Qty", "Cost", "Proceeds",
                   "Booked gain", "Grandfathered", "FMV used", "Taxable gain", "Attribution"]
    for col, h in enumerate(lot_headers, start=1):
        sw.cell(col, h)
    sw.row += 1
    for lot in cg.lot_rows:
        sw.cell(1, lot.scrip)
        sw.cell(2, lot.sale_date.isoformat())
        sw.cell(3, lot.buy_date.isoformat() if lot.buy_date else None)
        sw.cell(4, lot.term)
        sw.cell(5, lot.qty)
        sw.cell(6, lot.cost, number_format=INR_FORMAT)
        sw.cell(7, lot.proceeds, number_format=INR_FORMAT)
        sw.cell(8, lot.booked_gain, number_format=INR_FORMAT)
        sw.cell(9, lot.grandfathered)
        sw.cell(10, lot.fmv_used)
        sw.cell(11, lot.taxable_gain, number_format=INR_FORMAT)
        sw.cell(12, lot.attribution)
        sw.row += 1
    return layout


def write_exempt_income_sheet(wb: Workbook, ei: sch.ExemptIncomeSchedule) -> dict:
    ws = wb.create_sheet("ExemptIncome")
    sw = _SheetWriter(ws)
    layout = {}
    sw.header("Exempt Income")
    layout["ppf"] = sw.label_value("PPF interest", ei.ppf_interest)
    layout["firm_profit"] = sw.label_value("Share of firm profit (s.10(2A))", ei.share_of_firm_profit)
    return layout


def write_taxes_paid_sheet(wb: Workbook, tp: sch.TaxesPaidSchedule) -> dict:
    ws = wb.create_sheet("TaxesPaid")
    sw = _SheetWriter(ws)
    layout = {}
    sw.header("Taxes Paid")
    layout["advance"] = sw.label_value("Advance tax", tp.advance_tax)
    layout["sat"] = sw.label_value("Self-assessment tax", tp.self_assessment_tax)
    layout["tds_salary"] = sw.label_value("TDS on salary", tp.tds_salary)
    layout["tds_interest"] = sw.label_value("TDS on interest", tp.tds_interest)
    layout["tds_dividend"] = sw.label_value("TDS on dividend", tp.tds_dividend)
    layout["tcs"] = sw.label_value("TCS", tp.tcs)
    sw.blank()
    sw.header("26AS tie-out")
    if not tp.as26_available:
        sw.label_value("26AS tie-out", "no 26AS workbook supplied -- skipped", number_format=None)
    else:
        sw.label_value("26AS TDS on interest (194A)", tp.as26_tds_interest)
        sw.label_value("26AS TDS on dividend (194/194K)", tp.as26_tds_dividend)
        sw.label_value("Tie-out status", "OK" if tp.tie_out_ok else f"{len(tp.tie_out_conflicts)} CONFLICT(S)", number_format=None)
        for c in tp.tie_out_conflicts:
            sw.label_value(
                f"  CONFLICT -- {c['category']}",
                f"book={c['book']:.2f} 26AS={c['as26']:.2f} diff={c['diff']:.2f}",
                number_format=None,
            )
    sw.blank()
    sw.cell(1, "Total taxes paid")
    sw.cell(2, tp.total, number_format=INR_FORMAT)
    layout["total"] = f"B{sw.row}"
    sw.row += 1
    return layout


def write_deductions_sheet(wb: Workbook, ded: sch.DeductionsSchedule) -> dict:
    ws = wb.create_sheet("Deductions")
    sw = _SheetWriter(ws)
    layout = {}
    sw.header("Deductions (Chapter VI-A)")
    if ded.regime_na:
        sw.label_value("Status", ded.regime_na_note, number_format=None)
    for label, candidates in (("80C", ded.candidates_80c), ("80D", ded.candidates_80d), ("80G", ded.candidates_80g)):
        for c in candidates:
            sw.label_value(f"{label} candidate -- {c.path} (CA to confirm)", c.amount)
    layout["80c"] = sw.label_value("80C claimed (capped)", ded.total_80c_claimed)
    layout["80d"] = sw.label_value("80D claimed (capped)", ded.total_80d_claimed)
    layout["80tta_ttb"] = sw.label_value("80TTA/80TTB claimed (capped)", ded.total_80tta_ttb_claimed)
    layout["80g"] = sw.label_value("80G claimed (capped)", ded.total_80g_claimed)
    sw.cell(1, "Total VI-A deductions")
    sw.cell(2, 0 if ded.regime_na else ded.total, number_format=INR_FORMAT)
    layout["total"] = f"B{sw.row}"
    sw.row += 1
    return layout


def write_schedule_al_sheet(wb: Workbook, al: sch.ScheduleALSchedule) -> dict:
    ws = wb.create_sheet("ScheduleAL")
    sw = _SheetWriter(ws)
    layout = {}
    sw.header(f"Schedule AL (required if Total Income > {al.threshold:,.2f}: {al.required})")
    for tag, total in al.buckets.items():
        sw.label_value(tag, total)
    layout["total_assets"] = sw.label_value("Total assets (at cost)", al.total_assets)
    layout["total_liabilities"] = sw.label_value("Total liabilities", al.total_liabilities)
    layout["net"] = sw.label_value("Net", al.net)
    return layout


def write_schedule_fa_sheet(wb: Workbook, fa: sch.ScheduleFASchedule) -> None:
    ws = wb.create_sheet("ScheduleFA")
    sw = _SheetWriter(ws)
    sw.header("Schedule FA -- pre-seeded from AL_FOREIGN-flagged holdings; detail cells manual")
    for col, h in enumerate(["Account", "Amount", "Country", "Peak balance", "Details"], start=1):
        sw.cell(col, h)
    sw.row += 1
    for r in fa.rows:
        sw.cell(1, r.path)
        sw.cell(2, r.amount, number_format=INR_FORMAT)
        sw.row += 1


def write_is_transcript(wb: Workbook, tree) -> None:
    ws = wb.create_sheet("IS_Transcript")
    sw = _SheetWriter(ws)
    sw.header("Retained Earnings -- verbatim transcript")
    for col, h in enumerate(["Path", "Total", "GUID"], start=1):
        sw.cell(col, h)
    sw.row += 1
    for n in tree.all_nodes():
        if n.section.startswith("RetainedEarnings") and not n.children:
            sw.cell(1, n.path)
            sw.cell(2, n.total, number_format=INR_FORMAT)
            sw.cell(3, n.guid)
            sw.row += 1


def write_bs_transcript(wb: Workbook, tree) -> None:
    ws = wb.create_sheet("BS_Transcript")
    sw = _SheetWriter(ws)
    sw.header("Balance Sheet -- verbatim transcript")
    for col, h in enumerate(["Path", "Total", "GUID", "Section"], start=1):
        sw.cell(col, h)
    sw.row += 1
    for n in tree.all_nodes():
        if not n.section.startswith("RetainedEarnings") and not n.children:
            sw.cell(1, n.path)
            sw.cell(2, n.total, number_format=INR_FORMAT)
            sw.cell(3, n.guid)
            sw.cell(4, n.section)
            sw.row += 1
    sw.blank()
    sw.label_value("Assets - Equity - NetIncome (self-check, should be 0)", tree.section_totals.get("Assets Accounts", 0.0)
                    - tree.section_totals.get("Equity, Trading, and Liabilities", 0.0)
                    - tree.section_totals.get("Retained Earnings", 0.0))


# ---------------------------------------------------------------------------
# Computation sheet -- the formula-driven backbone
# ---------------------------------------------------------------------------

def _slab_tax_formula(income_cell: str, slab_refs: dict) -> str:
    """SUM of MAX(0, MIN(income, upto_i) - upto_(i-1)) * rate_i across the
    padded slab rows in `slab_refs` (Rules-sheet cell refs) -- a generic
    Excel slab-tax formula that never hardcodes a rate or threshold."""
    terms = []
    prev_upto_ref = None
    for i in sorted(slab_refs):
        upto_ref, rate_ref = slab_refs[i]
        lower = f"'Rules'!{prev_upto_ref}" if prev_upto_ref else "0"
        terms.append(f"(MAX(0,MIN({income_cell},'Rules'!{upto_ref})-{lower}))*'Rules'!{rate_ref}")
        prev_upto_ref = upto_ref
    return "+".join(terms)


def _surcharge_formula(income_cell: str, tax_cell: str, surcharge_refs: dict) -> str:
    """Nested-IF surcharge-rate lookup (highest matching band) * tax."""
    expr = "0"
    for i in sorted(surcharge_refs):
        above_ref, rate_ref = surcharge_refs[i]
        expr = f"IF({income_cell}>'Rules'!{above_ref},'Rules'!{rate_ref},{expr})"
    return f"({expr})*{tax_cell}"


def _rebate_formula(normal_income_cell: str, tax_normal_cell: str, max_ti_ref: str, max_amt_ref: str, marginal_ref: str) -> str:
    base = f"MIN('Rules'!{max_amt_ref},{tax_normal_cell})"
    marginal = f"MAX(0,{tax_normal_cell}-({normal_income_cell}-'Rules'!{max_ti_ref}))"
    return (
        f"IF({normal_income_cell}<='Rules'!{max_ti_ref},{base},"
        f"IF('Rules'!{marginal_ref},{marginal},0))"
    )


def write_computation_sheet(
    wb: Workbook, model: sch.ITRModel, rules_layout: dict, entity_layout: dict,
    salary_layout: dict, business_layout: dict, hp_layout: dict, os_layout: dict,
    cg_layout: dict, ded_layout: dict, tp_layout: dict,
) -> None:
    ws = wb.create_sheet("Computation")
    sw = _SheetWriter(ws)
    regime_cell = f"'Entity'!{entity_layout['regime'].coordinate}"

    sw.header("Computation")
    salary_cell = sw.label_value("Salary income", f"='Salary'!{salary_layout['income_chargeable'].coordinate}")
    hp_cell = sw.label_value("Income from house property", f"='HouseProperty'!{hp_layout['income'].coordinate}")
    biz_cell = sw.label_value("Business/profession income", f"='BusinessPL'!{business_layout['net']}")
    os_cell = sw.label_value("Other sources income", f"='OtherSources'!{os_layout['taxable_total']}")
    cg_lt_cell = sw.label_value("Capital gains -- LT (special rate)", f"='CapitalGains'!{cg_layout['lt_taxable_gross'].coordinate}-'CapitalGains'!{cg_layout['lt_exemption'].coordinate}")
    cg_st_cell = sw.label_value("Capital gains -- ST (special rate)", f"='CapitalGains'!{cg_layout['st_taxable_gross'].coordinate}")

    sw.cell(1, "Normal-rate income (Salary+HP+Business+OS)")
    normal_cell_ref = sw.cell(
        2, f"={salary_cell.coordinate}+{hp_cell.coordinate}+{biz_cell.coordinate}+{os_cell.coordinate}",
        number_format=INR_FORMAT,
    ).coordinate
    sw.row += 1

    sw.cell(1, "Gross Total Income")
    gti_cell = sw.cell(2, f"={normal_cell_ref}+{cg_lt_cell.coordinate}+{cg_st_cell.coordinate}", number_format=INR_FORMAT).coordinate
    sw.row += 1

    via_cell = sw.label_value("Chapter VI-A deductions", f"='Deductions'!{ded_layout['total']}")

    sw.cell(1, "Total Income (before s.288A rounding)")
    ti_raw_cell = sw.cell(2, f"={gti_cell}-{via_cell.coordinate}", number_format=INR_FORMAT).coordinate
    sw.row += 1

    sw.cell(1, "Total Income (rounded, s.288A)")
    ti_cell = sw.cell(2, f"=MROUND({ti_raw_cell},'Rules'!{rules_layout['round_ti_nearest'].coordinate})", number_format=INR_FORMAT).coordinate
    sw.row += 1
    sw.blank()

    sw.cell(1, "Normal income net of special-rate CG (used for slab tax)")
    normal_for_slab_cell = sw.cell(2, f"=MAX(0,{ti_cell}-{cg_lt_cell.coordinate}-{cg_st_cell.coordinate})", number_format=INR_FORMAT).coordinate
    sw.row += 1

    def _regime_block(prefix: str, slab_refs: dict, rebate_max_ti: str, rebate_max_amt: str,
                       rebate_marginal: str, surcharge_refs: dict) -> str:
        sw.header(f"{prefix} regime")
        tax_normal_cell = sw.cell(2, f"={_slab_tax_formula(normal_for_slab_cell, slab_refs)}", number_format=INR_FORMAT)
        sw.cell(1, "Tax on normal-rate income")
        tax_normal_ref = tax_normal_cell.coordinate
        sw.row += 1

        sw.cell(1, "112A/111A tax on special-rate CG")
        cg_tax_cell = sw.cell(2, f"='CapitalGains'!{cg_layout['special_tax_cell']}", number_format=INR_FORMAT)
        cg_tax_ref = cg_tax_cell.coordinate
        sw.row += 1

        sw.cell(1, "87A rebate")
        rebate_cell = sw.cell(2, f"={_rebate_formula(normal_for_slab_cell, tax_normal_ref, rebate_max_ti, rebate_max_amt, rebate_marginal)}", number_format=INR_FORMAT)
        rebate_ref = rebate_cell.coordinate
        sw.row += 1

        sw.cell(1, "Tax after rebate + special-rate tax")
        after_rebate_cell = sw.cell(2, f"=MAX(0,{tax_normal_ref}-{rebate_ref})+{cg_tax_ref}", number_format=INR_FORMAT)
        after_rebate_ref = after_rebate_cell.coordinate
        sw.row += 1

        sw.cell(1, "Surcharge")
        surcharge_cell = sw.cell(2, f"={_surcharge_formula(f'({normal_for_slab_cell}+{cg_lt_cell.coordinate}+{cg_st_cell.coordinate})', after_rebate_ref, surcharge_refs)}", number_format=INR_FORMAT)
        surcharge_ref = surcharge_cell.coordinate
        sw.row += 1

        sw.cell(1, "Cess")
        cess_cell = sw.cell(2, f"=({after_rebate_ref}+{surcharge_ref})*'Rules'!{rules_layout['cess_rate'].coordinate}", number_format=INR_FORMAT)
        cess_ref = cess_cell.coordinate
        sw.row += 1

        sw.cell(1, "Tax liability")
        liability_cell = sw.cell(2, f"={after_rebate_ref}+{surcharge_ref}+{cess_ref}", number_format=INR_FORMAT)
        sw.row += 1
        return liability_cell.coordinate

    new_liability = _regime_block(
        "New", rules_layout["new_slabs"],
        rules_layout["new_rebate_max_ti"].coordinate, rules_layout["new_rebate_max_amt"].coordinate,
        rules_layout["new_rebate_marginal"].coordinate, rules_layout["new_surcharge"],
    )
    old_liability = _regime_block(
        "Old", rules_layout["old_slabs"],
        rules_layout["old_rebate_max_ti"].coordinate, rules_layout["old_rebate_max_amt"].coordinate,
        rules_layout["old_rebate_marginal"].coordinate, rules_layout["old_surcharge"],
    )
    sw.blank()

    sw.cell(1, "Selected regime (from Entity sheet)")
    sw.cell(2, f"={regime_cell}", number_format=None)
    sw.row += 1

    sw.cell(1, "Tax liability (selected regime)")
    selected_liability_cell = sw.cell(2, f"=IF({regime_cell}=\"old\",{old_liability},{new_liability})", number_format=INR_FORMAT).coordinate
    sw.row += 1

    taxes_paid_cell = sw.label_value("Taxes paid", f"='TaxesPaid'!{tp_layout['total']}")

    sw.cell(1, "Refund (+) / Payable (-), s.288B rounded")
    sw.cell(2, f"=MROUND({taxes_paid_cell.coordinate}-{selected_liability_cell},'Rules'!{rules_layout['round_tax_nearest'].coordinate})", number_format=INR_FORMAT)
    sw.row += 1


# ---------------------------------------------------------------------------
# Reconciliation sheet
# ---------------------------------------------------------------------------

def write_reconciliation_sheet(
    wb: Workbook, tree, identity_failures: list, book_cross_check: list, form16_cross_check: list,
    unmapped: list, mapping_version: str, rules_version: str, run_timestamp: str, input_hashes: dict,
    cg_reconciliation_ok: bool,
    split_year_exemption_prorated: bool = False, split_year_exemption_ratio: float = 1.0,
    as26_available: bool = False, as26_tie_out_ok: bool = True, as26_conflicts: list | None = None,
) -> None:
    ws = wb.create_sheet("Reconciliation")
    sw = _SheetWriter(ws)
    sw.header("Reconciliation")
    sw.label_value("Imbalance Amount", tree.imbalance)
    sw.label_value("Identity checks", "OK" if not identity_failures else f"{len(identity_failures)} FAILURE(S)", number_format=None)
    for f in identity_failures:
        sw.label_value("  failure", f, number_format=None)
    sw.label_value("Capital Gains reconciliation (lots vs control)", "OK" if cg_reconciliation_ok else "MISMATCH", number_format=None)
    if split_year_exemption_prorated:
        sw.label_value(
            "ASSUMPTION APPLIED -- split-year 112A exemption pro-rated across "
            "before/after split-date buckets (see Rules sheet)",
            split_year_exemption_ratio, number_format="0.0000",
        )
    if as26_available:
        sw.label_value("Book<->26AS TDS tie-out (see TaxesPaid sheet)", "OK" if as26_tie_out_ok else f"{len(as26_conflicts)} CONFLICT(S)", number_format=None)
        for c in as26_conflicts or []:
            sw.label_value(
                f"  CONFLICT -- {c['category']}",
                f"book={c['book']:.2f} 26AS={c['as26']:.2f} diff={c['diff']:.2f}",
                number_format=None,
            )
    sw.blank()

    sw.header("Book<->HTML cross-check")
    for r in book_cross_check:
        sw.label_value(r.name, "OK" if r.ok else f"MISMATCH html={r.html_total:.2f} book={r.book_total:.2f}", number_format=None)
    sw.blank()

    sw.header("Book<->Form16 cross-check")
    for r in form16_cross_check:
        sw.label_value(r.label, "OK" if r.ok else f"MISMATCH mapped={r.mapped_total:.2f} form16={r.form16_total:.2f}", number_format=None)
    sw.blank()

    sw.header(f"Unmapped accounts ({len(unmapped)})")
    for leaf in unmapped:
        sw.label_value(leaf.path, leaf.total)
    sw.blank()

    sw.label_value("Mapping version", mapping_version, number_format=None)
    sw.label_value("Rules version", rules_version, number_format=None)
    sw.label_value("Run timestamp", run_timestamp, number_format=None)
    for name, digest in input_hashes.items():
        sw.label_value(f"Input hash -- {name}", digest, number_format=None)


# ---------------------------------------------------------------------------
# Mapping Review sheet
# ---------------------------------------------------------------------------

_REVIEW_HEADERS = [
    "Account path", "FY amount", "Tag", "Destination", "Treatment", "Suggested-by", "Correction", "GUID",
]


def _suggested_by(guid: str, mapping_entries: dict | None) -> str:
    """Best-effort provenance label for one resolved leaf. A leaf with no
    entry of its own inherited its tag from an ancestor (mapping.py's
    nearest-ancestor rule); a leaf WITH an entry is heuristic/llm/approved
    depending on how that entry got there (bootstrap_mappings.py's
    heuristic scratch script vs. an LLM suggestion vs. a human-approved
    entry with neither marker)."""
    if not mapping_entries or guid not in mapping_entries:
        return "inherited"
    entry = mapping_entries[guid]
    if entry.suggested_by_llm:
        return "llm"
    if "heuristic" in (entry.note or "").lower():
        return "heuristic"
    return "approved"


def _destination(tag: str) -> tuple:
    meta = tag_vocab.TAGS.get(tag)
    if meta is None:
        return "?", tag, ""
    return meta.target, _TAG_LINE_LABEL.get(tag, tag), meta.treatment


def write_mapping_review_sheet(
    wb: Workbook, tree, resolved: dict, unmapped: list, mapping_entries: dict | None = None,
) -> None:
    """One row per mapped leaf (plus a highlighted unmapped block at top),
    grouped by destination sheet with a subtotal per group -- an
    Excel-native surface for a non-technical reviewer to sanity-check every
    mapping.yaml tag against the plain-English treatment it drives, and
    record corrections for scripts/apply_mapping_corrections.py to apply
    back to the mapping file."""
    ws = wb.create_sheet("Mapping Review")
    sw = _SheetWriter(ws)

    guid_amount = {n.guid: (n.total or 0.0) for n in tree.all_nodes() if n.guid and not n.children}

    def _header_row():
        for col, h in enumerate(_REVIEW_HEADERS, start=1):
            sw.cell(col, h, bold=True)
        sw.row += 1

    if unmapped:
        sw.header(f"UNMAPPED ({len(unmapped)}) -- needs a mapping entry before a full workbook can build")
        _header_row()
        for leaf in unmapped:
            row = sw.row
            sw.cell(1, leaf.path)
            sw.cell(2, leaf.total, number_format=INR_FORMAT)
            sw.cell(3, "REPLACE_ME")
            sw.cell(4, "(unresolved)")
            sw.cell(5, "(unresolved -- pick a tag from tags.py)")
            sw.cell(6, "unmapped")
            sw.cell(7, None)
            sw.cell(8, leaf.guid)
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = _UNMAPPED_FILL
            sw.row += 1
        sw.blank()

    groups: dict[str, list] = {}
    for guid, rl in resolved.items():
        target, _, _ = _destination(rl.tag)
        groups.setdefault(target, []).append(rl)

    for target in sorted(groups):
        sw.header(f"Destination: {target}")
        _header_row()
        subtotal = 0.0
        for rl in sorted(groups[target], key=lambda r: r.path):
            _, line_label, treatment = _destination(rl.tag)
            amount = guid_amount.get(rl.guid, 0.0)
            subtotal += amount
            sw.cell(1, rl.path)
            sw.cell(2, amount, number_format=INR_FORMAT)
            sw.cell(3, rl.tag)
            sw.cell(4, f"{target} - {line_label}")
            sw.cell(5, treatment)
            sw.cell(6, _suggested_by(rl.guid, mapping_entries))
            sw.cell(7, None)
            sw.cell(8, rl.guid)
            sw.row += 1
        sw.cell(1, f"Subtotal -- {target}", bold=True)
        sw.cell(2, subtotal, bold=True, number_format=INR_FORMAT)
        sw.row += 1
        sw.blank()


# ---------------------------------------------------------------------------
# Top-level orchestrator
# ---------------------------------------------------------------------------

def write_workbook(
    output_path: str, tree, model: sch.ITRModel, rules: rules_engine.RulesConfig, user_rules: list,
    entity, regime: str, year_key: str, form16_election: str | None,
    identity_failures: list, book_cross_check: list, form16_cross_check: list,
    unmapped: list, mapping_version: str, run_timestamp: str, input_hashes: dict,
    resolved: dict | None = None, mapping_entries: dict | None = None,
) -> None:
    fy_end = date(int(year_key[:4]) + 1, 3, 31) if year_key else date.today()

    wb = Workbook()
    wb.remove(wb.active)

    rules_layout = write_rules_sheet(wb, rules, user_rules, entity.status, entity.dob, fy_end)
    entity_layout = write_entity_sheet(wb, entity, year_key, rules, regime, form16_election)
    salary_layout = write_salary_sheet(wb, model.salary)
    business_layout = write_business_sheet(wb, model.business)
    hp_layout = write_house_property_sheet(wb, model.house_property)
    write_schedule_fa_sheet(wb, model.schedule_fa)
    os_layout = write_other_sources_sheet(wb, model.other_sources)
    cg_layout = write_capital_gains_sheet(wb, model.capital_gains, rules_layout)
    write_exempt_income_sheet(wb, model.exempt_income)
    tp_layout = write_taxes_paid_sheet(wb, model.taxes_paid)
    ded_layout = write_deductions_sheet(wb, model.deductions)
    write_schedule_al_sheet(wb, model.schedule_al)
    write_is_transcript(wb, tree)
    write_bs_transcript(wb, tree)

    write_computation_sheet(
        wb, model, rules_layout, entity_layout, salary_layout, business_layout, hp_layout,
        os_layout, cg_layout, ded_layout, tp_layout,
    )

    write_reconciliation_sheet(
        wb, tree, identity_failures, book_cross_check, form16_cross_check, unmapped,
        mapping_version, rules.version, run_timestamp, input_hashes, model.capital_gains.reconciliation_ok,
        model.capital_gains.split_year_exemption_prorated, model.capital_gains.split_year_exemption_ratio,
        model.taxes_paid.as26_available, model.taxes_paid.tie_out_ok, model.taxes_paid.tie_out_conflicts,
    )

    write_mapping_review_sheet(wb, tree, resolved or {}, unmapped, mapping_entries)

    wb.save(output_path)
