"""
outlier_report.py -- Part 3: diffs an ITRModel (schedules.py's computed
figures) against a filed-return reference (filed_return.FiledReturn) or an
equivalent lower-confidence reference built from a CA working Excel (for an
entity with no filed return yet), and writes a per-entity
<entity>-outliers.xlsx: Section A (line-level MATCH/DIFF/NO-REFERENCE),
Section B (drill-down to the mapped accounts feeding each DIFF line),
Section C (conservative classification), plus a Summary sheet.

Classification is deliberately conservative -- UNCLASSIFIED unless a narrow,
specific signal justifies mapping-suspect / rules-suspect / data-gap /
filed-side-note (spec: "default to UNCLASSIFIED when unsure"). A DIFF whose
reference came from a CA working Excel (not an actual filed return) is
always filed-side-note, never asserted as our error or theirs.
"""
from __future__ import annotations

from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

NOT_EXTRACTED = "NOT-EXTRACTED"
TOLERANCE_RUPEE = 1.0  # filed returns are whole-rupee; our figures may carry paisa

MAPPING_SUSPECT = "mapping-suspect"
RULES_SUSPECT = "rules-suspect"
DATA_GAP = "data-gap"
FILED_SIDE_NOTE = "filed-side-note"
UNCLASSIFIED = "UNCLASSIFIED"

CA_WORKING_CONFIDENCE = "ca-working-lower-confidence"
FILED_RETURN_CONFIDENCE = "filed-return"

# Lines whose value is a direct sum of one or more mapping tags (drillable).
# Composite lines (GTI, Total Income, Net Tax Liability, Refund/Payable,
# Capital Gains, Deductions) are derived from multiple schedules and have no
# single tag set -- Section B notes them as "see component schedules".
LINE_TAGS = {
    "Salary income": ["SALARY_GROSS"],
    "House Property income": ["HP_RENT", "HP_MUNICIPAL_TAX", "HP_INTEREST"],
    "Other Sources (total)": [
        "OS_INTEREST_SB", "OS_INTEREST_BANK", "OS_INTEREST_NBFC",
        "OS_INTEREST_EPF_TAXABLE", "OS_REFUND_INTEREST", "OS_DIVIDEND", "OS_SLBS",
    ],
}

COMPOSITE_LINES = (
    "Capital Gains (total)", "Gross Total Income", "Deductions (Chapter VI-A)",
    "Total Income", "Net Tax Liability", "Refund/Payable",
)


@dataclass
class ReferenceLine:
    value: object  # float, or NOT_EXTRACTED
    confidence: str = FILED_RETURN_CONFIDENCE
    note: str = ""


@dataclass
class ReferenceSet:
    entity_key: str
    source_kind: str  # "filed-json" | "filed-pdf" | "ca-working"
    source_path: str
    lines: dict  # line_name -> ReferenceLine
    regime: str = NOT_EXTRACTED
    regime_note: str = ""


@dataclass
class LineDiff:
    line_name: str
    our_value: float
    ref_value: object
    diff: object
    status: str  # "MATCH" | "DIFF" | "NO-REFERENCE"
    confidence: str


@dataclass
class DrillRow:
    account_path: str
    tag: str
    total: float


def build_our_lines(model) -> dict:
    """Extracts the comparable computation-schedule figures from an
    ITRModel (schedules.py), keyed the same way as ReferenceSet.lines."""
    c = model.computation
    return {
        "Salary income": c.salary_income,
        "House Property income": c.house_property_income,
        "Capital Gains (total)": c.capital_gains_lt + c.capital_gains_st,
        "Other Sources (total)": c.other_sources_income,
        "Gross Total Income": c.gti,
        "Deductions (Chapter VI-A)": c.via_deductions,
        "Total Income": c.total_income_rounded,
        "Net Tax Liability": c.tax_block.tax_liability,
        "Refund/Payable": c.refund_or_payable,
    }


def reference_from_filed_return(fr) -> ReferenceSet:
    """Builds a ReferenceSet from a filed_return.FiledReturn (JSON or PDF
    path). Any field the parser itself flagged NOT-EXTRACTED stays
    NOT-EXTRACTED here -- never guessed."""
    heads = fr.heads
    tl = fr.tax_liability

    refund_due = tl.get("refund_due", NOT_EXTRACTED)
    bal_payable = tl.get("balance_tax_payable", NOT_EXTRACTED)
    if refund_due not in (NOT_EXTRACTED, None) and refund_due:
        refund_or_payable = refund_due
    elif bal_payable not in (NOT_EXTRACTED, None):
        refund_or_payable = -bal_payable
    else:
        refund_or_payable = NOT_EXTRACTED

    def rl(v):
        return ReferenceLine(value=v, confidence=FILED_RETURN_CONFIDENCE)

    lines = {
        "Salary income": rl(heads.get("Salaries", NOT_EXTRACTED)),
        "House Property income": rl(heads.get("IncomeFromHP", NOT_EXTRACTED)),
        "Capital Gains (total)": rl(heads.get("CapitalGainsTotal", NOT_EXTRACTED)),
        "Other Sources (total)": rl(heads.get("OtherSourcesTotal", NOT_EXTRACTED)),
        "Gross Total Income": rl(heads.get("GrossTotalIncome", NOT_EXTRACTED)),
        "Deductions (Chapter VI-A)": rl(heads.get("DeductionsVIA", NOT_EXTRACTED)),
        "Total Income": rl(heads.get("TotalIncome", NOT_EXTRACTED)),
        "Net Tax Liability": rl(tl.get("net_tax_liability", NOT_EXTRACTED)),
        "Refund/Payable": rl(refund_or_payable),
    }
    return ReferenceSet(
        entity_key=fr.entity_key, source_kind=f"filed-{fr.source_format}",
        source_path=fr.source_path, lines=lines,
        regime=fr.regime, regime_note=fr.regime_note,
    )


def reference_from_ca_working_excel(path: str | Path, sheet: str, entity_key: str) -> ReferenceSet:
    """Best-effort reference for an entity with NO filed return, built from a
    CA's own working Excel. openpyxl reads cached formula values only
    (data_only=True) -- when the sheet's own subtotal/total formulas were
    never recalculated by Excel (observed on Khyati's sheet: Gross Total
    Income / Total Income / Total Tax / REFUND DUE cells all cache as None),
    this falls back to summing the labeled COMPONENT cells that do carry a
    cached value, and leaves a line NOT-EXTRACTED when even its components
    are missing. Every derived figure is a sum of visible component cells,
    never a guess -- but confidence is always CA_WORKING_CONFIDENCE (lower),
    per spec addendum: reference = CA working, lower confidence, lean
    Section-C toward UNCLASSIFIED/filed-side-note rather than asserting
    either side is wrong. Label matching is heuristic and tuned to the
    common Indian CA "COMPUTATION OF TOTAL INCOME" sheet layout; it may not
    generalize to an arbitrarily-formatted working sheet."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[sheet]

    def row_label(row):
        for cell in row:
            if isinstance(cell, str) and cell.strip():
                return cell.strip()
        return None

    def row_last_number(row):
        for cell in reversed(row):
            if isinstance(cell, (int, float)):
                return float(cell)
        return None

    labeled: dict[str, float] = {}
    for row in ws.iter_rows(values_only=True):
        label = row_label(row)
        if label is None:
            continue
        val = row_last_number(row)
        if val is not None:
            labeled.setdefault(label, val)

    def find(*substrings):
        for label, val in labeled.items():
            low = label.lower()
            if all(s.lower() in low for s in substrings):
                return val
        return None

    salary_recd = find("salary", "recd")
    perquisites = find("value of perquis")
    std_ded = find("standard deduction")
    prof_tax = find("professional tax")
    salary_gross = (salary_recd or 0.0) + (perquisites or 0.0)
    # s.10 exemption components (LTA/HRA etc.) are summed as "Less: Allowance
    # u/s 10" sub-rows; without a stable label per sub-row this is left
    # NOT-EXTRACTED rather than guessed, so salary_income here is gross-less-
    # standard-deduction-and-prof-tax only when no s10 rows are involved.
    salary_income = NOT_EXTRACTED
    if salary_recd is not None and std_ded is not None:
        salary_income = salary_gross - (std_ded + (prof_tax or 0.0))

    os_components = [
        find("interest received on fd"),
        find("dividend on shares"),
        find("interest on refund"),
        find("interest on sb", "minor"),
        find("s/b a/c interest"),
    ]
    os_present = [v for v in os_components if v is not None]
    other_sources_total = sum(os_present) if os_present else NOT_EXTRACTED
    os_gap_note = "" if all(v is not None for v in os_components) else "one or more Other Sources component rows had no cached value"

    ded_80c = find("80 c") or find("80c")
    ded_80d = find("80d")
    ded_80g = find("80g")
    ded_80tta = find("80tta")
    via_components = [ded_80c, ded_80d, ded_80g, ded_80tta]
    via_total = sum(v for v in via_components if v is not None) if any(v is not None for v in via_components) else NOT_EXTRACTED

    gross_total_income = NOT_EXTRACTED
    if salary_income != NOT_EXTRACTED and other_sources_total != NOT_EXTRACTED:
        gross_total_income = salary_income + other_sources_total

    total_income = NOT_EXTRACTED
    if gross_total_income != NOT_EXTRACTED and via_total != NOT_EXTRACTED:
        total_income = gross_total_income - via_total

    def rl(v, note=""):
        return ReferenceLine(value=v, confidence=CA_WORKING_CONFIDENCE, note=note)

    lines = {
        "Salary income": rl(salary_income, "derived: Salary Recd + perquisites - standard deduction - professional tax"),
        "House Property income": rl(NOT_EXTRACTED, "no House Property section in this CA sheet"),
        "Capital Gains (total)": rl(NOT_EXTRACTED, "CA sheet's Capital Gain section had no line items this year"),
        "Other Sources (total)": rl(other_sources_total, os_gap_note),
        "Gross Total Income": rl(gross_total_income, "derived (sheet's own GTI formula cell has no cached value)"),
        "Deductions (Chapter VI-A)": rl(via_total, "derived from 80C/80D/80G/80TTA component cells"),
        "Total Income": rl(total_income, "derived (sheet's own Total Income formula cell has no cached value)"),
        "Net Tax Liability": rl(NOT_EXTRACTED, "sheet's Education Cess/Total Tax cells have no cached value"),
        "Refund/Payable": rl(NOT_EXTRACTED, "sheet's REFUND DUE/TAX PAYABLE cells have no cached value"),
    }
    return ReferenceSet(
        entity_key=entity_key, source_kind="ca-working", source_path=str(path),
        lines=lines, regime=NOT_EXTRACTED,
        regime_note="CA working Excel is not a filed return -- regime not independently confirmable from it",
    )


def compare_lines(our_lines: dict, reference: ReferenceSet) -> list[LineDiff]:
    diffs = []
    for name, our_val in our_lines.items():
        ref_line = reference.lines.get(name)
        if ref_line is None or ref_line.value == NOT_EXTRACTED:
            confidence = ref_line.confidence if ref_line is not None else "n/a"
            diffs.append(LineDiff(name, our_val, NOT_EXTRACTED, NOT_EXTRACTED, "NO-REFERENCE", confidence))
            continue
        d = our_val - ref_line.value
        status = "MATCH" if abs(d) <= TOLERANCE_RUPEE else "DIFF"
        diffs.append(LineDiff(name, our_val, ref_line.value, d, status, ref_line.confidence))
    return diffs


def drill_down(line_name: str, resolved: dict, node_by_guid: dict) -> list[DrillRow]:
    tags = LINE_TAGS.get(line_name, [])
    rows = []
    for leaf in resolved.values():
        if leaf.tag in tags and leaf.guid in node_by_guid:
            node = node_by_guid[leaf.guid]
            rows.append(DrillRow(account_path=node.path, tag=leaf.tag, total=node.total or 0.0))
    return rows


def classify(diff: LineDiff, drill_rows: list[DrillRow]) -> str:
    """Conservative by design -- default UNCLASSIFIED. Only two narrow,
    specific signals are asserted automatically:
      - reference came from a CA working Excel (not a filed return): always
        filed-side-note, never our-side/their-side fault.
      - a drillable line (single tag set) has a filed DIFF but literally no
        mapped account feeding it: data-gap (nothing was ever mapped here).
    Everything else -- including any line where drill_rows exist but still
    don't reconcile -- stays UNCLASSIFIED; that judgment call needs a human.
    """
    if diff.status != "DIFF":
        return ""
    if diff.confidence == CA_WORKING_CONFIDENCE:
        return FILED_SIDE_NOTE
    if diff.line_name not in COMPOSITE_LINES and not drill_rows:
        return DATA_GAP
    return UNCLASSIFIED


def write_outlier_workbook(
    output_path: str, reference: ReferenceSet, diffs: list[LineDiff],
    drill: dict, classifications: dict,
) -> None:
    wb = openpyxl.Workbook()
    header_font = Font(bold=True)

    ws_a = wb.active
    ws_a.title = "Section A - Comparison"
    ws_a.append(["Line", "Our Value", "Reference Value", "Diff", "Status", "Reference Confidence"])
    for cell in ws_a[1]:
        cell.font = header_font
    for d in diffs:
        ws_a.append([d.line_name, d.our_value, d.ref_value, d.diff, d.status, d.confidence])

    ws_b = wb.create_sheet("Section B - Drilldown")
    ws_b.append(["Line", "Account Path", "Tag", "Account Total"])
    for cell in ws_b[1]:
        cell.font = header_font
    for d in diffs:
        if d.status != "DIFF":
            continue
        rows = drill.get(d.line_name, [])
        if not rows:
            ws_b.append([d.line_name, "(composite/derived line -- see component schedules)", "", ""])
        for r in rows:
            ws_b.append([d.line_name, r.account_path, r.tag, r.total])

    ws_c = wb.create_sheet("Section C - Classification")
    ws_c.append(["Line", "Status", "Classification", "Reference Confidence", "Note"])
    for cell in ws_c[1]:
        cell.font = header_font
    for d in diffs:
        cls = classifications.get(d.line_name, "")
        note = "reference = CA working, lower confidence" if d.confidence == CA_WORKING_CONFIDENCE else ""
        ws_c.append([d.line_name, d.status, cls, d.confidence, note])

    ws_s = wb.create_sheet("Summary")
    ws_s.append(["Entity", reference.entity_key])
    ws_s.append(["Reference source", reference.source_kind])
    ws_s.append(["Reference path", reference.source_path])
    ws_s.append(["Regime (filed/reference)", reference.regime])
    ws_s.append(["Regime note", reference.regime_note])
    ws_s.append([])
    match_n = sum(1 for d in diffs if d.status == "MATCH")
    diff_n = sum(1 for d in diffs if d.status == "DIFF")
    noref_n = sum(1 for d in diffs if d.status == "NO-REFERENCE")
    ws_s.append(["Lines compared", len(diffs)])
    ws_s.append(["MATCH", match_n])
    ws_s.append(["DIFF", diff_n])
    ws_s.append(["NO-REFERENCE", noref_n])
    ws_s.append([])
    ws_s.append(["Classification", "Count"])
    counts = Counter(classifications.get(d.line_name, "") for d in diffs if d.status == "DIFF")
    for cls, n in counts.items():
        ws_s.append([cls or "(n/a)", n])

    wb.save(output_path)


def run_outlier_report(model, reference: ReferenceSet, resolved: dict, node_by_guid: dict, output_path: str) -> list[LineDiff]:
    our_lines = build_our_lines(model)
    diffs = compare_lines(our_lines, reference)
    drill = {d.line_name: drill_down(d.line_name, resolved, node_by_guid) for d in diffs if d.status == "DIFF"}
    classifications = {d.line_name: classify(d, drill.get(d.line_name, [])) for d in diffs if d.status == "DIFF"}
    write_outlier_workbook(output_path, reference, diffs, drill, classifications)
    return diffs
