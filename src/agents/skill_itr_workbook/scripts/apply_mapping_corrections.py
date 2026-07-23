"""
apply_mapping_corrections.py -- reads a reviewed workbook's "Mapping Review"
sheet (write_workbook.write_mapping_review_sheet) and applies every
non-blank Correction cell that names a valid tag back into the entity's
mapping YAML. A touched entry (whether it was previously unmapped or
already mapped with a different tag) is marked approved -- suggested_by_llm
cleared, note replaced -- since a human just confirmed it via the
Correction cell.

By default (no `output_yaml` given) this writes IN PLACE to `mapping_file`,
after first writing a timestamped `.bak-<ISO8601>` backup alongside it --
the same discipline ui/tabs/itr_mapping_review.py's Save button already
uses. This is deliberate: a prior version of this script always wrote to a
separate `output_yaml` that required a manual review-and-rename step to
take effect, and in practice that manual step was never completed for any
real entity -- every correction ever run through the old flow was silently
stranded, and the mapping files kept resolving 100% "heuristic" forever.
Writing in place by default closes that gap. Pass an explicit `output_yaml`
only if you deliberately want a side file to review before it goes live
(e.g. a dry run) -- doing so does NOT update the live mapping file, and
this script says so loudly. A Correction cell naming an unknown tag is
reported, never applied.

Usage (from the repo root, with the venv active):
    python src/agents/skill_itr_workbook/scripts/apply_mapping_corrections.py \\
        Data/itr/mappings/Harshal.mapping.yaml \\
        reviewed/Harshal-ITR.xlsx

    # dry run -- writes a side file instead of updating the live mapping:
    python src/agents/skill_itr_workbook/scripts/apply_mapping_corrections.py \\
        Data/itr/mappings/Harshal.mapping.yaml \\
        reviewed/Harshal-ITR.xlsx \\
        --output Harshal.mapping.proposed.yaml
"""
from __future__ import annotations

import argparse
import datetime
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import openpyxl  # noqa: E402

import tags as tag_vocab  # noqa: E402
from configs import MappingEntry, dump_mapping_entries, load_mapping  # noqa: E402

REVIEW_SHEET = "Mapping Review"
_APPROVED_NOTE = "approved via Mapping Review sheet correction"


def read_corrections(reviewed_xlsx: str) -> tuple[dict, list]:
    """Scan the reviewed workbook's Mapping Review sheet. Returns
    (valid: {guid: (tag, path)}, invalid: [(path, guid, bad_tag)]) for every
    row with a non-blank Correction cell."""
    wb = openpyxl.load_workbook(reviewed_xlsx, data_only=True)
    ws = wb[REVIEW_SHEET]

    valid: dict[str, tuple[str, str]] = {}
    invalid: list[tuple] = []
    in_data_block = False

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if row[0] == "Account path":
            in_data_block = True
            continue
        if not in_data_block:
            continue
        if row[0] is None or (isinstance(row[0], str) and row[0].startswith(("Destination:", "UNMAPPED", "Subtotal"))):
            in_data_block = False
            continue

        path, guid = row[0], row[7] if len(row) > 7 else None
        correction = row[6] if len(row) > 6 else None
        if not guid or correction is None:
            continue
        correction = str(correction).strip()
        if not correction:
            continue

        if tag_vocab.is_valid_tag(correction):
            valid[guid] = (correction, str(path))
        else:
            invalid.append((path, guid, correction))

    return valid, invalid


def read_current_paths(reviewed_xlsx: str) -> dict:
    """Scan the reviewed workbook's Mapping Review sheet and return
    {guid: current_path} for every mapped data row. write_mapping_review_sheet
    always writes the resolved leaf's CURRENT tree path into the "Account
    path" column, so this is the tree's live guid -> path state without
    needing to re-parse the original HTML/tree (2026-07-23 path-drift fix) --
    the reviewed workbook already carries everything needed to self-heal
    stale `path:` entries in the mapping file."""
    wb = openpyxl.load_workbook(reviewed_xlsx, data_only=True)
    ws = wb[REVIEW_SHEET]

    known_paths: dict[str, str] = {}
    in_data_block = False

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if row[0] == "Account path":
            in_data_block = True
            continue
        if not in_data_block:
            continue
        if row[0] is None or (isinstance(row[0], str) and row[0].startswith(("Destination:", "UNMAPPED", "Subtotal"))):
            in_data_block = False
            continue

        path, guid = row[0], row[7] if len(row) > 7 else None
        if guid:
            known_paths[guid] = str(path)

    return known_paths


def refresh_drifted_paths(entries: dict, known_paths: dict) -> tuple[dict, int]:
    """Return a new {guid: MappingEntry} dict with every entry's stored
    `path` refreshed to match `known_paths` wherever the GUID is present
    there and the path differs -- a benign rename (see configs.load_mapping's
    "benign rename, auto-fixable" wording). GUIDs absent from `known_paths`
    are left completely untouched: that case is a real problem (deleted
    account, or the wrong book was loaded), never auto-healed. Returns
    (new entries dict, count of entries refreshed); `entries` itself is not
    mutated."""
    refreshed = dict(entries)
    count = 0
    for guid, entry in entries.items():
        new_path = known_paths.get(guid)
        if new_path is not None and new_path != entry.path:
            refreshed[guid] = MappingEntry(
                guid=entry.guid,
                path=new_path,
                tag=entry.tag,
                flags=entry.flags,
                note=entry.note,
                suggested_by_llm=entry.suggested_by_llm,
            )
            count += 1
    return refreshed, count


def apply_corrections_map(
    mapping_file: str,
    corrections: dict,
    output_yaml: str,
    paths: dict | None = None,
    known_paths: dict | None = None,
) -> tuple[int, list]:
    """Load `mapping_file` (an empty/missing file is treated as zero
    entries -- lets a true cold-start entity, which has no mapping file yet,
    be corrected straight into existence), apply every entry in
    `corrections` ({guid: tag}) that names a valid tag, and write the result
    to `output_yaml`. NEVER touches `mapping_file` in place -- the caller
    decides whether/how `output_yaml` replaces it (see the UI's backup
    discipline in ui/tabs/itr_mapping_review.py).

    A touched entry (whether previously unmapped or already mapped with a
    different tag) is marked approved: `suggested_by_llm` cleared, `note`
    replaced with `_APPROVED_NOTE` -- a human just confirmed it.

    `paths` supplies the account path for a guid that has no existing entry
    in `mapping_file` (a previously-unmapped leaf, whose path is only known
    from the proposed-mappings snippet / parsed tree, not the mapping file
    itself). A guid with neither an existing entry nor a supplied path is
    written with an empty path.

    `known_paths`, if supplied (guid -> current tree path, e.g. from
    read_current_paths()), refreshes any drifted `path:` on EVERY entry
    being written -- not just the ones touched by `corrections` -- before
    the file is written (2026-07-23 path-drift fix: "heal for free" on the
    next correction run). Omitting it (the default) leaves stored paths
    exactly as loaded, same as before this option existed.

    Returns (count applied, invalid corrections as [(path, guid, bad_tag)]),
    mirroring apply_corrections()'s return shape.
    """
    mp = Path(mapping_file)
    if mp.is_file():
        entries = dict(load_mapping(mapping_file).entries)
    else:
        entries = {}
    paths = paths or {}

    applied = 0
    invalid: list[tuple] = []
    for guid, tag in corrections.items():
        if not guid or tag is None:
            continue
        tag = str(tag).strip()
        if not tag:
            continue
        existing = entries.get(guid)
        path = paths.get(guid) or (existing.path if existing else "")
        if not tag_vocab.is_valid_tag(tag):
            invalid.append((path, guid, tag))
            continue
        entries[guid] = MappingEntry(
            guid=guid,
            path=path,
            tag=tag,
            flags=existing.flags if existing else [],
            note=_APPROVED_NOTE,
            suggested_by_llm=None,
        )
        applied += 1

    if known_paths:
        entries, _ = refresh_drifted_paths(entries, known_paths)

    Path(output_yaml).write_text(dump_mapping_entries(list(entries.values())), encoding="utf-8")
    return applied, invalid


def backup_mapping_file(mapping_file: str) -> str | None:
    """Write a timestamped `.bak-<ISO8601>` copy of `mapping_file` alongside
    it before an in-place overwrite (mirrors the UI's save discipline).
    Returns the backup path, or None if `mapping_file` doesn't exist yet
    (a true cold-start entity has nothing to back up)."""
    mp = Path(mapping_file)
    if not mp.is_file():
        return None
    stamp = datetime.datetime.now().strftime("%Y%m%dT%H%M%S")
    backup_path = mp.with_name(f"{mp.name}.bak-{stamp}")
    backup_path.write_text(mp.read_text(encoding="utf-8"), encoding="utf-8")
    return str(backup_path)


def apply_corrections(mapping_file: str, reviewed_xlsx: str, output_yaml: str | None = None) -> tuple[int, list, str | None]:
    """Load `mapping_file`, apply every valid correction from
    `reviewed_xlsx`, and write the result. If `output_yaml` is None
    (the default), writes IN PLACE to `mapping_file` -- after first taking a
    timestamped backup via `backup_mapping_file` -- so the correction is
    live for the very next run, no manual rename step required. If
    `output_yaml` is given explicitly, writes there instead and leaves
    `mapping_file` untouched (a deliberate dry run).

    Returns (count applied, invalid corrections, backup path or None).

    Thin wrapper around `apply_corrections_map`: reads the reviewed
    workbook's Correction cells (which already validate tags via
    `read_corrections`), then delegates the entries build + write. Every
    entry's `path:` (not only the ones being corrected) is also refreshed
    against the reviewed workbook's current guid -> path state
    (read_current_paths) -- so any account renamed since the mapping file
    was last written self-heals for free on this write, without a separate
    step (2026-07-23 path-drift fix)."""
    valid, invalid = read_corrections(reviewed_xlsx)
    corrections = {guid: tag for guid, (tag, _path) in valid.items()}
    paths = {guid: path for guid, (_tag, path) in valid.items()}
    known_paths = read_current_paths(reviewed_xlsx)

    backup_path = None
    target = output_yaml
    if target is None:
        target = mapping_file
        backup_path = backup_mapping_file(mapping_file)

    applied, _ = apply_corrections_map(mapping_file, corrections, target, paths=paths, known_paths=known_paths)
    return applied, invalid, backup_path


def refresh_paths(mapping_file: str, known_paths: dict, output_yaml: str | None = None) -> tuple[int, str | None]:
    """Refresh every drifted `path:` in `mapping_file` against `known_paths`
    with no tag corrections involved -- lets path drift clear on its own
    even when no reviewer correction is pending (the `--refresh-paths` CLI
    mode). A TRUE no-op when nothing has drifted: no backup is taken and
    nothing is written. When something did drift, mirrors
    apply_corrections()'s discipline -- backs up then writes in place unless
    `output_yaml` is given explicitly (a dry run that leaves `mapping_file`
    untouched).

    Returns (count refreshed, backup path or None)."""
    entries = dict(load_mapping(mapping_file).entries)
    refreshed, count = refresh_drifted_paths(entries, known_paths)
    if count == 0:
        return 0, None

    backup_path = None
    target = output_yaml
    if target is None:
        target = mapping_file
        backup_path = backup_mapping_file(mapping_file)

    Path(target).write_text(dump_mapping_entries(list(refreshed.values())), encoding="utf-8")
    return count, backup_path


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("mapping_file", help="The entity's current Data/itr/mappings/<entity>.mapping.yaml")
    parser.add_argument("reviewed_xlsx", help="The ITR workbook, reviewed, with Correction cells filled in")
    parser.add_argument(
        "output_yaml", nargs="?", default=None,
        help="Optional dry-run path. If omitted (recommended), corrections are written IN PLACE "
             "to mapping_file (with an automatic .bak-<timestamp> backup) and take effect on the "
             "next run. If given, corrections are written here instead and mapping_file is left "
             "unchanged -- they will NOT take effect until you apply them yourself.",
    )
    parser.add_argument(
        "--refresh-paths", action="store_true",
        help="Refresh stale path: entries against reviewed_xlsx's current guid -> path state and "
             "exit -- no Correction cells are read or applied. Use this to clear path-drift "
             "warnings (renamed accounts) on their own, with no pending tag correction. A true "
             "no-op (no backup, no write) when nothing has drifted.",
    )
    args = parser.parse_args()

    if args.refresh_paths:
        known_paths = read_current_paths(args.reviewed_xlsx)
        count, backup_path = refresh_paths(args.mapping_file, known_paths, args.output_yaml)
        target = args.output_yaml or args.mapping_file
        if count == 0:
            print("No drifted path(s) found -- nothing to refresh (no write).")
        else:
            if backup_path:
                print(f"Backed up previous mapping to {backup_path}")
            print(f"Refreshed {count} drifted path(s) -> {target}")
        return 0

    applied, invalid, backup_path = apply_corrections(args.mapping_file, args.reviewed_xlsx, args.output_yaml)
    if args.output_yaml is None:
        if backup_path:
            print(f"Backed up previous mapping to {backup_path}")
        print(f"Applied {applied} correction(s) -> {args.mapping_file} (LIVE -- takes effect on the next run)")
    else:
        print(f"Applied {applied} correction(s) -> {args.output_yaml}")
        print(f"NOTE: {args.mapping_file} was NOT modified. These corrections do not take effect "
              f"until you apply them (e.g. re-run without the output_yaml argument).")
    if invalid:
        print(f"{len(invalid)} correction(s) NOT applied (unknown tag):")
        for path, guid, tag in invalid:
            print(f"  - {path} (guid {guid}): {tag!r} is not a valid tag")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
