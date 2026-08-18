"""
parse_foreign.py -- parser for the broker "consolidated tax report" (INDmoney
and similar) that this skill uses to fill in Schedule FA, Schedule FSI,
Form 67 and Schedule TR for foreign holdings and foreign-sourced income.

Split the way skill_mf_cas/parser.py is split: a thin I/O layer opens the
workbook with openpyxl (data_only=True) and turns each sheet into a plain
``list[list]`` grid (merged-cell ranges pre-filled so every cell in the range
carries the anchor's value -- openpyxl otherwise returns ``None`` for the
non-anchor cells of a merged range); everything below that boundary is pure
Python operating on grids, never on openpyxl objects.

The real report's layout (transcribed into
``tests/skill_itr_workbook/fixture_gen_fa.py``, which is the authoritative
spec this module is written against) has several traps a naive parser walks
straight into:

  - headers are NOT at a fixed row -- three blank rows, a Client Name/PAN/
    (Financial Year or Report as on Date) label-value block, a title row,
    THEN the real header. Every header is found by scanning for its label
    text, never by a hardcoded row index.
  - two-row headers with horizontally merged parents (Schedule FA's "Gross
    amount paid/credited" over "Nature of Amount"/"Amount"; Form 67's "Tax
    paid outside India" over "Amount"/"Rate").
  - Schedule FSI's country/TIN block is merged VERTICALLY across four fixed
    head-of-income rows, and carries a "(a)".."(f)" column-letter row between
    the header and the data -- columns are resolved off those letters, not
    off position.
  - country cells carry an embedded newline (``"002\\nUnited States of
    America"``).
  - placeholder glyphs ("-", "#", "*") and whole cells of instructional
    prose ("To be filled by user basis Individual slab rates") sit where a
    reader expects a number. ``float()`` on these raises; silently coercing
    them to 0 is worse -- a 0 on a tax return is an assertion of fact. Every
    numeric field in this module is ``float | None``, parsed through the one
    shared ``parse_numeric_cell`` below, which returns ``(None, flag)``
    instead of ever guessing.
  - Schedule FA's A3 block is one row PER ACQUISITION LOT, not per entity --
    the same entity can appear on more than one row. Never dedupe by entity
    name.
  - "Amount (₹)" headers carry a rupee sign; header/label matching is by
    prefix/substring, never an exact string.

Two ties this module deliberately does NOT enforce, because they are
correct and "fixing" them would be wrong:

  - Schedule FA reports the CALENDAR year; Dividend/Interest report the
    FINANCIAL year -- FA's gross-amount figures are not expected to equal
    the Dividend/Interest totals.
  - Schedule FSI's "Other Sources" income IS expected to tie to the FY
    dividend total plus the FY interest total -- callers may check this via
    ``ScheduleFSIData.other_sources_income`` against ``DividendData.total +
    InterestData.total``.

Two deliberate parsing choices worth naming, since the fixture does not
settle them on its own (real-report variation was not something a synthetic
fixture can fully anticipate):

  - a REQUIRED sheet missing entirely (Schedule FA/FSI/Form 67/Schedule TR/
    Dividend/Interest) raises ``ForeignReportError`` -- that is a structural
    problem serious enough that the caller should stop and look, not carry
    on with a silently empty schedule.
  - a header row that cannot be found WITHIN a present sheet does not raise;
    it returns that schedule's dataclass with ``parsed=False`` and
    ``unparsed_reason`` set, and a note appended to the report's
    ``warnings`` list. A late change to one sheet's layout (e.g. a renamed
    column) should not stop every other sheet in the same report from being
    read.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

REQUIRED_SHEETS = (
    "Schedule FA", "Schedule FSI", "Form 67", "Schedule TR", "Dividend", "Interest",
)

_PLACEHOLDER_GLYPHS = {"-", "#", "*"}


class ForeignReportError(Exception):
    """Raised when the workbook cannot be opened, or a REQUIRED sheet is
    missing outright. NOT raised when a present sheet's header can't be
    found -- see module docstring."""


# ---------------------------------------------------------------------------
# Shared value parser -- never guesses, always flags.
# ---------------------------------------------------------------------------

def parse_numeric_cell(raw) -> tuple:
    """(value, flag). ``value`` is a float only when ``raw`` was a genuine
    number (or a numeric-looking string, comma-grouping tolerated).
    Otherwise ``value`` is None and ``flag`` names why:

      - "empty"            -- None or an all-whitespace string
      - "placeholder:<g>"  -- one of the glyphs "-", "#", "*"
      - "prose:<text>"     -- any other non-numeric string (instructional
                               text, e.g. "To be filled by user basis
                               Individual slab rates")

    Never returns 0.0 for a non-numeric cell -- a 0 on a tax return is an
    assertion of fact and must never be invented.
    """
    if raw is None:
        return None, "empty"
    if isinstance(raw, bool):
        return None, f"prose:{raw!r}"
    if isinstance(raw, (int, float)):
        return float(raw), None
    if isinstance(raw, str):
        s = raw.strip()
        if s == "":
            return None, "empty"
        if s in _PLACEHOLDER_GLYPHS:
            return None, f"placeholder:{s}"
        cleaned = s.replace(",", "")
        try:
            return float(cleaned), None
        except ValueError:
            return None, f"prose:{s[:120]}"
    return None, f"prose:{raw!r}"


def _clean_str(raw) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    return s or None


def split_country_cell(raw) -> tuple:
    """"002\\nUnited States of America" -> ("002", "United States of
    America"). Tolerates a stray leading space after the newline (Schedule
    TR's fixture cell has one, Schedule FSI's does not -- both are real
    layout variance, not something to "fix" away). Returns (None, None) for
    an empty cell, or (None, text) when there is no embedded newline to
    split on."""
    if raw is None:
        return None, None
    text = str(raw)
    if "\n" in text:
        code, _, name = text.partition("\n")
        return code.strip() or None, name.strip() or None
    return None, text.strip() or None


# ---------------------------------------------------------------------------
# I/O layer -- openpyxl in, plain grids out. Nothing below this point touches
# an openpyxl object.
# ---------------------------------------------------------------------------

def _grid_from_sheet(ws) -> list:
    """A plain list-of-lists snapshot of the sheet, with every cell of every
    merged range filled in with that range's top-left (anchor) value."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    grid = [
        [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col_r, max_row_r = rng.bounds
        if min_row > max_row or min_col > max_col:
            continue
        anchor = grid[min_row - 1][min_col - 1]
        for r in range(min_row, min(max_row_r, max_row) + 1):
            for c in range(min_col, min(max_col_r, max_col) + 1):
                grid[r - 1][c - 1] = anchor
    return grid


def _load_grids(path) -> dict:
    p = Path(path)
    try:
        wb = openpyxl.load_workbook(str(p), data_only=True, read_only=False)
    except Exception as e:  # noqa: BLE001 -- any openpyxl/zip failure is a report problem
        raise ForeignReportError(f"{p}: could not open workbook ({e})") from e

    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        raise ForeignReportError(
            f"{p}: required sheet(s) missing: {', '.join(missing)} "
            f"(found: {', '.join(wb.sheetnames)})"
        )
    return {name: _grid_from_sheet(wb[name]) for name in REQUIRED_SHEETS}


# ---------------------------------------------------------------------------
# Pure label-scan helpers (grid in, no openpyxl).
# ---------------------------------------------------------------------------

def _label_value(grid: list, label_substr: str, max_row: int = 15) -> str | None:
    """Scan the first `max_row` rows for a cell containing `label_substr`;
    return the next non-empty cell to its right on the same row."""
    for row in grid[:max_row]:
        for c, v in enumerate(row):
            if v is not None and label_substr in str(v):
                for v2 in row[c + 1:]:
                    if v2 is not None and str(v2).strip() != "":
                        return str(v2).strip()
    return None


def _find_header_row(grid: list, required_substrings, start_row: int = 1) -> int | None:
    """1-based row number of the first row whose cells, taken together,
    contain every one of `required_substrings` as a substring somewhere.
    Never a hardcoded row index -- callers only supply label text."""
    for r in range(start_row, len(grid) + 1):
        joined = " | ".join(str(v) for v in grid[r - 1] if v is not None)
        if all(sub in joined for sub in required_substrings):
            return r
    return None


def _col_index(row: list, *substrings, start_col: int = 1, occurrence: int = 1) -> int | None:
    """1-based column of the Nth (1-based `occurrence`) cell in `row` whose
    text contains ALL of `substrings`."""
    seen = 0
    for c in range(start_col, len(row) + 1):
        v = row[c - 1]
        if v is None:
            continue
        text = str(v)
        if all(sub in text for sub in substrings):
            seen += 1
            if seen == occurrence:
                return c
    return None


def _row_first_cell_text(row: list) -> str | None:
    return _clean_str(row[0]) if row else None


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class ForeignCustodialAccount:
    """One row of Schedule FA's A2 block."""
    country: str | None
    country_code: str | None
    institution: str | None
    address: str | None
    zip_code: str | None
    account_number: str | None
    status: str | None
    opening_date: str | None
    peak_balance: float | None
    closing_balance: float | None
    gross_amount_nature: str | None
    gross_amount: float | None
    flags: dict = field(default_factory=dict)


@dataclass
class ForeignEquityLot:
    """One row of Schedule FA's A3 block -- ONE ROW PER ACQUISITION LOT. The
    same entity may legitimately appear on more than one row; never dedupe
    by entity name."""
    country: str | None
    country_code: str | None
    entity: str | None
    address: str | None
    zip_code: str | None
    nature_of_entity: str | None
    date_acquired: str | None
    initial_value: float | None
    peak_value: float | None
    closing_value: float | None
    gross_amount_credited: float | None
    gross_proceeds: float | None
    broker: str | None
    flags: dict = field(default_factory=dict)


@dataclass
class ScheduleFAData:
    custodial_accounts: list = field(default_factory=list)   # list[ForeignCustodialAccount]
    equity_lots: list = field(default_factory=list)           # list[ForeignEquityLot]
    report_as_on: str | None = None
    parsed: bool = True
    unparsed_reason: str | None = None


@dataclass
class ScheduleFSIRow:
    head_of_income: str | None
    country_code: str | None
    country: str | None
    tin: str | None
    income: float | None
    tax_paid_outside: float | None
    tax_payable_india: float | None
    tax_relief: float | None
    dtaa_article: str | None
    flags: dict = field(default_factory=dict)


@dataclass
class ScheduleFSIData:
    rows: list = field(default_factory=list)   # list[ScheduleFSIRow]
    financial_year: str | None = None
    parsed: bool = True
    unparsed_reason: str | None = None

    @property
    def other_sources_income(self) -> float | None:
        for r in self.rows:
            if r.head_of_income and "Other Sources" in r.head_of_income:
                return r.income
        return None


@dataclass
class Form67Row:
    sr_no: int | None
    country: str | None
    source_of_income: str | None
    income: float | None
    tax_paid_amount: float | None
    tax_paid_rate: str | None
    dtaa_article: str | None
    dtaa_rate: str | None
    # Everything else on the row (tax-payable-under-normal-provisions,
    # 115JB/JC, credit-claimed-under-90/91, total-FTC-claimed, ...) is real
    # report content but the fixture's own header/sub-header rows disagree
    # with each other about which text describes which column -- rather than
    # guess a semantics the fixture doesn't settle, every other cell is kept
    # verbatim here, keyed by its HEADER-ROW label.
    other_cells: dict = field(default_factory=dict)
    flags: dict = field(default_factory=dict)


@dataclass
class Form67Data:
    rows: list = field(default_factory=list)   # list[Form67Row]
    financial_year: str | None = None
    parsed: bool = True
    unparsed_reason: str | None = None


@dataclass
class ScheduleTRRow:
    country_code: str | None
    country: str | None
    tin: str | None
    total_taxes_paid_outside: float | None
    total_relief_available: str | None   # usually prose, e.g. "From 'Schedule
                                          # FSI' sheet: Total of column (e)"
    relief_claimed_under_section: float | None
    dtaa_relief: str | None
    non_dtaa_relief: float | None
    refunded_or_credited: str | None
    flags: dict = field(default_factory=dict)


@dataclass
class ScheduleTRData:
    rows: list = field(default_factory=list)   # list[ScheduleTRRow]
    financial_year: str | None = None
    parsed: bool = True
    unparsed_reason: str | None = None


@dataclass
class QuarterlyBucket:
    label: str
    amount: float | None
    flag: str | None = None
    series: str | None = None   # e.g. "Dividend Income", "Dividend u/s 2(22)(f)"


@dataclass
class DividendPayment:
    company: str | None
    date_received: str | None
    fx_rate: float | None
    amount_usd: float | None
    amount_inr: float | None
    tax_usd: float | None
    tax_inr: float | None
    broker: str | None
    flags: dict = field(default_factory=dict)


@dataclass
class DividendData:
    payments: list = field(default_factory=list)   # list[DividendPayment]
    total: float | None = None
    quarterly: list = field(default_factory=list)   # list[QuarterlyBucket]
    financial_year: str | None = None
    parsed: bool = True
    unparsed_reason: str | None = None


@dataclass
class InterestReceipt:
    date_received: str | None
    fx_rate: float | None
    amount_usd: float | None
    amount_inr: float | None
    broker: str | None
    flags: dict = field(default_factory=dict)


@dataclass
class InterestData:
    receipts: list = field(default_factory=list)   # list[InterestReceipt]
    total: float | None = None
    financial_year: str | None = None
    parsed: bool = True
    unparsed_reason: str | None = None


@dataclass
class VendorLimit:
    key: str
    description: str
    present: bool


@dataclass
class ForeignIncomeReport:
    source_path: str
    financial_year: str | None
    report_as_on: str | None
    schedule_fa: ScheduleFAData
    schedule_fsi: ScheduleFSIData
    form_67: Form67Data
    schedule_tr: ScheduleTRData
    dividend: DividendData
    interest: InterestData
    vendor_limits: list = field(default_factory=list)   # list[VendorLimit]
    warnings: list = field(default_factory=list)


# ---------------------------------------------------------------------------
# Vendor-stated limitations (disclaimers) -- matched by stable substring,
# flagged when an EXPECTED one is absent rather than failing.
# ---------------------------------------------------------------------------

VENDOR_LIMIT_KEYS = (
    ("indglobal_dividends_excluded",
     "Report excludes INDglobal dividends (credited directly to the user's "
     "bank) -- add these manually.", "INDglobal dividends"),
    ("indglobal_us_tax_not_captured",
     "US tax withheld on INDglobal holdings is not captured by this report.",
     "US tax withheld on INDglobal"),
    ("delisted_stock_peak_closing_omitted",
     "Peak/closing value of delisted stock is omitted from Schedule FA.",
     "delisted stock"),
    ("corporate_actions_stock_splits_only",
     "Corporate actions covered are limited to stock splits/reverse splits.",
     "Stock Splits"),
    ("sbi_tt_buy_rate",
     "USD->INR conversion uses the SBI TT Buy rate.", "SBI TT Buy Rate"),
    ("us_timezone",
     "Report is prepared on US timezone (may not match an India-timezone "
     "statement).", "US timezone"),
)


def _detect_vendor_limits(grids: dict) -> list:
    blob = "\n".join(
        str(v) for grid in grids.values() for row in grid for v in row if isinstance(v, str)
    )
    return [
        VendorLimit(key=key, description=desc, present=substr in blob)
        for key, desc, substr in VENDOR_LIMIT_KEYS
    ]


# ---------------------------------------------------------------------------
# Schedule FA
# ---------------------------------------------------------------------------

def _parse_schedule_fa(grid: list, warnings: list) -> ScheduleFAData:
    report_as_on = _label_value(grid, "Report as on Date:") or _label_value(grid, "Financial Year:")

    header_a2 = _find_header_row(grid, ["Country Name", "Account Number", "Gross amount"])
    accounts: list[ForeignCustodialAccount] = []
    if header_a2 is None:
        warnings.append("Schedule FA: could not find the A2 (custodial accounts) header row.")
    else:
        header_row = grid[header_a2 - 1]
        sub_row = grid[header_a2] if header_a2 < len(grid) else []
        c_country = _col_index(header_row, "Country Name")
        c_country_code = _col_index(header_row, "Country Code")
        c_institution = _col_index(header_row, "Name of financial institution")
        c_address = _col_index(header_row, "Address of financial institution")
        c_zip = _col_index(header_row, "ZIP Code")
        c_account = _col_index(header_row, "Account Number")
        c_status = _col_index(header_row, "Status")
        c_opening = _col_index(header_row, "Account Opening date")
        c_peak = _col_index(header_row, "Peak Balance")
        c_closing = _col_index(header_row, "Closing Balance")
        c_nature = _col_index(sub_row, "Nature of Amount")
        # "Nature of Amount" itself contains the substring "Amount", so the
        # bare "Amount" sub-header must be matched as the SECOND occurrence
        # on this row, not the first.
        c_gross = _col_index(sub_row, "Amount", occurrence=2)

        def _get(row, col):
            return row[col - 1] if col and col - 1 < len(row) else None

        for r in range(header_a2 + 2, len(grid) + 1):
            row = grid[r - 1]
            first = _row_first_cell_text(row)
            if first is None or first.startswith("Disclaimer") or first.startswith("A3."):
                break
            peak_v, peak_f = parse_numeric_cell(_get(row, c_peak))
            close_v, close_f = parse_numeric_cell(_get(row, c_closing))
            gross_v, gross_f = parse_numeric_cell(_get(row, c_gross))
            flags = {}
            if peak_f:
                flags["peak_balance"] = peak_f
            if close_f:
                flags["closing_balance"] = close_f
            if gross_f:
                flags["gross_amount"] = gross_f
            accounts.append(ForeignCustodialAccount(
                country=_clean_str(_get(row, c_country)),
                country_code=_clean_str(_get(row, c_country_code)),
                institution=_clean_str(_get(row, c_institution)),
                address=_clean_str(_get(row, c_address)),
                zip_code=_clean_str(_get(row, c_zip)),
                account_number=_clean_str(_get(row, c_account)),
                status=_clean_str(_get(row, c_status)),
                opening_date=_clean_str(_get(row, c_opening)),
                peak_balance=peak_v, closing_balance=close_v,
                gross_amount_nature=_clean_str(_get(row, c_nature)),
                gross_amount=gross_v,
                flags=flags,
            ))

    header_a3 = _find_header_row(grid, ["Name of the entity", "Broker Name",
                                          "Total gross amount paid/credited"])
    lots: list[ForeignEquityLot] = []
    if header_a3 is None:
        warnings.append("Schedule FA: could not find the A3 (equity/debt interest) header row.")
    else:
        header_row = grid[header_a3 - 1]
        c_country = _col_index(header_row, "Country Name")
        c_country_code = _col_index(header_row, "Country Code")
        c_entity = _col_index(header_row, "Name of the entity")
        c_address = _col_index(header_row, "Address of the entity")
        c_zip = _col_index(header_row, "ZIP Code")
        c_nature = _col_index(header_row, "Nature of entity")
        c_acquired = _col_index(header_row, "Date of acquiring interest")
        c_initial = _col_index(header_row, "Initial value")
        c_peak = _col_index(header_row, "Peak value")
        c_closing = _col_index(header_row, "Closing Balance")
        c_gross_credited = _col_index(header_row, "Total gross amount paid/credited")
        c_gross_proceeds = _col_index(header_row, "Total gross proceeds")
        c_broker = _col_index(header_row, "Broker Name")

        def _get(row, col):
            return row[col - 1] if col and col - 1 < len(row) else None

        for r in range(header_a3 + 1, len(grid) + 1):
            row = grid[r - 1]
            first = _row_first_cell_text(row)
            if first is None or first.startswith("Disclaimer"):
                break
            initial_v, initial_f = parse_numeric_cell(_get(row, c_initial))
            peak_v, peak_f = parse_numeric_cell(_get(row, c_peak))
            close_v, close_f = parse_numeric_cell(_get(row, c_closing))
            gross_c_v, gross_c_f = parse_numeric_cell(_get(row, c_gross_credited))
            gross_p_v, gross_p_f = parse_numeric_cell(_get(row, c_gross_proceeds))
            flags = {}
            for name, flag in (
                ("initial_value", initial_f), ("peak_value", peak_f),
                ("closing_value", close_f), ("gross_amount_credited", gross_c_f),
                ("gross_proceeds", gross_p_f),
            ):
                if flag:
                    flags[name] = flag
            lots.append(ForeignEquityLot(
                country=_clean_str(_get(row, c_country)),
                country_code=_clean_str(_get(row, c_country_code)),
                entity=_clean_str(_get(row, c_entity)),
                address=_clean_str(_get(row, c_address)),
                zip_code=_clean_str(_get(row, c_zip)),
                nature_of_entity=_clean_str(_get(row, c_nature)),
                date_acquired=_clean_str(_get(row, c_acquired)),
                initial_value=initial_v, peak_value=peak_v, closing_value=close_v,
                gross_amount_credited=gross_c_v, gross_proceeds=gross_p_v,
                broker=_clean_str(_get(row, c_broker)),
                flags=flags,
            ))

    parsed = not (header_a2 is None and header_a3 is None)
    reason = None if parsed else "neither the A2 nor the A3 header row could be found"
    return ScheduleFAData(
        custodial_accounts=accounts, equity_lots=lots, report_as_on=report_as_on,
        parsed=parsed, unparsed_reason=reason,
    )


# ---------------------------------------------------------------------------
# Schedule FSI
# ---------------------------------------------------------------------------

def _parse_schedule_fsi(grid: list, warnings: list) -> ScheduleFSIData:
    financial_year = _label_value(grid, "Financial Year:")
    header = _find_header_row(grid, ["Head of Income", "Tax paid outside India",
                                       "Country Code"])
    if header is None:
        warnings.append("Schedule FSI: could not find the header row.")
        return ScheduleFSIData(financial_year=financial_year, parsed=False,
                                unparsed_reason="header row not found")

    letters_row = grid[header] if header < len(grid) else []
    letter_cols = {}
    for c, v in enumerate(letters_row, start=1):
        text = _clean_str(v)
        if text in ("(a)", "(b)", "(c)", "(d)", "(e)", "(f)"):
            letter_cols[text] = c
    if "(a)" not in letter_cols:
        warnings.append("Schedule FSI: could not find the (a)..(f) column-letter row.")
        return ScheduleFSIData(financial_year=financial_year, parsed=False,
                                unparsed_reason="(a)..(f) column-letter row not found")

    c_head = letter_cols["(a)"]
    c_income = letter_cols.get("(b)")
    c_tax_paid = letter_cols.get("(c)")
    c_tax_payable = letter_cols.get("(d)")
    c_tax_relief = letter_cols.get("(e)")
    c_article = letter_cols.get("(f)")
    header_row = grid[header - 1]
    c_country_code = _col_index(header_row, "Country Code")
    c_tin = _col_index(header_row, "Tax Identification Number")

    def _get(row, col):
        return row[col - 1] if col and col - 1 < len(row) else None

    rows: list[ScheduleFSIRow] = []
    # Four fixed heads of income (Salary, House Property, Capital Gains,
    # Other Sources) -- fixed by the report's own layout, not guessed; the
    # STARTING row is found by label scan above, only the count is a domain
    # fact.
    first = header + 2
    for r in range(first, first + 4):
        if r > len(grid):
            break
        row = grid[r - 1]
        income_v, income_f = parse_numeric_cell(_get(row, c_income))
        tax_paid_v, tax_paid_f = parse_numeric_cell(_get(row, c_tax_paid))
        tax_payable_v, tax_payable_f = parse_numeric_cell(_get(row, c_tax_payable))
        tax_relief_v, tax_relief_f = parse_numeric_cell(_get(row, c_tax_relief))
        flags = {}
        for name, flag in (
            ("income", income_f), ("tax_paid_outside", tax_paid_f),
            ("tax_payable_india", tax_payable_f), ("tax_relief", tax_relief_f),
        ):
            if flag:
                flags[name] = flag
        code, name_text = split_country_cell(_get(row, c_country_code))
        rows.append(ScheduleFSIRow(
            head_of_income=_clean_str(_get(row, c_head)),
            country_code=code, country=name_text,
            tin=_clean_str(_get(row, c_tin)),
            income=income_v, tax_paid_outside=tax_paid_v,
            tax_payable_india=tax_payable_v, tax_relief=tax_relief_v,
            dtaa_article=_clean_str(_get(row, c_article)),
            flags=flags,
        ))

    return ScheduleFSIData(rows=rows, financial_year=financial_year)


# ---------------------------------------------------------------------------
# Form 67
# ---------------------------------------------------------------------------

def _parse_form_67(grid: list, warnings: list) -> Form67Data:
    financial_year = _label_value(grid, "Financial Year:")
    header = _find_header_row(grid, ["Name of the country", "Tax paid outside India",
                                       "Source of Income"])
    if header is None:
        warnings.append("Form 67: could not find the header row.")
        return Form67Data(financial_year=financial_year, parsed=False,
                           unparsed_reason="header row not found")

    header_row = grid[header - 1]
    sub_row = grid[header] if header < len(grid) else []
    c_sr = _col_index(header_row, "Sr No")
    c_country = _col_index(header_row, "Name of the country")
    c_source = _col_index(header_row, "Source of Income")
    c_income = _col_index(header_row, "Income from outside India")
    c_tax_amount = _col_index(sub_row, "Amount")
    c_tax_rate = _col_index(sub_row, "Rate")
    c_dtaa_article = _col_index(sub_row, "Article No.")
    c_dtaa_rate = _col_index(sub_row, "Rate of tax as per")

    named_cols = {c for c in (c_sr, c_country, c_source, c_income, c_tax_amount,
                                c_tax_rate, c_dtaa_article, c_dtaa_rate) if c}

    def _get(row, col):
        return row[col - 1] if col and col - 1 < len(row) else None

    rows: list[Form67Row] = []
    for r in range(header + 2, len(grid) + 1):
        row = grid[r - 1]
        first = _row_first_cell_text(row)
        if first is None or first.startswith("Disclaimer"):
            break
        sr_v, sr_f = parse_numeric_cell(_get(row, c_sr))
        income_v, income_f = parse_numeric_cell(_get(row, c_income))
        tax_amount_v, tax_amount_f = parse_numeric_cell(_get(row, c_tax_amount))
        flags = {}
        if sr_f:
            flags["sr_no"] = sr_f
        if income_f:
            flags["income"] = income_f
        if tax_amount_f:
            flags["tax_paid_amount"] = tax_amount_f
        other_cells = {}
        for c in range(1, len(row) + 1):
            if c in named_cols:
                continue
            label = _clean_str(_get(header_row, c)) or f"column {c}"
            value = row[c - 1] if c - 1 < len(row) else None
            if value is not None:
                other_cells[label] = value
        rows.append(Form67Row(
            sr_no=int(sr_v) if sr_v is not None else None,
            country=_clean_str(_get(row, c_country)),
            source_of_income=_clean_str(_get(row, c_source)),
            income=income_v,
            tax_paid_amount=tax_amount_v,
            tax_paid_rate=_clean_str(_get(row, c_tax_rate)),
            dtaa_article=_clean_str(_get(row, c_dtaa_article)),
            dtaa_rate=_clean_str(_get(row, c_dtaa_rate)),
            other_cells=other_cells,
            flags=flags,
        ))

    return Form67Data(rows=rows, financial_year=financial_year)


# ---------------------------------------------------------------------------
# Schedule TR
# ---------------------------------------------------------------------------

def _parse_schedule_tr(grid: list, warnings: list) -> ScheduleTRData:
    financial_year = _label_value(grid, "Financial Year:")
    header = _find_header_row(grid, ["Total taxes paid outside India",
                                       "Tax relief claimed under section"])
    if header is None:
        warnings.append("Schedule TR: could not find the header row.")
        return ScheduleTRData(financial_year=financial_year, parsed=False,
                               unparsed_reason="header row not found")

    letters_row = grid[header] if header < len(grid) else []
    letter_cols = {}
    for c, v in enumerate(letters_row, start=1):
        text = _clean_str(v)
        if text in ("1(a)", "1(b)", "1(c)", "1(d)", "1(e)", "2", "3", "4"):
            letter_cols.setdefault(text, c)

    def _get(row, col):
        return row[col - 1] if col and col - 1 < len(row) else None

    c_country = letter_cols.get("1(a)")
    c_tin = letter_cols.get("1(b)")
    c_taxes_paid = letter_cols.get("1(c)")
    c_relief_available = letter_cols.get("1(d)")
    c_relief_claimed = letter_cols.get("1(e)")
    c_dtaa_relief = letter_cols.get("2")
    c_non_dtaa_relief = letter_cols.get("3")
    c_refunded = letter_cols.get("4")

    rows: list[ScheduleTRRow] = []
    for r in range(header + 2, len(grid) + 1):
        row = grid[r - 1]
        first = _row_first_cell_text(row)
        if first is None or first.startswith("Disclaimer"):
            break
        taxes_v, taxes_f = parse_numeric_cell(_get(row, c_taxes_paid))
        relief_claimed_v, relief_claimed_f = parse_numeric_cell(_get(row, c_relief_claimed))
        non_dtaa_v, non_dtaa_f = parse_numeric_cell(_get(row, c_non_dtaa_relief))
        flags = {}
        if taxes_f:
            flags["total_taxes_paid_outside"] = taxes_f
        if relief_claimed_f:
            flags["relief_claimed_under_section"] = relief_claimed_f
        if non_dtaa_f:
            flags["non_dtaa_relief"] = non_dtaa_f
        code, name_text = split_country_cell(_get(row, c_country))
        rows.append(ScheduleTRRow(
            country_code=code, country=name_text,
            tin=_clean_str(_get(row, c_tin)),
            total_taxes_paid_outside=taxes_v,
            total_relief_available=_clean_str(_get(row, c_relief_available)),
            relief_claimed_under_section=relief_claimed_v,
            dtaa_relief=_clean_str(_get(row, c_dtaa_relief)),
            non_dtaa_relief=non_dtaa_v,
            refunded_or_credited=_clean_str(_get(row, c_refunded)),
            flags=flags,
        ))

    return ScheduleTRData(rows=rows, financial_year=financial_year)


# ---------------------------------------------------------------------------
# Dividend
# ---------------------------------------------------------------------------

def _parse_dividend(grid: list, warnings: list) -> DividendData:
    financial_year = _label_value(grid, "Financial Year:")
    summary_header = _find_header_row(grid, ["Particulars", "Amount", "ITR Section"])
    total = None
    total_flag = None
    if summary_header is None:
        warnings.append("Dividend: could not find the summary (Particulars/Amount/ITR "
                         "Section) header row.")
    else:
        for r in range(summary_header + 1, len(grid) + 1):
            row = grid[r - 1]
            first = _row_first_cell_text(row)
            if first == "Total":
                total, total_flag = parse_numeric_cell(row[1] if len(row) > 1 else None)
                break
            if first is None:
                break

    quarterly: list[QuarterlyBucket] = []
    qb_title = _find_header_row(grid, ["Quaterly Break", "Up"])
    if qb_title is None:
        # Real-world reports may fix the "Quaterly" typo; tolerate either spelling.
        qb_title = _find_header_row(grid, ["Quarterly Break", "Up"])
    if qb_title is None:
        warnings.append("Dividend: could not find the Quarterly Break-Up title row.")
    else:
        # The label row ("Particulars" + the five period headers) is NOT
        # guaranteed to sit immediately under the title row -- the real
        # report inserts a completely blank spacer row first. Scan forward
        # for the first row with at least two non-empty cells from column 2
        # onward, bounded so a missing block can't run into an unrelated
        # section further down the sheet.
        label_row_num = None
        for r in range(qb_title + 1, min(qb_title + 5, len(grid)) + 1):
            non_empty = sum(1 for v in grid[r - 1][1:] if _clean_str(v))
            if non_empty >= 2:
                label_row_num = r
                break
        if label_row_num is None:
            warnings.append(
                "Dividend: found the Quarterly Break-Up title row but no label row "
                "(\"Particulars\" + period headers) within 5 rows below it."
            )
        else:
            label_row = grid[label_row_num - 1]
            periods = {}
            for c in range(2, len(label_row) + 1):
                label = _clean_str(label_row[c - 1])
                if label:
                    periods[c] = label
            # Each row after the label row is one SERIES (e.g. "Dividend
            # Income", "Dividend u/s 2(22)(f)" -- buyback dividend, taxed
            # differently) -- not just the first. Stop at the first row whose
            # first cell is empty, bounded so a malformed sheet can't run
            # away past this block.
            for r in range(label_row_num + 1, min(label_row_num + 10, len(grid)) + 1):
                row = grid[r - 1]
                series_name = _row_first_cell_text(row)
                if series_name is None:
                    break
                for c, label in periods.items():
                    value = row[c - 1] if c - 1 < len(row) else None
                    v, f = parse_numeric_cell(value)
                    quarterly.append(QuarterlyBucket(label=label, amount=v, flag=f,
                                                       series=series_name))
            if not quarterly:
                warnings.append(
                    "Dividend: found the Quarterly Break-Up label row but no series data "
                    "rows beneath it."
                )

    us_header_row_idx = _find_header_row(grid, ["Dividends From US Stocks"])
    payments: list[DividendPayment] = []
    if us_header_row_idx is not None:
        header = _find_header_row(grid, ["Name of the company", "Date Received", "Broker Name"],
                                   start_row=us_header_row_idx)
        if header is not None:
            header_row = grid[header - 1]
            c_name = _col_index(header_row, "Name of the company")
            c_date = _col_index(header_row, "Date Received")
            c_rate = _col_index(header_row, "Exchange Rate")
            c_amount = _col_index(header_row, "Amount")
            c_tax = _col_index(header_row, "Tax deducted")
            c_broker = _col_index(header_row, "Broker Name")

            def _get(row, col):
                return row[col - 1] if col and col - 1 < len(row) else None

            for r in range(header + 2, len(grid) + 1):
                row = grid[r - 1]
                first = _row_first_cell_text(row)
                if first is None or first == "Total" or first.startswith("Disclaimer"):
                    break
                rate_v, rate_f = parse_numeric_cell(_get(row, c_rate))
                usd_v, usd_f = parse_numeric_cell(_get(row, c_amount))
                inr_v, inr_f = parse_numeric_cell(_get(row, c_amount + 1) if c_amount else None)
                tax_usd_v, tax_usd_f = parse_numeric_cell(_get(row, c_tax))
                tax_inr_v, tax_inr_f = parse_numeric_cell(_get(row, c_tax + 1) if c_tax else None)
                flags = {}
                for name, fl in (
                    ("fx_rate", rate_f), ("amount_usd", usd_f), ("amount_inr", inr_f),
                    ("tax_usd", tax_usd_f), ("tax_inr", tax_inr_f),
                ):
                    if fl:
                        flags[name] = fl
                payments.append(DividendPayment(
                    company=_clean_str(_get(row, c_name)),
                    date_received=_clean_str(_get(row, c_date)),
                    fx_rate=rate_v, amount_usd=usd_v, amount_inr=inr_v,
                    tax_usd=tax_usd_v, tax_inr=tax_inr_v,
                    broker=_clean_str(_get(row, c_broker)),
                    flags=flags,
                ))
        else:
            warnings.append("Dividend: could not find the 'Dividends From US Stocks' detail header.")

    parsed = total is not None or bool(payments) or bool(quarterly)
    reason = None if parsed else "no dividend total, quarterly block, or US-stock payment rows found"
    return DividendData(
        payments=payments, total=total, quarterly=quarterly, financial_year=financial_year,
        parsed=parsed, unparsed_reason=reason,
    )


# ---------------------------------------------------------------------------
# Interest
# ---------------------------------------------------------------------------

def _parse_interest(grid: list, warnings: list) -> InterestData:
    financial_year = _label_value(grid, "Financial Year:")
    total = None
    for row in grid:
        first = _row_first_cell_text(row)
        if first and "Total Interest received" in first:
            total, _ = parse_numeric_cell(row[1] if len(row) > 1 else None)
            break
    if total is None:
        warnings.append("Interest: could not find the 'Total Interest received' summary cell.")

    receipts: list[InterestReceipt] = []
    header = _find_header_row(grid, ["Date Received", "Exchange Rate", "Broker Name"])
    if header is not None:
        header_row = grid[header - 1]
        c_date = _col_index(header_row, "Date Received")
        c_rate = _col_index(header_row, "Exchange Rate")
        c_amount = _col_index(header_row, "Amount")
        c_broker = _col_index(header_row, "Broker Name")

        def _get(row, col):
            return row[col - 1] if col and col - 1 < len(row) else None

        for r in range(header + 2, len(grid) + 1):
            row = grid[r - 1]
            first = _row_first_cell_text(row)
            if first is None or first == "Total":
                break
            rate_v, rate_f = parse_numeric_cell(_get(row, c_rate))
            usd_v, usd_f = parse_numeric_cell(_get(row, c_amount))
            inr_v, inr_f = parse_numeric_cell(_get(row, c_amount + 1) if c_amount else None)
            flags = {}
            for name, fl in (("fx_rate", rate_f), ("amount_usd", usd_f), ("amount_inr", inr_f)):
                if fl:
                    flags[name] = fl
            receipts.append(InterestReceipt(
                date_received=_clean_str(_get(row, c_date)),
                fx_rate=rate_v, amount_usd=usd_v, amount_inr=inr_v,
                broker=_clean_str(_get(row, c_broker)),
                flags=flags,
            ))
    else:
        warnings.append("Interest: could not find the Transactions detail header.")

    parsed = total is not None or bool(receipts)
    reason = None if parsed else "no interest total or transaction rows found"
    return InterestData(receipts=receipts, total=total, financial_year=financial_year,
                         parsed=parsed, unparsed_reason=reason)


# ---------------------------------------------------------------------------
# Top-level entry point
# ---------------------------------------------------------------------------

def load_foreign_income_report(path) -> ForeignIncomeReport:
    """The only I/O + pure-parse orchestration point. Raises
    ``ForeignReportError`` if the workbook can't be opened or a required
    sheet is missing outright (see module docstring); otherwise always
    returns a ``ForeignIncomeReport`` -- individual schedules that couldn't
    be located within their sheet come back with ``parsed=False`` and a
    note in ``.warnings`` rather than raising."""
    grids = _load_grids(path)
    warnings: list[str] = []

    schedule_fa = _parse_schedule_fa(grids["Schedule FA"], warnings)
    schedule_fsi = _parse_schedule_fsi(grids["Schedule FSI"], warnings)
    form_67 = _parse_form_67(grids["Form 67"], warnings)
    schedule_tr = _parse_schedule_tr(grids["Schedule TR"], warnings)
    dividend = _parse_dividend(grids["Dividend"], warnings)
    interest = _parse_interest(grids["Interest"], warnings)

    financial_year = (
        schedule_fsi.financial_year or form_67.financial_year or schedule_tr.financial_year
        or dividend.financial_year or interest.financial_year
    )
    vendor_limits = _detect_vendor_limits(grids)
    for vl in vendor_limits:
        if not vl.present:
            warnings.append(f"Vendor limitation not found in source (expected, verify manually): "
                             f"{vl.description}")

    return ForeignIncomeReport(
        source_path=str(path), financial_year=financial_year, report_as_on=schedule_fa.report_as_on,
        schedule_fa=schedule_fa, schedule_fsi=schedule_fsi, form_67=form_67, schedule_tr=schedule_tr,
        dividend=dividend, interest=interest, vendor_limits=vendor_limits, warnings=warnings,
    )
