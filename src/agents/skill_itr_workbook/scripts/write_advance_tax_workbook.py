"""
write_advance_tax_workbook.py -- openpyxl writer for the Advance Tax
Estimator output (2026-08 advance-tax-estimator work, PR1/engine).

DESIGN NOTE (brief silent on mechanism; deviation from write_workbook.py's
full Rules-sheet-driven slab/surcharge formula system, documented per
constraint 0.14): the tax liability itself (slabs/rebate/surcharge/cess) is
computed once in Python via the single shared tax_core.compute_tax_on and
written to the Computation sheet as plain values -- same treatment
write_workbook.py itself gives "schedule sheets' own detail rows (parsed
input values -- these ARE the traceable source data)". Reproducing that
entire formula system for a brand-new dual-regime, per-CG-item estimator
was judged disproportionate; what the brief actually enumerates as required
output -- the cumulative statutory instalment schedule, required-cumulative
vs paid-cumulative vs shortfall, and "pay this much by this date" -- IS
fully formula-driven here, referencing ONLY cells on this workbook (the
Computation sheet's Tax Liability/TDS cells and the Advance Tax Paid input
cells), never a baked number for that arithmetic.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import advance_tax as adv

INR_FORMAT = "#,##,##0.00"
_FLAG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FLAG_FONT = Font(color="9C0006", bold=True)

_INSTALMENT_LABELS = ("Q1 (by 15 Jun)", "Q2 (by 15 Sep)", "Q3 (by 15 Dec)", "Q4 (by 15 Mar)")
_INSTALMENT_PCT = (0.15, 0.45, 0.75, 1.00)


def _bold(cell) -> None:
    cell.font = Font(bold=True)


def _flag(cell) -> None:
    cell.fill = _FLAG_FILL
    cell.font = _FLAG_FONT


class _SW:
    def __init__(self, ws):
        self.ws = ws
        self.row = 1

    def cell(self, col: int, value=None, *, bold: bool = False, number_format: str | None = None, flag: bool = False):
        c = self.ws.cell(row=self.row, column=col, value=value)
        if bold:
            _bold(c)
        if number_format:
            c.number_format = number_format
        if flag:
            _flag(c)
        return c

    def label_value(self, label: str, value=None, *, number_format: str | None = INR_FORMAT):
        self.cell(1, label)
        c = self.cell(2, value, number_format=number_format)
        self.row += 1
        return c

    def header(self, text: str):
        self.cell(1, text, bold=True)
        self.row += 1

    def blank(self):
        self.row += 1


def _write_computation_sheet(ws, estimate: "adv.AdvanceTaxEstimate", regime_estimate: "adv.RegimeEstimate", label: str) -> dict:
    """Writes the (Python-computed, plain-value) tax computation for one
    regime. Returns {"tax_liability": "<A1-ref>", "tds": "<A1-ref>"} so the
    Instalment Schedule sheet can reference these cells by formula."""
    sw = _SW(ws)
    sw.header(f"Computation -- {label} regime")
    sw.blank()
    sw.label_value("Salary (full year, incl. projection)", estimate.input_data.salary.full_year)
    sw.label_value("Remuneration + interest on capital (PGBP)", estimate.input_data.remuneration.full_year)
    sw.label_value("House property income", estimate.input_data.rent.full_year)
    sw.label_value("Other-sources interest", estimate.input_data.interest.full_year)
    sw.label_value("Other sources (dividend etc., actual to date only)", estimate.input_data.other_sources_other)
    sw.blank()
    sw.label_value("Normal income (after VIA deductions)", regime_estimate.normal_income)
    sw.label_value("LTCG 112A (after exemption)", regime_estimate.cg_tax.lt_taxable_after_exemption)
    sw.label_value("STCG 111A", regime_estimate.cg_tax.st_taxable)
    sw.label_value("Special-rate tax (112A + 111A)", regime_estimate.cg_tax.special_rate_tax)
    sw.blank()
    tax_row = sw.row
    sw.label_value("Tax on normal income", regime_estimate.tax_block.tax_on_normal_income)
    sw.label_value("Rebate u/s 87A", regime_estimate.tax_block.rebate_87a)
    sw.label_value("Surcharge", regime_estimate.tax_block.surcharge)
    sw.label_value("Cess", regime_estimate.tax_block.cess)
    liability_row = sw.row
    sw.label_value("Tax liability", regime_estimate.tax_block.tax_liability, number_format=INR_FORMAT)
    _bold(ws.cell(row=liability_row, column=1))
    _bold(ws.cell(row=liability_row, column=2))
    tds_row = sw.row
    sw.label_value("TDS credited to date", estimate.input_data.tds_to_date)
    sw.blank()

    sheet_name = ws.title
    return {
        "tax_liability": f"'{sheet_name}'!B{liability_row}",
        "tds": f"'{sheet_name}'!B{tds_row}",
    }


def _write_instalment_sheet(ws, estimate: "adv.AdvanceTaxEstimate", regime_estimate: "adv.RegimeEstimate", refs: dict, label: str) -> None:
    sw = _SW(ws)
    sw.header(f"Advance Tax Instalment Schedule -- {label} regime")
    sw.blank()

    sw.cell(1, "Tax due on returned income (= Tax liability - TDS to date)", bold=True)
    due_row = sw.row
    sw.cell(2, f"=MAX(0,{refs['tax_liability']}-{refs['tds']})", number_format=INR_FORMAT)
    sw.row += 2

    sw.cell(1, "Instalment", bold=True)
    sw.cell(2, "Due date", bold=True)
    sw.cell(3, "Required %", bold=True)
    sw.cell(4, "Excluded (s.234C 1st proviso)", bold=True)
    sw.cell(5, "Required (cumulative)", bold=True)
    sw.cell(6, "Paid (cumulative)", bold=True)
    sw.cell(7, "Shortfall", bold=True)
    sw.cell(8, "Interest u/s 234C accrued", bold=True)
    sw.row += 1

    due_cell = f"$B${due_row}"
    fy_start = int(estimate.input_data.fy[:4])
    due_dates = [
        f"{fy_start if m != 3 else fy_start + 1}-{m:02d}-{d:02d}"
        for (m, d) in ((6, 15), (9, 15), (12, 15), (3, 15))
    ]
    for i, (inst_label, pct, due_date) in enumerate(zip(_INSTALMENT_LABELS, _INSTALMENT_PCT, due_dates)):
        inst = regime_estimate.interest_234c.instalments[i]
        row = sw.row
        sw.cell(1, inst_label)
        sw.cell(2, due_date)
        sw.cell(3, pct)
        sw.cell(4, _excluded_amount(estimate, regime_estimate, i), number_format=INR_FORMAT)
        sw.cell(5, f"=MAX(0,({due_cell}-D{row})*{pct})", number_format=INR_FORMAT)
        sw.cell(6, estimate.advance_tax_cumulative_paid[i], number_format=INR_FORMAT)
        sw.cell(7, f"=MAX(0,E{row}-F{row})", number_format=INR_FORMAT)
        sw.cell(8, inst.amount, number_format=INR_FORMAT)
        sw.row += 1


def _excluded_amount(estimate, regime_estimate, index: int) -> float:
    # Recomputed here (not re-exported by advance_tax.py's private helper)
    # from the already-built Interest234C result: required_amount at safe_pct
    # scaling isn't the exclusion itself, so instead read it back from the
    # instalment's own considered-vs-full base.
    full_base = max(0.0, regime_estimate.tax_block.tax_liability - estimate.input_data.tds_to_date)
    considered = regime_estimate.interest_234c.instalments[index].tax_due_considered
    return max(0.0, full_base - considered)


def _write_flags_sheet(ws, estimate: "adv.AdvanceTaxEstimate") -> None:
    sw = _SW(ws)
    sw.header("Flags")
    sw.blank()

    if estimate.unprojected_heads:
        c = sw.cell(1, "UNPROJECTED HEADS (blank remainder-of-year, treated as NOT estimated -- fill in or run Annualise):")
        _flag(c)
        sw.row += 1
        for name in estimate.unprojected_heads:
            sw.cell(1, f"  - {name}")
            sw.row += 1
        sw.blank()
    else:
        sw.cell(1, "All recurring heads have a projected remainder.")
        sw.row += 1
        sw.blank()

    if estimate.undated_capital_gains:
        c = sw.cell(
            1,
            f"UNDATED CAPITAL GAINS ITEMS: {len(estimate.undated_capital_gains)} item(s), "
            f"total {estimate.undated_cg_amount:,.2f} -- NO s.234C first-proviso relief given "
            "(conservative default; verify sale dates before filing).",
        )
        _flag(c)
        sw.row += 1
        for item in estimate.undated_capital_gains:
            sw.cell(1, f"  - {item.gain_type}: {item.amount:,.2f}" + (f" ({item.note})" if item.note else ""))
            sw.row += 1
    else:
        sw.cell(1, "No undated capital-gains items.")
        sw.row += 1


def write_workbook(estimate: "adv.AdvanceTaxEstimate", out_path) -> None:
    """Writes the Advance Tax Estimate .xlsx: New-regime and Old-regime
    Computation + Instalment Schedule sheets side by side, plus a Flags
    sheet. Instalment-schedule arithmetic (required/paid/shortfall) is
    live Excel formulas over this workbook's own cells -- see module
    docstring for the one documented deviation (tax liability itself is a
    Python-computed value, not a Rules-sheet-driven formula)."""
    wb = Workbook()
    wb.remove(wb.active)

    ws_new_comp = wb.create_sheet("Computation (New)")
    ws_old_comp = wb.create_sheet("Computation (Old)")
    ws_new_inst = wb.create_sheet("Instalments (New)")
    ws_old_inst = wb.create_sheet("Instalments (Old)")
    ws_flags = wb.create_sheet("Flags")

    refs_new = _write_computation_sheet(ws_new_comp, estimate, estimate.new_regime, "New")
    refs_old = _write_computation_sheet(ws_old_comp, estimate, estimate.old_regime, "Old")
    _write_instalment_sheet(ws_new_inst, estimate, estimate.new_regime, refs_new, "New")
    _write_instalment_sheet(ws_old_inst, estimate, estimate.old_regime, refs_old, "Old")
    _write_flags_sheet(ws_flags, estimate)

    wb.save(out_path)
