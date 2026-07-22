"""
presentation.py -- the deliverable/presentation layer over the calculation
engine (2026-07-19 "presentable workbook" prompt).

The workbook write_workbook.py produces is a *calculation engine*: 17 sheets,
every figure traceable, none of it printable. This module ADDS presentable
sheets in front of those 17, in this order -- `Statement of Income`, `BS`,
`IS`, `PL for Business` (present only for an entity with a configured
`business_subtree`, 2026-07-19 "PL for Business" prompt), `CG` -- and hides
the four raw working sheets. It changes no computation, no rule, no rate and
no tax logic, and it overwrites no existing sheet's values.

**Every money cell written here is an Excel formula pointing back at the
existing sheets** (`Computation`, `CapitalGains`, `OtherSources`,
`IS_Transcript`, `BS_Transcript`, `TaxesPaid`, `Entity`). Nothing is
recomputed and nothing is hardcoded -- that formula link is the audit trail
this tool has and a hand-built workbook does not. The only literals this
module writes are labels, headings and captions.

Two items are PARKED by Harshal's explicit direction (2026-07-19) and render
as a label plus an empty, visibly-styled value cell so the layout is final and
they can be filled later without a re-layout: Father's Name and Aadhaar No.
A third, residential status, is an *assumed* constant `R/OR` -- it renders a
value a reader would take as computed, so it carries a footnote marker and an
Assumptions note (see `_STATUS_FOOTNOTE`).

Brought-forward-loss set-off was PARKED through 2026-07-19; as of the
2026-07-20 on-page-totals change it is a REAL, editable input cell
(`_INPUT_FILL`, defaulting to 0) instead -- see `write_statement_of_income`.
That same change moves the income-ladder totals (Gross Total Income, Total
Income, and the normal-income/special-rate-CG split feeding `Computation`'s
slab tax) onto this sheet as on-page formulas, so an override anywhere in the
ladder -- a leaf, Chapter VI-A, or the b/f-loss cell -- now flows end-to-end
through tax, cess and refund. See
`docs/history/2026-07-20-itr-onpage-totals-plan.md`.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import interest_234
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

INR_FORMAT = "#,##,##0.00"
DATE_FORMAT = "DD-MM-YYYY"

FONT_NAME = "Arial"
INDENT = " " * 6          # CA file indents ~6 spaces per hierarchy level

#: Canonical order for the deliverable sheets. This is the ONLY place order is
#: expressed: `move_presentation_sheets_first` positions whatever subset of
#: these actually got created, so a sheet may be conditionally omitted (`CG`
#: and `PL for Business` already are) and a further sheet may later be
#: inserted at any position by adding its name here -- nothing downstream
#: assumes a fixed count or that any particular sheet is present.
PRESENTATION_SHEETS = ("Statement of Income", "BS", "IS", "PL for Business", "CG")

#: Raw working sheets hidden (never deleted) once the four above exist.
HIDDEN_SHEETS = ("Rules", "Mapping Review", "IS_Transcript", "BS_Transcript")

_STATUS_FOOTNOTE = "*"
_ASSUMPTION_NOTE = (
    "* Residential status is ASSUMED to be R/OR (Resident and Ordinarily Resident). "
    "It is not determined by this tool -- no day-count test or RNOR analysis is "
    "performed. Confirm before filing: R/OR vs RNOR vs NR changes what income is "
    "taxable at all.\n"
    "* Interest u/s 234A / 234B / 234C IS computed and is included in the "
    "Aggregate liability and Refund lines above. It depends on inputs the tool "
    "cannot know: the ACTUAL DATE OF FILING (assumed to be the due date unless "
    "you set it), the due date itself (31 July, i.e. the non-audit case), and "
    "the instalment-wise advance tax. All are editable in the Workings section "
    "-- check them, since 234A and 234B both grow for every part-month of delay."
)

#: Fail-loud (2026-07-19 CG gain-split-vs-action fix), "banner, no abort":
#: schedules.py's build_capital_gains already computes reconciliation_ok /
#: reconciliation_diff correctly -- this is only about making a mismatch
#: IMPOSSIBLE TO MISS on the deliverable sheets, never about refusing to
#: produce the workbook. `CG_RECONCILIATION_ERROR_MARKER` is the substring
#: agent.py greps its own run() summary for to decide the process's exit
#: code, so the wording here and the summary line it feeds must both
#: contain it verbatim.
CG_RECONCILIATION_ERROR_MARKER = "ERROR: Capital Gains do not reconcile to books"
_CG_ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_CG_ERROR_FONT_COLOR = "9C0006"


#: Sign-safe replacement for Excel's MROUND(number, multiple): MROUND raises
#: #NUM! whenever `number` and `multiple` have OPPOSITE signs. Every
#: statutory rounding checkpoint in this workbook (s.288A on Total Income,
#: s.288B on the refund/tax-payable line) rounds to a POSITIVE Rules-sheet
#: constant (`round_ti_nearest` / `round_tax_nearest`, both 10), but the
#: number being rounded goes negative for a loss year (s.288A) or a
#: tax-payable (as opposed to refund) assessee (s.288B) -- so the pre-existing
#: `=MROUND(x,m)` formula produced #NUM! on every such return (bug fix,
#: 2026-07-22). `ROUND(x/m,0)*m` reproduces MROUND's round-half-away-from-zero
#: behaviour for BOTH signs and never errors:
#:   MROUND(7.5,10)=10   == ROUND(0.75,0)*10=10
#:   MROUND(-5,-10)=-10  == ROUND(-0.5,0)*10=-10
#:   MROUND(0,10)=0      == ROUND(0,0)*10=0
#: `number_expr` is wrapped in its own parentheses before the division so a
#: compound expression (e.g. "'TaxesPaid'!B15-E39") isn't corrupted by
#: operator precedence; `multiple_ref` is a single cell reference and is
#: safe to repeat.
def mround_safe(number_expr: str, multiple_ref: str) -> str:
    return f"ROUND(({number_expr})/{multiple_ref},0)*{multiple_ref}"


#: Row offsets, from the start of the "Interest u/s 234" workings block, of
#: every cell the visible ladder references. The block has a FIXED shape
#: (unlike the income sections, none of it is conditional), so these offsets
#: are the whole layout contract -- `_i234_refs` turns them into coordinates
#: before the block is written and `write_234_workings` asserts the real
#: render lands on them.
_I234_ROW = {
    "title": 0,
    "due_date": 1, "filing_date": 2,
    "tax": 3, "tds": 4, "advance": 5, "assessed": 6,
    "a_months": 7, "a_interest": 8,
    "b_months": 9, "b_interest": 10,
    "c_header": 11, "c_q1": 12, "c_q2": 13, "c_q3": 14, "c_q4": 15,
    "c_total": 16, "total": 17,
}
I234_BLOCK_ROWS = 18

#: Columns within the block. The 234C instalment rows use three of them:
#: cumulative advance tax paid (editable), the first-proviso exclusion
#: (editable), and the resulting interest (formula).
_I234_COL_PAID = 3      # INNER
_I234_COL_VALUE = 4     # SUB -- single-value rows, and the 234C exclusion
_I234_COL_OUT = 5       # OUTER -- the interest amounts


def _i234_refs(start_row: int) -> dict:
    """Coordinates of the 234 workings cells, known before the block is
    written so the ladder above can reference them."""
    col_v = get_column_letter(_I234_COL_VALUE)
    col_o = get_column_letter(_I234_COL_OUT)
    col_p = get_column_letter(_I234_COL_PAID)
    refs = {}
    for key, off in _I234_ROW.items():
        r = start_row + off
        if key in ("a_interest", "b_interest", "c_total", "total",
                   "c_q1", "c_q2", "c_q3", "c_q4"):
            refs[key] = f"{col_o}{r}"
        else:
            refs[key] = f"{col_v}{r}"
        refs[key + "_row"] = r
    refs["_paid_col"] = col_p
    refs["_value_col"] = col_v
    return refs


def _excel_months_between(start_ref: str, end_ref: str) -> str:
    """Excel equivalent of interest_234.months_between: whole months with ANY
    part of a month counted as a full one. Deliberately not DAYS()/30 or
    DATEDIF's "m" alone -- both undercount the part-month the statute
    charges for."""
    return (
        f"IF({end_ref}<={start_ref},0,"
        f"(YEAR({end_ref})-YEAR({start_ref}))*12+(MONTH({end_ref})-MONTH({start_ref}))"
        f"+IF(DAY({end_ref})>DAY({start_ref}),1,0))"
    )


def _excel_round_down_100(expr: str) -> str:
    """Rule 119A. `MAX(0,...)` is not cosmetic: Excel's FLOOR raises #NUM!
    when number and significance have opposite signs -- the same defect class
    as the MROUND bug fixed above -- so the argument is clamped non-negative
    before FLOOR ever sees it. A negative base means nothing is owed anyway."""
    return f"FLOOR(MAX(0,{expr}),100)"


def write_234_workings(ws, start_row: int, model, total_liability_ref: str) -> int:
    """Render the live "Interest u/s 234" block inside Workings/Inputs.

    Everything here is a real Excel formula over editable input cells, not a
    frozen Python result: changing the filing date or an instalment figure
    must move the interest AND the Refund line, because a date cell that
    looks editable but silently does nothing is worse than no cell at all.
    The Python computation in interest_234.py seeds the defaults and is the
    control the tests check these formulas against.
    """
    refs = _i234_refs(start_row)
    tp = model.taxes_paid
    i234 = model.interest_234
    # The due date is 31 July of the ASSESSMENT year, so its year IS the AY --
    # which is where s.234B's clock starts (1 April of the AY).
    ay_start_year = i234.due_date.year

    ws.cell(row=refs["title_row"], column=2,
            value="Interest u/s 234A / 234B / 234C -- inputs & workings").font = _font(11, bold=True)

    def label(key: str, text: str) -> None:
        ws.cell(row=refs[key + "_row"], column=2, value=INDENT + text).font = _font()

    def date_input(key: str, value) -> None:
        c = ws.cell(row=refs[key + "_row"], column=_I234_COL_VALUE, value=value)
        c.number_format = "dd-mmm-yyyy"
        c.border = _INPUT_BORDER
        c.fill = _INPUT_FILL
        c.font = _font()

    # -- inputs -------------------------------------------------------------
    label("due_date", "Due date for furnishing the return")
    date_input("due_date", i234.due_date)
    label("filing_date", "Actual date of filing (edit me)")
    date_input("filing_date", i234.filing_date or i234.due_date)

    label("tax", "Tax on total income (incl. surcharge & cess)")
    ws.cell(row=refs["tax_row"], column=_I234_COL_VALUE,
            value=f"={total_liability_ref}").number_format = INR_FORMAT

    label("tds", "Less: TDS / TCS credit allowed")
    _input_cell(ws, refs["tds_row"], _I234_COL_VALUE,
                default_value=i234.result.tds_credit.amount if i234.computed else 0)

    label("advance", "Advance tax paid (full year)")
    _input_cell(ws, refs["advance_row"], _I234_COL_VALUE, default_value=tp.advance_tax)

    label("assessed", "Assessed tax (tax less TDS/TCS)")
    ws.cell(row=refs["assessed_row"], column=_I234_COL_VALUE,
            value=f"=MAX(0,{refs['tax']}-{refs['tds']})").number_format = INR_FORMAT

    # -- 234A ---------------------------------------------------------------
    label("a_months", "234A - months late (part month = full month)")
    ws.cell(row=refs["a_months_row"], column=_I234_COL_VALUE,
            value=f"={_excel_months_between(refs['due_date'], refs['filing_date'])}")

    label("a_interest", "Interest u/s 234A")
    base_a = _excel_round_down_100(f"{refs['tax']}-{refs['tds']}-{refs['advance']}")
    ws.cell(row=refs["a_interest_row"], column=_I234_COL_OUT,
            value=f"={base_a}*{interest_234.RATE_PER_MONTH}*{refs['a_months']}"
            ).number_format = INR_FORMAT

    # -- 234B ---------------------------------------------------------------
    # Runs from 1 April of the ASSESSMENT year, not the income year.
    apr1 = f"DATE({ay_start_year},4,1)"
    label("b_months", f"234B - months from 01-Apr-{ay_start_year}")
    ws.cell(row=refs["b_months_row"], column=_I234_COL_VALUE,
            value=f"={_excel_months_between(apr1, refs['filing_date'])}")

    label("b_interest", f"Interest u/s 234B (nil if advance tax >= "
                        f"{interest_234.S234B_THRESHOLD:.0%} of assessed tax)")
    base_b = _excel_round_down_100(f"{refs['assessed']}-{refs['advance']}")
    ws.cell(
        row=refs["b_interest_row"], column=_I234_COL_OUT,
        value=(f"=IF({refs['advance']}>={interest_234.S234B_THRESHOLD}*{refs['assessed']},0,"
               f"{base_b}*{interest_234.RATE_PER_MONTH}*{refs['b_months']})"),
    ).number_format = INR_FORMAT

    # -- 234C instalment table ---------------------------------------------
    # Column headings go in the LABEL column, not above the money columns:
    # every cell in columns C-E below the caption must be a formula or a
    # styled input cell (test_statement_money_columns_below_caption_are_all_
    # formulas), and a bare heading string there would break that invariant.
    hdr = refs["c_header_row"]
    ws.cell(
        row=hdr, column=2,
        value=INDENT + "234C instalments  [cumulative advance tax | proviso exclusion | interest]",
    ).font = _font(bold=True)

    cum = list(tp.advance_tax_cumulative) + [0.0] * 4
    for q, (lbl, _mo, _dy, req_pct, safe_pct, months) in enumerate(interest_234.S234C_INSTALMENTS):
        r = refs[f"c_q{q + 1}_row"]
        ws.cell(row=r, column=2,
                value=INDENT * 2 + f"{lbl} -- {req_pct:.0%} required").font = _font()
        _input_cell(ws, r, _I234_COL_PAID, default_value=cum[q])
        # The first-proviso exclusion for this instalment, as an editable
        # figure: it depends on WHEN the capital gain arose, which the tool
        # infers rather than knows.
        _input_cell(ws, r, _I234_COL_VALUE, default_value=_seed_exclusion(i234, q))
        paid = f"{refs['_paid_col']}{r}"
        exc = f"{refs['_value_col']}{r}"
        considered = f"MAX(0,{refs['assessed']}-{exc})"
        shortfall = _excel_round_down_100(f"{considered}*{req_pct}-{paid}")
        ws.cell(
            row=r, column=_I234_COL_OUT,
            value=(f"=IF({paid}>={considered}*{safe_pct},0,"
                   f"{shortfall}*{interest_234.RATE_PER_MONTH}*{months})"),
        ).number_format = INR_FORMAT

    label("c_total", "Interest u/s 234C")
    ws.cell(
        row=refs["c_total_row"], column=_I234_COL_OUT,
        value=f"=SUM({refs['c_q1']}:{refs['c_q4']})",
    ).number_format = INR_FORMAT

    label("total", "Total interest u/s 234A + 234B + 234C")
    c = ws.cell(
        row=refs["total_row"], column=_I234_COL_OUT,
        value=f"={refs['a_interest']}+{refs['b_interest']}+{refs['c_total']}",
    )
    c.number_format = INR_FORMAT
    c.font = _font(bold=True)

    return start_row + I234_BLOCK_ROWS


def _seed_exclusion(i234, q: int) -> float:
    """Default for the 234C first-proviso exclusion cell: the amount
    interest_234 actually excluded for that instalment, so the sheet opens
    agreeing with the computed figure and the filer edits from there."""
    if not i234.computed or q >= len(i234.result.s234c.instalments):
        return 0.0
    inst = i234.result.s234c.instalments[q]
    assessed = max(
        inst.tax_due_considered,
        i234.result.s234b.assessed_tax,
    )
    return max(0.0, assessed - inst.tax_due_considered)


def cg_mismatch_banner_text(cg_schedule) -> str:
    return (
        f"*** {CG_RECONCILIATION_ERROR_MARKER} "
        f"(diff {cg_schedule.reconciliation_diff:,.2f}) -- DO NOT FILE ***"
    )


def _write_cg_error_banner(ws, row: int, cg_schedule, ncols: int) -> int:
    """Top-of-sheet ERROR banner, present only when the CG lot-sum fails to
    reconcile to the books' control total (schedules.py build_capital_gains,
    diff magnitude > 0.01). Modeled on the CapitalGains sheet's existing
    'Unresolved FMV scrips (fail-loud, review)' row, but rendered as a
    prominent merged banner at the very top of a presentation sheet instead
    of a label/value cell buried in a working sheet -- this is what a CA or
    a bank actually opens. The workbook is ALWAYS still produced ("banner,
    no abort"); omitted entirely (returns `row` unchanged) when reconciled."""
    if cg_schedule.reconciliation_ok:
        return row
    c = ws.cell(row=row, column=1, value=cg_mismatch_banner_text(cg_schedule))
    c.font = Font(name=FONT_NAME, size=11, bold=True, color=_CG_ERROR_FONT_COLOR)
    c.fill = _CG_ERROR_FILL
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols, 1))
    return row + 2

#: Fail-loud (2026-07-22 salary-gross fix), mirrors CG_RECONCILIATION_ERROR_MARKER
#: above -- the control that would have caught the bug where the Salary
#: sheet's displayed gross (17(1)+17(2)+17(3)) silently dropped perquisites.
#: schedules.py's build_salary computes reconciliation_ok/reconciliation_diff
#: (gross - s.10 exempt - std deduction - prof tax vs income chargeable) on
#: the Form16 path; this is only about making a mismatch impossible to miss.
#: Same "banner, no abort" contract and agent.py exit-code convention as CG.
SALARY_RECONCILIATION_ERROR_MARKER = "ERROR: Salary sheet does not reconcile to Form 16"


def salary_mismatch_banner_text(salary_schedule) -> str:
    return (
        f"*** {SALARY_RECONCILIATION_ERROR_MARKER} "
        f"(gross - s.10 exempt - std deduction - prof tax vs income chargeable, "
        f"diff {salary_schedule.reconciliation_diff:,.2f}) -- DO NOT FILE ***"
    )


def _write_salary_error_banner(ws, row: int, salary_schedule, ncols: int) -> int:
    """Top-of-sheet ERROR banner, modeled directly on `_write_cg_error_banner`
    above -- present only when the Salary sheet's own arithmetic fails to
    reconcile to Form 16's income chargeable (schedules.py build_salary,
    diff magnitude > 0.01). 'Banner, no abort': the workbook is ALWAYS still
    produced; omitted entirely (returns `row` unchanged) when reconciled, and
    when there is nothing to check (book-only / manual path, where
    reconciliation_ok stays at its default True)."""
    if salary_schedule.reconciliation_ok:
        return row
    c = ws.cell(row=row, column=1, value=salary_mismatch_banner_text(salary_schedule))
    c.font = Font(name=FONT_NAME, size=11, bold=True, color=_CG_ERROR_FONT_COLOR)
    c.fill = _CG_ERROR_FILL
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols, 1))
    return row + 2


_PARKED_NOTE = "(to be filled)"

#: Human labels for rules.resolve_age_class()'s return values. 'general'
#: deliberately renders no suffix (prompt section 2b).
_AGE_LABELS = {"general": "", "senior": "Senior Citizen", "super_senior": "Super Senior Citizen"}

_THIN = Side(style="thin")
_TOP_RULE = Border(top=_THIN)
_BOTTOM_RULE = Border(bottom=_THIN)
#: The empty-but-visible cell used for every PARKED value (prompt 2a/2d).
_PARKED_BORDER = Border(bottom=Side(style="dotted"))
_PARKED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# A real, editable INPUT cell (2026-07-20 on-page-totals change) -- distinct
# from a PARKED cell: it always carries a literal value (never blank) and
# feeds live downstream formulas. Styled with a solid border (not dotted) and
# a different fill so a reader -- and a test -- can tell "type here, it
# means something" apart from "not filled in yet". See b/f-loss cell below.
_INPUT_BORDER = Border(bottom=Side(style="thin"))
_INPUT_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")


def _input_cell(ws, row: int, col: int, default_value=0):
    """A real, editable input cell: a literal value (default 0), visibly
    styled so it reads as "preparer-editable", never a formula. Unlike
    `_parked_cell` this is wired into downstream formulas -- overriding it
    must move every total below it."""
    c = ws.cell(row=row, column=col, value=default_value)
    c.number_format = INR_FORMAT
    c.border = _INPUT_BORDER
    c.fill = _INPUT_FILL
    c.font = _font()
    return c


def _font(size: int = 10, *, bold: bool = False, underline: str | None = None,
          italic: bool = False) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold, underline=underline, italic=italic)


def status_line(age_class: str, residency: str = "R/OR", *, declared: bool = False) -> str:
    """The header block's residential-status line. `residency` is the
    resolved value (R/OR / RNOR / NR) from `rules.resolve_residency()`; the
    age half is whatever `rules.resolve_age_class()` returned. 'general'
    renders no suffix. The footnote marker is dropped once `declared` is
    True -- a reader must be able to tell "someone asserted this" from "the
    tool fell back" (2026-07-19 residency prompt, section 2)."""
    label = _AGE_LABELS.get(age_class, "")
    base = f"{residency} - {label}" if label else residency
    return base if declared else base + " " + _STATUS_FOOTNOTE


# ---------------------------------------------------------------------------
# Hierarchy rebuild -- IS_Transcript / BS_Transcript `Path` column -> tree
# ---------------------------------------------------------------------------

@dataclass
class HNode:
    """One node of a GnuCash account tree rebuilt from transcript `Path`
    strings. Leaves carry `source` (the A1 ref of the transcript cell holding
    their amount); groups carry children and get their own subtotal row."""
    name: str
    source: str | None = None
    children: dict = field(default_factory=dict)   # name -> HNode, insertion-ordered

    @property
    def is_leaf(self) -> bool:
        return not self.children


def build_hierarchy(entries) -> HNode:
    """`entries` is [(path, source_cell), ...] where `path` is a GnuCash
    account path like 'Assets/Current Assets/Cash and Bank/BOB - <acct>'.
    Splits on '/' to rebuild the full tree -- depth is driven by the data, so
    a deeper or shallower book renders correctly with no hardcoded level
    count."""
    root = HNode("")
    for path, source in entries:
        parts = [p for p in str(path).split("/") if p != ""]
        if not parts:
            continue
        node = root
        for part in parts[:-1]:
            node = node.children.setdefault(part, HNode(part))
        leaf = node.children.setdefault(parts[-1], HNode(parts[-1]))
        leaf.source = source
    return root


def max_group_level(root: HNode) -> int:
    """Deepest *group* (non-leaf) level below `root`, 0-based. Drives how many
    tiered subtotal columns the sheet needs -- never a hardcoded count."""
    best = -1

    def walk(node: HNode, level: int) -> None:
        nonlocal best
        for child in node.children.values():
            if not child.is_leaf:
                best = max(best, level)
                walk(child, level + 1)

    walk(root, 0)
    return best


def render_hierarchy(ws, start_row: int, root: HNode, source_sheet: str,
                     label_col: int = 1, money_col: int = 2) -> tuple:
    """Render `root`'s children as an indented, tiered-subtotal block.

    Leaf amounts sit in `money_col`; a group's subtotal sits one column
    further out per level of nesting, so a leaf, its group total and a
    higher-level total occupy three different columns -- exactly how the CA
    reference builds `C14 = SUM(B10:B13)` and tiers above it.

    Every amount cell is a formula: leaves point at `source_sheet`, group
    subtotals SUM their own children's cells on this sheet.

    Returns (next_row, [(row, col) for each top-level group/leaf]).
    """
    depth = max_group_level(root)
    row = start_row

    def col_for_group(level: int) -> int:
        return money_col + (depth - level) + 1

    def emit(node: HNode, level: int) -> tuple:
        nonlocal row
        label = INDENT * level + node.name
        if node.is_leaf:
            ws.cell(row=row, column=label_col, value=label).font = _font()
            c = ws.cell(row=row, column=money_col, value=f"='{source_sheet}'!{node.source}")
            c.number_format = INR_FORMAT
            c.font = _font()
            here = (row, money_col)
            row += 1
            return here

        head = ws.cell(row=row, column=label_col, value=label)
        head.font = _font(bold=True)
        row += 1

        child_cells = [emit(child, level + 1) for child in node.children.values()]

        total_col = col_for_group(level)
        tc = ws.cell(row=row, column=label_col, value=INDENT * level + f"Total {node.name}")
        tc.font = _font(bold=True)
        if child_cells and all(c[1] == money_col for c in child_cells):
            first, last = child_cells[0][0], child_cells[-1][0]
            letter = get_column_letter(money_col)
            formula = f"=SUM({letter}{first}:{letter}{last})"
        else:
            formula = "=" + "+".join(f"{get_column_letter(c)}{r}" for r, c in child_cells) if child_cells else "=0"
        t = ws.cell(row=row, column=total_col, value=formula)
        t.number_format = INR_FORMAT
        t.font = _font(bold=True)
        t.border = _TOP_RULE
        here = (row, total_col)
        row += 1
        return here

    tops = [emit(child, 0) for child in root.children.values()]
    return row, tops


# ---------------------------------------------------------------------------
# Shared formatting / print setup
# ---------------------------------------------------------------------------

def apply_sheet_chrome(ws, widths: dict, last_row: int, last_col: int, *,
                       freeze: str | None = None, landscape: bool = False,
                       print_title: str = "") -> None:
    """Column widths, gridlines off, freeze panes, print area and A4
    fit-to-one-page-wide print setup. No sheet ships with default widths."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False
    if freeze:
        ws.freeze_panes = freeze

    ws.print_area = f"A1:{get_column_letter(last_col)}{max(last_row, 1)}"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = ws.page_margins.right = 0.5
    ws.page_margins.top = ws.page_margins.bottom = 0.6
    if print_title:
        ws.oddHeader.center.text = print_title
        ws.oddFooter.right.text = "Page &P of &N"


def fit_label_width(ws, col: int, minimum: float, cap: float = 90.0) -> float:
    """Label-column width: wide enough for the longest label actually present,
    never narrower than the CA reference's own width, never absurdly wide.

    The CA file's fixed A=38.1/47.8 were tuned to that one entity's account
    names; a book with deeper nesting or longer names would truncate at those
    widths, and a truncated label is the single most visible defect this whole
    exercise exists to fix. Formulas are skipped -- their rendered text is not
    the formula string.
    """
    longest = 0
    for row in ws.iter_rows(min_col=col, max_col=col):
        for c in row:
            if isinstance(c.value, str) and not c.value.startswith("="):
                longest = max(longest, len(c.value))
    return max(minimum, min(cap, longest + 2))


def _aadhaar_formula(ref: str) -> str:
    """CA-file space-grouped Aadhaar (`NNNN NNNN NNNN`), built as a formula
    over the raw digits stored on the Entity sheet -- never a second literal
    copy of the number (2026-07-19 residency prompt, section 1)."""
    return f'=LEFT({ref},4)&" "&MID({ref},5,4)&" "&MID({ref},9,4)'


def _parked_cell(ws, row: int, col: int):
    """A PARKED value: empty, but visibly styled -- dotted rule plus a light
    fill -- so the reader can see the field exists and is unfilled (prompt
    2a/2d). Never carries a value, and the row is never dropped. The
    "(to be filled)" wording goes in the LABEL, so no stray text ever lands in
    a money column."""
    c = ws.cell(row=row, column=col, value=None)
    c.border = _PARKED_BORDER
    c.fill = _PARKED_FILL
    c.font = _font(italic=True)
    return c


# ---------------------------------------------------------------------------
# Tax-computation formula builders (2026-07-21 on-page-totals change, design
# doc section 11.1). Moved here from write_workbook.py -- these are pure
# Excel-formula-string builders over cell-reference strings, no hardcoded
# rate or threshold, and this module needs to call them directly to write the
# on-page tax block. write_workbook.py `import presentation`s this module (to
# call `write_statement_of_income`), so the dependency cannot run the other
# way: write_workbook.py's own `write_computation_tail` now calls
# `presentation._slab_tax_formula` / `presentation._surcharge_formula`
# instead of keeping a second copy.
# ---------------------------------------------------------------------------

def _slab_tax_formula(income_cell: str, slab_refs: dict) -> str:
    """SUM of MAX(0, MIN(income, upto_i) - upto_(i-1)) * rate_i across the
    padded slab rows in `slab_refs` (Rules-sheet cell refs) -- a generic
    Excel slab-tax formula that never hardcodes a rate or threshold."""
    terms = []
    prev_upto_ref = None
    for i in sorted(slab_refs):
        upto_ref, rate_ref = slab_refs[i]
        lower = f"'Rules'!{prev_upto_ref}" if prev_upto_ref else "0"
        terms.append(f"(MAX(0,MIN({income_cell},'Rules'!{upto_ref})-{lower}))*'Rules'!{rate_ref}")
        prev_upto_ref = upto_ref
    return "+".join(terms)


def _surcharge_formula(income_cell: str, tax_cell: str, surcharge_refs: dict) -> str:
    """Nested-IF surcharge-rate lookup (highest matching band) * tax."""
    expr = "0"
    for i in sorted(surcharge_refs):
        above_ref, rate_ref = surcharge_refs[i]
        expr = f"IF({income_cell}>'Rules'!{above_ref},'Rules'!{rate_ref},{expr})"
    return f"({expr})*{tax_cell}"


def _rebate_base_formula(normal_income_cell: str, tax_normal_cell: str, max_ti_ref: str,
                         max_amt_ref: str) -> str:
    """s.87A base rebate ONLY (no marginal relief) -- the design doc's ladder
    (section 11.1) wants this as its own line item, split out of what used to
    be one combined "87A rebate" cell (write_workbook.py's `_rebate_formula`,
    still used unchanged by `Computation`'s own audit copy). Mutually
    exclusive with `_marginal_relief_formula` by construction (one of the two
    is always 0), so `tax - rebate_base - marginal_relief` reproduces the old
    combined formula's result exactly -- see the paired function below."""
    return f"IF({normal_income_cell}<='Rules'!{max_ti_ref},MIN('Rules'!{max_amt_ref},{tax_normal_cell}),0)"


def _marginal_relief_formula(normal_income_cell: str, tax_normal_cell: str, max_ti_ref: str,
                             marginal_ref: str) -> str:
    """s.87A marginal relief ONLY -- the other half of the split described in
    `_rebate_base_formula`'s docstring. Zero whenever income is within the
    rebate-eligible band (that's `_rebate_base_formula`'s job) or the regime
    has no marginal-relief flag set."""
    return (
        f"IF(AND({normal_income_cell}>'Rules'!{max_ti_ref},'Rules'!{marginal_ref}),"
        f"MAX(0,{tax_normal_cell}-({normal_income_cell}-'Rules'!{max_ti_ref})),0)"
    )


#: Helper columns for the on-page tax block's both-regimes working (design
#: doc section 11.1) -- past OUTER (column E), so `apply_sheet_chrome`'s print
#: area (always "A1:<last_col><last_row>" with last_col=OUTER) never includes
#: them. This is genuinely-live, page-local arithmetic -- not a mirror of
#: `Computation`, which keeps its own separately-anchored copy for audit only.
_TAX_HELPER_COL = 8   # column H


def _write_regime_tax_workings(ws, start_row: int, col: int, prefix: str, normal_ref: str,
                               special_cg_ref: str, cg_tax_ref: str, cess_rate_ref: str,
                               slab_refs: dict, rebate_max_ti: str, rebate_max_amt: str,
                               rebate_marginal: str, surcharge_refs: dict) -> tuple:
    """Writes one regime's FULL on-page tax working -- slab tax -> s.87A
    rebate -> marginal relief -> tax after rebate -> + special-rate CG tax ->
    surcharge -> cess -> total tax liability -- each its own live-formula
    cell, chained cell-to-cell (never re-embedded as one giant nested
    string, which is how `Computation`'s own regime block stays readable).
    The visible ladder in `write_statement_of_income` just IF()-picks between
    this regime's final cells and the other regime's at every line.

    Numeric-parity note (design doc section 11.4, "default parity to the
    paisa"): schedules.py's compute_tax() applies surcharge and cess to
    (tax-after-rebate + special-rate CG tax) COMBINED, not to normal-rate tax
    alone -- so special-rate CG tax is added in BEFORE surcharge/cess here,
    matching schedules.py exactly, even though the design doc's line order
    lists "add special-rate CG tax" after "Total tax liability". Both regimes
    are always computed regardless of which is selected -- see
    `write_statement_of_income`'s `sel()`.

    Returns (parts, next_row) where `parts` has keys slab/rebate/marginal/
    cg_tax/surcharge/cess/liability -> this regime's cell coordinates.
    """
    r = start_row

    # NOTE: these working-cell labels deliberately avoid reusing any phrase
    # from the VISIBLE ladder below (e.g. never literally "Tax on total
    # income" or "Brought forward losses set off") -- tests locate visible
    # rows by a case-insensitive label SUBSTRING search over the *whole*
    # sheet, and these helper cells sit earlier in row order (row 2 onward)
    # than the visible ladder, so a shared phrase would shadow the real row.
    def put(label: str, formula: str) -> str:
        nonlocal r
        ws.cell(row=r, column=col, value=f"{prefix} regime working -- {label}").font = _font(8, italic=True)
        c = ws.cell(row=r, column=col + 1, value=formula)
        c.number_format = INR_FORMAT
        c.font = _font(8)
        r += 1
        return c.coordinate

    slab_ref = put("gross tax on income (slab)", "=" + _slab_tax_formula(normal_ref, slab_refs))
    rebate_ref = put("rebate u/s 87A",
                     "=" + _rebate_base_formula(normal_ref, slab_ref, rebate_max_ti, rebate_max_amt))
    marginal_ref = put("marginal relief u/s 87A",
                       "=" + _marginal_relief_formula(normal_ref, slab_ref, rebate_max_ti, rebate_marginal))
    after_rebate_ref = put("tax after rebate and marginal relief",
                           f"=MAX(0,{slab_ref}-{rebate_ref}-{marginal_ref})")
    combined_ref = put("plus special-rate CG tax", f"={after_rebate_ref}+{cg_tax_ref}")
    surcharge_income_expr = f"({normal_ref}+{special_cg_ref})"
    surcharge_ref = put("surcharge on tax",
                        "=" + _surcharge_formula(surcharge_income_expr, combined_ref, surcharge_refs))
    cess_ref = put("health and education cess", f"=({combined_ref}+{surcharge_ref})*{cess_rate_ref}")
    liability_ref = put("tax liability, combined", f"={combined_ref}+{surcharge_ref}+{cess_ref}")
    return {
        "slab": slab_ref, "rebate": rebate_ref, "marginal": marginal_ref,
        "cg_tax": cg_tax_ref, "surcharge": surcharge_ref, "cess": cess_ref,
        "liability": liability_ref,
    }, r


# ---------------------------------------------------------------------------
# Statutory per-bucket brought-forward-loss set-off (2026-07-21 on-page-
# totals change, design doc section 11.2). Set BY THE INCOME TAX ACT, not a
# design choice: each bucket sets off ONLY against its own income head/
# gain-type, capped at that head's available (non-negative) income for the
# year -- it can never drive a head negative and it can never leak into
# another head. Set-off happens here, at the head/gain-type level, BEFORE
# aggregation into Gross Total Income (the pre-2026-07-21 behaviour -- one
# lump cell subtracted from normal income after GTI -- was not what the
# statute requires; see write_statement_of_income's docstring).
# ---------------------------------------------------------------------------

#: Simple single-head buckets: each just caps its own bf-loss input against
#: MAX(head, 0) via `_capped_setoff_expr`. Data-driven so a future bucket
#: (s.73 speculation-business loss, s.73A specified-business loss, s.74A
#: race-horse-owning loss) is a one-line addition here once the tool models
#: that income source -- intentionally NOT built now (design doc section
#: 11.2 lists them but only the first four buckets ship in this change).
#: Each tuple is (bucket key, on-page label, comp_layout leaf key it nets).
_SIMPLE_BF_BUCKETS = (
    ("bf_hp", "b/f House Property loss (s.71B) -- HP income only", "hp"),
    ("bf_business", "b/f Business loss (s.72) -- Business income only", "business"),
)


def _capped_setoff_expr(head_ref: str, bf_ref: str) -> str:
    """Reduces `head_ref` by up to `bf_ref`, capped at MAX(head_ref,0) -- the
    statutory cap (design doc section 11.2): the bucket cannot drive the head
    negative, and it cannot apply at all when the head is already <=0 (that's
    the head's OWN current-year loss, a separate matter from a brought-
    forward loss set off against it). Returns a bare expression (no leading
    "=") so callers can embed it directly in `item()`'s formula string."""
    return f"({head_ref})-MIN({bf_ref},MAX({head_ref},0))"


def _cg_setoff_exprs(stcg_ref: str, ltcg_ref: str, bf_stcl_ref: str, bf_ltcl_ref: str) -> tuple:
    """s.74 capital-loss set-off: b/f Short-term capital loss sets off
    against STCG first, any remainder spills over against LTCG; b/f
    Long-term capital loss sets off against LTCG ONLY (never STCG, never
    normal income), applied AFTER the STCL spillover. Both capped so a
    set-off can never drive an already-positive head negative and can never
    leak outside the CG head.

    Deliberately NOT floored via an unconditional MAX(0,...) on the raw
    head value: when a head is already a current-year LOSS (e.g. LTCG < 0)
    and there is nothing to set off against it (bf buckets at their default
    0, or a positive spillover with no positive LTCG available to absorb
    it), the raw signed value must pass through UNCHANGED so the default
    (no-override) case stays paisa-exact with schedules.py's own ground
    truth (which flows the raw signed CG figures into GTI/Total Income and
    only nets LTCG+STCG together for the special-rate base -- see
    test_default_case_ladder_matches_schedules_engine_to_the_paisa). The
    cap only ever *reduces towards* zero an amount that was actually
    available (positive) before the set-off touched it; it never invents a
    floor on an untouched negative starting point. Returns
    (net_stcg_expr, net_ltcg_expr), bare expressions (no leading "=")."""
    stcg_avail = f"MAX({stcg_ref},0)"
    stcl_used = f"MIN({bf_stcl_ref},{stcg_avail})"
    net_stcg = f"({stcg_ref})-{stcl_used}"
    stcl_spill = f"MAX(0,{bf_stcl_ref}-{stcg_avail})"
    ltcg_avail = f"MAX({ltcg_ref},0)"
    spill_used = f"MIN({stcl_spill},{ltcg_avail})"
    remaining_after_spill = f"({ltcg_avail})-{spill_used}"
    ltcl_used = f"MIN({bf_ltcl_ref},{remaining_after_spill})"
    net_ltcg = f"({ltcg_ref})-{spill_used}-{ltcl_used}"
    return net_stcg, net_ltcg


# ---------------------------------------------------------------------------
# Sheet 1 -- Statement of Income
# ---------------------------------------------------------------------------

def write_statement_of_income(wb, model, entity_layout: dict, comp_layout: dict,
                              os_layout: dict, tp_layout: dict, ded_layout: dict,
                              rules_layout: dict, cg_layout: dict,
                              age_class: str, period_label: str, print_title: str,
                              computation_tail_fn, *,
                              father_name: str | None = None, aadhaar: str | None = None,
                              residency_value: str = "R/OR", residency_declared: bool = False):
    """Mirrors the CA reference's `ITWorking`: a letterhead header block, then
    three money columns (line items / sub-totals / running total) showing ONLY
    the selected regime.

    2026-07-20/21 on-page-totals change: the income ladder (GTI -> Chapter
    VI-A -> Total Income -> normal/special-CG split) AND the standard tax
    computation (slab tax -> s.87A rebate -> marginal relief -> + special-
    rate CG tax -> surcharge -> cess -> Total tax liability -> refund) are
    now computed ON THIS PAGE from on-page cells, rather than mirroring a
    hidden `Computation` sheet. `comp_layout` here is the LEAF layout only
    (salary, hp, business, os, cg_lt, cg_st -- see
    `write_computation_leaf_cells`). `cg_layout` supplies
    `special_tax_cell` -- the 112A/111A tax on special-rate CG, read
    directly from `CapitalGains`, regime-independent.

    Brought-forward losses (design doc section 11.2) are FOUR statutory
    per-bucket input cells -- HP (s.71B), Business (s.72), STCL and LTCL
    (both s.74) -- each capped and routed to its own head/gain-type BEFORE
    aggregation into Gross Total Income; see `_SIMPLE_BF_BUCKETS`,
    `_capped_setoff_expr` and `_cg_setoff_exprs`. This replaces the single
    lump "reduces normal income first" input cell from 2026-07-20.

    `computation_tail_fn(page_layout)` is still called, once this page has
    written its own `normal_income_base` coordinate, so `write_computation_tail`
    keeps writing its re-anchored slab-tax machinery onto `Computation` as a
    parallel, hidden backing/audit sheet (design doc section 11.1) -- but this
    page's own tax/cess/liability/refund lines no longer read that return
    value; they are built independently here via `_write_regime_tax_workings`
    so the page is genuinely live, not a mirror. See
    `docs/history/2026-07-20-itr-onpage-totals-plan.md` section 11.
    """
    ws = wb.create_sheet("Statement of Income")
    ent = lambda key: f"='Entity'!{entity_layout[key].coordinate}"  # noqa: E731
    # comp_layout values are already sheet-qualified expressions
    # (e.g. "'Computation'!B12" or "'Computation'!B45+'Computation'!B46").
    comp = lambda key: f"={comp_layout[key]}"                        # noqa: E731

    regime_ref = f"'Entity'!{entity_layout['regime'].coordinate}"

    def sel(new_expr: str, old_expr: str) -> str:
        """Selected-regime picker -- the both-regimes comparison stays on the
        Computation working sheet; it is not deliverable content."""
        return f"=IF({regime_ref}=\"old\",{old_expr},{new_expr})"

    LBL, INNER, SUB, OUTER = 2, 3, 4, 5

    # -- header block -------------------------------------------------------
    row = 1
    row = _write_cg_error_banner(ws, row, model.capital_gains, OUTER)
    row = _write_salary_error_banner(ws, row, model.salary, OUTER)
    title = ws.cell(row=row, column=1, value="STATEMENT OF INCOME")
    title.font = _font(14, bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=OUTER)
    title.alignment = Alignment(horizontal="center")
    row += 2

    def left(label: str, formula: str | None):
        ws.cell(row=row, column=1, value=label).font = _font(bold=True)
        if formula is None:
            _parked_cell(ws, row, 2)
        else:
            ws.cell(row=row, column=2, value=formula).font = _font()

    def right(label: str, value, *, parked: bool = False):
        ws.cell(row=row, column=4, value=label).font = _font(bold=True)
        if parked:
            _parked_cell(ws, row, 5)
        else:
            ws.cell(row=row, column=5, value=value).font = _font()

    left("Name", ent("name"))
    right("Previous Year", period_label)
    row += 1
    # Father's Name / Aadhaar are OPTIONAL entity fields (2026-07-19 residency
    # prompt, section 1 -- previously PARKED, Harshal 2026-07-19). Per field:
    # present -> a real formula + label with "(to be filled)" dropped; absent
    # -> label keeps "(to be filled)" and the cell stays empty-but-styled.
    if father_name:
        left("Father's Name", ent("father_name"))
    else:
        left(f"Father's Name {_PARKED_NOTE}", None)
    right("PAN", ent("pan"))
    row += 1
    left("Address", ent("address"))
    if aadhaar:
        right("Aadhaar No.", _aadhaar_formula(f"'Entity'!{entity_layout['aadhaar'].coordinate}"))
    else:
        right(f"Aadhaar No. {_PARKED_NOTE}", None, parked=True)
    row += 1
    left("", "");                             right("Date of Birth", ent("dob"));          row += 1
    ws.cell(row=row, column=4, value="Status").font = _font(bold=True)
    ws.cell(row=row, column=5, value=ent("status")).font = _font()
    row += 1
    ws.cell(row=row, column=4, value="Residential Status").font = _font(bold=True)
    ws.cell(row=row, column=5,
            value=status_line(age_class, residency_value, declared=residency_declared)).font = _font()
    row += 1
    ws.cell(row=row, column=4, value="Regime").font = _font(bold=True)
    ws.cell(row=row, column=5,
            value=f"=IF({regime_ref}=\"old\",\"Old Regime\",\"Tax u/s 115BAC\")").font = _font()
    row += 2

    header_end = row
    for col in (INNER, SUB, OUTER):
        c = ws.cell(row=row, column=col, value="Rs.")
        c.font = _font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = _BOTTOM_RULE
    row += 1

    # -- body ---------------------------------------------------------------
    def section(name: str):
        c = ws.cell(row=row, column=1, value="•")
        c.font = _font(bold=True)
        h = ws.cell(row=row, column=LBL, value=name)
        h.font = _font(bold=True, underline="single")

    def money(col: int, formula: str, *, bold: bool = False, rule: Border | None = None):
        c = ws.cell(row=row, column=col, value=formula)
        c.number_format = INR_FORMAT
        c.font = _font(bold=bold)
        if rule is not None:
            c.border = rule
        return c

    def item(label: str, formula: str):
        ws.cell(row=row, column=LBL, value=INDENT + label).font = _font()
        money(INNER, formula)

    def line(label: str, col: int, formula: str, *, bold: bool = False,
             rule: Border | None = None):
        ws.cell(row=row, column=LBL, value=label).font = _font(bold=bold)
        money(col, formula, bold=bold, rule=rule)

    subtotal_cells: list[str] = []

    def close_section(first_item_row: int):
        """Section items live in the inner column; the section total lands one
        column out."""
        nonlocal row
        letter = get_column_letter(INNER)
        c = ws.cell(row=row, column=SUB, value=f"=SUM({letter}{first_item_row}:{letter}{row - 1})")
        c.number_format = INR_FORMAT
        c.font = _font(bold=True)
        c.border = _TOP_RULE
        subtotal_cells.append(f"{get_column_letter(SUB)}{row}")
        row += 1

    # -- Workings/Inputs placement (2026-07-22 layout fix) ------------------
    # The "Brought forward losses set off" input block and the New/Old regime
    # tax-working formulas used to sit HERE, at the very top of the page,
    # before any income was even shown -- confusing for a reader (raw
    # statutory set-off inputs and helper tax machinery ahead of the actual
    # income statement). Both are now relocated to a labelled "Workings /
    # Inputs" section BELOW the "Refund Due / (Tax Payable)" line. The catch:
    # their cell coordinates are consumed by formulas written EARLIER in this
    # function than that -- the b/f buckets by the HP/Business/CG
    # net-of-set-off expressions immediately below, the tax workings by the
    # visible tax ladder further down -- earlier, that is, than the point
    # where their real target row would become known by simply running this
    # function top-to-bottom.
    #
    # Rather than a full two-pass render, the row arithmetic for everything
    # BETWEEN the body start and the Refund line is fully deterministic from
    # `model`/`tp` alone (each section either renders a fixed row count or is
    # skipped entirely) -- so that row count is predicted analytically
    # (`_predict_pre_workings_rows`, right below) and the Workings/Inputs
    # section's start row is computed UP FRONT, before any body content is
    # written. The b/f input cells' coordinates and the tax-workings' start
    # row are both derived from that prediction and used immediately by the
    # formulas that need them; the actual `section()`/`_input_cell()` calls
    # that render the b/f block are deferred to later in this function, once
    # the real `row` cursor reaches that point. A runtime assertion right
    # after the Refund line confirms the real cursor lands exactly where
    # predicted -- if the two ever drift (e.g. a future edit adds a row to
    # some section above and forgets to update the prediction), workbook
    # generation raises immediately instead of silently shipping a
    # `#REF!`-producing offset.
    os_ = model.other_sources
    os_items = (
        ("Savings bank interest", "sb", os_.interest_sb),
        ("Bank FD interest", "bank", os_.interest_bank),
        ("NBFC/HFC interest", "nbfc", os_.interest_nbfc),
        ("EPF taxable interest", "epf", os_.interest_epf_taxable),
        ("Interest on Income Tax refund", "refund_interest", os_.refund_interest),
        ("Dividend income (gross)", "dividend", os_.dividend_gross),
        ("SLBS income", "slbs", os_.slbs),
    )
    tp = model.taxes_paid
    prepaid = (
        ("TDS on salary", "tds_salary", tp.tds_salary),
        ("TDS on interest", "tds_interest", tp.tds_interest),
        ("TDS on dividend", "tds_dividend", tp.tds_dividend),
        ("TCS", "tcs", tp.tcs),
        ("Advance Tax", "advance", tp.advance_tax),
        ("Self-assessment Tax", "sat", tp.self_assessment_tax),
    )

    def _predict_pre_workings_rows() -> int:
        """Deterministic row count from the body start row (right after the
        'Rs.' header rule) through the Refund line, both inclusive --
        EXCLUDING the b/f-loss block and regime tax workings, both relocated
        to the Workings/Inputs section below Refund. Every term here mirrors
        an actual write further down this function -- see the runtime
        assertion after the Refund line, which is what actually guarantees
        this stays correct."""
        n = 0
        cg = model.capital_gains
        if model.salary.income_chargeable:
            n += 3   # section + 1 item + close_section total
        if model.house_property.income:
            n += 3
        if model.business.remuneration or model.business.expenses_total:
            n += 3
        if cg.lt_taxable_gross or cg.st_taxable_gross:
            n += 2 + (1 if cg.lt_taxable_gross else 0) + (1 if cg.st_taxable_gross else 0)
        if any(v for _, _, v in os_items):
            n += 2 + sum(1 for _, _, v in os_items if v)
        n += 1                                  # blank before GTI
        # GTI, VIA, TI, special-CG, normal-income -- 5 lines, each its own row,
        # PLUS the one extra blank row the normal-income line's `row += 2`
        # (instead of `+= 1`) leaves behind.
        n += 6
        # Tax ladder -- 7 lines (slab/rebate/marginal/cg/surcharge/cess/
        # liability), each its own row, then the four interest lines (234A/
        # 234B/234C/aggregate liability), PLUS the one extra blank row the
        # aggregate line's `row += 2` (instead of `+= 1`) leaves behind.
        n += 12
        n += 2 + max(sum(1 for _, _, v in prepaid if v), 1)  # prepaid section (incl. close_section total)
        n += 1                                  # "Total prepaid taxes" line
        n += 1                                  # Refund Due / (Tax Payable)
        return n

    body_start_row = row
    predicted_post_refund_row = body_start_row + _predict_pre_workings_rows()
    workings_start_row = predicted_post_refund_row + 2  # mirrors the old 2-blank-row gap

    # Brought-forward-loss bucket coordinates -- predicted now (see above),
    # written for real once the cursor reaches the Workings/Inputs section
    # further down. Layout inside that section: "Workings / Inputs" header
    # (+1 row), one blank row, "Brought forward losses..." section label
    # (+1 row), then the four input rows.
    bf_first_row = workings_start_row + 3
    bf_refs: dict[str, str] = {}
    _bf_row = bf_first_row
    for key, _label, _head in _SIMPLE_BF_BUCKETS:
        bf_refs[key] = f"{get_column_letter(SUB)}{_bf_row}"
        _bf_row += 1
    for key in ("bf_stcl", "bf_ltcl"):
        bf_refs[key] = f"{get_column_letter(SUB)}{_bf_row}"
        _bf_row += 1
    bf_last_row = _bf_row - 1

    # Interest u/s 234 workings -- same deferred-write pattern as the b/f
    # buckets: the visible ladder references these cells, but the block is
    # RENDERED further down, inside Workings/Inputs. Its shape is fixed, so
    # the coordinates are knowable now; `write_234_workings` asserts the real
    # render lands on them.
    i234_start_row = bf_last_row + 2
    i234_refs = _i234_refs(i234_start_row)

    # Net-of-set-off expressions -- computed ONCE, unconditionally (comp_layout's
    # leaf cells are always present even when a head's section below is not
    # rendered), so "(of which) Special-rate Capital Gains" further down can
    # reuse the same net CG figures the CG section itself shows. When a bucket
    # is left at its default 0, every expression below reduces to the original
    # (un-netted) leaf value -- preserving exact parity with the pre-2026-07-21
    # default case.
    net_hp_expr = _capped_setoff_expr(comp_layout["hp"], bf_refs["bf_hp"])
    net_business_expr = _capped_setoff_expr(comp_layout["business"], bf_refs["bf_business"])
    net_stcg_expr, net_ltcg_expr = _cg_setoff_exprs(
        comp_layout["cg_st"], comp_layout["cg_lt"], bf_refs["bf_stcl"], bf_refs["bf_ltcl"],
    )

    #: Short pointer left on every income-section head whose figure is netted
    #: against a b/f-loss bucket above, since the input cells themselves no
    #: longer sit right next to these sections (2026-07-22 layout fix).
    _BF_NOTE = "  (b/f set-off applied -- see Workings below)"

    # Heads of income -- rendered only when the underlying figure is non-zero.
    # HP / Business / CG items use the NET-of-set-off expressions above so the
    # statutory routing happens at the head/gain-type level, before
    # aggregation into Gross Total Income (design doc section 11.2) -- not as
    # one lump deduction against Total Income (the pre-2026-07-21 behaviour).
    if model.salary.income_chargeable:
        section("Income from Salary"); row += 1
        first = row
        item("Income chargeable under Salaries", comp("salary")); row += 1
        close_section(first)

    if model.house_property.income:
        section("Income from House Property" + _BF_NOTE); row += 1
        first = row
        item("Income from house property", f"={net_hp_expr}"); row += 1
        close_section(first)

    if model.business.remuneration or model.business.expenses_total:
        section("Income from Business or Profession" + _BF_NOTE); row += 1
        first = row
        item("Net business income", f"={net_business_expr}"); row += 1
        close_section(first)

    cg = model.capital_gains
    if cg.lt_taxable_gross or cg.st_taxable_gross:
        section("Capital Gains" + _BF_NOTE); row += 1
        first = row
        if cg.lt_taxable_gross:
            item("Long Term Capital Gain / (Loss)", f"={net_ltcg_expr}"); row += 1
        if cg.st_taxable_gross:
            item("Short Term Capital Gain / (Loss)", f"={net_stcg_expr}"); row += 1
        close_section(first)

    if any(v for _, _, v in os_items):
        section("Income from other sources"); row += 1
        first = row
        for label, key, value in os_items:
            if value:
                item(label, f"='OtherSources'!{os_layout[key].coordinate}")
                row += 1
        close_section(first)

    row += 1
    # Gross Total Income -- ON-PAGE, summed from this page's own section
    # subtotal cells (2026-07-20 on-page-totals change). Previously this
    # mirrored a hidden Computation-sheet total; now it IS the total, so an
    # override to any leaf item flows through automatically.
    # Guard against the (theoretical) case of every head-of-income section
    # being zero and therefore unrendered -- "=SUM()" is not a valid Excel
    # formula, so fall back to a literal 0 range rather than an empty SUM().
    gti_formula = f"=SUM({','.join(subtotal_cells)})" if subtotal_cells else "=0"
    line("Total", OUTER, gti_formula, bold=True, rule=_TOP_RULE)
    gti_row = row
    gti_ref = f"{get_column_letter(OUTER)}{gti_row}"
    row += 1

    # Chapter VI-A deductions -- a leaf reference to the Deductions sheet
    # (individual items still come from the source schedule; only the total
    # is aggregated here).
    line("Less - Chapter VI-A deductions", SUB, f"='Deductions'!{ded_layout['total']}")
    via_row = row
    via_ref = f"{get_column_letter(SUB)}{via_row}"
    row += 1

    # Brought-forward losses (design doc section 11.2) are now netted at the
    # HEAD level above, before this SUM -- so Gross Total Income is already
    # net of every bucket's statutory set-off. There is no separate lump
    # subtraction at the Total Income level any more (that was the
    # pre-2026-07-21 behaviour); see the b/f-loss block above.

    # s.288A rounding (nearest, from Rules) is applied here, exactly as the
    # pre-change Computation-sheet formula did (MROUND(ti_raw, round_ti_nearest))
    # -- dropping it would silently break paisa-exact parity with today's
    # workbook for the default (no-override) case. Uses `mround_safe` (not
    # Excel's MROUND) because Total Income can go negative in a loss year,
    # which would otherwise raise #NUM! (bug fix, 2026-07-22).
    round_ti_ref = f"'Rules'!{rules_layout['round_ti_nearest'].coordinate}"
    line("Total Income", OUTER, f"={mround_safe(f'{gti_ref}-{via_ref}', round_ti_ref)}",
         bold=True, rule=_TOP_RULE)
    ti_row = row
    ti_ref = f"{get_column_letter(OUTER)}{ti_row}"
    row += 1

    # Special-rate CG base and normal-income base -- the correctness-trap
    # carve-out (design doc section 5): Total Income (above) INCLUDES
    # special-rate CG (112A/111A LTCG/STCG, net of the CG b/f-loss buckets);
    # the slab-tax BASE that `Computation` reads must exclude it. Reuses the
    # SAME net_ltcg_expr/net_stcg_expr the CG section itself renders (when
    # rendered) so the b/f CG buckets interact correctly with the carve-out
    # -- always safe to evaluate even when the CG section is not rendered
    # on-page (both leaves, and hence both net expressions, are 0 in that
    # case).
    line("  (of which) Special-rate Capital Gains", SUB,
         f"=({net_ltcg_expr})+({net_stcg_expr})")
    special_cg_row = row
    special_cg_ref = f"{get_column_letter(SUB)}{special_cg_row}"
    row += 1
    line("  (of which) Normal income (slab-tax base)", SUB,
         f"=MAX(0,{ti_ref}-{special_cg_ref})")
    normal_income_row = row
    normal_income_ref = f"{get_column_letter(SUB)}{normal_income_row}"
    row += 2

    # Re-anchor point: hand the page's own coordinates to `Computation` so
    # its slab/rebate/surcharge/cess formulas read THIS page instead of
    # recomputing GTI/TI from source leaves (design doc section 7 step 2).
    page_layout = {
        "gti": gti_ref,
        "total_income": ti_ref,
        "normal_income_base": normal_income_ref,
        "special_cg_base": special_cg_ref,
    }
    # `Computation`'s parallel backing/audit copy of the tax-slab machinery
    # still gets written (re-anchored to read `normal_income_base` above) --
    # design doc section 11.1 keeps `Computation` as a hidden backing sheet.
    # Its return value is no longer consumed for this page's own tax lines
    # (see below): those are built independently via
    # `_write_regime_tax_workings` so this page is genuinely live, not a
    # mirror of a hidden sheet.
    computation_tail_fn(page_layout)

    # Standard tax computation -- ON-PAGE, live formulas (2026-07-21 on-page-
    # totals change, design doc section 11.1): slab tax -> s.87A rebate ->
    # marginal relief -> + special-rate CG tax -> surcharge -> cess -> Total
    # tax liability. Both regimes' full working is computed in helper cells
    # past the print area (see `_write_regime_tax_workings`); the visible
    # lines below just IF()-pick between the two at every step, exactly like
    # the rest of this page's regime selector.
    # Regime tax workings now live in the Workings/Inputs section below
    # Refund (2026-07-22 layout fix) -- written here (their call site is
    # unchanged) but targeting the PREDICTED `workings_start_row` (columns
    # H/I, so this never collides with the visible A-F ladder regardless of
    # which real row the cursor is on at this point in the function).
    cg_tax_ref = f"'CapitalGains'!{cg_layout['special_tax_cell']}"
    cess_rate_ref = f"'Rules'!{rules_layout['cess_rate'].coordinate}"
    tax_workings_row = workings_start_row + 1
    ws.cell(row=workings_start_row, column=_TAX_HELPER_COL,
            value="Regime comparison workings (reference only)").font = _font(9, bold=True, italic=True)
    new_tax_parts, _helper_row = _write_regime_tax_workings(
        ws, tax_workings_row, _TAX_HELPER_COL, "New", normal_income_ref, special_cg_ref, cg_tax_ref, cess_rate_ref,
        rules_layout["new_slabs"], rules_layout["new_rebate_max_ti"].coordinate,
        rules_layout["new_rebate_max_amt"].coordinate, rules_layout["new_rebate_marginal"].coordinate,
        rules_layout["new_surcharge"],
    )
    old_tax_parts, tax_workings_end_row = _write_regime_tax_workings(
        ws, _helper_row, _TAX_HELPER_COL, "Old", normal_income_ref, special_cg_ref, cg_tax_ref, cess_rate_ref,
        rules_layout["old_slabs"], rules_layout["old_rebate_max_ti"].coordinate,
        rules_layout["old_rebate_max_amt"].coordinate, rules_layout["old_rebate_marginal"].coordinate,
        rules_layout["old_surcharge"],
    )

    def tax_sel(key: str) -> str:
        return sel(new_tax_parts[key], old_tax_parts[key])

    line("Tax on total income", SUB, tax_sel("slab"))
    row += 1
    line("Less - s.87A rebate", SUB, tax_sel("rebate"))
    row += 1
    line("Less - Marginal relief (s.87A)", SUB, tax_sel("marginal"))
    row += 1
    line("Add - Special-rate Capital Gains tax (112A/111A)", SUB, f"={cg_tax_ref}")
    row += 1
    line("Add - Surcharge", SUB, tax_sel("surcharge"))
    row += 1
    line("Add - Health & Education Cess", SUB, tax_sel("cess"))
    row += 1
    line("Total tax liability", OUTER, tax_sel("liability"), bold=True, rule=_TOP_RULE)
    total_liability_ref = f"{get_column_letter(OUTER)}{row}"
    row += 1

    # Interest u/s 234 -- added to tax BEFORE prepaid taxes are deducted, the
    # order PartB-TTI uses, so the Refund line below is net of interest. Each
    # line mirrors a live formula in the Workings block; none is a frozen
    # value, so editing the filing date there moves the Refund here.
    line("Add - Interest u/s 234A (late filing)", SUB, f"={i234_refs['a_interest']}")
    row += 1
    line("Add - Interest u/s 234B (advance tax default)", SUB, f"={i234_refs['b_interest']}")
    row += 1
    line("Add - Interest u/s 234C (instalment deferment)", SUB, f"={i234_refs['c_total']}")
    row += 1
    line("Aggregate liability (tax + interest)", OUTER,
         f"={total_liability_ref}+{i234_refs['total']}", bold=True, rule=_TOP_RULE)
    aggregate_liability_ref = f"{get_column_letter(OUTER)}{row}"
    row += 2

    # `tp` and `prepaid` are already defined near the top of this function
    # (hoisted alongside `os_items` so `_predict_pre_workings_rows` can use
    # them too -- see the Workings/Inputs placement comment above).
    section("Less - Prepaid Taxes"); row += 1
    first = row
    for label, key, value in prepaid:
        if value:
            item(label, f"='TaxesPaid'!{tp_layout[key].coordinate}")
            row += 1
    if row == first:                      # nothing prepaid -- still show the head
        item("Prepaid taxes", f"='TaxesPaid'!{tp_layout['total']}")
        row += 1
    close_section(first)

    line("Total prepaid taxes", OUTER, f"='TaxesPaid'!{tp_layout['total']}", bold=True)
    row += 1
    # s.288B rounding (nearest, from Rules) -- matches Computation's own
    # refund formula exactly (MROUND(taxes_paid - liability, round_tax_nearest)),
    # now built directly from this page's own `total_liability_ref` instead of
    # reading `Computation`'s tail. Uses `mround_safe` (not Excel's MROUND)
    # because this expression goes negative for every tax-payable (as opposed
    # to refund) assessee, which would otherwise raise #NUM! on the headline
    # "Refund Due / (Tax Payable)" cell (bug fix, 2026-07-22).
    round_tax_ref = f"'Rules'!{rules_layout['round_tax_nearest'].coordinate}"
    # Net of interest u/s 234 (2026-07-22): the refund/payable line measures
    # prepaid taxes against the AGGREGATE liability, not the bare tax, or the
    # headline figure would understate what is actually due at filing.
    taxes_paid_minus_liability = f"'TaxesPaid'!{tp_layout['total']}-{aggregate_liability_ref}"
    refund_formula = f"={mround_safe(taxes_paid_minus_liability, round_tax_ref)}"
    line("Refund Due / (Tax Payable)", OUTER, refund_formula, bold=True, rule=_TOP_RULE)
    row += 1

    # --- Workings/Inputs section (below Refund) -----------------------------
    # This is the runtime safety net for the row prediction made at the top
    # of this function: if it ever fires, `_predict_pre_workings_rows` has
    # drifted from the body it predicts (most likely: a row was added to/
    # removed from some section above without updating the prediction) --
    # fail loudly here rather than silently write a workbook whose b/f-loss
    # cells or regime tax workings are one row off from what the HP/Business/
    # CG/tax-ladder formulas above actually reference (a `#REF!`-shaped bug
    # that would otherwise only surface much later, if at all, on manual
    # inspection).
    assert row == predicted_post_refund_row, (
        f"Statement of Income layout drift: _predict_pre_workings_rows() "
        f"predicted the row right after Refund would be {predicted_post_refund_row}, "
        f"but it is actually {row}. Fix the prediction (it must mirror every "
        "row written between the body start and the Refund line) before this "
        "ships -- otherwise the Workings/Inputs section below is placed on "
        "top of the wrong cells."
    )
    row += 2  # mirrors the pre-2026-07-22 2-blank-row gap before Assumptions
    assert row == workings_start_row, "Workings/Inputs start-row drift (see assertion above)."

    header = ws.cell(row=row, column=1, value="Workings / Inputs")
    header.font = _font(12, bold=True, underline="single")
    row += 2
    section("Brought forward losses set off (statutory routing, s.71B / s.72 / s.74)")
    row += 1
    assert row == bf_first_row, "b/f-loss block start-row drift (see assertion above)."
    for key, label, _head in _SIMPLE_BF_BUCKETS:
        ws.cell(row=row, column=LBL, value=INDENT + label).font = _font()
        _input_cell(ws, row, SUB, default_value=0)
        assert bf_refs[key] == f"{get_column_letter(SUB)}{row}"
        row += 1
    for key, label in (
        ("bf_stcl", "b/f Short-term capital loss (s.74) -- STCG first, remainder to LTCG"),
        ("bf_ltcl", "b/f Long-term capital loss (s.74) -- LTCG only"),
    ):
        ws.cell(row=row, column=LBL, value=INDENT + label).font = _font()
        _input_cell(ws, row, SUB, default_value=0)
        assert bf_refs[key] == f"{get_column_letter(SUB)}{row}"
        row += 1
    assert row - 1 == bf_last_row, "b/f-loss block end-row drift (see assertion above)."
    row += 1

    # Interest u/s 234 block. Rendered whether or not it could be computed:
    # the ladder above references fixed coordinates inside it, and leaving
    # them unwritten would silently read as zero interest. When there is no
    # income year to compute against, the block is replaced by a note of the
    # same height so the rows still line up and the reader can see WHY the
    # ladder shows nothing.
    assert row == i234_start_row, "Interest u/s 234 block start-row drift (see assertion above)."
    if model.interest_234.computed:
        row = write_234_workings(ws, i234_start_row, model, total_liability_ref)
    else:
        note = ws.cell(
            row=i234_start_row, column=2,
            value="Interest u/s 234A / 234B / 234C not computed -- no income year supplied.",
        )
        note.font = _font(bold=True)
        row = i234_start_row + I234_BLOCK_ROWS
    # The regime tax workings (columns H/I) were already written earlier in
    # this function, at `workings_start_row` -- may run taller than the b/f
    # block above in columns A-F, so advance the cursor past whichever block
    # is taller for a correct `last_row` / print area.
    row = max(row, tax_workings_end_row)
    row += 1

    # Assumptions note only renders while residency is DEFAULTED -- a reader
    # must be able to tell "someone asserted this" from "the tool fell back"
    # (2026-07-19 residency prompt, section 2). Dropped entirely once the
    # entity declares R/OR / RNOR / NR.
    if not residency_declared:
        ws.cell(row=row, column=1, value="Assumptions").font = _font(bold=True, underline="single")
        row += 1
        note = ws.cell(row=row, column=1, value=_ASSUMPTION_NOTE)
        note.font = _font(8, italic=True)
        note.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=OUTER)
        row += 2

    apply_sheet_chrome(
        ws,
        {"A": 3.5, "B": fit_label_width(ws, LBL, 46), "C": 15, "D": 15, "E": 16, "F": 12},
        last_row=row, last_col=OUTER,
        freeze=f"A{header_end + 1}", print_title=print_title,
    )
    return ws


# ---------------------------------------------------------------------------
# Sheets 2 and 3 -- IS and BS
# ---------------------------------------------------------------------------

def _write_hierarchy_sheet(wb, sheet_name: str, title_formula: str, entries,
                           source_sheet: str, label_width: float, print_title: str, *,
                           extra_row_fn=None):
    """`extra_row_fn(ws, next_row, tops, last_col) -> new_next_row`, if given,
    runs after the hierarchy body and before sheet chrome is applied -- e.g.
    `PL for Business`'s net-income row, which is not part of the generic
    hierarchy shape but still needs `last_row`/widths sized to include it."""
    ws = wb.create_sheet(sheet_name)
    t = ws.cell(row=1, column=1, value=title_formula)
    t.font = _font(12, bold=True)
    root = build_hierarchy(entries)
    depth = max_group_level(root)
    last_col = max(2 + depth + 1, 3)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    t.alignment = Alignment(horizontal="center")

    for col in range(2, last_col + 1):
        c = ws.cell(row=3, column=col, value="Rs.")
        c.font = _font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = _BOTTOM_RULE

    next_row, tops = render_hierarchy(ws, 4, root, source_sheet, label_col=1, money_col=2)
    if extra_row_fn is not None:
        next_row = extra_row_fn(ws, next_row, tops, last_col)

    widths = {"A": fit_label_width(ws, 1, label_width)}
    for col in range(2, last_col + 1):
        widths[get_column_letter(col)] = 15
    apply_sheet_chrome(ws, widths, last_row=next_row, last_col=last_col,
                       freeze="A4", print_title=print_title)
    return ws


def write_is_sheet(wb, entries, entity_layout: dict, period_text: str, print_title: str):
    """`IS` -- clustered as closely to GnuCash as the transcript's own `Path`
    column allows. `IS` is naturally shallow; it is not forced to look
    symmetrical with `BS`."""
    title = (f"='Entity'!{entity_layout['name'].coordinate}"
             f"&\" Income Statement For Period Covering {period_text}\"")
    return _write_hierarchy_sheet(wb, "IS", title, entries, "IS_Transcript", 38.1, print_title)


def write_bs_sheet(wb, entries, entity_layout: dict, as_at_text: str, print_title: str):
    """`BS` -- a GnuCash view. Every intermediate group keeps its own row and
    its own subtotal; sibling groups are never merged. In particular
    `Fixed Deposits` stays a sibling of `Cash and Bank` -- Schedule AL's
    statutory buckets do combine them, but that is a different sheet with a
    different purpose and does not license combining them here."""
    title = (f"='Entity'!{entity_layout['name'].coordinate}"
             f"&\" Balance Sheet as at {as_at_text}\"")
    return _write_hierarchy_sheet(wb, "BS", title, entries, "BS_Transcript", 47.8, print_title)


# ---------------------------------------------------------------------------
# PL for Business -- nets one entity's business income against its expenses
# ---------------------------------------------------------------------------

class BusinessSubtreeError(Exception):
    """Raised when an entity's `business_subtree` config (Data/itr/entities.yaml)
    is set but this FY's `IS` entries contain nothing under it.

    A configured `business_subtree` is a structural fact about that entity's
    GnuCash chart of accounts, not a per-year on/off signal -- once set, it is
    expected to keep matching every year. A zero-match year is therefore far
    more likely a GnuCash rename or a typo in the config than a genuine
    zero-business year, and gate 4 (2026-07-19 PL for Business prompt) forbids
    treating the two the same way: silently rendering an empty/omitted sheet
    here would hide exactly the failure this check exists to catch.
    """


def resolve_business_entries(is_entries, business_subtree: str | None):
    """Filter `is_entries` -- the same `[(path, cell), ...]` list `write_is_sheet`
    already consumes -- down to the leaves under `business_subtree` (a GnuCash
    account path prefix, e.g. `"Income/xBusiness Income"`), via a plain
    path-prefix subtree walk. Never a keyword/name match: an out-of-subtree
    account that merely *sounds* business-related (e.g. `Expense/Professional
    Tax`) must never leak in just because its name contains "business"-ish
    words.

    Returns None when `business_subtree` is not configured for this entity --
    the ordinary, per-FY "this entity has no business" case, evaluated fresh
    from config every run (never an entity-level flag baked in elsewhere,
    never cross-year-persisted, never inferred from a prior year or from
    whether a CA reference workbook happened to include the sheet). The
    `PL for Business` sheet is simply omitted; no error.

    Raises BusinessSubtreeError when `business_subtree` IS configured but this
    FY's `is_entries` matches nothing under it (see BusinessSubtreeError).
    """
    if not business_subtree:
        return None
    prefix = business_subtree.rstrip("/") + "/"
    matches = [(path, cell) for path, cell in is_entries if str(path).startswith(prefix)]
    if not matches:
        raise BusinessSubtreeError(
            f"business_subtree {business_subtree!r} is configured for this entity "
            f"but no IS entries under it were found for this FY -- check for a "
            f"GnuCash rename before assuming a zero-business year"
        )
    return matches


def write_pl_for_business_sheet(wb, entries, entity_layout: dict, period_text: str,
                                print_title: str):
    """`PL for Business` -- nets one entity's business income against its
    business expenses via a plain subtree walk (see `resolve_business_entries`).
    Reuses `_write_hierarchy_sheet` exactly like `IS`/`BS`: the income leaf and
    the nested expense group render as an ordinary hierarchy, sourced from
    `IS_Transcript` since `entries` is a filtered slice of the same
    `is_entries` list `IS` renders from. Expense leaves already carry the
    book's negative sign (HTML convention), so the net row is a plain SUM/`+`
    of the top-level cells -- never restated as a subtraction.
    """
    title = (f"='Entity'!{entity_layout['name'].coordinate}"
             f"&\" Profit and Loss for Business For Period Covering {period_text}\"")

    def net_row(ws, row, tops, last_col):
        label = ws.cell(row=row, column=1, value="Net Business Income / (Loss)")
        label.font = _font(bold=True)
        formula = ("=" + "+".join(f"{get_column_letter(c)}{r}" for r, c in tops)) if tops else "=0"
        cell = ws.cell(row=row, column=last_col, value=formula)
        cell.number_format = INR_FORMAT
        cell.font = _font(bold=True)
        cell.border = _TOP_RULE
        return row + 1

    return _write_hierarchy_sheet(wb, "PL for Business", title, entries, "IS_Transcript",
                                  38.1, print_title, extra_row_fn=net_row)


# ---------------------------------------------------------------------------
# Sheet 4 -- CG
# ---------------------------------------------------------------------------

_CG_HEADERS = [
    ("Sr.", "No.", ""),
    ("", "Name", ""),
    ("No. of", "Shares /", "Units"),
    ("Date of", "Purchase", ""),
    ("Cost", "Price", ""),
    ("Valuation", "as of", "31-01-2018"),
    ("Date of", "Sale", ""),
    ("Selling", "Price", ""),
    ("Capital", "Gain /", "(Loss)"),
    ("Taxable", "Capital Gain", "/ (Loss)"),
]

#: CapitalGains 'Lot detail' column letters -- this sheet is a VIEW over that
#: sheet, never a reimplementation. The CA reference hardcodes 31-Jan-2018 FMV
#: prices as literals inside each row's formula and its grandfathering
#: arithmetic is inconsistent between rows (K9 = I9-MAX(F9,G9) but
#: K10 = J10-G10, which subtracts FMV in addition to cost). Neither trait is
#: copied: every figure below is a direct reference to the CapitalGains sheet,
#: which already applies grandfathering consistently via the real fmv_tables
#: lookups.
_CG_SRC = {
    "scrip": "A", "sale_date": "B", "buy_date": "C", "term": "D", "qty": "E",
    "cost": "F", "proceeds": "G", "booked_gain": "H", "fmv_used": "J",
    "taxable_gain": "K",
}


def _cg_block(ws, row: int, banner: str, lot_rows, lot_start_row: int, term: str) -> int:
    b = ws.cell(row=row, column=1, value=banner)
    b.font = _font(11, bold=True)
    b.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(_CG_HEADERS))
    row += 1

    header_top = row
    for col, parts in enumerate(_CG_HEADERS, start=1):
        for offset, word in enumerate(parts):
            c = ws.cell(row=header_top + offset, column=col, value=word or None)
            c.font = _font(bold=True)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
    for col in range(1, len(_CG_HEADERS) + 1):
        ws.cell(row=header_top + 2, column=col).border = _BOTTOM_RULE
    row = header_top + 3

    first_data_row = row
    sr = 0
    for offset, lot in enumerate(lot_rows):
        if lot.term != term:
            continue
        sr += 1
        src = lot_start_row + offset

        def ref(key: str) -> str:
            return f"'CapitalGains'!{_CG_SRC[key]}{src}"

        # Sr. No. is a formula too, so "no numeric literal on this sheet"
        # stays a clean, assertable invariant.
        ws.cell(row=row, column=1, value=f"=ROW()-{first_data_row - 1}").font = _font()
        ws.cell(row=row, column=2, value=f"={ref('scrip')}").font = _font()
        c = ws.cell(row=row, column=3, value=f"={ref('qty')}")
        c.number_format = "#,##0.###"
        for col, key in ((4, "buy_date"), (7, "sale_date")):
            d = ws.cell(row=row, column=col,
                        value=f"=IF({ref(key)}=\"\",\"\",DATEVALUE({ref(key)}))")
            d.number_format = DATE_FORMAT
            d.font = _font()
            d.alignment = Alignment(horizontal="center")
        for col, key in ((5, "cost"), (8, "proceeds"), (9, "booked_gain"), (10, "taxable_gain")):
            m = ws.cell(row=row, column=col, value=f"={ref(key)}")
            m.number_format = INR_FORMAT
            m.font = _font()
        # FMV is READ from CapitalGains (which resolved it from fmv_tables),
        # never restated as a literal price in this formula.
        f = ws.cell(row=row, column=6, value=f"=IF({ref('fmv_used')}=\"\",\"\",{ref('fmv_used')})")
        f.number_format = INR_FORMAT
        f.font = _font()
        row += 1

    if sr == 0:
        ws.cell(row=row, column=2, value="(none)").font = _font(italic=True)
        row += 1

    ws.cell(row=row, column=2, value="SUM").font = _font(bold=True)
    for col in (5, 8, 9, 10):
        letter = get_column_letter(col)
        t = ws.cell(row=row, column=col, value=f"=SUM({letter}{first_data_row}:{letter}{row - 1})")
        t.number_format = INR_FORMAT
        t.font = _font(bold=True)
        t.border = _TOP_RULE
    return row + 2


def has_capital_gains_activity(cg_schedule) -> bool:
    """Does the financial year being generated have any capital-gains activity?

    Evaluated PER RUN, from this year's own data -- lots/disposals, book
    control totals, or taxable figures. Deliberately NOT an entity-level flag,
    config toggle, or anything else that persists one year's answer into the
    next: an entity with no gains this year may well have a disposal next
    year, and a cached "this entity doesn't do CG" would silently drop a real
    CG sheet the year it finally matters. It is also never inferred from a
    prior year's output or from whether a CA's reference workbook happened to
    include the sheet.
    """
    return bool(cg_schedule.lot_rows) or any((
        cg_schedule.lt_control, cg_schedule.st_control,
        cg_schedule.lt_taxable_gross, cg_schedule.st_taxable_gross,
    ))


def write_cg_sheet(wb, cg_schedule, entity_layout: dict, lot_start_row: int, print_title: str):
    ws = wb.create_sheet("CG")
    ncols = len(_CG_HEADERS)

    row = 1
    row = _write_cg_error_banner(ws, row, cg_schedule, ncols)
    title_row = row
    t = ws.cell(row=title_row, column=1, value=f"='Entity'!{entity_layout['name'].coordinate}")
    t.font = _font(12, bold=True)
    t.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=ncols)

    row = title_row + 2
    row = _cg_block(
        ws, row,
        "Details of Long Term Capital Gain / (Loss) on Shares and MF during the year",
        cg_schedule.lot_rows, lot_start_row, "LT",
    )
    row = _cg_block(
        ws, row,
        "Details of Short Term Capital Gain / (Loss) on Shares and MF during the year",
        cg_schedule.lot_rows, lot_start_row, "ST",
    )

    # The Name column holds formulas, so its rendered width can't be measured
    # from the sheet -- size it from the source scrip names instead.
    name_width = max([30.0] + [len(str(r.scrip)) + 2 for r in cg_schedule.lot_rows])
    widths = {"A": 5, "B": min(name_width, 90.0), "C": 12, "D": 13, "E": 14, "F": 13,
              "G": 13, "H": 14, "I": 15, "J": 16}
    apply_sheet_chrome(ws, widths, last_row=row, last_col=ncols,
                       landscape=True, print_title=print_title)
    return ws


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def hide_working_sheets(wb) -> None:
    """Hidden, not gone -- the audit trail stays in the file."""
    for name in HIDDEN_SHEETS:
        if name in wb.sheetnames:
            wb[name].sheet_state = "hidden"


def move_presentation_sheets_first(wb) -> None:
    front = [wb[n] for n in PRESENTATION_SHEETS if n in wb.sheetnames]
    rest = [ws for ws in wb._sheets if ws not in front]
    wb._sheets = front + rest


def build_presentation_layer(wb, model, entity_layout: dict, comp_layout: dict,
                             os_layout: dict, tp_layout: dict, ded_layout: dict,
                             rules_layout: dict, cg_layout: dict,
                             is_entries, bs_entries,
                             lot_start_row: int, age_class: str, year_key: str,
                             year_label: str, computation_tail_fn, *,
                             father_name: str | None = None,
                             aadhaar: str | None = None, residency_value: str = "R/OR",
                             residency_declared: bool = False,
                             business_subtree: str | None = None) -> None:
    """Add the four deliverable sheets in front of the calculation sheets and
    hide the raw working sheets. Adds only -- restyles and overwrites nothing.

    `comp_layout` is the LEAF-only Computation layout (2026-07-20 on-page-totals
    change); `computation_tail_fn` is called from inside `write_statement_of_income`
    once the page's own ladder coordinates exist, to keep `Computation`'s
    parallel backing copy of the tax-slab machinery current. `rules_layout` is
    needed on-page now too, for the s.288A/s.288B rounding (MROUND) and the
    full slab/rebate/surcharge/cess machinery that used to live only on
    `Computation`. `cg_layout` supplies `special_tax_cell` (112A/111A tax on
    special-rate CG) for the on-page tax block. See
    `write_statement_of_income`'s docstring."""
    start_year = int(year_key[:4]) if year_key else 0
    period_text = f"01-04-{start_year} to 31-03-{start_year + 1}"
    as_at_text = f"31-03-{start_year + 1}"
    print_title = f"&\"{FONT_NAME},Bold\"&A  --  {year_label}"

    write_statement_of_income(
        wb, model, entity_layout, comp_layout, os_layout, tp_layout, ded_layout,
        rules_layout, cg_layout, age_class, period_text, print_title, computation_tail_fn,
        father_name=father_name, aadhaar=aadhaar,
        residency_value=residency_value, residency_declared=residency_declared,
    )
    write_is_sheet(wb, is_entries, entity_layout, period_text, print_title)
    write_bs_sheet(wb, bs_entries, entity_layout, as_at_text, print_title)
    # PL for Business is omitted entirely when this entity has no
    # business_subtree configured (the ordinary case for every entity except
    # Harshal's) -- see resolve_business_entries. When configured but this
    # FY's data matches nothing under it, resolve_business_entries raises
    # rather than silently rendering a zero/omitted sheet.
    business_entries = resolve_business_entries(is_entries, business_subtree)
    if business_entries is not None:
        write_pl_for_business_sheet(wb, business_entries, entity_layout, period_text, print_title)
    # CG is omitted entirely when this FY has no capital-gains activity --
    # mirroring what the CA actually produced for such a year. A blank grid on
    # a document handed to a CA or a bank is worse than no sheet.
    if has_capital_gains_activity(model.capital_gains):
        write_cg_sheet(wb, model.capital_gains, entity_layout, lot_start_row, print_title)

    hide_working_sheets(wb)
    move_presentation_sheets_first(wb)
