"""
as26.py -- optional reader for the 26AS skill's own output workbook (plan
section 2.2 row 9 / OQ-5 / D17): "Part I" sheet already carries a
"Transaction Date" column per row (verified against
skill_26as/scripts/extract_26as_to_xlsx.py -- no column enhancement was
needed there), so 234C quarter-bucket attribution and the TaxesPaid tie-out
can both be driven from it directly.

Part I layout (skill_26as's build_part_i): header row 3, data from row 4;
each deductor's rows are followed by a "Sub-total -- <name>" row (col A
starts with '#', col B starts with 'Sub-total') which this reader skips.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime

import openpyxl

_HEADER_ROW = 3
_DATA_START_ROW = 4

_DATE_FORMATS = ("%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d")


@dataclass
class As26Transaction:
    tan: str
    deductor_name: str
    section: str | None
    txn_date: date | None
    amount: float
    tax_deducted: float
    tds_deposited: float


@dataclass
class As26Data:
    transactions: list = field(default_factory=list)   # list[As26Transaction]


def _parse_date(text) -> date | None:
    if text is None or text == "":
        return None
    if isinstance(text, datetime):
        return text.date()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(str(text).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def classify_section(section: str | None, tds_sections: dict) -> str | None:
    """Classify a 26AS transaction's TDS section code into 'dividend' /
    'interest' / None using the Rules-config-driven tds_sections map (CF5) --
    the section->category mapping itself is a statutory fact, never a rate,
    but is still sourced from the Rules yaml rather than hardcoded here."""
    if not section:
        return None
    section = str(section).strip()
    for category, codes in tds_sections.items():
        if section in codes:
            return category
    return None


def parse_as26_workbook(path: str) -> As26Data:
    """Read Part I of a 26AS skill output workbook into a flat transaction
    list, skipping the per-deductor sub-total rows."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Part I" not in wb.sheetnames:
        return As26Data()
    ws = wb["Part I"]

    transactions: list[As26Transaction] = []
    for row in ws.iter_rows(min_row=_DATA_START_ROW, values_only=True):
        if not row or row[0] is None:
            continue
        col_a = str(row[0])
        col_b = str(row[1]) if row[1] is not None else ""
        if col_a.startswith("#") or col_b.startswith("Sub-total"):
            continue  # per-deductor sub-total row
        if col_b == "No Transactions Present":
            continue

        deductor_name, tan = row[1], row[2]
        section, txn_date_raw, amount, tax_deducted, tds_deposited = row[7], row[8], row[12], row[13], row[14]
        if tan is None:
            continue

        transactions.append(As26Transaction(
            tan=str(tan), deductor_name=str(deductor_name or ""), section=section,
            txn_date=_parse_date(txn_date_raw), amount=_to_float(amount),
            tax_deducted=_to_float(tax_deducted), tds_deposited=_to_float(tds_deposited),
        ))
    return As26Data(transactions=transactions)
