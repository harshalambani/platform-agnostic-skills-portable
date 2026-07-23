"""
agent.py -- ITR Workbook -- DIRECT mode, no LLM in the deterministic core.

Batch 1 of 7: parses the eguile Balance Sheet HTML into an account tree and
runs the built-in identity checks (pattern: skill_26as/agent.py).
Batch 2 of 7: when the optional .gnucash book_file is supplied, also parses
it and runs the book<->HTML cross-check (plan section 1.2, point 4), printed
alongside the HTML-only verification summary.
Batch 3 of 7: when the optional mapping_file is supplied, resolves every
HTML leaf to a tag (scripts/mapping.py, nearest-ancestor). Any unmapped
leaf blocks the run (status BLOCKED-FOR-REVIEW, plan section 4.1
fail-loud + learning loop): an LLM tag suggestion is attempted for each
(scripts/suggest.py, degrades to no-suggestion with no endpoint
configured) and a <output>-proposed-mappings.yaml snippet is written next
to the output file, ready to review and paste into the mapping file. No
value from an unmapped leaf reaches any downstream schedule.
Batch 4 of 7: when the optional form16_pdf is supplied, parses the TRACES
Part B/Annexure-I salary computation (scripts/parse_form16.py) and prints
the employer TAN, the 115BAC(1A) opt-out election, and the form's own
internal consistency checks (flagged, never "corrected"). When a
mapping_file is ALSO supplied, additionally runs the two Book<->Form16
cross-checks (scripts/verify.py: 17(1) vs SALARY_GROSS, net tax payable
vs TAXPAID_TDS_SALARY). form16_pan is the decryption password for an
encrypted certificate (TRACES convention: password == employee PAN) --
never stored, used only for this one parse call.
Batch 6 of 7: when mapping_file resolves every leaf (STATUS OK), builds the
full schedule model (scripts/schedules.py, fed by scripts/rules.py's
tax-rules config for the FY inferred from the HTML) and writes the
standardized, formula-driven ITR workbook (scripts/write_workbook.py) --
14 sheets, Computation cells all `='<sheet>'!<cell>` formulas, a
both-regimes tax-block comparison keyed off the Entity sheet's regime
cell. CG-lot reconstruction (scripts/lots.py) + FMV/scrips lookup feed
CapitalGains; dividend/interest quarter bucketing (scripts/quarters.py)
feeds OtherSources, driven by 26AS transaction dates when an as26_workbook
is supplied (CF5, scripts/as26.py), else by book posting dates.
entities_path/entity_key/rules_dir/scrips_path are keyword-only run()
params for this engine; entity_key defaults to the mapping_file's stem.
Falls back to the Batch-1-style stub workbook whenever mapping is
absent/blocked or no rules config matches the FY.
Batch 7 of 7 (final): entity/ay/regime dropdowns land in skill.yaml
(options_from: itr_entities / itr_ay_years; entity_key/ay/regime_override
run() params). The entity is now resolved up front, so an encrypted
Form 16's decrypt password auto-derives from the resolved entity's PAN
(form16_pan removed as a UI input, still available as a programmatic
override). A selected `ay` that disagrees with the HTML's own inferred
income year hard-fails the run rather than silently building against the
wrong year's rules.
Batch 8: entities_path/rules_dir/scrips_path are now anchored (by the UI's
`{data_root}` run_args token, see ui/tabs/_generic.py) to the same data root
ui/_config.data_root_dir() resolves in both source and frozen builds, fixing
a frozen-build bug where the CWD-relative defaults doubled up as
Data/Data/itr/... _resolve_entity() now fails loud (EntityResolutionError)
when an explicitly selected entity_key can't be resolved, instead of
silently substituting a generic UNKNOWN profile. A blank mapping_file with
a selected entity now auto-derives that entity's
<data_root>/itr/mappings/<entity>.mapping.yaml when it exists; true cold
start (no mapping anywhere) now routes into the same BLOCKED-FOR-REVIEW +
proposed-mappings-snippet flow as a partially mapped file, rather than
silently reporting STATUS: OK on an empty one-sheet stub.
2026-07-16 ITR Best-Effort Workbook, Part 1: BLOCKED-FOR-REVIEW-with-nothing-
built is gone. Any unmapped leaf (partial mapping OR true cold start) now
still builds the full BS + P&L + IT working workbook -- every unmapped leaf
routes into scripts/schedules.py's UNCLASSIFIED/REVIEW bucket
(build_unclassified) instead of silently being dropped, so BS/P&L totals
still tally with those amounts included. The IT working shows a DRAFT tax
(resolved items only) plus a worst-case upper bound (every unclassified
INCOME-type leaf taxed at the top slab; unclassified expense/deduction
items are never assumed to reduce tax) -- both loudly labelled, neither
presented as filing-ready. STATUS is now "BUILT -- N REVIEW ITEM(S)" (N>0)
instead of BLOCKED-FOR-REVIEW; the <output>-proposed-mappings.yaml
learning-loop snippet is still written whenever N>0. Hard-error paths
(unparseable HTML, unresolved entity, AY-vs-HTML mismatch, mapping
VALIDATION ERROR) are unchanged -- still stub + ERROR, no workbook.
"""
from __future__ import annotations

import argparse
import datetime
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / "scripts"))
from parse_eguile import parse_file, verify  # noqa: E402
import parse_gnucash as pg  # noqa: E402
import parse_form16  # noqa: E402
import verify as book_verify  # noqa: E402
import configs  # noqa: E402
import mapping as mapping_engine  # noqa: E402
import suggest  # noqa: E402
import rules as rules_engine  # noqa: E402
import schedules as sch  # noqa: E402
import write_workbook  # noqa: E402
import presentation  # noqa: E402
import as26 as as26_engine  # noqa: E402
from openpyxl.utils.exceptions import InvalidFileException  # noqa: E402

_AS_OF_RE = re.compile(r"Balance Sheet \(eguile\)[^0-9]*(\d{2})-(\d{2})-(\d{4})")

OK = "OK"
BLOCKED_FOR_REVIEW = "BLOCKED-FOR-REVIEW"   # hard-error paths only (bad HTML, unresolved
                                             # entity, AY mismatch, mapping VALIDATION ERROR) --
                                             # unmapped leaves no longer route here (2026-07-16
                                             # Part 1: best-effort build-with-call-outs instead).
BUILT_WITH_REVIEW_PREFIX = "BUILT --"


def _infer_year_key(bs_html: str) -> str | None:
    """Read the eguile HTML's 'Balance Sheet (eguile) 31-03-YYYY' as-of date
    and derive the canonical income-year key (plan section 5.1, D19), e.g.
    31-03-2025 -> '2024-25'. Returns None if the date can't be found."""
    text = Path(bs_html).read_text(encoding="utf-8", errors="replace")
    m = _AS_OF_RE.search(text)
    if not m:
        return None
    year = int(m.group(3))
    return f"{year - 1}-{str(year)[-2:]}"


def _fy_to_ay(fy_year_key: str) -> str:
    """Convert a canonical income-year (FY) key, e.g. '2024-25', to the
    Assessment Year key format used by entities.yaml's regime_by_ay,
    e.g. '2025-26'."""
    start = int(fy_year_key[:4])
    return f"{start + 1}-{str(start + 2)[-2:]}"


def _provenance_summary_lines(tree, resolved: dict, mapping_entries: dict | None) -> list[str]:
    """Fail-loud (2026-07-19 mapping-precedence prompt item 1b): report the
    heuristic-vs-approved split in the run's own text summary, not only on
    the Mapping Review sheet -- and surface any heuristic-tagged INCOME
    account by name, since an unverified guess there can silently misstate
    taxable income. Returns [] when nothing has resolved yet (cold start)."""
    if not resolved:
        return []
    counts = write_workbook.provenance_counts(resolved, mapping_entries)
    lines = [
        f"Mapping: tag provenance -- {counts.get('approved', 0)} approved, "
        f"{counts.get('heuristic', 0)} heuristic (unverified), {counts.get('llm', 0)} llm-suggested, "
        f"{counts.get('inherited', 0)} inherited (of {len(resolved)} resolved leaf(ves))."
    ]
    income_leaves = write_workbook.heuristic_income_leaves(tree, resolved, mapping_entries)
    if income_leaves:
        lines.append(
            f"  WARNING: {len(income_leaves)} INCOME account(s) resolved via an unreviewed "
            "heuristic guess (not human-approved) -- verify before filing:"
        )
        for path, tag, amount in income_leaves:
            lines.append(f"    - {path} -> {tag} (FY total {amount})")
    return lines


def _mapping_summary(
    tree,
    mapping_file: str | None,
    output_path: str,
    config_path: str,
    model_override: str | None,
):
    """Resolve every HTML leaf to a tag via the supplied mapping file (plan
    sections 3.1, 4.1). Returns (summary_lines, status, result_or_None).
    When mapping_file is blank (cold start -- no mapping supplied and none
    could be auto-derived for the entity), resolution still runs against an
    empty mapping so every leaf is treated as unmapped: this routes into the
    same BLOCKED-FOR-REVIEW + proposed-mappings snippet flow as a partially
    mapped file, rather than silently reporting OK with nothing built (a
    mapping-less run must never green-light an empty stub). status is OK
    only when every leaf actually resolved to a tag; BLOCKED-FOR-REVIEW
    whenever any leaf is unmapped -- in which case an LLM suggestion is
    attempted for each (degrades gracefully with no endpoint) and a
    <output>-proposed-mappings.yaml snippet is written next to output_path,
    ready to review and paste into the mapping file.
    result_or_None is the ResolutionResult (needed for the Form16
    cross-checks) when resolution ran at all, else None. loaded_entries is
    the raw guid -> MappingEntry dict (Mapping Review sheet's Suggested-by
    column needs the note/suggested_by_llm fields that ResolvedLeaf doesn't
    carry), or None when mapping_file is invalid."""
    known_paths = {n.guid: n.path for n in tree.all_nodes() if n.guid}
    cold_start = not mapping_file
    if cold_start:
        loaded = configs.MappingLoadResult(entries={}, warnings=[])
    else:
        try:
            loaded = configs.load_mapping(mapping_file, known_paths=known_paths)
        except configs.MappingValidationError as e:
            return [f"Mapping: VALIDATION ERROR: {e}"], BLOCKED_FOR_REVIEW, None, None

    result = mapping_engine.resolve_tree(tree, loaded)
    lines = [f"Mapping: {len(result.resolved)} leaf(ves) resolved, {len(result.unmapped)} unmapped."]
    for w in result.warnings:
        lines.append(f"  WARNING: {w}")
    lines.extend(_provenance_summary_lines(tree, result.resolved, loaded.entries))

    if not result.blocked:
        lines.append("Mapping: OK -- every leaf resolved to a tag.")
        return lines, OK, result, loaded.entries

    if cold_start:
        lines.insert(
            0,
            "Mapping: no mapping_file supplied and none found for this entity -- "
            "cold start: treating every leaf as unmapped.",
        )
    n_unmapped = len(result.unmapped)
    unmapped_total = sum(leaf.total or 0.0 for leaf in result.unmapped)
    status = f"{BUILT_WITH_REVIEW_PREFIX} {n_unmapped} REVIEW ITEM(S)"
    lines.append(
        f"Mapping: {n_unmapped} unmapped account(s) (Rs {unmapped_total:,.2f} total) -- routed to the "
        "UNCLASSIFIED/REVIEW bucket; the workbook still builds (best-effort, plan decision locked "
        "2026-07-16). See the Unclassified/Computation sheets for the DRAFT tax and worst-case "
        "upper bound. Unmapped accounts:"
    )
    for leaf in result.unmapped:
        lines.append(f"  - {leaf.path} (FY total {leaf.total})")

    suggestions = suggest.suggest_for_unmapped(result.unmapped, result.resolved, config_path, model_override)
    snippet = mapping_engine.proposed_mapping_snippet(result, suggestions)
    snippet_path = f"{output_path}-proposed-mappings.yaml"
    Path(snippet_path).write_text(snippet, encoding="utf-8")
    lines.append(f"Proposed mappings written to {snippet_path} -- review and paste into the mapping file.")
    if suggestions:
        lines.append(f"  ({len(suggestions)} of {n_unmapped} unmapped account(s) got an LLM suggestion.)")
    else:
        lines.append("  (No LLM suggestions -- no endpoint configured, or the endpoint call failed.)")

    return lines, status, result, loaded.entries


def _form16_summary(tree, data, parse_error: str | None, resolved: dict | None) -> list[str]:
    """Print the employer TAN, the 115BAC(1A) opt-out election, and the
    form's own internal consistency checks (flagged, never "corrected") for
    an already-parsed Form16Data. When `resolved` (the mapping engine's
    guid -> ResolvedLeaf) is available, also runs the two Book<->Form16
    cross-checks."""
    if parse_error is not None:
        return [f"Form16: PARSE ERROR: {parse_error}"]
    if data is None:
        return ["Form16: no form16_pdf supplied -- skipped."]

    lines = [
        f"Form16: employer TAN {data.tan}, certificate {data.certificate_no}, AY {data.assessment_year}.",
        f"Form16: opted out of 115BAC(1A)? {data.opted_out_115bac}",
    ]
    if data.extra_certificates:
        lines.append(
            f"Form16: {len(data.extra_certificates)} extra certificate(s) in this PDF (not parsed): "
            + ", ".join(f"{cert or '?'}/{tan}" for cert, tan in data.extra_certificates)
        )

    if data.identity_ok:
        lines.append(f"Form16: internal consistency OK ({len(data.identity_checks)} check(s)).")
    else:
        lines.append("Form16: internal consistency FAILURES (flagged, not corrected):")
        for c in data.identity_checks:
            if not c.ok:
                lines.append(f"  MISMATCH {c.label}: expected={c.expected:.2f} actual={c.actual:.2f}")

    if resolved is not None:
        form16_results = book_verify.cross_check_form16(tree, resolved, data)
        lines.append(book_verify.summarize_form16(form16_results))
    else:
        lines.append("Book<->Form16 cross-check: no mapping_file supplied -- skipped.")

    return lines


def _verify_summary(
    bs_html: str,
    book_file: str | None,
    mapping_file: str | None,
    form16_pdf: str | None,
    form16_pan: str | None,
    output_path: str,
    config_path: str,
    model_override: str | None,
):
    """Returns (summary_text, tree, book, failures, book_cross_check_results,
    resolution_result_or_None, form16_data_or_None, year_key_or_None,
    status, mapping_entries_or_None). `tree`/`failures` are None-safe
    placeholders when the HTML itself failed to parse (summary_text starts
    with 'ERROR:')."""
    try:
        tree = parse_file(bs_html)
    except ValueError as e:
        return f"ERROR: {e}", None, None, [], [], None, None, None, BLOCKED_FOR_REVIEW, None

    failures = verify(tree)
    node_count = sum(1 for _ in tree.all_nodes())
    lines = [
        f"Parsed {node_count} account rows from {Path(bs_html).name}.",
        f"Imbalance Amount: {tree.imbalance:.2f}",
        f"Total Assets Accounts: {tree.section_totals.get('Assets Accounts', 0.0):.2f}",
        f"Total Equity, Trading, and Liabilities: "
        f"{tree.section_totals.get('Equity, Trading, and Liabilities', 0.0):.2f}",
        f"Total Retained Earnings: {tree.section_totals.get('Retained Earnings', 0.0):.2f}",
    ]
    if failures:
        lines.append("Verify: VALIDATION ERRORS:")
        lines.extend(f"  - {f}" for f in failures)
    else:
        lines.append("Verify: OK -- all identity checks passed.")

    book = None
    book_cross_check = []
    year_key = _infer_year_key(bs_html)
    if book_file:
        lines.append("")
        if year_key is None:
            lines.append("Book<->HTML cross-check: could not infer the FY from the HTML "
                         "as-of date -- skipped.")
        else:
            book = pg.parse_book(book_file)
            book_cross_check = book_verify.cross_check(tree, book, year_key)
            lines.append(f"[FY {year_key}] {book_verify.summarize(book_cross_check)}")

    lines.append("")
    mapping_lines, status, result, mapping_entries = _mapping_summary(tree, mapping_file, output_path, config_path, model_override)
    lines.append(f"STATUS: {status}")
    lines.extend(mapping_lines)

    lines.append("")
    # The Book<->Form16 cross-check needs tags from an ACTUAL mapping file;
    # a cold-start ResolutionResult (mapping_file blank, resolved against an
    # empty mapping) is not one -- pass None so it reports "skipped" rather
    # than a spurious 0-vs-Form16 MISMATCH for every tag nothing was ever
    # supplied to resolve.
    resolved = result.resolved if (result is not None and mapping_file) else None
    form16_data = None
    form16_error = None
    if form16_pdf:
        try:
            form16_data = parse_form16.parse_form16(form16_pdf, pan=form16_pan)
        except parse_form16.Form16ParseError as e:
            form16_error = str(e)
    lines.extend(_form16_summary(tree, form16_data, form16_error, resolved))

    return "\n".join(lines), tree, book, failures, book_cross_check, result, form16_data, year_key, status, mapping_entries


def _write_stub_workbook(output_path: str, summary: str) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"
    ws["A1"] = "ITR Workbook -- scaffold (mapping/rules unavailable this run)"
    for i, line in enumerate(summary.splitlines(), start=2):
        ws.cell(row=i, column=1, value=line)
    wb.save(output_path)


class EntityResolutionError(Exception):
    """Raised by _resolve_entity when an explicitly selected entity_key
    cannot be resolved (entities.yaml missing/unreadable, or the key isn't
    in it). Fail-loud: a missing entity must never silently degrade to a
    generic Individual/new-regime profile, since that would silently pick
    the wrong regime/age band for the whole run."""


def _resolve_entity(mapping_file: str | None, entities_path: str, entity_key: str | None) -> configs.EntityProfile:
    """Resolve the entity profile driving rules.py/schedules.py's regime and
    age-class decisions (and, since Batch 7, the Form16 decrypt PAN).

    entity_key is normally the skill.yaml `entity` dropdown's value -- an
    explicit user selection. When it is set but cannot be resolved (the
    entities.yaml at entities_path is missing/unreadable, or the key isn't
    in it), this is fail-loud: EntityResolutionError is raised naming the
    resolved path that was looked at, rather than silently substituting a
    generic UNKNOWN/Individual/new-regime profile (which would silently
    pick the wrong regime/age band for the run).

    When entity_key is omitted, it defaults to the mapping file's stem
    (Data/itr/mappings/<entity>.mapping.yaml convention) as a best-effort
    lookup only -- not a user selection -- so an unresolved stem-derived
    key still degrades gracefully to a generic profile rather than failing
    a run where no entity was actually chosen."""
    key = entity_key or None   # the UI's select input sends "" when unset, not None
    explicit = key is not None
    if key is None and mapping_file:
        key = Path(mapping_file).stem
        if key.endswith(".mapping"):
            key = key[: -len(".mapping")]

    try:
        entities = configs.load_entities(entities_path)
    except (OSError, configs.ConfigValidationError) as e:
        if explicit:
            raise EntityResolutionError(
                f"entity {key!r} could not be resolved: entities.yaml unreadable at "
                f"{Path(entities_path).resolve()} ({e})"
            ) from e
        return configs.EntityProfile(key=key or "UNKNOWN", name=key or "Unknown", pan="", status="Individual")

    if key and key in entities:
        return entities[key]

    if explicit:
        raise EntityResolutionError(
            f"entity {key!r} not found in {Path(entities_path).resolve()}"
        )
    return configs.EntityProfile(key=key or "UNKNOWN", name=key or "Unknown", pan="", status="Individual")


def _build_and_write_workbook(
    tree, book, result, form16_data, year_key: str | None, failures: list, book_cross_check: list,
    output_path: str, mapping_file: str | None, entity: "configs.EntityProfile",
    rules_dir: str, scrips_path: str, as26_workbook: str | None = None,
    regime_override: str | None = None, mapping_entries: dict | None = None,
) -> list[str]:
    """Builds the full schedule model + formula-driven workbook (plan
    sections 2.2/3) whenever mapping resolution ran at all. Returns extra
    summary lines describing what happened; the caller still writes a stub
    workbook on any failure so a run never crashes for a schedule/rules
    problem outside this batch's control (e.g. no rules config for this
    year yet). `entity` is resolved once by the caller (run()) rather than
    re-resolved here, so the Form16 PAN auto-derivation (B7) and this
    schedule build always agree on the same profile.

    2026-07-16 Part 1: a resolved-but-partial mapping (result.blocked,
    i.e. some leaves unmapped) NO LONGER skips the build -- every unmapped
    leaf routes into schedules.py's UNCLASSIFIED/REVIEW bucket instead
    (build_unclassified) and the full workbook still builds with loud
    call-outs (write_workbook.py). Only result is None (a mapping
    VALIDATION ERROR -- a config bug, not a review case) or a missing
    year_key still skips the build entirely."""
    if result is None or year_key is None:
        return []

    try:
        rules = rules_engine.load_rules(rules_dir, year_key)
    except rules_engine.RulesError as e:
        return [f"Workbook: rules config unavailable ({e}) -- stub workbook only."]

    regime = regime_override or entity.regime_by_ay.get(_fy_to_ay(year_key), entity.default_regime)

    scrips = {}
    if Path(scrips_path).is_file():
        scrips = configs.load_scrips(scrips_path)
    fmv_tables = sch.load_fmv_tables()

    as26_data = None
    as26_lines: list[str] = []
    if as26_workbook:
        try:
            as26_data = as26_engine.parse_as26_workbook(as26_workbook)
            as26_lines.append(
                f"26AS: loaded {len(as26_data.transactions)} transaction(s) from "
                f"{Path(as26_workbook).name} -- driving 234C quarter buckets + TaxesPaid tie-out."
            )
        except (OSError, KeyError, InvalidFileException) as e:
            as26_lines.append(f"26AS: workbook unreadable ({e}) -- book-date buckets only, tie-out skipped.")

    model = sch.build_all_schedules(
        tree, result.resolved, book, form16_data, year_key, rules, regime,
        entity.status, entity.dob, scrips, fmv_tables, as26_data, result.unmapped,
        residency=entity.residency,
    )

    form16_cross_check = book_verify.cross_check_form16(tree, result.resolved, form16_data) if form16_data else []
    user_rules = rules_engine.load_user_rules(str(Path(rules_dir) / "user_rules.yaml"))

    write_workbook.write_workbook(
        output_path, tree, model, rules, user_rules, entity, regime, year_key,
        form16_data.opted_out_115bac if form16_data else None,
        failures, book_cross_check, form16_cross_check, result.unmapped,
        Path(mapping_file).name if mapping_file else "none",
        datetime.datetime.now().isoformat(timespec="seconds"), {},
        result.resolved, mapping_entries,
    )
    lines = [
        f"Workbook: full schedule model built for {rules.year_label} (regime={regime}), "
        f"written to {Path(output_path).name}.",
    ]
    if result.unmapped:
        n = len(result.unmapped)
        total = sum(leaf.total or 0.0 for leaf in result.unmapped)
        lines.append(
            f"Workbook: {n} account(s) unclassified (Rs {total:,.2f} total) -- routed to the "
            "UNCLASSIFIED/REVIEW bucket; DRAFT tax + worst-case upper bound shown on Computation "
            "(NOT filing-ready). See Unclassified / Mapping Review sheets."
        )
    lines.extend(as26_lines)
    if model.taxes_paid.as26_available and not model.taxes_paid.tie_out_ok:
        for c in model.taxes_paid.tie_out_conflicts:
            lines.append(
                f"26AS tie-out CONFLICT -- {c['category']}: book={c['book']:.2f} "
                f"26AS={c['as26']:.2f} diff={c['diff']:.2f} (see TaxesPaid/Reconciliation sheets)."
            )
    if model.capital_gains.unresolved_scrips:
        lines.append(
            "Workbook: FMV lookup FAILED for scrip(s) (fail-loud, review scrips.yaml): "
            + ", ".join(model.capital_gains.unresolved_scrips)
        )
    if not model.capital_gains.reconciliation_ok:
        # "Banner, no abort" (2026-07-19 CG gain-split-vs-action fix): the
        # workbook above has ALREADY been written in full, with a matching
        # top-of-sheet ERROR banner on the CG and Statement of Income sheets
        # (see presentation.py's _write_cg_error_banner). This line's job is
        # only to carry presentation.CG_RECONCILIATION_ERROR_MARKER into the
        # run() summary so that main()'s process-level exit code below (and
        # any other caller grepping the summary) can detect the mismatch.
        lines.append(
            f"{presentation.CG_RECONCILIATION_ERROR_MARKER} "
            f"(diff {model.capital_gains.reconciliation_diff:,.2f}) -- workbook still written "
            "(see ERROR banner on CG / Statement of Income sheets) -- DO NOT FILE without review."
        )
    if model.taxes_paid.unclassified_sections:
        # "Banner, no abort" (2026-07-23 26AS s.193-drop fix), same contract
        # as the CG/Salary checks above: the workbook has ALREADY been
        # written in full, with a matching top-of-sheet ERROR banner on the
        # Statement of Income sheet (see
        # presentation._write_taxes_paid_unclassified_banner) and a detail
        # row on the TaxesPaid working sheet. This line's job is only to
        # carry presentation.TAXES_PAID_UNCLASSIFIED_SECTION_ERROR_MARKER
        # into the run() summary so main()'s process-level exit code below
        # (and any other caller grepping the summary) can detect it.
        total_at_stake = sum(i["amount"] for i in model.taxes_paid.unclassified_sections)
        sections = ", ".join(sorted({i["section"] for i in model.taxes_paid.unclassified_sections}))
        lines.append(
            f"{presentation.TAXES_PAID_UNCLASSIFIED_SECTION_ERROR_MARKER} "
            f"(section(s) {sections}, Rs {total_at_stake:,.2f} TDS at stake) -- workbook still "
            "written (see ERROR banner on Statement of Income / TaxesPaid sheets) -- DO NOT FILE "
            "without review."
        )
    if not model.salary.reconciliation_ok:
        # "Banner, no abort" (2026-07-22 salary-gross fix), same contract as
        # the CG reconciliation check above: the workbook has ALREADY been
        # written in full, with a matching top-of-sheet ERROR banner on the
        # Salary and Statement of Income sheets (see
        # presentation._write_salary_error_banner). This line's job is only
        # to carry presentation.SALARY_RECONCILIATION_ERROR_MARKER into the
        # run() summary so that main()'s process-level exit code below (and
        # any other caller grepping the summary) can detect the mismatch.
        lines.append(
            f"{presentation.SALARY_RECONCILIATION_ERROR_MARKER} "
            f"(diff {model.salary.reconciliation_diff:,.2f}) -- workbook still written "
            "(see ERROR banner on Salary / Statement of Income sheets) -- DO NOT FILE without review."
        )
    return lines


def run(
    bs_html: str,
    output_path: str,
    book_file: str | None = None,
    mapping_file: str | None = None,
    form16_pdf: str | None = None,
    form16_pan: str | None = None,
    as26_workbook: str | None = None,
    config_path: str = "config.yaml",
    model_override: str | None = None,
    entities_path: str = "Data/itr/entities.yaml",
    entity_key: str | None = None,
    rules_dir: str = "Data/itr/rules",
    scrips_path: str = "Data/itr/scrips.yaml",
    ay: str | None = None,
    regime_override: str | None = None,
) -> str:
    """
    Parse the eguile Balance Sheet HTML at bs_html, run the identity checks,
    and (when a mapping_file resolves every leaf) build the full schedule
    model and write the standardized, formula-driven ITR workbook (plan
    section 2.2) at output_path. Deterministic core -- no LLM unless
    mapping_file surfaces unmapped accounts, in which case an optional tag
    suggestion is attempted (scripts/suggest.py) and degrades gracefully to
    none when no LLM endpoint is configured.

    When book_file is supplied, also runs the book<->HTML cross-check and
    feeds CG-lot reconstruction + dividend/interest quarter bucketing. When
    mapping_file is supplied, resolves every leaf to a tag (plan section
    3.1); any unmapped leaf (2026-07-16 Part 1) routes into the
    UNCLASSIFIED/REVIEW bucket instead of blocking the build, sets the
    run's status to "BUILT -- N REVIEW ITEM(S)", and writes a
    <output_path>-proposed-mappings.yaml snippet -- the full workbook still
    builds, with a DRAFT tax figure (resolved items only) and a worst-case
    upper bound shown alongside, neither presented as filing-ready.

    When form16_pdf is supplied, parses the Part B/Annexure-I salary
    computation and feeds the Salary schedule + the two Book<->Form16
    cross-checks (when mapping_file is ALSO supplied). form16_pan is the
    decryption password (TRACES convention: password == employee PAN);
    never stored. Batch 7: form16_pan is no longer a skill.yaml UI input --
    the entity resolved from entity_key/entities.yaml already carries its
    PAN, so it is used automatically as the decrypt password when
    form16_pan is not explicitly passed (the explicit kwarg still wins, for
    programmatic callers/tests that need to override it).

    entities_path/entity_key/rules_dir/scrips_path remain keyword-only
    conveniences for the schedules/rules engine; entity_key is now also the
    ITR Workbook skill.yaml's `entity` dropdown value (Batch 7,
    options_from: itr_entities), defaulting to the mapping_file's stem when
    omitted. They default to the production Data/itr/ convention paths.

    ay (Batch 7), when supplied, is the skill.yaml `ay` dropdown's selected
    canonical income-year key (options_from: itr_ay_years, e.g. "2025-26").
    If it disagrees with the year inferred from the Balance Sheet HTML's
    own as-of date, the run hard-fails (ERROR summary + stub workbook only)
    rather than silently building against the wrong year's rules -- a
    year-mismatch is a real filing risk, not a warning. Omit to skip this
    check (year is inferred from the HTML alone, as in prior batches).

    regime_override (Batch 7), when supplied, is the skill.yaml `regime`
    dropdown's value ("new"/"old") and overrides the entity's configured
    default_regime/regime_by_ay for this run only; leave unset/blank to use
    the entity's configured regime.

    as26_workbook (CF5), when supplied, is the 26AS skill's own output
    workbook (scripts/as26.py reads its Part I sheet). Its transaction
    dates become the authoritative source for the dividend/interest 234C
    quarter buckets on the OtherSources sheet (overriding the book-date
    buckets, plan D17), and its per-transaction tax_deducted totals feed
    the TaxesPaid sheet's 26AS tie-out -- any book-vs-26AS TDS mismatch on
    interest or dividend is flagged as a CONFLICT on both the TaxesPaid and
    Reconciliation sheets, never silently reconciled. Falls back to
    book-date buckets and a skipped tie-out when omitted or unreadable.
    """
    try:
        entity = _resolve_entity(mapping_file, entities_path, entity_key)
    except EntityResolutionError as e:
        summary = f"ERROR: {e}"
        _write_stub_workbook(output_path, summary)
        return summary

    auto_derived_note: str | None = None
    if not mapping_file and entity_key:
        # B(i): the entity dropdown never auto-attached the entity's own
        # mapping file, so a mapping-less run for an entity that already has
        # an approved mapping silently produced an empty stub. Default to
        # <entities_path's data-root>/itr/mappings/<entity>.mapping.yaml
        # when it exists, and proceed as if it had been supplied.
        candidate = Path(entities_path).parent / "mappings" / f"{entity.key}.mapping.yaml"
        if candidate.is_file():
            mapping_file = str(candidate)
            auto_derived_note = (
                f"Mapping: auto-derived {candidate.name} for entity {entity.key!r} "
                "(Entity mapping box was empty)."
            )

    effective_form16_pan = form16_pan if form16_pan is not None else (entity.pan or None)

    summary, tree, book, failures, book_cross_check, result, form16_data, year_key, status, mapping_entries = _verify_summary(
        bs_html, book_file, mapping_file, form16_pdf, effective_form16_pan, output_path, config_path, model_override,
    )
    if auto_derived_note:
        summary = auto_derived_note + "\n\n" + summary

    if tree is None:
        _write_stub_workbook(output_path, summary)
        return summary

    if ay and year_key and ay != year_key:
        summary = summary + "\n\n" + (
            f"ERROR: selected Assessment Year does not match the Balance Sheet HTML -- "
            f"the HTML's as-of date infers income year {year_key!r} but {ay!r} was "
            f"selected. Pick the matching year, or upload the HTML for the selected "
            f"year, then run again."
        )
        _write_stub_workbook(output_path, summary)
        return summary

    extra_lines = _build_and_write_workbook(
        tree, book, result, form16_data, year_key, failures, book_cross_check, output_path,
        mapping_file, entity, rules_dir, scrips_path, as26_workbook, regime_override, mapping_entries,
    )
    if extra_lines:
        summary = summary + "\n\n" + "\n".join(extra_lines)
    else:
        _write_stub_workbook(output_path, summary)

    return summary


def main(argv: list[str] | None = None) -> int:
    """Standalone CLI wrapper around run() (this skill is normally invoked
    in-process via skill.yaml's `mode: "direct"` / `entry_point: "agent:run"`,
    through the registry or the Gradio UI's background thread -- see
    ui/_runner.py -- so no OS-process exit code previously existed for it).
    Added for the CG-reconciliation "banner, no abort" fail-loud requirement
    (2026-07-19 CG gain-split-vs-action fix): the workbook is ALWAYS written
    in full, but a genuine CG control mismatch must still be impossible to
    miss for a process/CI caller, so this wrapper greps run()'s returned
    summary for presentation.CG_RECONCILIATION_ERROR_MARKER and turns that
    into a non-zero exit code plus a stderr line, mirroring this repo's
    other skills' `def main() -> int: ... print("ERROR: ...", file=sys.stderr);
    return 1` convention (e.g. skill_krc/scripts/parse_krc_ledger.py)."""
    parser = argparse.ArgumentParser(
        description="Build the ITR workbook from the command line (wraps agent.run()).",
    )
    parser.add_argument("bs_html", help="Path to the eguile Balance Sheet HTML.")
    parser.add_argument("output_path", help="Path to write the .xlsx workbook to.")
    parser.add_argument("--book-file", dest="book_file", default=None)
    parser.add_argument("--mapping-file", dest="mapping_file", default=None)
    parser.add_argument("--form16-pdf", dest="form16_pdf", default=None)
    parser.add_argument("--form16-pan", dest="form16_pan", default=None)
    parser.add_argument("--as26-workbook", dest="as26_workbook", default=None)
    parser.add_argument("--config-path", dest="config_path", default="config.yaml")
    parser.add_argument("--model-override", dest="model_override", default=None)
    parser.add_argument("--entities-path", dest="entities_path", default="Data/itr/entities.yaml")
    parser.add_argument("--entity-key", dest="entity_key", default=None)
    parser.add_argument("--rules-dir", dest="rules_dir", default="Data/itr/rules")
    parser.add_argument("--scrips-path", dest="scrips_path", default="Data/itr/scrips.yaml")
    parser.add_argument("--ay", dest="ay", default=None)
    parser.add_argument("--regime-override", dest="regime_override", default=None)
    args = parser.parse_args(argv)

    summary = run(
        args.bs_html, args.output_path, book_file=args.book_file, mapping_file=args.mapping_file,
        form16_pdf=args.form16_pdf, form16_pan=args.form16_pan, as26_workbook=args.as26_workbook,
        config_path=args.config_path, model_override=args.model_override,
        entities_path=args.entities_path, entity_key=args.entity_key, rules_dir=args.rules_dir,
        scrips_path=args.scrips_path, ay=args.ay, regime_override=args.regime_override,
    )
    print(summary)
    exit_code = 0
    if presentation.CG_RECONCILIATION_ERROR_MARKER in summary:
        print(
            f"{presentation.CG_RECONCILIATION_ERROR_MARKER} -- workbook was written but "
            "MUST be reviewed before filing (see ERROR banner on CG / Statement of Income "
            "sheets).",
            file=sys.stderr,
        )
        exit_code = 1
    if presentation.SALARY_RECONCILIATION_ERROR_MARKER in summary:
        print(
            f"{presentation.SALARY_RECONCILIATION_ERROR_MARKER} -- workbook was written but "
            "MUST be reviewed before filing (see ERROR banner on Salary / Statement of Income "
            "sheets).",
            file=sys.stderr,
        )
        exit_code = 1
    if presentation.TAXES_PAID_UNCLASSIFIED_SECTION_ERROR_MARKER in summary:
        print(
            f"{presentation.TAXES_PAID_UNCLASSIFIED_SECTION_ERROR_MARKER} -- workbook was "
            "written but MUST be reviewed before filing (see ERROR banner on Statement of "
            "Income / TaxesPaid sheets).",
            file=sys.stderr,
        )
        exit_code = 1
    return exit_code


if __name__ == "__main__":
    sys.exit(main())
