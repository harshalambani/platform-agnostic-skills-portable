"""
build_krc_gnucash.py — Part III of the KR Choksey broker-import skill.

Turn the Part II "Bills" sheet into importable GnuCash multi-split CSVs. Each
Bills row maps to exactly one entry, by Type/Direction:

  Purchase (TRADE, Payable/Dr):
      Dr  <security account>      = Net/Bill Amount   (Shares = +Quantity)
      Cr  <broker account>        = Net/Bill Amount
  SLBM (Type SLBM — the netting game; net proceeds booked as income):
      Dr  <broker account>        = Net/Bill Amount
      Cr  Income from SLBS        = Net/Bill Amount
  Sale (TRADE, Receivable/Cr) — multi-split with FIFO cost basis:
      Dr  <broker account>        = Net/Bill Amount   (sale proceeds)
      Cr  <security account>      = FIFO cost basis    (Shares = -Quantity)
      Cr  Long/Short Term Capital Gain = proceeds - cost basis (per-lot split;
          negative split value = a gain credit, positive = a loss debit)

FIFO cost basis and holding period come from the security's prior purchase lots
in the supplied .gnucash book (and any earlier purchase in this same run).
Lots held strictly longer than `long_term_threshold_months` (config) are Long
Term; the gain is apportioned per consumed lot into LTCG / STCG.

Output (one CSV per type, multi-split layout — consecutive rows = one txn):
    Date, Transaction ID, Description, Account, Amount, Shares, Currency
GnuCash sign convention: Debit positive, Credit negative; per txn the Amount
column sums to zero. Account = full colon path WITHOUT "Root Account:".
Import each file with the "Multi-split" box ticked, skipping 1 header line; map
Amount->Amount and Shares->Shares (do NOT use a Price column).

Usage:
    python build_krc_gnucash.py <bills_xlsx> <gnucash_path> <out_dir> [config_yaml]

Exit codes: 0 ok; 1 bad args/inputs; 2 produced but some rows need review.
"""
from __future__ import annotations
import csv, gzip, re, sys
from datetime import date
from fractions import Fraction
from pathlib import Path

DEFAULT_CONFIG = "Data/settings/krc_gnucash_config.yaml"


# ---------- config ----------
def load_config(explicit: str | None):
    import yaml
    path = Path(explicit) if explicit else Path(DEFAULT_CONFIG)
    defaults = {
        "long_term_threshold_months": 12,
        "accounts": {
            "broker": "Assets:Current Assets:Brokers:KR Choksey",
            "slbs_income": "Income:Income from SLBS",
            "ltcg": "Income:Long Term Capital Gain",
            "stcg": "Income:Short Term Capital Gain",
        },
        "currency": "INR",
        "security_aliases": {},
    }
    if path.is_file():
        try:
            user = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
            for k, v in user.items():
                if k == "accounts" and isinstance(v, dict):
                    defaults["accounts"].update(v)
                else:
                    defaults[k] = v
        except Exception as e:
            print(f"WARNING: could not read config {path}: {e}; using defaults.",
                  file=sys.stderr)
    else:
        # write a template so the user has something to edit next time
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(yaml.safe_dump(defaults, sort_keys=False), encoding="utf-8")
            print(f"NOTE: wrote a default config template to {path}")
        except Exception:
            pass
    return defaults, str(path)


# ---------- dates ----------
_MONTHS = {m: i for i, m in enumerate(
    ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"], 1)}

def parse_date(s):
    if not s:
        return None
    s = str(s).strip()
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)      # DD/MM/YYYY (Bills)
    if m:
        d, mo, y = map(int, m.groups()); return date(y, mo, d)
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)           # YYYY-MM-DD (GnuCash)
    if m:
        y, mo, d = map(int, m.groups()); return date(y, mo, d)
    m = re.match(r"([A-Za-z]{3})\s+(\d{1,2})\s+(\d{4})", s)  # "Dec 7 2023" (old CN)
    if m:
        mo = _MONTHS.get(m.group(1).lower()); 
        if mo:
            return date(int(m.group(3)), mo, int(m.group(2)))
    return None

def shift_months_back(d: date, n: int) -> date:
    total = (d.year * 12 + (d.month - 1)) - n
    y, mo = divmod(total, 12)
    mo += 1
    day = min(d.day, [31,29 if y%4==0 and (y%100!=0 or y%400==0) else 28,31,30,31,30,31,31,30,31,30,31][mo-1])
    return date(y, mo, day)


# ---------- gnucash book ----------
# GnuCash "special account types" (KVP slots) are NOT valid posting targets and
# must never be offered as a security-match candidate: a placeholder is a header
# account, a hidden account is a delisted/retired holding. These are the boolean
# 'true' slots; 'opening-balance' is an 'equity-type' string slot.
#
# Self-contained mirror of agents.gnucash_accounts.BOOL_FLAG_KEYS (this script
# runs as a stand-alone subprocess). tests/test_gnucash_accounts.py guards the
# two lists against drift.
SPECIAL_BOOL_FLAG_KEYS = ("placeholder", "hidden", "tax-related",
                          "auto-interest-transfer")
_SPECIAL_TRUE_VALUES = ("true", "t", "1", "yes", "y")


def _is_special(account_xml):
    """True if a <gnc:account> XML block carries any special-type flag slot."""
    for mm in re.finditer(
        r"<slot:key>(.*?)</slot:key>\s*<slot:value[^>]*>(.*?)</slot:value>",
        account_xml, re.S,
    ):
        key = mm.group(1).strip()
        val = mm.group(2).strip().lower()
        if key in SPECIAL_BOOL_FLAG_KEYS and val in _SPECIAL_TRUE_VALUES:
            return True
        if key == "equity-type" and val == "opening-balance":
            return True
    return False


def load_book(gnucash_path):
    raw = Path(gnucash_path).read_bytes()
    xml = (gzip.decompress(raw) if raw[:2] == b"\x1f\x8b" else raw).decode("utf-8", "replace")
    acc = {}
    for a in re.findall(r"<gnc:account[^>]*>.*?</gnc:account>", xml, re.S):
        gid = (re.search(r"<act:id[^>]*>(.*?)</act:id>", a) or [None, ""])[1]
        name = (re.search(r"<act:name>(.*?)</act:name>", a) or [None, ""])[1]
        typ = (re.search(r"<act:type>(.*?)</act:type>", a) or [None, ""])[1]
        par = (re.search(r"<act:parent[^>]*>(.*?)</act:parent>", a) or [None, None])[1]
        acc[gid] = {"name": _unescape(name), "typ": typ, "par": par,
                    "special": _is_special(a)}

    def path(g):
        parts = []
        while g and g in acc and acc[g]["typ"] != "ROOT":
            parts.append(acc[g]["name"]); g = acc[g]["par"]
        return ":".join(reversed(parts))

    paths = {g: path(g) for g in acc}
    stock_guids = {g for g, d in acc.items() if d["typ"] in ("STOCK", "MUTUAL")}

    # FIFO holdings: replay all txns in date order; per stock account keep
    # remaining lots [date, shares_remaining(Fraction), price(Fraction)].
    txs = re.findall(r"<gnc:transaction[^>]*>.*?</gnc:transaction>", xml, re.S)
    def txdate(t):
        m = re.search(r"<trn:date-posted>.*?<ts:date>(.*?)</ts:date>", t, re.S)
        return m.group(1)[:10] if m else ""
    lots = {g: [] for g in stock_guids}
    for t in sorted(txs, key=txdate):
        d = txdate(t)
        for sp in re.findall(r"<trn:split>.*?</trn:split>", t, re.S):
            ga = (re.search(r"<split:account[^>]*>(.*?)</split:account>", sp) or [None, ""])[1]
            if ga not in stock_guids:
                continue
            val = _frac((re.search(r"<split:value>(.*?)</split:value>", sp) or [None, "0/1"])[1])
            qty = _frac((re.search(r"<split:quantity>(.*?)</split:quantity>", sp) or [None, "0/1"])[1])
            if qty > 0:
                lots[ga].append([parse_date(d), qty, (val / qty if qty else Fraction(0))])
            elif qty < 0:
                need = -qty
                while need > 0 and lots[ga]:
                    lot = lots[ga][0]
                    take = min(lot[1], need); lot[1] -= take; need -= take
                    if lot[1] == 0:
                        lots[ga].pop(0)
    # holdings keyed by full account path
    holdings = {}
    for g, ls in lots.items():
        rem = [l for l in ls if l[1] > 0]
        if rem:
            holdings[paths[g]] = rem
    return acc, paths, stock_guids, holdings

def _frac(s):
    n, d = s.split("/"); return Fraction(int(n), int(d))

def _unescape(s):
    return (s.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
             .replace("&apos;", "'").replace("&quot;", '"'))


# ---------- security name matching ----------
_DROP = {"LTD", "LTD.", "LIMITED", "THE", "LI", "L", "AND", "&", "CO", "COMPANY",
         "INDIA", "(INDIA)", "PVT", "PRIVATE"}

def _tokens(name):
    # strip apostrophes/periods WITHIN words first ("Reddy's"->"REDDYS",
    # "Dr."->"DR") so they don't fragment tokens, then split on other punctuation.
    name = (name or "").upper().replace("'", "").replace("\u2019", "").replace(".", "")
    name = re.sub(r"[^A-Za-z0-9 ]", " ", name)
    return {w for w in name.split() if w and w not in _DROP}

def match_security(bills_name, stock_paths, aliases):
    if bills_name in aliases:
        return aliases[bills_name], "alias"
    bt = _tokens(bills_name)
    if not bt:
        return None, "no-tokens"
    best, score = None, 0.0
    for p in stock_paths:
        leaf = p.split(":")[-1]
        st = _tokens(leaf)
        if not st:
            continue
        j = len(bt & st) / len(bt | st)
        if j > score:
            best, score = p, j
    return (best, f"fuzzy:{score:.2f}") if score >= 0.5 else (None, f"fuzzy:{score:.2f}")


# ---------- bills ----------
def load_bills(xlsx):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Bills"] if "Bills" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = {h: i for i, h in enumerate(rows[0])}
    def col(r, name):
        i = hdr.get(name)
        return r[i] if i is not None and i < len(r) else None
    bills = []
    for r in rows[1:]:
        if not any(r):
            continue
        bills.append({
            "cn_no": col(r, "CN No"),
            "type": col(r, "Type"),
            "date": col(r, "Date"),
            "settlement": col(r, "Settlement No"),
            "security": col(r, "Security"),
            "quantity": col(r, "Quantity"),
            "net": col(r, "Net/Bill Amount"),
            "direction": col(r, "Direction"),
        })
    return bills


# ---------- entry building ----------
def f2(x):
    return round(float(x), 2)

def build_entries(bills, cfg, paths, holdings):
    acct = cfg["accounts"]
    stock_paths = [p for g, p in paths.items()]  # all paths; matcher filters by leaf tokens
    # restrict to STOCK/MUTUAL leaves for matching
    stock_only = [p for p in holdings.keys()] or stock_paths
    # better: pass actual stock account paths
    purchases, sales, slbm, review = [], [], [], []
    thr = int(cfg.get("long_term_threshold_months", 12))

    # process chronologically so an in-run purchase feeds a later sale's FIFO
    def bkey(b):
        d = parse_date(b["date"]); return d or date.min
    for b in sorted(bills, key=bkey):
        net = b["net"]
        if net is None:
            review.append((b, "no net amount")); continue
        net = f2(net)
        dt = b["date"]
        if b["type"] == "SLBM":
            slbm.append({
                "id": f"KRC-L{b['settlement'] or b['cn_no']}",
                "date": dt, "desc": f"SLBM {b['security']} (CN {b['cn_no']}, settle {b['settlement']})",
                "splits": [
                    (acct["broker"], net, None),
                    (acct["slbs_income"], -net, None),
                ],
            })
            continue
        # TRADE -> Purchase or Sale by direction
        sec_path, how = match_security(b["security"], _stock_paths, cfg.get("security_aliases") or {})
        if sec_path is None:
            review.append((b, f"no security account match ({how}) for {b['security']!r}")); continue
        qty = b["quantity"]
        is_sale = str(b["direction"] or "").startswith("Receiv")
        if not is_sale:
            # Purchase
            purchases.append({
                "id": f"KRC-P{b['cn_no']}",
                "date": dt, "desc": f"Buy {b['security']} (CN {b['cn_no']})",
                "splits": [
                    (sec_path, net, qty),
                    (acct["broker"], -net, None),
                ],
            })
            # add lot to holdings for later sales in this run
            q = Fraction(str(qty)) if qty else Fraction(0)
            if q > 0:
                holdings.setdefault(sec_path, []).append([parse_date(dt), q, Fraction(str(net)) / q])
        else:
            # Sale — FIFO cost basis + LTCG/STCG split
            q = abs(Fraction(str(qty))) if qty else Fraction(0)
            lots = holdings.get(sec_path, [])
            avail = sum(l[1] for l in lots)
            if q == 0 or avail < q:
                review.append((b, f"insufficient FIFO lots for sale of {b['security']} "
                                  f"(need {float(q)}, have {float(avail)})")); continue
            sale_dt = parse_date(dt)
            cutoff = shift_months_back(sale_dt, thr) if sale_dt else None
            proceeds_per_share = Fraction(str(net)) / q
            need = q; cost = Fraction(0); lt = Fraction(0); st = Fraction(0)
            while need > 0 and lots:
                lot = lots[0]
                take = min(lot[1], need)
                lot_cost = take * lot[2]
                lot_proceeds = take * proceeds_per_share
                gain = lot_proceeds - lot_cost
                cost += lot_cost
                if cutoff and lot[0] and lot[0] < cutoff:
                    lt += gain
                else:
                    st += gain
                lot[1] -= take; need -= take
                if lot[1] == 0:
                    lots.pop(0)
            splits = [
                (acct["broker"], net, None),                 # Dr proceeds
                (sec_path, -f2(cost), -abs(float(qty))),     # Cr cost basis, -shares
            ]
            if round(float(lt), 2) != 0:
                splits.append((acct["ltcg"], -f2(lt), None))
            if round(float(st), 2) != 0:
                splits.append((acct["stcg"], -f2(st), None))
            # rounding guard: force exact zero-sum onto the last gain split
            drift = round(sum(s[1] for s in splits), 2)
            if drift and len(splits) > 2:
                a, amt, sh = splits[-1]
                splits[-1] = (a, round(amt - drift, 2), sh)
            sales.append({
                "id": f"KRC-S{b['cn_no']}",
                "date": dt,
                "desc": f"Sell {b['security']} (CN {b['cn_no']})",
                "splits": splits,
            })
    return purchases, slbm, sales, review


# ---------- csv ----------
def _isod(val):
    d = parse_date(val)
    return d.isoformat() if d else str(val)


def write_csv(entries, out_path, currency):
    # Stock entries (Purchase/Sale) carry share quantities and need the
    # Value/Amount/Price columns. SLBM is currency-only (no security account)
    # and uses the simple signed Amount column (same shape as the 26AS journal).
    has_stock = any(sh not in (None, "") for e in entries for (_a, _m, sh) in e["splits"])
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if has_stock:
            # Value = money (INR, sums to zero). Amount = account-commodity
            # units (SHARES for stock rows, INR for cash rows) -> exact share
            # count, no division. Price (per-share) is a FALLBACK: map EITHER
            # Value+Amount (recommended) OR Value+Price, not all three.
            w.writerow(["Date", "Transaction ID", "Number", "Description",
                        "Account", "Value", "Amount", "Price", "Currency"])
            for e in entries:
                dstr = _isod(e["date"])
                for acc, amt, sh in e["splits"]:
                    if sh not in (None, ""):                 # stock row
                        units = _shares(sh)
                        shf = abs(float(sh))
                        price = f"{abs(amt) / shf:.6f}" if shf else ""
                    else:                                    # cash/income/gain row
                        units, price = f"{amt:.2f}", ""
                    w.writerow([dstr, e["id"], e["id"], e["desc"], acc,
                                f"{amt:.2f}", units, price, currency])
        else:
            # SLBM: currency-only -> simple signed Amount (Dr +, Cr -, sums to 0).
            w.writerow(["Date", "Transaction ID", "Number", "Description",
                        "Account", "Amount", "Currency"])
            for e in entries:
                dstr = _isod(e["date"])
                for acc, amt, sh in e["splits"]:
                    w.writerow([dstr, e["id"], e["id"], e["desc"], acc,
                                f"{amt:.2f}", currency])


def _shares(sh):
    v = float(sh)
    return f"{v:.4f}".rstrip("0").rstrip(".") if v != int(v) else str(int(v))


# ---------- main ----------
def main():
    if len(sys.argv) < 4:
        print("Usage: build_krc_gnucash.py <bills_xlsx> <gnucash_path> <out_dir> [config_yaml]",
              file=sys.stderr)
        return 1
    bills_xlsx, gnucash_path, out_dir = sys.argv[1:4]
    config_yaml = sys.argv[4] if len(sys.argv) > 4 else None
    for pth, label in [(bills_xlsx, "Bills xlsx"), (gnucash_path, "GnuCash file")]:
        if not Path(pth).is_file():
            print(f"ERROR: {label} not found: {pth}", file=sys.stderr); return 1
    out = Path(out_dir); out.mkdir(parents=True, exist_ok=True)

    cfg, cfg_path = load_config(config_yaml)
    acc, paths, stock_guids, holdings = load_book(gnucash_path)
    global _stock_paths
    # Match candidates exclude placeholder/hidden STOCK/MUTUAL accounts (not
    # valid post targets). FIFO holdings (stock_guids) are left whole so a
    # since-hidden holding's prior lots still feed cost-basis calculations.
    _stock_paths = [paths[g] for g in stock_guids if not acc[g]["special"]]
    bills = load_bills(bills_xlsx)

    purchases, slbm, sales, review = build_entries(bills, cfg, paths, holdings)

    cur = cfg.get("currency", "INR")
    # Only write files that actually contain entries (no empty 0-txn CSVs).
    written = []
    for _name, _entries in [("Purchase", purchases), ("SLBM", slbm), ("Sale", sales)]:
        if _entries:
            write_csv(_entries, out / f"{_name}.csv", cur)
            written.append(f"{_name}.csv")
    if review:
        with open(out / "Review.csv", "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f); w.writerow(["CN No", "Type", "Security", "Net", "Reason"])
            for b, why in review:
                w.writerow([b["cn_no"], b["type"], b["security"], b["net"], why])

    # verification: every transaction's Amount column sums to zero
    bad = []
    for label, entries in [("Purchase", purchases), ("SLBM", slbm), ("Sale", sales)]:
        for e in entries:
            tot = round(sum(s[1] for s in e["splits"]), 2)
            if abs(tot) >= 0.01:
                bad.append(f"{label} {e['id']} sums to {tot}")

    print(f"Config: {cfg_path} (long-term threshold {cfg['long_term_threshold_months']} months)")
    print(f"Bills read: {len(bills)}")
    def _ln(name, entries):
        return (f"{name} entries: {len(entries)} -> {name}.csv" if entries
                else f"{name} entries: 0 (no file written)")
    print(_ln("Purchase", purchases))
    print(_ln("SLBM", slbm))
    print(_ln("Sale", sales))
    print(f"Files written: {', '.join(written) if written else '(none)'}")
    print(f"Needs review:     {len(review)}" + (" -> Review.csv" if review else ""))
    for b, why in review:
        print(f"  REVIEW: CN {b['cn_no']} ({b['type']}) {b['security']}: {why}")
    if bad:
        print("BALANCE ERRORS:")
        for m in bad:
            print("  " + m)
    else:
        print("All transactions balance (Amount splits sum to zero).")
    return 0 if (not review and not bad) else 2


if __name__ == "__main__":
    raise SystemExit(main())
