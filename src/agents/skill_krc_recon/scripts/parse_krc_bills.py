"""
parse_krc_bills.py — Part II of the KR Choksey broker-import skill: reconcile
contract-note bills against the Part I "Simplified Ledger" workbook.

Reads:
  * a folder of KR Choksey contract notes (PAN-protected PDFs) — SLBM
    confirmation memos and equity trade notes; and
  * the Part I output xlsx (the "Simplified Ledger" sheet, plus the
    "References" sheet carrying recovered bill-settlement anchors and
    NEFT/RTGS payout UTRs).

Writes one workbook with sheets [Bills, Trade Lines, Reconciliation]:
  * Bills        — one row per contract note (CN no, date, settlement,
                   security, charge breakdown, net/bill amount).
  * Trade Lines  — per-security legs (borrow/return for SLBM, buy/sell for
                   trades) for Part III's GnuCash canonical.
  * Reconciliation — every Simplified-Ledger row tagged & matched: bills are
                   matched to a ledger movement by AMOUNT (primary key) and
                   corroborated by SETTLEMENT NUMBER via the References sheet;
                   non-bill rows reuse the Part I Tag (Opening / Demat Charge /
                   Bank Pay-In / Bank Pay-Out). Bank rows are flagged
                   Unreconciled until the GnuCash bank import exists. A
                   'Settlement Movement' row with no matching bill is flagged
                   REVIEW (a trade whose contract note is missing).

Why amount is primary: in this statement the settlement label printed on a
surviving amount row is OFFSET (it belongs to a different settlement); the
true bill->settlement link lives only in the dropped BILL-ENTRY anchor rows,
now preserved on the Part I References sheet.

Usage:
    python parse_krc_bills.py <cn_dir> <ledger_xlsx> <password> <out_xlsx>

Exit codes:
    0  success — all bills matched, no review rows
    1  bad arguments / inputs not found / decryption failed
    2  parsed & written, but some bills unmatched or rows need review
"""
from __future__ import annotations
import re, sys, glob, subprocess
from pathlib import Path


# ---------- helpers ----------
def num(s):
    if s is None:
        return None
    s = str(s).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def last_num_on_line(line):
    nums = re.findall(r"-?\d[\d,]*\.?\d*", line)
    return num(nums[-1]) if nums else None


# Broker's own legal name — never a traded security. Guards the anchored
# security fallback below from grabbing the letterhead.
_BROKER_NAME_TOKENS = ("KRCHOKSEY", "PRIVATE LIMITED")


_ANCHOR_RE = re.compile(r"([A-Z][A-Z&. ]*?(?:LTD|LIMITED))\s+([BS])\s+(\d+)\s+([\d.]+)")


def anchored_trade_lines(t):
    """Recover trade legs — security, side, quantity, rate — from each trade
    line's economic shape: the company name printed immediately before the
    Buy/Sell flag, quantity and rate (e.g. '... RAMCO CEMENTS LIMITED S 250
    993.0000'). This survives the text-extraction garbling seen on forced
    square-off notes, whose rotated 'Remark' column collides the 16-digit order
    number with the trade time and wraps the ISIN-anchored name across lines — so
    both the equity-header parse and the order-number trade-line parse miss it.
    Excludes the broker's own letterhead name (which never precedes a Buy/Sell +
    qty + rate)."""
    legs = []
    for m in _ANCHOR_RE.finditer(t):
        name = m.group(1).strip()
        if any(tok in name for tok in _BROKER_NAME_TOKENS):
            continue
        legs.append({
            "security": name,
            "bs": "BUY" if m.group(2) == "B" else "SELL",
            "quantity": num(m.group(3)),
            "rate": num(m.group(4)),
        })
    return legs


def anchored_securities(t):
    """Distinct traded-security names recovered from the trade lines' economic
    shape (see anchored_trade_lines). Kept as the security-name fallback."""
    return sorted({leg["security"] for leg in anchored_trade_lines(t)})


def _anchored_line(cn_no, leg):
    """Build a standard trade-line dict from a recovered anchored leg. Used when
    the structured trade-line regex found nothing (square-off / garbled notes),
    so the bill still carries a share quantity for Part III's FIFO booking."""
    return {
        "cn_no": cn_no, "type": "TRADE", "security": leg["security"],
        "series": None, "reversal_date": None,
        "quantity": leg["quantity"], "rate": leg["rate"],
        "net_rate": None, "proc_charge": None,
        "net_amount": None, "bs": leg["bs"],
    }


def decrypt(pdf, pw, out):
    r = subprocess.run(["qpdf", f"--password={pw}", "--decrypt", pdf, out],
                       capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(f"qpdf failed on {pdf}: {r.stderr.strip()}")


def text_of(pdf):
    import pdfplumber
    with pdfplumber.open(pdf) as p:
        return "\n".join(pg.extract_text() or "" for pg in p.pages)


# ---------- contract-note parsing ----------
def parse_slbm(t, fname):
    def g(pat, grp=1):
        m = re.search(pat, t)
        return m.group(grp).strip() if m else None

    def lineval(label):
        for ln in t.splitlines():
            if label in ln:
                return last_num_on_line(ln)
        return None

    securities = sorted({m.group(1) for m in re.finditer(
        r"\d{2}:\d{2}:\d{2}\s+([A-Z]{3,})\s+\d", t)})
    rec = {
        "type": "SLBM", "file": fname,
        "cn_no": g(r"Confirmation Memo No\.\s*:\s*(\d+)"),
        "date": g(r"Transaction Date\s*:\s*([\d/]+)"),
        "settlement": g(r"Settlement Number\s*:\s*(\d+)"),
        "payinout": g(r"Pay In-Pay Out Date\s*:\s*(\d{2}/\d{2}/\d{4})"),
        "gst_invoice": g(r"GST Invoice No\s*:\s*(\S+)"),
        "security": ", ".join(securities) if securities else None,
        "lending_fees": lineval("Total Lending fees Rs."),
        "processing": lineval("Total Processing Charges Rs."),
        "cgst": lineval("CGST ("), "sgst": lineval("SGST ("),
        "igst": lineval("IGST ("),
        "stt": None, "stamp": None, "brokerage": None,
        "net_amount": num(g(r"Total Bill Amount Rs\.?\s*:\s*([\d.,]+)")),
        "direction": "Receivable (Cr)",
    }
    lines = []
    for m in re.finditer(r"(\d{12})(\d{2}:\d{2}:\d{2})\s+(\d+)\s+(\d{2}:\d{2}:\d{2})\s+"
                         r"([A-Z]{3,})\s+(\d{1,2}[A-Za-z]{3}\d{4})\s+(\w+)\s+(-?\d+)"
                         r"(?:\s+([\d.]+)\s+([\d.]+)\s+([\d.]+))?", t):
        lines.append({
            "cn_no": rec["cn_no"], "type": "SLBM", "security": m.group(5),
            "series": m.group(7), "reversal_date": m.group(6),
            "quantity": num(m.group(8)), "rate": num(m.group(9)),
            "net_rate": num(m.group(10)), "proc_charge": num(m.group(11)),
            "net_amount": None, "bs": None,
        })
    return rec, lines


def parse_trade(t, fname):
    def g(pat, grp=1):
        m = re.search(pat, t)
        return m.group(grp).strip() if m else None

    def oblig(label):
        for ln in t.splitlines():
            if label in ln:
                m = re.findall(r"([\d,]+\.\d+)\s*(DR|CR)", ln)
                if m:
                    v, d = m[-1]
                    return num(v), d
        return None, None

    net_v, net_d = oblig("Net Amount Receivable/Payable By Client")
    brk, _ = oblig("Taxable Value Of Supply (Brokerage)")
    cgst, _ = oblig("CGST*")
    sgst, _ = oblig("SGST*")
    stamp, _ = oblig("Stamp Duty")
    stt, _ = oblig("Securities Transactions Tax")
    secs = sorted({m.group(1).strip() for m in re.finditer(
        r"[A-Z]{2}\w{9,10}\s+([A-Z][A-Z .]+?(?:LTD|LIMITED)\.?)", t)})
    rec = {
        "type": "TRADE", "file": fname,
        "cn_no": g(r"CONTRACT NOTE NO\s*:\s*(\d+)"),
        "date": g(r"Trade Date\s*:\s*([\d/]+)"),
        "settlement": g(r"Settlement No\s*(\d+)"),
        "payinout": g(r"Settlement\s+(\d{2}/\d{2}/\d{4})"),
        "gst_invoice": g(r"Invoice Reference Number\s*:\s*(\S+)"),
        "security": ", ".join(secs) if secs else None,
        "lending_fees": None, "processing": None,
        "brokerage": brk, "cgst": cgst, "sgst": sgst, "igst": None,
        "stt": stt, "stamp": stamp,
        "net_amount": net_v,
        "direction": "Payable (Dr)" if net_d == "DR" else "Receivable (Cr)",
    }
    lines = []
    for m in re.finditer(
        r"(\d{16})\s+(\d{2}:\d{2}:\d{2})\s+(\d+)\s+(\d{2}:\d{2}:\d{2})\s+"
        r"(.+?)\s+([BS])\s+(\d+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([DC])", t):
        lines.append({
            "cn_no": rec["cn_no"], "type": "TRADE", "security": m.group(5).strip(),
            "series": None, "reversal_date": None,
            "quantity": num(m.group(7)), "rate": num(m.group(8)),
            "net_rate": num(m.group(9)), "proc_charge": None,
            "net_amount": num(m.group(10)), "bs": "BUY" if m.group(6) == "B" else "SELL",
        })
    # Square-off / garbled notes: the structured trade-line regex above finds
    # nothing (the order number collides with the trade time). Recover the legs
    # from their economic shape so the bill still carries a share quantity —
    # Part III (GnuCash Import) needs it to book a FIFO sale.
    if not lines:
        lines = [_anchored_line(rec["cn_no"], leg) for leg in anchored_trade_lines(t)]
    # Bill-level security: prefer the trade-annexure line items (robust) when the
    # equity-segment header parse missed the name (it wraps across PDF lines).
    if (not rec["security"]) and lines:
        line_secs = sorted({ln["security"] for ln in lines if ln["security"]})
        if line_secs:
            rec["security"] = ", ".join(line_secs)
    # Last resort: recover the name from the trade line's economic shape (name
    # before Buy/Sell + qty + rate). Handles square-off notes where the two
    # parses above both miss (see anchored_securities).
    if not rec["security"]:
        anc = anchored_securities(t)
        if anc:
            rec["security"] = ", ".join(anc)
    return rec, lines


def parse_trade_old(t, fname):
    """
    Older KRChoksey contract-note layout: 2-page, inline trade-line table
    (Order/Trade/Security/BUY-SELL/Qty/.../Net Total), ISIN on its own line,
    and 'Net Amount Receivable/Payable by Client'. The Net Total column is
    already signed (BUY negative, SELL positive).
    """
    def g(pat, grp=1):
        m = re.search(pat, t)
        return m.group(grp).strip() if m else None

    def lv(label_pat):
        m = re.search(label_pat + r"\s+([\d.,]+)", t)
        return num(m.group(1)) if m else None

    recv = re.search(r"Net Amount (Receivable|Payable) by Client\s+([\d.,]+)", t)
    net_v = num(recv.group(2)) if recv else None
    direction = "Receivable (Cr)" if (recv and recv.group(1) == "Receivable") else "Payable (Dr)"
    rec = {
        "type": "TRADE", "file": fname,
        "cn_no": g(r"CONTRACT NOTE NO\.?\s*:?\s*(\d+)"),
        "date": g(r"TRADE DATE\s+([A-Za-z]{3}\s+\d{1,2}\s+\d{4})"),
        "settlement": g(r"SETTLEMENT NO\.?\s*:?\s*(\d+)"),
        "payinout": g(r"SETTLEMENT DATE\.?\s+(\d{2}/\d{2}/\d{4})"), "gst_invoice": None, "security": None,
        "lending_fees": None, "processing": None, "brokerage": None,
        "cgst": lv(r"CGST\s*\(@?[\d.]+%\)"),
        "sgst": lv(r"SGST\s*\(@?[\d.]+%\)"),
        "igst": None,
        "stt": lv(r"Securities Transaction Tax"),
        "stamp": lv(r"Stamp Duty"),
        "net_amount": net_v,
        "direction": direction,
    }
    lines = []
    for m in re.finditer(
        r"(\d{10,16})\s+(\d{2}:\d{2}:\d{2})\s+(\d+)\s+(\d{2}:\d{2}:\d{2})\s+"
        r"(.+?)\s+(BUY|SELL)\s+(\d+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+(-?[\d.]+)\s+(\S+)", t):
        lines.append({
            "cn_no": rec["cn_no"], "type": "TRADE", "security": m.group(5).strip(),
            "series": None, "reversal_date": None,
            "quantity": num(m.group(7)), "rate": num(m.group(8)),
            "net_rate": num(m.group(10)), "proc_charge": None,
            "net_amount": num(m.group(12)),  # signed Net Total
            "bs": "BUY" if m.group(6) == "BUY" else "SELL",
        })
    # Square-off / garbled notes: recover legs (with quantity) from their
    # economic shape when the structured trade-line regex found nothing.
    if not lines:
        lines = [_anchored_line(rec["cn_no"], leg) for leg in anchored_trade_lines(t)]
    if (not rec["security"]) and lines:
        secs = sorted({ln["security"] for ln in lines if ln["security"]})
        rec["security"] = ", ".join(secs) if secs else None
    if not rec["security"]:
        anc = anchored_securities(t)
        if anc:
            rec["security"] = ", ".join(anc)
    return rec, lines


def parse_cn(pdf, pw):
    scratch = f"/tmp/_cn_{Path(pdf).stem}.pdf"
    decrypt(pdf, pw, scratch)
    t = text_of(scratch)
    try:
        Path(scratch).unlink()
    except OSError:
        pass
    fname = Path(pdf).name
    if "SLB Session" in t or "LENDING & BORROWING" in t:
        return parse_slbm(t, fname)
    if "Net Amount Receivable/Payable By Client" in t:
        return parse_trade(t, fname)
    return parse_trade_old(t, fname)


# ---------- Part I workbook (Simplified Ledger + References) ----------
def load_simplified(xlsx):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Simplified Ledger"] if "Simplified Ledger" in wb.sheetnames else wb.active
    rows = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        r = (tuple(r) + (None,) * 9)[:9]
        rows.append({"date": r[0], "vno": r[1], "particulars": r[2], "chqno": r[3],
                     "debit": num(r[4]), "credit": num(r[5]),
                     "balance": r[6], "drcr": r[7], "tag": r[8]})
    return rows


def load_references(xlsx):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx)
    settlements, utrs = set(), []
    if "References" in wb.sheetnames:
        for r in list(wb["References"].iter_rows(values_only=True))[1:]:
            r = (tuple(r) + (None,) * 5)[:5]
            rtype, date, vno, exch, val = r
            if rtype == "Bill Anchor" and val is not None:
                settlements.add(str(val).strip())
            elif rtype and str(rtype).startswith("Bank Payment"):
                utrs.append({"date": date, "utr": val})
    return settlements, utrs


# ---------- classification + matching ----------
def classify_and_match(led_rows, bills, settlements, utrs):
    by_amt = {}
    for b in bills:
        by_amt.setdefault(round(b["net_amount"], 2), []).append(b)
    utr_dates = {u["date"] for u in utrs}

    out, used = [], set()
    for r in led_rows:
        amt = r["debit"] or r["credit"] or 0.0
        tag = r["tag"] or ""
        cat = bill_cn = bill_setl = match = note = None

        cand = [b for b in by_amt.get(round(amt, 2), []) if b["cn_no"] not in used]
        if cand:
            b = cand[0]
            used.add(b["cn_no"])
            bill_cn, bill_setl = b["cn_no"], b["settlement"]
            cat = "Trade Bill" if b["type"] == "TRADE" else "SLBM Bill"
            setl_ok = str(b["settlement"]) in settlements
            match = ("Matched (amount + settlement confirmed)" if setl_ok
                     else "Matched (amount only)")
            if b["type"] == "TRADE":
                note = "Settled via incoming bank credit; bank leg pending GnuCash import"
                if not setl_ok:
                    note += "; trade booked under a BSE 'M-' bill anchor in the ledger"
        elif "OPENING BALANCE" in (r["particulars"] or "") or tag == "Opening Balance":
            cat, match = "Opening Balance", "No bill expected"
        elif tag == "Demat Charge":
            cat, match = "Demat Charge", "No bill expected (flagged)"
        elif tag in ("Bank Pay-In", "Bank Pay-Out"):
            cat = tag
            match = "Unreconciled (pending GnuCash bank import)"
            for u in utrs:
                if u["date"] == r["date"]:
                    note = f"NEFT/RTGS ref {u['utr']}"
                    break
        elif tag == "Settlement Movement":
            cat = "Settlement Movement"
            match = "REVIEW — settlement row with no matching contract note (bill missing?)"
        else:
            cat = "Unclassified"
            match = "REVIEW — no bill, no Part I tag"

        out.append({**r, "category": cat, "bill_cn": bill_cn,
                    "bill_settlement": bill_setl, "match": match, "note": note})

    unmatched = [b for b in bills if b["cn_no"] not in used]
    return out, unmatched


# ---------- charge apportionment (trade bills) ----------
def apportion_charges(bills, trade_lines):
    """
    Spread each TRADE bill's total levies across its trade lines so the per-line
    balanced nets sum EXACTLY to the bill net. Lumped charge apportioned by
    turnover (|signed gross|); charge is an ADD-ON to a BUY (raises cost) and a
    DEDUCTION from a SELL (lowers proceeds): net = signed_gross - charge. The
    last line absorbs the rounding residual so the total ties exactly.
    signed_gross = -|value| for BUY, +|value| for SELL (works for both layouts).
    Returns [(cn_no, sum_net, bill_net, ties_ok), ...].
    """
    by_cn = {}
    for ln in trade_lines:
        by_cn.setdefault(ln["cn_no"], []).append(ln)
    checks = []
    for b in bills:
        if b["type"] != "TRADE" or b.get("net_amount") is None:
            continue
        lns = [leg for leg in by_cn.get(b["cn_no"], []) if leg.get("net_amount") is not None]
        if not lns:
            continue
        for leg in lns:
            leg["signed_gross"] = round((-1 if leg["bs"] == "BUY" else 1) * abs(leg["net_amount"]), 4)
        bill_net = round(b["net_amount"] * (1 if b["direction"].startswith("Receiv") else -1), 2)
        gross_sum = sum(leg["signed_gross"] for leg in lns)
        total_charge = gross_sum - bill_net
        turnover = sum(abs(leg["signed_gross"]) for leg in lns) or 1.0
        for leg in lns[:-1]:
            c = round(total_charge * abs(leg["signed_gross"]) / turnover, 2)
            leg["apportioned_charge"] = c
            leg["net_balanced"] = round(leg["signed_gross"] - c, 2)
        prev = sum(leg["net_balanced"] for leg in lns[:-1])
        last = lns[-1]
        last["net_balanced"] = round(bill_net - prev, 2)
        last["apportioned_charge"] = round(last["signed_gross"] - last["net_balanced"], 2)
        net_sum = round(sum(leg["net_balanced"] for leg in lns), 2)
        checks.append((b["cn_no"], net_sum, bill_net, abs(net_sum - bill_net) < 0.005))
    return checks


# ---------- workbook ----------
def write_workbook(out_path, bills, trade_lines, recon, unmatched, diag):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    hdr = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="305496")

    def style_header(ws, n):
        for c in ws[1][:n]:
            c.font = hdr
            c.fill = hfill
            c.alignment = Alignment(vertical="center")

    ws = wb.active
    ws.title = "Bills"
    bcols = ["CN No", "Type", "File", "Date", "Settlement No", "Pay-In/Out Date",
             "Security", "Quantity", "Direction", "Brokerage/Lending Fees", "Processing",
             "CGST", "SGST", "IGST", "STT", "Stamp", "Net/Bill Amount"]
    ws.append(bcols)
    sec_col = bcols.index("Security") + 1
    missing_fill = PatternFill("solid", fgColor="FFF2CC")  # amber: fill me in
    for b in sorted(bills, key=lambda x: x["date"] or ""):
        ws.append([b["cn_no"], b["type"], b["file"], b["date"], b["settlement"],
                   b.get("payinout"), b["security"], b.get("quantity"), b["direction"],
                   b.get("brokerage") or b.get("lending_fees"), b.get("processing"),
                   b.get("cgst"), b.get("sgst"), b.get("igst"), b.get("stt"),
                   b.get("stamp"), b["net_amount"]])
        # Highlight a blank Security cell so the user can spot & fill it by hand.
        if not b.get("security"):
            ws.cell(ws.max_row, sec_col).fill = missing_fill
    style_header(ws, len(bcols))
    for i, w in enumerate([9, 7, 26, 11, 13, 20, 16, 10, 16, 18, 12, 9, 9, 7, 9, 8, 15], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    ws2 = wb.create_sheet("Trade Lines")
    tcols = ["CN No", "Type", "B/S", "Security", "Series", "Reversal Date",
             "Quantity", "Rate", "Net Rate", "Proc Charge", "Net Amount",
             "Signed Gross", "Apportioned Charge", "Net (balanced)"]
    ws2.append(tcols)
    for ln in trade_lines:
        ws2.append([ln["cn_no"], ln["type"], ln["bs"], ln["security"], ln["series"],
                    ln["reversal_date"], ln["quantity"], ln["rate"], ln["net_rate"],
                    ln["proc_charge"], ln["net_amount"],
                    ln.get("signed_gross"), ln.get("apportioned_charge"),
                    ln.get("net_balanced")])
    style_header(ws2, len(tcols))
    for i, w in enumerate([9, 7, 6, 22, 8, 13, 11, 10, 10, 11, 13, 14, 16, 14], 1):
        ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w

    ws3 = wb.create_sheet("Reconciliation")
    fills = {"SLBM Bill": "E2EFDA", "Trade Bill": "E2EFDA", "Demat Charge": "FFF2CC",
             "Bank Pay-In": "DDEBF7", "Bank Pay-Out": "DDEBF7",
             "Opening Balance": "F2F2F2", "Settlement Movement": "FFC7CE",
             "Unclassified": "FFC7CE"}
    ws3.append(["KR Choksey — Bills vs Ledger Reconciliation"])
    ws3["A1"].font = Font(bold=True, size=13)
    ws3.append(["Bills parsed", diag["n_bills"], "", "Matched two ways", diag["matched_both"]])
    ws3.append(["Ledger rows", diag["n_rows"], "", "Unmatched bills", len(unmatched)])
    ws3.append(["Bills total", round(diag["bills_total"], 2), "",
                "Matched ledger total", round(diag["matched_total"], 2)])
    ws3.append([])
    head_row = ws3.max_row + 1
    rcols = ["Date", "V.No", "Particulars", "ChqNo", "Debit", "Credit", "Balance",
             "Dr/Cr", "Tag (Part I)", "Category", "Bill CN", "Bill Settlement",
             "Match Status", "Notes"]
    ws3.append(rcols)
    for c in ws3[head_row][:len(rcols)]:
        c.font = hdr
        c.fill = hfill
    for r in recon:
        ws3.append([r["date"], r["vno"], r["particulars"], r["chqno"], r["debit"],
                    r["credit"], r["balance"], r["drcr"], r["tag"], r["category"],
                    r["bill_cn"], r["bill_settlement"], r["match"], r["note"]])
        f = fills.get(r["category"])
        if f:
            for c in ws3[ws3.max_row][:len(rcols)]:
                c.fill = PatternFill("solid", fgColor=f)
    for i, w in enumerate([11, 13, 42, 12, 12, 12, 11, 6, 16, 16, 9, 13, 40, 30], 1):
        ws3.column_dimensions[ws3.cell(head_row, i).column_letter].width = w
    ws3.freeze_panes = ws3.cell(head_row + 1, 1)

    if unmatched:
        ws4 = wb.create_sheet("Exceptions")
        ws4.append(["Unmatched bills (no ledger row found)"])
        ws4.append(["CN No", "Type", "Amount", "Settlement"])
        for b in unmatched:
            ws4.append([b["cn_no"], b["type"], b["net_amount"], b["settlement"]])

    wb.save(out_path)


# ---------- main ----------
def main() -> int:
    if len(sys.argv) < 5:
        print("Usage: parse_krc_bills.py <cn_dir> <ledger_xlsx> <password> <out_xlsx>",
              file=sys.stderr)
        return 1
    cn_dir, ledger_xlsx, pw, out_xlsx = sys.argv[1:5]

    if not Path(cn_dir).is_dir():
        print(f"ERROR: contract-note folder not found: {cn_dir}", file=sys.stderr)
        return 1
    if not Path(ledger_xlsx).is_file():
        print(f"ERROR: Part I ledger workbook not found: {ledger_xlsx}", file=sys.stderr)
        return 1

    pdfs = sorted(p for p in glob.glob(str(Path(cn_dir) / "*"))
                  if p.lower().endswith(".pdf"))
    if not pdfs:
        print(f"ERROR: no PDF contract notes in {cn_dir}", file=sys.stderr)
        return 1

    bills, trade_lines, skipped = [], [], []
    for pdf in pdfs:
        name = Path(pdf).name
        try:
            rec, lines = parse_cn(pdf, pw)
        except Exception as e:  # noqa: BLE001 — one bad note must not sink the batch
            # Decryption or parse failure on a single note: record it and carry
            # on so the workbook is still produced for every note that did parse.
            skipped.append((name, str(e)))
            print(f"WARNING: could not parse {name}: {e}", file=sys.stderr)
            continue
        if rec.get("net_amount") is None:
            skipped.append((name, "no bill amount found"))
            print(f"WARNING: could not read a bill amount from {name}; skipped.",
                  file=sys.stderr)
            continue
        bills.append(rec)
        trade_lines.extend(lines)

    if not bills:
        print("ERROR: no usable bills parsed from the contract notes.", file=sys.stderr)
        return 1

    # Bill-level quantity (trading volume): net signed for trades
    # (BUY +, SELL -); gross lent volume for SLBM (borrow+return nets to 0).
    for b in bills:
        lns = [leg for leg in trade_lines
               if leg["cn_no"] == b["cn_no"] and leg.get("quantity") is not None]
        if not lns:
            b["quantity"] = None
        elif b["type"] == "SLBM":
            # lent volume = one-side total (borrow == return); use the larger
            # side so borrow-only and round-trip memos both report correctly.
            pos = sum(leg["quantity"] for leg in lns if leg["quantity"] > 0)
            neg = sum(-leg["quantity"] for leg in lns if leg["quantity"] < 0)
            b["quantity"] = max(pos, neg) or None
        else:
            b["quantity"] = sum(
                abs(leg["quantity"]) if leg.get("bs") == "BUY" else -abs(leg["quantity"])
                for leg in lns)

    settlements, utrs = load_references(ledger_xlsx)
    led_rows = load_simplified(ledger_xlsx)
    recon, unmatched = classify_and_match(led_rows, bills, settlements, utrs)
    apport_checks = apportion_charges(bills, trade_lines)

    matched = [r for r in recon if r["bill_cn"]]
    diag = {
        "n_bills": len(bills), "n_rows": len(led_rows),
        "matched_both": sum(1 for r in matched if "settlement confirmed" in (r["match"] or "")),
        "bills_total": sum(b["net_amount"] for b in bills),
        "matched_total": sum((r["debit"] or r["credit"] or 0) for r in matched),
    }
    write_workbook(out_xlsx, bills, trade_lines, recon, unmatched, diag)

    review = [r for r in recon if str(r["match"]).startswith("REVIEW")]
    missing_sec = [b for b in bills if not b.get("security")]
    print(f"Bills parsed: {len(bills)}  | trade lines: {len(trade_lines)}")
    print(f"Ledger rows: {len(led_rows)}")
    print(f"Matched bills: {len(matched)}/{len(bills)}  "
          f"(amount + settlement confirmed: {diag['matched_both']})")
    print(f"Unmatched bills: {len(unmatched)}")
    print(f"Bills total: {diag['bills_total']:.2f}  | matched ledger total: "
          f"{diag['matched_total']:.2f}  | "
          f"{'OK' if abs(diag['bills_total'] - diag['matched_total']) < 0.02 else 'MISMATCH'}")
    if missing_sec:
        cns = ", ".join(str(b["cn_no"]) for b in missing_sec)
        print(f"Bills with unparsed security (blank in the workbook — fill the "
              f"Security column manually): {len(missing_sec)} — CN {cns}")
    if skipped:
        print(f"Contract notes skipped (not in the workbook): {len(skipped)}")
        for nm, why in skipped:
            print(f"  SKIPPED: {nm} — {why}")
    if unmatched:
        for b in unmatched:
            print(f"  UNMATCHED BILL: CN {b['cn_no']} {b['type']} {b['net_amount']}")
    if review:
        for r in review:
            print(f"  REVIEW: {r['date']} D={r['debit']} C={r['credit']} {r['category']}")
    if apport_checks:
        bad = [c for c in apport_checks if not c[3]]
        print(f"Trade-bill apportionment: {len(apport_checks)} bill(s); "
              f"{len(apport_checks) - len(bad)} balanced exactly to bill net"
              + (f"; {len(bad)} OFF" if bad else ""))
        for cn, snet, bnet, ok in apport_checks:
            if not ok:
                print(f"  APPORTION OFF: CN {cn} lines sum {snet} != bill net {bnet}")
    print(f"Wrote reconciliation workbook to {out_xlsx}")

    # The workbook is always written above; a non-zero code only signals that
    # some rows need a human look (unmatched/review), a note was skipped, or a
    # security couldn't be parsed and needs filling in by hand.
    return 0 if not (unmatched or review or missing_sec or skipped) else 2


if __name__ == "__main__":
    raise SystemExit(main())
