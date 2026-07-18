#!/usr/bin/env python3
"""
GnuCash Import Pipeline
End-to-end: raw bank statement → GnuCash-ready mapped CSV.

Chain:
  1. Bank parse → canonical. Every dedicated bank (DEDICATED_BANKS, derived
                  from agents.banks.discover() — currently ICICI / Bank of
                  Baroda / HSBC / HDFC / Kotak) is dispatched purely through
                  the agents.banks registry: BankSkill.parse() returns
                  canonical rows in memory, and this module writes the
                  canonical CSV + sidecar once via the shared canonical_io
                  tail — no bank writes its own CSV. "Other Bank (CSV)" is
                  the one path that still uses LLM-assisted column
                  normalisation.
  2. Account mapping   (skill_gnucash_account_mapper)

Public surface:
    run() — PA Skills UI entry point.
"""

import csv
import gzip
import json
import logging
import re
import sys
import tempfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

from agents.balance_utils import (
    verify_running_balance,
    extract_opening_closing,
    format_balance_summary,
    _safe_float,
)
from agents.banks import discover as discover_banks, load_bank_skill
from agents.canonical_io import (
    read_sidecar as _ci_read_sidecar,
    write_canonical_csv,
    write_sidecar,
)
from agents.skill_gnucash_reconciler.agent import (
    parse_gnucash_for_reconcile,
    reconcile,
    detect_contra_entries,
)

log = logging.getLogger(__name__)


def _emit_progress(step: int, message: str) -> None:
    """Push a progress event to the UI streaming queue (if active)."""
    try:
        from agents.base_agent import get_progress_queue
        q = get_progress_queue()
        if q is not None:
            q.put({"step": step, "type": "pipeline", "snippet": message})
    except Exception:
        pass  # queue not available — running outside UI

# GnuCash XML namespaces
_NS = {
    'gnc': '{http://www.gnucash.org/XML/gnc}',
    'act': '{http://www.gnucash.org/XML/act}',
    'trn': '{http://www.gnucash.org/XML/trn}',
    'split': '{http://www.gnucash.org/XML/split}',
    'ts': '{http://www.gnucash.org/XML/ts}',
    'slot': '{http://www.gnucash.org/XML/slot}',
    'cmdty': '{http://www.gnucash.org/XML/cmdty}',
}

def _resolve_single_file(path_or_dir: str, extensions: tuple[str, ...]) -> str:
    """If *path_or_dir* is a directory (staged uploads), return the first
    matching file inside it; otherwise return the path unchanged."""
    p = Path(path_or_dir)
    if p.is_dir():
        for ext in extensions:
            matches = sorted(p.glob(f"*{ext}"))
            if matches:
                return str(matches[0])
        # Fall back to any file at all
        children = sorted(p.iterdir())
        if children:
            return str(children[0])
    return path_or_dir


def _read_sidecar(canonical_path: str) -> dict | None:
    """Read the _summary.json sidecar if it exists (shared canonical_io tail)."""
    return _ci_read_sidecar(canonical_path)


def _apply_confirmed_contras(output_path: str, contra_flags: dict) -> int:
    """Re-map the Account of confirmed contras to their counterparty bank.

    A confirmed contra (status == "confirmed", i.e. a reference-matched
    cross-bank transfer) must post against the other bank rather than the
    category the mapper guessed. Possible contras are left untouched. Mutates
    ``contra_flags`` in place to record the mapper's original account
    (``mapped_account``) and what was applied (``applied_account``), and
    rewrites ``output_path`` only when at least one row changed.

    contra_flags keys are 0-based row indices into the output CSV (the mapper
    preserves canonical order 1:1). Returns the number of rows re-mapped.
    """
    confirmed = {
        int(idx): c for idx, c in contra_flags.items()
        if isinstance(c, dict) and c.get("status") == "confirmed"
    }
    if not confirmed:
        return 0

    with open(output_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)

    remapped = 0
    for idx, c in confirmed.items():
        if not (0 <= idx < len(rows)):
            continue
        bank_acct = c.get("contra_account", "") or ""
        if bank_acct.startswith("Root Account:"):
            bank_acct = bank_acct[len("Root Account:"):]
        if not bank_acct:
            continue
        c["mapped_account"] = rows[idx].get("Account", "")
        c["applied_account"] = bank_acct
        rows[idx]["Account"] = bank_acct
        remapped += 1

    if remapped:
        with open(output_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
    return remapped


# Banks with dedicated extraction skills — registry-driven (agents.banks
# discover()) rather than a hardcoded literal, so onboarding a new bank
# (adding its skill.yaml `bank: true` manifest) automatically extends both
# this gating list and the pipeline's `bank` dropdown (ui/tabs/_generic.py's
# "banks" options_from source) without any edit here. See
# 2026-07-18-bank-registry-gating-followup-prompt.md — this list previously
# diverged from the dropdown's options: list (Kotak was added to the
# dropdown in #89 but not here), causing an offer-then-reject bug at the
# SUPPORTED_BANKS guard below.
DEDICATED_BANKS = [b.display_name for b in discover_banks()]
# Banks that go through the generic CSV normalisation path
CSV_BANKS = ["Other Bank (CSV)"]
SUPPORTED_BANKS = DEDICATED_BANKS + CSV_BANKS

# Canonical schema (order matters — GnuCash importer expects this sequence)
CANONICAL_COLS = [
    "Date",
    "Transaction ID",
    "Description",
    "Account",
    "Deposit",
    "Withdrawal",
    "Balance",
    "Currency",
]

# ---------------------------------------------------------------------------
# GnuCash ledger balance extraction
# ---------------------------------------------------------------------------

def _normalize_digits(s: str | None) -> str:
    """Strip everything but digits, e.g. for comparing account numbers that
    may carry spaces/dashes/masking in either the statement or GnuCash."""
    return "".join(c for c in (s or "") if c.isdigit())


def _get_gnucash_account_balance(
    gnucash_file: str, bank_name: str, account_number: str | None = None
) -> dict:
    """
    Parse a .gnucash XML file and find the bank account's ledger balance.

    Candidate accounts are those whose name contains bank_name
    (case-insensitive) under Assets. When ``account_number`` is supplied
    (from statement metadata), it is normalised to digits and matched
    against digits embedded in each candidate's account name (e.g.
    "BOB - 760001001951") — this disambiguates multiple accounts at the
    same bank and is preferred over a bare name match. If no account
    number is given, or none of the candidates' digits match it, falls
    back to the first name match and reports that in "match_warning" so
    callers can surface it (a name-only match is a guess when there are
    multiple accounts for the same bank).

    Returns:
        {
            "found": bool,
            "account_name": str,
            "balance": float,
            "last_txn_date": str or None,  # YYYY-MM-DD
            "match_warning": str or None,
        }
    """
    try:
        with gzip.open(gnucash_file, 'rt', encoding='utf-8') as f:
            tree = ET.parse(f)
    except Exception:
        return {
            "found": False, "account_name": "", "balance": 0.0,
            "last_txn_date": None, "match_warning": None,
        }

    root = tree.getroot()

    # Build account map: id → {name, parent_id, type}
    acc_map = {}
    for acc in root.findall(f'.//{_NS["gnc"]}account'):
        aid = acc.findtext(f'{_NS["act"]}id', '')
        aname = acc.findtext(f'{_NS["act"]}name', '')
        atype = acc.findtext(f'{_NS["act"]}type', '')
        parent_el = acc.find(f'{_NS["act"]}parent')
        parent_id = parent_el.text if parent_el is not None else None
        acc_map[aid] = {"name": aname, "parent_id": parent_id, "type": atype}

    def _full_path(aid):
        parts = []
        visited = set()
        while aid and aid in acc_map and aid not in visited:
            visited.add(aid)
            parts.append(acc_map[aid]["name"])
            aid = acc_map[aid]["parent_id"]
        return ":".join(reversed(parts))

    # Candidate accounts: name contains the bank, under Assets/Bank.
    bank_lower = bank_name.lower()
    candidates = [
        (aid, _full_path(aid))
        for aid, info in acc_map.items()
        if bank_lower in _full_path(aid).lower() and info["type"] in ("BANK", "ASSET")
    ]

    target_id = None
    target_name = ""
    match_warning = None
    norm_number = _normalize_digits(account_number)

    if norm_number:
        number_matches = [
            (aid, full) for aid, full in candidates
            if _normalize_digits(full) and (
                norm_number in _normalize_digits(full)
                or _normalize_digits(full) in norm_number
            )
        ]
        if number_matches:
            target_id, target_name = number_matches[0]
            if len(number_matches) > 1:
                match_warning = (
                    f"Multiple GnuCash accounts matched account number "
                    f"'{account_number}'; using '{target_name}'."
                )

    if target_id is None and candidates:
        target_id, target_name = candidates[0]
        if norm_number:
            match_warning = (
                f"Could not match account number '{account_number}' to any "
                f"GnuCash account digits; fell back to name match on "
                f"'{bank_name}' -> '{target_name}'. Verify this is the "
                f"correct account."
            )
        elif len(candidates) > 1:
            match_warning = (
                f"{len(candidates)} GnuCash accounts match bank name "
                f"'{bank_name}'; using '{target_name}' by name only — no "
                f"account number was available to disambiguate."
            )

    if not target_id:
        return {
            "found": False, "account_name": "", "balance": 0.0,
            "last_txn_date": None, "match_warning": None,
        }

    # Sum splits for this account
    balance = 0.0
    last_date = None

    for trn in root.findall(f'.//{_NS["gnc"]}transaction'):
        date_el = trn.find(f'{_NS["trn"]}date-posted/{_NS["ts"]}date')
        trn_date = date_el.text[:10] if date_el is not None else None

        for sp in trn.findall(f'{_NS["trn"]}splits/{_NS["trn"]}split'):
            sp_acc = sp.findtext(f'{_NS["split"]}account', '')
            if sp_acc == target_id:
                val_str = sp.findtext(f'{_NS["split"]}value', '0/1')
                # GnuCash stores values as "num/denom" e.g. "150000/100"
                parts = val_str.split('/')
                if len(parts) == 2:
                    try:
                        balance += int(parts[0]) / int(parts[1])
                    except (ValueError, ZeroDivisionError):
                        pass
                if trn_date:
                    if last_date is None or trn_date > last_date:
                        last_date = trn_date

    return {
        "found": True,
        "account_name": target_name,
        "balance": round(balance, 2),
        "last_txn_date": last_date,
        "match_warning": match_warning,
    }


def _reconcile_opening_balance(
    canonical_rows: list[dict],
    gnucash_file: str,
    bank_name: str,
    account_number: str | None = None,
) -> dict:
    """
    Reconcile the canonical CSV's opening balance against GnuCash ledger.

    Three scenarios:
      A. Statement has entries dated before/on GnuCash's last txn date
         → these are duplicates already posted → skip them
      B. GnuCash has entries not in the statement (previous statement omitted)
         → cannot detect without prior statement → flag warning
      C. Some other error → flag error

    Returns:
        {
            "ok": bool,
            "message": str,
            "rows_skipped": int,           # scenario A duplicates removed
            "filtered_rows": list[dict],   # rows after removing duplicates
            "gnucash_balance": float,
            "statement_opening": float,
        }
    """
    if not canonical_rows:
        return {
            "ok": False,
            "message": "No rows in canonical CSV.",
            "rows_skipped": 0,
            "filtered_rows": [],
            "gnucash_balance": 0.0,
            "statement_opening": 0.0,
            "match_warning": None,
        }

    gc = _get_gnucash_account_balance(gnucash_file, bank_name, account_number)
    if not gc["found"]:
        log.warning(
            "Could not find %s account in GnuCash — skipping opening balance check",
            bank_name,
        )
        return {
            "ok": True,
            "message": f"GnuCash account for '{bank_name}' not found — skipping balance reconciliation.",
            "rows_skipped": 0,
            "filtered_rows": canonical_rows,
            "gnucash_balance": 0.0,
            "statement_opening": 0.0,
            "account_found": False,
            "match_warning": None,
        }

    gc_balance = gc["balance"]
    gc_last_date = gc["last_txn_date"]

    # Derive statement opening balance
    oc = extract_opening_closing(canonical_rows)
    stmt_opening = oc["opening_balance"]

    diff = abs(gc_balance - stmt_opening)

    if diff <= 0.02:
        # Perfect match — no duplicates, no gap
        return {
            "ok": True,
            "message": (
                f"Opening balance matches GnuCash: "
                f"GnuCash={gc_balance:.2f}, Statement={stmt_opening:.2f}"
            ),
            "rows_skipped": 0,
            "filtered_rows": canonical_rows,
            "gnucash_balance": gc_balance,
            "statement_opening": stmt_opening,
            "account_found": True,
            "match_warning": gc.get("match_warning"),
        }

    # Scenario A check: are there rows in the statement dated on or before
    # GnuCash's last transaction date? If so, they're likely duplicates.
    if gc_last_date:
        filtered = []
        skipped = 0
        running_bal = stmt_opening

        for row in canonical_rows:
            row_date = row.get("Date", "")
            # Normalise date for comparison — handle DD/MM/YYYY and YYYY-MM-DD
            try:
                if "/" in row_date:
                    parts = row_date.split("/")
                    if len(parts[2]) == 4:
                        cmp_date = f"{parts[2]}-{parts[1]}-{parts[0]}"
                    else:
                        cmp_date = f"20{parts[2]}-{parts[1]}-{parts[0]}"
                else:
                    cmp_date = row_date
            except (IndexError, ValueError):
                cmp_date = row_date

            if cmp_date <= gc_last_date:
                skipped += 1
                continue
            filtered.append(row)

        if skipped > 0 and filtered:
            # Re-check: does the new opening balance match GnuCash now?
            new_oc = extract_opening_closing(filtered)
            new_diff = abs(gc_balance - new_oc["opening_balance"])
            if new_diff <= 0.02:
                return {
                    "ok": True,
                    "message": (
                        f"Scenario A: {skipped} entries already in GnuCash (dated ≤ {gc_last_date}) — skipped.\n"
                        f"Opening balance now matches: GnuCash={gc_balance:.2f}, "
                        f"Statement (after skip)={new_oc['opening_balance']:.2f}"
                    ),
                    "rows_skipped": skipped,
                    "filtered_rows": filtered,
                    "gnucash_balance": gc_balance,
                    "statement_opening": new_oc["opening_balance"],
                    "account_found": True,
                    "match_warning": gc.get("match_warning"),
                }

    # Scenario B or C — gap we can't resolve
    return {
        "ok": False,
        "message": (
            f"OPENING BALANCE MISMATCH: GnuCash ({gc['account_name']}) "
            f"shows {gc_balance:.2f} but statement opens at {stmt_opening:.2f} "
            f"(diff={diff:.2f}).\n"
            f"Possible causes: (B) prior statement entries were omitted, "
            f"or (C) there's a data error. Please investigate manually."
        ),
        "rows_skipped": 0,
        "filtered_rows": canonical_rows,
        "gnucash_balance": gc_balance,
        "statement_opening": stmt_opening,
        "account_found": True,
        "match_warning": gc.get("match_warning"),
    }


def final_closing_balance_verdict(
    recon: dict,
    final_rows: list[dict],
    stmt_closing: float | None,
    unresolved_opening_gap: float | None,
) -> str:
    """
    Compute the final closing-balance verdict message.

    The verdict compares the actual POST-IMPORT GnuCash balance (pre-import
    book balance + net of the rows just imported) against the STATEMENT's own
    closing balance — an independent source when present, since it comes from
    the bank, not from the book. Comparing the statement-derived canonical CSV
    against itself is circular and can't catch a real opening-balance gap. Any
    gap left unexplained by dedup (``unresolved_opening_gap``) must propagate
    here too — it must never be silently dropped just because dedup found
    zero overlapping rows, and it always wins over a coincidentally-matching
    closing balance (an AMBER/RED verdict, never a false green).
    """
    if recon.get("account_found") is False:
        return "⚠ GnuCash account not found; post-import balance could not be verified."

    if stmt_closing is None:
        return "⚠ No independent statement closing balance available; nothing to verify the book against."

    net_imported = sum(
        _safe_float(r.get("Deposit", 0)) - _safe_float(r.get("Withdrawal", 0))
        for r in final_rows
    )
    post_import_balance = recon["gnucash_balance"] + net_imported
    closing_diff = abs(post_import_balance - stmt_closing)

    if unresolved_opening_gap is not None and unresolved_opening_gap > 0.02:
        return (
            f"⚠ unreconciled {unresolved_opening_gap:.2f} — opening-balance adjustment or "
            f"investigate. (Post-import book={post_import_balance:.2f}, "
            f"statement closing={stmt_closing:.2f}, diff={closing_diff:.2f}.)"
        )
    if closing_diff <= 0.02:
        return (
            f"Closing balance VERIFIED (independent): post-import book="
            f"{post_import_balance:.2f} matches statement closing={stmt_closing:.2f}."
        )
    return (
        f"❌ CLOSING BALANCE MISMATCH: post-import book={post_import_balance:.2f}, "
        f"statement closing={stmt_closing:.2f} (diff={closing_diff:.2f})."
    )


_NORMALISE_PROMPT = """\
You are a data normalisation assistant. Your job is to map the columns of a bank
statement CSV to a fixed canonical schema.

Canonical columns (in order):
  Date, Transaction ID, Description, Account, Deposit, Withdrawal, Balance, Currency

The user's CSV has these headers:
{headers}

Here are the first few sample rows so you can see the data format:
{sample}

Rules:
- Return ONLY a valid JSON object — no markdown, no explanation, no code fences.
- Map each canonical column name to the best-matching header from the user's CSV.
- If there is no reasonable match for a canonical column, map it to null.
- "Deposit" and "Withdrawal" are credit and debit amounts respectively. They may
  appear as a single "Amount" column with sign — if so, map both to that column
  and include a "sign_convention" key: "positive_is_deposit" or "negative_is_deposit".
- Do not invent column names; only use names that appear verbatim in the headers list.

Example response:
{{
  "Date": "Txn Date",
  "Transaction ID": "Ref No / Cheque No",
  "Description": "Narration",
  "Account": null,
  "Deposit": "Credit",
  "Withdrawal": "Debit",
  "Balance": "Balance (INR)",
  "Currency": null
}}
"""


def _read_tabular_rows(file_path: str) -> list[list[str]]:
    """Read a CSV or XLS/XLSX file into a flat list of string rows (no header
    split — see _find_generic_header_row for locating the header row)."""
    p = Path(file_path)
    suffix = p.suffix.lower()

    if suffix in (".xls", ".xlsx"):
        if suffix == ".xls":
            import xlrd
            wb = xlrd.open_workbook(str(p))
            ws = wb.sheet_by_index(0)
            rows = []
            for r in range(ws.nrows):
                row = []
                for c in range(ws.ncols):
                    cell = ws.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        row.append(dt.strftime("%d/%m/%Y"))
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        v = cell.value
                        row.append(str(int(v)) if v == int(v) else str(v))
                    elif cell.ctype == xlrd.XL_CELL_EMPTY:
                        row.append("")
                    else:
                        row.append(str(cell.value).strip())
                rows.append(row)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(str(p), data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c).strip() if c is not None else "" for c in row])
        return rows

    else:  # CSV
        with open(file_path, newline="", encoding="utf-8-sig") as fh:
            reader = csv.reader(fh)
            return [[str(c).strip() for c in row] for row in reader]


_HEADER_DATEISH_RE = re.compile(r'^\d{1,4}[/-]\d{1,2}[/-]\d{1,4}$')


def _find_generic_header_row(rows: list[list[str]]) -> int:
    """Locate the header row in an arbitrary bank CSV/XLS export, tolerant of
    preamble rows (account info, statement period, blank rows, '****'
    separators) above it — mirrors skill_hdfc's header detection but without
    assuming any particular bank's column names. Heuristic: the first row
    with >=2 non-empty cells where most cells look like text labels (contain
    a letter and aren't themselves a date), i.e. a plausible header row."""
    for i, row in enumerate(rows):
        cells = [str(c).strip() for c in row]
        non_empty = [c for c in cells if c]
        if len(non_empty) < 2:
            continue
        text_like = sum(
            1 for c in non_empty
            if re.search(r'[A-Za-z]', c) and not _HEADER_DATEISH_RE.match(c)
        )
        if text_like >= max(2, (len(non_empty) + 1) // 2):
            return i
    return 0  # fallback: no clear header row found, assume the first row


_CANONICAL_MAPPING_KEYS = {
    "Date", "Transaction ID", "Description", "Account",
    "Deposit", "Withdrawal", "Balance", "Currency",
}


def _sanitize_and_validate_mapping(raw_reply: str, headers: list[str]) -> dict:
    """Parse + sanitize an LLM column-mapping reply: strip accidental
    markdown fences, drop unknown keys, and verify every mapped value is
    either null or a header that actually appears (verbatim) in the file.
    Raises ValueError naming the offending keys/values on any violation —
    callers retry once on this error, then hard-fail."""
    raw = raw_reply.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    try:
        mapping = json.loads(raw.strip())
    except json.JSONDecodeError as e:
        raise ValueError(f"LLM reply was not valid JSON: {e}") from e
    if not isinstance(mapping, dict):
        raise ValueError("LLM reply must be a JSON object")

    sign_convention = mapping.get("sign_convention")
    sanitized = {k: v for k, v in mapping.items() if k in _CANONICAL_MAPPING_KEYS}

    header_set = set(headers)
    bad_values = {
        k: v for k, v in sanitized.items()
        if v is not None and v not in header_set
    }
    if bad_values:
        raise ValueError(
            "LLM mapped canonical column(s) to header(s) not present in the "
            f"file: {bad_values}. Real headers are: {headers}"
        )

    sanitized["sign_convention"] = sign_convention
    return sanitized


def _normalise_to_canonical(
    input_file: str,
    output_path: str,
    bank_name: str,
    config_path: str,
    model_override: str,
) -> None:
    """
    LLM-assisted column normalisation: maps arbitrary CSV/XLS to the canonical
    8-column schema and writes the result to output_path. Reserved for
    "Other Bank (CSV)" — HDFC always uses skill_hdfc's deterministic parser.
    """
    agents_root = Path(__file__).resolve().parent.parent
    if str(agents_root) not in sys.path:
        sys.path.insert(0, str(agents_root))
    from agents.base_agent import run_direct  # noqa: E402

    rows = _read_tabular_rows(input_file)
    if not rows:
        raise ValueError(f"Could not read any data from {input_file}")

    header_idx = _find_generic_header_row(rows)
    headers = rows[header_idx]
    sample = rows[header_idx + 1: header_idx + 6]
    all_rows = rows[header_idx + 1:]
    if not headers or not any(str(h).strip() for h in headers):
        raise ValueError(f"Could not find a header row in {input_file}")

    headers_str = json.dumps(headers)
    sample_str = "\n".join(
        "  " + ", ".join(f"{h}={v}" for h, v in zip(headers, row))
        for row in sample
    )

    prompt = _NORMALISE_PROMPT.format(headers=headers_str, sample=sample_str)
    system = (
        f"You are normalising a {bank_name} bank statement CSV to canonical format. "
        "Output ONLY raw JSON — no markdown, no explanation."
    )

    mapping = None
    last_error = None
    for attempt in range(2):  # one retry after an invalid/hallucinated reply
        user_message = prompt
        if attempt == 1:
            user_message = (
                prompt
                + f"\n\nYour previous reply was rejected: {last_error}\n"
                + "Only use header names that appear verbatim in the headers list above."
            )
        raw = run_direct(
            user_message=user_message,
            system_prompt=system,
            config_path=config_path,
            model_override=model_override,
        )
        try:
            mapping = _sanitize_and_validate_mapping(raw, headers)
            break
        except ValueError as e:
            last_error = e
            mapping = None

    if mapping is None:
        raise ValueError(
            f"LLM column mapping failed validation after retry: {last_error}"
        )

    sign_convention = mapping.pop("sign_convention", None)
    col_idx = {h: i for i, h in enumerate(headers)}

    def _get(row, col_name):
        if col_name is None:
            return ""
        idx = col_idx.get(col_name)
        if idx is None or idx >= len(row):
            return ""
        return row[idx]

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as out:
        writer = csv.writer(out)
        writer.writerow(CANONICAL_COLS)
        for row in all_rows:
            # Handle single-amount-column case
            deposit_src = mapping.get("Deposit")
            withdrawal_src = mapping.get("Withdrawal")
            deposit_val = _get(row, deposit_src)
            withdrawal_val = _get(row, withdrawal_src)

            if deposit_src and deposit_src == withdrawal_src and sign_convention:
                # Same source column — split by sign
                try:
                    amt = float(deposit_val.replace(",", "") or "0")
                except ValueError:
                    amt = 0.0
                if sign_convention == "positive_is_deposit":
                    deposit_val = str(amt) if amt > 0 else ""
                    withdrawal_val = str(abs(amt)) if amt < 0 else ""
                else:  # negative_is_deposit
                    deposit_val = str(abs(amt)) if amt < 0 else ""
                    withdrawal_val = str(amt) if amt > 0 else ""

            out_row = [
                _get(row, mapping.get("Date")),
                _get(row, mapping.get("Transaction ID")),
                _get(row, mapping.get("Description")),
                _get(row, mapping.get("Account")),
                deposit_val,
                withdrawal_val,
                _get(row, mapping.get("Balance")),
                _get(row, mapping.get("Currency")) or "INR",
            ]
            writer.writerow(out_row)


def run(
    bank: str,
    statement_files: str,
    gnucash_file: str,
    output_path: str,
    config_path: str = None,
    model_override: str = None,
    pdf_password: str = None,
) -> str:
    """
    Run the full GnuCash import pipeline.

    Args:
        bank:            Bank name — one of SUPPORTED_BANKS.
        statement_files: Path or comma-separated paths to uploaded statement file(s).
                         XLS for ICICI; PDF(s) for BoB / HSBC;
                         CSV or XLS/XLSX for HDFC / Other Bank.
        gnucash_file:    Path to .gnucash book (must be closed in GnuCash).
        output_path:     Path for the final mapped CSV.
        config_path:     Passed through to sub-skills and LLM calls.
        model_override:  Passed through to sub-skills and LLM calls.
        pdf_password:    Optional statement password, forwarded to skill_hdfc
                         for password-protected HDFC PDFs (for HDFC often the
                         Cust ID). Never logged.

    Returns:
        Human-readable summary string for the UI.
    """
    agents_root = Path(__file__).resolve().parent.parent
    if str(agents_root) not in sys.path:
        sys.path.insert(0, str(agents_root))

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    bank = (bank or "").strip()
    if bank not in SUPPORTED_BANKS:
        return (
            f"❌ Unknown bank **{bank!r}**.\n\n"
            f"Supported banks: {', '.join(SUPPORTED_BANKS)}"
        )

    log_lines = []

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        canonical_path = str(tmp_path / "canonical.csv")

        # ── Step 1: Bank extraction → canonical CSV ───────────────────────────
        # Every dedicated bank (DEDICATED_BANKS) is dispatched purely through
        # the agents.banks registry: display_name matches this pipeline's
        # `bank` dispatch string exactly (verified 2026-07-17), so no mapping
        # table is needed. Each BankSkill.parse() returns canonical rows in
        # memory only; the canonical CSV + sidecar is written once, here, via
        # the shared canonical_io tail — no bank writes its own CSV anymore.

        bank_info = next((b for b in discover_banks() if b.display_name == bank), None)

        if bank_info is not None:
            bank_input = (
                statement_files[0] if isinstance(statement_files, list)
                else statement_files.split(",")[0].strip()
            )
            # Per-bank input shaping — the one piece of bank-specific logic
            # that can't be pushed into a uniform call, since each skill's
            # parse() expects a different shape for a staged multi-file
            # upload (single resolved file vs. a whole PDF directory).
            if bank == "ICICI":
                bank_input = _resolve_single_file(bank_input, (".xls", ".xlsx"))
            elif bank == "HDFC":
                bank_input = _resolve_single_file(bank_input, (".csv", ".xls", ".xlsx", ".pdf"))
            elif bank == "HSBC":
                # If input is a directory (staged uploads), use it as-is;
                # if it's a single file, use its parent directory — HSBC's
                # parse() OCRs every PDF in the directory it's given.
                hsbc_src = Path(bank_input)
                bank_input = hsbc_src if hsbc_src.is_dir() else hsbc_src.parent
            elif bank == "Bank of Baroda":
                bob_src = Path(bank_input)
                if bob_src.is_dir() and not sorted(bob_src.glob("*.pdf")):
                    return (
                        f"## {bank} → no PDFs found\n\n"
                        f"❌ The staged upload directory contains no .pdf files:\n"
                        f"`{bank_input}`"
                    )
                bank_input = bob_src

            _emit_progress(1, f"{bank}: extracting statement to canonical CSV")
            log_lines.append(f"**Step 1** — {bank}: extracting statement to canonical CSV")
            try:
                skill = load_bank_skill(bank_info)
                bank_result = skill.parse(bank_input, password=pdf_password)
                write_canonical_csv(bank_result.rows, canonical_path)
                write_sidecar(
                    canonical_path, bank_info.display_name, "derived",
                    bank_result.opening_balance, bank_result.closing_balance,
                    bank_result.row_count,
                    account_number=(bank_result.meta.account_number if bank_result.meta else None),
                )
                log.info("%s skill: %d canonical rows (balance_ok=%s)",
                         bank, bank_result.row_count, bank_result.balance_check.ok)
            except Exception as e:
                log.error("%s extraction failed: %s", bank, e, exc_info=True)
                return (
                    f"## {bank} → extraction error\n\n"
                    f"❌ {bank} skill raised an exception:\n```\n{e}\n```"
                )

        elif bank == "Other Bank (CSV)":
            input_file = (
                statement_files[0] if isinstance(statement_files, list)
                else statement_files.split(",")[0].strip()
            )
            input_file = _resolve_single_file(input_file, (".csv", ".xls", ".xlsx"))
            _emit_progress(1, f"Other Bank: reading CSV/XLS statement")
            log_lines.append("**Step 1** — Other Bank: reading CSV/XLS statement")
            _emit_progress(2, f"Other Bank: LLM normalising columns → canonical schema")
            log_lines.append("**Step 2** — Other Bank: LLM normalising columns → canonical schema")
            try:
                _normalise_to_canonical(
                    input_file=input_file,
                    output_path=canonical_path,
                    bank_name=bank,
                    config_path=config_path,
                    model_override=model_override,
                )
            except Exception as e:
                log.error("Other Bank normalisation failed: %s", e, exc_info=True)
                return (
                    f"## {bank} → normalisation error\n\n"
                    f"❌ Column normalisation raised an exception:\n```\n{e}\n```"
                )

        # ── Verify extraction produced output ────────────────────────────────
        if not Path(canonical_path).is_file():
            steps_summary = "\n".join(f"🟢 {line}" for line in log_lines)
            return (
                f"## {bank} → extraction failed\n\n"
                f"{steps_summary}\n\n"
                f"🔴 Bank extraction did not produce a canonical CSV.\n"
                f"Check the console log for errors from the {bank} skill."
            )

        # ── Balance verification on canonical CSV ─────────────────────────────
        _emit_progress(3, f"{bank}: verifying balances")
        # Read the canonical CSV back for balance checks
        with open(canonical_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            canonical_rows = list(reader)

        # Running balance check (Intervention 2)
        running = verify_running_balance(canonical_rows)
        if running["ok"]:
            log_lines.append(
                f"Running balance: OK ({running['opening_balance']:.2f} → "
                f"{running['closing_balance']:.2f})"
            )
        else:
            log_lines.append(
                f"Running balance: {running['mismatches']} mismatch(es)"
            )

        # ── Opening balance reconciliation with GnuCash (Intervention 1) ──
        # Prefer the account number from statement metadata (if the adapter
        # captured one) over a bare bank-name match — disambiguates multiple
        # accounts at the same bank.
        stmt_sidecar = _read_sidecar(canonical_path)
        stmt_account_number = stmt_sidecar.get("account_number") if stmt_sidecar else None
        recon = _reconcile_opening_balance(canonical_rows, gnucash_file, bank, stmt_account_number)
        if recon.get("match_warning"):
            log_lines.append(f"⚠ Account match: {recon['match_warning']}")

        if recon["rows_skipped"] > 0:
            log_lines.append(
                f"Skipped {recon['rows_skipped']} duplicate entries "
                f"already in GnuCash (date-based filter)"
            )
            with open(canonical_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=CANONICAL_COLS)
                writer.writeheader()
                writer.writerows(recon["filtered_rows"])
            canonical_rows = recon["filtered_rows"]

        # Track any unexplained opening-balance gap so it survives into the
        # final closing-balance verdict below, instead of silently evaporating
        # if dedup happens to find zero overlapping rows.
        unresolved_opening_gap = None
        if recon["ok"]:
            log_lines.append(f"**Balance check** — {recon['message']}")
        else:
            unresolved_opening_gap = abs(recon["gnucash_balance"] - recon["statement_opening"])
            log_lines.append(
                f"**Balance check** — ⚠ Opening balance gap detected "
                f"(GnuCash={recon['gnucash_balance']:.2f}, "
                f"statement={recon['statement_opening']:.2f}, diff={unresolved_opening_gap:.2f}). "
                f"Dedup below will attempt to reconcile overlapping transactions; "
                f"if it doesn't, this gap carries into the final verdict."
            )

        # ── Duplicate detection (Phase 4 Lite) ─────────────────────────────────
        # Compare canonical CSV against GnuCash book to flag duplicates
        _emit_progress(4, f"{bank}: checking for duplicates in GnuCash")

        gnucash_data = None  # set here so contra detection can use it even if dup-check fails
        try:
            # Parse GnuCash file to get existing transactions
            gnucash_data = parse_gnucash_for_reconcile(gnucash_file)

            # Convert canonical_rows to reconcile format
            # (canonical_rows are already dicts with 'Date', 'Deposit', 'Withdrawal' keys)
            reconcile_rows = []
            for idx, row in enumerate(canonical_rows, 1):
                try:
                    deposit = _safe_float(row.get('Deposit', 0))
                    withdrawal = _safe_float(row.get('Withdrawal', 0))
                    reconcile_rows.append({
                        'row_num': idx,
                        'date': row.get('Date', ''),
                        'description': row.get('Description', ''),
                        'deposit': deposit,
                        'withdrawal': withdrawal,
                    })
                except (ValueError, KeyError):
                    reconcile_rows.append({
                        'row_num': idx,
                        'date': row.get('Date', ''),
                        'description': row.get('Description', ''),
                        'deposit': 0.0,
                        'withdrawal': 0.0,
                    })

            # Run reconciliation
            report, dedup_summary = reconcile(reconcile_rows, gnucash_data)

            matched_count = dedup_summary.get('matched', 0)
            duplicate_count = dedup_summary.get('duplicates', 0)
            new_count = dedup_summary.get('new', 0)
            total_duplicates = matched_count + duplicate_count

            # Filter to keep only "New" rows
            new_rows = [
                canonical_rows[i] for i, r in enumerate(report)
                if r.get('status') == 'New'
            ]

            # Edge case: all rows are duplicates
            if total_duplicates > 0 and new_count == 0:
                return (
                    f"## {bank} → GnuCash pipeline — all transactions already in GnuCash\n\n"
                    f"Duplicate check — All {len(canonical_rows)} transaction(s) are already "
                    f"in your GnuCash book. Nothing to import.\n\n"
                    f"---\n\n"
                    f"**Next:** If you expected new transactions, check that:\n"
                    f"1. Your GnuCash file is current\n"
                    f"2. Your bank statement covers the right period\n"
                    f"3. Transactions match by date + amount (GnuCash matching logic)"
                )

            # Rewrite canonical CSV with only new rows
            if total_duplicates > 0:
                log_lines.append(
                    f"Duplicate check — {total_duplicates} already in GnuCash "
                    f"(removed), {new_count} new (will be mapped)"
                )
                with open(canonical_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=CANONICAL_COLS)
                    writer.writeheader()
                    writer.writerows(new_rows)
                canonical_rows = new_rows
            else:
                log_lines.append(
                    f"Duplicate check — {new_count} new transactions (none in GnuCash)"
                )

        except Exception as e:
            log_lines.append(
                f"⚠️ Duplicate check skipped — {e}. Proceeding with all rows."
            )
            log.warning(f"Duplicate detection failed: {e}")

        # ── Resolve GnuCash bank account path (for CSV Account column) ────
        gc_info = _get_gnucash_account_balance(gnucash_file, bank)
        gnucash_bank_account = ""
        if gc_info["found"]:
            raw_path = gc_info["account_name"]
            # Strip "Root Account:" prefix — GnuCash CSV importer doesn't want it
            if raw_path.startswith("Root Account:"):
                gnucash_bank_account = raw_path[len("Root Account:"):]
            else:
                gnucash_bank_account = raw_path
            log_lines.append(f"Bank account: `{gnucash_bank_account}`")
        else:
            # No matching bank account in the .gnucash book. This quietly turns
            # off three things, so say so plainly rather than leaving the user to
            # wonder why this bank's output looks reduced: (1) Transfer Account is
            # left blank, (2) cross-bank transfer (contra) detection is skipped,
            # (3) the opening-balance reconciliation check is skipped.
            log_lines.append(
                f"⚠️ **Couldn't find a '{bank}' bank account in the GnuCash book.** "
                f"Transfer Account is left blank, and cross-bank transfer (contra) "
                f"detection and the opening-balance check are skipped. To enable them, "
                f"add or rename a bank-typed account in your `.gnucash` that matches "
                f"'{bank}' (e.g. `Assets:…:Cash and Bank:{bank} - <account-number>`)."
            )

        # ── Contra detection (cross-bank transfer matching) ──────────────────
        contra_flags: dict[int, dict] = {}  # row_idx → contra info
        try:
            if gnucash_bank_account and gnucash_data:
                contras = detect_contra_entries(
                    canonical_rows, gnucash_data, gnucash_bank_account
                )
                if contras:
                    for c in contras:
                        contra_flags[c["row_idx"]] = c
                    confirmed = sum(1 for c in contras if c["confidence"] == "high")
                    possible = len(contras) - confirmed
                    parts = []
                    if confirmed:
                        parts.append(f"{confirmed} confirmed (account set to bank)")
                    if possible:
                        parts.append(f"{possible} possible")
                    log_lines.append(
                        f"Contra check — {len(contras)} cross-bank "
                        f"transfer(s) detected ({', '.join(parts)}). "
                        f"Review in **Review Mappings** tab."
                    )
                    _emit_progress(4, f"{bank}: {len(contras)} contra(s) flagged")
                else:
                    log_lines.append("Contra check — no cross-bank transfers detected")
        except Exception as e:
            log.warning(f"Contra detection failed: {e}")
            log_lines.append(f"⚠️ Contra check skipped — {e}")

        # ── Step 3: Account mapping ───────────────────────────────────────────
        if bank == "ICICI":
            step_n = 2
        elif bank in CSV_BANKS:
            step_n = 3
        else:
            step_n = 3

        _emit_progress(5, f"{bank}: mapping accounts from {Path(gnucash_file).name}")
        log_lines.append(
            f"**Step {step_n}** — GnuCash: mapping accounts from "
            f"`{Path(gnucash_file).name}`"
        )
        from skill_gnucash_account_mapper.agent import run as mapper_run  # noqa: E402
        mapping_result = mapper_run(
            gnucash_file=gnucash_file,
            canonical_csv=canonical_path,
            output_path=output_path,
            config_path=config_path,
            model_override=model_override,
            bank_name=bank,
            gnucash_bank_account=gnucash_bank_account,
        )

        # ── Apply confirmed contras to the mapped output ────────────────────
        # For confirmed (high-confidence, reference-matched) transfers, book the
        # row against the counterparty bank instead of whatever category the
        # mapper guessed — a genuine bank-to-bank transfer must not land in
        # income/expense/investment. Possible (medium) contras are left alone:
        # they stay a review hint and keep the mapper's account. contra row_idx
        # is 0-based into canonical_rows, which the mapper preserves 1:1.
        if contra_flags:
            try:
                remapped = _apply_confirmed_contras(output_path, contra_flags)
                if remapped:
                    log_lines.append(
                        f"Contra — booked {remapped} confirmed transfer(s) "
                        f"to the counterparty bank account."
                    )
            except Exception as e:
                log.warning(f"Could not apply confirmed contras: {e}")
                log_lines.append(f"⚠️ Contra remap skipped — {e}")

        # Write contra flags sidecar (if any) alongside the output CSV
        if contra_flags:
            contra_sidecar = Path(output_path).with_suffix('.contra.json')
            try:
                import json as _json
                with open(contra_sidecar, 'w', encoding='utf-8') as cf:
                    _json.dump(contra_flags, cf, indent=2, default=str)
            except Exception as e:
                log.warning(f"Could not write contra sidecar: {e}")

        _emit_progress(6, f"{bank}: final balance verification")
        try:
            with open(output_path, "r", encoding="utf-8") as f:
                final_rows = list(csv.DictReader(f))
            sidecar = _read_sidecar(canonical_path)
            stmt_closing = sidecar.get("closing_balance") if sidecar else None
            log_lines.append(
                "**Final check** — "
                + final_closing_balance_verdict(recon, final_rows, stmt_closing, unresolved_opening_gap)
            )
        except Exception as e:
            log_lines.append(f"**Final check** — Could not verify closing balance: {e}")

    # Color-code each log line: green = OK, amber = warning, red = error
    _WARN_KEYS = ("mismatch", "gap detected", "skipped", "⚠", "warning")
    _ERR_KEYS = ("❌", "error", "failed", "could not")

    formatted_lines = []
    for line in log_lines:
        # Strip markdown bold for cleaner one-line display
        clean = line.replace("**", "")
        low = clean.lower()
        if any(k in low for k in _ERR_KEYS):
            formatted_lines.append(f"🔴 {clean}")
        elif any(k in low for k in _WARN_KEYS):
            formatted_lines.append(f"🟡 {clean}")
        else:
            formatted_lines.append(f"🟢 {clean}")

    steps_summary = "  \n".join(formatted_lines)  # MD line break (two spaces + \n)
    return (
        f"## {bank} → GnuCash pipeline complete\n\n"
        f"{steps_summary}\n\n"
        f"---\n\n"
        f"{mapping_result}\n\n"
        f"**Next:** check the **Review Mappings** tab to verify/correct account assignments, then import into GnuCash."
    )
