#!/usr/bin/env python3
"""
Create CC Transaction List - Implementation Script

Parses bank-specific credit card statement PDFs and consolidates all
transactions into a single Excel workbook.

Supported banks and the pdftotext layout each uses:
  SBM-Global      : multi-line — date alone, description next line, amount next line
  YES-Reserv      : single-line — DD/MM/YYYY + description, amount with Dr/Cr on same line
  HDFC-Regalia    : single-line — DD/MM/YYYY [HH:MM:SS] + description + amount [Cr]
  Axis-Flipkart   : columnar — pdftotext serialises columns: all dates, then all descs, then amounts
  Axis-IndianOil  : same columnar layout as Axis-Flipkart
  ICICI-Amazon    : tabular — DD/MM/YYYY on one line, txn-id + description on next, amount standalone
  ICICI-Sapphiro  : same layout as ICICI-Amazon
  SBI-BPCL-Octane : multi-line — DD Mon YY + description block + amount with D/C suffix
  HSBC-Premier    : single-line — DDMMM date + description + amount (no Cr/Dr suffix)
"""

import os
import re
import subprocess
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ---------------------------------------------------------------------------
# PDF text extraction
# ---------------------------------------------------------------------------

def extract_text_from_pdf(pdf_path):
    """Extract text from PDF using pdftotext (UTF-8, replace bad bytes)."""
    try:
        result = subprocess.run(
            ['pdftotext', str(pdf_path), '-'],
            capture_output=True,
            timeout=30,
        )
        if result.returncode != 0:
            return ""
        return result.stdout.decode('utf-8', errors='replace')
    except Exception as e:
        print(f"  [WARN] pdftotext failed for {Path(pdf_path).name}: {e}")
        return ""


# ---------------------------------------------------------------------------
# Bank/card detection from folder path
# ---------------------------------------------------------------------------

def get_bank_and_card_type(pdf_path):
    """
    Extract bank name and card type from the immediate parent folder of the PDF.
    The folder is expected to be named 'Bank-CardType' (e.g. 'Axis-Flipkart').
    The parent folder is the direct parent directory of the PDF file.
    """
    parent = Path(pdf_path).parent.name
    if '-' in parent:
        bank = parent.split('-', 1)[0]
        card_type = parent[len(bank)+1:]
        return bank, card_type
    return "Unknown", "Unknown"


# ---------------------------------------------------------------------------
# Parser: SBM Global
# ---------------------------------------------------------------------------

def parse_sbm_transactions(text, bank_name, card_type):
    """
    SBM Global layout (pdftotext output):

        Transactions
        Date
        15-Apr-2025
        Description text here
        1234.56
        ...

    Date line: DD-MMM-YYYY
    Direction: 'Cr' if description contains 'Repayment'/'Credit'/'Cashback', else 'Dr'
    """
    transactions = []
    lines = [l.strip() for l in text.split('\n')]

    in_section = False
    i = 0
    while i < len(lines):
        line = lines[i]

        # Enter section on "Transactions" followed by "Date"
        if line == 'Transactions' and i + 1 < len(lines) and lines[i + 1] == 'Date':
            in_section = True
            i += 2
            continue

        if in_section and ('Important Information' in line or 'Page' in line):
            in_section = False
            i += 1
            continue

        if not in_section or not line:
            i += 1
            continue

        date_match = re.match(r'^(\d{2}-[A-Za-z]{3}-\d{4})$', line)
        if date_match:
            date = date_match.group(1)
            description = None
            amount = None

            j = i + 1
            # Next non-blank non-numeric line → description
            while j < len(lines) and j < i + 10:
                nl = lines[j]
                if nl and not re.match(r'^\d[\d,]*\.?\d*$', nl):
                    description = nl
                    break
                j += 1

            # Next numeric-only line → amount
            j2 = j + 1 if j > i else i + 1
            while j2 < len(lines) and j2 < i + 15:
                nl = lines[j2]
                if re.match(r'^\d[\d,]*\.?\d*$', nl):
                    amount = nl.replace(',', '')
                    break
                j2 += 1

            if description and amount:
                cr_keywords = r'repayment|credit|cashback|refund|reversal'
                direction = 'Cr' if re.search(cr_keywords, description, re.I) else 'Dr'
                transactions.append({
                    'date': date, 'description': description,
                    'amount': amount, 'direction': direction,
                    'bank': bank_name, 'card_type': card_type,
                })
            i += 1
            continue

        i += 1

    return transactions


# ---------------------------------------------------------------------------
# Parser: YES Bank Reserv
# ---------------------------------------------------------------------------

def parse_yes_bank_transactions(text, bank_name, card_type):
    """
    YES Bank layout:

        Statement Details
        Date    Transaction Details    Merchant Category    Amount (Rs.)
        08/10/2025 Swiggy Limited Bangalore IN - Ref No: VT...    Retail Outlet Services    395.00 Dr
        ------------------End of the Statement------------------

    Date and description are on the same line; amount with Dr/Cr is at the end
    of the same line OR on a subsequent line within the next 15 lines.
    """
    transactions = []
    lines = text.split('\n')

    in_section = False
    i = 0
    while i < len(lines):
        line = lines[i]

        if 'Statement Details' in line:
            in_section = True
            i += 1
            continue

        if in_section and ('End of the Statement' in line or '---' in line):
            in_section = False
            i += 1
            continue

        if not in_section or not line.strip():
            i += 1
            continue

        # Match DD/MM/YYYY at start of line
        date_match = re.match(r'^(\d{2}/\d{2}/\d{4})\s+(.+)', line.strip())
        if date_match:
            date, rest = date_match.groups()

            # Amount may be at end of same line or in next few lines
            amt_match = re.search(r'([\d,]+\.?\d*)\s*(Cr|Dr)\s*$', rest)
            if amt_match:
                amount = amt_match.group(1).replace(',', '')
                direction = amt_match.group(2)
                description = re.sub(r'([\d,]+\.?\d*)\s*(Cr|Dr)\s*$', '', rest).strip()
                # Strip trailing merchant category (everything after last " - Ref No:" or similar)
                description = re.sub(r'\s+-\s+Ref No:.*$', '', description).strip()
                transactions.append({
                    'date': date, 'description': description,
                    'amount': amount, 'direction': direction,
                    'bank': bank_name, 'card_type': card_type,
                })
            else:
                # Look ahead for the amount line
                description = rest.strip()
                found = False
                j = i + 1
                while j < min(i + 15, len(lines)) and not found:
                    nl = lines[j].strip()
                    if not nl:
                        j += 1
                        continue
                    if re.match(r'^\d{2}/\d{2}/\d{4}', nl):
                        break
                    amt_match2 = re.search(r'([\d,]+\.?\d*)\s*(Cr|Dr)', nl)
                    if amt_match2:
                        amount = amt_match2.group(1).replace(',', '')
                        direction = amt_match2.group(2)
                        description = re.sub(r'\s+-\s+Ref No:.*$', '', description).strip()
                        transactions.append({
                            'date': date, 'description': description,
                            'amount': amount, 'direction': direction,
                            'bank': bank_name, 'card_type': card_type,
                        })
                        found = True
                    j += 1

        i += 1

    return transactions


# ---------------------------------------------------------------------------
# Parser: HDFC Regalia / TataNeu
# ---------------------------------------------------------------------------

def parse_hdfc_transactions(text, bank_name, card_type):
    """
    HDFC layout — pdftotext serialises columns separately:

        Domestic Transactions
        Date
        16/04/2025          ← all dates together
        16/04/2025
        19/04/2025 14:55:04
        ...
        Transaction Description
        JOHN DOE      ← account name — skip
        REDEMPTION PROC FEE (Ref# ...)
        TELE TRANSFER CREDIT
        ...
        Amount (in Rs.)
        99.00
        2,950.00 Cr
        ...

    Strategy: collect dates, descriptions, amounts in separate passes then zip.
    """
    transactions = []
    lines = [l.strip() for l in text.split('\n')]

    # Find the transaction section boundaries
    start_idx = -1
    end_idx = len(lines)
    for i, l in enumerate(lines):
        if re.search(r'Domestic Transactions|International Transactions', l) and start_idx == -1:
            start_idx = i
        if start_idx != -1 and re.search(r'Reward Points Summary|Points expiring', l):
            end_idx = i
            break

    if start_idx == -1:
        return transactions

    section = lines[start_idx:end_idx]

    # --- collect dates (DD/MM/YYYY with optional HH:MM:SS) ---
    dates = []
    in_date_col = False
    for l in section:
        if l == 'Date':
            in_date_col = True
            continue
        if l in ('Feature Reward', 'Transaction Description', 'Amount (in Rs.)'):
            in_date_col = False
            continue
        if in_date_col and re.match(r'^\d{2}/\d{2}/\d{4}', l):
            # Strip optional timestamp
            dates.append(re.match(r'^(\d{2}/\d{2}/\d{4})', l).group(1))

    # --- collect descriptions ---
    descriptions = []
    in_desc_col = False
    skip_name = False
    for l in section:
        if l == 'Transaction Description':
            in_desc_col = True
            skip_name = True
            continue
        if l in ('Amount (in Rs.)', 'Reward Points Summary'):
            in_desc_col = False
            continue
        if in_desc_col:
            if skip_name:
                skip_name = False
                continue  # skip account holder name
            if l and not re.match(r'^\d', l):
                descriptions.append(l)

    # --- collect amounts + directions ---
    amounts = []
    directions = []
    in_amt_col = False
    for l in section:
        if l == 'Amount (in Rs.)':
            in_amt_col = True
            continue
        if l in ('Reward Points Summary',):
            in_amt_col = False
            continue
        if in_amt_col:
            amt_match = re.match(r'^([\d,]+\.?\d*)\s*(Cr)?\s*$', l)
            if amt_match:
                amounts.append(amt_match.group(1).replace(',', ''))
                directions.append('Cr' if amt_match.group(2) else 'Dr')

    count = min(len(dates), len(descriptions), len(amounts))
    for k in range(count):
        if float(amounts[k]) > 0:
            transactions.append({
                'date': dates[k], 'description': descriptions[k],
                'amount': amounts[k], 'direction': directions[k],
                'bank': bank_name, 'card_type': card_type,
            })

    return transactions


# ---------------------------------------------------------------------------
# Parser: Axis Bank (Flipkart / IndianOil)
# ---------------------------------------------------------------------------

def parse_axis_transactions(text, bank_name, card_type):
    """
    Axis Bank — pdftotext serialises this 3-column table column-by-column:

        Account Summary
        DATE
        Card No:
        18/11/2025        ← all dates first
        20/11/2025
        ...

        TRANSACTION DETAILS
        440006******5138  ← card number — skip
        FLIPKART PAYMENTS,GURGAON   ← descriptions
        FLIPKART PAYMENTS,BANGALORE
        ...

        MERCHANT CATEGORY
        AMOUNT (Rs.)
        CASHBACK EARNED
        Name
        HARSHAL ...       ← account name — skip
        MISC STORE        ← merchant category — skip
        8.00 Dr           ← amounts
        249.00 Cr
        ...

    Strategy: collect dates, descriptions, amounts/directions in separate passes
    then zip them together.
    """
    transactions = []

    # Locate 'Account Summary' ... '**** End of Statement ****'
    start = text.find('Account Summary')
    end = text.find('**** End of Statement ****')
    if start == -1:
        return transactions
    section = text[start: end if end != -1 else len(text)]

    lines = [l.strip() for l in section.split('\n') if l.strip()]

    # --- collect dates ---
    dates = []
    i = 0
    # Dates appear between 'DATE' / 'Card No:' header and 'TRANSACTION DETAILS'
    in_date_block = False
    for line in lines:
        if line == 'DATE':
            in_date_block = True
            continue
        if line == 'TRANSACTION DETAILS':
            in_date_block = False
            continue
        if in_date_block and re.match(r'^\d{2}/\d{2}/\d{4}$', line):
            dates.append(line)

    # --- collect descriptions ---
    descriptions = []
    in_desc_block = False
    skip_next = False
    for line in lines:
        if line == 'TRANSACTION DETAILS':
            in_desc_block = True
            skip_next = True  # next line is card number
            continue
        if line in ('MERCHANT CATEGORY', 'AMOUNT (Rs.)', 'CASHBACK EARNED'):
            in_desc_block = False
            continue
        if in_desc_block:
            if skip_next:
                skip_next = False
                continue  # skip card number line
            # Skip blank-ish separators
            if re.match(r'^[\d*X]+$', line):
                continue
            descriptions.append(line)

    # --- collect amounts ---
    amounts = []
    directions = []
    in_amount_block = False
    # Account holder name appears right after MERCHANT CATEGORY block
    skip_name = False
    for line in lines:
        if line == 'AMOUNT (Rs.)':
            in_amount_block = True
            skip_name = True  # next meaningful line may be account name
            continue
        if line == 'CASHBACK EARNED':
            # amounts still interleaved with cashback; stop after cashback section
            continue
        if line == '**** End of Statement ****':
            break
        if in_amount_block:
            # Skip "Name", account holder name, "MISC STORE" category lines
            if line in ('Name', 'MERCHANT CATEGORY', 'CASHBACK EARNED'):
                skip_name = False
                continue
            # Skip lines that look like merchant categories (all caps, no numbers)
            if re.match(r'^[A-Z][A-Z &]+$', line) and not re.search(r'\d', line):
                continue
            amt_match = re.match(r'^([\d,]+\.?\d*)\s*(Dr|Cr)$', line)
            if amt_match:
                if skip_name:
                    skip_name = False
                amounts.append(amt_match.group(1).replace(',', ''))
                directions.append(amt_match.group(2))

    # Zip together — use min length in case of parsing drift
    count = min(len(dates), len(descriptions), len(amounts))
    for k in range(count):
        try:
            amt_val = float(amounts[k])
        except (ValueError, TypeError):
            amt_val = 0
        if amt_val == 0:
            continue  # skip zero-amount rows (cashback placeholders etc.)
        transactions.append({
            'date': dates[k],
            'description': descriptions[k],
            'amount': amounts[k],
            'direction': directions[k],
            'bank': bank_name,
            'card_type': card_type,
        })

    return transactions


# ---------------------------------------------------------------------------
# Parser: ICICI (Amazon / Sapphiro)
# ---------------------------------------------------------------------------

def parse_icici_transactions(text, bank_name, card_type):
    """
    ICICI layout — pdftotext serialises columns:

        Transaction Details
        ...header junk...
        [txn-id]         ← e.g. 13086609419
        AMAZON PAY IN GROCERY BANGALORE IN   ← description
        12               ← reward points (skip)
        243.00           ← amount
        25/03/2026       ← date
        [txn-id]
        AMAZON PAY...
        150
        3,004.00 CR      ← amount with optional CR suffix

    Pattern: txn-id → description (1-2 lines) → points → amount → date
    Repeat. Collect all four columns then zip.
    """
    transactions = []
    lines = [l.strip() for l in text.split('\n')]

    # Find section start
    start_idx = -1
    for i, l in enumerate(lines):
        if l == 'Transaction Details':
            start_idx = i + 1
            break
    if start_idx == -1:
        return transactions

    # Find section end
    end_idx = len(lines)
    for i, l in enumerate(lines[start_idx:], start=start_idx):
        if re.search(r'TOTAL AMOUNT DUE|Total Amount Due|End of Statement', l, re.I):
            end_idx = i
            break

    section = lines[start_idx:end_idx]

    # Each "record" in ICICI column layout repeats in this order:
    # txn_id (8+ digit number) → description → [points] → amount [CR] → date DD/MM/YYYY
    # Walk through and detect each record start by txn_id

    txn_id_re = re.compile(r'^\d{10,}$')
    date_re   = re.compile(r'^\d{2}/\d{2}/\d{4}$')
    amt_re    = re.compile(r'^([\d,]+\.\d{2})\s*(CR|DR)?\s*$', re.I)
    pts_re    = re.compile(r'^-?\d{1,5}$')   # reward points: small integer, possibly negative

    i = 0
    while i < len(section):
        l = section[i]

        if not txn_id_re.match(l):
            i += 1
            continue

        # Found a txn-id; collect the record
        txn_id = l
        desc_parts = []
        amount = None
        direction = 'Dr'
        date = None

        j = i + 1
        while j < len(section) and j < i + 12:
            nl = section[j]
            if not nl:
                j += 1
                continue
            if txn_id_re.match(nl):
                break   # next record
            if date_re.match(nl):
                date = nl
                j += 1
                break
            amt_match = amt_re.match(nl)
            if amt_match and amount is None:
                amount = amt_match.group(1).replace(',', '')
                direction = 'Cr' if (amt_match.group(2) or '').upper() == 'CR' else 'Dr'
                j += 1
                continue
            if pts_re.match(nl) and amount is None:
                # reward points — skip
                j += 1
                continue
            desc_parts.append(nl)
            j += 1

        description = ' '.join(desc_parts).strip()
        # Auto-detect credits even without CR suffix
        cr_keywords = r'payment|credit|refund|reversal|cashback|bbps'
        if re.search(cr_keywords, description, re.I):
            direction = 'Cr'

        if date and description and amount:
            transactions.append({
                'date': date, 'description': description,
                'amount': amount, 'direction': direction,
                'bank': bank_name, 'card_type': card_type,
            })

        i = j

    return transactions


# ---------------------------------------------------------------------------
# Parser: SBI BPCL Octane
# ---------------------------------------------------------------------------

def parse_sbi_transactions(text, bank_name, card_type):
    """
    SBI layout — pdftotext serialises columns:

        Date           (column 1, before 'Transaction Details')
        20 Mar 26
        07 Mar 26

        Transaction Details   for Statement Period: ...
        ANNUAL FEE CHARGED (EXCL TAX 269.82)    ← description lines
        IGST DB @ 18.00%
        TRANSACTIONS FOR JOHN DOE          ← header — skip
        WWWBIGBASKETCOM
        GURGAON
        IN

        Amount ( ` )   (column 3)
        1,499.00
        269.82

        D              (column 4 — direction: C or D)
        D

        145.00
        D

    Strategy:
    1. Dates are in their own column, appearing BEFORE 'Transaction Details'
       under a 'Date' header.
    2. Descriptions are between 'Transaction Details' header and 'Amount ( ` )'.
       'TRANSACTIONS FOR JOHN DOE' is a header — skip it.
    3. Amounts and directions are after 'Amount ( ` )'.
       Each amount is immediately followed by its C/D direction.
    """
    transactions = []
    lines = [l.strip() for l in text.split('\n')]

    # --- collect dates from the Date column (before Transaction Details) ---
    dates = []
    in_date_col = False
    txn_details_idx = -1
    for i, l in enumerate(lines):
        if l == 'Date':
            in_date_col = True
            continue
        if 'Transaction Details' in l:
            in_date_col = False
            txn_details_idx = i
            continue
        if in_date_col and re.match(r'^\d{1,2}\s+[A-Za-z]{3}\s+\d{2}$', l):
            dates.append(l)

    if txn_details_idx == -1:
        return transactions

    # --- collect descriptions (between Transaction Details and Amount) ---
    descriptions = []
    amt_col_idx = -1
    in_desc = False
    for i, l in enumerate(lines[txn_details_idx:], start=txn_details_idx):
        if i == txn_details_idx:
            in_desc = True
            continue
        if 'Amount ( ` )' in l:
            in_desc = False
            amt_col_idx = i
            break
        if in_desc:
            # Skip blank lines and the "TRANSACTIONS FOR <NAME>" header
            if not l or re.match(r'^TRANSACTIONS FOR\b', l, re.I):
                continue
            # Skip "for Statement Period:" line
            if l.startswith('for Statement Period'):
                continue
            descriptions.append(l)

    if amt_col_idx == -1:
        return transactions

    # --- collect amounts + directions (after 'Amount ( ` )') ---
    # Format: number, then 'D' or 'C' on the next non-blank line
    amounts = []
    directions = []
    end_marker = re.compile(r'Transactions highlighted|Important Messages|SAVINGS AND BENEFITS')
    i = amt_col_idx + 1
    while i < len(lines):
        l = lines[i]
        if end_marker.search(l):
            break
        amt_match = re.match(r'^([\d,]+\.?\d{2})$', l)
        if amt_match:
            amount = amt_match.group(1).replace(',', '')
            # Look for C/D on the next non-blank line
            j = i + 1
            while j < len(lines) and not lines[j].strip():
                j += 1
            if j < len(lines) and lines[j].strip() in ('C', 'D'):
                direction = 'Cr' if lines[j].strip() == 'C' else 'Dr'
                amounts.append(amount)
                directions.append(direction)
                i = j + 1
                continue
        i += 1

    count = min(len(dates), len(amounts))

    # SBI descriptions can span multiple lines per transaction.
    # If description count > transaction count, collapse them into 'count' groups.
    # Strategy: join ALL description lines into one string then split by known
    # section markers. Simpler: just join all into 'count' chunks evenly.
    if len(descriptions) > count and count > 0:
        # Evenly distribute, joining extras into the last chunk
        chunk = len(descriptions) // count
        remainder = len(descriptions) % count
        grouped = []
        idx = 0
        for k in range(count):
            take = chunk + (1 if k < remainder else 0)
            part = ' '.join(descriptions[idx:idx+take])
            grouped.append(part)
            idx += take
        descriptions = grouped
    elif len(descriptions) < count:
        # Pad with empty strings if fewer descriptions than expected
        descriptions += [''] * (count - len(descriptions))

    for k in range(count):
        desc = descriptions[k] if k < len(descriptions) else ''
        # Strip reference codes like 000DP015...
        desc = re.sub(r'\b000[A-Z0-9]{8,}\b', '', desc).strip()
        # Strip trailing country abbreviation (e.g. " IN" at end)
        desc = re.sub(r'\s+\bIN\b\s*$', '', desc).strip()
        transactions.append({
            'date': dates[k], 'description': desc,
            'amount': amounts[k], 'direction': directions[k],
            'bank': bank_name, 'card_type': card_type,
        })

    return transactions


# ---------------------------------------------------------------------------
# Parser: HSBC Premier
# ---------------------------------------------------------------------------

def parse_hsbc_transactions(text, bank_name, card_type):
    """
    HSBC layout — each record in pdftotext output:

        13DEC              ← date alone on its line
                           ← blank
        51xx xxxx xxxx 3084 JOHN DOE    ← card+name line (skip) OR merchant
        ADANI DIGITAL LABS PVT Ahmedabad IND  ← merchant (when card line present)
                           ← blank
        13,097.12          ← amount alone on its line

        12DEC
                           ← blank
        AARO FOODS PRIVATE LI MUMBAI    ← merchant (no card line)
                           ← blank
        MAH                ← state/country code (skip)
                           ← blank
        550.00             ← amount

    Strategy: walk line by line. When we see a DDMMM date, the *next* significant
    block until a standalone number is the description (ignoring card-number lines
    and 2-3 letter country/state codes). The standalone number is the amount.
    """
    transactions = []
    lines = text.split('\n')

    # Parse statement period to assign correct years to DDMMM dates
    # e.g. "09 DEC 2025 To 08 JAN 2026"
    period_match = re.search(
        r'(\d{2}\s+[A-Z]{3}\s+(20\d{2}))\s+To\s+(\d{2}\s+[A-Z]{3}\s+(20\d{2}))',
        text, re.I
    )
    month_to_num = {
        'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,
        'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12,
    }
    # Build a dict: month_abbr → year for this statement
    month_year = {}
    if period_match:
        start_year = int(period_match.group(2))
        end_year   = int(period_match.group(4))
        start_month_str = period_match.group(1).split()[1].upper()
        end_month_str   = period_match.group(3).split()[1].upper()
        sm = month_to_num.get(start_month_str, 1)
        em = month_to_num.get(end_month_str, 1)
        # Walk from start to end
        m, y = sm, start_year
        while True:
            month_year[list(month_to_num.keys())[m-1]] = y
            if m == em and y == end_year:
                break
            m += 1
            if m > 12:
                m = 1
                y += 1
            if y > end_year + 1:
                break
    default_year = period_match.group(2) if period_match else str(
        re.search(r'20\d{2}', text).group() if re.search(r'20\d{2}', text) else '2025'
    )

    # Start after "Interest Rate applicable" (which appears right before transactions)
    start_idx = 0
    for idx, l in enumerate(lines):
        if 'Interest Rate applicable' in l:
            start_idx = idx + 1
            break

    date_re  = re.compile(r'^(\d{2}[A-Z]{3})$')
    amt_re   = re.compile(r'^([\d,]+\.\d{2})\s*(CR|DR)?\s*$', re.I)
    # Card line: masked card numbers like "51xx xxxx xxxx 3084 NAME" or "51xx xxxx xxxx 1749 ..."
    card_re  = re.compile(r'^[45X\*][1-9xX\*]xx\s+xxxx\s+xxxx\s+\d{4}', re.I)
    # Country/state code: 2-3 uppercase letters only
    code_re  = re.compile(r'^[A-Z]{2,3}$')

    i = start_idx
    while i < len(lines):
        l = lines[i].strip()

        date_m = date_re.match(l)
        if date_m:
            ddmmm = date_m.group(1)
            mon = ddmmm[2:]  # e.g. 'DEC'
            yr = month_year.get(mon, default_year)
            date = f"{ddmmm} {yr}"
            desc_parts = []
            amount = None
            direction = 'Dr'

            j = i + 1
            while j < len(lines):
                nl = lines[j].strip()
                j += 1

                if not nl:
                    continue

                # Stop at next date
                if date_re.match(nl):
                    j -= 1  # back up so outer loop sees this date
                    break

                # Amount line
                amt_m = amt_re.match(nl)
                if amt_m:
                    amount = amt_m.group(1).replace(',', '')
                    if (amt_m.group(2) or '').upper() == 'CR':
                        direction = 'Cr'
                    break

                # Skip card-number lines and standalone country/state codes
                if card_re.match(nl) or code_re.match(nl):
                    continue

                desc_parts.append(nl)

            description = ' '.join(desc_parts).strip()
            # Auto-detect credits
            cr_keywords = r'payment|credit|refund|reversal|cashback|bbps'
            if re.search(cr_keywords, description, re.I):
                direction = 'Cr'

            if description and amount:
                transactions.append({
                    'date': date, 'description': description,
                    'amount': amount, 'direction': direction,
                    'bank': bank_name, 'card_type': card_type,
                })

            i = j
            continue

        i += 1

    return transactions


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

def extract_all_transactions(pdf_dir):
    """Walk Bank-CardType/ subfolders, route each PDF to the right parser."""
    all_transactions = []
    bank_counts = defaultdict(int)
    processed = 0

    pdf_dir = Path(pdf_dir)
    all_found = sorted(pdf_dir.rglob('*.pdf')) + sorted(pdf_dir.rglob('*.PDF'))

    # Deduplicate: if the same Bank-CardType/filename appears more than once
    # (e.g. from two cc-sort runs into the same folder), keep only the first.
    seen = set()
    pdf_files = []
    for p in all_found:
        key = (p.parent.name, p.name)
        if key not in seen:
            seen.add(key)
            pdf_files.append(p)

    dupes = len(all_found) - len(pdf_files)
    if dupes:
        print(f"[INFO] Skipped {dupes} duplicate PDF(s) (same bank folder + filename)\n")

    print(f"Processing {len(pdf_files)} PDFs:\n")

    for pdf_path in pdf_files:
        bank, card_type = get_bank_and_card_type(pdf_path)
        text = extract_text_from_pdf(pdf_path)
        if not text:
            print(f"  [SKIP] {pdf_path.name} — no text extracted")
            continue

        b = bank.lower()
        ct = card_type.lower()

        if b == 'sbm':
            transactions = parse_sbm_transactions(text, bank, card_type)
        elif b == 'yes':
            transactions = parse_yes_bank_transactions(text, bank, card_type)
        elif b == 'hdfc':
            transactions = parse_hdfc_transactions(text, bank, card_type)
        elif b == 'axis':
            transactions = parse_axis_transactions(text, bank, card_type)
        elif b == 'icici':
            transactions = parse_icici_transactions(text, bank, card_type)
        elif b == 'sbi':
            transactions = parse_sbi_transactions(text, bank, card_type)
        elif b == 'hsbc':
            transactions = parse_hsbc_transactions(text, bank, card_type)
        else:
            transactions = []

        count = len(transactions)
        if count:
            print(f"  {bank:10} {card_type:20}: {count:3} transactions")
            all_transactions.extend(transactions)
            bank_counts[f"{bank}-{card_type}"] += count
            processed += 1
        else:
            print(f"  {bank:10} {card_type:20}: --- (0 transactions) [{pdf_path.name[:50]}]")

    print(f"\nProcessed {processed} PDFs with transactions")
    return all_transactions, bank_counts


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def create_excel_file(transactions, output_path):
    """Write Transactions + Summary sheets to an .xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ['Bank', 'Card Type', 'Date', 'Description', 'Amount', 'Direction']
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    sorted_trans = sorted(transactions, key=lambda x: x.get('date', ''), reverse=True)
    for trans in sorted_trans:
        try:
            amount_val = float(trans['amount']) if trans.get('amount') else 0
        except (ValueError, TypeError):
            amount_val = 0
        ws.append([
            trans['bank'],
            trans['card_type'],
            trans['date'],
            trans['description'],
            amount_val,
            trans['direction'],
        ])

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10

    for row in ws.iter_rows(min_row=2, max_row=len(sorted_trans) + 1, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(['Bank-Card Type', 'Count'])
    summary_data = defaultdict(int)
    for t in transactions:
        summary_data[f"{t['bank']}-{t['card_type']}"] += 1
    for key in sorted(summary_data):
        ws_sum.append([key, summary_data[key]])
    ws_sum.column_dimensions['A'].width = 28
    ws_sum.column_dimensions['B'].width = 10

    wb.save(output_path)
    return len(sorted_trans)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python create_cc_transaction_list.py <pdf_directory> [output_excel]")
        sys.exit(1)

    pdf_directory = sys.argv[1]
    output_excel = sys.argv[2] if len(sys.argv) > 2 else "Consolidated_Transactions.xlsx"

    # If a directory was passed as output, append a default filename
    output_path = Path(output_excel)
    if output_path.is_dir():
        output_excel = str(output_path / "Consolidated_Transactions.xlsx")
        print(f"Output path was a folder — saving to: {output_excel}\n")

    print("=" * 80)
    print("BANK STATEMENT TRANSACTION EXTRACTION")
    print("=" * 80)
    print()

    transactions, bank_counts = extract_all_transactions(pdf_directory)

    print()
    print("=" * 80)
    print(f"TOTAL EXTRACTED: {len(transactions)} transactions")
    print("=" * 80)
    print("\nBreakdown by bank:")
    for bank_key in sorted(bank_counts):
        print(f"  {bank_key:35}: {bank_counts[bank_key]:4}")

    if transactions:
        total = create_excel_file(transactions, output_excel)
        print(f"\n✓ Created: {output_excel}")
        print(f"✓ Spreadsheet contains {total} transactions")
    else:
        print("\n✗ No transactions extracted")
