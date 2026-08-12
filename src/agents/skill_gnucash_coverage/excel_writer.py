"""
excel_writer.py -- writes the Coverage Gap Detector's report workbook.

Two sheets, mirroring skill_ais_reconcile/excel_writer.py's conventions
(same font, same header/confidence fills, same autosize/freeze-pane habits)
so every GnuCash-family skill's output looks and behaves the same way:

  1. "Gaps"    -- one row per suspected gap month: entity, book, account,
                  account class, month, confidence, the account's median
                  (so the grading is auditable), and a prominent TRAILING
                  flag/highlight for gaps after an account's last posting.
  2. "Summary" -- one row per in-scope account with transaction history:
                  first/last transaction date, active-window end, median,
                  confidence, and gap/trailing/suppressed counts.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"

_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
_HEADER_FONT = Font(name=FONT_NAME, bold=True)

_RED_FILL = PatternFill("solid", fgColor="FFC7CE")
_RED_FONT_COLOR = "9C0006"
_YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
_YELLOW_FONT_COLOR = "9C6500"


def _font(bold: bool = False, color: str | None = None) -> Font:
    return Font(name=FONT_NAME, bold=bold, color=color)


def _set(ws, row: int, col: int, value, *, bold: bool = False,
         fill: PatternFill | None = None, color: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, color=color)
    if fill is not None:
        cell.fill = fill


def _write_header(ws, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def _autosize(ws, ncols: int, min_width: int = 10, max_width: int = 44) -> None:
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        longest = 0
        for cell in ws[letter]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, longest + 2))


def _write_gaps_sheet(wb: Workbook, gaps: list) -> None:
    ws = wb.create_sheet("Gaps")
    headers = ["Entity", "Book", "Account", "Account Class", "Month",
               "Confidence", "Account Median (txns/mo)", "Trailing Gap"]
    _write_header(ws, 1, headers)

    row = 2
    if not gaps:
        _set(ws, row, 1, "No suspected coverage gaps found in the selected book(s).")
        _autosize(ws, len(headers))
        return

    # Trailing gaps first (highest-value signal), then interior gaps; within
    # each, HIGH confidence first -- keeps the most actionable rows on top.
    ordered = sorted(
        gaps,
        key=lambda g: (not g.trailing, g.confidence != "HIGH", g.entity, g.book,
                       g.account_path, g.month),
    )
    for g in ordered:
        is_high = g.confidence == "HIGH"
        fill = _RED_FILL if (g.trailing and is_high) else (_YELLOW_FILL if is_high else None)
        color = (_RED_FONT_COLOR if (g.trailing and is_high)
                 else (_YELLOW_FONT_COLOR if is_high else None))
        _set(ws, row, 1, g.entity)
        _set(ws, row, 2, g.book)
        _set(ws, row, 3, g.account_path)
        _set(ws, row, 4, g.account_class)
        _set(ws, row, 5, g.month)
        _set(ws, row, 6, g.confidence, fill=fill, color=color, bold=is_high)
        _set(ws, row, 7, round(g.median, 2))
        _set(ws, row, 8, "TRAILING" if g.trailing else "",
             fill=(_RED_FILL if g.trailing else None),
             color=(_RED_FONT_COLOR if g.trailing else None),
             bold=g.trailing)
        row += 1

    _autosize(ws, len(headers))


def _write_summary_sheet(wb: Workbook, stats: list) -> None:
    ws = wb.create_sheet("Summary")
    headers = ["Entity", "Book", "Account", "Account Class", "First Txn",
               "Last Txn", "Window End", "Median (txns/mo)", "Confidence",
               "Zero Months", "Trailing Zero Months", "FY-Boundary Suppressed"]
    _write_header(ws, 1, headers)

    row = 2
    for s in sorted(stats, key=lambda s: (s.entity, s.book, s.account_path)):
        _set(ws, row, 1, s.entity)
        _set(ws, row, 2, s.book)
        _set(ws, row, 3, s.account_path)
        _set(ws, row, 4, s.account_class)
        _set(ws, row, 5, s.first_date)
        _set(ws, row, 6, s.last_date)
        _set(ws, row, 7, s.window_end)
        _set(ws, row, 8, round(s.median, 2))
        _set(ws, row, 9, s.confidence, bold=(s.confidence == "HIGH"))
        _set(ws, row, 10, s.zero_months)
        _set(ws, row, 11, s.trailing_zero_months,
             bold=s.trailing_zero_months > 0,
             color=(_RED_FONT_COLOR if s.trailing_zero_months > 0 else None))
        _set(ws, row, 12, s.suppressed_boundary_gaps)
        row += 1

    if not stats:
        _set(ws, row, 1, "No in-scope accounts with transaction history were found.")

    _autosize(ws, len(headers))


def write_gaps_workbook(gaps: list, stats: list, out_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _write_gaps_sheet(wb, gaps)
    _write_summary_sheet(wb, stats)
    wb.save(out_path)
