"""
cg_writer.py -- writes the KFintech Capital Gain / Loss Statement skill's
report workbook + CSV siblings.

Separate from excel_writer.py deliberately -- excel_writer.py writes the
older CAS-text/FIFO-derived shape (Transactions/Holdings/RealisedGains/
Exceptions, driven by parser.SchemeBlock + lots.SchemeReconciliation) and is
kept intact for that still-supported path (see AGENT.md). This module
writes the RTA-already-matched shape driven by cg_parser's dataclasses.
Same styling conventions (font, header fill, autosize/freeze-pane helpers)
reused from excel_writer.py so both skill outputs look alike.

Five sheets:
  1. "MatchedLots"     -- every RTA-matched acquisition/outflow/gain row,
                           as-is (one row per consumed buy lot).
  2. "SchemeSummary"    -- Scheme_Level_Summary, as given.
  3. "PeriodSummary"    -- the five s.234C bucket figures per section, from
                           both Summary - Equity and Summary - NonEquity.
  4. "Reconciliation"   -- the three-way agree/variance/cannot-reconcile
                           result per gain category.
  5. "Exceptions"       -- reconciliation categories that did not agree or
                           could not be reconciled, collected in one place.

CSV sibling for MatchedLots is written alongside the XLSX (same stem,
"_matched_lots.csv" suffix), mirroring excel_writer.py's sidecar pattern.

No PAN, investor name, address, email, or mobile number is ever written
here -- cg_parser.parse_transaction_details never reads rows 1-2 of
Trasaction_Details (where that PII lives) into any field in the first
place, so there is nothing to filter at this layer.
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .cg_parser import (
    MatchedLotRow, ReconciliationResult, SchemeSummaryRow, SummarySheetData,
)

FONT_NAME = "Arial"

_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
_HEADER_FONT = Font(name=FONT_NAME, bold=True)

_RED_FILL = PatternFill("solid", fgColor="FFC7CE")
_RED_FONT_COLOR = "9C0006"
_YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
_YELLOW_FONT_COLOR = "9C6500"


def _font(bold: bool = False, color: str | None = None) -> Font:
    return Font(name=FONT_NAME, bold=bold, color=color)


def _set(ws, row: int, col: int, value, *, bold: bool = False,
         fill: PatternFill | None = None, color: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, color=color)
    if fill is not None:
        cell.fill = fill


def _write_header(ws, row: int, headers: list) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def _autosize(ws, ncols: int, min_width: int = 10, max_width: int = 44) -> None:
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        longest = 0
        for cell in ws[letter]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, longest + 2))


_MATCHED_LOT_HEADERS = [
    "Fund", "Fund Name", "Folio", "Scheme", "ISIN",
    "Acq Trxn Type", "Acq Date", "Current Units", "Source Scheme Units",
    "Original Purchase Cost", "Original Cost Amount",
    "Grandfathered NAV", "Grandfathered Cost Value",
    "IT Applicable NAV", "IT Applicable Cost Value",
    "Outflow Trxn Type", "Outflow Date", "Units", "Amount", "Price",
    "Tax Perc", "Tax", "Short Term", "Indexed Cost",
    "Long Term With Index", "Long Term Without Index",
]


def _matched_lot_values(m: MatchedLotRow) -> list:
    return [
        m.fund_code, m.fund_name, m.folio, m.scheme_name, m.isin,
        m.acq_trxn_type, m.acq_date.isoformat() if m.acq_date else "",
        m.current_units, m.source_scheme_units,
        m.original_purchase_cost, m.original_cost_amount,
        m.grandfathered_nav, m.grandfathered_cost_value,
        m.it_applicable_nav, m.it_applicable_cost_value,
        m.outflow_trxn_type, m.outflow_date.isoformat() if m.outflow_date else "",
        m.units, m.amount, m.price, m.tax_perc, m.tax,
        m.short_term, m.indexed_cost, m.long_term_with_index, m.long_term_without_index,
    ]


def _write_matched_lots_sheet(wb: Workbook, rows: list) -> None:
    ws = wb.create_sheet("MatchedLots")
    _write_header(ws, 1, _MATCHED_LOT_HEADERS)
    row = 2
    for m in rows:
        for col, val in enumerate(_matched_lot_values(m), start=1):
            _set(ws, row, col, val)
        row += 1
    if row == 2:
        _set(ws, row, 1, "No matched lots parsed.")
    _autosize(ws, len(_MATCHED_LOT_HEADERS))


def _write_scheme_summary_sheet(wb: Workbook, rows: list, total: SchemeSummaryRow | None) -> None:
    ws = wb.create_sheet("SchemeSummary")
    headers = ["Scheme", "ISIN", "Count", "Outflow Amount", "Net Value",
               "Fair Market NAV", "Fair Market Value", "Short Gain",
               "Long Gain With Index", "Long Gain Without Index"]
    _write_header(ws, 1, headers)
    row = 2
    for s in rows:
        _set(ws, row, 1, s.scheme_name)
        _set(ws, row, 2, s.isin)
        _set(ws, row, 3, s.count)
        _set(ws, row, 4, s.outflow_amount)
        _set(ws, row, 5, s.net_value)
        _set(ws, row, 6, s.fair_market_nav)
        _set(ws, row, 7, s.fair_market_value)
        _set(ws, row, 8, s.short_gain)
        _set(ws, row, 9, s.long_gain_with_index)
        _set(ws, row, 10, s.long_gain_without_index)
        row += 1
    if total is not None:
        _set(ws, row, 1, "Total", bold=True)
        _set(ws, row, 3, total.count, bold=True)
        _set(ws, row, 4, total.outflow_amount, bold=True)
        _set(ws, row, 5, total.net_value, bold=True)
        _set(ws, row, 6, total.fair_market_nav, bold=True)
        _set(ws, row, 7, total.fair_market_value, bold=True)
        _set(ws, row, 8, total.short_gain, bold=True)
        _set(ws, row, 9, total.long_gain_with_index, bold=True)
        _set(ws, row, 10, total.long_gain_without_index, bold=True)
        row += 1
    if row == 2:
        _set(ws, row, 1, "No scheme-level summary rows parsed.")
    _autosize(ws, len(headers))


def _write_period_summary_sheet(wb: Workbook, equity: SummarySheetData, nonequity: SummarySheetData) -> None:
    ws = wb.create_sheet("PeriodSummary")
    headers = ["Sheet", "Section", "Row Label", "01/04-15/06", "16/06-15/09",
               "16/09-15/12", "16/12-15/03", "16/03-31/03", "Total"]
    _write_header(ws, 1, headers)
    row = 2
    for sheet_data in (equity, nonequity):
        for r in sheet_data.rows:
            _set(ws, row, 1, sheet_data.sheet_label)
            _set(ws, row, 2, r.section)
            _set(ws, row, 3, r.label)
            buckets = list(r.buckets.values())
            for i, v in enumerate(buckets):
                _set(ws, row, 4 + i, v)
            _set(ws, row, 9, r.total)
            row += 1
    if row == 2:
        _set(ws, row, 1, "No period summary rows parsed.")
    _autosize(ws, len(headers))


def _write_reconciliation_sheet(wb: Workbook, results: list) -> None:
    ws = wb.create_sheet("Reconciliation")
    headers = ["Category", "Trasaction_Details Sum", "Scheme_Level_Summary Total",
               "Summary Sheets Total", "Agreement", "Note"]
    _write_header(ws, 1, headers)
    row = 2
    for r in results:
        undecidable = r.agree is None
        breach = r.agree is False
        _set(ws, row, 1, r.category)
        _set(ws, row, 2, r.txn_details_total)
        _set(ws, row, 3, r.scheme_summary_total)
        _set(ws, row, 4, r.summary_sheet_total)
        status = "CANNOT RECONCILE" if undecidable else ("AGREE" if r.agree else "VARIANCE")
        _set(ws, row, 5, status,
             fill=(_YELLOW_FILL if undecidable else (_RED_FILL if breach else None)),
             color=(_YELLOW_FONT_COLOR if undecidable else (_RED_FONT_COLOR if breach else None)),
             bold=(undecidable or breach))
        _set(ws, row, 6, r.note or "")
        row += 1
    _autosize(ws, len(headers))


def _write_exceptions_sheet(wb: Workbook, results: list) -> None:
    ws = wb.create_sheet("Exceptions")
    headers = ["Category", "Status", "Detail"]
    _write_header(ws, 1, headers)
    row = 2
    for r in results:
        if r.agree is False:
            _set(ws, row, 1, r.category)
            _set(ws, row, 2, "VARIANCE", fill=_RED_FILL, color=_RED_FONT_COLOR, bold=True)
            _set(ws, row, 3, r.note or "")
            row += 1
        elif r.agree is None:
            _set(ws, row, 1, r.category)
            _set(ws, row, 2, "CANNOT RECONCILE", fill=_YELLOW_FILL, color=_YELLOW_FONT_COLOR, bold=True)
            _set(ws, row, 3, r.note or "")
            row += 1
    if row == 2:
        _set(ws, row, 1, "No exceptions -- all reconciliation categories agreed.")
    _autosize(ws, len(headers))


def write_report_workbook(
    matched_lots: list,
    scheme_rows: list,
    scheme_total,
    summary_equity: SummarySheetData,
    summary_nonequity: SummarySheetData,
    reconciliation: list,
    out_path: str,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _write_matched_lots_sheet(wb, matched_lots)
    _write_scheme_summary_sheet(wb, scheme_rows, scheme_total)
    _write_period_summary_sheet(wb, summary_equity, summary_nonequity)
    _write_reconciliation_sheet(wb, reconciliation)
    _write_exceptions_sheet(wb, reconciliation)
    wb.save(out_path)

    stem = Path(out_path).with_suffix("")
    _write_matched_lots_csv(matched_lots, f"{stem}_matched_lots.csv")


def _write_matched_lots_csv(rows: list, out_path: str) -> None:
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(_MATCHED_LOT_HEADERS)
        for m in rows:
            w.writerow(_matched_lot_values(m))
