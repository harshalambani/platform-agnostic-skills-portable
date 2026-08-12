"""
excel_writer.py -- writes the MF CAS skill's report workbook + CSV siblings.

Follows skill_gnucash_coverage/excel_writer.py's conventions (same font,
same header fill, same autosize/freeze-pane helpers) so this skill's output
looks and behaves like the rest of the GnuCash/ITR-family skills.

Four sheets:
  1. "Transactions"   -- every parsed transaction row, as-is.
  2. "Holdings"        -- one row per folio+scheme: opening/closing units,
                           reconciliation OK/breach.
  3. "RealisedGains"   -- one row per CONSUMED FIFO lot (raw facts only --
                           no tax/rate/LT-ST computation, see AGENT.md).
  4. "Exceptions"      -- every flagged row (unattributed, grandfathering,
                           reconciliation breach, RTA variance) collected in
                           one place for a reviewer to work through.

CSV siblings for Transactions and RealisedGains are written alongside the
XLSX (same stem, "_transactions.csv" / "_realised_gains.csv" suffix) since
registry.Output only carries a single extension -- mirrors the sidecar-file
pattern used elsewhere in this repo for multi-artifact skills.

No PAN, investor name, address, email, or mobile number is ever written
here -- the upstream parser never captures them in the first place (see
parser.py's module docstring), so there is nothing to filter at this layer.
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .lots import SchemeReconciliation

FONT_NAME = "Arial"

_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
_HEADER_FONT = Font(name=FONT_NAME, bold=True)

_RED_FILL = PatternFill("solid", fgColor="FFC7CE")
_RED_FONT_COLOR = "9C0006"
_YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
_YELLOW_FONT_COLOR = "9C6500"


def _font(bold: bool = False, color: str | None = None) -> Font:
    return Font(name=FONT_NAME, bold=bold, color=color)


def _set(ws, row: int, col: int, value, *, bold: bool = False,
         fill: PatternFill | None = None, color: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, color=color)
    if fill is not None:
        cell.fill = fill


def _write_header(ws, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def _autosize(ws, ncols: int, min_width: int = 10, max_width: int = 44) -> None:
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        longest = 0
        for cell in ws[letter]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, longest + 2))


def _write_transactions_sheet(wb: Workbook, schemes) -> None:
    ws = wb.create_sheet("Transactions")
    headers = ["Folio", "AMC", "Registrar", "Scheme", "ISIN", "Scheme Type",
               "Date", "Description", "Amount", "Units", "NAV", "Balance", "Kind"]
    _write_header(ws, 1, headers)
    row = 2
    for s in schemes:
        for t in s.transactions:
            _set(ws, row, 1, s.folio)
            _set(ws, row, 2, s.amc)
            _set(ws, row, 3, s.rta)
            _set(ws, row, 4, s.scheme_name)
            _set(ws, row, 5, s.isin)
            _set(ws, row, 6, s.scheme_type)
            _set(ws, row, 7, t.date.isoformat())
            _set(ws, row, 8, t.description)
            _set(ws, row, 9, t.amount)
            _set(ws, row, 10, t.units)
            _set(ws, row, 11, t.nav)
            _set(ws, row, 12, t.balance)
            _set(ws, row, 13, t.kind)
            row += 1
    if row == 2:
        _set(ws, row, 1, "No transactions parsed.")
    _autosize(ws, len(headers))


def _write_holdings_sheet(wb: Workbook, recons: list[SchemeReconciliation]) -> None:
    ws = wb.create_sheet("Holdings")
    headers = ["Folio", "Scheme", "ISIN", "Expected Closing Units",
               "Statement Closing Units", "Units Diff", "Reconciliation"]
    _write_header(ws, 1, headers)
    row = 2
    for r in sorted(recons, key=lambda r: (r.folio, r.scheme)):
        breach = not r.units_ok
        _set(ws, row, 1, r.folio)
        _set(ws, row, 2, r.scheme)
        _set(ws, row, 3, r.isin)
        _set(ws, row, 4, r.units_expected_closing)
        _set(ws, row, 5, r.units_actual_closing)
        _set(ws, row, 6, r.units_diff)
        _set(ws, row, 7, "OK" if r.units_ok else "BREACH",
             fill=(_RED_FILL if breach else None),
             color=(_RED_FONT_COLOR if breach else None), bold=breach)
        row += 1
    if row == 2:
        _set(ws, row, 1, "No folios/schemes parsed.")
    _autosize(ws, len(headers))


def _write_realised_gains_sheet(wb: Workbook, recons: list[SchemeReconciliation]) -> None:
    ws = wb.create_sheet("RealisedGains")
    headers = ["Folio", "Scheme", "ISIN", "Scheme Type", "Units", "Buy Date",
               "Buy NAV", "Buy Cost", "Sell Date", "Sell NAV", "Sale Proceeds",
               "Holding Days", "Gain", "Flags"]
    _write_header(ws, 1, headers)
    row = 2
    for r in recons:
        for dl in r.disposal_lots:
            flagged = bool(dl.flags)
            _set(ws, row, 1, dl.folio)
            _set(ws, row, 2, dl.scheme)
            _set(ws, row, 3, dl.isin)
            _set(ws, row, 4, dl.scheme_type)
            _set(ws, row, 5, dl.units)
            _set(ws, row, 6, dl.buy_date.isoformat() if dl.buy_date else "")
            _set(ws, row, 7, dl.buy_nav)
            _set(ws, row, 8, dl.buy_cost)
            _set(ws, row, 9, dl.sell_date.isoformat())
            _set(ws, row, 10, dl.sell_nav)
            _set(ws, row, 11, dl.sale_proceeds)
            _set(ws, row, 12, dl.holding_days)
            _set(ws, row, 13, dl.gain)
            _set(ws, row, 14, "; ".join(dl.flags),
                 fill=(_YELLOW_FILL if flagged else None),
                 color=(_YELLOW_FONT_COLOR if flagged else None), bold=flagged)
            row += 1
    if row == 2:
        _set(ws, row, 1, "No disposals found -- no realised gains to report.")
    _autosize(ws, len(headers))


def _write_exceptions_sheet(wb: Workbook, recons: list[SchemeReconciliation]) -> None:
    ws = wb.create_sheet("Exceptions")
    headers = ["Folio", "Scheme", "Category", "Detail"]
    _write_header(ws, 1, headers)
    row = 2
    for r in recons:
        if not r.units_ok:
            _set(ws, row, 1, r.folio); _set(ws, row, 2, r.scheme)
            _set(ws, row, 3, "Units reconciliation breach",
                 fill=_RED_FILL, color=_RED_FONT_COLOR, bold=True)
            _set(ws, row, 4, f"expected closing={r.units_expected_closing}, "
                              f"statement closing={r.units_actual_closing}, "
                              f"diff={r.units_diff}")
            row += 1
        if not r.matched_vs_disposed_ok:
            _set(ws, row, 1, r.folio); _set(ws, row, 2, r.scheme)
            _set(ws, row, 3, "Matched-vs-disposed units breach",
                 fill=_RED_FILL, color=_RED_FONT_COLOR, bold=True)
            _set(ws, row, 4, "sum of matched lot units != disposed units "
                              f"(unattributed shortfall={r.unattributed_shortfall_units})")
            row += 1
        if r.rta_gain_note:
            # Explicit "cannot cross-check" state -- must never render as a
            # blank/absent row, which would read as "no variance found"
            # when the truth is "not derivable from this statement".
            _set(ws, row, 1, r.folio); _set(ws, row, 2, r.scheme)
            _set(ws, row, 3, "RTA realised-gain cross-check",
                 fill=_YELLOW_FILL, color=_YELLOW_FONT_COLOR, bold=True)
            _set(ws, row, 4, f"RTA reported={r.rta_gain_reported}, {r.rta_gain_note}")
            row += 1
        elif r.rta_gain_variance not in (None, 0) and abs(r.rta_gain_variance) > 0.01:
            _set(ws, row, 1, r.folio); _set(ws, row, 2, r.scheme)
            _set(ws, row, 3, "RTA realised-gain variance",
                 fill=_YELLOW_FILL, color=_YELLOW_FONT_COLOR, bold=True)
            _set(ws, row, 4, f"RTA reported={r.rta_gain_reported}, "
                              f"FIFO derived={r.rta_gain_derived}, "
                              f"variance={r.rta_gain_variance}")
            row += 1
        for dl in r.disposal_lots:
            for flag in dl.flags:
                _set(ws, row, 1, dl.folio); _set(ws, row, 2, dl.scheme)
                _set(ws, row, 3, flag,
                     fill=_YELLOW_FILL, color=_YELLOW_FONT_COLOR, bold=True)
                _set(ws, row, 4,
                     f"sell {dl.sell_date.isoformat()}, units={dl.units}, "
                     f"buy_date={dl.buy_date.isoformat() if dl.buy_date else 'unknown'}")
                row += 1
    if row == 2:
        _set(ws, row, 1, "No exceptions -- every disposal was fully attributed and reconciled.")
    _autosize(ws, len(headers))


def write_report_workbook(schemes, recons: list[SchemeReconciliation], out_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _write_transactions_sheet(wb, schemes)
    _write_holdings_sheet(wb, recons)
    _write_realised_gains_sheet(wb, recons)
    _write_exceptions_sheet(wb, recons)
    wb.save(out_path)

    stem = Path(out_path).with_suffix("")
    _write_transactions_csv(schemes, f"{stem}_transactions.csv")
    _write_realised_gains_csv(recons, f"{stem}_realised_gains.csv")


def _write_transactions_csv(schemes, out_path: str) -> None:
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Folio", "AMC", "Registrar", "Scheme", "ISIN", "Scheme Type",
                    "Date", "Description", "Amount", "Units", "NAV", "Balance", "Kind"])
        for s in schemes:
            for t in s.transactions:
                w.writerow([s.folio, s.amc, s.rta, s.scheme_name, s.isin, s.scheme_type,
                            t.date.isoformat(), t.description, t.amount, t.units, t.nav,
                            t.balance, t.kind])


def _write_realised_gains_csv(recons: list[SchemeReconciliation], out_path: str) -> None:
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Folio", "Scheme", "ISIN", "Scheme Type", "Units", "Buy Date",
                    "Buy NAV", "Buy Cost", "Sell Date", "Sell NAV", "Sale Proceeds",
                    "Holding Days", "Gain", "Flags"])
        for r in recons:
            for dl in r.disposal_lots:
                w.writerow([
                    dl.folio, dl.scheme, dl.isin, dl.scheme_type, dl.units,
                    dl.buy_date.isoformat() if dl.buy_date else "", dl.buy_nav, dl.buy_cost,
                    dl.sell_date.isoformat(), dl.sell_nav, dl.sale_proceeds,
                    dl.holding_days, dl.gain, "; ".join(dl.flags),
                ])
