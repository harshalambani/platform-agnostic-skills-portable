"""
agent.py -- KFintech Capital Gain / Loss Statement skill (mutual funds).
DIRECT mode, no LLM, no network.

See AGENT.md for why this directory is still named skill_mf_cas even though
it no longer parses a CAS (Consolidated Account Statement) -- in short, the
real document a user actually receives for this purpose is KFintech's
"Capital Gain / Loss Statement", a different, RTA-already-matched xlsx, not
the raw transaction-history CAS this skill originally targeted. The old CAS
PDF path silently produced an empty report on every real document (0
schemes, 0 transactions, no error) because no real CAS actually has the text
shape `parser.parse_cas_text`'s regexes expect -- this rewrite fixes that
silent-empty defect by validating workbook shape up front and failing loud.

Architecture is split hard at the file-access boundary, per this skill's
original design brief and preserved here: `_load_workbook` below is the ONLY
function in this package that touches msoffcrypto, a password, or the
filesystem. Everything downstream -- `cg_parser`'s parse/reconcile functions
-- is pure, takes an in-memory openpyxl Workbook (or synthetic
openpyxl-built workbooks in tests), and never touches I/O.

`parser.py` (CAS-text regex parser) and `lots.py` (FIFO lot derivation) are
INTENTIONALLY UNTOUCHED and UNUSED on this path -- the KFintech Capital
Gain / Loss Statement already contains RTA-matched buy/sell lots and
RTA-computed gains, so there is no FIFO reconstruction to do here. Both
modules and their existing tests are kept intact for a possible future
CAMS/MF-Central transaction-level statement, which would need exactly that
FIFO-from-raw-transactions logic (see AGENT.md).

N1 assumption (unchanged from the original skill): output lands as
STANDALONE ARTIFACTS ONLY. No .gnucash writes, no ITR workbook injection.

Out of scope (unchanged, now explicit at the file-type boundary): the
`*_TXN.pdf` depository CAS PDF and any PDF parsing of the CG statement are
not supported -- a `.pdf` input fails loud with a message pointing at the
xlsx requirement. This skill also only covers KFINTECH-serviced funds --
see skill.yaml's help text for the CAMS-fund caveat.
"""
from __future__ import annotations

from pathlib import Path

from agents.bank_common import password as _password

from .cg_parser import (
    CGParseError,
    parse_scheme_level_summary,
    parse_summary_sheet,
    parse_transaction_details,
    reconcile,
    validate_workbook_shape,
    SHEET_SUMMARY_EQUITY,
    SHEET_SUMMARY_NONEQUITY,
    SHEET_SCHEME_SUMMARY,
    SHEET_TXN_DETAILS,
    SHEET_TXN_DETAILS_ALT,
)
from .cg_parser import find_sheet
from .cg_writer import write_report_workbook


def _load_workbook(xlsx_path: str, password: str | None):
    """The ONLY function in this package that touches msoffcrypto, a
    password, or the filesystem for the xlsx path. Returns an in-memory
    openpyxl Workbook (data_only=True -- formula results, not formulas).

    Raises ValueError (never leaking the password) on a wrong/missing
    password, or if the file is not actually an encrypted OOXML workbook
    msoffcrypto can decrypt.
    """
    import io

    import msoffcrypto
    import openpyxl

    with open(xlsx_path, "rb") as f:
        try:
            office_file = msoffcrypto.OfficeFile(f)
            office_file.load_key(password=password or "", verify_password=True)
            buf = io.BytesIO()
            office_file.decrypt(buf)
        except Exception as e:
            if _password.is_password_error(e):
                raise ValueError(
                    _password.password_error_message(
                        "the statement's own open password, as supplied by KFintech",
                        doc_type="xlsx",
                    )
                ) from e
            raise ValueError(
                f"Could not open this file as an encrypted xlsx workbook: {e}"
            ) from e

    buf.seek(0)
    try:
        return openpyxl.load_workbook(buf, data_only=True)
    except Exception as e:
        raise ValueError(
            f"Decryption succeeded but the result is not a readable xlsx workbook: {e}"
        ) from e


def run(
    cas_path: str,
    output_path: str,
    config_path: str = None,
    model_override: str = None,
    cas_password: str = None,
) -> str:
    """Parse a KFintech Capital Gain / Loss Statement (encrypted .xlsx),
    cross-check its three internally-redundant figures (Trasaction_Details
    row sums, Scheme_Level_Summary totals, Summary sheet totals), and write
    a report workbook (MatchedLots/SchemeSummary/PeriodSummary/
    Reconciliation/Exceptions) plus a MatchedLots CSV sibling to
    output_path. Returns a text summary.

    Parameter names are unchanged from the skill's original CAS-PDF
    incarnation (cas_path/cas_password) for skill.yaml run_args
    compatibility -- they now refer to the xlsx statement and its password.
    """
    input_path = Path(cas_path)
    if not input_path.is_file():
        return f"ERROR: file not found: {cas_path}"

    if input_path.suffix.lower() == ".pdf":
        return (
            "ERROR: this skill no longer accepts a PDF. It now parses the KFintech "
            "Capital Gain / Loss Statement, which KFintech issues as an encrypted "
            ".xlsx workbook, not a PDF. Depository *_TXN.pdf files and PDF parsing "
            "of the CG statement are out of scope -- please supply the .xlsx statement."
        )
    if input_path.suffix.lower() != ".xlsx":
        return (
            f"ERROR: unsupported file type '{input_path.suffix}'. This skill accepts "
            "only the KFintech Capital Gain / Loss Statement as an .xlsx file."
        )

    try:
        wb = _load_workbook(str(input_path), cas_password)
    except ValueError as e:
        return f"ERROR: {e}"

    try:
        validate_workbook_shape(wb)
        ws_equity = find_sheet(wb, SHEET_SUMMARY_EQUITY)
        ws_nonequity = find_sheet(wb, SHEET_SUMMARY_NONEQUITY)
        ws_scheme = find_sheet(wb, SHEET_SCHEME_SUMMARY)
        ws_txn = find_sheet(wb, SHEET_TXN_DETAILS, SHEET_TXN_DETAILS_ALT)

        summary_equity = parse_summary_sheet(ws_equity, "Equity")
        summary_nonequity = parse_summary_sheet(ws_nonequity, "NonEquity")
        scheme_rows, scheme_total = parse_scheme_level_summary(ws_scheme)
        matched_lots = parse_transaction_details(ws_txn)
    except CGParseError as e:
        return f"ERROR: {e}"

    reconciliation = reconcile(
        summary_equity, summary_nonequity, scheme_rows, scheme_total, matched_lots,
    )

    write_report_workbook(
        matched_lots, scheme_rows, scheme_total,
        summary_equity, summary_nonequity, reconciliation, output_path,
    )

    variances = [r for r in reconciliation if r.agree is False]
    undecidable = [r for r in reconciliation if r.agree is None]

    lines_out = [
        f"KFintech Capital Gain / Loss Statement parsed -- {len(matched_lots)} "
        f"matched-lot row(s), {len(scheme_rows)} scheme(s).",
    ]
    if variances:
        lines_out.append(
            f"  WARNING: three-way reconciliation variance in {len(variances)} "
            "category(ies) -- see Reconciliation/Exceptions sheets."
        )
    if undecidable:
        lines_out.append(
            f"  NOTE: {len(undecidable)} category(ies) could not be reconciled -- "
            "see Reconciliation/Exceptions sheets for the explicit reason."
        )
    if not variances and not undecidable:
        lines_out.append("  Three-way reconciliation: all categories agree.")
    lines_out.append(f"  Workbook: {output_path}")
    return "\n".join(lines_out)
