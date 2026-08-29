"""
writer.py -- writes the Partner Compensation Reconciliation workbook.

The only module in this package that touches openpyxl. Takes an
engine.Report (pure data) and writes a 10-sheet workbook:

  1. Logic            -- the seven legs and the governing identity, in prose.
  2. Drivers           -- every rate/period/date as a labelled amber input
                           cell; the ONLY sheet where a rate appears as a
                           literal. Other sheets reference these cells by
                           formula rather than repeating the literal.
  3. Monthly grid       -- one row per line item, one column per month, plus
                           a total column (a live =SUM(...) formula).
  4. Payroll stream     -- only written if the input supplied payroll rows;
                           omitted entirely otherwise.
  5. One-offs           -- the gross-up and the roundness indicator.
  6. Cohorts            -- the incentive cohort ledger, with the FY-assignment
                           (reporting/prior/future) column.
  7. Capital            -- the rule as a formula off the Drivers sheet, the
                           Advisory comparison, and the rate-change detector.
  8. Reconciliation     -- the leg-vs-leg matrix: agree / variance /
                           cannot-reconcile (skill_mf_cas idiom).
  9. Exceptions         -- every non-agreeing / non-reconcilable row,
                           collected in one place.
  10. Open items        -- anything the engine could not resolve, with what
                           would close it.

Style vocabulary (given literally in the spec this package was built from
-- do not invent alternates):

    H  = bold white text, size 10        (header row text)
    HF = navy solid fill "1F3864"        (header row fill)
    SF = pale-blue solid fill "D9E2F3"   (section band)
    TF = amber solid fill "FFF2CC"       (total / driver-input cell)
    OK = green solid fill "C6EFCE"
    BAD= red solid fill "FFC7CE"
    N  = '#,##0;(#,##0);"-"'             (number format)
    P  = '0.0000%'                       (percentage format)

Repo convention (s.6.1 of the spec this was built from): only leaf items
are ever hardcoded; every total, subtotal and cross-foot is a live Excel
formula (=SUM(...), =C8*C9, ...) so the workbook recomputes when a driver
cell is edited.

The "=" trap (s.6.2): openpyxl stores any string cell that *starts* with
"=" as a formula, even plain text like "= Gross incentive" -- producing a
workbook that saves without error but that Excel reports as damaged. Every
label in this module is routed through `_text()`, which inserts a leading
space in front of a literal leading "=" so it is stored as ordinary text.
tests/test_skill_partner_comp_recon.py validates the saved workbook's raw
XML to guard against a regression of this exact defect.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .engine import Report

FONT_NAME = "Arial"

H = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
HF = PatternFill("solid", fgColor="1F3864")
SF = PatternFill("solid", fgColor="D9E2F3")
TF = PatternFill("solid", fgColor="FFF2CC")
OK = PatternFill("solid", fgColor="C6EFCE")
BAD = PatternFill("solid", fgColor="FFC7CE")
N = '#,##0;(#,##0);"-"'
P = '0.0000%'


def _text(value):
    """Route every label/string through here. A literal leading '=' would
    otherwise be stored by openpyxl as a formula (the '=' trap -- see this
    module's docstring); insert a leading space to keep it plain text."""
    if isinstance(value, str) and value.startswith("="):
        return " " + value
    return value


def _set(ws, row, col, value, *, font=None, fill=None, number_format=None,
         bold=False, wrap=False):
    cell = ws.cell(row=row, column=col, value=_text(value))
    if font is not None:
        cell.font = font
    elif bold:
        cell.font = Font(name=FONT_NAME, bold=True)
    else:
        cell.font = Font(name=FONT_NAME)
    if fill is not None:
        cell.fill = fill
    if number_format is not None:
        cell.number_format = number_format
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def _write_header(ws, row, headers):
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=_text(text))
        cell.font = H
        cell.fill = HF
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def _autosize(ws, ncols, min_width=10, max_width=48):
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        longest = 0
        for cell in ws[letter]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, longest + 2))


def _status_fill(agree):
    if agree is True:
        return OK, "AGREE"
    if agree is False:
        return BAD, "VARIANCE"
    return TF, "CANNOT RECONCILE"


# ---------------------------------------------------------------------------
# 1. Logic
# ---------------------------------------------------------------------------

def _write_logic_sheet(wb, report: Report):
    ws = wb.create_sheet("Logic")
    lines = [
        ("Partner Compensation Reconciliation -- Logic", True),
        (f"Financial year: {report.financial_year}", False),
        ("", False),
        ("The seven legs (each an independent source; see Reconciliation):", True),
        ("1. Year-end position -- Compensation Advisory + LLP capital/current account statement.", False),
        ("2. Incentive schedule -- the firm's payment-schedule PDF (per-month grid).", False),
        ("3. Monthly payouts -- 12 monthly payout advices, plus payslips for any payroll months.", False),
        ("4. Capital contribution -- derived by rule, cross-checked to the Advisory's stated closing.", False),
        ("5. Bank -- the partner's bank statement credits.", False),
        ("6. One-offs -- special incentive / ex-gratia, appearing on a single month's advice.", False),
        ("7. The return -- filed ITR + computation + Form 26AS + Form 16.", False),
        ("", False),
        ("Governing identity (derived every month, never read off the advice):", True),
        ("misc = total_paid - remuneration - share_of_profit_gross - additional_share_of_profit", False),
        ("", False),
        ("A one-off ('Additional Share of Profit') is shown net of the firm's tax; it is", False),
        ("grossed up as gross = net / (1 - firms_tax_rate), then checked for roundness", False),
        ("(distance to the nearest Rs 100,000) -- a one-off is always awarded round.", False),
        ("", False),
        ("Capital rule: required_cumulative_capital = target_compensation *", False),
        ("(months_achieved / months_total) * capital_rate. Deducted only when an", False),
        ("instalment is actually paid, never when awarded.", False),
        ("", False),
        ("Non-goals: no tax computation, no .gnucash writes, no ITR workbook injection.", True),
        ("No rate, percentage or period is ever a constant -- every one is read from the", False),
        ("Drivers sheet / this run's input, and a missing one produces an explicit", False),
        ("CANNOT RECONCILE row rather than a default or a guess.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        _set(ws, i, 1, text, bold=bold, wrap=True)
    ws.column_dimensions["A"].width = 100


# ---------------------------------------------------------------------------
# 2. Drivers -- the only sheet a rate/period/date is ever a literal on.
# ---------------------------------------------------------------------------

def _write_drivers_sheet(wb, report: Report):
    ws = wb.create_sheet("Drivers")
    _write_header(ws, 1, ["Driver", "Value", "Format"])
    d = report.drivers
    rows = [
        ("Financial year", report.financial_year, None),
        ("Firm's tax rate", d.get("firms_tax_rate"), P),
        ("Capital contribution rate", d.get("capital_rate"), P),
        ("Capital months (total)", d.get("capital_months_total"), N),
        ("Capital months (achieved)", d.get("capital_months_achieved"), N),
        ("Target compensation", d.get("target_compensation"), N),
        ("Remuneration TDS section", d.get("remuneration_tds_section"), None),
        ("Remuneration TDS rate", d.get("remuneration_tds_rate"), P),
        ("Remuneration TDS start date", d.get("remuneration_tds_start_date"), None),
    ]
    cell_refs = {}
    row = 2
    for label, value, fmt in rows:
        _set(ws, row, 1, label, bold=True)
        cell = _set(ws, row, 2, "-- not supplied --" if value is None else value,
                    fill=TF, number_format=fmt)
        cell_refs[label] = cell.coordinate
        row += 1
    _autosize(ws, 3)
    return cell_refs


# ---------------------------------------------------------------------------
# 3. Monthly grid
# ---------------------------------------------------------------------------

_MONTHLY_LINE_LABELS = [
    ("remuneration", "Remuneration"),
    ("share_of_profit_gross", "Share of profit (gross)"),
    ("additional_share_of_profit", "Additional share of profit (one-off, net)"),
    ("firms_tax", "Firm's tax"),
    ("tds", "TDS"),
    ("capital_transferred", "Capital transferred"),
    ("total_paid", "Total paid"),
    ("misc", "Misc adjustment (derived, never read off the advice)"),
]


def _write_monthly_grid_sheet(wb, report: Report):
    ws = wb.create_sheet("Monthly grid")
    months = [m.month for m in report.monthly]
    _write_header(ws, 1, ["Line item"] + months + ["Total"])
    for r, (attr, label) in enumerate(_MONTHLY_LINE_LABELS, start=2):
        _set(ws, r, 1, label, bold=True)
        for c, m in enumerate(report.monthly, start=2):
            _set(ws, r, c, getattr(m, attr), number_format=N)
        first_col = get_column_letter(2)
        last_col = get_column_letter(1 + len(report.monthly))
        total_col = 2 + len(report.monthly)
        _set(ws, r, total_col, f"=SUM({first_col}{r}:{last_col}{r})",
             fill=TF, number_format=N, bold=True)
    _autosize(ws, 2 + len(report.monthly))


# ---------------------------------------------------------------------------
# 4. Payroll stream -- conditional.
# ---------------------------------------------------------------------------

def _write_payroll_sheet(wb, report: Report):
    if not report.payroll:
        return
    ws = wb.create_sheet("Payroll stream")
    headers = ["Month", "Gross salary", "TDS (salary)", "Net paid"]
    _write_header(ws, 1, headers)
    row = 2
    for p in report.payroll:
        _set(ws, row, 1, p.get("month"))
        _set(ws, row, 2, p.get("gross_salary"), number_format=N)
        _set(ws, row, 3, p.get("tds"), number_format=N)
        _set(ws, row, 4, p.get("net_paid"), number_format=N)
        row += 1
    total_row = row
    _set(ws, total_row, 1, "Total", bold=True)
    for col in (2, 3, 4):
        letter = get_column_letter(col)
        _set(ws, total_row, col, f"=SUM({letter}2:{letter}{total_row - 1})",
             fill=TF, number_format=N, bold=True)
    _autosize(ws, len(headers))


# ---------------------------------------------------------------------------
# 5. One-offs
# ---------------------------------------------------------------------------

def _write_one_offs_sheet(wb, report: Report):
    ws = wb.create_sheet("One-offs")
    headers = ["Net (as shown on advice)", "Firm's tax rate used", "Gross (derived)",
               "Roundness (distance to nearest Rs 1,00,000)", "Status"]
    _write_header(ws, 1, headers)
    row = 2
    for o in report.one_offs:
        _set(ws, row, 1, o.net, number_format=N)
        _set(ws, row, 2, o.firms_tax_rate, number_format=P)
        _set(ws, row, 3, o.gross, number_format=N)
        _set(ws, row, 4, o.roundness, number_format=N)
        fill = OK if o.status == "CONFIRMED" else (TF if o.status == CANNOT_RECONCILE_LABEL else BAD)
        _set(ws, row, 5, o.status, fill=fill, bold=True)
        row += 1
    if row == 2:
        _set(ws, row, 1, "No one-off (additional share of profit) rows in this year's monthly grid.")
    _autosize(ws, len(headers))


CANNOT_RECONCILE_LABEL = "CANNOT RECONCILE"


# ---------------------------------------------------------------------------
# 6. Cohorts
# ---------------------------------------------------------------------------

def _write_cohorts_sheet(wb, report: Report):
    ws = wb.create_sheet("Cohorts")
    headers = ["Award FY", "Payment date", "Instalment FY", "Membership (this reporting FY)",
               "Gross", "Firm's tax", "Capital deducted", "Net"]
    _write_header(ws, 1, headers)
    row = 2
    for inst in report.cohort_instalments:
        _set(ws, row, 1, inst.award_fy)
        _set(ws, row, 2, inst.payment_date.isoformat())
        _set(ws, row, 3, inst.instalment_fy)
        fill = SF if inst.membership == "reporting" else TF
        _set(ws, row, 4, inst.label, fill=fill)
        _set(ws, row, 5, inst.gross, number_format=N)
        _set(ws, row, 6, inst.firms_tax, number_format=N)
        _set(ws, row, 7, inst.capital, number_format=N)
        _set(ws, row, 8, inst.net, number_format=N)
        row += 1
    if row == 2:
        _set(ws, row, 1, "No incentive cohorts in this run's input.")
    _autosize(ws, len(headers))


# ---------------------------------------------------------------------------
# 7. Capital
# ---------------------------------------------------------------------------

def _write_capital_sheet(wb, report: Report, driver_refs: dict):
    ws = wb.create_sheet("Capital")
    row = 1
    _set(ws, row, 1, "Required cumulative capital (rule, as a formula off Drivers)", bold=True)
    row += 1
    tc_ref = driver_refs.get("Target compensation")
    ma_ref = driver_refs.get("Capital months (achieved)")
    mt_ref = driver_refs.get("Capital months (total)")
    rate_ref = driver_refs.get("Capital contribution rate")
    if report.capital_rule.status == "OK" and all((tc_ref, ma_ref, mt_ref, rate_ref)):
        formula = f"=Drivers!{tc_ref}*(Drivers!{ma_ref}/Drivers!{mt_ref})*Drivers!{rate_ref}"
        _set(ws, row, 1, "Required cumulative capital")
        _set(ws, row, 2, formula, fill=TF, number_format=N, bold=True)
    else:
        _set(ws, row, 1, "Required cumulative capital")
        _set(ws, row, 2, report.capital_rule.reason or CANNOT_RECONCILE_LABEL, fill=TF)
    row += 2

    _set(ws, row, 1, "Cross-check vs Advisory's stated closing capital", bold=True)
    row += 1
    advisory_stated = (report.reconciliation and next(
        (r for r in report.reconciliation if r.category.startswith("Closing capital")), None
    ))
    if advisory_stated is not None:
        for label, value in advisory_stated.sources.items():
            _set(ws, row, 1, label)
            _set(ws, row, 2, "-- not supplied --" if value is None else value, number_format=N)
            row += 1
        status_fill, status_text = _status_fill(advisory_stated.agree)
        _set(ws, row, 1, "Status", bold=True)
        _set(ws, row, 2, status_text, fill=status_fill, bold=True)
        row += 1
        _set(ws, row, 1, advisory_stated.note, wrap=True)
        row += 1
    row += 1

    _set(ws, row, 1, "Mid-year rate-change detector", bold=True)
    row += 1
    if not report.rate_change_suspects:
        _set(ws, row, 1, "No cohort in this run shows unequal capital-deducted "
                         "instalments -- no rate change suspected.")
        row += 1
    else:
        headers = ["Instalment count", "First instalment capital", "Implied old rate",
                   "Implied new rate", "Actual total capital", "Note"]
        _write_header(ws, row, headers)
        row += 1
        for s in report.rate_change_suspects:
            _set(ws, row, 1, s.instalment_count, number_format=N)
            _set(ws, row, 2, s.first_instalment_capital, number_format=N)
            _set(ws, row, 3, s.implied_old_rate, number_format=P)
            _set(ws, row, 4, s.implied_new_rate, number_format=P)
            _set(ws, row, 5, s.actual_total, number_format=N)
            _set(ws, row, 6, s.note, fill=BAD, wrap=True, bold=True)
            row += 1
    _autosize(ws, 6)


# ---------------------------------------------------------------------------
# 8. Reconciliation / 9. Exceptions
# ---------------------------------------------------------------------------

def _write_reconciliation_sheet(wb, report: Report):
    ws = wb.create_sheet("Reconciliation")
    headers = ["Category", "Sources", "Status", "Note"]
    _write_header(ws, 1, headers)
    row = 2
    for r in report.reconciliation:
        _set(ws, row, 1, r.category, wrap=True)
        sources_text = "; ".join(
            f"{k} = {'not supplied' if v is None else format(v, ',.2f')}"
            for k, v in r.sources.items()
        )
        _set(ws, row, 2, sources_text, wrap=True)
        fill, text = _status_fill(r.agree)
        _set(ws, row, 3, text, fill=fill, bold=True)
        _set(ws, row, 4, r.note, wrap=True)
        row += 1
    _autosize(ws, len(headers))


def _write_exceptions_sheet(wb, report: Report):
    ws = wb.create_sheet("Exceptions")
    headers = ["Category", "Status", "Detail"]
    _write_header(ws, 1, headers)
    row = 2
    for r in report.reconciliation:
        if r.agree is not True:
            fill, text = _status_fill(r.agree)
            _set(ws, row, 1, r.category, wrap=True)
            _set(ws, row, 2, text, fill=fill, bold=True)
            _set(ws, row, 3, r.note, wrap=True)
            row += 1
    for s in report.rate_change_suspects:
        _set(ws, row, 1, "Capital -- mid-year rate change", wrap=True)
        _set(ws, row, 2, "RATE CHANGE SUSPECTED", fill=BAD, bold=True)
        _set(ws, row, 3, s.note, wrap=True)
        row += 1
    for o in report.one_offs:
        if o.status == "SUSPECT":
            _set(ws, row, 1, "One-off gross-up roundness", wrap=True)
            _set(ws, row, 2, "SUSPECT", fill=BAD, bold=True)
            _set(ws, row, 3, f"Net {o.net:,.2f} grossed up at {o.firms_tax_rate:.4%} gives "
                              f"{o.gross:,.2f}, {o.roundness:,.2f} from the nearest lakh -- "
                              "the rate for this year is probably wrong.", wrap=True)
            row += 1
    for msg in report.tds_month_exceptions:
        _set(ws, row, 1, "Remuneration TDS applicability", wrap=True)
        _set(ws, row, 2, "EXCEPTION", fill=BAD, bold=True)
        _set(ws, row, 3, msg, wrap=True)
        row += 1
    if row == 2:
        _set(ws, row, 1, "No exceptions -- all reconciliation categories agreed and no "
                         "detector fired.")
    _autosize(ws, len(headers))


# ---------------------------------------------------------------------------
# 10. Open items
# ---------------------------------------------------------------------------

def _write_open_items_sheet(wb, report: Report):
    ws = wb.create_sheet("Open items")
    headers = ["Item", "Why it is open", "What would close it"]
    _write_header(ws, 1, headers)
    row = 2
    for r in report.reconciliation:
        if r.agree is None:
            _set(ws, row, 1, r.category, wrap=True)
            _set(ws, row, 2, r.note, wrap=True)
            _set(ws, row, 3, "Supply the missing source figure for this financial year.",
                 wrap=True)
            row += 1
    if report.capital_rule.status != "OK":
        _set(ws, row, 1, "Required cumulative capital", wrap=True)
        _set(ws, row, 2, report.capital_rule.reason, wrap=True)
        _set(ws, row, 3, "Supply the missing capital driver(s) in the Drivers block.",
             wrap=True)
        row += 1
    if row == 2:
        _set(ws, row, 1, "No open items -- every category was either resolved or produced "
                         "an explicit exception above.")
    _autosize(ws, len(headers))


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def write_report_workbook(report: Report, out_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _write_logic_sheet(wb, report)
    driver_refs = _write_drivers_sheet(wb, report)
    _write_monthly_grid_sheet(wb, report)
    _write_payroll_sheet(wb, report)
    _write_one_offs_sheet(wb, report)
    _write_cohorts_sheet(wb, report)
    _write_capital_sheet(wb, report, driver_refs)
    _write_reconciliation_sheet(wb, report)
    _write_exceptions_sheet(wb, report)
    _write_open_items_sheet(wb, report)
    wb.save(out_path)
