"""
build_tds_journals.py — Form 26AS (Parts I & VI) -> GnuCash multi-split journal CSV.

Reads the Part I (TDS) and Part VI (TCS) sub-totals from a Convert-tab 26AS
workbook and the user's GnuCash account tree, and emits one balanced journal
transaction per deductor / collector, in four categories driven by the section:

  A. Interest   (sections 194A, 193)
       Dr  Expense:TDS on Interest                  = Tax Deducted          (a)
       Dr  Income:Interest Income:Interest on FD    = Amount - Tax  (c - a)     [fixed generic]
       Cr  <matched interest income account>        = Amount Paid/Credited  (c)
  B. Dividend   (section 194)
       Dr  Expense:TDS on Dividend                  = Tax Deducted          (a)
       Cr  <matched dividend income account>        = Tax Deducted          (a)
  C. Partnership (section 194T)
       Dr  Expense:TDS on Partnership Payments      = Tax Deducted          (a)   [emitted as-is]
       Cr  <matched remuneration income account>    = Tax Deducted          (a)
  T. TCS         (Part VI, sections 206C*)
       Dr  Expense:TCS on Foreign Trip              = Tax Collected         (a)
       Cr  Expense:Drawings                         = Tax Collected         (a)
     Only the TAX moves — the spend it was collected on is already booked.
     Both accounts are discovered from the chart, and the credit leg is
     caller-configurable (Bank instead of Drawings when the TCS was paid
     across separately).
  G. 15G/15H    (Part II — TDS for 15G/15H)
       Dr  Expense:TDS on Interest                  = Tax Deducted          (a)   [normally 0]
       Dr  Income:Interest Income:Interest on FD    = Amount - Tax  (c - a)     [fixed generic]
       Cr  <matched interest income account>        = Amount Paid/Credited  (c)
     A RECLASSIFICATION, not new income: the 15G/15H interest is already
     booked in the generic FD bucket, so this moves it into the specific
     NBFC/deductor account. Tax deducted is statutorily supposed to be zero
     for a 15G/15H deductor (that is the point of filing the form) — with
     a=0 this degenerates to a plain 2-split Dr Interest on FD / Cr NBFC. If
     tax is nonetheless non-zero it posts exactly like Category A's 3-split
     (same maths; only the category letter, id series and description
     differ — see CATEGORY_SERIES below). Category is decided by WHICH PART
     the row came from, NEVER by section lookup — Part II rows are almost
     always section 194A too, and a section-based lookup would silently
     merge them into genuine Part I TDS.

Amounts come from the per-party sub-totals (header totals) in Part I / Part VI.
Each transaction is dated 31-March of the current calendar year.

Account matching is DETERMINISTIC (token overlap + acronym + alias table +
single-candidate rule). Anything that cannot be matched with confidence is
routed to Liabilities:Suspense and flagged in the review for LLM/manual
resolution. No network or LLM calls happen here — this module is pure and
testable.

CSV dialect: GnuCash multi-split transaction import. One row per split. The
Date / Transaction ID / Description are REPEATED on every split row of a
transaction — GnuCash groups splits by matching transaction fields / shared
Transaction ID, and does NOT reliably attach blank-date continuation rows
(blank rows import as parse errors). A single signed "Amount" column holds the
split value using GnuCash's convention (Debit = positive, Credit = negative);
each transaction's Amounts sum to zero. Map it to the importer's "Amount"
column type (or "Amount (Negated)" if a build imports the signs reversed).
Transfer Amount / Transfer Account are NOT used: those exist only in two-split
mode (one row = one 2-split transaction), and the Interest journals have three
splits, so the file must be multi-split. Account is the full colon path WITHOUT
the "Root Account:" prefix. Import with the "Multi-split" box ticked, skipping
the 1 header line.

Usage:
  python build_tds_journals.py <26as.xlsx> <book.gnucash> <out.csv>
"""
from __future__ import annotations

import csv
import datetime as _dt
import gzip
import re
import sys
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

# -------------------- fixed account names --------------------

ACC_TDS_INTEREST = "Expense:TDS on Interest"
ACC_TDS_DIVIDEND = "Expense:TDS on Dividend"
ACC_TDS_PARTNERSHIP = "Expense:TDS on Partnership Payments"  # may not exist; emitted as-is
ACC_INTEREST_ON_FD = "Income:Interest Income:Interest on FD"  # fixed generic (Category A debit b)
ACC_SUSPENSE = "Liabilities:Suspense"

# Category T (TCS, Part VI). The debit is the tax the collector took off the
# assessee — a credit claimable in the return, so it must reach the books or
# the refund is understated. The contra is the personal spending the collection
# rode on (a foreign tour package, an LRS remittance), which in these books is
# already sitting in Drawings: the journal reclassifies that slice of Drawings
# into a recoverable tax. Both are defaults — the caller may override either,
# e.g. to credit Bank when the TCS was paid separately rather than bundled into
# the spend.
ACC_TCS_DEFAULT = "Expense:TCS on Foreign Trip"
ACC_DRAWINGS = "Expense:Drawings"

CURRENCY = "INR"

# Section -> category. TCS sections (206C...) are matched by prefix, not by
# this table, because the sub-section suffix varies (206CQ foreign remittance /
# tour package, 206CR sale of goods, 206CL motor vehicle, ...) and they all
# post identically.
SECTION_CATEGORY = {
    "194A": "A", "193": "A",
    "194": "B",
    "194T": "C",
}

TCS_SECTION_PREFIX = "206C"

# Category -> Transaction ID series prefix. SINGLE SOURCE OF TRUTH: both this
# module's write_csv() and ui/tabs/tds_journal_review.py's _txn_id_for() must
# resolve a category to the same series, or the review screen's Save path
# builds a Transaction ID that doesn't exist in the journal CSV and silently
# skips the row (see series_for_category's docstring). The review tab imports
# this dict rather than restating it.
CATEGORY_SERIES = {
    "A": "TDSJ", "B": "TDSJ", "C": "TDSJ",   # Part I (TDS)
    "T": "TCSJ",                              # Part VI (TCS)
    "G": "15GJ",                              # Part II (15G/15H)
    "?": "TDSJ",  # unhandled/mixed section (build_journals' placeholder) —
                  # already routed Suspense/Suspense on both legs, so the
                  # series choice doesn't affect the books; kept as "TDSJ"
                  # only to preserve pre-existing behaviour, not because it
                  # means anything.
}


# Part II's series token, derived from CATEGORY_SERIES rather than restated
# as the literal "15GJ" -- split_part_ii() (below) and anything that needs to
# recognise a Part II Transaction ID must key off this constant, so a future
# rename of the series in CATEGORY_SERIES can't silently desync the splitter
# from the id the journal was actually built with.
PART_II_SERIES = CATEGORY_SERIES["G"]


def series_for_category(category: str) -> str:
    """Resolve a Journal.category to its Transaction ID series prefix.

    Raises ValueError on any category not listed in CATEGORY_SERIES rather
    than silently defaulting to "TDSJ" — a wrong series would build a
    Transaction ID that doesn't exist in the journal CSV, and the review
    screen's Save path would then skip the row without any indication of why
    (see ui/tabs/tds_journal_review.py's _find_credit_split). "?" (the
    placeholder build_journals() uses for an unhandled section) IS listed,
    mapped to "TDSJ", purely to preserve pre-existing behaviour for those
    already-Suspense-routed journals — not because it carries any category
    meaning of its own.
    """
    try:
        return CATEGORY_SERIES[(category or "").strip().upper()]
    except KeyError:
        raise ValueError(
            f"no Transaction ID series defined for category {category!r} — "
            "add it to CATEGORY_SERIES"
        ) from None

# Alias expansion for matching: maps a token found in an ACCOUNT name to the
# set of tokens it stands for in a DEDUCTOR name (and vice-versa via expansion).
ALIASES = {
    "BOB": {"BANK", "BARODA"},
    "EPF": {"PROVIDENT", "FUND"},
    "PF": {"PROVIDENT", "FUND"},
    "DRL": {"REDDY", "REDDYS", "LABORATORIES"},
    "HDFC": {"HDFC"},
    "ICICI": {"ICICI"},
    "SBM": {"SBM"},
    "PNB": {"PUNJAB", "NATIONAL"},
    "LIC": {"LIFE", "INSURANCE"},
}

# Legal / generic noise tokens stripped from deductor names before matching.
STOPWORDS = {
    "LIMITED", "LTD", "LIMITE", "COMPANY", "CO", "PRIVATE", "PVT", "LLP",
    "INVESTMENT", "INVESTMENTS", "FINANCE", "AND", "THE", "OF", "OFFICE",
    "INDIA", "SERVICES", "SERVICE", "REGIONAL", "COMMISSIONER", "BANDRA",
    "EAST", "WEST", "NORTH", "SOUTH", "INC", "CORPORATION", "CORP",
}

# Account-name decoration stripped before matching (keeps the entity token).
ACC_NOISE = {
    "INTEREST", "ON", "FROM", "FD", "BOND", "DIVIDEND", "SHARES", "REMUNERATION",
    "TAXABLE", "BANK", "PARTNERSHIP", "MF", "INCOME",
}


# -------------------- data classes --------------------

@dataclass
class Deductor:
    sr: int
    name: str
    sections: tuple
    amount_paid: float
    tax_deducted: float
    tds_deposited: float


@dataclass
class Account:
    path: str          # full path WITHOUT "Root Account:" prefix
    leaf: str
    type: str
    special: bool = False  # placeholder/hidden/etc. — not a valid post target


@dataclass
class Split:
    account: str
    debit: float = 0.0
    credit: float = 0.0


@dataclass
class Journal:
    sr: int
    deductor: str
    category: str
    section_label: str
    splits: list = field(default_factory=list)
    credit_account: str = ""
    credit_confidence: str = ""        # High / Medium / Low / Suspense
    credit_basis: str = ""
    candidates: list = field(default_factory=list)
    needs_review: bool = False

    @property
    def total_debit(self) -> float:
        return round(sum(s.debit for s in self.splits), 2)

    @property
    def total_credit(self) -> float:
        return round(sum(s.credit for s in self.splits), 2)

    @property
    def balanced(self) -> bool:
        return abs(self.total_debit - self.total_credit) < 0.005


# -------------------- 26AS xlsx parsing --------------------

def parse_part_i(xlsx_path: Path) -> tuple[list[Deductor], str]:
    """Return (deductors, financial_year) from a Convert-tab 26AS workbook."""
    wb = load_workbook(xlsx_path, data_only=True)
    if "Part I" not in wb.sheetnames:
        raise ValueError("Workbook has no 'Part I' sheet — is this a 26AS Convert output?")
    return _parse_party_sheet(wb["Part I"])


def parse_parts(xlsx_path: Path) -> tuple[list[Deductor], list[Deductor], list[Deductor], str]:
    """Return (part_i_deductors, part_ii_deductors, part_vi_collectors, financial_year).

    Any part may be empty — a 26AS with only TCS and no TDS is perfectly
    valid, and so is one with no 15G/15H deductors — but a workbook carrying
    NONE of the three sheets is not a Convert output and is rejected, so a
    wrong file never silently produces an empty journal.

    Part II shares columns 1/2/4/5/6/8 (Sr, Name, the three header totals,
    Section) with Part I/VI byte-for-byte — the missing "Status of Booking"
    column only shifts columns FROM 10 onward, which _parse_party_sheet never
    reads — so the same reader serves all three sheets unchanged.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    if not ({"Part I", "Part II", "Part VI"} & set(wb.sheetnames)):
        raise ValueError("Workbook has none of 'Part I' / 'Part II' / 'Part VI' "
                         "— is this a 26AS Convert output?")
    deductors, fy = ([], "")
    if "Part I" in wb.sheetnames:
        deductors, fy = _parse_party_sheet(wb["Part I"])
    g_deductors, fy2 = ([], "")
    if "Part II" in wb.sheetnames:
        g_deductors, fy2 = _parse_party_sheet(wb["Part II"])
    collectors, fy6 = ([], "")
    if "Part VI" in wb.sheetnames:
        collectors, fy6 = _parse_party_sheet(wb["Part VI"])
    return deductors, g_deductors, collectors, (fy or fy2 or fy6)


def _parse_party_sheet(ws) -> tuple[list[Deductor], str]:
    """Read a deductor/collector sheet (Part I, Part II or Part VI — identical
    geometry through column 8; see parse_parts).

    For Part VI the same fields carry the TCS equivalents: amount_paid is the
    amount paid/DEBITED and tax_deducted the tax COLLECTED.
    """
    # Financial year lives in the meta strip (row 2).
    fy = ""
    meta = ws.cell(2, 1).value or ""
    m = re.search(r"Financial Year:\s*(\d{4}-\d{2})", str(meta))
    if m:
        fy = m.group(1)

    from collections import OrderedDict
    hdr: "OrderedDict[int, list]" = OrderedDict()
    secs: dict[int, list] = {}
    for r in range(4, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if not isinstance(a, int):
            continue
        sr = a
        if sr not in hdr:
            hdr[sr] = [ws.cell(r, 2).value, ws.cell(r, 4).value,
                       ws.cell(r, 5).value, ws.cell(r, 6).value]
            secs[sr] = []
        sec = ws.cell(r, 8).value
        if sec is not None and str(sec) not in secs[sr]:
            secs[sr].append(str(sec))

    deductors = []
    for sr, (name, amt, tax, tds) in hdr.items():
        deductors.append(Deductor(
            sr=sr, name=str(name).strip(), sections=tuple(secs[sr]),
            amount_paid=float(amt or 0), tax_deducted=float(tax or 0),
            tds_deposited=float(tds or 0),
        ))
    return deductors, fy


# -------------------- gnucash account parsing --------------------

# GnuCash marks "special account types" via KVP slots on the account element.
# Such accounts are NOT valid posting targets and must never be offered as a
# credit-account candidate (a placeholder header like 'Income:Interest Income'
# cannot receive splits directly; a hidden account is retired). These keys are
# the boolean 'true'/'false' slots; 'opening-balance' is handled separately as
# an 'equity-type' string slot.
#
# This list is a SELF-CONTAINED mirror of agents.gnucash_accounts.BOOL_FLAG_KEYS
# — this script runs as a stand-alone subprocess and cannot rely on `agents`
# being importable in a frozen child. tests/test_gnucash_accounts.py asserts the
# two lists stay identical so they never drift.
SPECIAL_BOOL_FLAG_KEYS = ("placeholder", "hidden", "tax-related",
                          "auto-interest-transfer")
_SPECIAL_TRUE_VALUES = frozenset({"true", "t", "1", "yes", "y"})
_EQUITY_TYPE_KEY = "equity-type"
_OPENING_BALANCE_VALUE = "opening-balance"


def _account_is_special(acc_el: ET.Element, ns: dict) -> bool:
    """True if a <gnc:account> element carries any special-type flag slot."""
    slots = acc_el.find("act:slots", ns)
    if slots is None:
        return False
    for slot in list(slots):                       # direct <slot> children only
        key = value = ""
        for child in slot:
            local = child.tag.rsplit("}", 1)[-1]
            if local == "key":
                key = (child.text or "").strip()
            elif local == "value":
                value = (child.text or "").strip()
        if key in SPECIAL_BOOL_FLAG_KEYS and value.lower() in _SPECIAL_TRUE_VALUES:
            return True
        if key == _EQUITY_TYPE_KEY and value.lower() == _OPENING_BALANCE_VALUE:
            return True
    return False


def load_accounts(gnucash_path: Path) -> list[Account]:
    raw = gnucash_path.read_bytes()
    data = gzip.decompress(raw) if raw[:2] == b"\x1f\x8b" else raw
    root = ET.fromstring(data)
    ns = {"gnc": "http://www.gnucash.org/XML/gnc",
          "act": "http://www.gnucash.org/XML/act"}
    by_id = {}
    for a in root.iter("{http://www.gnucash.org/XML/gnc}account"):
        name = a.find("act:name", ns).text
        aid = a.find("act:id", ns).text
        typ = a.find("act:type", ns).text
        par = a.find("act:parent", ns)
        by_id[aid] = (name, par.text if par is not None else None, typ,
                      _account_is_special(a, ns))

    def full_path(aid):
        parts = []
        cur = aid
        while cur and cur in by_id:
            n, p, _t, _s = by_id[cur]
            parts.append(n)
            cur = p
        return list(reversed(parts))

    accounts = []
    for aid, (name, _p, typ, special) in by_id.items():
        parts = full_path(aid)
        # Drop the root account name ("Root Account") from the path.
        if parts and parts[0].lower().startswith("root"):
            parts = parts[1:]
        if not parts:
            continue
        accounts.append(Account(path=":".join(parts), leaf=parts[-1], type=typ,
                                special=special))
    return accounts


def find_account(accounts: list[Account], path: str) -> Optional[Account]:
    for a in accounts:
        if a.path == path:
            return a
    return None


# -------------------- matching --------------------

def _tokens(s: str) -> list[str]:
    return [t for t in re.split(r"[^A-Za-z0-9]+", s.upper()) if t]


def _sig_tokens(s: str, drop: set) -> set:
    return {t for t in _tokens(s) if t not in drop and len(t) > 1}


def _acronym(tokens: list[str]) -> str:
    return "".join(t[0] for t in tokens if t)


# Tokens that are generic to a fixed-deposit interest account (i.e. NOT a
# bank/entity name). Used to tell the generic 'Interest on FD' account apart
# from deductor-specific ones like 'Interest on BOB - FD'.
FD_NOISE = {"INTEREST", "ON", "FROM", "INCOME", "RECEIVED", "EARNED", "OTHER",
            "THE", "A", "OF", "FD", "FIXED", "DEPOSIT", "DEPOSITS"}


def find_generic_fd_account(accounts: list[Account]) -> Optional[str]:
    """Deterministically locate the GENERIC fixed-deposit interest income
    account — the Category-A second debit ('Interest on FD' in the spec).

    Its exact name varies per book: 'Interest on FD', 'Interest on Fixed
    Deposit', 'FD Interest', 'Interest on Fixed Deposits', etc. We fuzzy-match:
    an INCOME account under an interest subtree that mentions FD / Fixed Deposit
    and carries NO specific bank/entity token (a specific token means it's a
    deductor account like 'Interest on BOB - FD', not the generic one)."""
    best: Optional[Account] = None
    for a in accounts:
        if a.special:
            continue  # placeholder/hidden — not a valid post target
        if a.type != "INCOME" or "interest" not in a.path.lower():
            continue
        toks = _tokens(a.leaf)
        tokset = set(toks)
        has_fd = ("FD" in tokset) or ({"FIXED", "DEPOSIT"} <= tokset)
        if not has_fd:
            continue
        entity = {t for t in toks if t not in FD_NOISE and len(t) > 1}
        if entity:
            continue  # deductor-specific (e.g. 'Interest on BOB - FD')
        # If several generic candidates, prefer the shortest (simplest) leaf.
        if best is None or len(a.leaf) < len(best.leaf):
            best = a
    return best.path if best else None


def find_tcs_account(accounts: list[Account]) -> Optional[str]:
    """Locate the account that carries TCS suffered, e.g. 'TCS on Foreign Trip'.

    Its name varies by what the collection was on ('TCS on Foreign Trip', 'TCS
    on LRS Remittance', 'TCS Receivable'), so we match on the TCS token rather
    than a fixed path. Deliberately NOT restricted to one account type: books
    model this either as an expense reclassified out of Drawings or as an asset
    (a receivable from the department). Prefer the shortest leaf when several
    exist, and never a placeholder."""
    best: Optional[Account] = None
    for a in accounts:
        if a.special:
            continue
        if "TCS" not in _tokens(a.leaf):
            continue
        if best is None or len(a.leaf) < len(best.leaf):
            best = a
    return best.path if best else None


def find_drawings_account(accounts: list[Account]) -> Optional[str]:
    """Locate the Drawings account — the default TCS contra."""
    best: Optional[Account] = None
    for a in accounts:
        if a.special:
            continue
        if "DRAWINGS" not in _tokens(a.leaf):
            continue
        if best is None or len(a.leaf) < len(best.leaf):
            best = a
    return best.path if best else None


def _candidates_for(category: str, accounts: list[Account],
                    fd_exclude: str = "") -> list[Account]:
    """Income accounts in the relevant subtree. Subtree match is case-insensitive
    and lenient (handles 'Interest Income', 'Interest Received', 'Dividend
    Income', 'Dividend - Shares', etc.). The generic FD account (fd_exclude) is
    never offered as a credit candidate, so a deductor can't land a debit and
    credit on the same account."""
    out = []
    for a in accounts:
        if a.special:
            continue  # placeholder/hidden — never a credit candidate
        if a.type != "INCOME":
            continue
        pl = a.path.lower()
        if category == "A" and "interest" in pl:
            if fd_exclude and a.path == fd_exclude:
                continue
            out.append(a)
        elif category == "B" and "dividend" in pl:
            out.append(a)
        elif category == "C" and "remuneration" in pl:
            out.append(a)
    return out


def match_credit_account(deductor: str, category: str,
                         accounts: list[Account],
                         fd_account: str = "") -> tuple[Optional[str], str, str, list]:
    """Return (account_path|None, confidence, basis, candidate_paths)."""
    cands = _candidates_for(category, accounts, fd_account)
    cand_paths = [c.path for c in cands]
    if not cands:
        return None, "Suspense", "no income accounts in category subtree", []

    # Single-candidate rule (e.g. Category C has only 'Remuneration from Partnership').
    if len(cands) == 1:
        return cands[0].path, "Medium", "only candidate in category subtree", cand_paths

    d_tokens_list = [t for t in _tokens(deductor) if t not in STOPWORDS and len(t) > 1]
    d_tokens = set(d_tokens_list)
    d_acro = _acronym(d_tokens_list)

    best = None
    best_score = 0.0
    best_basis = ""
    for c in cands:
        c_tokens = _sig_tokens(c.leaf, ACC_NOISE)
        score = 0.0
        hits = []
        # 1) direct shared significant tokens
        shared = d_tokens & c_tokens
        if shared:
            score += 2.0 * len(shared)
            hits.append("token:" + "/".join(sorted(shared)))
        # 2) prefix / substring (CHOLA in CHOLAMANDALAM)
        for ct in c_tokens:
            for dt in d_tokens:
                if ct == dt:
                    continue
                if len(ct) >= 3 and (dt.startswith(ct) or ct.startswith(dt)):
                    score += 1.5
                    hits.append(f"prefix:{ct}~{dt}")
        # 3) acronym: account token equals deductor acronym (BOB, DRL)
        for ct in c_tokens:
            if len(ct) >= 2 and ct == d_acro:
                score += 2.5
                hits.append(f"acronym:{ct}")
        # 4) alias table
        for ct in c_tokens:
            alias = ALIASES.get(ct)
            if alias and (alias & d_tokens):
                score += 2.0
                hits.append(f"alias:{ct}")
        if score > best_score:
            best_score = score
            best = c
            best_basis = ", ".join(hits)

    if best is None or best_score < 1.5:
        return None, "Suspense", "no confident match", cand_paths
    conf = "High" if best_score >= 2.0 else "Medium"
    return best.path, conf, best_basis, cand_paths


# -------------------- journal construction --------------------

def categorize(sections: tuple) -> tuple[Optional[str], str]:
    """Return (category, label). label is the section string for the description."""
    cats = {SECTION_CATEGORY.get(s) for s in sections}
    cats.discard(None)
    label = "/".join(sections)
    if len(cats) == 1:
        return cats.pop(), label
    return None, label  # unknown or mixed -> needs review


def build_journals(deductors: list[Deductor], accounts: list[Account],
                   overrides: Optional[dict] = None) -> list[Journal]:
    """overrides: {deductor_sr (int) -> credit account full path}. Used by the
    LLM-fallback path to resolve deductors the deterministic matcher sent to
    Suspense. An override always wins over the deterministic choice."""
    overrides = overrides or {}
    # The generic FD-interest debit account ('Interest on FD' in the spec) is
    # fuzzy-found from the actual chart, since its name varies per book
    # ('Interest on Fixed Deposit', etc.). Falls back to the canonical name
    # (surfaced under 'accounts to create') only if no FD account exists.
    fd_account = find_generic_fd_account(accounts) or ACC_INTEREST_ON_FD
    journals = []
    for d in deductors:
        cat, label = categorize(d.sections)
        j = Journal(sr=d.sr, deductor=d.name, category=cat or "?", section_label=label)
        a = round(d.tax_deducted, 2)
        c = round(d.amount_paid, 2)

        if cat is None:
            # Unknown/mixed section — park the tax on Suspense, flag.
            j.splits = [Split(ACC_SUSPENSE, debit=a), Split(ACC_SUSPENSE, credit=a)]
            j.credit_account = ACC_SUSPENSE
            j.credit_confidence = "Suspense"
            j.credit_basis = f"unhandled section(s): {label}"
            j.needs_review = True
            journals.append(j)
            continue

        acct, conf, basis, cands = match_credit_account(d.name, cat, accounts, fd_account)
        j.candidates = cands
        if d.sr in overrides and overrides[d.sr]:
            credit_acc = overrides[d.sr]
            j.credit_account = credit_acc
            j.credit_confidence = "Override"
            j.credit_basis = "resolved by override/LLM"
            j.needs_review = False
        else:
            credit_acc = acct or ACC_SUSPENSE
            j.credit_account = credit_acc
            j.credit_confidence = conf
            j.credit_basis = basis
            j.needs_review = acct is None or conf in ("Low", "Suspense")

        if cat == "A":
            j.splits = [
                Split(ACC_TDS_INTEREST, debit=a),
                Split(fd_account, debit=round(c - a, 2)),
                Split(credit_acc, credit=c),
            ]
        elif cat == "B":
            j.splits = [
                Split(ACC_TDS_DIVIDEND, debit=a),
                Split(credit_acc, credit=a),
            ]
        elif cat == "C":
            j.splits = [
                Split(ACC_TDS_PARTNERSHIP, debit=a),
                Split(credit_acc, credit=a),
            ]
        journals.append(j)
    return journals


def build_15g_journals(deductors: list[Deductor], accounts: list[Account],
                       overrides: Optional[dict] = None) -> list[Journal]:
    """Part II (15G/15H) -> Category G journals.

    Reuses Category A's exact posting template and account-matching pool
    (interest income accounts) — no new maths, per the spec: with tax
    deducted (a) at its statutory value of 0 this degenerates to a plain
    2-split Dr Interest on FD / Cr NBFC transfer, and if a is ever nonzero it
    posts the identical 3-split Category A does.

    Category is hardcoded to "G" here — NEVER derived from
    categorize()/SECTION_CATEGORY. Part II deductors are almost always
    section 194A, which SECTION_CATEGORY maps to "A"; routing them through
    that lookup would silently merge 15G/15H reclassifications into the
    TDSJ id series, indistinguishable from genuine Part I TDS in the review
    CSV (see CATEGORY_SERIES / series_for_category).

    overrides: {deductor_sr (int) -> credit account full path}, same shape
    as build_journals' — Part II Sr numbers restart per part (like Part VI),
    so this is deliberately a separate map from the Part I overrides.
    """
    overrides = overrides or {}
    fd_account = find_generic_fd_account(accounts) or ACC_INTEREST_ON_FD
    journals = []
    for d in deductors:
        label = "/".join(d.sections)
        j = Journal(sr=d.sr, deductor=d.name, category="G", section_label=label)
        a = round(d.tax_deducted, 2)
        c = round(d.amount_paid, 2)

        # Account matching reuses Category A's candidate pool (interest income
        # accounts) — a 15G/15H deductor's amount is interest income exactly
        # like a Category A one; only the tax-withholding status differs.
        acct, conf, basis, cands = match_credit_account(d.name, "A", accounts, fd_account)
        j.candidates = cands
        if d.sr in overrides and overrides[d.sr]:
            credit_acc = overrides[d.sr]
            j.credit_account = credit_acc
            j.credit_confidence = "Override"
            j.credit_basis = "resolved by override/LLM"
            j.needs_review = False
        else:
            credit_acc = acct or ACC_SUSPENSE
            j.credit_account = credit_acc
            j.credit_confidence = conf
            j.credit_basis = basis
            j.needs_review = acct is None or conf in ("Low", "Suspense")

        if a:
            # Tax deducted despite 15G/15H (shouldn't happen, but if it does,
            # handle it exactly like Category A's 3-split).
            j.splits = [
                Split(ACC_TDS_INTEREST, debit=a),
                Split(fd_account, debit=round(c - a, 2)),
                Split(credit_acc, credit=c),
            ]
        else:
            # Statutory case: no tax withheld -> a plain reclassification,
            # Dr Interest on FD / Cr the specific NBFC account. Total income
            # is unchanged, this only moves it out of the generic bucket.
            j.splits = [
                Split(fd_account, debit=c),
                Split(credit_acc, credit=c),
            ]
        journals.append(j)
    return journals


def is_tcs_section(sections: tuple) -> bool:
    """True if every section on the collector block is a 206C sub-section."""
    secs = [str(s).strip().upper() for s in sections if str(s).strip()]
    return bool(secs) and all(s.startswith(TCS_SECTION_PREFIX) for s in secs)


def build_tcs_journals(collectors: list[Deductor], accounts: list[Account],
                       tcs_account: str = "", credit_account: str = "",
                       overrides: Optional[dict] = None) -> list[Journal]:
    """One 2-split journal per collector:

        Dr  <TCS account>       = Tax Collected
        Cr  <credit account>    = Tax Collected

    Only the TAX is journalled, never the amount paid/debited: the underlying
    spend is already in the books (that is what the collection rode on), so
    posting the gross would double-count it.

    Both accounts default to what the chart actually contains and fall back to
    the canonical names, which then surface under 'accounts to create' rather
    than importing silently wrong. credit_account is a parameter because the
    contra is a bookkeeping choice, not a statutory one — Drawings when the TCS
    was bundled into personal spending, Bank when it was paid across separately.

    match_credit_account is deliberately NOT used here: it searches INCOME
    subtrees, and a collection at source is a spending-side event with no
    income leg to match.

    overrides: {collector_sr -> credit account path}, same shape as the TDS path.
    """
    overrides = overrides or {}
    dr = tcs_account or find_tcs_account(accounts) or ACC_TCS_DEFAULT
    cr_default = credit_account or find_drawings_account(accounts) or ACC_DRAWINGS

    journals = []
    for c in collectors:
        label = "/".join(c.sections)
        j = Journal(sr=c.sr, deductor=c.name, category="T", section_label=label)
        tax = round(c.tax_deducted, 2)

        if c.sr in overrides and overrides[c.sr]:
            cr = overrides[c.sr]
            j.credit_confidence = "Override"
            j.credit_basis = "resolved by override/LLM"
        else:
            cr = cr_default
            j.credit_confidence = "High" if credit_account or \
                find_drawings_account(accounts) else "Medium"
            j.credit_basis = ("TCS contra — " + (
                "credit account supplied by caller" if credit_account
                else f"defaulted to {cr}"))

        if not is_tcs_section(c.sections):
            # A non-206C section under Part VI means the sheet was misread or
            # the department used a section we don't model. Park it loudly
            # rather than posting a tax credit we can't justify.
            j.splits = [Split(ACC_SUSPENSE, debit=tax), Split(ACC_SUSPENSE, credit=tax)]
            j.credit_account = ACC_SUSPENSE
            j.credit_confidence = "Suspense"
            j.credit_basis = f"unexpected non-206C section(s) in Part VI: {label}"
            j.needs_review = True
            journals.append(j)
            continue

        j.credit_account = cr
        j.splits = [Split(dr, debit=tax), Split(cr, credit=tax)]
        journals.append(j)
    return journals


# -------------------- output --------------------

def journal_date() -> str:
    """31-March of the current calendar year, ISO format."""
    return _dt.date(_dt.date.today().year, 3, 31).isoformat()


def fy_prefix(fy: str) -> str:
    """Compact financial-year prefix for Transaction IDs: '2025-26' -> '2526'.

    Keeps Transaction IDs unique across years (next year's TDSJ01 would
    otherwise collide). Falls back to the journal date's FY when fy is blank
    (31-March of year Y closes FY (Y-1)-(Y))."""
    m = re.match(r"\s*(\d{4})-(\d{2})\s*$", fy or "")
    if m:
        return m.group(1)[2:] + m.group(2)
    y = _dt.date.today().year
    return f"{(y - 1) % 100:02d}{y % 100:02d}"


def _fmt(x: float) -> str:
    return f"{x:.2f}" if x else ""


# Column order for the GnuCash multi-split journal CSV. Shared by
# build_csv_rows()/write_csv() here and by ui/tabs/tds_journal_review.py's
# _JOURNAL_HEADERS (restated there, not imported, because that module reads
# rows back with csv.DictReader rather than building them from Journal
# objects -- but the two lists must stay in the same order and shape).
JOURNAL_HEADERS = ["Date", "Transaction ID", "Number", "Description", "Account",
                   "Amount", "Currency"]


def build_csv_rows(journals: list[Journal], fy: str) -> list[dict]:
    """Build one row dict per split, in JOURNAL_HEADERS order/shape.

    This is the single place that turns Journal objects into the CSV row
    shape everything else operates on: write_csv() writes these rows
    verbatim, and split_part_ii() (below) partitions this exact shape -- the
    review screen's Save path partitions the same shape read back off disk
    via csv.DictReader, so one splitter serves both call sites.
    """
    date = journal_date()
    fy_pfx = fy_prefix(fy)
    rows: list[dict] = []
    for j in journals:
        # Each category gets its own ID series (see CATEGORY_SERIES): Part
        # I Sr.1, Part II Sr.1 and Part VI Sr.1 are different parties, and
        # a shared prefix would make GnuCash's multi-split importer fuse
        # their splits into one unbalanced transaction.
        kind = series_for_category(j.category)
        if j.category == "T":
            tag, part_label = "TCS", "Part VI"
        elif j.category == "G":
            tag, part_label = "15G/15H TDS", "Part II (15G/15H)"
        else:
            tag, part_label = "TDS", "Part I"
        txn_id = f"{fy_pfx}-{kind}{j.sr:02d}"
        desc = (f"{part_label} {tag} FY {fy} - {j.deductor} (Sec {j.section_label})"
                 if fy else
                 f"{part_label} {tag} - {j.deductor} (Sec {j.section_label})")
        # Repeat Date / Transaction ID / Description on EVERY split row.
        # GnuCash's multi-split importer groups splits by matching
        # transaction fields (and the Transaction ID), NOT by blank-date
        # continuation rows -- blank rows show up as parse errors.
        for s in j.splits:
            signed = round(s.debit - s.credit, 2)   # Dr +, Cr -
            # Number duplicates Transaction ID -> GnuCash visible Num field.
            rows.append({
                "Date": date, "Transaction ID": txn_id, "Number": txn_id,
                "Description": desc, "Account": s.account,
                "Amount": f"{signed:.2f}", "Currency": CURRENCY,
            })
    return rows


def write_rows_csv(rows: list[dict], out_path: Path) -> None:
    """Write JOURNAL_HEADERS-shaped row dicts to out_path (multi-split
    GnuCash layout). Shared by write_csv() and the Part-I-only split file so
    the two files are byte-for-byte the same dialect."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=JOURNAL_HEADERS)
        w.writeheader()
        w.writerows(rows)


def write_csv(journals: list[Journal], out_path: Path, fy: str) -> None:
    # GnuCash-native multi-split layout. One signed "Amount" column per split
    # (maps to the importer's "Amount" column type). GnuCash's sign
    # convention: Debit = positive, Credit = negative; per transaction the
    # Amounts sum to zero. (If a given build of GnuCash imports the signs
    # reversed, map this column to "Amount (Negated)" instead.)
    write_rows_csv(build_csv_rows(journals, fy), out_path)


def _series_token(txn_id: str) -> Optional[str]:
    """Parse the series token out of a Transaction ID shaped
    f"{fy_prefix}-{kind}{sr:02d}" -- everything after the LAST '-', with
    trailing digits stripped. fy_prefix itself may contain '-' (it doesn't
    today, but nothing guarantees that), so splitting on the FIRST '-' would
    be wrong; rsplit(..., 1) is deliberate. Returns None if txn_id has no '-'
    or the tail is entirely digits (no series letters survive the strip)."""
    txn_id = (txn_id or "").strip()
    if "-" not in txn_id:
        return None
    tail = txn_id.rsplit("-", 1)[1]
    token = tail.rstrip("0123456789")
    return token or None


def split_part_ii(journal_rows: list[dict]) -> tuple[list[dict], list[dict], list[str]]:
    """Partition JOURNAL_HEADERS-shaped CSV row dicts into
    (part_i_rows, part_ii_rows, problems).

    Whole TRANSACTIONS are partitioned, never individual splits: every row
    sharing a Transaction ID goes to the same side, decided once per id (the
    first row seen for that id) rather than per row -- splitting a
    transaction across both files would silently unbalance whichever file
    lost a split, the exact failure this feature exists to prevent.

    A Transaction ID whose series token cannot be parsed is kept in
    part_i_rows -- NOT dropped. Silently discarding a real journal entry
    because its id looked strange is a far worse failure than leaving it in
    a file the user reviews before import; the id is also appended to
    `problems` so the anomaly is never silent either.
    """
    problems: list[str] = []
    side_by_txn: dict[str, str] = {}   # txn_id -> "II" or "I"
    for row in journal_rows:
        txn_id = (row.get("Transaction ID") or "").strip()
        if txn_id in side_by_txn:
            continue
        token = _series_token(txn_id)
        if token is None:
            side_by_txn[txn_id] = "I"
            problems.append(
                f"Transaction ID {txn_id!r} has no parseable series -- "
                "kept in the Part I file rather than dropped"
            )
        elif token == PART_II_SERIES:
            side_by_txn[txn_id] = "II"
        else:
            side_by_txn[txn_id] = "I"

    part_i_rows: list[dict] = []
    part_ii_rows: list[dict] = []
    for row in journal_rows:
        txn_id = (row.get("Transaction ID") or "").strip()
        if side_by_txn.get(txn_id) == "II":
            part_ii_rows.append(row)
        else:
            part_i_rows.append(row)
    return part_i_rows, part_ii_rows, problems


def part_i_path_for(out_path: Path) -> Path:
    """<stem>-tds-journals.csv -> <stem>-tds-journals-partI.csv (sibling),
    same naming convention as the existing -review.csv sibling."""
    return out_path.with_name(out_path.stem + "-partI.csv")


def write_part_i_split(rows: list[dict], out_path: Path) -> tuple[Optional[str], list[str]]:
    """Write/delete the Part-I-only sibling of out_path, given the FULL set
    of just-written journal rows.

    Returns (part_i_output_path_or_None, problems). If no Part II rows exist
    in `rows`, no file is written -- and if a stale one from a PREVIOUS run
    is sitting there, it is deleted, because a leftover Part-I file the
    current run didn't need is exactly the hand-filtered-copy-goes-stale bug
    this feature replaces. Called both by run() (generation time) and by
    ui/tabs/tds_journal_review.py's _save_changes() (every re-save), so a
    hand-filtered copy can never exist without being regenerated in lockstep.
    """
    part_i_path = part_i_path_for(out_path)
    part_i_rows, part_ii_rows, problems = split_part_ii(rows)
    if not part_ii_rows:
        if part_i_path.exists():
            part_i_path.unlink()
        return None, problems
    write_rows_csv(part_i_rows, part_i_path)
    return str(part_i_path), problems


def write_review(journals: list[Journal], path: Path, accounts: list[Account]) -> None:
    existing = {a.path for a in accounts}
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Sr", "Deductor", "Section", "Category", "Credit Account",
                    "Confidence", "Account Exists", "Balanced", "Debit", "Credit",
                    "Needs Review", "Basis"])
        for j in journals:
            w.writerow([
                j.sr, j.deductor, j.section_label, j.category, j.credit_account,
                j.credit_confidence, "yes" if j.credit_account in existing else "NO",
                "yes" if j.balanced else "NO", f"{j.total_debit:.2f}",
                f"{j.total_credit:.2f}", "yes" if j.needs_review else "",
                j.credit_basis,
            ])


# -------------------- orchestration --------------------

def run(xlsx_path: Path, gnucash_path: Path, out_path: Path,
        overrides: Optional[dict] = None, tcs_credit_account: str = "",
        tcs_overrides: Optional[dict] = None,
        g_overrides: Optional[dict] = None) -> dict:
    deductors, g_deductors, collectors, fy = parse_parts(xlsx_path)
    accounts = load_accounts(gnucash_path)
    journals = build_journals(deductors, accounts, overrides)
    # Part II (15G/15H). Sr numbers restart per part, so this is a separate
    # overrides map — a shared one would let Part I Sr.2 silently redirect
    # Part II Sr.2.
    journals += build_15g_journals(g_deductors, accounts, overrides=g_overrides)
    # Part VI TCS. Sr numbers restart per part, so TCS overrides are a separate
    # map — a shared one would let Part I Sr.2 silently redirect Part VI Sr.2.
    journals += build_tcs_journals(collectors, accounts,
                                   credit_account=tcs_credit_account,
                                   overrides=tcs_overrides)
    csv_rows = build_csv_rows(journals, fy)
    write_rows_csv(csv_rows, out_path)

    review_path = out_path.with_name(out_path.stem + "-review.csv")
    write_review(journals, review_path, accounts)

    # Part-I-only sibling for users who post 15G/15H (Part II) by hand in
    # GnuCash: importing the full journal's 15GJ rows on top of a hand-posted
    # entry double-books the Interest-on-FD -> NBFC reclassification. This
    # file is ALWAYS regenerated from the journal just written (never a
    # stale hand-filtered copy) -- see write_part_i_split's docstring.
    part_i_output, part_ii_problems = write_part_i_split(csv_rows, out_path)

    existing = {a.path for a in accounts}

    # Any split account (debit or credit) not present in the book must be
    # created by the user before import (e.g. 'Expense:TDS on Partnership
    # Payments'). Surface them explicitly so nothing imports silently wrong.
    missing = sorted({s.account for j in journals for s in j.splits
                      if s.account not in existing})

    review_rows = []
    for j in journals:
        review_rows.append({
            "sr": j.sr, "deductor": j.deductor, "section": j.section_label,
            "category": j.category, "credit_account": j.credit_account,
            "confidence": j.credit_confidence,
            "account_exists": j.credit_account in existing,
            "balanced": j.balanced, "needs_review": j.needs_review,
            "candidates": j.candidates, "basis": j.credit_basis,
        })

    return {
        "fy": fy,
        "journal_date": journal_date(),
        "deductors": len(deductors),
        "part_ii_deductors": len(g_deductors),
        "collectors": len(collectors),
        "balanced_all": all(j.balanced for j in journals),
        "needs_review": [r for r in review_rows if r["needs_review"]],
        "missing_accounts": missing,
        "rows": review_rows,
        "output": str(out_path),
        "review": str(review_path),
        "part_i_output": part_i_output,
        "part_ii_problems": part_ii_problems,
    }


def main(argv: list[str]) -> int:
    if len(argv) not in (4, 5):
        print("Usage: python build_tds_journals.py <26as.xlsx> <book.gnucash> "
              "<out.csv> [overrides.json]", file=sys.stderr)
        return 2
    overrides = None
    if len(argv) == 5:
        import json
        raw = json.loads(Path(argv[4]).read_text(encoding="utf-8"))
        # Keys are deductor Sr numbers. Be tolerant of keys that carry extra
        # text (a tool-calling model may echo the display label "Sr 7" rather
        # than "7"): pull the digits out and skip any key with no number so a
        # stray label never crashes the build with int("Sr 7").
        overrides = {}
        for k, v in raw.items():
            m = re.search(r"\d+", str(k))
            if m:
                overrides[int(m.group(0))] = v
    stats = run(Path(argv[1]), Path(argv[2]), Path(argv[3]), overrides)
    print(f"FY {stats['fy']}  date {stats['journal_date']}  "
          f"deductors {stats['deductors']}  part_ii_deductors {stats['part_ii_deductors']}  "
          f"collectors {stats['collectors']}  balanced_all {stats['balanced_all']}")
    for r in stats["rows"]:
        flag = "  <-- REVIEW" if r["needs_review"] else ""
        print(f"  Sr{r['sr']:>2} [{r['category']}/{r['section']:<5}] {r['deductor'][:34]:34} "
              f"-> {r['credit_account']}  ({r['confidence']}){flag}")
        if r["needs_review"] and r["candidates"]:
            print(f"        candidates: {', '.join(r['candidates'])}")
    if stats["needs_review"]:
        print(f"\n{len(stats['needs_review'])} deductor(s) need review (Suspense/low confidence).")
    if stats["missing_accounts"]:
        print("\nAccounts to CREATE in GnuCash before import:")
        for acc in stats["missing_accounts"]:
            print(f"  - {acc}")
    if stats["part_ii_problems"]:
        print("\nWarning: could not determine the series for some Transaction "
              "ID(s) -- kept in the Part I file, not dropped:")
        for p in stats["part_ii_problems"]:
            print(f"  - {p}")
    print(f"Saved: {stats['output']}")
    print(f"Review: {stats['review']}")
    if stats["part_i_output"]:
        print(f"Part I only (excludes 15G/15H, for hand-posted Part II): "
              f"{stats['part_i_output']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
