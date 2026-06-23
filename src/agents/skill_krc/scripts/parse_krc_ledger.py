"""
parse_krc_ledger.py — Simplify a KR Choksey broker annual ledger statement.

Part I of the KR Choksey broker-import skill: extract the ledger table from a
password-protected statement PDF, drop the broker's internal segment-to-segment
transfer entries (NSE_CASH / BSE_CASH / NSE_SLBM journal hops), and produce a
single-sheet "Simplified Ledger" Excel workbook whose running balance ties out
to the statement's own printed closing balance.

Usage (called by tools.py via subprocess, or directly for testing):
    python parse_krc_ledger.py <pdf_path> <password> <output_xlsx_path>

Exit codes:
    0  success (closing balance reconciles)
    1  bad arguments / file not found / decryption failed
    2  parsing succeeded but the recomputed closing balance does NOT match
       the statement's printed closing balance (output is still written,
       flagged, so a human can inspect it)

Algorithm notes
----------------
The PDF has no embedded table structure — rows are reconstructed from word
x0/top coordinates via pdfplumber. Column x0 ranges (validated against the
sample KR Choksey AC109 statement):
    Date         x0  ~13-14
    V.No         x0  ~58-79
    Particulars  x0  ~106+   (free text, may wrap across several lines)
    ChqNo        x0  ~360-405
    Debit        x0  ~410-440
    Credit       x0  ~460-490
    Balance      x0  ~525-555
    Dr/Cr suffix x0  ~558-580

Row boundaries are NOT simply "one PDF line = one row" — Particulars text
wraps, and a new logical row can start either at a Date/V.No pair, a bare
V.No (same-day continuation), or a line whose first word is an exchange
segment token (NSE_CASH / BSE_CASH / NSE_SLBM and their no-underscore
variants) when the row already in progress has accumulated more than just
a bare segment-token label. A small number of real entries (OPENING BALANCE,
*PAYMENT PAID BY NEFT/RTGS, BILL ENTRY FOR L2-...) are themselves zero-value
reference/annotation lines in this statement's print layout — their actual
value movement is recorded on an adjacent line (e.g. the bill referenced by
a "BILL ENTRY FOR L2-..." line is actually booked on an earlier JV row with
no bill number of its own). build_simplified_ledger() drops these
zero-value rows from the final output (confirmed acceptable by the user,
2026-06-22, even though it loses the bill-number link to the row that
carries the matching amount) — every row in the Simplified Ledger sheet
has a real Debit or Credit.

Self-verifying oracle: for every row with a printed Balance, the signed
balance (positive if suffix is "Cr", negative if "Dr") must equal
previous_signed_balance - debit + credit. This is checked across the FULL
original row set before any filtering, and again on the simplified
("real entries only") set against the statement's own printed closing
balance line.

Internal-transfer elimination rule (locked decision, confirmed against real
data — literal V.No/JV-number pairing does NOT hold reliably):
    Drop every row whose Particulars contains the marker text
    "INTER EXCHANGE SETL". These are broker-internal segment-to-segment
    journal transfers, not real economic transactions.
"""
from __future__ import annotations

import re
import subprocess
import sys
from collections import defaultdict
from pathlib import Path

DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
SEGMENT_TOKEN_RE = re.compile(r"^(NSE_CASH|BSE_CASH|NSE_SLBM|NSECASH|BSECASH|NSESLBM)$")
INTERNAL_TRANSFER_MARKER = "INTER EXCHANGE SETL"

COLUMN_HEADER_LINE = "Date V. No Particulars ChqNo. Debit Credit Balance"
BOILERPLATE_MARKERS = (
    "Company :Grp1",
    "Closing Balance",
    "Yours faithfully",
    "Authorised Signatory",
    "Pan No",
    "Account Ledger Page",
    "Print Date",
    "Running Account Authorisation",
    "This is computer generated",
    "I / We herby confirm",
)
SKIP_LINE_SUBSTRINGS = (
    "Dear Sir",
    "Sub: Confirmation",
    "Please find herewith",
    "If any discrepancies",
)

# Column x0 ranges, validated against the sample statement.
VNO_X0 = (50, 100)
CHQNO_X0 = (360, 405)
DEBIT_X0 = (410, 440)
CREDIT_X0 = (460, 490)
BALANCE_X0 = (525, 555)
DRCR_X0 = (558, 580)


def _num(s: str | None) -> float | None:
    if not s:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def _cluster_lines(words: list[dict], tol: float = 2.0) -> list[list[dict]]:
    """Group words into logical PDF lines by rounded `top` coordinate."""
    buckets: dict[float, list[dict]] = defaultdict(list)
    for w in words:
        key = round(w["top"] / tol) * tol
        buckets[key].append(w)
    return [sorted(buckets[k], key=lambda w: w["x0"]) for k in sorted(buckets.keys())]


def decrypt_pdf(pdf_path: str, password: str, scratch_path: str) -> None:
    """Decrypt a password-protected PDF via qpdf. Raises on failure."""
    result = subprocess.run(
        ["qpdf", f"--password={password}", "--decrypt", pdf_path, scratch_path],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"qpdf decryption failed (wrong password, or qpdf not installed): "
            f"{result.stderr.strip()}"
        )


def extract_rows(decrypted_pdf_path: str) -> tuple[list[dict], dict]:
    """
    Reconstruct logical ledger rows from the decrypted PDF.

    Returns (rows, statement_totals) where statement_totals is the parsed
    "Closing Balance: <debit_total> <credit_total> <balance> <Dr|Cr>" line,
    or {} if not found.
    """
    import pdfplumber

    rows: list[dict] = []
    statement_totals: dict = {}

    with pdfplumber.open(decrypted_pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages):
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            lines = _cluster_lines(words)

            started = False  # becomes True once we pass this page's column-header line
            cur_date: str | None = None
            cur_vno = ""
            cur_particulars: list[str] = []
            cur_chqno = ""
            seen_segment_fragment = False

            def flush_no_amount() -> None:
                nonlocal cur_vno, cur_particulars, cur_chqno, seen_segment_fragment
                rows.append(
                    {
                        "date": cur_date,
                        "vno": cur_vno,
                        "particulars": " ".join(cur_particulars),
                        "chqno": cur_chqno,
                        "debit": None,
                        "credit": None,
                        "balance": None,
                        "drcr": None,
                        "page": page_no + 1,
                    }
                )
                cur_vno, cur_particulars, cur_chqno = "", [], ""
                seen_segment_fragment = False

            for line in lines:
                line_text = " ".join(w["text"] for w in line)

                if line_text.strip() == COLUMN_HEADER_LINE:
                    started = True
                    continue
                if not started:
                    continue
                if any(marker in line_text for marker in BOILERPLATE_MARKERS):
                    if "Closing Balance" in line_text:
                        nums = [_num(w["text"]) for w in line if _num(w["text"]) is not None]
                        drcr_words = [w["text"] for w in line if w["text"] in ("Dr", "Cr")]
                        if len(nums) >= 3:
                            statement_totals = {
                                "total_debit": nums[0],
                                "total_credit": nums[1],
                                "closing_balance": nums[2],
                                "closing_drcr": drcr_words[0] if drcr_words else None,
                            }
                        # Everything after "Closing Balance" on this page is
                        # covering-letter / authorisation-clause / signature
                        # boilerplate that wraps across several lines, none of
                        # which repeat a BOILERPLATE_MARKERS substring (only
                        # the first line of each wrapped sentence does) - so
                        # it would otherwise fall through and get appended as
                        # row continuation text, eventually landing words like
                        # "SHARES"/"SECURITIES"/"PRIVATE" in the Debit/Credit/
                        # Balance x0 ranges purely by coincidence (verified
                        # against the AC109 sample, page 3, BP09634444 row).
                        # Flush whatever's buffered (consistent with how every
                        # other zero-value reference row gets flushed) and
                        # stop parsing this page entirely - it's never another
                        # ledger row past this point.
                        if cur_particulars:
                            flush_no_amount()
                        break
                    continue
                if any(s in line_text for s in SKIP_LINE_SUBSTRINGS):
                    continue

                debit = credit = balance = drcr = None
                remaining: list[dict] = []
                for w in line:
                    x = w["x0"]
                    if DEBIT_X0[0] <= x <= DEBIT_X0[1]:
                        debit = w["text"]
                    elif CREDIT_X0[0] <= x <= CREDIT_X0[1]:
                        credit = w["text"]
                    elif BALANCE_X0[0] <= x <= BALANCE_X0[1]:
                        balance = w["text"]
                    elif DRCR_X0[0] <= x <= DRCR_X0[1] and w["text"] in ("Cr", "Dr"):
                        drcr = w["text"]
                    elif CHQNO_X0[0] <= x <= CHQNO_X0[1]:
                        cur_chqno = (cur_chqno + " " + w["text"]).strip()
                    else:
                        remaining.append(w)

                is_new_row_prefix = bool(remaining) and (
                    DATE_RE.match(remaining[0]["text"])
                    or (VNO_X0[0] <= remaining[0]["x0"] <= VNO_X0[1])
                )

                # Baseline-collision guard: this statement sometimes prints a
                # row's amount columns ~1pt above the FOLLOWING row's
                # Date/V.No label - close enough that line-clustering (tol=2.0)
                # merges them into one logical "line". When that happens, the
                # amount found on this merged line belongs to the row already
                # buffered in cur_particulars, not to the new label we're
                # about to parse. Flush the pending buffer with this line's
                # amount first, then start the new row fresh (no amount of
                # its own from this line).
                if is_new_row_prefix and (debit or credit or balance) and cur_particulars:
                    rows.append(
                        {
                            "date": cur_date,
                            "vno": cur_vno,
                            "particulars": " ".join(cur_particulars),
                            "chqno": cur_chqno,
                            "debit": debit,
                            "credit": credit,
                            "balance": balance,
                            "drcr": drcr,
                            "page": page_no + 1,
                        }
                    )
                    cur_vno, cur_particulars, cur_chqno = "", [], ""
                    seen_segment_fragment = False
                    debit = credit = balance = drcr = None  # already consumed above

                if remaining and DATE_RE.match(remaining[0]["text"]):
                    cur_date = remaining[0]["text"]
                    remaining = remaining[1:]
                    if remaining and VNO_X0[0] <= remaining[0]["x0"] <= VNO_X0[1]:
                        cur_vno = remaining[0]["text"]
                        remaining = remaining[1:]
                    seen_segment_fragment = False
                elif remaining and VNO_X0[0] <= remaining[0]["x0"] <= VNO_X0[1]:
                    cur_vno = remaining[0]["text"]
                    remaining = remaining[1:]
                    seen_segment_fragment = False

                # A line starting with an exchange-segment token, when the
                # current row already has more than a bare segment label
                # accumulated, marks the start of a NEW logical row (the
                # broker re-prints the segment label at the start of each
                # fresh narration line). Flush what's accumulated so far
                # (amount-less; its value, if any, was already captured on
                # an earlier line in this same buffer).
                if remaining and SEGMENT_TOKEN_RE.match(remaining[0]["text"]):
                    if seen_segment_fragment and len(cur_particulars) > 1:
                        flush_no_amount()
                    seen_segment_fragment = True

                cur_particulars.extend(w["text"] for w in remaining)

                if debit or credit or balance:
                    rows.append(
                        {
                            "date": cur_date,
                            "vno": cur_vno,
                            "particulars": " ".join(cur_particulars),
                            "chqno": cur_chqno,
                            "debit": debit,
                            "credit": credit,
                            "balance": balance,
                            "drcr": drcr,
                            "page": page_no + 1,
                        }
                    )
                    cur_vno, cur_particulars, cur_chqno = "", [], ""
                    seen_segment_fragment = False

    return rows, statement_totals


def verify_balance_invariant(rows: list[dict]) -> list[str]:
    """
    Check, for every row with a printed balance, that
    signed_balance == prev_signed_balance - debit + credit.
    Returns a list of human-readable mismatch descriptions (empty = all OK).
    """
    mismatches = []
    prev_signed: float | None = None
    for i, r in enumerate(rows):
        b = _num(r["balance"])
        if b is None:
            continue
        d = _num(r["debit"]) or 0.0
        c = _num(r["credit"]) or 0.0
        signed = b if r["drcr"] == "Cr" else -b
        if prev_signed is not None:
            expected = prev_signed - d + c
            if abs(expected - signed) > 0.02:
                mismatches.append(
                    f"Row {i} (date={r['date']}, vno={r['vno']}): "
                    f"expected balance {expected:.2f}, got {signed:.2f} "
                    f"-- {r['particulars'][:60]!r}"
                )
        prev_signed = signed
    return mismatches


_ACCOUNTCODE_PREFIX_RE = re.compile(
    r"^ACCCOUNTCODE:\S+,ACCOUNTNAME:.*?,FROM\s*:CDSL,TO\s*"
)
_LEADING_DOUBLE_SEGMENT_RE = re.compile(
    r"^(?:NSE_CASH|BSE_CASH|NSE_SLBM|NSECASH|BSECASH|NSESLBM)\s+"
    r"(?=(?:NSE_CASH|BSE_CASH|NSE_SLBM|NSECASH|BSECASH|NSESLBM)\b)"
)


def _clean_particulars(text: str) -> str:
    """
    Strip two display-only print-layout artifacts that occasionally glue two
    narration fragments together with no separator on this statement (the
    underlying Date/V.No/amounts are unaffected either way - this only
    tidies the Particulars text used in the final output):

    1. A DEMAT-transfer row's "ACCCOUNTCODE:...,FROM :CDSL,TO" boilerplate
       sometimes has the row's real description (e.g. "BSE_CASH DP BALANCE
       TRANSFER") appended right after "TO". When extra text follows "TO",
       drop the boilerplate prefix and keep just the real description. When
       nothing follows "TO" (the boilerplate IS the whole narration, e.g. a
       plain DEMAT fund transfer with no further label), leave it untouched.
    2. A bare leading segment token (e.g. "NSESLBM") immediately followed by
       a second, different segment token (e.g. "BSE_CASH DP BALANCE
       TRANSFER") is a leftover first-line label from a multi-line
       narration; the second token is the row's real leading label, so the
       first is dropped.
    """
    stripped = _ACCOUNTCODE_PREFIX_RE.sub("", text, count=1)
    if stripped.strip():
        text = stripped
    text = _LEADING_DOUBLE_SEGMENT_RE.sub("", text, count=1)
    return text.strip()


# ---------------------------------------------------------------------------
# Part II support (added 2026-06-22): coarse row tagging + recovery of the
# bill-settlement / NEFT-RTGS linkage rows that the simplified ledger drops.
# ---------------------------------------------------------------------------
_DEMAT_MARKER = "DP BALANCE TRANSFER"
_OPENING_MARKER = "OPENING BALANCE"
_BANK_MARKERS = ("RECEIVED AMOUNT", "HDFC-CMS", "ATOM", "NEFT", "RTGS")
_BILL_ANCHOR_RE = re.compile(r"BILL ENTRY FOR (L2|M)-(\d+)")
_NEFT_MARKER = "*PAYMENT PAID BY NEFT/RTGS"


def _tag_row(row: dict, utr_dates: set) -> str:
    """
    Classify one real (amount-bearing) ledger row into a coarse category used
    as a *hint* by Phase II reconciliation. This is heuristic, not
    authoritative: Phase II still matches contract-note bills by amount and may
    reclassify a row (e.g. a trade settled by an incoming bank credit is tagged
    'Bank Pay-*' here but becomes 'Trade Bill' once a same-amount bill is
    found). A row with no opening / demat / bank signal is left as a bill
    candidate ('Settlement Movement').
    """
    p = row["particulars"] or ""
    if _OPENING_MARKER in p:
        return "Opening Balance"
    if _DEMAT_MARKER in p and not row["chqno"]:
        return "Demat Charge"
    is_bank = (
        bool(row["chqno"])
        or any(k in p for k in _BANK_MARKERS)
        or row["date"] in utr_dates
    )
    if is_bank:
        return "Bank Pay-In" if row["credit"] else "Bank Pay-Out"
    return "Settlement Movement"


def collect_references(rows: list[dict]) -> list[dict]:
    """
    Recover the zero-value linkage rows that build_simplified_ledger() drops:

      * Bill-settlement anchors - 'NSE_SLBM BILL ENTRY FOR L2-<setl>' and
        'BSE_CASH BILL ENTRY FOR M-<setl>' - give each bill's settlement number.
      * NEFT/RTGS payout refs   - '*PAYMENT PAID BY NEFT/RTGS' - give the bank
        UTR / instrument number for each pay-out.

    These carry no Debit/Credit of their own (their value moves on an adjacent
    row), so they are intentionally absent from the Simplified Ledger sheet.
    They are written to a separate 'References' sheet so Phase II can match
    bills two ways (amount AND settlement number) without re-parsing the PDF.
    """
    refs: list[dict] = []
    for r in rows:
        p = r["particulars"] or ""
        m = _BILL_ANCHOR_RE.search(p)
        if m:
            refs.append({
                "ref_type": "Bill Anchor",
                "date": r["date"],
                "vno": r["vno"],
                "exchange": "NSE_SLBM" if m.group(1) == "L2" else "BSE_CASH",
                "settlement_or_utr": m.group(2),
            })
        elif _NEFT_MARKER in p:
            refs.append({
                "ref_type": "Bank Payment (NEFT/RTGS)",
                "date": r["date"],
                "vno": r["vno"],
                "exchange": None,
                "settlement_or_utr": (r["chqno"] or "").replace(" ", ""),
            })
    return refs


def build_simplified_ledger(rows: list[dict], utr_dates: set | None = None) -> tuple[list[dict], float]:
    """
    Drop internal-transfer rows (marker-tagged) and zero-value reference/label
    rows; recompute a running balance from the remaining real Debit/Credit
    rows. Each surviving row additionally carries a coarse 'tag' (see
    _tag_row) used as a Phase II matching hint.

    Returns (simplified_rows, recomputed_closing_balance).
    """
    if utr_dates is None:
        utr_dates = set()
    real_rows = [
        r for r in rows
        if INTERNAL_TRANSFER_MARKER not in r["particulars"]
        and (r["debit"] or r["credit"])
    ]

    simplified = []
    running_balance = 0.0
    for r in real_rows:
        d = _num(r["debit"]) or 0.0
        c = _num(r["credit"]) or 0.0
        running_balance = running_balance - d + c
        drcr = "Cr" if running_balance >= 0 else "Dr"
        simplified.append(
            {
                "date": r["date"],
                "vno": r["vno"],
                "particulars": _clean_particulars(r["particulars"]),
                "chqno": r["chqno"],
                "debit": d if d else None,
                "credit": c if c else None,
                "balance": abs(running_balance),
                "drcr": drcr,
                "tag": _tag_row(r, utr_dates),
            }
        )
    return simplified, running_balance


def write_xlsx(simplified_rows: list[dict], references: list[dict], output_path: str) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Simplified Ledger"

    headers = ["Date", "V.No", "Particulars", "ChqNo", "Debit", "Credit", "Balance", "Dr/Cr", "Tag"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in simplified_rows:
        ws.append(
            [
                r["date"],
                r["vno"],
                r["particulars"],
                r["chqno"],
                r["debit"],
                r["credit"],
                round(r["balance"], 2) if r["balance"] is not None else None,
                r["drcr"],
                r.get("tag"),
            ]
        )

    widths = [12, 14, 55, 14, 14, 14, 14, 8, 20]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

    # Second sheet: recovered bill-settlement anchors + NEFT/RTGS payout refs.
    ws2 = wb.create_sheet("References")
    rheaders = ["Ref Type", "Date", "V.No", "Exchange", "Settlement / UTR"]
    ws2.append(rheaders)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for ref in references:
        ws2.append(
            [
                ref["ref_type"],
                ref["date"],
                ref["vno"],
                ref["exchange"],
                ref["settlement_or_utr"],
            ]
        )
    rwidths = [24, 12, 16, 12, 22]
    for col_idx, width in enumerate(rwidths, start=1):
        ws2.column_dimensions[ws2.cell(1, col_idx).column_letter].width = width

    wb.save(output_path)


def main() -> int:
    if len(sys.argv) < 4:
        print("Usage: parse_krc_ledger.py <pdf_path> <password> <output_xlsx_path>", file=sys.stderr)
        return 1

    pdf_path, password, output_path = sys.argv[1], sys.argv[2], sys.argv[3]

    if not Path(pdf_path).is_file():
        print(f"ERROR: PDF not found: {pdf_path}", file=sys.stderr)
        return 1

    scratch = str(Path(output_path).with_suffix(".decrypted.tmp.pdf"))
    try:
        decrypt_pdf(pdf_path, password, scratch)
    except RuntimeError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1

    try:
        rows, statement_totals = extract_rows(scratch)
    finally:
        Path(scratch).unlink(missing_ok=True)

    if not rows:
        print("ERROR: no ledger rows could be extracted from the PDF.", file=sys.stderr)
        return 1

    mismatches = verify_balance_invariant(rows)
    print(f"Extracted {len(rows)} ledger rows.")
    if mismatches:
        print(f"WARNING: {len(mismatches)} row(s) failed the balance invariant check:")
        for m in mismatches[:10]:
            print(f"  {m}")
    else:
        print("Balance invariant check: all rows reconcile.")

    internal_count = sum(1 for r in rows if INTERNAL_TRANSFER_MARKER in r["particulars"])
    print(f"Internal transfer rows removed: {internal_count}")

    references = collect_references(rows)
    utr_dates = {r["date"] for r in references if r["ref_type"].startswith("Bank Payment")}

    simplified, recomputed_closing = build_simplified_ledger(rows, utr_dates)
    zero_value_count = (len(rows) - internal_count) - len(simplified)
    print(f"Zero-value reference rows dropped: {zero_value_count}")
    print(f"Real rows kept: {len(simplified)}")
    print(f"References recovered (bill anchors + NEFT/RTGS): {len(references)}")

    write_xlsx(simplified, references, output_path)
    print(f"Wrote {len(simplified)} ledger rows + {len(references)} references to {output_path}")

    rc = 0
    if statement_totals:
        stated_closing = statement_totals["closing_balance"]
        stated_signed = stated_closing if statement_totals["closing_drcr"] == "Cr" else -stated_closing
        if abs(stated_signed - recomputed_closing) > 0.02:
            print(
                f"WARNING: recomputed closing balance ({recomputed_closing:.2f}) does NOT match "
                f"the statement's printed closing balance ({stated_signed:.2f})."
            )
            rc = 2
        else:
            print(
                f"Closing balance verified: recomputed {recomputed_closing:.2f} matches "
                f"statement's printed closing balance ({stated_signed:.2f})."
            )
    else:
        print("WARNING: could not locate a 'Closing Balance' line in the PDF to verify against.")
        rc = 2

    return rc


if __name__ == "__main__":
    raise SystemExit(main())
